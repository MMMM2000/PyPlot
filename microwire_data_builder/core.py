"""Core data processing for the microwire database builder."""

from __future__ import annotations

import csv
import hashlib
import importlib.util
import json
import sys
import logging
import math
import os
import re
import shutil
import subprocess
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, Hashable, Iterable, Iterator, List, Mapping, Optional, Sequence, Set, Tuple, cast
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import numpy as np
import pandas as pd

from plotting.shared.transition_analysis import estimate_temperature_transition_points

try:
    from .video import extract_video_metrics
except ImportError:
    module_name = "microwire_data_builder.video"
    module_path = Path(__file__).with_name("video.py")
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec and spec.loader:
        module = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = module
        spec.loader.exec_module(module)
        extract_video_metrics = module.extract_video_metrics
    else:
        raise

try:
    from .manual_diameters import MANUAL_DIAMETER_OVERRIDES
except ImportError:
    module_name = "microwire_data_builder.manual_diameters"
    module_path = Path(__file__).with_name("manual_diameters.py")
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec and spec.loader:
        module = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = module
        spec.loader.exec_module(module)
        MANUAL_DIAMETER_OVERRIDES = module.MANUAL_DIAMETER_OVERRIDES
    else:
        raise

sys.modules.setdefault("microwire_data_builder.core", sys.modules[__name__])

os.environ.setdefault("MPLBACKEND", "Agg")

LOGGER_NAME = "microwire_data_builder"
R_CHECK_THRESHOLD = 0.05
DEFAULT_OUTPUT_NAME = "microwire_database"
PLOT_DIR_NAME = "plots"
ORIGIN_DIR_NAME = "origin_objects"
WORD_REPORT_DIR_NAME = "word_reports"


class BuildCancelledError(Exception):
    """Raised when a build is cancelled by the caller."""

MICRO_SIGN = "µ"
MICROSCOPE_RESIZE_TARGET = 2200
FOCUS_ROI_LIMIT = 3

CURRENT_DENSITY_EXTRA_COLUMNS = [
    "As1 (mA)",
    "Af1 (mA)",
    "Ms1 (mA)",
    "Mf1 (mA)",
    "As2 (mA)",
    "Af2 (mA)",
    "Ms2 (mA)",
    "Mf2 (mA)",
    "As current density (A/mm^2)",
    "Ms current density (A/mm^2)",
    "J_As1 (A/mm^2)",
    "J_Af1 (A/mm^2)",
    "J_Ms1 (A/mm^2)",
    "J_Mf1 (A/mm^2)",
    "J_As2 (A/mm^2)",
    "J_Af2 (A/mm^2)",
    "J_Ms2 (A/mm^2)",
    "J_Mf2 (A/mm^2)",
    "As2-As1 (mA)",
    "Af2-Af1 (mA)",
    "Ms2-Ms1 (mA)",
    "Mf2-Mf1 (mA)",
    "Mf1-Af1 (mA)",
    "Mf2-Af2 (mA)",
    "Setpoints (mA)",
    "Sources",
    "Current annealing transition status",
    "Current annealing transition review counts",
]

SHAPE_MEMORY_VALUE_COLUMNS = [
    "Displacement (mm)",
    "Load (g)",
    "Strain (%)",
    "Stress (MPa)",
]

SHAPE_MEMORY_FRACTURE_COLUMNS = [
    "Fracture load (g)",
    "Fracture strain (%)",
    "Fracture stress (MPa)",
]

SHAPE_MEMORY_ENTRY_ALIASES = {
    "Shape memory displacement (mm)": "Displacement (mm)",
    "Shape memory load (g)": "Load (g)",
    "Shape memory strain (%)": "Strain (%)",
    "Shape memory stress (MPa)": "Stress (MPa)",
    "Shape memory fracture load (g)": "Fracture load (g)",
    "Shape memory fracture strain (%)": "Fracture strain (%)",
    "Shape memory fracture stress (MPa)": "Fracture stress (MPa)",
}

STRAIN_ENTRY_ALIASES = {
    "Strain": "Legacy strain",
    "Stress (MPa)": "Legacy stress (MPa)",
}

EA_VALENCE = {
    "Ni": 10,
    "Fe": 8,
    "Co": 9,
    "Cu": 11,
    "Ga": 3,
    "Ge": 4,
    "Si": 4,
    "Sn": 4,
}
EA_COMPOSITION_RE = re.compile(r"([A-Z][a-z]?)(\d+(?:\.\d+)?)")


def _compute_ea_from_composition(composition: str) -> Optional[float]:
    if not composition:
        return None
    matches = EA_COMPOSITION_RE.findall(str(composition))
    if not matches:
        return None
    total = 0.0
    weighted = 0.0
    for element, count_text in matches:
        if element not in EA_VALENCE:
            return None
        try:
            count = float(count_text)
        except (TypeError, ValueError):
            return None
        total += count
        weighted += count * EA_VALENCE[element]
    if total <= 0:
        return None
    return round(weighted / total, 2)


def _estimate_transition_temp_c(ea_value: Optional[float]) -> Optional[float]:
    if ea_value is None:
        return None
    try:
        numeric = float(ea_value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(numeric):
        return None
    temp_k = 3000.0 * numeric - 22880.0
    temp_c = temp_k - 273.15
    return round(temp_c, 1)

STRAIN_EXTRA_COLUMNS = [
    "Calc mode",
    "Clamp span (mm)",
    "m",
    "Legacy stress (MPa)",
    "M length",
    "A length",
    "Broke",
]

CORE_TEMPERATURE_COLUMN = "Core temperature (°C)"
GLASS_TEMPERATURE_COLUMN = "Glass temperature (°C)"
ESTIMATED_TRANSITION_COLUMN = "Tt est (°C)"
GLASS_PULL_COLUMN = "Glass pull-off"
VIDEO_END_LENGTH_COLUMN = "Video end length (m)"
VIDEO_MW_LENGTH_COLUMN = "Video wire range (m)"
ANNEALING_TRANSITION_COLUMN = "Current annealing transition currents"
CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN = "Current annealing transition status"
CURRENT_ANNEALING_TRANSITION_COUNTS_COLUMN = "Current annealing transition review counts"
VSM_TRANSITION_TEMP_STATUS_COLUMN = "VSM transition temp status"
VSM_TRANSITION_TEMP_COUNTS_COLUMN = "VSM transition temp review counts"
MINI_DMA_TRANSITION_STATUS_COLUMN = "Mini DMA transition status"
MINI_DMA_TRANSITION_COUNTS_COLUMN = "Mini DMA transition review counts"

OUTPUT_COLUMNS = [
    "Composition",
    "Microwire",
    "e/a",
    ESTIMATED_TRANSITION_COLUMN,
    "d (µm)",
    "D (µm)",
    "d/D",
    "Brittle",
    "Legacy strain",
    *STRAIN_EXTRA_COLUMNS,
    "As (mA)",
    "Ms (mA)",
    *CURRENT_DENSITY_EXTRA_COLUMNS,
    "As (°C)",
    "Af (°C)",
    "Ms (°C)",
    "Mf (°C)",
    "Length (m)",
    "Production datetime",
    "Mass (g)",
    "Resistance (Ω)",
    CORE_TEMPERATURE_COLUMN,
    GLASS_TEMPERATURE_COLUMN,
    "Winding speed (m/min)",
    "Glass feeding (mm/min)",
    "Underpressure",
    GLASS_PULL_COLUMN,
    VIDEO_END_LENGTH_COLUMN,
    VIDEO_MW_LENGTH_COLUMN,
    "Notes",
    "Data source",
    "File 1000 mA",
    "Other annealing files",
    ANNEALING_TRANSITION_COLUMN,
    "Figure — 1000 mA",
    "Figure — other annealing",
    "Figure — 1000 mA (Origin)",
    "Figure — other annealing (Origin)",
    "VSM hysteresis graphs",
    "VSM temperature scan graphs",
    "DMA iso-stress graphs",
    "Mini DMA graphs",
    "Manual stress/strain graphs",
    *SHAPE_MEMORY_VALUE_COLUMNS,
    *SHAPE_MEMORY_FRACTURE_COLUMNS,
    "FMR graphs",
]

DIAMETER_COLUMN = "d (µm)"
GLASS_DIAMETER_COLUMN = "D (µm)"
DIAMETER_RATIO_COLUMN = "d/D"
BRITTLE_COLUMN = "Brittle"

MICROSCOPE_IMAGE_COLUMNS = (
    "d (µm) image",
    "D (µm) image",
)

FIGURE_COLUMNS = (
    "Figure — 1000 mA",
    "Figure — other annealing",
)

VSM_HYSTERESIS_COLUMN = "VSM hysteresis graphs"
VSM_TEMPERATURE_SCAN_COLUMN = "VSM temperature scan graphs"
DMA_ISOSTRESS_COLUMN = "DMA iso-stress graphs"
MINI_DMA_COLUMN = "Mini DMA graphs"
MINI_DMA_STRAIN_COLUMN = "Mini DMA strain by stress/load"
MINI_DMA_TRANSITION_COLUMN = "Mini DMA transition currents by stress/load"
MINI_DMA_BREAK_COLUMN = "Mini DMA break point"
SHAPE_MEMORY_STRESS_STRAIN_COLUMN = "Manual stress/strain graphs"
LEGACY_SHAPE_MEMORY_STRESS_STRAIN_COLUMN = "Shape memory stress/strain graphs"
VSM_HYSTERESIS_ORIGIN_COLUMN = "VSM hysteresis graphs (Origin)"
VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN = "VSM temperature scan graphs (Origin)"
DMA_ISOSTRESS_ORIGIN_COLUMN = "DMA iso-stress graphs (Origin)"
MINI_DMA_ORIGIN_COLUMN = "Mini DMA graphs (Origin)"
SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN = "Manual stress/strain graphs (Origin)"
LEGACY_SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN = "Shape memory stress/strain graphs (Origin)"
SHAPE_MEMORY_DISPLACEMENT_COLUMN = "Displacement (mm)"
SHAPE_MEMORY_LOAD_COLUMN = "Load (g)"
SHAPE_MEMORY_STRAIN_COLUMN = "Strain (%)"
SHAPE_MEMORY_STRESS_COLUMN = "Stress (MPa)"
SHAPE_MEMORY_FRACTURE_LOAD_COLUMN = "Fracture load (g)"
SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN = "Fracture strain (%)"
SHAPE_MEMORY_FRACTURE_STRESS_COLUMN = "Fracture stress (MPa)"
FMR_COLUMN = "FMR graphs"
FMR_ORIGIN_COLUMN = "FMR graphs (Origin)"
RVT_FILE_COLUMN = "R vs T files"
RVT_GRAPH_COLUMN = "R vs T graphs"
RVT_ORIGIN_COLUMN = "R vs T graphs (Origin)"
RVT_RESIDUAL_ORIGIN_COLUMN = "R vs T residual graphs (Origin)"
RVT_POINT_COUNT_COLUMN = "R vs T points"
RVT_TEMPERATURE_RANGE_COLUMN = "R vs T temperature range (deg C)"
RVT_RESISTANCE_RANGE_COLUMN = "R vs T resistance range (Ohm)"

TRANSITION_TEMP_AS_COLUMN = "As (°C)"
TRANSITION_TEMP_AF_COLUMN = "Af (°C)"
TRANSITION_TEMP_MS_COLUMN = "Ms (°C)"
TRANSITION_TEMP_MF_COLUMN = "Mf (°C)"
TRANSITION_TEMP_COLUMNS = (
    TRANSITION_TEMP_AS_COLUMN,
    TRANSITION_TEMP_AF_COLUMN,
    TRANSITION_TEMP_MS_COLUMN,
    TRANSITION_TEMP_MF_COLUMN,
)

STRAIN_COLUMN = "Legacy strain"

_REVIEW_COUNT_KEYS = (
    "total",
    "accepted",
    "manual",
    "no_transition",
    "excluded",
    "needs_attention",
    "unreviewed",
    "auto_candidates",
)


def _empty_transition_review_counts() -> Dict[str, int]:
    return {key: 0 for key in _REVIEW_COUNT_KEYS}


def _coerce_review_count(value: object) -> int:
    try:
        numeric = int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0
    return max(numeric, 0)


def _normalise_transition_status(raw_status: object) -> str:
    text = str(raw_status or "").strip()
    folded = text.casefold().replace("-", "_").replace(" ", "_")
    mapping = {
        "accepted": "Accepted auto",
        "accepted_auto": "Accepted auto",
        "auto_accepted": "Accepted auto",
        "manual": "Manual adjusted",
        "manual_adjusted": "Manual adjusted",
        "no_transition": "No transition",
        "excluded": "Excluded",
        "needs_attention": "Needs attention",
        "partly_reviewed": "Partly reviewed",
        "partial": "Partly reviewed",
        "auto_candidate": "Auto candidate",
        "auto_candidates": "Auto candidate",
        "unreviewed": "Unreviewed",
        "not_measured": "Not measured",
        "no_scans": "Not measured",
    }
    return mapping.get(folded, text or "Unreviewed")


def _review_counts_from_payload(payload: object) -> Dict[str, int]:
    counts = _empty_transition_review_counts()
    if not isinstance(payload, Mapping):
        return counts
    aliases = {
        "accepted_auto": "accepted",
        "accepted": "accepted",
        "manual_adjusted": "manual",
        "manual": "manual",
        "no-transition": "no_transition",
        "no transition": "no_transition",
        "excluded": "excluded",
        "needs attention": "needs_attention",
        "needs_attention": "needs_attention",
        "unreviewed": "unreviewed",
        "auto": "auto_candidates",
        "auto_candidates": "auto_candidates",
        "total": "total",
    }
    for raw_key, value in payload.items():
        key = aliases.get(str(raw_key).strip().casefold())
        if key in counts:
            counts[key] += _coerce_review_count(value)
    if counts["total"] <= 0:
        counts["total"] = (
            counts["accepted"]
            + counts["manual"]
            + counts["no_transition"]
            + counts["excluded"]
            + counts["needs_attention"]
            + counts["unreviewed"]
        )
    return counts


def _format_transition_review_counts(counts: Mapping[str, int]) -> str:
    total = int(counts.get("total", 0) or 0)
    if total <= 0:
        return ""
    return "; ".join(
        f"{key}={int(counts.get(key, 0) or 0)}"
        for key in _REVIEW_COUNT_KEYS
    )


def _aggregate_transition_review_status(counts: Mapping[str, int]) -> str:
    total = int(counts.get("total", 0) or 0)
    if total <= 0:
        return "Not measured"
    manual = int(counts.get("manual", 0) or 0)
    accepted = int(counts.get("accepted", 0) or 0)
    no_transition = int(counts.get("no_transition", 0) or 0)
    excluded = int(counts.get("excluded", 0) or 0)
    needs_attention = int(counts.get("needs_attention", 0) or 0)
    unreviewed = int(counts.get("unreviewed", 0) or 0)
    negative = no_transition + excluded
    reviewed = manual + accepted + negative
    if needs_attention and needs_attention == total:
        return "Needs attention"
    if manual == total:
        return "Manual adjusted"
    if accepted == total:
        return "Accepted auto"
    if negative == total:
        return "No transition" if no_transition else "Excluded"
    if reviewed or needs_attention:
        return "Partly reviewed"
    if int(counts.get("auto_candidates", 0) or 0):
        return "Auto candidate"
    if unreviewed or total:
        return "Unreviewed"
    return "Not measured"


def _increment_review_count_for_status(
    counts: Dict[str, int],
    status: object,
    *,
    has_manual_values: bool = False,
    has_auto_values: bool = False,
) -> None:
    normalised = _normalise_transition_status(status)
    counts["total"] += 1
    if normalised == "Manual adjusted" or has_manual_values:
        counts["manual"] += 1
    elif normalised == "Accepted auto":
        counts["accepted"] += 1
    elif normalised == "No transition":
        counts["no_transition"] += 1
    elif normalised == "Excluded":
        counts["excluded"] += 1
    elif normalised == "Needs attention":
        counts["needs_attention"] += 1
    else:
        counts["unreviewed"] += 1
        if has_auto_values:
            counts["auto_candidates"] += 1


def _clean_review_values(values: object) -> Dict[str, float]:
    if not isinstance(values, Mapping):
        return {}
    cleaned: Dict[str, float] = {}
    for label in ("As", "Af", "Ms", "Mf", "As1", "Af1", "Ms1", "Mf1", "As2", "Af2", "Ms2", "Mf2"):
        value = values.get(label)
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            cleaned[label] = float(value)
    return cleaned

for _graph_column, _origin_column in (
    (VSM_HYSTERESIS_COLUMN, VSM_HYSTERESIS_ORIGIN_COLUMN),
    (VSM_TEMPERATURE_SCAN_COLUMN, VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN),
    (DMA_ISOSTRESS_COLUMN, DMA_ISOSTRESS_ORIGIN_COLUMN),
    (MINI_DMA_COLUMN, MINI_DMA_ORIGIN_COLUMN),
    (SHAPE_MEMORY_STRESS_STRAIN_COLUMN, SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN),
    (FMR_COLUMN, FMR_ORIGIN_COLUMN),
):
    if _graph_column in OUTPUT_COLUMNS and _origin_column not in OUTPUT_COLUMNS:
        OUTPUT_COLUMNS.insert(OUTPUT_COLUMNS.index(_graph_column) + 1, _origin_column)

for _mini_dma_extra_column in reversed(
    (
        MINI_DMA_STRAIN_COLUMN,
        MINI_DMA_TRANSITION_COLUMN,
        MINI_DMA_TRANSITION_STATUS_COLUMN,
        MINI_DMA_TRANSITION_COUNTS_COLUMN,
        MINI_DMA_BREAK_COLUMN,
    )
):
    if _mini_dma_extra_column not in OUTPUT_COLUMNS:
        OUTPUT_COLUMNS.insert(OUTPUT_COLUMNS.index(MINI_DMA_ORIGIN_COLUMN) + 1, _mini_dma_extra_column)

for _transition_temp_status_column in reversed(
    (VSM_TRANSITION_TEMP_STATUS_COLUMN, VSM_TRANSITION_TEMP_COUNTS_COLUMN)
):
    if _transition_temp_status_column not in OUTPUT_COLUMNS:
        OUTPUT_COLUMNS.insert(
            OUTPUT_COLUMNS.index(TRANSITION_TEMP_MF_COLUMN) + 1,
            _transition_temp_status_column,
        )

ORIGIN_FIGURE_COLUMNS = tuple(
    column
    for column in OUTPUT_COLUMNS
    if column.startswith("Figure") and "(Origin)" in column
) + (
    RVT_ORIGIN_COLUMN,
    RVT_RESIDUAL_ORIGIN_COLUMN,
    VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN,
    VSM_HYSTERESIS_ORIGIN_COLUMN,
    DMA_ISOSTRESS_ORIGIN_COLUMN,
    MINI_DMA_ORIGIN_COLUMN,
    SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN,
    FMR_ORIGIN_COLUMN,
)

WORD_MICROWIRE_DATA_COLUMNS = (
    "Composition",
    "Microwire",
    "e/a",
    "Strain (%)",
    "Stress (MPa)",
    "Load (g)",
    "Stress/strain current (mA)",
    "Stress/strain current density (A/mm^2)",
    "Fracture strain (%)",
    "Fracture stress (MPa)",
    "Fracture load (g)",
    "Fracture stress/strain current (mA)",
    "Fracture stress/strain current density (A/mm^2)",
    DIAMETER_COLUMN,
    GLASS_DIAMETER_COLUMN,
    DIAMETER_RATIO_COLUMN,
    TRANSITION_TEMP_AS_COLUMN,
    TRANSITION_TEMP_AF_COLUMN,
    TRANSITION_TEMP_MS_COLUMN,
    TRANSITION_TEMP_MF_COLUMN,
    "As1 (mA)",
    "Af1 (mA)",
    "Ms1 (mA)",
    "Mf1 (mA)",
    "As2 (mA)",
    "Af2 (mA)",
    "Ms2 (mA)",
    "Mf2 (mA)",
    "Mf2-Mf1 (mA)",
    "As2-As1 (mA)",
    "Mf1-Af1 (mA)",
    "Af2-Af1 (mA)",
    "Mf2-Af2 (mA)",
    "Ms2-Ms1 (mA)",
    "As current density (A/mm^2)",
    "Ms current density (A/mm^2)",
    "Length (m)",
    "Production datetime",
    "Mass (g)",
    "Resistance (Ω)",
    CORE_TEMPERATURE_COLUMN,
    GLASS_TEMPERATURE_COLUMN,
    "Winding speed (m/min)",
    "Glass feeding (mm/min)",
    ESTIMATED_TRANSITION_COLUMN,
    "Underpressure",
    GLASS_PULL_COLUMN,
    BRITTLE_COLUMN,
)

MICROWIRE_LABEL_RE = re.compile(r"(\d+)\s*[/\-]\s*(\d+)")
MICROWIRE_SORT_RE = re.compile(
    r"^\s*(\d+)\s*[/\-]\s*(\d+)\s*([A-Za-z][A-Za-z0-9]*)?\s*$"
)
MICROWIRE_TOKEN_RE = re.compile(
    r"(\d+)\s*[/\-_]\s*(\d+)([A-Za-z][A-Za-z0-9]*)?"
)

_INVALID_FILENAME_CHARS = set('<>:"/\\|?*')

DRAW_NUMERIC_FIELDS = {
    "mass_g",
    "fabrication_resistance_ohm",
    "winding_speed_m_per_min",
    "glass_feed_mm_per_min",
    "underpressure",
}
PIECE_NUMERIC_FIELDS = {
    "piece_turns",
    "length_m",
    "d_um",
    "D_um",
    "d_over_D",
}
DIMENSION_FIELDS = {"d_um", "D_um", "d_over_D"}
DATETIME_FIELDS = {"production_datetime"}
DATE_FIELDS = {"piece_date"}
RAW_VALUE_FIELDS = (
    DRAW_NUMERIC_FIELDS
    | PIECE_NUMERIC_FIELDS
    | DATETIME_FIELDS
    | DATE_FIELDS
    | {"fabrication_temperature_c"}
)

HEADER_HINTS: Dict[str, str] = {
    "Dtum": "piece_date",
    "dtum": "piece_date",
    "Poet otok": "piece_turns",
    "P.": "piece_y",
    "zloÅ¾enie": "composition_label",
    "zlozenie": "composition_label",
    "composition": "composition_label",
    "dÃ¡tum a Äas vÃ½roby": "production_datetime",
    "datum a cas vyroby": "production_datetime",
    "datumacasvyroby": "production_datetime",
    "hmotnosÅ¥": "mass_g",
    "hmotnost": "mass_g",
    "mass": "mass_g",
    "odpor": "fabrication_resistance_ohm",
    "resistance": "fabrication_resistance_ohm",
    "teplota": "fabrication_temperature_c",
    "temperature": "fabrication_temperature_c",
    "winding speed (m/min)": "winding_speed_m_per_min",
    "glass feeding (mm/min)": "glass_feed_mm_per_min",
    "underpressure": "underpressure",
    "glass pull-off": "glass_pull_off",
    "glass pull off": "glass_pull_off",
    "glass pull-away": "glass_pull_off",
    "glass pull away": "glass_pull_off",
    "bistabilny/nebistabilny": "bistable_status",
    "poznÃ¡mka": "notes",
    "PoznÃ¡mka": "notes",
    "poznamka": "notes",
    "Poznamka": "notes",
    "pozn.": "notes",
    "pozn": "notes",
    "poznÃ¡mky": "notes",
    "p.Ä": "piece_y",
    "p.c": "piece_y",
    "p.Ä.": "piece_y",
    "poÄet otÃ¡Äok": "piece_turns",
    "pocet otacok": "piece_turns",
    "dÄºÅ¾ka (m)": "length_m",
    "dlzka (m)": "length_m",
    "d (Âµm)": "d_um",
    "d (um)": "d_um",
    "d (Î¼m)": "d_um",
    "d (μm)": "d_um",
    "d (�m)": "d_um",
    "d (µm)": "d_um",
    "d(Âµm)": "d_um",
    "d(um)": "d_um",
    "d(Î¼m)": "d_um",
    "d": "d_um",
    "D (Âµm)": "D_um",
    "D (um)": "D_um",
    "D (Î¼m)": "D_um",
    "D (μm)": "D_um",
    "D (�m)": "D_um",
    "D (µm)": "D_um",
    "D(Âµm)": "D_um",
    "D(um)": "D_um",
    "D(Î¼m)": "D_um",
    "D": "D_um",
    "d/D": "d_over_D",
    "d/d": "d_over_D",
    "Datum": "piece_date",
    "DÃ¡tum": "piece_date",
    "datum": "piece_date",
    "dÃ¡tum": "piece_date",
}

ANNEALING_COLUMNS = ["I_A", "V_V", "R_ohm"]

COMPOSITION_TOKEN_PATTERN = re.compile(r"^(?P<composition>(?:[A-Z][a-z]?\d+)+)")
DRAW_PATTERN = re.compile(r"^(?P<draw>\d+)")
PIECE_PATTERN = re.compile(r"^(?P<piece>\d+)")
XY_PATTERN = re.compile(r"(\d+)_+(\d+)")
DRAW_PIECE_AFTER_COMPOSITION_PATTERN = re.compile(
    r"(?:^|[\s_-])(?P<draw>\d{1,3})[_-](?P<piece>\d{1,3})(?=$|[\s_-])"
)
MICROSCOPE_PAIR_PATTERNS: Tuple[re.Pattern[str], ...] = (
    re.compile(r"(\d+)_+(\d+)"),
    re.compile(r"(\d+)[/-](\d+)")
)
MICROSCOPE_WHITESPACE_PAIR = re.compile(r"\b(\d{1,3})\s+(\d{1,3})\b")
MICROWIRE_SHORTCUT_STOP_NAMES = {".shortcut-targets-by-id"}
SETPOINT_PATTERN = re.compile(r"(\d{1,4})mA", re.IGNORECASE)
ALT_VARIANT_PATTERN = re.compile(r"(?:s\d+|\d+_\d+)a(?!\w)", re.IGNORECASE)
DOT_DATE_PATTERN = re.compile(r"\d{1,2}\.\d{1,2}\.\d{2,4}")
SLASH_DATE_PATTERN = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")

MICROSCOPE_MARKER_PATTERN = re.compile(r"[\[{(]\s*(?P<digit>[12Il])\s*[\]})1Il]?")
MICROSCOPE_PRIMARY_HINT = re.compile(r"(\[\s*[1Il]|[1Il]\])$")
MICROSCOPE_PRIMARY_PATTERN = re.compile(
    rf"\[1]\s*(?P<value>\d+(?:[.,]\d+)?)\s*(?:u?m|{MICRO_SIGN}m)",
    re.IGNORECASE,
)
MICROSCOPE_VALUE_PATTERN = re.compile(
    rf"(?P<value>\d+(?:[.,]\d+)?)\s*(?:u?m|{MICRO_SIGN}m)",
    re.IGNORECASE,
)
MICROSCOPE_SECONDARY_PREFIX = re.compile(r"\[2]\s*$", re.IGNORECASE)
MICROSCOPE_NUMBER_TOKEN = re.compile(r"^\d+(?:[.,]\d+)?$")
MICROSCOPE_UNIT_HINTS = ("Âµm", "um", "Î¼m")

KNOWN_TIMEZONE_TOKENS = {
    "UTC",
    "GMT",
    "CET",
    "CEST",
    "EET",
    "EEST",
    "BST",
    "IST",
    "WEST",
    "WET",
    "EST",
    "EDT",
    "CST",
    "CDT",
    "MST",
    "MDT",
    "PST",
    "PDT",
}

MICROSCOPE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp")


DEFAULT_FIGSIZE: Tuple[float, float] = (10.0, 6.0)


def _normalise_figsize(value: Sequence[float] | Tuple[float, float] | None) -> Tuple[float, float]:
    default = DEFAULT_FIGSIZE
    if not value:
        return default
    try:
        width, height = value  # type: ignore[misc]
    except Exception:
        return default
    try:
        width_f = float(width)
        height_f = float(height)
    except Exception:
        return default
    if not math.isfinite(width_f) or width_f <= 0:
        width_f = default[0]
    if not math.isfinite(height_f) or height_f <= 0:
        height_f = default[1]
    return (max(0.5, width_f), max(0.5, height_f))

EXCEL_EMBED_DPI = 96.0
EMU_PER_INCH = 914400


def _normalise_dpi(value: object) -> Tuple[float, float]:
    """Return ``(x, y)`` DPI extracted from ``value``.

    ``value`` may be a number, a 2-tuple, or ``None``.  When a component is
    missing or non-finite the Excel embed DPI is used as a sensible default.
    """

    default = EXCEL_EMBED_DPI
    if isinstance(value, (tuple, list)) and value:
        try:
            dpi_x = float(value[0])
        except Exception:
            dpi_x = default
        try:
            dpi_y = float(value[1]) if len(value) > 1 else float(value[0])
        except Exception:
            dpi_y = default
    elif isinstance(value, (int, float)):
        try:
            dpi_x = dpi_y = float(value)
        except Exception:
            dpi_x = dpi_y = default
    else:
        dpi_x = dpi_y = default

    if not math.isfinite(dpi_x) or dpi_x <= 0:
        dpi_x = default
    if not math.isfinite(dpi_y) or dpi_y <= 0:
        dpi_y = default
    return dpi_x, dpi_y


def _image_metrics(image_path: Path) -> Tuple[int, int, float, float]:
    """Return pixel dimensions and DPI metadata for ``image_path``."""

    try:
        from PIL import Image as PILImage
    except Exception:
        return 0, 0, EXCEL_EMBED_DPI, EXCEL_EMBED_DPI

    try:
        with PILImage.open(image_path) as pil_image:
            width_px, height_px = pil_image.size
            dpi_x, dpi_y = _normalise_dpi(pil_image.info.get("dpi"))
    except Exception:
        return 0, 0, EXCEL_EMBED_DPI, EXCEL_EMBED_DPI

    return int(width_px or 0), int(height_px or 0), dpi_x, dpi_y


def _excel_pixel_limits(figure_size: Tuple[float, float]) -> Tuple[int, int]:
    width_in, height_in = figure_size
    width_px = max(int(round(width_in * EXCEL_EMBED_DPI)), 1)
    height_px = max(int(round(height_in * EXCEL_EMBED_DPI)), 1)
    return width_px, height_px
def _excel_row_height(height_in: float) -> float:
    return max(height_in * 72.0, 18.0)


def _column_width_from_pixels(pixels: float) -> float:
    if pixels <= 0:
        return 0.0
    if pixels <= 12.0:
        return pixels / 12.0
    return (pixels - 5.0) / 7.0


def _excel_column_width(width_in: float) -> float:
    pixels = max(width_in * EXCEL_EMBED_DPI, 1.0)
    return max(_column_width_from_pixels(pixels), 8.43)


def _adjust_drawing_ext_dimensions(
    excel_path: Path,
    figure_size: Tuple[float, float],
    log: logging.Logger,
) -> None:
    """Update the drawing CX/CY extents so the width matches the requested figure size."""

    if not excel_path.exists():
        return
    width_in, height_in = figure_size
    width_emu = int(round(width_in * EMU_PER_INCH))
    height_emu = int(round(height_in * EMU_PER_INCH))
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    temp_file.close()
    temp_path = Path(temp_file.name)
    try:
        with ZipFile(excel_path, "r") as source, ZipFile(temp_path, "w") as target:
            for info in source.infolist():
                data = source.read(info.filename)
                if info.filename.startswith("xl/drawings/") and info.filename.endswith(".xml"):
                    try:
                        tree = ET.fromstring(data)
                    except ET.ParseError:
                        pass
                    else:
                        updated = False
                        for namespace in (
                            "http://schemas.openxmlformats.org/drawingml/2006/main",
                            "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                        ):
                            for ext in tree.findall(f".//{{{namespace}}}ext"):
                                ext.set("cx", str(width_emu))
                                ext.set("cy", str(height_emu))
                                updated = True
                        if updated:
                            data = ET.tostring(tree, encoding="utf-8")
                target.writestr(info, data)
        os.replace(temp_path, excel_path)
    except Exception as exc:
        log.warning("Failed to adjust drawing metadata for %s: %s", excel_path, exc)
        try:
            temp_path.unlink()
        except Exception:
            pass


@dataclass
class BuilderConfig:
    """Configuration for the database builder."""

    fabrication_files: List[Path]
    annealing_files: List[Path]
    output_dir: Path
    microscope_files: List[Path] = field(default_factory=list)
    video_files: List[Path] = field(default_factory=list)
    strain_files: List[Path] = field(default_factory=list)
    make_plots: bool = False
    export_formats: Tuple[str, ...] = ("csv",)
    plot_dir_name: str = PLOT_DIR_NAME
    origin_dir_name: str = ORIGIN_DIR_NAME
    output_name: str = DEFAULT_OUTPUT_NAME
    plot_backends: Tuple[str, ...] = ()
    export_behaviour: Optional[Dict[str, str]] = None
    matplotlib_figsize: Tuple[float, float] = DEFAULT_FIGSIZE
    include_microscope_crops: bool = True
    highlight_ocr_values: bool = True
    column_filter: Optional[Tuple[str, ...]] = None
    column_order: Optional[Tuple[str, ...]] = None
    sort_spec: Optional[Tuple[Tuple[str, bool], ...]] = None


@dataclass
class BuildStats:
    """Accumulates processing statistics."""

    parsed: int = 0
    skipped: int = 0
    missing_draw: int = 0
    missing_piece: int = 0
    resistance_checks_failed: int = 0
    rows_built: int = 0
    missing_high_measurement: int = 0
    missing_low_measurement: int = 0


@dataclass
class OriginArtifact:
    """Metadata describing an Origin object exported for Excel embedding."""

    descriptor: str
    object_path: Optional[Path]
    graph_name: Optional[str] = None
    workbook_name: Optional[str] = None
    worksheet_name: Optional[str] = None
    display_text: Optional[str] = None
    clipboard_fallback: bool = False


@dataclass(frozen=True)
class WordOleInsertion:
    """Origin object insertion request for a generated Word report."""

    bookmark_name: str
    object_path: Path
    label: str
    clipboard_fallback: bool = False
    graph_name: Optional[str] = None
    descriptor: Optional[str] = None


@dataclass(frozen=True)
class WordOleEmbeddingResult:
    """Auditable result for one attempted Word Origin OLE insertion."""

    bookmark_name: str
    descriptor: str
    label: str
    object_path: str
    attempted: bool
    inserted: bool
    status: str
    reason: str = ""


@dataclass(frozen=True)
class WordPictureInsertion:
    """Image insertion request for a generated Word report."""

    bookmark_name: str
    image_path: Path
    label: str


@dataclass
class BuildResult:
    """Return value from :func:`build_database`."""

    dataframe: pd.DataFrame
    exports: Dict[str, Path]
    plot_paths: List[str]
    origin_artifacts: Dict[str, OriginArtifact]
    stats: BuildStats
    microscope_crops: Dict[str, Path] = field(default_factory=dict)
    ocr_highlights: Dict[str, Set[int]] = field(default_factory=dict)
    word_reports: List[Path] = field(default_factory=list)


@dataclass
class StrainRecord:
    """Single strain measurement parsed from the worksheet."""

    composition: str
    draw: Optional[int]
    piece: Optional[int]
    microwire_label: str
    m_length: Optional[float]
    a_length: Optional[float]
    percent: Optional[float]
    broke: bool
    source: Path


@dataclass
class VsmHysteresisRecord:
    """Parsed VSM hysteresis measurement for a single file."""

    path: Path
    sample: str
    data: pd.DataFrame
    temperature: Optional[float] = None
    angle: Optional[float] = None
    key: Optional[Tuple[str, int, int]] = None
    label: Optional[str] = None


@dataclass
class VsmTemperatureScanRecord:
    """Parsed VSM temperature scan measurement for a single file."""

    path: Path
    sample: str
    data: pd.DataFrame
    key: Optional[Tuple[str, int, int]] = None
    label: Optional[str] = None


@dataclass
class DmaIsoStressRecord:
    """Parsed DMA iso-stress measurement for a single file."""

    path: Path
    sample: str
    datasets: Dict[int, Tuple[List[float], List[float]]]
    key: Optional[Tuple[str, int, int]] = None
    label: Optional[str] = None


@dataclass
class MiniDmaRecord:
    """Parsed Mini DMA run for a single measurement folder."""

    path: Path
    sample: str
    data: pd.DataFrame
    key: Optional[Tuple[str, int, int]] = None
    label: Optional[str] = None
    strain_summary: Tuple[str, ...] = ()
    transition_summary: Tuple[str, ...] = ()
    break_summary: str = ""


@dataclass
class ShapeMemoryStressStrainRecord:
    """Parsed shape-memory stress/strain measurement for a single file."""

    path: Path
    sample: str
    data: pd.DataFrame
    key: Optional[Tuple[str, int, int]] = None
    label: Optional[str] = None


@dataclass
class FmrRecord:
    """Parsed FMR measurement for a single file."""

    path: Path
    sample: str
    data: pd.DataFrame
    key: Optional[Tuple[str, int, int]] = None
    label: Optional[str] = None


class FabricationIndex:
    """Lookup tables populated from fabrication spreadsheets."""

    def __init__(self) -> None:
        self.draw_level: Dict[Tuple[str, int], Dict[str, object]] = {}
        self.piece_level: Dict[Tuple[str, int, int], Dict[str, object]] = {}

    def set_draw(self, composition: str, draw_x: int, data: Dict[str, object]) -> None:
        key = (composition, draw_x)
        existing = self.draw_level.get(key, {})
        for field, value in data.items():
            if field.endswith("__display") and isinstance(value, (list, tuple)):
                merged: List[object] = []
                for source in (existing.get(field), value):
                    if isinstance(source, (list, tuple)):
                        for item in source:
                            if item not in merged:
                                merged.append(item)
                existing[field] = merged
            else:
                if _has_meaningful_value(value) or field not in existing:
                    existing[field] = value
        self.draw_level[key] = existing

    def set_piece(self, composition: str, draw_x: int, piece_y: int, data: Dict[str, object]) -> None:
        key = (composition, draw_x, piece_y)
        existing = self.piece_level.get(key, {})
        for field, value in data.items():
            if field.endswith("__display") and isinstance(value, (list, tuple)):
                merged: List[object] = []
                for source in (existing.get(field), value):
                    if isinstance(source, (list, tuple)):
                        for item in source:
                            if item not in merged:
                                merged.append(item)
                existing[field] = merged
            else:
                if _has_meaningful_value(value) or field not in existing:
                    existing[field] = value
        self.piece_level[key] = existing

    def get_draw(self, composition: str, draw_x: Optional[int]) -> Dict[str, object]:
        if draw_x is None:
            return {}
        return self.draw_level.get((composition, draw_x), {})

    def get_piece(self, composition: str, draw_x: Optional[int], piece_y: Optional[int]) -> Dict[str, object]:
        if draw_x is None or piece_y is None:
            return {}
        return self.piece_level.get((composition, draw_x, piece_y), {})


@dataclass
class MicroscopeDetection:
    """OCR evidence extracted from a microscope image."""

    value: float
    image_path: Optional[Path]
    bbox: Optional[Tuple[int, int, int, int]] = None  # (left, top, right, bottom)
    text: Optional[str] = None
    source: str = "ocr"
    confidence: Optional[float] = None
    marker: Optional[int] = None
    crop_path: Optional[Path] = None
    category: Optional[str] = None

    def matches(self, value: float, *, tol: float = 0.25) -> bool:
        try:
            delta = abs(float(self.value) - float(value))
        except Exception:
            return False
        return delta <= tol

    def ensure_crop(self, output_dir: Path, prefix: str) -> Optional[Path]:
        """Persist a cropped preview of the detection region."""

        if self.crop_path is not None and self.crop_path.exists():
            return self.crop_path
        if self.image_path is None:
            return None
        try:
            from PIL import Image  # type: ignore[import-not-found]
        except ImportError:
            return None

        crop_bbox = self.bbox
        try:
            with Image.open(self.image_path) as img:
                width = img.width
                height = img.height
                if crop_bbox is not None:
                    left, top, right, bottom = crop_bbox
                    pad = int(round(max(right - left, bottom - top) * 0.25))
                    if pad > 0:
                        left = max(left - pad, 0)
                        top = max(top - pad, 0)
                        right = min(right + pad, width)
                        bottom = min(bottom + pad, height)
                    if right <= left or bottom <= top:
                        crop_bbox = None
                if crop_bbox is None:
                    left, top, right, bottom = 0, 0, width, height
                output_dir.mkdir(parents=True, exist_ok=True)
                safe_prefix = _safe_plot_stem(prefix)
                stem = f"{safe_prefix}_{self.value:.2f}"
                candidate = output_dir / f"{stem}.png"
                counter = 1
                while candidate.exists():
                    candidate = output_dir / f"{stem}_{counter}.png"
                    counter += 1
                region = img.crop((left, top, right, bottom))
                region.save(candidate)
                self.crop_path = candidate
                return candidate
        except Exception:
            return None
        return None


@dataclass
class MicroscopeOCRResult:
    """Container of raw OCR candidates extracted from an image."""

    values: List[float] = field(default_factory=list)
    detections: List[MicroscopeDetection] = field(default_factory=list)
    texts: List[str] = field(default_factory=list)

    def __iter__(self) -> Iterator[float]:  # pragma: no cover - convenience shim
        return iter(self.values)

    def append_value(self, value: float) -> None:
        if not isinstance(value, (int, float)):
            return
        numeric = float(value)
        if not math.isfinite(numeric) or numeric <= 0:
            return
        self.values.append(numeric)


@dataclass
class MicroscopeCacheEntry:
    """Serialized OCR output for a single microscope capture."""

    path: str
    mtime: float
    size: int
    values: List[float] = field(default_factory=list)
    detections: List[Dict[str, Any]] = field(default_factory=list)
    texts: List[str] = field(default_factory=list)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "path": self.path,
            "mtime": float(self.mtime),
            "size": int(self.size),
            "values": list(self.values),
            "detections": list(self.detections),
            "texts": list(self.texts),
        }

    @classmethod
    def from_dict(cls, payload: Mapping[str, Any]) -> Optional["MicroscopeCacheEntry"]:
        if not isinstance(payload, Mapping):
            return None
        path = str(payload.get("path") or "")
        try:
            mtime = float(payload.get("mtime", 0.0))
        except (TypeError, ValueError):
            mtime = 0.0
        try:
            size = int(payload.get("size", 0))
        except (TypeError, ValueError):
            size = 0
        values_raw = payload.get("values") or []
        detections_raw = payload.get("detections") or []
        texts_raw = payload.get("texts") or []
        values: List[float] = []
        for entry in values_raw:
            try:
                numeric = float(entry)
            except (TypeError, ValueError):
                continue
            values.append(numeric)
        detections: List[Dict[str, Any]] = []
        for entry in detections_raw:
            if isinstance(entry, Mapping):
                detections.append(dict(entry))
        texts: List[str] = []
        for entry in texts_raw:
            text = str(entry)
            if text:
                texts.append(text)
        return cls(path=path, mtime=mtime, size=size, values=values, detections=detections, texts=texts)

    @classmethod
    def from_result(
        cls,
        path: Path,
        mtime: float,
        size: int,
        result: MicroscopeOCRResult,
    ) -> "MicroscopeCacheEntry":
        detections_payload: List[Dict[str, Any]] = []
        for detection in result.detections:
            payload: Dict[str, Any] = {
                "value": float(detection.value),
                "bbox": list(detection.bbox) if detection.bbox else None,
                "text": detection.text,
                "source": detection.source,
                "confidence": detection.confidence,
                "marker": detection.marker,
                "category": detection.category,
                "image_path": str(detection.image_path) if detection.image_path else None,
                "crop_path": str(detection.crop_path) if detection.crop_path else None,
            }
            detections_payload.append(payload)
        return cls(
            path=str(path),
            mtime=float(mtime),
            size=int(size),
            values=list(result.values),
            detections=detections_payload,
            texts=list(result.texts),
        )

    def is_current(self, path: Path) -> bool:
        try:
            stat_result = path.stat()
        except OSError:
            return False
        return (
            math.isclose(float(stat_result.st_mtime), float(self.mtime), rel_tol=0.0, abs_tol=1e-6)
            and int(stat_result.st_size) == int(self.size)
        )

    def to_result(self, path: Path) -> MicroscopeOCRResult:
        detections: List[MicroscopeDetection] = []
        for entry in self.detections:
            value = entry.get("value")
            if not isinstance(value, (int, float)):
                continue
            bbox_source = entry.get("bbox")
            bbox: Optional[Tuple[int, int, int, int]] = None
            if isinstance(bbox_source, (list, tuple)) and len(bbox_source) == 4:
                try:
                    bbox = tuple(int(round(float(coord))) for coord in bbox_source)
                except (TypeError, ValueError):
                    bbox = None
            image_token = entry.get("image_path")
            if image_token:
                try:
                    image_path = Path(str(image_token))
                except Exception:
                    image_path = path
            else:
                image_path = path
            crop_token = entry.get("crop_path")
            crop_path: Optional[Path]
            if crop_token:
                try:
                    crop_candidate = Path(str(crop_token))
                except Exception:
                    crop_candidate = None
                crop_path = crop_candidate
            else:
                crop_path = None
            detection = MicroscopeDetection(
                value=float(value),
                image_path=image_path,
                bbox=bbox,
                text=entry.get("text"),
                source=str(entry.get("source") or "ocr"),
                confidence=float(entry["confidence"]) if isinstance(entry.get("confidence"), (int, float)) else None,
                marker=int(entry["marker"]) if isinstance(entry.get("marker"), (int, float)) else None,
                crop_path=crop_path,
            )
            category = entry.get("category")
            if category is not None:
                detection.category = str(category)
            detections.append(detection)
        return MicroscopeOCRResult(
            values=list(self.values),
            detections=detections,
            texts=list(self.texts),
        )


@dataclass
class MicroscopeMeasurements:
    """Diameter samples gathered from microscope images."""

    core: List[MicroscopeDetection] = field(default_factory=list)
    glass: List[MicroscopeDetection] = field(default_factory=list)
    other: List[MicroscopeDetection] = field(default_factory=list)
    brittle: bool = False

    def _target(self, category: str) -> List[MicroscopeDetection]:
        if category == "core":
            return self.core
        if category == "glass":
            return self.glass
        return self.other

    def add_placeholder(self, category: str, image_path: Path) -> None:
        """Ensure at least one detection entry exists for the given image.

        When OCR fails to extract a numeric value we still want the
        microscope worksheet to list the image so the operator can review it
        manually.  A placeholder detection keeps track of the originating
        file without affecting downstream ratio calculations.
        """

        if not image_path:
            return
        target = self._target(category)
        try:
            image_path = Path(image_path)
        except Exception:
            return
        for detection in target:
            existing = getattr(detection, "image_path", None)
            if existing is None:
                continue
            try:
                if Path(existing) == image_path:
                    return
            except Exception:
                continue
        placeholder = MicroscopeDetection(
            value=float("nan"),
            image_path=image_path,
            source="placeholder",
        )
        placeholder.category = category
        target.append(placeholder)

    def extend(
        self,
        category: str,
        values: Iterable[float],
        detections: Iterable[MicroscopeDetection] | None = None,
    ) -> None:
        target = self._target(category)
        evidence = list(detections or [])
        for detection in evidence:
            detection.category = category
        source = "manual"
        if detections is not None:
            source = "ocr"
        for value in values:
            if not isinstance(value, (int, float)):
                continue
            numeric = float(value)
            if not math.isfinite(numeric) or numeric <= 0:
                continue
            match = None
            for candidate in list(evidence):
                if candidate.matches(numeric):
                    match = candidate
                    evidence.remove(candidate)
                    break
            if match is None:
                match = MicroscopeDetection(
                    value=numeric,
                    image_path=None,
                    source=source,
                )
                match.category = category
            target.append(match)

    def _best_detection_sequence(
        self, categories: Sequence[str], prefer: str
    ) -> Optional[MicroscopeDetection]:
        for category in categories:
            target = self._target(category)
            if not target:
                continue
            value = _select_microscope_value([entry.value for entry in target], prefer)
            if value is None:
                continue
            chosen = self._match_detection(target, value)
            if chosen is not None:
                return chosen
        return None

    @staticmethod
    def _match_detection(
        candidates: Sequence[MicroscopeDetection], value: float
    ) -> Optional[MicroscopeDetection]:
        best: MicroscopeDetection | None = None
        best_delta: float | None = None
        for detection in candidates:
            try:
                delta = abs(float(detection.value) - float(value))
            except Exception:
                continue
            if best is None or (best_delta is not None and delta < best_delta) or best_delta is None:
                best = detection
                best_delta = delta
        return best

    def best_core(self) -> Optional[float]:
        detection = self.best_core_detection()
        return detection.value if detection is not None else None

    def best_core_detection(self) -> Optional[MicroscopeDetection]:
        return self._best_detection_sequence(("core", "other", "glass"), "min")

    def best_glass_detection(self) -> Optional[MicroscopeDetection]:
        return self._best_detection_sequence(("glass", "other", "core"), "max")

    def best_glass(self) -> Optional[float]:
        detection = self.best_glass_detection()
        return detection.value if detection is not None else None


@dataclass
class VideoMetricsSummary:
    """Aggregated metrics extracted from fabrication videos."""

    temperatures: List[float] = field(default_factory=list)
    underpressures: List[float] = field(default_factory=list)
    winding_speeds: List[float] = field(default_factory=list)
    glass_feeds: List[float] = field(default_factory=list)
    sources: set[Path] = field(default_factory=set)

    def record(self, result, *, source: Optional[Path] = None) -> None:
        temp = getattr(result, "median_temperature", None)
        if callable(temp):
            value = temp()
        else:
            value = None
        if value is not None and isinstance(value, (int, float)) and math.isfinite(float(value)):
            self.temperatures.append(float(value))
        under_fn = getattr(result, "median_underpressure", None)
        if callable(under_fn):
            under_value = under_fn()
        else:
            under_value = None
        if under_value is not None and isinstance(under_value, (int, float)) and math.isfinite(float(under_value)):
            self.underpressures.append(float(under_value))
        speed_fn = getattr(result, "median_winding_speed", None)
        if callable(speed_fn):
            speed_value = speed_fn()
        else:
            speed_value = None
        if speed_value is not None and isinstance(speed_value, (int, float)) and math.isfinite(float(speed_value)):
            self.winding_speeds.append(float(speed_value))
        feed_fn = getattr(result, "median_glass_feed", None)
        if callable(feed_fn):
            feed_value = feed_fn()
        else:
            feed_value = None
        if feed_value is not None and isinstance(feed_value, (int, float)) and math.isfinite(float(feed_value)):
            self.glass_feeds.append(float(feed_value))
        if source is not None:
            try:
                self.sources.add(Path(source))
            except Exception:
                pass

    def temperature(self) -> Optional[float]:
        return _select_microscope_value(self.temperatures, "median", allow_negative=True)

    def underpressure(self) -> Optional[float]:
        return _select_microscope_value(self.underpressures, "median", allow_negative=True)

    def winding_speed(self) -> Optional[float]:
        return _select_microscope_value(self.winding_speeds, "median")

    def glass_feed(self) -> Optional[float]:
        return _select_microscope_value(self.glass_feeds, "median")


def _logger(logger: Optional[logging.Logger]) -> logging.Logger:
    if logger is not None:
        return logger
    return logging.getLogger(LOGGER_NAME)


def _normalise_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("Âµ", "µ").replace("�", "µ").replace("ï¿½", "µ")
    return text


def _is_unit_suffix(text: str) -> bool:
    """Return True when *text* only contains unit markers such as µm."""

    if not text:
        return False
    stripped = unicodedata.normalize("NFKC", text).strip()
    if not stripped:
        return False
    ascii_text = unicodedata.normalize("NFKD", stripped)
    ascii_text = "".join(ch for ch in ascii_text if not unicodedata.combining(ch))
    simplified = re.sub(r"\s+", "", ascii_text).lower()
    canonical_units = {"um", "µm", "μm"}
    if simplified in canonical_units | {"(um)", "(µm)", "(μm)", "um)", "µm)", "μm)"}:
        return True
    unit_chars = set("uµμm()/.-[]{}")
    if simplified and all(ch in unit_chars for ch in simplified):
        return True
    return False

def _merged_header_row(
    df: pd.DataFrame, header_idx: int, *, lookback: int = 3
) -> List[object]:
    """Return a header row with empty cells backfilled from previous rows."""

    if header_idx < 0 or header_idx >= len(df):
        return []

    header = list(df.iloc[header_idx])
    if header_idx == 0:
        return header

    max_offset = min(max(header_idx, 0), max(0, lookback))
    if max_offset <= 0:
        return header

    for col_idx in range(len(header)):
        value = header[col_idx]
        text = _normalise_text(value) if value is not None and not _is_nan(value) else ""
        if text and _is_unit_suffix(text):
            prefix: Optional[str] = None
            for offset in range(1, max_offset + 1):
                prev_idx = header_idx - offset
                if prev_idx < 0:
                    break
                try:
                    candidate = df.iat[prev_idx, col_idx]
                except (IndexError, ValueError):
                    candidate = None
                if candidate is None or _is_nan(candidate):
                    continue
                candidate_text = _normalise_text(candidate)
                if not candidate_text or _is_unit_suffix(candidate_text):
                    continue
                prefix = candidate_text
                break
            if prefix:
                combined = f"{prefix} {text}".strip()
                header[col_idx] = combined
                continue
        if text:
            continue
        for offset in range(1, max_offset + 1):
            try:
                candidate = df.iat[header_idx - offset, col_idx]
            except (IndexError, ValueError):
                candidate = None
            if candidate is None or _is_nan(candidate):
                continue
            candidate_text = _normalise_text(candidate)
            if candidate_text:
                header[col_idx] = candidate_text
                break
    return header


def _header_key(value: object) -> Optional[str]:
    text = _normalise_text(value)
    if not text:
        return None
    hint = HEADER_HINTS.get(text)
    if hint:
        return hint
    if text in {"d (�m)", "D (�m)"}:
        text = text.replace("�", "µ")
        hint = HEADER_HINTS.get(text)
        if hint:
            return hint
    ascii_text = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(ch for ch in ascii_text if not unicodedata.combining(ch))
    ascii_hint = HEADER_HINTS.get(ascii_text)
    if ascii_hint:
        return ascii_hint
    ascii_hint = HEADER_HINTS.get(ascii_text.lower())
    if ascii_hint:
        return ascii_hint
    lowered = ascii_text.lower()
    simple = lowered.replace("\u00a0", " ")
    simple = re.sub(r"\s+", " ", simple)
    simple_compact = re.sub(r"[^a-z0-9]+", "", lowered)
    if "zlozen" in lowered:
        return "composition_label"
    if "datum" in lowered and "cas" in lowered:
        return "production_datetime"
    if "hmotn" in lowered or "mass" in lowered:
        return "mass_g"
    if "odpor" in lowered or "resistance" in lowered:
        return "fabrication_resistance_ohm"
    if "teplot" in lowered or "temp" in lowered:
        return "fabrication_temperature_c"
    if "winding" in lowered:
        return "winding_speed_m_per_min"
    if "glass" in lowered and ("feed" in lowered or "feeding" in lowered):
        return "glass_feed_mm_per_min"
    if "glass" in lowered and "pull" in lowered:
        return "glass_pull_off"
    if "underpressure" in lowered:
        return "underpressure"
    if "bistabil" in lowered:
        return "bistable_status"
    if "dtum" in lowered and "cas" not in lowered:
        return "piece_date"
    if "p.c" in lowered or "p.Ä" in lowered or simple_compact == "pc":
        return "piece_y"
    if "pocet" in lowered and "otac" in lowered:
        return "piece_turns"
    if "poet" in lowered and "otok" in lowered:
        return "piece_turns"
    if "dlzk" in lowered or "dlÅ¾k" in lowered:
        return "length_m"
    micron_hint = any(token in lowered for token in ("Âµm", "µm", "um", "Î¼m", "mikro", "micro"))
    ascii_simple = ascii_text.replace("\u00a0", " ")
    if micron_hint:
        if re.search(r"\bD\d*\b", ascii_text):
            return "D_um"
        if re.search(r"\bd\d*\b", ascii_text):
            return "d_um"
        if "glass" in lowered or "sklo" in lowered:
            return "D_um"
        if any(token in lowered for token in ("jadro", "jadra", "jadier", "core")):
            return "d_um"
        if re.search(r"\bd\s*/\s*D\b", ascii_simple, re.IGNORECASE):
            return "d_over_D"
    if lowered.strip().startswith("d") and any(token in lowered for token in ("Âµm", "µm")):
        first = ascii_text.strip()[:1]
        if first == "D":
            return "D_um"
        return "d_um"
    if lowered.strip().startswith("d") and "um" in lowered and all(token not in lowered for token in ("Âµ", "µ")):
        first = ascii_text.strip()[:1]
        if first == "D":
            return "D_um"
        return "d_um"
    if micron_hint and "d" in lowered and "d" in ascii_simple and "D" in ascii_simple:
        return "d_over_D"
    if "d/d" in lowered or simple_compact == "dd":
        return "d_over_D"
    if any(token in lowered for token in ("jadro", "jadra", "jadier", "core")) and re.search(r"\bd\d*\b", ascii_text):
        return "d_um"
    if any(token in lowered for token in ("sklo", "skla", "glass", "clad", "cladding", "sheath")) and re.search(r"\bD\d*\b", ascii_text):
        return "D_um"
    if ("datum" in lowered or "dÃ¡tum" in lowered) and "cas" not in lowered:
        return "piece_date"
    return None


def _is_nan(value: object) -> bool:
    if isinstance(value, (float, np.floating)):
        try:
            return math.isnan(float(value))
        except ValueError:
            return True
    if value is None:
        return False
    try:
        if value is pd.NA:  # type: ignore[attr-defined]
            return True
    except AttributeError:
        pass
    try:
        result = pd.isna(value)
    except Exception:
        return False
    if isinstance(result, (bool, np.bool_)):
        return bool(result)
    return False


def _raw_string(value: object) -> Optional[str]:
    if value is None or _is_nan(value):
        return None
    text = _normalise_text(value)
    return text if text else None


def _parse_numeric(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not _is_nan(value):
        return float(value)
    text = _normalise_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    number = match.group(0).replace(",", ".")
    try:
        return float(number)
    except ValueError:
        return None


def _clean_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)) and not _is_nan(value):
        text = f"{value}"
    else:
        text = _normalise_text(value)
    return text.strip() if text else ""


def _has_meaningful_value(value: object) -> bool:
    if value is None:
        return False
    if _is_nan(value):
        return False
    if isinstance(value, str):
        return bool(_clean_str(value))
    return True


def _parse_strain_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not _is_nan(value):
        return float(value)
    text = _clean_str(value)
    if not text:
        return None
    upper = text.upper()
    if "DIV/0" in upper or "INF" in upper or "BROKE" in upper:
        return None
    return _parse_numeric(text)


def _microwire_tuple_from_label(label: str) -> Optional[Tuple[int, int]]:
    match = MICROWIRE_LABEL_RE.search(label)
    if not match:
        return None
    try:
        return int(match.group(1)), int(match.group(2))
    except ValueError:
        return None


def _microwire_parts_from_label(
    label: str,
) -> Optional[Tuple[int, int, Optional[str]]]:
    if not label:
        return None
    cleaned = _clean_str(label)
    if not cleaned:
        return None
    match = MICROWIRE_SORT_RE.match(cleaned)
    if not match:
        match = MICROWIRE_TOKEN_RE.search(cleaned)
    if not match:
        return None
    try:
        draw = int(match.group(1))
        piece = int(match.group(2))
    except (TypeError, ValueError):
        return None
    suffix = match.group(3)
    suffix = suffix.strip() if isinstance(suffix, str) else ""
    return draw, piece, suffix or None


def _split_microwire_key(
    value: object,
) -> Optional[Tuple[str, int, int, Optional[str]]]:
    def _coerce_index(raw: object) -> Optional[int]:
        if isinstance(raw, bool):
            return None
        if isinstance(raw, (int, np.integer)):
            return int(raw)
        if isinstance(raw, (float, np.floating)):
            numeric = float(raw)
            if not math.isfinite(numeric) or not numeric.is_integer():
                return None
            return int(numeric)
        if isinstance(raw, str):
            text = raw.strip()
            if not text:
                return None
            if re.fullmatch(r"[+-]?\d+", text):
                try:
                    return int(text)
                except (TypeError, ValueError):
                    return None
            try:
                numeric = float(text)
            except (TypeError, ValueError):
                return None
            if not math.isfinite(numeric) or not numeric.is_integer():
                return None
            return int(numeric)
        return None

    if isinstance(value, tuple):
        if len(value) == 3:
            composition = str(value[0]).strip()
            draw = _coerce_index(value[1])
            piece = _coerce_index(value[2])
            if draw is None or piece is None:
                return None
            if composition:
                return composition, draw, piece, None
        if len(value) == 4:
            composition = str(value[0]).strip()
            draw = _coerce_index(value[1])
            piece = _coerce_index(value[2])
            if draw is None or piece is None:
                return None
            suffix = value[3]
            suffix_text = str(suffix).strip() if suffix is not None else ""
            if composition:
                return composition, draw, piece, suffix_text or None
    if isinstance(value, str):
        return _microwire_key_from_string(value)
    return None


def _microwire_key_to_str(
    key: Tuple[str, int, int, Optional[str]] | Tuple[str, int, int],
) -> str:
    if len(key) == 3:
        composition, draw, piece = key
        suffix = None
    else:
        composition, draw, piece, suffix = key
    base = f"{composition}|{draw}|{piece}"
    if suffix:
        return f"{base}|{suffix}"
    return base


def _microwire_key_from_string(
    value: str,
) -> Optional[Tuple[str, int, int, Optional[str]]]:
    if not value:
        return None
    parts = [part.strip() for part in value.split("|")]
    if len(parts) not in {3, 4}:
        return None
    composition = parts[0]
    if not composition:
        return None
    parsed = _split_microwire_key((composition, parts[1], parts[2], parts[3] if len(parts) == 4 else None))
    if parsed is None:
        return None
    _composition, draw, piece, suffix = parsed
    if not _composition:
        return None
    return composition, draw, piece, suffix


def _format_strain_value(record: StrainRecord) -> Optional[str]:
    if record.broke:
        return "broke"
    if record.percent is None:
        return None
    return f"{record.percent:.3f}%"


def _sanitize_datetime_text(text: str) -> str:
    cleaned = text.strip()
    cleaned = re.sub(r"(?<=\d)/(?=\d{1,2}(?::\d{1,2})?)", " ", cleaned)
    cleaned = re.sub(r"\s*\([^)]*\)\s*$", "", cleaned)
    tokens: list[str] = []
    for raw_token in cleaned.split():
        token = raw_token.strip().strip(",.;")
        if not token:
            continue
        tokens.append(token)
    while tokens:
        candidate = tokens[-1]
        upper = candidate.upper()
        if any(char.isdigit() for char in candidate):
            break
        if upper in KNOWN_TIMEZONE_TOKENS:
            break
        tokens.pop()
    return " ".join(tokens)


def _parse_datetime(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    text = _normalise_text(value)
    if not text:
        return None
    text = _sanitize_datetime_text(text)
    try:
        dayfirst: Optional[bool]
        if DOT_DATE_PATTERN.search(text):
            dayfirst = True
        elif SLASH_DATE_PATTERN.search(text):
            dayfirst = False
        else:
            dayfirst = None
        if dayfirst is None:
            dt = pd.to_datetime(text, dayfirst=False, errors="coerce")
            if pd.isna(dt):
                dt = pd.to_datetime(text, dayfirst=True, errors="coerce")
        else:
            dt = pd.to_datetime(text, dayfirst=dayfirst, errors="coerce")
    except (TypeError, ValueError):
        return None
    if pd.isna(dt):
        return None
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    if isinstance(dt, datetime):
        return dt.isoformat(sep=" ")
    return None


def _parse_date(value: object) -> Optional[str]:
    parsed = _parse_datetime(value)
    if not parsed:
        return None
    try:
        dt = datetime.fromisoformat(parsed)
    except ValueError:
        return parsed
    return dt.date().isoformat()


def _select_microscope_value(
    values: Sequence[float], prefer: str, *, allow_negative: bool = False
) -> Optional[float]:
    cleaned: List[float] = []
    for v in values:
        if not isinstance(v, (int, float)):
            continue
        value = float(v)
        if not math.isfinite(value):
            continue
        if not allow_negative and value <= 0:
            continue
        cleaned.append(value)
    if not cleaned:
        return None
    if prefer == "min":
        return float(min(cleaned))
    if prefer == "max":
        return float(max(cleaned))
    cleaned.sort()
    mid = len(cleaned) // 2
    if len(cleaned) % 2:
        return float(cleaned[mid])
    return float((cleaned[mid - 1] + cleaned[mid]) / 2.0)


def _extract_field_value(field: str, value: object) -> Tuple[Optional[object], Optional[str]]:
    raw = _raw_string(value)
    if field in DATETIME_FIELDS:
        return _parse_datetime(value), raw
    if field in DATE_FIELDS:
        return _parse_date(value), raw
    if field in DRAW_NUMERIC_FIELDS or field in PIECE_NUMERIC_FIELDS:
        return _parse_numeric(value), raw
    if field == "fabrication_temperature_c":
        numeric = _parse_numeric(value)
        return numeric if numeric is not None else raw, raw
    if value is None or _is_nan(value):
        return None, raw
    return _normalise_text(value), raw


def _format_dimension_display(field: str, value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not _is_nan(value):
        numeric = float(value)
        if not math.isfinite(numeric):
            return None
        precision = 3
        formatted = f"{numeric:.{precision}f}".rstrip("0").rstrip(".")
        return formatted or f"{numeric:.{precision}f}"
    if isinstance(value, str):
        numeric = _parse_numeric(value)
        if numeric is not None:
            precision = 3
            formatted = f"{numeric:.{precision}f}".rstrip("0").rstrip(".")
            return formatted or f"{numeric:.{precision}f}"
    text = _clean_str(value)
    return text or None


def _canonical_dimension_field(field: Optional[str]) -> Optional[str]:
    """Map parsed spreadsheet headers to canonical diameter buckets."""

    if not field:
        return None
    if field in DIMENSION_FIELDS:
        return field
    base = field[:-4] if field.endswith("_raw") else field
    simplified = base.replace("__", "_")
    lowered = simplified.lower()
    cleaned = simplified.strip()
    lowered_clean = cleaned.lower()

    if simplified in DIMENSION_FIELDS:
        return simplified
    lowered_map = {name.lower(): name for name in DIMENSION_FIELDS}
    matched = lowered_map.get(lowered)
    if matched is not None:
        return matched

    if "d_over_d" in lowered or "doverd" in lowered or "ratio_d" in lowered:
        return "d_over_D"

    if lowered_clean in {"d", "d.", "d:"}:
        return "d_um"
    if cleaned in {"D", "D.", "D:"}:
        return "D_um"

    has_micron_hint = any(token in lowered for token in ("_um", " um", "Âµm", "Î¼m", "mic"))
    context_hint = any(
        token in lowered
        for token in (
            "core",
            "jadro",
            "jadra",
            "jadier",
            "glass",
            "sklo",
            "clad",
            "cladding",
            "sheath",
            "outer",
            "inner",
        )
    )
    if context_hint and any(term in lowered for term in ("feed", "feeding", "speed")):
        context_hint = False

    if not (has_micron_hint or context_hint):
        return None

    if any(token in lowered for token in ("core", "jadro", "jadra", "jadier", "inner")):
        return "d_um"
    if any(token in lowered for token in ("glass", "sklo", "clad", "cladding", "sheath", "outer")):
        return "D_um"
    first = simplified[:1]
    if first == "d":
        return "d_um"
    if first == "D":
        return "D_um"
    return None


def _append_dimension_display(
    record: Dict[str, object],
    field: str,
    parsed: Optional[object],
    raw: Optional[str],
) -> None:
    canonical = _canonical_dimension_field(field)
    if canonical not in DIMENSION_FIELDS:
        return
    key = f"{canonical}__display"
    bucket = record.get(key)
    if not isinstance(bucket, list):
        bucket = [] if bucket is None else list(bucket if isinstance(bucket, (list, tuple)) else [bucket])
    for candidate in (parsed, raw):
        text = _format_dimension_display(canonical, candidate)
        if not text:
            continue
        if text not in bucket:
            bucket.append(text)
    record[key] = bucket


def _composition_from_path(path: Path) -> str:
    stem = path.stem
    return stem.split()[0]


def _extract_composition_token(text: str) -> Optional[str]:
    if not text:
        return None
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9.]*", text)
    if not tokens:
        return None
    token = tokens[0].rstrip(".")
    return token or None


def _identifier_candidates_from_path(path: Path) -> List[str]:
    candidates: List[str] = []
    for candidate in (path.stem, path.name):
        text = str(candidate or "").strip()
        if text:
            candidates.append(text)
    for parent in path.parents:
        name = str(parent.name or "").strip()
        if not name:
            continue
        if name.lower() in MICROWIRE_SHORTCUT_STOP_NAMES:
            break
        try:
            parent_stop_name = str(parent.parent.name or "").strip().lower()
        except Exception:
            parent_stop_name = ""
        if parent_stop_name in MICROWIRE_SHORTCUT_STOP_NAMES:
            continue
        candidates.append(name)
    return candidates


def _microscope_key(path: Path) -> Optional[Tuple[str, int, int, Optional[str]]]:
    def _suffix_from_text(text: str, start_idx: int) -> Optional[str]:
        if start_idx >= len(text):
            return None
        suffix_chars: List[str] = []
        started = False
        for ch in text[start_idx:]:
            if ch.isalnum():
                if not started:
                    if not ch.isalpha():
                        break
                    started = True
                suffix_chars.append(ch)
                continue
            break
        suffix = "".join(suffix_chars).strip()
        return suffix or None

    def _match(text: str) -> Optional[Tuple[str, int, int, Optional[str]]]:
        if not text:
            return None
        composition = _extract_composition_token(text)
        if not composition or not any(ch.isdigit() for ch in composition):
            return None

        def _to_pair(
            match: re.Match[str], source_text: str
        ) -> Optional[Tuple[str, int, int, Optional[str]]]:
            try:
                draw_x = int(match.group(1))
                piece_y = int(match.group(2))
            except (TypeError, ValueError):
                return None
            suffix = _suffix_from_text(source_text, match.end())
            return composition, draw_x, piece_y, suffix

        for pattern in MICROSCOPE_PAIR_PATTERNS:
            normalised = text
            if pattern is MICROSCOPE_PAIR_PATTERNS[1]:
                normalised = normalised.replace("-", "/")
            match = pattern.search(normalised)
            if match:
                if pattern is MICROSCOPE_PAIR_PATTERNS[1]:
                    following = normalised[match.end(): match.end() + 1]
                    if following and following in "-/":
                        continue
                pair = _to_pair(match, normalised)
                if pair is not None:
                    return pair

        for match in MICROSCOPE_WHITESPACE_PAIR.finditer(text):
            pair = _to_pair(match, text)
            if pair is not None:
                return pair

        tokens = re.split(r"\s+", text)
        for token in tokens:
            if not token:
                continue
            normalised = token.replace("-", "_").replace("/", "_")
            match = XY_PATTERN.search(normalised)
            if not match:
                continue
            pair = _to_pair(match, normalised)
            if pair is not None:
                return pair
        return None

    for candidate in _identifier_candidates_from_path(path):
        result = _match(candidate)
        if result is not None:
            return result
    return None


def _suffix_from_path(
    path: Optional[Path],
    draw_x: Optional[int],
    piece_y: Optional[int],
) -> Optional[str]:
    if path is None or draw_x is None or piece_y is None:
        return None
    try:
        parsed = _microscope_key(path)
    except Exception:
        return None
    if parsed is None:
        return None
    _, parsed_draw, parsed_piece, suffix = parsed
    if parsed_draw == draw_x and parsed_piece == piece_y:
        return suffix
    return None


def _microscope_category(path: Path) -> str:
    stem = path.stem.lower()
    if "core" in stem:
        return "core"
    if "glass" in stem:
        return "glass"
    return "other"


def _microscope_is_brittle(path: Path) -> bool:
    try:
        return "brittle" in path.stem.lower()
    except Exception:
        return False


def _iter_fragment_files(
    root: Path,
    fragment: str,
    extensions: Sequence[str],
    *,
    max_depth: int = 3,
    limit: Optional[int] = None,
) -> List[Path]:
    if root is None:
        return []
    try:
        if not root.exists():
            return []
    except OSError:
        return []

    fragment_lower = fragment.lower()
    matches: List[Path] = []
    stack: List[Tuple[Path, int]] = [(root, 0)]
    visited: Set[Path] = set()
    while stack:
        current, depth = stack.pop()
        try:
            entries = list(current.iterdir())
        except OSError:
            continue
        for entry in entries:
            try:
                if entry.is_dir():
                    if depth >= max_depth:
                        continue
                    try:
                        key = entry.resolve()
                    except OSError:
                        key = entry
                    if key in visited:
                        continue
                    visited.add(key)
                    name_lower = entry.name.lower()
                    if depth == 0 or fragment_lower in name_lower:
                        stack.append((entry, depth + 1))
                elif entry.is_file():
                    if entry.suffix.lower() in extensions and fragment_lower in entry.name.lower():
                        matches.append(entry)
                        if limit is not None and len(matches) >= limit:
                            return matches
            except OSError:
                continue
    return matches


def _auto_discover_microscope_paths(
    annealing_files: Sequence[Path], logger: logging.Logger
) -> List[Path]:
    discovered: List[Path] = []
    seen: Set[Path] = set()
    for raw_path in annealing_files:
        path = Path(raw_path)
        key = _microscope_key(path)
        if key is None:
            continue
        parts = _split_microwire_key(key)
        if parts is None:
            continue
        composition, draw_x, piece_y, suffix = parts
        fragment = f"{draw_x}_{piece_y}{suffix or ''}"
        try:
            resolved = path.resolve()
            parents = list(resolved.parents)
        except OSError:
            parents = list(path.parents)
        candidate_roots: List[Path] = []
        for parent in parents:
            for name in ("microscope", "Microscope"):
                candidate = parent / name
                try:
                    if candidate.is_dir():
                        candidate_roots.append(candidate)
                except OSError:
                    continue
            if candidate_roots:
                break
        if not candidate_roots:
            continue
        lowered = composition.lower()
        search_dirs: List[Path] = []
        for root_dir in candidate_roots:
            search_dirs.append(root_dir)
            try:
                direct = root_dir / composition
                if direct.is_dir():
                    search_dirs.append(direct)
                for child in root_dir.iterdir():
                    if child.is_dir() and child.name.lower().startswith(lowered):
                        search_dirs.append(child)
            except OSError:
                continue
        for search_dir in dict.fromkeys(search_dirs):
            matches = _iter_fragment_files(search_dir, fragment, MICROSCOPE_EXTENSIONS, limit=20)
            for match in matches:
                try:
                    resolved_match = match.resolve()
                except OSError:
                    resolved_match = match
                if resolved_match in seen:
                    continue
                seen.add(resolved_match)
                discovered.append(match)
    if discovered:
        logger.debug("Auto-discovered %s microscope image(s)", len(discovered))
    return discovered


def _normalise_microscope_text(text: str) -> str:
    cleaned = unicodedata.normalize("NFKC", text or "")
    cleaned = cleaned.replace("Î¼", MICRO_SIGN)
    cleaned = cleaned.replace("|", "1")
    cleaned = re.sub(r"(^|\s)(?:1|I|l){1,2}\]", lambda m: f"{m.group(1)}[1]", cleaned)
    cleaned = re.sub(r"(^|\s)(?:2|Z)\]", lambda m: f"{m.group(1)}[2]", cleaned)

    def _marker_replacer(match: re.Match[str]) -> str:
        digit = match.group("digit")
        if digit in {"I", "l"}:
            digit = "1"
        if digit not in {"1", "2"}:
            digit = "1"
        return f"[{digit}]"

    cleaned = MICROSCOPE_MARKER_PATTERN.sub(_marker_replacer, cleaned)
    cleaned = re.sub(r"\[1(?=\s*\d)", "[1]", cleaned)
    cleaned = re.sub(r"\[2(?=\s*\d)", "[2]", cleaned)
    return cleaned


def _is_secondary_prefix(prefix: str) -> bool:
    return bool(MICROSCOPE_SECONDARY_PREFIX.search(prefix.strip()))


def _has_primary_marker(prefix: str) -> bool:
    snippet = unicodedata.normalize("NFKC", prefix[-8:] if prefix else "")
    if not snippet:
        return False
    snippet = snippet.replace("Î¼", MICRO_SIGN)
    snippet = snippet.replace("|", "1").replace("I", "1").replace("l", "1")
    snippet = snippet.replace("{", "[").replace("(", "[")
    snippet = snippet.replace("}", "]").replace(")", "]")
    snippet = snippet.strip()
    if not snippet:
        return False
    if "[1" in snippet:
        return True
    if "1]" in snippet:
        return True
    return bool(MICROSCOPE_PRIMARY_HINT.search(snippet))


def _parse_microscope_candidates(texts: Iterable[str]) -> List[float]:
    preferred: Dict[float, float] = {}
    fallback_with_marker: Dict[float, float] = {}
    fallback_loose: Dict[float, float] = {}
    for raw_text in texts:
        if not raw_text:
            continue
        text = _normalise_microscope_text(raw_text)
        found_primary = False
        for match in MICROSCOPE_PRIMARY_PATTERN.finditer(text):
            raw_value = match.group("value").replace(",", ".")
            try:
                value = float(raw_value)
            except ValueError:
                continue
            if not math.isfinite(value) or value <= 0:
                continue
            key = round(value, 2)
            preferred.setdefault(key, value)
            found_primary = True
        if found_primary:
            continue
        if preferred:
            continue
        for match in MICROSCOPE_VALUE_PATTERN.finditer(text):
            start = max(match.start() - 6, 0)
            prefix = text[start:match.start()]
            if _is_secondary_prefix(prefix):
                continue
            raw_value = match.group("value").replace(",", ".")
            try:
                value = float(raw_value)
            except ValueError:
                continue
            if not math.isfinite(value) or value <= 0:
                continue
            if value > 1000:
                continue
            key = round(value, 2)
            if _has_primary_marker(prefix):
                fallback_with_marker.setdefault(key, value)
            else:
                fallback_loose.setdefault(key, value)
    if preferred:
        selected = preferred
    elif fallback_with_marker:
        selected = fallback_with_marker
    else:
        selected = fallback_loose
    return [float(v) for v in selected.values()]


def _extract_microscope_diameters(
    path: Path,
    logger: logging.Logger,
) -> MicroscopeOCRResult:
    """Automatic microscope OCR has been retired."""

    _ = path, logger
    return MicroscopeOCRResult()
    @dataclass
    class _OCRWord:
        text: str
        left: int
        top: int
        width: int
        height: int
        conf: Optional[float]

    def _group_paddle_words(raw_result) -> List[List[_OCRWord]]:
        words: List[_OCRWord] = []
        for entry in raw_result or []:
            if not entry:
                continue
            if isinstance(entry, dict):
                texts = entry.get("rec_texts") or []
                scores = entry.get("rec_scores") or []
                polys = entry.get("rec_polys") or entry.get("dt_polys") or []
                boxes = entry.get("rec_boxes")
                for idx, text in enumerate(texts):
                    token = (text or "").strip()
                    if not token:
                        continue
                    score = None
                    if idx < len(scores):
                        try:
                            score = float(scores[idx])
                        except (TypeError, ValueError):
                            score = None
                    points_seq = []
                    poly = None
                    if isinstance(polys, (list, tuple)):
                        if idx < len(polys):
                            poly = polys[idx]
                        elif polys:
                            poly = polys[0]
                    elif polys is not None:
                        poly = polys
                    if poly is not None:
                        try:
                            iterable = poly.tolist()
                        except AttributeError:
                            iterable = list(poly)
                        for pt in iterable or []:
                            if not pt:
                                continue
                            if isinstance(pt, (list, tuple)) and len(pt) >= 2:
                                try:
                                    points_seq.append([float(pt[0]), float(pt[1])])
                                except (TypeError, ValueError):
                                    continue
                    if not points_seq and boxes is not None:
                        box = None
                        if isinstance(boxes, (list, tuple)):
                            if idx < len(boxes):
                                box = boxes[idx]
                            elif boxes:
                                box = boxes[0]
                        else:
                            box = boxes
                        if box is not None:
                            try:
                                flat = box.tolist()
                            except AttributeError:
                                flat = list(box)
                            if len(flat) == 4:
                                x0, y0, x1, y1 = flat
                                points_seq = [
                                    [float(x0), float(y0)],
                                    [float(x1), float(y0)],
                                    [float(x1), float(y1)],
                                    [float(x0), float(y1)],
                                ]
                    if not points_seq:
                        continue
                    xs = [float(pt[0]) for pt in points_seq if pt]
                    ys = [float(pt[1]) for pt in points_seq if pt]
                    if not xs or not ys:
                        continue
                    left = min(xs)
                    top = min(ys)
                    right = max(xs)
                    bottom = max(ys)
                    width = max(int(round(right - left)), 0)
                    height = max(int(round(bottom - top)), 0)
                    conf = None
                    if score is not None and math.isfinite(score):
                        conf = float(score)
                    words.append(
                        _OCRWord(
                            text=token,
                            left=int(round(left)),
                            top=int(round(top)),
                            width=width,
                            height=height,
                            conf=conf,
                        )
                    )
                continue
            for detection in entry:
                if not detection:
                    continue
                try:
                    points, data = detection
                except (TypeError, ValueError):
                    continue
                if not data:
                    continue
                token = (data[0] or "").strip()
                if not token:
                    continue
                score = None
                if len(data) > 1:
                    try:
                        score = float(data[1])
                    except (TypeError, ValueError):
                        score = None
                xs = [float(pt[0]) for pt in (points or []) if pt]
                ys = [float(pt[1]) for pt in (points or []) if pt]
                if not xs or not ys:
                    continue
                left = min(xs)
                top = min(ys)
                right = max(xs)
                bottom = max(ys)
                width = max(int(round(right - left)), 0)
                height = max(int(round(bottom - top)), 0)
                conf = None
                if score is not None and math.isfinite(score):
                    conf = float(score)
                words.append(
                    _OCRWord(
                        text=token,
                        left=int(round(left)),
                        top=int(round(top)),
                        width=width,
                        height=height,
                        conf=conf,
                    )
                )
        if not words:
            return []
        words.sort(key=lambda w: (w.top, w.left))
        lines: List[List[_OCRWord]] = []
        for word in words:
            placed = False
            for line in lines:
                baseline = line[0]
                baseline_center = baseline.top + baseline.height / 2.0
                word_center = word.top + word.height / 2.0
                threshold = max(baseline.height, word.height, 1) * 0.6
                if abs(word_center - baseline_center) <= threshold:
                    line.append(word)
                    placed = True
                    break
            if not placed:
                lines.append([word])
        for line in lines:
            line.sort(key=lambda w: w.left)
        return lines

    seen_detections: Set[Tuple[float, int, int, int, int]] = set()
    seen_texts: Set[str] = set()

    def _extract_line_detections(
        words: Sequence[_OCRWord],
        variant_size: Tuple[int, int],
        *,
        offset: Tuple[int, int] = (0, 0),
        reference_size: Optional[Tuple[int, int]] = None,
    ) -> List[MicroscopeDetection]:
        detections: List[MicroscopeDetection] = []
        vw, vh = variant_size
        if vw <= 0 or vh <= 0:
            return detections
        ref_w, ref_h = reference_size or variant_size
        if ref_w <= 0 or ref_h <= 0:
            return detections
        scale_ref_x = ref_w / float(vw)
        scale_ref_y = ref_h / float(vh)
        offset_x, offset_y = offset
        normalised_words = [_normalise_microscope_text(word.text) for word in words]
        lowered_words = [text.lower().replace("Î¼", "Âµ") for text in normalised_words]
        for idx, word in enumerate(words):
            normalised = normalised_words[idx]
            marker_match = re.match(r"^\[\s*([12Il])\s*\]\s*", normalised)
            marker_offset = 0
            marker: Optional[int] = None
            if marker_match:
                marker_offset = marker_match.end()
                digit = marker_match.group(1)
                if digit in {"I", "l"}:
                    digit = "1"
                try:
                    marker = int(digit)
                except (TypeError, ValueError):
                    marker = None
            candidate_segment = normalised[marker_offset:]
            context_prefix = " ".join(normalised_words[max(0, idx - 3) : idx + 1])
            match = re.search(r"(\d+(?:[.,]\d+)?)", candidate_segment)
            if not match:
                continue
            raw_value = match.group(1).replace(",", ".")
            try:
                value = float(raw_value)
            except ValueError:
                continue
            if not math.isfinite(value) or value <= 0 or value > 1_000:
                continue
            suffix = normalised[marker_offset + match.end():].lower()
            unit_idx: Optional[int] = None
            if any(hint in suffix for hint in MICROSCOPE_UNIT_HINTS):
                unit_idx = idx
            else:
                for offset in range(1, 4):
                    probe_idx = idx + offset
                    if probe_idx >= len(words):
                        break
                    probe_text = lowered_words[probe_idx]
                    if any(hint in probe_text for hint in MICROSCOPE_UNIT_HINTS):
                        unit_idx = probe_idx
                        break
            if unit_idx is None and marker is not None:
                unit_idx = idx
            if marker is None:
                context_normalised = context_prefix.replace("Î¼", MICRO_SIGN)
                if any(hint in context_normalised for hint in ("[1", "1]", "[1]")):
                    marker = 1
                elif any(hint in context_normalised for hint in ("[2", "2]", "[2]")):
                    marker = 2
            if unit_idx is None:
                lookahead = " ".join(lowered_words[idx : min(len(words), idx + 3)])
                if any(hint in lookahead for hint in MICROSCOPE_UNIT_HINTS):
                    unit_idx = idx
            if unit_idx is None:
                continue
            start_idx = idx
            prefix = normalised[: marker_offset + match.start()]
            if any(hint in prefix for hint in ("[1", "1]", "[1]")) or "[1]" in normalised:
                marker = 1
            elif any(hint in prefix for hint in ("[2", "2]", "[2]")) or "[2]" in normalised:
                marker = 2
            if marker is None:
                for offset in range(1, 3):
                    prev_idx = idx - offset
                    if prev_idx < 0:
                        break
                    prev_text = normalised_words[prev_idx]
                    if any(hint in prev_text for hint in ("[1", "1]", "[1]")):
                        marker = 1
                        start_idx = min(start_idx, prev_idx)
                        break
                    if any(hint in prev_text for hint in ("[2", "2]", "[2]")):
                        marker = 2
                        start_idx = min(start_idx, prev_idx)
                        break
            tokens = words[start_idx : unit_idx + 1]
            if not tokens:
                continue
            left = min(entry.left for entry in tokens)
            top = min(entry.top for entry in tokens)
            right = max(entry.left + entry.width for entry in tokens)
            bottom = max(entry.top + entry.height for entry in tokens)
            left_ref = (left * scale_ref_x) + offset_x
            top_ref = (top * scale_ref_y) + offset_y
            right_ref = (right * scale_ref_x) + offset_x
            bottom_ref = (bottom * scale_ref_y) + offset_y
            left_base = max(int(round(left_ref * scale_to_original_x)), 0)
            top_base = max(int(round(top_ref * scale_to_original_y)), 0)
            right_base = min(
                int(round(right_ref * scale_to_original_x)), original_width
            )
            bottom_base = min(
                int(round(bottom_ref * scale_to_original_y)), original_height
            )
            if right_base <= left_base or bottom_base <= top_base:
                continue
            confidences = [
                entry.conf
                for entry in tokens
                if isinstance(entry.conf, (int, float)) and entry.conf >= 0
            ]
            confidence = None
            if confidences:
                confidence = float(sum(confidences) / len(confidences))
            detection = MicroscopeDetection(
                value=value,
                image_path=path,
                bbox=(left_base, top_base, right_base, bottom_base),
                text=" ".join(entry.text for entry in tokens),
                source="ocr",
                confidence=confidence,
                marker=marker,
            )
            detections.append(detection)
        return detections

    def _consume_ocr_output(
        label: str,
        raw_result,
        variant_size: Tuple[int, int],
        *,
        offset: Tuple[int, int] = (0, 0),
        reference_size: Optional[Tuple[int, int]] = None,
    ) -> bool:
        lines = _group_paddle_words(raw_result)
        if not lines:
            return False
        combined_lines: List[str] = []
        for line in lines:
            if not line:
                continue
            line_text_raw = " ".join(word.text for word in line)
            line_text = " ".join(line_text_raw.split())
            if not line_text:
                continue
            _append_candidate(line_text)
            tagged = f"{label}: {line_text}"
            if tagged not in seen_texts:
                result.texts.append(tagged)
                seen_texts.add(tagged)
            combined_lines.append(line_text)
            for detection in _extract_line_detections(
                line,
                variant_size,
                offset=offset,
                reference_size=reference_size,
            ):
                bbox = detection.bbox or (0, 0, 0, 0)
                key = (
                    round(detection.value, 2),
                    bbox[0],
                    bbox[1],
                    bbox[2],
                    bbox[3],
                )
                if key in seen_detections:
                    continue
                if detection.text:
                    detection.text = " ".join(str(detection.text).split())
                result.detections.append(detection)
                seen_detections.add(key)
                result.append_value(detection.value)
        if combined_lines:
            _append_candidate("\n".join(combined_lines))
            return True
        return False

    processed_any_variant = False
    if ocr is not None:
        try:
            direct_result = ocr.ocr(str(path))
        except Exception:
            direct_result = None
        if direct_result:
            if _consume_ocr_output("original", direct_result, (original_width, original_height)):
                processed_any_variant = True

        for label, variant, offset, ref_size in _iter_variant_entries():
            if variant is None:
                continue
            variant_size = variant.size
            variant_rgb = variant.convert("RGB")
            variant_array = np.array(variant_rgb)
            if variant_array.ndim == 2:
                variant_array = np.stack([variant_array] * 3, axis=-1)
            elif variant_array.ndim == 3 and variant_array.shape[2] == 4:
                variant_array = variant_array[:, :, :3]
            if variant_array.ndim != 3 or variant_array.shape[2] != 3:
                continue
            variant_array = variant_array.astype("uint8", copy=False)
            bgr_image = variant_array[:, :, ::-1].copy()
            try:
                ocr_result = ocr.ocr(bgr_image)
            except Exception:
                log.debug(
                    "PaddleOCR failed while processing %s; skipping this variant",
                    path,
                    exc_info=True,
                )
                continue
            if _consume_ocr_output(
                label,
                ocr_result,
                variant_size,
                offset=offset,
                reference_size=ref_size,
            ):
                processed_any_variant = True

            if candidates and any("[1" in candidate or "1]" in candidate for candidate in candidates):
                values = _parse_microscope_candidates(candidates)
                if values:
                    for value in values:
                        result.append_value(value)
                    break
            if result.values:
                break

    if not processed_any_variant:
        log.debug("PaddleOCR produced no candidate text for %s", path)

    if not result.values:
        parsed = _parse_microscope_candidates(candidates)
        for value in parsed:
            result.append_value(value)

    if not result.values and result.detections:
        seen_keys: Dict[float, float] = {}
        for detection in result.detections:
            key = round(detection.value, 2)
            seen_keys.setdefault(key, detection.value)
        result.values.extend(seen_keys.values())

    if not result.values:
        fallback_numbers: List[float] = []
        seen_numbers: Set[float] = set()
        pool: List[str] = []
        if candidates:
            pool.extend(candidates)
        if result.texts:
            pool.extend(result.texts)
        for raw_text in pool:
            if not raw_text:
                continue
            text = _normalise_microscope_text(raw_text)
            for match in re.finditer(r"\d+(?:[.,]\d+)?", text):
                start, end = match.start(), match.end()
                prefix = text[start - 1] if start > 0 else ""
                suffix = text[end] if end < len(text) else ""
                if prefix in "[{(" and suffix in "]})":
                    continue
                number = match.group(0).replace(",", ".")
                try:
                    value = float(number)
                except ValueError:
                    continue
                if not math.isfinite(value) or value <= 0 or value > 500:
                    continue
                key = round(value, 3)
                if key in seen_numbers:
                    continue
                seen_numbers.add(key)
                fallback_numbers.append(value)
        for value in fallback_numbers:
            result.append_value(value)

    if result.values:
        deduped: Dict[float, float] = {}
        for value in result.values:
            key = round(float(value), 2)
            deduped.setdefault(key, float(value))
        result.values = list(deduped.values())

    return result


class MicroscopeGroupingResult(tuple):
    """Combined grouping output that behaves like both tuple and mapping."""

    __slots__ = ()

    def __new__(
        cls,
        index: Dict[Tuple[str, int, int, Optional[str]], MicroscopeMeasurements],
        cache: Dict[str, MicroscopeCacheEntry],
    ) -> "MicroscopeGroupingResult":
        return super().__new__(cls, (index, cache))

    @property
    def index(self) -> Dict[Tuple[str, int, int, Optional[str]], MicroscopeMeasurements]:
        return super().__getitem__(0)

    @property
    def cache(self) -> Dict[str, MicroscopeCacheEntry]:
        return super().__getitem__(1)

    def __contains__(self, key: object) -> bool:
        return key in self.index

    def __getitem__(self, key):  # type: ignore[override]
        if isinstance(key, (int, slice)):
            return super().__getitem__(key)
        return self.index[key]

    def get(self, key, default=None):
        return self.index.get(key, default)

    def keys(self):
        return self.index.keys()

    def items(self):
        return self.index.items()

    def values(self):
        return self.index.values()

    def __repr__(self) -> str:
        return f"MicroscopeGroupingResult(index={self.index!r}, cache={self.cache!r})"


def _group_microscope_measurements(
    microscope_files: Sequence[Path],
    logger: Optional[logging.Logger],
    progress_callback: Optional[Callable[[int, int], None]] = None,
    debug_callback: Optional[Callable[[Path, MicroscopeOCRResult], None]] = None,
    update_callback: Optional[
        Callable[[Tuple[str, int, int, Optional[str]], MicroscopeMeasurements], None]
    ] = None,
    cache: Optional[Mapping[str, Any]] = None,
) -> Tuple[
    Dict[Tuple[str, int, int, Optional[str]], MicroscopeMeasurements],
    Dict[str, MicroscopeCacheEntry],
]:
    """Group microscope captures into microwire measurements while caching OCR output."""

    log = _logger(logger)
    combined: List[Path] = []
    seen: Set[Path] = set()
    for raw_path in microscope_files:
        path = Path(raw_path)
        try:
            resolved = path.resolve()
        except OSError:
            resolved = path
        if resolved in seen:
            continue
        seen.add(resolved)
        combined.append(path)

    def _cache_key(path: Path) -> str:
        try:
            return str(path.resolve())
        except Exception:
            return str(path)

    cache_lookup: Dict[str, MicroscopeCacheEntry] = {}
    if cache:
        for key, payload in cache.items():
            entry: Optional[MicroscopeCacheEntry]
            if isinstance(payload, MicroscopeCacheEntry):
                entry = payload
            elif isinstance(payload, Mapping):
                entry = MicroscopeCacheEntry.from_dict(payload)
            else:
                entry = None
            if entry is None:
                continue
            cache_lookup[str(key)] = entry

    total = len(combined)
    processed = 0
    grouped: Dict[Tuple[str, int, int, Optional[str]], MicroscopeMeasurements] = {}
    updated_cache: Dict[str, MicroscopeCacheEntry] = {}
    for raw_path in combined:
        path = raw_path.expanduser()
        processed += 1

        def _notify() -> None:
            if progress_callback is not None:
                try:
                    progress_callback(processed, total)
                except BuildCancelledError:
                    raise
                except Exception:
                    pass

        try:
            stat_result = path.stat()
        except OSError:
            log.debug("Microscope image %s does not exist; skipping", path)
            _notify()
            continue
        mtime = float(stat_result.st_mtime)
        size = int(stat_result.st_size)

        key = _microscope_key(path)
        if key is None:
            log.debug("Unable to derive microwire key from microscope image %s", path)
            _notify()
            continue
        category = _microscope_category(path)
        record = grouped.setdefault(key, MicroscopeMeasurements())
        if _microscope_is_brittle(path):
            record.brittle = True

        cache_token = _cache_key(path)
        cache_entry = cache_lookup.get(cache_token) or cache_lookup.get(str(path))
        raw_result: MicroscopeOCRResult | Iterable[float] | None = None
        if cache_entry is not None and cache_entry.is_current(path):
            try:
                raw_result = cache_entry.to_result(path)
            except Exception:
                raw_result = None
        if raw_result is None:
            try:
                raw_result = _extract_microscope_diameters(path, log)
            except BuildCancelledError:
                raise
            except Exception:
                log.exception("Microscope OCR failed for %s", path)
                _notify()
                continue

        if isinstance(raw_result, MicroscopeOCRResult):
            result = MicroscopeOCRResult(
                values=list(raw_result.values),
                detections=list(raw_result.detections),
                texts=list(raw_result.texts),
            )
        else:
            values = [
                float(v)
                for v in raw_result
                if isinstance(v, (int, float)) and math.isfinite(float(v)) and float(v) > 0
            ]
            result = MicroscopeOCRResult(values=values)

        values = list(result.values)
        detections = list(result.detections)
        debug_texts = list(result.texts)

        if debug_callback is not None:
            debug_payload = MicroscopeOCRResult(
                values=list(result.values),
                detections=list(result.detections),
                texts=list(debug_texts),
            )
            try:
                debug_callback(path, debug_payload)
            except Exception:
                log.debug(
                    "Microscope OCR debug callback failed for %s", path, exc_info=True
                )

        if detections:
            grouped_detections: Dict[str, List[MicroscopeDetection]] = {}
            for detection in detections:
                override_category = category
                if category == "glass":
                    if detection.marker == 2:
                        continue
                    if detection.marker == 1:
                        override_category = "glass"
                else:
                    if detection.marker == 1:
                        override_category = "core"
                    elif detection.marker == 2:
                        override_category = "glass"
                grouped_detections.setdefault(override_category, []).append(detection)
            used_keys: Set[float] = set()
            for det_category, det_list in grouped_detections.items():
                numeric_values = [
                    float(det.value)
                    for det in det_list
                    if isinstance(det.value, (int, float))
                ]
                if not numeric_values:
                    for detection in det_list:
                        fallback_path = getattr(detection, "image_path", None) or path
                        record.add_placeholder(det_category, fallback_path)
                    continue
                record.extend(det_category, numeric_values, det_list)
                for numeric in numeric_values:
                    used_keys.add(round(float(numeric), 3))
            residual_values: List[float] = []
            for value in values:
                if not isinstance(value, (int, float)):
                    continue
                numeric = float(value)
                rounded = round(numeric, 3)
                if rounded in used_keys:
                    continue
                residual_values.append(numeric)
            if residual_values:
                record.extend(category, residual_values, [])
        elif values:
            record.extend(category, values, detections)
        else:
            record.add_placeholder(category, path)
            cache_entry_final = MicroscopeCacheEntry.from_result(path, mtime, size, result)
            cache_entry_final.path = cache_token
            updated_cache[cache_token] = cache_entry_final
            _notify()
            continue
        if not detections and values:
            record.add_placeholder(category, path)
        if update_callback is not None:
            try:
                update_callback(key, record)
            except Exception:
                log.debug(
                    "Microscope OCR debug callback failed for %s", key, exc_info=True
                )

        cache_entry_final = MicroscopeCacheEntry.from_result(path, mtime, size, result)
        cache_entry_final.path = cache_token
        updated_cache[cache_token] = cache_entry_final
        _notify()

    for record in grouped.values():
        if not getattr(record, "brittle", False) and not record.core and bool(record.glass):
            record.brittle = True

    missing_references: List[str] = []
    mismatched_references: List[str] = []
    for key, override in MANUAL_DIAMETER_OVERRIDES.items():
        record = grouped.get(key)
        d_expected = override.get("d")
        D_expected = override.get("D")
        d_actual = record.best_core() if record else None
        D_actual = record.best_glass() if record else None
        comp, draw, piece = key
        label = f"{comp} {draw}/{piece}" if piece is not None else f"{comp} {draw}"
        if d_expected is not None:
            if d_actual is None:
                missing_references.append(f"{label} d={d_expected}")
            elif abs(d_actual - d_expected) > 0.5:
                mismatched_references.append(
                    f"{label} d expected {d_expected:.2f}µm got {d_actual:.2f}µm"
                )
        if D_expected is not None:
            if D_actual is None:
                missing_references.append(f"{label} D={D_expected}")
            elif abs(D_actual - D_expected) > 0.5:
                mismatched_references.append(
                    f"{label} D expected {D_expected:.2f}µm got {D_actual:.2f}µm"
                )
    if missing_references:
        preview = ", ".join(missing_references[:6])
        if len(missing_references) > 6:
            preview += ", …"
        log.info(
            "Microscope OCR is still missing %s manual reference(s): %s",
            len(missing_references),
            preview,
        )
    if mismatched_references:
        preview = ", ".join(mismatched_references[:6])
        if len(mismatched_references) > 6:
            preview += ", …"
        log.warning(
            "Microscope OCR deviates from manual references: %s", preview
        )
    return MicroscopeGroupingResult(grouped, updated_cache)

def _draw_key(path: Path) -> Optional[Tuple[str, int]]:
    for candidate in _identifier_candidates_from_path(path):
        if not candidate:
            continue
        composition = _extract_composition_token(candidate)
        if not composition or not any(ch.isdigit() for ch in composition):
            continue
        draw_match = re.search(r"(\d+)", candidate)
        if not draw_match:
            continue
        try:
            draw_x = int(draw_match.group(1))
        except (TypeError, ValueError):
            continue
        return composition, draw_x
    return None


def _video_summary_key(path: Path) -> Optional[Tuple[str, int, Optional[int]]]:
    key = _microscope_key(path)
    if key is not None:
        parts = _split_microwire_key(key)
        if parts is not None:
            composition, draw_x, piece_y, _suffix = parts
            return composition, draw_x, piece_y
    draw_key = _draw_key(path)
    if draw_key is None:
        return None
    composition, draw_x = draw_key
    return composition, draw_x, None


def _collect_video_sources(
    video_files: Sequence[Path],
    logger: Optional[logging.Logger],
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary]:
    log = _logger(logger)
    aggregated: Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary] = {}
    unique_files = list(dict.fromkeys(Path(p) for p in video_files))
    total = len(unique_files)
    processed = 0
    for raw_path in unique_files:
        path = raw_path.expanduser()
        processed += 1

        def _notify() -> None:
            if progress_callback is not None:
                try:
                    progress_callback(processed, total)
                except BuildCancelledError:
                    raise
                except Exception:
                    pass

        try:
            if not path.exists():
                log.debug("Video file %s does not exist; skipping", path)
                _notify()
                continue
        except OSError:
            _notify()
            continue
        summary_key = _video_summary_key(path)
        if summary_key is None:
            log.debug("Unable to derive microwire key from video %s", path)
            _notify()
            continue
        summary = aggregated.setdefault(summary_key, VideoMetricsSummary())
        summary.sources.add(path)
        _notify()
    return aggregated


def _collect_video_metrics(
    video_files: Sequence[Path],
    logger: Optional[logging.Logger],
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary]:
    log = _logger(logger)
    aggregated: Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary] = {}
    unique_files = list(dict.fromkeys(Path(p) for p in video_files))
    total = len(unique_files)
    processed = 0
    for raw_path in unique_files:
        path = raw_path.expanduser()
        processed += 1

        def _notify() -> None:
            if progress_callback is not None:
                try:
                    progress_callback(processed, total)
                except BuildCancelledError:
                    raise
                except Exception:
                    pass

        try:
            if not path.exists():
                log.debug("Video file %s does not exist; skipping", path)
                _notify()
                continue
        except OSError:
            _notify()
            continue
        summary_key = _video_summary_key(path)
        if summary_key is None:
            log.debug("Unable to derive microwire key from video %s", path)
            _notify()
            continue
        composition, draw_x, piece_y = summary_key
        try:
            result = extract_video_metrics(path, logger=log)
        except Exception:
            log.exception("Failed to analyse video %s", path)
            _notify()
            continue
        summary = aggregated.setdefault((composition, draw_x, piece_y), VideoMetricsSummary())
        summary.record(result, source=path)
        _notify()
    return aggregated


def _parse_draw_rows(
    df: pd.DataFrame,
    headers: List[Optional[str]],
    composition: str,
    index: FabricationIndex,
    logger: logging.Logger,
    source_path: Path,
) -> None:
    seen_draws: set[int] = set()
    for _, row in df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if first_cell is None or _is_nan(first_cell):
            continue
        text = _normalise_text(first_cell)
        if not text:
            continue
        draw_x: Optional[int] = None
        m = DRAW_PATTERN.match(text)
        if m:
            draw_x = int(m.group("draw"))
        elif text.lower().startswith(composition.lower()):
            draw_x = 1
        if draw_x is None:
            continue
        if draw_x in seen_draws:
            logger.debug("Duplicate draw %s in %s", draw_x, composition)
        seen_draws.add(draw_x)
        record: Dict[str, object] = {"_source_path": str(source_path)}
        for col_idx, field in enumerate(headers):
            if col_idx == 0 or not field:
                continue
            value = row.iloc[col_idx] if col_idx < len(row) else None
            parsed, raw = _extract_field_value(field, value)
            record[field] = parsed
            if field in RAW_VALUE_FIELDS:
                record[f"{field}_raw"] = raw
            _append_dimension_display(record, field, parsed, raw)
        index.set_draw(composition, draw_x, record)


def _fabrication_piece_has_meaningful_values(record: Optional[Mapping[str, object]]) -> bool:
    if not isinstance(record, Mapping):
        return False
    for field, value in record.items():
        if (
            not field
            or field.startswith("_")
            or field.endswith("_raw")
            or field.endswith("__display")
        ):
            continue
        if isinstance(value, str):
            if value.strip():
                return True
            continue
        if isinstance(value, (list, tuple, set, dict)):
            if value:
                return True
            continue
        if value is None or _is_nan(value):
            continue
        if isinstance(value, (int, float)):
            if float(value) != 0.0:
                return True
            continue
        return True
    return False


def _relevant_sibling_piece_candidates(
    available_pieces: Iterable[int],
    relevant_pieces: Optional[Iterable[Optional[int]]] = None,
    piece_records: Optional[Mapping[int, Mapping[str, object]]] = None,
) -> List[int]:
    """Return same-draw sibling pieces up to the last meaningful positive row.

    Piece workbooks often contain placeholder rows after the real pieces. The
    builder should therefore promote sibling pieces only up to the highest
    positive piece row that contains meaningful fabrication data.
    """

    available = sorted(
        {
            int(piece)
            for piece in available_pieces
            if piece is not None and int(piece) > 0
        }
    )
    relevant: List[int] = []
    if relevant_pieces is not None:
        for piece in relevant_pieces:
            if piece is None:
                continue
            try:
                piece_int = int(piece)
            except (TypeError, ValueError):
                continue
            if piece_int > 0:
                relevant.append(piece_int)
    if available:
        meaningful = []
        if piece_records:
            meaningful = [
                piece
                for piece in available
                if _fabrication_piece_has_meaningful_values(piece_records.get(piece))
            ]
        if meaningful:
            limit = max(meaningful)
            return [piece for piece in available if piece <= limit]
        return available
    if relevant:
        return sorted(set(relevant))
    return available


def _parse_piece_rows(
    df: pd.DataFrame,
    headers: List[Optional[str]],
    composition: str,
    draw_x: Optional[int],
    index: FabricationIndex,
    logger: logging.Logger,
    source_path: Path,
) -> None:
    if draw_x is None:
        logger.warning("Could not determine draw number for piece workbook %s", composition)
        return
    for _, row in df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if first_cell is None or _is_nan(first_cell):
            continue
        text = _normalise_text(first_cell)
        if not text:
            continue
        m = PIECE_PATTERN.match(text)
        if not m:
            continue
        piece_y = int(m.group("piece"))
        record: Dict[str, object] = {"_source_path": str(source_path)}
        for col_idx, field in enumerate(headers):
            if col_idx == 0 or not field:
                continue
            value = row.iloc[col_idx] if col_idx < len(row) else None
            parsed, raw = _extract_field_value(field, value)
            record[field] = parsed
            if field in RAW_VALUE_FIELDS:
                record[f"{field}_raw"] = raw
            _append_dimension_display(record, field, parsed, raw)
        index.set_piece(composition, draw_x, piece_y, record)


def _read_excel(path: Path, logger: Optional[logging.Logger] = None) -> pd.DataFrame:
    try:
        return pd.read_excel(path, header=None, dtype=object)
    except ImportError:
        raise
    except ValueError as exc:
        engines: List[str] = []
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            engines.append("openpyxl")
        if suffix in {".xls", ".xlsb"}:
            engines.append("xlrd")
        engines.extend(["openpyxl", "xlrd", "odf"])
        for engine in engines:
            try:
                return pd.read_excel(path, header=None, dtype=object, engine=engine)
            except ImportError:
                continue
            except ValueError:
                continue
            except Exception:
                continue
        if logger is not None:
            logger.warning(
                "%s: unable to determine Excel format (%s); skipping",
                path,
                exc,
            )
        return pd.DataFrame()
    except Exception as exc:
        if logger is not None:
            logger.exception("Failed to read %s", path)
        return pd.DataFrame()


def _parse_composition_workbook(path: Path, index: FabricationIndex, logger: logging.Logger) -> None:
    df = _read_excel(path, logger)
    if df.empty:
        logger.warning("%s is empty", path)
        return
    header_values = _merged_header_row(df, 0)
    headers = [_header_key(value) for value in header_values]
    data = df.iloc[1:].reset_index(drop=True)
    composition = _composition_from_path(path)
    _parse_draw_rows(data, headers, composition, index, logger, path)


def _is_piece_header_row(values: Sequence[object]) -> bool:
    tokens: List[str] = []
    collapsed: List[str] = []
    for value in values:
        text = _normalise_text(value)
        if not text:
            continue
        ascii_text = unicodedata.normalize("NFKD", text)
        ascii_text = "".join(ch for ch in ascii_text if not unicodedata.combining(ch)).lower()
        tokens.append(ascii_text)
        collapsed.append(re.sub(r"[^a-z0-9]", "", ascii_text))
    if not tokens:
        return False
    collapsed_set = {token for token in collapsed if token}
    token_set = {token for token in tokens if token}
    score = 0
    if any(token.startswith("p") and len(token) <= 3 for token in collapsed_set):
        score += 1
    if any("datum" in token for token in token_set):
        score += 1
    if any("dlka" in token or "dka" in token or "dlzka" in token for token in token_set):
        score += 1
    if any(token in {"dum", "d", "dmu"} or "dum" in token for token in collapsed_set):
        score += 1
    if any("d/d" in token or "d/" in token for token in token_set) or "dd" in collapsed_set:
        score += 1
    if any("intenzita" in token for token in token_set):
        score += 1
    if any("hsw" in token for token in token_set):
        score += 1
    return score >= 3


def _parse_piece_workbook(path: Path, index: FabricationIndex, logger: logging.Logger) -> None:
    df = _read_excel(path, logger)
    if df.empty:
        logger.warning("%s is empty", path)
        return
    best_idx: Optional[int] = None
    best_headers: List[Optional[str]] = []
    best_score = -1
    for idx, row in df.iterrows():
        if not _is_piece_header_row(row.tolist()):
            continue
        candidate_values = _merged_header_row(df, idx)
        candidate_headers = [_header_key(value) for value in candidate_values]
        score = sum(1 for field in candidate_headers if field)
        if score > best_score:
            best_idx = idx
            best_headers = candidate_headers
            best_score = score
    if best_idx is None:
        if len(df.index) > 1:
            best_idx = 1
            logger.warning("%s: unable to locate header row; using the second row as a fallback", path)
            candidate_values = _merged_header_row(df, best_idx)
            best_headers = [_header_key(value) for value in candidate_values]
        else:
            logger.warning("%s: unable to locate header row", path)
            return
    header_idx = best_idx
    header_values = _merged_header_row(df, header_idx)
    if not best_headers:
        best_headers = [_header_key(value) for value in header_values]
    headers = best_headers
    data = df.iloc[header_idx + 1 :].reset_index(drop=True)
    stem = path.stem
    match = re.search(r"(?P<draw>\d+)[._](?P<comp>[A-Za-z0-9]+)", stem)
    draw_x: Optional[int] = None
    composition = _composition_from_path(path)
    if match:
        draw_x = int(match.group("draw"))
        composition = match.group("comp")
    _parse_piece_rows(data, headers, composition, draw_x, index, logger, path)


def build_fabrication_index(
    fabrication_files: Sequence[Path],
    logger: Optional[logging.Logger] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    cancel_callback: Optional[Callable[[], bool]] = None,
) -> FabricationIndex:
    log = _logger(logger)
    index = FabricationIndex()
    unique_files = list(dict.fromkeys(Path(p) for p in fabrication_files))
    total = len(unique_files)
    for idx, path in enumerate(unique_files, start=1):
        if cancel_callback is not None and cancel_callback():
            raise BuildCancelledError()
        if path.suffix.lower() != ".xlsx":
            log.debug("Skipping non-Excel file %s", path)
            if progress_callback is not None:
                try:
                    progress_callback(idx, total)
                except Exception:
                    pass
            continue
        stem = path.stem
        parent_stem = path.parent.name
        if re.match(r"^\d+", stem) or re.match(r"^\d+", parent_stem):
            _parse_piece_workbook(path, index, log)
        else:
            _parse_composition_workbook(path, index, log)
        if progress_callback is not None:
            try:
                progress_callback(idx, total)
            except Exception:
                pass
    return index


@dataclass
class MeasurementMetadata:
    composition_token: str
    draw_x: Optional[int]
    piece_y: Optional[int]
    setpoint_mA: Optional[int]
    alt_variant: bool
    measurement_id: str
    file_name: str
    relpath: str
    timestamp_mtime_utc: str


@dataclass
class MeasurementRecord:
    path: Path
    metadata: MeasurementMetadata
    dataframe: pd.DataFrame
    sanity_ok: bool
    sanity_error: Optional[float]
    transition_summary: Tuple[str, ...] = ()


def _hash_file(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _metadata_from_path(path: Path, root: Optional[Path] = None) -> MeasurementMetadata:
    stat = path.stat()
    timestamp = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
    base = path.stem
    parts = base.split()
    composition_match = COMPOSITION_TOKEN_PATTERN.search(base)
    composition = (
        composition_match.group("composition")
        if composition_match
        else parts[0]
        if parts
        else base
    )
    identity_tail = base[composition_match.end():] if composition_match else base
    xy_match = DRAW_PIECE_AFTER_COMPOSITION_PATTERN.search(identity_tail)
    if xy_match is None:
        xy_match = XY_PATTERN.search(base)
    draw_x: Optional[int] = int(xy_match.group(1)) if xy_match else None
    piece_y: Optional[int] = int(xy_match.group(2)) if xy_match else None
    setpoint_match = SETPOINT_PATTERN.search(base)
    setpoint = int(setpoint_match.group(1)) if setpoint_match else None
    alt_variant = bool(ALT_VARIANT_PATTERN.search(base))
    relpath = os.fspath(path.relative_to(root)) if root and path.is_relative_to(root) else path.as_posix()
    return MeasurementMetadata(
        composition_token=composition,
        draw_x=draw_x,
        piece_y=piece_y,
        setpoint_mA=setpoint,
        alt_variant=alt_variant,
        measurement_id=_hash_file(path),
        file_name=path.name,
        relpath=relpath,
        timestamp_mtime_utc=timestamp,
    )


def _numeric_text_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace("\u2212", "-", regex=False)
        .str.replace(",", ".", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce")


def _normalise_annealing_columns(
    *,
    current_mA: pd.Series,
    voltage_v: pd.Series,
    resistance_ohm: pd.Series,
    cycle: Optional[pd.Series] = None,
) -> pd.DataFrame:
    df = pd.DataFrame(
        {
            "I_mA": _numeric_text_series(current_mA),
            "V_V": _numeric_text_series(voltage_v),
            "R_ohm": _numeric_text_series(resistance_ohm),
        }
    )
    if cycle is not None:
        df["Cycle"] = _numeric_text_series(cycle)
    df = df.replace([np.inf, -np.inf], np.nan)
    df = df.dropna(subset=["I_mA", "R_ohm"]).reset_index(drop=True)
    if df.empty:
        raise ValueError("no valid samples after parsing")
    df["I_A"] = df["I_mA"] / 1_000.0
    columns = ["I_A", "V_V", "R_ohm", "I_mA"]
    if "Cycle" in df.columns:
        columns.append("Cycle")
    return df.loc[:, columns]


def _load_annealing_dat(path: Path) -> pd.DataFrame:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    non_empty = [line for line in lines if line.strip()]
    if not non_empty:
        raise ValueError(f"{path}: no valid samples after parsing")
    first_line = non_empty[0]
    if "Cycle" in first_line and "Iset_mA" in first_line:
        df = pd.read_csv(path, sep=r"\s+", engine="python")
        required = {"Cycle", "Ireal_mA", "Voltage_V", "Resistance_Ohm"}
        missing = required.difference(df.columns)
        if missing:
            raise ValueError(f"{path}: missing Kosice cycle .dat columns {sorted(missing)}")
        return _normalise_annealing_columns(
            current_mA=df["Ireal_mA"],
            voltage_v=df["Voltage_V"],
            resistance_ohm=df["Resistance_Ohm"],
            cycle=df["Cycle"],
        )

    rows: List[List[float]] = []
    for line in non_empty:
        parts = line.split()
        if len(parts) != 4:
            continue
        try:
            rows.append([float(part.replace(",", ".")) for part in parts])
        except ValueError:
            continue
    if not rows:
        raise ValueError(f"{path}: no valid four-column Kosice .dat samples after parsing")
    df = pd.DataFrame(rows, columns=["Iset_A", "Ireal_A", "Voltage_V", "Resistance_Ohm"])
    return _normalise_annealing_columns(
        current_mA=df["Ireal_A"] * 1_000.0,
        voltage_v=df["Voltage_V"],
        resistance_ohm=df["Resistance_Ohm"],
    )


def _annealing_text_current_unit(path: Path) -> Optional[str]:
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError:
        return None
    for line in lines[:10]:
        text = line.strip().lstrip("#").strip().casefold()
        if not text:
            continue
        if "current" not in text and "ireal" not in text and "iset" not in text:
            continue
        if re.search(r"(?:^|[^a-z])m\s*a(?:[^a-z]|$)", text) or "_ma" in text:
            return "mA"
        if re.search(r"(?:^|[^a-z])a(?:[^a-z]|$)", text):
            return "A"
    return None


def _trim_annealing_burnthrough(df: pd.DataFrame) -> pd.DataFrame:
    try:
        from plotting.plugins.current_annealing.burnthrough import trim_burnthrough_glitch
    except ImportError:
        return df

    currents_mA = df["I_mA"].to_numpy(dtype=float)
    resistances = df["R_ohm"].to_numpy(dtype=float)
    trimmed_currents, trimmed_resistances = trim_burnthrough_glitch(currents_mA, resistances)
    trimmed_count = int(trimmed_currents.shape[0])
    if trimmed_count < currents_mA.shape[0]:
        df = df.iloc[:trimmed_count].copy()
        df.loc[:, "I_mA"] = trimmed_currents
        df.loc[:, "I_A"] = trimmed_currents / 1e3
        df.loc[:, "R_ohm"] = trimmed_resistances
        df = df.reset_index(drop=True)
    return df


def _load_annealing(
    path: Path,
    *,
    expected_setpoint_mA: Optional[float] = None,
) -> pd.DataFrame:
    if path.suffix.lower() == ".dat":
        return _trim_annealing_burnthrough(_load_annealing_dat(path))

    try:
        df = pd.read_csv(path, sep=None, engine="python", names=ANNEALING_COLUMNS, header=None)
    except (csv.Error, pd.errors.ParserError):
        df = pd.read_csv(
            path,
            sep=r"\s+",
            engine="python",
            names=ANNEALING_COLUMNS,
            header=None,
        )
    df = df.iloc[:, :3].copy()
    for column in ANNEALING_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce")
    df = df.dropna(subset=["I_A", "R_ohm"]).reset_index(drop=True)
    currents = df["I_A"].to_numpy(dtype=float)
    finite = currents[np.isfinite(currents)]
    scale = 1.0
    current_unit = _annealing_text_current_unit(path)
    if current_unit == "mA":
        scale = 1e-3
    if finite.size:
        max_abs = float(np.nanmax(np.abs(finite)))
        median_abs = float(np.nanmedian(np.abs(finite)))
        if scale == 1.0 and expected_setpoint_mA and expected_setpoint_mA > 0:
            expected_amp = expected_setpoint_mA / 1000.0
            if expected_amp > 0 and max_abs > expected_amp * 5:
                scale = 1e-3
        if scale == 1.0:
            if max_abs > 500:
                scale = 1e-3
            elif median_abs > 10:
                scale = 1e-3
    scaled_currents = currents * scale
    df.loc[:, "I_A"] = scaled_currents
    df.loc[:, "I_mA"] = scaled_currents * 1_000.0
    return _trim_annealing_burnthrough(df)


def _annealing_transition_summary(df: pd.DataFrame, *, label: str | None = None) -> Tuple[str, ...]:
    try:
        from plotting.plugins.current_annealing import core as annealing_core

        summaries = annealing_core.summarize_transition_loops(df)
        lines = annealing_core.format_transition_summaries(summaries, label=label)
    except Exception:
        return ()
    return tuple(str(line) for line in lines if str(line).strip())


def _series_to_mA(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    try:
        max_abs = float(numeric.abs().max(skipna=True) or 0.0)
    except Exception:
        max_abs = 0.0
    if max_abs <= 50:
        return numeric * 1e3
    return numeric


def _resistance_sanity_check(df: pd.DataFrame) -> Tuple[bool, Optional[float]]:
    currents = df["I_A"].to_numpy(dtype=float)
    voltages = df["V_V"].to_numpy(dtype=float)
    resistances = df["R_ohm"].to_numpy(dtype=float)
    with np.errstate(divide="ignore", invalid="ignore"):
        expected = np.divide(voltages, currents, out=np.full_like(resistances, np.nan), where=currents != 0)
    denom = np.maximum(np.abs(resistances), 1e-9)
    rel_error = np.abs(expected - resistances) / denom
    finite = rel_error[np.isfinite(rel_error)]
    if finite.size == 0:
        return False, None
    mean_error = float(np.mean(finite))
    return mean_error < R_CHECK_THRESHOLD, mean_error


def _value_for_output(record: Dict[str, object], field: str) -> Optional[object]:
    if not record:
        return None
    value = record.get(field)
    if value is None or _is_nan(value):
        raw = record.get(f"{field}_raw")
        if raw is None:
            return None
        text = str(raw).strip()
        return text or None
    return value


def _compose_notes(*records: Dict[str, object]) -> Optional[str]:
    notes: List[str] = []
    seen: set[str] = set()
    for record in records:
        if not record:
            continue
        for key in ("bistable_status", "notes"):
            candidate = _value_for_output(record, key)
            if candidate is None:
                continue
            text = str(candidate).strip()
            if not text:
                continue
            lowered = text.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            notes.append(text)
    return "; ".join(notes) if notes else None


def _load_strain_records(
    paths: Sequence[Path],
    log: logging.Logger,
) -> Dict[Tuple[str, int, int, Optional[str]], StrainRecord]:
    records: Dict[Tuple[str, int, int, Optional[str]], StrainRecord] = {}
    if not paths:
        return records
    seen: set[Path] = set()
    for raw_path in paths:
        path = Path(raw_path)
        if path in seen:
            continue
        seen.add(path)
        if not path.exists():
            log.warning("Strain worksheet %s not found; skipping", path)
            continue
        try:
            df = pd.read_excel(path)
        except Exception:
            log.exception("Failed to read strain worksheet %s", path)
            continue
        if df.empty:
            continue
        column_map: Dict[str, int] = {}
        for idx, header in enumerate(df.columns):
            key = _clean_str(header).lower()
            if not key:
                continue
            if "composition" in key:
                column_map.setdefault("composition", idx)
            elif "microwire" in key or "wire" in key:
                column_map.setdefault("microwire", idx)
            elif key.startswith("m") and "length" in key:
                column_map.setdefault("m_length", idx)
            elif key.startswith("a") and "length" in key:
                column_map.setdefault("a_length", idx)
            elif "strain" in key or "%" in key:
                column_map.setdefault("strain", idx)
            elif "broke" in key or "status" in key or "note" in key:
                column_map.setdefault("status", idx)
        comp_idx = column_map.get("composition")
        micro_idx = column_map.get("microwire")
        if comp_idx is None or micro_idx is None:
            log.warning("%s is missing composition or microwire columns; skipping", path)
            continue
        for row in df.itertuples(index=False, name=None):
            values = list(row)
            composition = _clean_str(values[comp_idx])
            if not composition or composition.lower() == "composition":
                continue
            microwire_label = _clean_str(values[micro_idx])
            if not microwire_label:
                continue
            draw_parts = _microwire_parts_from_label(microwire_label)
            if not draw_parts:
                continue
            draw, piece, suffix = draw_parts
            m_length = (
                _parse_strain_float(values[column_map["m_length"]])
                if "m_length" in column_map
                else None
            )
            a_length = (
                _parse_strain_float(values[column_map["a_length"]])
                if "a_length" in column_map
                else None
            )
            percent = (
                _parse_strain_float(values[column_map["strain"]])
                if "strain" in column_map
                else None
            )
            broke = False
            status_idx = column_map.get("status")
            if status_idx is not None:
                if _clean_str(values[status_idx]).lower() == "broke":
                    broke = True
            if not broke:
                for value in values:
                    if _clean_str(value).lower() == "broke":
                        broke = True
                        break
            if percent is None and not broke and m_length not in (None, 0) and a_length is not None:
                try:
                    percent = ((m_length - a_length) / m_length) * 100 if m_length else None
                except ZeroDivisionError:
                    percent = None
                if percent is not None and not math.isfinite(percent):
                    percent = None
            record = StrainRecord(
                composition=composition,
                draw=draw,
                piece=piece,
                microwire_label=microwire_label,
                m_length=m_length,
                a_length=a_length,
                percent=percent,
                broke=broke,
                source=path,
            )
            records[(composition, draw, piece, suffix)] = record
    return records


def _microwire_label(
    draw_x: Optional[int],
    piece_y: Optional[int],
    suffix: Optional[str] = None,
) -> str:
    if draw_x is None or piece_y is None:
        return ""
    suffix_text = str(suffix).strip() if suffix is not None else ""
    if suffix_text:
        return f"{draw_x}/{piece_y}{suffix_text}"
    return f"{draw_x}/{piece_y}"


def _select_high_measurement(records: List[MeasurementRecord]) -> Optional[MeasurementRecord]:
    if not records:
        return None
    exact_candidates = [
        record for record in records if getattr(record.metadata, "setpoint_mA", None) == 1000
    ]
    if not exact_candidates:
        return None

    def key(record: MeasurementRecord) -> Tuple[int, str]:
        variant_penalty = 0 if not record.metadata.alt_variant else 1
        return (variant_penalty, record.metadata.file_name.lower())

    return min(exact_candidates, key=key)


def _select_low_measurement(records: List[MeasurementRecord]) -> Optional[MeasurementRecord]:
    if not records:
        return None
    candidates = [r for r in records if r.metadata.setpoint_mA is not None]
    if not candidates:
        return None

    def key(record: MeasurementRecord) -> Tuple[int, int, str]:
        setpoint = record.metadata.setpoint_mA or 0
        variant_penalty = 0 if not record.metadata.alt_variant else 1
        return (setpoint, variant_penalty, record.metadata.file_name.lower())

    return min(candidates, key=key)


def _select_other_measurements(
    records: Sequence[MeasurementRecord],
    high_record: Optional[MeasurementRecord],
    low_record: Optional[MeasurementRecord] = None,
) -> List[MeasurementRecord]:
    excluded_ids = {id(high_record)}

    def key(record: MeasurementRecord) -> Tuple[int, float, str]:
        setpoint = getattr(getattr(record, "metadata", object()), "setpoint_mA", None)
        try:
            setpoint_value = float(setpoint) if setpoint is not None else math.inf
        except (TypeError, ValueError):
            setpoint_value = math.inf
        file_name = getattr(getattr(record, "metadata", object()), "file_name", "")
        return (0 if math.isfinite(setpoint_value) else 1, setpoint_value, str(file_name).lower())

    remaining = [record for record in records if id(record) not in excluded_ids]
    return sorted(remaining, key=key)


def _measurement_cache_key(record: MeasurementRecord) -> str:
    path = getattr(record, "path", None)
    if path:
        return str(path)
    metadata = getattr(record, "metadata", None)
    measurement_id = getattr(metadata, "measurement_id", None)
    if measurement_id:
        return str(measurement_id)
    return repr(record)


def _plot_measurement_matplotlib(
    df: pd.DataFrame,
    source: Path,
    plot_dir: Path,
    figsize: Tuple[float, float],
) -> Path:
    import matplotlib

    try:
        matplotlib.use("Agg", force=True)
    except Exception:
        pass
    import matplotlib.pyplot as plt

    matplotlib.rcParams["figure.max_open_warning"] = 0

    from plotting.plugins.current_annealing.core import plot_one
    from plotting.shared.toolkit import format_annealing_title

    plot_dir.mkdir(parents=True, exist_ok=True)
    title = format_annealing_title(source.stem)
    if "I_mA" in df.columns:
        currents = pd.to_numeric(df["I_mA"], errors="coerce")
    else:
        currents = _series_to_mA(df["I_A"])
    plot_df = pd.DataFrame({"I_mA": currents, "R_Ohm": pd.to_numeric(df["R_ohm"], errors="coerce")}).dropna()
    dpi_target = 192
    target_px = (
        max(int(figsize[0] * dpi_target), 200),
        max(int(figsize[1] * dpi_target), 120),
    )
    fig, fname = plot_one(plot_df, title, target_px=target_px, show_power_top_axis=True)
    safe_stem = _safe_plot_stem(fname)
    plot_path = plot_dir / f"{safe_stem}.png"
    plot_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(plot_path, dpi=300)
    plt.close(fig)
    return plot_path


def _normalise_sort_value(value: object) -> object:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, str):
        match = MICROWIRE_SORT_RE.match(value)
        if match:
            try:
                draw = int(match.group(1))
                piece = int(match.group(2))
            except (TypeError, ValueError):
                draw = piece = None
            else:
                suffix = match.group(3) or ""
                suffix = suffix.strip().lower()
                return f"{draw:05d}/{piece:05d}{suffix}"
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    return value


def _remap_highlights(
    highlights: Mapping[str, Set[int]],
    columns: Set[str],
    index_map: Optional[Dict[int, int]],
) -> Dict[str, Set[int]]:
    remapped: Dict[str, Set[int]] = {}
    for column, rows in highlights.items():
        if column not in columns:
            continue
        if index_map:
            updated = {index_map[row] for row in rows if row in index_map}
        else:
            updated = set(rows)
        if updated:
            remapped[column] = updated
    return remapped


def _apply_output_preferences(
    dataframe: pd.DataFrame,
    *,
    column_filter: Optional[Sequence[str]] = None,
    column_order: Optional[Sequence[str]] = None,
    sort_spec: Optional[Sequence[Tuple[str, bool]]] = None,
    highlight_map: Optional[Mapping[str, Set[int]]] = None,
) -> Tuple[pd.DataFrame, Dict[str, Set[int]]]:
    frame = dataframe.copy()
    highlights = dict(highlight_map) if highlight_map else {}
    index_map: Optional[Dict[int, int]] = None
    if sort_spec:
        sort_columns: List[str] = []
        ascending: List[bool] = []
        for entry in sort_spec:
            if not entry:
                continue
            column = entry[0]
            if not isinstance(column, str):
                continue
            if column not in frame.columns or column in sort_columns:
                continue
            sort_columns.append(column)
            ascending.append(bool(entry[1]) if len(entry) > 1 else True)
        if sort_columns:
            sorted_frame = frame.sort_values(
                by=sort_columns,
                ascending=ascending,
                kind="mergesort",
                key=lambda col: col.map(_normalise_sort_value) if hasattr(col, "map") else col,
            )
            index_map = {
                int(old_idx): int(new_idx)
                for new_idx, old_idx in enumerate(sorted_frame.index)
            }
            frame = sorted_frame.reset_index(drop=True)
    if column_filter:
        filtered = [column for column in column_filter if column in frame.columns]
        frame = frame.loc[:, filtered]
    if column_order:
        ordered = [column for column in column_order if column in frame.columns]
        remaining = [column for column in frame.columns if column not in ordered]
        frame = frame.loc[:, ordered + remaining]
    if highlights:
        highlights = _remap_highlights(highlights, set(frame.columns), index_map)
    return frame, highlights


def _update_existing_csv_with_strain(
    path: Path,
    strain_records: Dict[Tuple[str, int, int, Optional[str]], StrainRecord],
    output_columns: Sequence[str],
    log: logging.Logger,
) -> None:
    try:
        df = pd.read_csv(path)
    except Exception:
        log.exception("Failed to read existing CSV at %s; skipping update", path)
        return
    if df.empty:
        return
    for column_name in (
        VSM_HYSTERESIS_COLUMN,
        VSM_TEMPERATURE_SCAN_COLUMN,
        DMA_ISOSTRESS_COLUMN,
        FMR_COLUMN,
    ):
        if column_name in output_columns and column_name not in df.columns:
            df[column_name] = None
    if STRAIN_COLUMN not in df.columns:
        insert_index = len(df.columns)
        if STRAIN_COLUMN in output_columns:
            insert_index = min(len(df.columns), list(output_columns).index(STRAIN_COLUMN))
        df.insert(insert_index, STRAIN_COLUMN, None)
    if {"Composition", "Microwire"}.issubset(df.columns) and strain_records:
        for idx, row in df.iterrows():
            composition = _clean_str(row.get("Composition"))
            microwire_label = _clean_str(row.get("Microwire"))
            key_parts = _microwire_parts_from_label(microwire_label)
            if not composition or not key_parts:
                continue
            record = strain_records.get(
                (composition, key_parts[0], key_parts[1], key_parts[2])
            )
            if record is None:
                continue
            df.at[idx, STRAIN_COLUMN] = _format_strain_value(record)
    desired = [column for column in output_columns if column in df.columns]
    for column in df.columns:
        if column not in desired:
            desired.append(column)
    df = df[desired]
    df.to_csv(path, index=False)


def _update_existing_excel_with_strain(
    path: Path,
    strain_records: Dict[Tuple[str, int, int, Optional[str]], StrainRecord],
    log: logging.Logger,
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("openpyxl is required to update Excel exports") from exc

    wb = load_workbook(path)
    ws = cast(Any, wb.active)
    if ws.max_row < 1:
        wb.save(path)
        return

    def _headers() -> list[str]:
        return [str(cell.value) if cell.value is not None else "" for cell in ws[1]]

    headers = _headers()
    try:
        d_over_d_index = headers.index("d/D") + 1
    except ValueError:
        d_over_d_index = None
    strain_target = len(headers) + 1 if d_over_d_index is None else d_over_d_index + 1
    if d_over_d_index is not None:
        for column_name in FIGURE_COLUMNS:
            if column_name not in OUTPUT_COLUMNS:
                continue
            headers = _headers()
            if column_name not in headers:
                continue
            current_index = headers.index(column_name) + 1
            if current_index != strain_target:
                offset = strain_target - current_index
                col_letter = get_column_letter(current_index)
                ws.move_range(
                    f"{col_letter}1:{col_letter}{ws.max_row}",
                    rows=0,
                    cols=offset,
                )
            strain_target += 1

    for column_name in (
        VSM_HYSTERESIS_COLUMN,
        VSM_TEMPERATURE_SCAN_COLUMN,
        DMA_ISOSTRESS_COLUMN,
        FMR_COLUMN,
        STRAIN_COLUMN,
    ):
        if column_name not in OUTPUT_COLUMNS:
            continue
        headers = _headers()
        if column_name in headers:
            current_index = headers.index(column_name) + 1
            if current_index != strain_target:
                offset = strain_target - current_index
                col_letter = get_column_letter(current_index)
                ws.move_range(
                    f"{col_letter}1:{col_letter}{ws.max_row}",
                    rows=0,
                    cols=offset,
                )
        else:
            ws.insert_cols(strain_target)
            ws.cell(row=1, column=strain_target).value = column_name
        strain_target += 1

    headers = _headers()
    try:
        composition_index = headers.index("Composition") + 1
        microwire_index = headers.index("Microwire") + 1
        strain_index = headers.index(STRAIN_COLUMN) + 1
    except ValueError:
        log.warning("Unable to locate Composition/Microwire columns in %s; skipping strain update", path)
        wb.save(path)
        return

    if not strain_records:
        wb.save(path)
        return

    for row_idx in range(2, ws.max_row + 1):
        composition = _clean_str(ws.cell(row=row_idx, column=composition_index).value)
        microwire_label = _clean_str(ws.cell(row=row_idx, column=microwire_index).value)
        key_parts = _microwire_parts_from_label(microwire_label)
        if not composition or not key_parts:
            continue
        record = strain_records.get(
            (composition, key_parts[0], key_parts[1], key_parts[2])
        )
        if record is None:
            continue
        ws.cell(row=row_idx, column=strain_index).value = _format_strain_value(record)

    wb.save(path)


def _plot_measurement_origin(
    df: pd.DataFrame,
    source: Path,
    origin_dir: Path,
    log: Optional[logging.Logger] = None,
) -> Optional[OriginArtifact]:
    try:
        from plotting.plugins.current_annealing.core import plot_one_origin
        from plotting.shared.utils import format_annealing_title
    except ImportError as exc:  # pragma: no cover - depends on optional module
        raise RuntimeError("originpro is not available") from exc

    if "I_mA" in df.columns:
        currents = pd.to_numeric(df["I_mA"], errors="coerce")
    else:
        currents = _series_to_mA(df["I_A"])
    plot_df = pd.DataFrame({"I_mA": currents, "R_Ohm": pd.to_numeric(df["R_ohm"], errors="coerce")}).dropna()
    title = format_annealing_title(source.stem)
    handles = plot_one_origin(
        plot_df,
        title,
        source.name,
        show_power_top_axis=True,
        return_handles=True,
    )

    if not isinstance(handles, dict):
        return None

    safe_stem = _safe_plot_stem(source.stem)
    return export_origin_graph_artifact(
        handles=handles,
        descriptor_stem=safe_stem,
        origin_dir=origin_dir,
        display_text=None,
        log=log,
    )


def _describe_origin_handles(
    handles: Dict[str, object],
) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    graph = handles.get("graph")
    workbook = handles.get("workbook")
    worksheet = handles.get("worksheet")
    legend_label = handles.get("legend_label")

    parts: list[str] = []
    if legend_label:
        parts.append(str(legend_label))
    graph_name = _origin_object_name(graph)
    if graph_name:
        parts.append(f"Graph: {graph_name}")
    workbook_name = _origin_object_name(workbook)
    if workbook_name:
        parts.append(f"Book: {workbook_name}")
    worksheet_name = _origin_object_name(worksheet)
    if worksheet_name:
        parts.append(f"Sheet: {worksheet_name}")
    description = " | ".join(parts) if parts else None
    return description, graph_name, workbook_name, worksheet_name


def _export_origin_object(
    handles: Dict[str, object], target_path: Path, log: Optional[logging.Logger]
) -> Optional[Path]:
    origin_any = handles.get("origin")
    graph = handles.get("graph")
    if origin_any is None or graph is None:
        return None

    target_path = target_path.resolve()
    target_path.parent.mkdir(parents=True, exist_ok=True)
    attempted: list[str] = []

    def _path_created() -> Optional[Path]:
        candidates = [
            target_path,
            target_path.with_suffix(".oggu"),
            target_path.with_suffix(".opju"),
        ]
        for candidate in candidates:
            try:
                if candidate.exists():
                    return candidate
            except OSError:
                continue
        return None

    for attr in ("save_copy", "save_as", "save"):
        method = getattr(graph, attr, None)
        if callable(method):
            try:
                method(str(target_path))
                created = _path_created()
                if created is not None:
                    return created
            except Exception:
                attempted.append(attr)

    for attr in ("save_page", "save_window"):
        method = getattr(origin_any, attr, None)
        if callable(method):
            try:
                method(str(target_path))
                created = _path_created()
                if created is not None:
                    return created
            except Exception:
                attempted.append(attr)

    lt_executors = (
        getattr(graph, "lt_exec", None),
        getattr(origin_any, "lt_exec", None),
    )
    if any(callable(executor) for executor in lt_executors):
        graph_name = _origin_object_name(graph) or ""
        base = str(target_path)
        commands = []
        if graph_name:
            commands.append(f"page -s \"{graph_name}\"; save -i \"{base}\";")
            commands.append(f"page -s \"{graph_name}\"; save -ix \"{base}\";")
            commands.append(f"page -s \"{graph_name}\"; save -oggu \"{base}\";")
            commands.append(f"page -s \"{graph_name}\"; save -opju \"{base}\";")
        commands.append(f"save -i \"{base}\";")
        commands.append(f"save -ix \"{base}\";")
        commands.append(f"save -oggu \"{base}\";")
        commands.append(f"save -opju \"{base}\";")
        for command in commands:
            for lt_exec in lt_executors:
                if not callable(lt_exec):
                    continue
                try:
                    lt_exec(command)
                    created = _path_created()
                    if created is not None:
                        return created
                    attempted.append(f"{command} (no file)")
                except Exception:
                    attempted.append(command)

    if log is not None:
        log.warning(
            "Origin graph export failed for %s; attempted %s",
            target_path,
            ", ".join(attempted) if attempted else "no-export",
        )
    return None


def _collapse_asset_references(values: Sequence[str]) -> str | list[str] | None:
    cleaned: List[str] = []
    for value in values:
        if not isinstance(value, str):
            continue
        stripped = value.strip()
        if stripped and stripped not in cleaned:
            cleaned.append(stripped)
    if not cleaned:
        return None
    if len(cleaned) == 1:
        return cleaned[0]
    return cleaned


def _single_asset_reference(value: Any) -> Optional[str]:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, (list, tuple, set)):
        resolved: List[str] = []
        for item in value:
            candidate = _single_asset_reference(item)
            if candidate and candidate not in resolved:
                resolved.append(candidate)
        if len(resolved) == 1:
            return resolved[0]
    return None


def _copy_origin_graph_page(
    graph: object,
    log: Optional[logging.Logger] = None,
) -> bool:
    if graph is None:
        return False
    try:
        graph.activate()  # type: ignore[attr-defined]
    except Exception:
        pass
    lt_exec = getattr(graph, "lt_exec", None)
    if callable(lt_exec):
        try:
            # Keep master-page items from forcing an image-only clipboard payload.
            lt_exec("@GMC=1;")
        except Exception:
            pass
    copy_page = getattr(graph, "copy_page", None)
    if not callable(copy_page):
        return False
    try:
        copy_page("OLE", ratio=85)
        return True
    except TypeError:
        try:
            copy_page("OLE")
            return True
        except Exception:
            pass
    except Exception:
        pass
    if log is not None:
        log.debug("Failed to copy Origin graph page %s as OLE", _origin_object_name(graph))
    return False


def export_origin_graph_artifact(
    *,
    handles: Dict[str, object],
    descriptor_stem: str,
    origin_dir: Path,
    display_text: Optional[str] = None,
    log: Optional[logging.Logger] = None,
) -> Optional[OriginArtifact]:
    """Save and describe a live Origin graph page for Word/Excel embedding."""

    graph = handles.get("graph")
    if graph is None:
        return None
    description, graph_name, workbook_name, worksheet_name = _describe_origin_handles(handles)
    safe_stem = _safe_plot_stem(descriptor_stem)
    origin_dir.mkdir(parents=True, exist_ok=True)
    artifact_name = f"{safe_stem}.oggu"
    target_path = origin_dir / artifact_name
    exported_path = _export_origin_object(handles, target_path, log)
    clipboard_fallback = _copy_origin_graph_page(graph, log)
    return OriginArtifact(
        descriptor=artifact_name,
        object_path=exported_path,
        graph_name=graph_name,
        workbook_name=workbook_name,
        worksheet_name=worksheet_name,
        display_text=display_text or description,
        clipboard_fallback=clipboard_fallback,
    )


def _pump_qt_events(app: object, *, rounds: int = 3) -> None:
    processor = getattr(app, "processEvents", None)
    if not callable(processor):
        return
    for _ in range(max(1, int(rounds))):
        try:
            processor()
        except Exception:
            break


def export_pyplot_origin_artifacts_for_paths(
    *,
    paths: Sequence[Path | str],
    plugin_name: str,
    origin_dir: Path,
    descriptor_prefix: str,
    display_prefix: str,
    log: Optional[logging.Logger] = None,
    plot_mode: str | None = None,
) -> List[OriginArtifact]:
    """Generate PyPlot graphs for paths and save them as Origin graph artifacts."""

    filtered: List[Path] = []
    seen: Set[Path] = set()
    for raw_path in paths:
        path = Path(raw_path)
        if not path.exists():
            continue
        try:
            key = path.resolve()
        except OSError:
            key = path
        if key in seen:
            continue
        seen.add(key)
        filtered.append(path)
    if not filtered:
        return []

    try:
        from PyQt6 import QtWidgets
        from plotting.pyplot.app import PyPlotWorkbench
    except Exception as exc:  # pragma: no cover - depends on optional GUI stack
        raise RuntimeError("PyPlot automation is unavailable.") from exc

    created_app = False
    app = QtWidgets.QApplication.instance()
    if not isinstance(app, QtWidgets.QApplication):
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        app = QtWidgets.QApplication([sys.argv[0]])
        created_app = True
        try:
            app.setQuitOnLastWindowClosed(False)
        except Exception:
            pass

    window: Any | None = None
    artifacts: List[OriginArtifact] = []
    try:
        window = PyPlotWorkbench(initial_plotter=plugin_name)
        _pump_qt_events(app, rounds=4)
        selector = getattr(window, "automation_select_plugin", None)
        if not callable(selector):
            raise RuntimeError("PyPlot plugin selection automation is unavailable.")
        selector(plugin_name)

        plugin_getter = getattr(window, "_plugin_instance_for_name", None)
        plugin = plugin_getter(plugin_name) if callable(plugin_getter) else None
        use_live_vsm_origin_export = True
        if (
            use_live_vsm_origin_export
            and str(plugin_name).strip().casefold() == "vsm temperature scan"
            and plugin is not None
        ):
            processor = getattr(plugin, "_processor", None)
            for setter_name, value in (
                ("set_show_derivative", False),
                ("set_show_smoothed_derivative", False),
                ("set_show_overlay_derivative", False),
                ("set_show_smoothed", False),
            ):
                setter = getattr(processor, setter_name, None)
                if callable(setter):
                    try:
                        setter(value)
                    except Exception:
                        pass
            for attr in (
                "show_derivative",
                "show_smoothed_derivative",
                "show_overlay_derivative",
                "show_smoothed_plot",
            ):
                try:
                    setattr(processor, attr, False)
                except Exception:
                    pass
            try:
                setter = getattr(processor, "set_split_directions", None)
                if callable(setter):
                    setter(True)
                else:
                    setattr(processor, "split_directions", True)
            except Exception:
                pass
            try:
                setter = getattr(processor, "set_combine_fields", None)
                if callable(setter):
                    setter(True)
                else:
                    setattr(processor, "combine_fields", True)
            except Exception:
                pass

        plugin_key = str(plugin_name).strip().casefold()
        if plugin is not None and plugin_key in {"current annealing", "mini dma"}:
            try:
                panel = getattr(plugin, "panel_widget", None)
                if callable(panel):
                    panel()
            except Exception:
                pass
            try:
                settings = getattr(plugin, "settings_widget", None)
                if callable(settings):
                    settings()
            except Exception:
                pass
            setter = getattr(plugin, "_set_show_power_top_axis", None)
            if callable(setter):
                try:
                    setter(True)
                except Exception:
                    pass
            checkbox = getattr(plugin, "_show_power_top_axis_checkbox", None)
            if checkbox is not None:
                try:
                    checkbox.setChecked(True)
                except Exception:
                    pass

        importer = getattr(window, "automation_import_paths", None)
        if not callable(importer):
            raise RuntimeError("PyPlot import automation is unavailable.")
        importer(filtered)

        if (
            use_live_vsm_origin_export
            and str(plugin_name).strip().casefold() == "vsm temperature scan"
            and plugin is not None
        ):
            processor = getattr(plugin, "_processor", None)
            dataset = getattr(plugin, "_dataset", None)
            if not dataset:
                if log is not None:
                    log.warning("No VSM temperature scan dataset was available for Origin export.")
                return artifacts
            try:
                from plotting.shared.origin import _ensure_origin_sdk_on_path

                _ensure_origin_sdk_on_path()
            except Exception:
                pass
            try:
                import originpro as op  # type: ignore
            except Exception as exc:
                raise RuntimeError("originpro is not available") from exc
            try:
                before_graphs = list(op.graph_list("p", True))
            except Exception:
                before_graphs = []
            if processor is None or not hasattr(processor, "plot_origin"):
                raise RuntimeError("VSM temperature scan Origin export is unavailable.")
            try:
                returned_graphs = processor.plot_origin(dataset, release_origin_on_exit=False)
            except TypeError:
                returned_graphs = processor.plot_origin(dataset)
            _pump_qt_events(app, rounds=6)
            if isinstance(returned_graphs, list) and returned_graphs:
                new_graphs = returned_graphs
            else:
                try:
                    after_graphs = list(op.graph_list("p", True))
                except Exception:
                    after_graphs = []
                new_graphs = after_graphs[len(before_graphs) :] if len(after_graphs) >= len(before_graphs) else after_graphs
            for counter, graph in enumerate(new_graphs, start=1):
                fallback_graph_name = f"WRVSM{counter:02d}"
                if _origin_object_name(graph) is None:
                    for attr, value in (
                        ("name", fallback_graph_name),
                        ("lname", f"{display_prefix} {counter:02d}"),
                    ):
                        try:
                            setattr(graph, attr, value)
                        except Exception:
                            pass
                graph_name = _origin_object_name(graph) or fallback_graph_name
                artifact = export_origin_graph_artifact(
                    handles={
                        "origin": op,
                        "graph": graph,
                        "legend_label": f"{display_prefix}: {graph_name}",
                    },
                    descriptor_stem=_safe_plot_stem(
                        "_".join((descriptor_prefix, graph_name or f"{counter:02d}"))
                    ),
                    origin_dir=origin_dir,
                    display_text=f"{display_prefix}: {graph_name}",
                    log=log,
                )
                if artifact is not None and (
                    artifact.object_path is not None or artifact.clipboard_fallback
                ):
                    artifacts.append(artifact)
            return artifacts

        if str(plugin_name).strip().casefold() == "r vs t" and str(plot_mode or "").strip().casefold() == "residual":
            generator = getattr(plugin, "generate_residuals", None)
        else:
            generator = getattr(window, "automation_generate", None)
        if not callable(generator):
            raise RuntimeError("PyPlot generate automation is unavailable.")
        generator()
        _pump_qt_events(app, rounds=10)

        pruner = getattr(window, "_prune_shared_plot_workbooks", None)
        if callable(pruner):
            pruner()
        workbook_getter = getattr(window, "_shared_plot_workbooks_for_plugin", None)
        workbooks = workbook_getter(plugin_name) if callable(workbook_getter) else []
        if not workbooks:
            current_tab = getattr(window, "tab_widget", None)
            current_widget = current_tab.currentWidget() if current_tab is not None else None
            descriptor = getattr(window, "_tab_descriptors", {}).get(current_widget)
            registrar = getattr(window, "_register_shared_plot_workbook_for_tab", None)
            if current_widget is not None and descriptor is not None and callable(registrar):
                registrar(current_widget, descriptor)
                workbooks = workbook_getter(plugin_name) if callable(workbook_getter) else []
        if not workbooks:
            managed_keys = getattr(plugin, "_managed_workbooks", None)
            workbook_store = getattr(window, "_workbooks", {})
            if managed_keys and isinstance(workbook_store, dict):
                workbooks = [
                    workbook_store[key]
                    for key in sorted(managed_keys, key=str)
                    if key in workbook_store
                ]
        if str(plugin_name).strip().casefold() == "vsm temperature scan":
            workbooks = [
                workbook
                for workbook in workbooks
                if "(tscan)" in str(getattr(workbook, "name", "") or "").casefold()
                and not any(
                    token in str(getattr(workbook, "name", "") or "").casefold()
                    for token in ("derivative", "smoothed", "overlay")
                )
            ]
        requested_paths: List[Path] = []
        for path in filtered:
            try:
                requested_paths.append(path.resolve())
            except OSError:
                requested_paths.append(path)

        def _source_matches_requested(source: object) -> bool:
            if not source:
                return False
            try:
                source_path = Path(str(source)).resolve()
            except OSError:
                source_path = Path(str(source))
            candidates = [source_path]
            candidates.extend(source_path.parents)
            return any(candidate in requested_paths for candidate in candidates)

        matching_workbook_keys: Set[Hashable] = set()
        source_tagged_workbook_keys: Set[Hashable] = set()
        shared_by_tab = getattr(window, "_shared_plot_workbook_by_tab", {})
        tab_descriptors = getattr(window, "_tab_descriptors", {})
        if isinstance(shared_by_tab, Mapping) and isinstance(tab_descriptors, Mapping):
            for tab, key in shared_by_tab.items():
                descriptor = tab_descriptors.get(tab)
                metadata = getattr(descriptor, "metadata", None)
                if not isinstance(metadata, Mapping):
                    continue
                source_file = metadata.get("source_file")
                if not source_file:
                    continue
                source_tagged_workbook_keys.add(key)
                if _source_matches_requested(source_file):
                    matching_workbook_keys.add(key)
        if source_tagged_workbook_keys:
            workbooks = [
                workbook
                for workbook in workbooks
                if getattr(workbook, "key", None) in matching_workbook_keys
            ]
        if not workbooks:
            if log is not None:
                log.warning("No PyPlot workbooks were available for %s Origin export.", plugin_name)
            return artifacts

        counter = 0

        def _capture_artifact(
            origin_any: object,
            graph: object,
            workbook: object,
            worksheet: object,
        ) -> None:
            nonlocal counter
            counter += 1
            workbook_name = str(getattr(workbook, "name", "") or "").strip()
            worksheet_name = str(getattr(worksheet, "name", "") or "").strip()
            label_parts = [part for part in (display_prefix, workbook_name, worksheet_name) if part]
            display_text = (
                ": ".join([label_parts[0], " - ".join(label_parts[1:])])
                if label_parts
                else display_prefix
            )
            descriptor_parts = [
                descriptor_prefix,
                workbook_name or "graph",
                worksheet_name or f"{counter:02d}",
            ]
            artifact = export_origin_graph_artifact(
                handles={
                    "origin": origin_any,
                    "graph": graph,
                    "workbook": workbook,
                    "worksheet": worksheet,
                    "legend_label": display_text,
                },
                descriptor_stem=_safe_plot_stem("_".join(descriptor_parts)),
                origin_dir=origin_dir,
                display_text=display_text,
                log=log,
            )
            if artifact is not None and (
                artifact.object_path is not None or artifact.clipboard_fallback
            ):
                artifacts.append(artifact)

        pusher = getattr(window, "_push_workbooks_to_origin", None)
        if not callable(pusher):
            raise RuntimeError("PyPlot Origin workbook export is unavailable.")
        _exported, _plotted, errors = pusher(
            workbooks,
            create_graphs=True,
            graph_callback=_capture_artifact,
            keep_origin_open=True,
            release_origin_on_exit=False,
        )
        if log is not None:
            for error in errors or []:
                log.warning("%s Origin export warning: %s", plugin_name, error)
        _pump_qt_events(app, rounds=4)
        return artifacts
    finally:
        if window is not None:
            try:
                clear_dirty = getattr(window, "_clear_project_dirty", None)
                if callable(clear_dirty):
                    clear_dirty()
            except Exception:
                pass
            try:
                window.close()
            except Exception:
                pass
        if isinstance(app, QtWidgets.QApplication):
            _pump_qt_events(app, rounds=4)
            if created_app:
                try:
                    app.quit()
                except Exception:
                    pass


def _asset_references(value: Any) -> List[str]:
    if isinstance(value, str):
        stripped = value.strip()
        return [stripped] if stripped else []
    if isinstance(value, (list, tuple, set)):
        resolved: List[str] = []
        for item in value:
            for candidate in _asset_references(item):
                if candidate and candidate not in resolved:
                    resolved.append(candidate)
        return resolved
    return []


_WORD_REPORT_LABELS: Dict[str, str] = {
    DIAMETER_COLUMN: "d (um)",
    GLASS_DIAMETER_COLUMN: "D (um)",
    DIAMETER_RATIO_COLUMN: "d/D",
    RVT_FILE_COLUMN: "R vs T source files",
    RVT_GRAPH_COLUMN: "R vs T graphs",
    RVT_ORIGIN_COLUMN: "R vs T graphs (Origin)",
    RVT_RESIDUAL_ORIGIN_COLUMN: "R vs T residual graphs (Origin)",
    VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN: "VSM temperature scan graphs (Origin)",
    VSM_HYSTERESIS_ORIGIN_COLUMN: "VSM hysteresis graphs (Origin)",
    DMA_ISOSTRESS_ORIGIN_COLUMN: "DMA iso-stress graphs (Origin)",
    MINI_DMA_ORIGIN_COLUMN: "Mini DMA graphs (Origin)",
    SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN: "Manual stress/strain graphs (Origin)",
    FMR_ORIGIN_COLUMN: "FMR graphs (Origin)",
    RVT_POINT_COUNT_COLUMN: "R vs T points",
    RVT_TEMPERATURE_RANGE_COLUMN: "R vs T temperature range (deg C)",
    RVT_RESISTANCE_RANGE_COLUMN: "R vs T resistance range (Ohm)",
    CORE_TEMPERATURE_COLUMN: "Core temperature (deg C)",
    GLASS_TEMPERATURE_COLUMN: "Glass temperature (deg C)",
    "Resistance (Î©)": "Resistance (Ohm)",
    "As (Â°C)": "As (deg C)",
    "Af (Â°C)": "Af (deg C)",
    "Ms (Â°C)": "Ms (deg C)",
    "Mf (Â°C)": "Mf (deg C)",
    "Figure â€” 1000 mA": "Figure - 1000 mA",
    "Figure â€” other annealing": "Figure - other annealing",
    "Figure â€” 1000 mA (Origin)": "Figure - 1000 mA (Origin)",
    "Figure â€” other annealing (Origin)": "Figure - other annealing (Origin)",
}

_WORD_IDENTITY_COLUMNS: Tuple[str, ...] = (
    "Composition",
    "Microwire",
    "e/a",
    ESTIMATED_TRANSITION_COLUMN,
)

_WORD_DIMENSION_COLUMNS: Tuple[str, ...] = (
    DIAMETER_COLUMN,
    GLASS_DIAMETER_COLUMN,
    DIAMETER_RATIO_COLUMN,
)

_WORD_FABRICATION_COLUMNS: Tuple[str, ...] = (
    "Length (m)",
    "Production datetime",
    "Mass (g)",
    "Resistance (Î©)",
    CORE_TEMPERATURE_COLUMN,
    GLASS_TEMPERATURE_COLUMN,
    "Winding speed (m/min)",
    "Glass feeding (mm/min)",
    "Underpressure",
    GLASS_PULL_COLUMN,
    VIDEO_END_LENGTH_COLUMN,
    VIDEO_MW_LENGTH_COLUMN,
)

_WORD_FUNCTIONAL_COLUMNS: Tuple[str, ...] = (
    BRITTLE_COLUMN,
    STRAIN_COLUMN,
    "As (mA)",
    "Ms (mA)",
    *CURRENT_DENSITY_EXTRA_COLUMNS,
    *TRANSITION_TEMP_COLUMNS,
    ANNEALING_TRANSITION_COLUMN,
    *STRAIN_EXTRA_COLUMNS,
    *SHAPE_MEMORY_VALUE_COLUMNS,
    *SHAPE_MEMORY_FRACTURE_COLUMNS,
)

_WORD_PROVENANCE_COLUMNS: Tuple[str, ...] = (
    "File 1000 mA",
    "Other annealing files",
    "Data source",
    RVT_FILE_COLUMN,
    RVT_POINT_COUNT_COLUMN,
    RVT_TEMPERATURE_RANGE_COLUMN,
    RVT_RESISTANCE_RANGE_COLUMN,
)

_WORD_GRAPH_COLUMNS: Tuple[str, ...] = (
    *FIGURE_COLUMNS,
    RVT_GRAPH_COLUMN,
    RVT_ORIGIN_COLUMN,
    VSM_HYSTERESIS_COLUMN,
    VSM_HYSTERESIS_ORIGIN_COLUMN,
    VSM_TEMPERATURE_SCAN_COLUMN,
    VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN,
    DMA_ISOSTRESS_COLUMN,
    DMA_ISOSTRESS_ORIGIN_COLUMN,
    MINI_DMA_COLUMN,
    MINI_DMA_ORIGIN_COLUMN,
    SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
    SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN,
    LEGACY_SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
    LEGACY_SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN,
    FMR_COLUMN,
    FMR_ORIGIN_COLUMN,
)

_CURRENT_ANNEALING_ORIGIN_COLUMNS = tuple(f"{column} (Origin)" for column in FIGURE_COLUMNS)

_WORD_GRAPH_SECTIONS: Tuple[
    Tuple[str, Tuple[str, ...], Tuple[str, ...]],
    ...,
] = (
    ("Current annealing", _CURRENT_ANNEALING_ORIGIN_COLUMNS, FIGURE_COLUMNS),
    ("R vs T", (RVT_ORIGIN_COLUMN, RVT_RESIDUAL_ORIGIN_COLUMN), (RVT_GRAPH_COLUMN,)),
    ("VSM temperature scan", (VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN,), (VSM_TEMPERATURE_SCAN_COLUMN,)),
    ("VSM hysteresis loops", (VSM_HYSTERESIS_ORIGIN_COLUMN,), (VSM_HYSTERESIS_COLUMN,)),
    ("DMA iso-stress", (DMA_ISOSTRESS_ORIGIN_COLUMN,), (DMA_ISOSTRESS_COLUMN,)),
    ("Mini DMA", (MINI_DMA_ORIGIN_COLUMN,), (MINI_DMA_COLUMN,)),
    (
        "Manual stress/strain",
        (
            SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN,
            LEGACY_SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN,
        ),
        (
            SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
            LEGACY_SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
        ),
    ),
    ("FMR", (FMR_ORIGIN_COLUMN,), (FMR_COLUMN,)),
)


@dataclass(frozen=True)
class WordGraphSectionEvaluation:
    title: str
    included: bool
    status: str
    reason: str
    accepted_origin_descriptors: Tuple[str, ...] = ()
    accepted_references: Tuple[str, ...] = ()
    invalid_origin_descriptors: Tuple[str, ...] = ()
    missing_origin_descriptors: Tuple[str, ...] = ()
    invalid_references: Tuple[str, ...] = ()


_WORD_ALWAYS_OMIT_COLUMNS: Tuple[str, ...] = (
    *_WORD_PROVENANCE_COLUMNS,
    *_WORD_GRAPH_COLUMNS,
    *ORIGIN_FIGURE_COLUMNS,
    *MICROSCOPE_IMAGE_COLUMNS,
)


def _word_label(column: str) -> str:
    return _WORD_REPORT_LABELS.get(column, column)


def _word_xml_escape(value: object) -> str:
    text = str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _word_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        stripped = value.strip()
        return bool(stripped) and stripped.lower() != "nan"
    if isinstance(value, (list, tuple, set)):
        return any(_word_has_value(item) for item in value)
    try:
        missing = pd.isna(value)
    except Exception:
        missing = False
    if isinstance(missing, (bool, np.bool_)) and missing:
        return False
    return True


def _word_format_value(value: Any) -> str:
    if not _word_has_value(value):
        return ""
    if isinstance(value, (list, tuple, set)):
        parts = [_word_format_value(item) for item in value if _word_has_value(item)]
        return ", ".join(part for part in parts if part)
    if isinstance(value, (pd.Timestamp, datetime)):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return value.isoformat()
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        numeric = float(value)
        if not math.isfinite(numeric):
            return ""
        return f"{numeric:.6g}"
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _word_format_value_cells(value: Any) -> List[str]:
    if not _word_has_value(value):
        return [""]
    if isinstance(value, (list, tuple, set)):
        cells = [_word_format_value(item) for item in value if _word_has_value(item)]
        return cells or [""]
    return [_word_format_value(value)]


def _word_key_values(row: pd.Series, columns: Sequence[str]) -> List[Tuple[str, str]]:
    values: List[Tuple[str, str]] = []
    for column in columns:
        if column not in row.index:
            continue
        value = row.get(column)
        if not _word_has_value(value):
            continue
        formatted = _word_format_value(value)
        if formatted:
            values.append((_word_label(column), formatted))
    return values


def _word_sample_title(row: pd.Series, fallback_index: int) -> str:
    composition = _word_format_value(row.get("Composition"))
    microwire = _word_format_value(row.get("Microwire"))
    title = " ".join(part for part in (composition, microwire) if part).strip()
    if title:
        return title
    return f"Sample {fallback_index + 1}"


def _word_report_filename(row: pd.Series, fallback_index: int) -> str:
    title = _word_sample_title(row, fallback_index).replace("/", "-")
    return f"{_safe_plot_stem(title).replace(' ', '_')}.docx"


def _word_run(text: str, *, bold: bool = False, size: int | None = None) -> str:
    props: List[str] = []
    if bold:
        props.append("<w:b/>")
    if size is not None:
        props.append(f'<w:sz w:val="{int(size)}"/>')
    prop_xml = f"<w:rPr>{''.join(props)}</w:rPr>" if props else ""
    return f"<w:r>{prop_xml}<w:t>{_word_xml_escape(text)}</w:t></w:r>"


def _word_field_run(field_name: str) -> str:
    safe_field = re.sub(r"[^A-Z0-9_ ]", "", str(field_name or "").upper()).strip()
    return (
        "<w:r><w:fldChar w:fldCharType=\"begin\"/></w:r>"
        f"<w:r><w:instrText xml:space=\"preserve\"> {safe_field} </w:instrText></w:r>"
        "<w:r><w:fldChar w:fldCharType=\"separate\"/></w:r>"
        "<w:r><w:t>1</w:t></w:r>"
        "<w:r><w:fldChar w:fldCharType=\"end\"/></w:r>"
    )


def _word_paragraph(
    text: str = "",
    *,
    bold: bool = False,
    size: int | None = None,
    spacing_after: int = 160,
    bookmark_name: str | None = None,
    bookmark_id: int | None = None,
    style: str | None = None,
) -> str:
    style_xml = f'<w:pStyle w:val="{_word_xml_escape(style)}"/>' if style else ""
    ppr = f'<w:pPr>{style_xml}<w:spacing w:after="{int(spacing_after)}"/></w:pPr>'
    bookmark_start = ""
    bookmark_end = ""
    if bookmark_name and bookmark_id is not None:
        safe_name = re.sub(r"[^A-Za-z0-9_]", "_", bookmark_name)
        bookmark_start = f'<w:bookmarkStart w:id="{int(bookmark_id)}" w:name="{safe_name}"/>'
        bookmark_end = f'<w:bookmarkEnd w:id="{int(bookmark_id)}"/>'
    return f"<w:p>{ppr}{bookmark_start}{_word_run(text, bold=bold, size=size)}{bookmark_end}</w:p>"


def _word_page_break() -> str:
    return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'


def _word_header_xml(title: str) -> str:
    paragraph = (
        '<w:p><w:pPr><w:pStyle w:val="Header"/>'
        '<w:jc w:val="right"/></w:pPr>'
        f'{_word_run(title, size=18)}</w:p>'
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:hdr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"{paragraph}</w:hdr>"
    )


def _word_footer_xml() -> str:
    paragraph = (
        '<w:p><w:pPr><w:pStyle w:val="Footer"/>'
        '<w:jc w:val="center"/></w:pPr>'
        f'{_word_run("Page ", size=18)}{_word_field_run("PAGE")}'
        f'{_word_run(" of ", size=18)}{_word_field_run("NUMPAGES")}</w:p>'
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"{paragraph}</w:ftr>"
    )


def _word_table(rows: Sequence[Tuple[str, str]]) -> str:
    if not rows:
        return ""
    border = (
        '<w:top w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="E5E7EB"/>'
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="E5E7EB"/>'
    )
    table_rows: List[str] = []
    for label, value in rows:
        label_cell = (
            '<w:tc><w:tcPr><w:tcW w:w="2600" w:type="dxa"/>'
            '<w:shd w:fill="F3F4F6"/></w:tcPr>'
            f'{_word_paragraph(label, bold=True, spacing_after=0)}</w:tc>'
        )
        value_cell = (
            '<w:tc><w:tcPr><w:tcW w:w="6200" w:type="dxa"/></w:tcPr>'
            f'{_word_paragraph(value, spacing_after=0)}</w:tc>'
        )
        table_rows.append(f"<w:tr>{label_cell}{value_cell}</w:tr>")
    return (
        "<w:tbl>"
        '<w:tblPr><w:tblW w:w="0" w:type="auto"/>'
        f"<w:tblBorders>{border}</w:tblBorders>"
        '<w:tblCellMar><w:top w:w="90" w:type="dxa"/><w:left w:w="90" w:type="dxa"/>'
        '<w:bottom w:w="90" w:type="dxa"/><w:right w:w="90" w:type="dxa"/></w:tblCellMar>'
        "</w:tblPr>"
        f"{''.join(table_rows)}"
        "</w:tbl>"
        + _word_paragraph("", spacing_after=120)
    )


def _word_microwire_data_table(rows: Sequence[Tuple[str, Sequence[str]]]) -> str:
    if not rows:
        return ""
    border = (
        '<w:top w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="D0D7DE"/>'
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="E5E7EB"/>'
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="E5E7EB"/>'
    )
    def _cell(text: str, *, label: bool = False, width: int = 2100) -> str:
        shading = '<w:shd w:fill="F3F4F6"/>' if label else ""
        return (
            '<w:tc><w:tcPr>'
            f'<w:tcW w:w="{int(width)}" w:type="dxa"/>{shading}</w:tcPr>'
            f'{_word_paragraph(text, bold=label, size=16, spacing_after=0)}</w:tc>'
        )

    table_rows: List[str] = []
    pending_pairs: List[Tuple[str, str]] = []

    def _flush_pairs() -> None:
        nonlocal pending_pairs
        if not pending_pairs:
            return
        midpoint = (len(pending_pairs) + 1) // 2
        left_pairs = pending_pairs[:midpoint]
        right_pairs = pending_pairs[midpoint:]
        pending_pairs = []
        for index, pair in enumerate(left_pairs):
            cells = []
            label, value = pair
            cells.append(_cell(label, label=True, width=1900))
            cells.append(_cell(value, width=2500))
            if index < len(right_pairs):
                label, value = right_pairs[index]
                cells.append(_cell(label, label=True, width=1900))
                cells.append(_cell(value, width=2500))
            while len(cells) < 4:
                cells.append(_cell("", label=True, width=1900))
                cells.append(_cell("", width=2500))
            table_rows.append(f"<w:tr>{''.join(cells)}</w:tr>")

    for label, raw_values in rows:
        values = [str(value) for value in list(raw_values)[:6]]
        if not values:
            values = [""]
        if len(values) == 1:
            pending_pairs.append((label, values[0]))
            continue
        _flush_pairs()
        value_width = max(1100, int(6200 / len(values)))
        label_cell = (
            '<w:tc><w:tcPr><w:tcW w:w="2600" w:type="dxa"/>'
            '<w:shd w:fill="F3F4F6"/></w:tcPr>'
            f'{_word_paragraph(label, bold=True, size=16, spacing_after=0)}</w:tc>'
        )
        value_cells = "".join(
            '<w:tc><w:tcPr>'
            f'<w:tcW w:w="{value_width}" w:type="dxa"/></w:tcPr>'
            f'{_word_paragraph(value, size=16, spacing_after=0)}</w:tc>'
            for value in values
        )
        table_rows.append(f"<w:tr>{label_cell}{value_cells}</w:tr>")
    _flush_pairs()
    return (
        "<w:tbl>"
        '<w:tblPr><w:tblW w:w="0" w:type="auto"/>'
        f"<w:tblBorders>{border}</w:tblBorders>"
        '<w:tblCellMar><w:top w:w="35" w:type="dxa"/><w:left w:w="45" w:type="dxa"/>'
        '<w:bottom w:w="35" w:type="dxa"/><w:right w:w="45" w:type="dxa"/></w:tblCellMar>'
        "</w:tblPr>"
        f"{''.join(table_rows)}"
        "</w:tbl>"
        + _word_paragraph("", spacing_after=120)
    )


def _word_section(
    title: str,
    rows: Sequence[Tuple[str, str]],
    *,
    empty_text: str = "",
) -> str:
    parts = [_word_paragraph(title, bold=True, size=28, spacing_after=120, style="Heading1")]
    if rows:
        parts.append(_word_table(rows))
    elif empty_text:
        parts.append(_word_paragraph(empty_text))
    return "".join(parts)


def _word_microwire_data_section(row: pd.Series) -> str:
    rows: List[Tuple[str, List[str]]] = []
    for column in WORD_MICROWIRE_DATA_COLUMNS:
        column_text = str(column)
        lowered = column_text.casefold()
        if (
            column_text in _WORD_ALWAYS_OMIT_COLUMNS
            or column_text.startswith("_")
            or column_text.endswith("(Origin)")
        ):
            continue
        if "file" in lowered or "path" in lowered or "image" in lowered:
            continue
        rows.append((_word_label(column_text), _word_format_value_cells(row.get(column))))
    parts = [_word_paragraph("Microwire data", bold=True, size=24, spacing_after=70, style="Heading1")]
    if rows:
        parts.append(_word_microwire_data_table(rows))
    else:
        parts.append(_word_paragraph("No microwire values available."))
    return "".join(parts)


def _word_resolve_picture_path(
    value: Any,
    microscope_crops: Mapping[str, Path],
) -> Optional[Path]:
    for descriptor in _asset_references(value):
        crop = microscope_crops.get(descriptor)
        if crop is not None and Path(crop).exists():
            return Path(crop)
        candidate = Path(descriptor)
        if candidate.exists():
            return candidate
    return None


def _word_microscope_section(
    row: pd.Series,
    microscope_crops: Mapping[str, Path],
    bookmark_start: int,
) -> Tuple[str, List[WordPictureInsertion], int]:
    parts = [
        _word_page_break(),
        _word_paragraph(
            "Microscope and dimensions",
            bold=True,
            size=28,
            spacing_after=120,
            style="Heading1",
        )
    ]
    picture_insertions: List[WordPictureInsertion] = []
    rows = _word_key_values(row, _WORD_DIMENSION_COLUMNS)
    if rows:
        parts.append(_word_table(rows))
    bookmark_id = bookmark_start
    image_labels = {
        MICROSCOPE_IMAGE_COLUMNS[0]: "Core diameter image",
        MICROSCOPE_IMAGE_COLUMNS[1]: "Glass diameter image",
    }
    added_image = False
    for column, label in image_labels.items():
        if column not in row.index:
            continue
        image_path = _word_resolve_picture_path(row.get(column), microscope_crops)
        if image_path is None:
            continue
        added_image = True
        parts.append(_word_paragraph(label, bold=True, spacing_after=60))
        bookmark_name = f"MicroscopeImage{bookmark_id}"
        parts.append(
            _word_paragraph(
                f"[Microscope image placeholder: {label}]",
                spacing_after=220,
                bookmark_name=bookmark_name,
                bookmark_id=bookmark_id,
            )
        )
        picture_insertions.append(
            WordPictureInsertion(
                bookmark_name=bookmark_name,
                image_path=image_path,
                label=label,
            )
        )
        bookmark_id += 1
    if not rows and not added_image:
        parts.append(_word_paragraph("Not measured yet."))
    return "".join(parts), picture_insertions, bookmark_id


def _word_additional_values(row: pd.Series, used_columns: Sequence[str]) -> List[Tuple[str, str]]:
    used = set(used_columns) | set(_WORD_ALWAYS_OMIT_COLUMNS)
    values: List[Tuple[str, str]] = []
    for column in row.index:
        column_text = str(column)
        lowered = column_text.casefold()
        if column_text in used or column_text.startswith("_"):
            continue
        if "file" in lowered or "path" in lowered or "image" in lowered:
            continue
        value = row.get(column)
        formatted = _word_format_value(value) if _word_has_value(value) else ""
        values.append((_word_label(column_text), formatted))
    return values


def _word_assemble_values(row: pd.Series) -> List[Tuple[str, str]]:
    values: List[Tuple[str, str]] = []
    for column in WORD_MICROWIRE_DATA_COLUMNS:
        column_text = str(column)
        lowered = column_text.casefold()
        if (
            column_text in _WORD_ALWAYS_OMIT_COLUMNS
            or column_text.startswith("_")
            or column_text.endswith("(Origin)")
        ):
            continue
        if "file" in lowered or "path" in lowered or "image" in lowered:
            continue
        value = row.get(column)
        formatted = _word_format_value(value) if _word_has_value(value) else ""
        values.append((_word_label(column_text), formatted))
    return values


def _word_normalise_sample_text(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _word_row_composition_token(row: pd.Series) -> str:
    return _word_normalise_sample_text(row.get("Composition"))


def _word_row_microwire_pair(row: pd.Series) -> Tuple[str, str] | None:
    text = str(row.get("Microwire") or "").strip()
    match = re.search(r"(\d+)\s*[/_\-]\s*(\d+)", text)
    if not match:
        return None
    return match.group(1), match.group(2)


def _word_reference_conflicts_with_sample(row: pd.Series, *values: object) -> bool:
    text = " ".join(str(value or "") for value in values if value not in (None, ""))
    if not text:
        return False
    lowered = text.casefold()
    row_composition = _word_row_composition_token(row)
    compositions = [
        _word_normalise_sample_text(match.group(0))
        for match in re.finditer(r"Ni\d+(?:[A-Za-z]{1,2}\d+)+", text)
    ]
    if row_composition and compositions:
        if all(composition != row_composition for composition in compositions):
            return True
    pair = _word_row_microwire_pair(row)
    if pair is not None:
        microwire_matches = re.findall(r"(?<!\d)(\d+)\s*[/_\-]\s*(\d+)(?!\d)", lowered)
        if microwire_matches and all(match != pair for match in microwire_matches):
            return True
    return False


def _word_origin_artifact_is_usable(artifact: OriginArtifact | None) -> bool:
    if artifact is None:
        return False
    if bool(getattr(artifact, "clipboard_fallback", False)):
        return True
    object_path = getattr(artifact, "object_path", None)
    return object_path is not None


def _word_origin_artifact_conflicts_with_sample(
    row: pd.Series,
    descriptor: str,
    artifact: OriginArtifact | None,
) -> bool:
    values: List[object] = [descriptor]
    if artifact is not None:
        object_path = getattr(artifact, "object_path", None)
        object_name = Path(object_path).name if object_path not in (None, "") else None
        values.extend(
            [
                artifact.display_text,
                object_name,
                artifact.graph_name,
                artifact.workbook_name,
                artifact.worksheet_name,
            ]
        )
    return _word_reference_conflicts_with_sample(row, *values)


def _word_evaluate_graph_sections(
    row: pd.Series,
    origin_artifacts: Mapping[str, OriginArtifact],
) -> List[WordGraphSectionEvaluation]:
    evaluations: List[WordGraphSectionEvaluation] = []
    for title, origin_columns, reference_columns in _WORD_GRAPH_SECTIONS:
        accepted_origin: List[str] = []
        invalid_origin: List[str] = []
        missing_origin: List[str] = []
        accepted_references: List[str] = []
        invalid_references: List[str] = []

        for column in origin_columns:
            if column not in row.index:
                continue
            for descriptor in _asset_references(row.get(column)):
                artifact = origin_artifacts.get(descriptor)
                if _word_origin_artifact_conflicts_with_sample(row, descriptor, artifact):
                    invalid_origin.append(descriptor)
                    continue
                if _word_origin_artifact_is_usable(artifact):
                    if descriptor not in accepted_origin:
                        accepted_origin.append(descriptor)
                else:
                    missing_origin.append(descriptor)

        for column in reference_columns:
            if column not in row.index:
                continue
            for descriptor in _asset_references(row.get(column)):
                if _word_reference_conflicts_with_sample(row, descriptor):
                    invalid_references.append(descriptor)
                    continue
                if descriptor not in accepted_references:
                    accepted_references.append(descriptor)

        included = bool(accepted_origin or accepted_references)
        if accepted_origin:
            status = "included"
            reason = "accepted_origin_object"
        elif accepted_references:
            status = "included"
            reason = "reference_content"
        elif invalid_origin or invalid_references:
            status = "invalid"
            reason = "content_failed_sample_validation"
        elif missing_origin:
            status = "invalid"
            reason = "origin_descriptor_missing_artifact"
        else:
            status = "skipped"
            reason = "no_section_content"

        evaluations.append(
            WordGraphSectionEvaluation(
                title=title,
                included=included,
                status=status,
                reason=reason,
                accepted_origin_descriptors=tuple(accepted_origin),
                accepted_references=tuple(accepted_references),
                invalid_origin_descriptors=tuple(dict.fromkeys(invalid_origin)),
                missing_origin_descriptors=tuple(dict.fromkeys(missing_origin)),
                invalid_references=tuple(dict.fromkeys(invalid_references)),
            )
        )
    return evaluations


def _word_iter_ole_embedding_results(ole_embedding_results: Any) -> List[Any]:
    if ole_embedding_results is None:
        return []
    if isinstance(ole_embedding_results, Mapping):
        items: List[Any] = []
        for value in ole_embedding_results.values():
            if isinstance(value, (list, tuple, set)):
                items.extend(value)
            else:
                items.append(value)
        return items
    if isinstance(ole_embedding_results, (list, tuple, set)):
        return list(ole_embedding_results)
    return [ole_embedding_results]


def _word_result_value(result: Any, field_name: str, default: Any = None) -> Any:
    if isinstance(result, Mapping):
        return result.get(field_name, default)
    return getattr(result, field_name, default)


def _word_ole_result_manifest(result: Any) -> Dict[str, object]:
    attempted = bool(_word_result_value(result, "attempted", False))
    inserted = bool(_word_result_value(result, "inserted", False))
    status = str(_word_result_value(result, "status", "") or "").strip()
    if not status:
        status = "succeeded" if inserted else ("failed" if attempted else "skipped")
    return {
        "descriptor": str(_word_result_value(result, "descriptor", "") or ""),
        "bookmark": str(_word_result_value(result, "bookmark_name", "") or ""),
        "label": str(_word_result_value(result, "label", "") or ""),
        "object_path": str(_word_result_value(result, "object_path", "") or ""),
        "attempted": attempted,
        "inserted": inserted,
        "status": status,
        "reason": str(_word_result_value(result, "reason", "") or ""),
    }


def _word_ole_results_by_descriptor(ole_embedding_results: Any) -> Dict[str, Dict[str, object]]:
    results: Dict[str, Dict[str, object]] = {}
    for result in _word_iter_ole_embedding_results(ole_embedding_results):
        manifest = _word_ole_result_manifest(result)
        descriptor = str(manifest.get("descriptor") or "").strip()
        if descriptor:
            results[descriptor] = manifest
    return results


def word_report_section_manifest_for_row(
    row: pd.Series,
    origin_artifacts: Mapping[str, OriginArtifact] | None = None,
    ole_embedding_results: Sequence[Any] | Mapping[str, Any] | None = None,
) -> List[Dict[str, object]]:
    """Return data-driven Word measurement-section decisions for one report row."""

    summaries: List[Dict[str, object]] = []
    ole_results_provided = ole_embedding_results is not None
    ole_results_by_descriptor = _word_ole_results_by_descriptor(ole_embedding_results)
    for evaluation in _word_evaluate_graph_sections(row, origin_artifacts or {}):
        accepted_origin_descriptors = list(evaluation.accepted_origin_descriptors)
        missing_origin_descriptors = list(evaluation.missing_origin_descriptors)
        ole_insertions: List[Dict[str, object]] = []
        ole_insertions_attempted: List[str] = []
        ole_insertions_succeeded: List[str] = []
        ole_insertions_failed: List[str] = []
        ole_insertions_skipped: List[str] = []
        ole_insertions_missing_artifact: List[str] = list(missing_origin_descriptors)

        for descriptor in accepted_origin_descriptors:
            result = ole_results_by_descriptor.get(descriptor)
            if result is None:
                if ole_results_provided:
                    ole_insertions_skipped.append(descriptor)
                    ole_insertions.append(
                        {
                            "descriptor": descriptor,
                            "bookmark": "",
                            "label": descriptor,
                            "object_path": "",
                            "attempted": False,
                            "inserted": False,
                            "status": "skipped",
                            "reason": "ole_embedding_result_not_recorded",
                        }
                    )
                continue
            ole_insertions.append(result)
            status = str(result.get("status") or "").strip()
            attempted = bool(result.get("attempted"))
            inserted = bool(result.get("inserted"))
            if attempted and descriptor not in ole_insertions_attempted:
                ole_insertions_attempted.append(descriptor)
            if status == "succeeded" or inserted:
                ole_insertions_succeeded.append(descriptor)
            elif status == "missing_artifact":
                if descriptor not in ole_insertions_missing_artifact:
                    ole_insertions_missing_artifact.append(descriptor)
            elif status == "skipped":
                ole_insertions_skipped.append(descriptor)
            else:
                ole_insertions_failed.append(descriptor)

        summaries.append(
            {
                "title": evaluation.title,
                "included": bool(evaluation.included),
                "status": evaluation.status,
                "reason": evaluation.reason,
                "origin_descriptors": accepted_origin_descriptors,
                "origin_artifacts_accepted": accepted_origin_descriptors,
                "origin_artifacts_attempted": ole_insertions_attempted,
                "ole_insertions": ole_insertions,
                "ole_insertions_attempted": ole_insertions_attempted,
                "ole_insertions_succeeded": ole_insertions_succeeded,
                "ole_insertions_failed": ole_insertions_failed,
                "ole_insertions_skipped": ole_insertions_skipped,
                "ole_insertions_missing_artifact": ole_insertions_missing_artifact,
                "references": list(evaluation.accepted_references),
                "invalid_origin_descriptors": list(evaluation.invalid_origin_descriptors),
                "missing_origin_descriptors": missing_origin_descriptors,
                "invalid_references": list(evaluation.invalid_references),
            }
        )
    return summaries


def _word_graph_sections(
    row: pd.Series,
    origin_artifacts: Mapping[str, OriginArtifact],
    bookmark_start: int,
) -> Tuple[str, List[WordOleInsertion], int]:
    parts: List[str] = []
    insertions: List[WordOleInsertion] = []
    bookmark_id = bookmark_start
    for section in _word_evaluate_graph_sections(row, origin_artifacts):
        if not section.included:
            continue
        parts.append(_word_page_break())
        parts.append(_word_paragraph(section.title, bold=True, size=28, spacing_after=120, style="Heading1"))
        for descriptor in section.accepted_origin_descriptors:
            artifact = origin_artifacts.get(descriptor)
            if artifact is None:
                continue
            display = artifact.display_text if artifact.display_text else descriptor
            bookmark_name = f"OriginGraph{bookmark_id}"
            parts.append(
                _word_paragraph(
                    f"[Origin object placeholder: {descriptor}]",
                    spacing_after=220,
                    bookmark_name=bookmark_name,
                    bookmark_id=bookmark_id,
                )
            )
            insertions.append(
                WordOleInsertion(
                    bookmark_name=bookmark_name,
                    object_path=(
                        Path(artifact.object_path)
                        if artifact.object_path is not None
                        else Path(artifact.descriptor)
                    ),
                    label=str(display or descriptor),
                    clipboard_fallback=bool(getattr(artifact, "clipboard_fallback", False)),
                    graph_name=artifact.graph_name,
                    descriptor=descriptor,
                )
            )
            bookmark_id += 1
        if section.accepted_references and not section.accepted_origin_descriptors:
            parts.append(_word_table([("Graphs in Assemble", ", ".join(section.accepted_references))]))
            parts.append(
                _word_paragraph(
                    "Editable Origin object was not generated for this graph yet."
                )
            )
    return "".join(parts), insertions, bookmark_id


def _word_document_xml(
    row: pd.Series,
    fallback_index: int,
    origin_artifacts: Mapping[str, OriginArtifact],
    microscope_crops: Mapping[str, Path],
) -> Tuple[str, List[WordOleInsertion], List[WordPictureInsertion]]:
    title = _word_sample_title(row, fallback_index)
    body: List[str] = [
        _word_paragraph(title, bold=True, size=40, spacing_after=220, style="Title"),
        _word_microwire_data_section(row),
    ]
    microscope_xml, picture_insertions, bookmark_id = _word_microscope_section(
        row,
        microscope_crops,
        1,
    )
    body.append(microscope_xml)
    origin_xml, origin_insertions, _ = _word_graph_sections(row, origin_artifacts, bookmark_id)
    body.append(origin_xml)
    body.append(
        '<w:sectPr><w:headerReference w:type="default" r:id="rId3"/>'
        '<w:footerReference w:type="default" r:id="rId4"/>'
        '<w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="900" w:right="900" w:bottom="900" w:left="900" '
        'w:header="720" w:footer="720" w:gutter="0"/></w:sectPr>'
    )
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<w:body>{''.join(body)}</w:body></w:document>"
    )
    return xml, origin_insertions, picture_insertions


def _write_word_docx(path: Path, document_xml: str, *, sample_title: str = "") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '<Override PartName="/word/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
        '<Override PartName="/word/settings.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.settings+xml"/>'
        '<Override PartName="/word/header1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.header+xml"/>'
        '<Override PartName="/word/footer1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.footer+xml"/>'
        '<Override PartName="/docProps/core.xml" '
        'ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        "</Types>"
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" '
        'Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" '
        'Target="docProps/app.xml"/>'
        "</Relationships>"
    )
    document_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/settings" '
        'Target="settings.xml"/>'
        '<Relationship Id="rId3" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/header" '
        'Target="header1.xml"/>'
        '<Relationship Id="rId4" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer" '
        'Target="footer1.xml"/>'
        "</Relationships>"
    )
    styles = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:style w:type="paragraph" w:default="1" w:styleId="Normal">'
        '<w:name w:val="Normal"/><w:qFormat/></w:style>'
        '<w:style w:type="paragraph" w:styleId="Header">'
        '<w:name w:val="header"/><w:basedOn w:val="Normal"/>'
        '<w:pPr><w:tabs><w:tab w:val="right" w:pos="9360"/></w:tabs>'
        '<w:spacing w:after="0"/></w:pPr><w:rPr><w:sz w:val="18"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Footer">'
        '<w:name w:val="footer"/><w:basedOn w:val="Normal"/>'
        '<w:pPr><w:tabs><w:tab w:val="center" w:pos="4680"/></w:tabs>'
        '<w:spacing w:after="0"/></w:pPr><w:rPr><w:sz w:val="18"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Title">'
        '<w:name w:val="Title"/><w:basedOn w:val="Normal"/><w:next w:val="Normal"/>'
        '<w:qFormat/><w:pPr><w:spacing w:after="220"/></w:pPr>'
        '<w:rPr><w:b/><w:sz w:val="40"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Heading1">'
        '<w:name w:val="heading 1"/><w:basedOn w:val="Normal"/><w:next w:val="Normal"/>'
        '<w:qFormat/><w:pPr><w:keepNext/><w:outlineLvl w:val="0"/>'
        '<w:spacing w:before="80" w:after="90"/></w:pPr>'
        '<w:rPr><w:b/><w:sz w:val="28"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Heading2">'
        '<w:name w:val="heading 2"/><w:basedOn w:val="Normal"/><w:next w:val="Normal"/>'
        '<w:qFormat/><w:pPr><w:keepNext/><w:outlineLvl w:val="1"/>'
        '<w:spacing w:before="120" w:after="80"/></w:pPr>'
        '<w:rPr><w:b/><w:sz w:val="24"/></w:rPr></w:style>'
        "</w:styles>"
    )
    settings = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:settings xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:defaultTabStop w:val="720"/></w:settings>'
    )
    created = datetime.now(timezone.utc).isoformat(timespec="seconds")
    core_props = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<dc:title>Microwire sample report</dc:title>"
        "<dc:creator>PyPlot Microwire Data Builder</dc:creator>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{created}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created}</dcterms:modified>'
        "</cp:coreProperties>"
    )
    app_props = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>PyPlot</Application></Properties>"
    )
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/_rels/document.xml.rels", document_rels)
        archive.writestr("word/styles.xml", styles)
        archive.writestr("word/settings.xml", settings)
        archive.writestr("word/header1.xml", _word_header_xml(sample_title))
        archive.writestr("word/footer1.xml", _word_footer_xml())
        archive.writestr("docProps/core.xml", core_props)
        archive.writestr("docProps/app.xml", app_props)


def _copy_origin_graph_to_clipboard_by_name(
    graph_name: Optional[str],
    log: logging.Logger,
) -> bool:
    name = str(graph_name or "").strip()
    if not name:
        return False
    try:
        try:
            from plotting.shared.origin import _ensure_origin_sdk_on_path

            _ensure_origin_sdk_on_path()
        except Exception:
            pass
        import originpro as op  # type: ignore
    except Exception as exc:
        log.debug("Unable to import originpro for Word OLE copy: %s", exc)
        return False
    lt_int = getattr(op, "lt_int", None)
    probe_ok = False
    if callable(lt_int):
        try:
            lt_int("@V")
            probe_ok = True
        except Exception:
            probe_ok = False
    if not probe_ok:
        log.debug("Origin automation is not active for Word OLE copy of %s", name)
        return False
    try:
        op.set_show()
    except Exception:
        pass
    try:
        op.lt_exec("@GMC=1;")
    except Exception:
        pass
    graph = None
    finder = getattr(op, "find_graph", None)
    if callable(finder):
        try:
            graph = finder(name)
        except Exception:
            graph = None
    if graph is None:
        graph_list = getattr(op, "graph_list", None)
        if callable(graph_list):
            try:
                candidates = graph_list()
            except Exception:
                candidates = []
            for candidate in candidates:
                candidate_names = {
                    text
                    for text in (
                        _origin_object_name(candidate),
                        getattr(candidate, "name", None),
                        getattr(candidate, "lname", None),
                        getattr(candidate, "long_name", None),
                        str(candidate),
                    )
                    if isinstance(text, str) and text.strip()
                }
                if name in candidate_names:
                    graph = candidate
                    break
    if graph is None:
        log.debug("Origin graph %s was not found for Word OLE copy", name)
        return False
    return _copy_origin_graph_page(graph, log)


def _copy_origin_graph_file_to_clipboard(
    object_path: Path,
    log: logging.Logger,
) -> bool:
    try:
        try:
            from plotting.shared.origin import _ensure_origin_sdk_on_path

            _ensure_origin_sdk_on_path()
        except Exception:
            pass
        import originpro as op  # type: ignore
    except Exception as exc:
        log.debug("Unable to import originpro for Word OLE file copy: %s", exc)
        return False
    try:
        opened = op.open(str(object_path), asksave=False)
    except Exception as exc:
        log.debug("Unable to open Origin graph %s for Word OLE copy: %s", object_path, exc)
        return False
    if opened is False:
        log.debug("Origin refused to open graph %s for Word OLE copy", object_path)
        return False
    try:
        op.set_show()
    except Exception:
        pass
    try:
        graphs = op.graph_list("p", True)
    except Exception:
        graphs = []
    graph = graphs[-1] if graphs else None
    if graph is None:
        log.debug("No graph page was available after opening %s", object_path)
        return False
    return _copy_origin_graph_page(graph, log)


def _embed_origin_objects_with_word(
    docx_path: Path,
    insertions: Sequence[WordOleInsertion],
    log: logging.Logger,
) -> List[WordOleEmbeddingResult]:
    if not insertions:
        return []
    docx_path = docx_path.resolve()
    results: List[WordOleEmbeddingResult] = []
    if os.name != "nt":
        log.warning("Word OLE embedding is only available on Windows; left placeholders in %s", docx_path)
        return [
            WordOleEmbeddingResult(
                bookmark_name=insertion.bookmark_name,
                descriptor=str(insertion.descriptor or insertion.object_path.name),
                label=insertion.label,
                object_path=str(insertion.object_path),
                attempted=False,
                inserted=False,
                status="skipped",
                reason="word_ole_windows_only",
            )
            for insertion in insertions
        ]
    powershell = shutil.which("powershell.exe") or shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:
        log.warning("PowerShell is unavailable; left Origin placeholders in %s", docx_path)
        return [
            WordOleEmbeddingResult(
                bookmark_name=insertion.bookmark_name,
                descriptor=str(insertion.descriptor or insertion.object_path.name),
                label=insertion.label,
                object_path=str(insertion.object_path),
                attempted=False,
                inserted=False,
                status="skipped",
                reason="powershell_unavailable",
            )
            for insertion in insertions
        ]
    script = r"""
param(
    [Parameter(Mandatory=$true)][string]$DocxPath,
    [Parameter(Mandatory=$true)][string]$PayloadPath
)
$ErrorActionPreference = 'Stop'
$wdCollapseStart = 1
$wdInLine = 0
$wdPasteOLEObject = 0
$missing = [Type]::Missing
$result = [ordered]@{
    inserted = 0
    attempted = 0
    warnings = @()
}
$word = $null
$doc = $null
try {
    $items = Get-Content -Raw -LiteralPath $PayloadPath | ConvertFrom-Json
    $word = New-Object -ComObject Word.Application
    $word.Visible = $false
    $doc = $word.Documents.Open($DocxPath)
    foreach ($item in @($items)) {
        $bookmarkName = [string]$item.bookmark
        $objectPath = [string]$item.path
        $preferClipboardValue = [bool]$item.preferClipboard
        $allowFileValue = [bool]$item.allowFile
        if ((-not $preferClipboardValue) -and (-not (Test-Path -LiteralPath $objectPath))) {
            $result.warnings += "Object not found: " + $objectPath
            continue
        }
        if (-not $doc.Bookmarks.Exists($bookmarkName)) {
            $result.warnings += "Bookmark not found: " + $bookmarkName
            continue
        }
        $result.attempted += 1
        $bookmarkObj = $null
        for ($bookmarkIndex = 1; $bookmarkIndex -le $doc.Bookmarks.Count; $bookmarkIndex++) {
            $candidate = $doc.Bookmarks.Item($bookmarkIndex)
            if ([string]$candidate.Name -eq $bookmarkName) {
                $bookmarkObj = $candidate
                break
            }
        }
        if ($bookmarkObj -eq $null) {
            $result.warnings += "Bookmark not found after lookup: " + $bookmarkName
            continue
        }
        $insertRange = $bookmarkObj.Range
        $insertRange.Text = ""
        $insertRange.Collapse($wdCollapseStart)
        $insertRange.Select() | Out-Null
        $selection = $word.Selection
        $shape = $null
        if ($preferClipboardValue) {
            try {
                $before = $doc.InlineShapes.Count
                try {
                    $selection.PasteSpecial($false, $false, $wdInLine, $false, $wdPasteOLEObject, "", "") | Out-Null
                } catch {
                    $selection.Paste() | Out-Null
                }
                $after = $doc.InlineShapes.Count
                if ($after -gt $before) {
                    $shape = $doc.InlineShapes.Item($after)
                }
            } catch {
                $result.warnings += "Clipboard paste failed for " + $bookmarkName + ": " + $_.Exception.Message
            }
        }
        if ($shape -eq $null -and $allowFileValue -and (Test-Path -LiteralPath $objectPath)) {
            $lastError = $null
            foreach ($classType in @("Origin95.Graph", "Origin.Graph", "", $null)) {
                try {
                    if ($classType -eq $null) {
                        $shape = $doc.InlineShapes.AddOLEObject($missing, $objectPath, $false, $false, $missing, $missing, $missing, $insertRange)
                    } else {
                        $shape = $doc.InlineShapes.AddOLEObject([string]$classType, $objectPath, $false, $false, $missing, $missing, $missing, $insertRange)
                    }
                    break
                } catch {
                    $lastError = $_
                }
            }
            if ($shape -eq $null) {
                $message = "unknown error"
                if ($lastError -ne $null -and $lastError.Exception -ne $null) {
                    $message = $lastError.Exception.Message
                }
                $result.warnings += "File insert failed for " + $bookmarkName + ": " + $message
            }
        }
        if ($shape -ne $null) {
            try {
                $shape.LockAspectRatio = $true
                if ($shape.Width -gt 430) {
                    $shape.Width = 430
                }
                if ($shape.Height -gt 330) {
                    $shape.Height = 330
                }
            } catch {
            }
            $result.inserted += 1
        }
    }
    $doc.Save()
} catch {
    $line = ""
    try {
        $line = [string]$_.InvocationInfo.ScriptLineNumber
    } catch {
    }
    if ($line) {
        $result.warnings += "PowerShell line " + $line + ": " + $_.Exception.Message
    } else {
        $result.warnings += $_.Exception.Message
    }
} finally {
    if ($doc -ne $null) {
        $doc.Close($true) | Out-Null
    }
    if ($word -ne $null) {
        $word.Quit() | Out-Null
    }
}
$result | ConvertTo-Json -Compress
"""
    inserted_count = 0
    attempted_count = 0
    with tempfile.TemporaryDirectory(prefix="pyplot-word-ole-") as tmp:
        tmp_path = Path(tmp)
        payload_path = tmp_path / "origin_objects.json"
        script_path = tmp_path / "embed_origin_objects.ps1"
        script_path.write_text(script, encoding="utf-8")
        for insertion in insertions:
            object_path = insertion.object_path
            try:
                if object_path.exists():
                    object_path = object_path.resolve()
            except OSError:
                pass
            object_exists = object_path.exists()
            copied_to_clipboard = False
            if object_exists:
                copied_to_clipboard = _copy_origin_graph_file_to_clipboard(
                    object_path,
                    log,
                )
            if not copied_to_clipboard and (insertion.clipboard_fallback or insertion.graph_name):
                copied_to_clipboard = _copy_origin_graph_to_clipboard_by_name(
                    insertion.graph_name,
                    log,
                )
            if not object_exists and not copied_to_clipboard:
                log.warning(
                    "Origin object for %s is unavailable; left placeholder in %s",
                    insertion.label,
                    docx_path,
                )
                results.append(
                    WordOleEmbeddingResult(
                        bookmark_name=insertion.bookmark_name,
                        descriptor=str(insertion.descriptor or insertion.object_path.name),
                        label=insertion.label,
                        object_path=str(object_path),
                        attempted=False,
                        inserted=False,
                        status="missing_artifact",
                        reason="origin_object_unavailable",
                    )
                )
                continue
            attempted_count += 1
            payload = [
                {
                    "bookmark": insertion.bookmark_name,
                    "path": str(object_path),
                    "label": insertion.label,
                    "preferClipboard": bool(copied_to_clipboard),
                    "allowFile": bool(object_exists),
                }
            ]
            payload_path.write_text(json.dumps(payload), encoding="utf-8")
            completed = subprocess.run(
                [
                    powershell,
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(script_path),
                    "-DocxPath",
                    str(docx_path),
                    "-PayloadPath",
                    str(payload_path),
                ],
                capture_output=True,
                text=True,
                timeout=180,
                check=False,
            )
            detail = (completed.stderr or completed.stdout or "").strip()
            if completed.returncode != 0:
                log.warning(
                    "Failed to embed Origin object %s into %s: %s",
                    insertion.label,
                    docx_path,
                    detail or completed.returncode,
                )
                results.append(
                    WordOleEmbeddingResult(
                        bookmark_name=insertion.bookmark_name,
                        descriptor=str(insertion.descriptor or insertion.object_path.name),
                        label=insertion.label,
                        object_path=str(object_path),
                        attempted=True,
                        inserted=False,
                        status="failed",
                        reason=detail or str(completed.returncode),
                    )
                )
                continue
            result: Dict[str, Any] = {}
            try:
                result = json.loads(completed.stdout or "{}")
            except Exception:
                result = {}
            item_inserted = int(result.get("inserted") or 0) > 0
            inserted_count += 1 if item_inserted else 0
            warnings = [str(warning) for warning in (result.get("warnings") or [])]
            for warning in warnings:
                log.warning("Origin object embedding warning for %s: %s", docx_path, warning)
            results.append(
                WordOleEmbeddingResult(
                    bookmark_name=insertion.bookmark_name,
                    descriptor=str(insertion.descriptor or insertion.object_path.name),
                    label=insertion.label,
                    object_path=str(object_path),
                    attempted=True,
                    inserted=item_inserted,
                    status="succeeded" if item_inserted else "failed",
                    reason="; ".join(warnings),
                )
            )
    if attempted_count and inserted_count != attempted_count:
        log.warning(
            "Embedded %s of %s Origin object(s) into %s",
            inserted_count,
            attempted_count,
            docx_path,
        )
    return results


def _synthetic_word_ole_results(
    insertions: Sequence[WordOleInsertion],
    *,
    status: str,
    attempted: bool,
    inserted: bool,
    reason: str = "",
) -> List[WordOleEmbeddingResult]:
    return [
        WordOleEmbeddingResult(
            bookmark_name=insertion.bookmark_name,
            descriptor=str(insertion.descriptor or insertion.object_path.name),
            label=insertion.label,
            object_path=str(insertion.object_path),
            attempted=attempted,
            inserted=inserted,
            status=status,
            reason=reason,
        )
        for insertion in insertions
    ]


def _embed_pictures_with_word(
    docx_path: Path,
    insertions: Sequence[WordPictureInsertion],
    log: logging.Logger,
) -> None:
    if not insertions:
        return
    docx_path = docx_path.resolve()
    if os.name != "nt":
        log.warning("Word image embedding is only available on Windows; left placeholders in %s", docx_path)
        return
    powershell = shutil.which("powershell.exe") or shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:
        log.warning("PowerShell is unavailable; left microscope image placeholders in %s", docx_path)
        return
    payload = [
        {
            "bookmark": insertion.bookmark_name,
            "path": str(insertion.image_path),
            "label": insertion.label,
        }
        for insertion in insertions
        if insertion.image_path.exists()
    ]
    if not payload:
        return
    script = r"""
param(
    [Parameter(Mandatory=$true)][string]$DocxPath,
    [Parameter(Mandatory=$true)][string]$PayloadPath
)
$ErrorActionPreference = 'Stop'
$items = Get-Content -Raw -LiteralPath $PayloadPath | ConvertFrom-Json
$word = $null
$doc = $null
try {
    $word = New-Object -ComObject Word.Application
    $word.Visible = $false
    $doc = $word.Documents.Open($DocxPath)
    foreach ($item in @($items)) {
        if (-not (Test-Path -LiteralPath $item.path)) {
            continue
        }
        if (-not $doc.Bookmarks.Exists($item.bookmark)) {
            continue
        }
        $bookmark = $doc.Bookmarks.Item($item.bookmark)
        $range = $bookmark.Range
        $range.Text = ""
        $shape = $doc.InlineShapes.AddPicture([string]$item.path, $false, $true, $range)
        try {
            $shape.LockAspectRatio = $true
            if ($shape.Width -gt 260) {
                $shape.Width = 260
            }
        } catch {
        }
    }
    $doc.Save()
} finally {
    if ($doc -ne $null) {
        $doc.Close($true) | Out-Null
    }
    if ($word -ne $null) {
        $word.Quit() | Out-Null
    }
}
"""
    with tempfile.TemporaryDirectory(prefix="pyplot-word-images-") as tmp:
        tmp_path = Path(tmp)
        payload_path = tmp_path / "pictures.json"
        script_path = tmp_path / "embed_pictures.ps1"
        payload_path.write_text(json.dumps(payload), encoding="utf-8")
        script_path.write_text(script, encoding="utf-8")
        completed = subprocess.run(
            [
                powershell,
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-DocxPath",
                str(docx_path),
                "-PayloadPath",
                str(payload_path),
            ],
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
        )
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "").strip()
        log.warning("Failed to embed microscope images into %s: %s", docx_path, detail or completed.returncode)


def _export_word_reports(
    dataframe: pd.DataFrame,
    output_dir: Path,
    origin_artifacts: Mapping[str, OriginArtifact],
    log: logging.Logger,
    microscope_crops: Mapping[str, Path] | None = None,
    ole_embedding_results: Dict[Path, List[WordOleEmbeddingResult]] | None = None,
) -> List[Path]:
    if dataframe.empty:
        return []
    output_dir.mkdir(parents=True, exist_ok=True)
    reports: List[Path] = []
    used_names: Set[str] = set()
    used_live_origin_clipboard = False
    try:
        for index, (_, row) in enumerate(dataframe.reset_index(drop=True).iterrows()):
            filename = _word_report_filename(row, index)
            stem = Path(filename).stem
            suffix = Path(filename).suffix or ".docx"
            candidate = filename
            duplicate_index = 2
            while candidate.lower() in used_names:
                candidate = f"{stem}_{duplicate_index}{suffix}"
                duplicate_index += 1
            used_names.add(candidate.lower())
            report_path = output_dir / candidate
            document_xml, origin_insertions, picture_insertions = _word_document_xml(
                row,
                index,
                origin_artifacts,
                microscope_crops or {},
            )
            used_live_origin_clipboard = used_live_origin_clipboard or any(
                insertion.clipboard_fallback for insertion in origin_insertions
            )
            _write_word_docx(
                report_path,
                document_xml,
                sample_title=_word_sample_title(row, index),
            )
            try:
                _embed_pictures_with_word(report_path, picture_insertions, log)
            except Exception:
                log.exception("Failed to run Word image embedding for %s", report_path)
            try:
                result_payload = _embed_origin_objects_with_word(report_path, origin_insertions, log)
            except Exception as exc:
                log.exception("Failed to run Word OLE embedding for %s", report_path)
                result_payload = _synthetic_word_ole_results(
                    origin_insertions,
                    status="failed",
                    attempted=True,
                    inserted=False,
                    reason=str(exc) or exc.__class__.__name__,
                )
            if result_payload is None:
                # Older tests and downstream callers monkeypatch this helper as
                # side-effect-only. Preserve that behavior while still letting
                # manifests record a successful attempted insertion.
                result_payload = _synthetic_word_ole_results(
                    origin_insertions,
                    status="succeeded",
                    attempted=True,
                    inserted=True,
                )
            if ole_embedding_results is not None:
                ole_embedding_results[report_path] = list(result_payload)
            reports.append(report_path)
    finally:
        if used_live_origin_clipboard:
            try:
                from plotting.shared.origin import release_origin

                release_origin()
            except Exception:
                log.debug("Failed to release live Origin session after Word export", exc_info=True)
    return reports


def export_word_reports(
    dataframe: pd.DataFrame,
    output_dir: Path | str,
    *,
    origin_artifacts: Mapping[str, OriginArtifact] | None = None,
    microscope_crops: Mapping[str, Path] | None = None,
    ole_embedding_results: Dict[Path, List[WordOleEmbeddingResult]] | None = None,
    logger: logging.Logger | None = None,
) -> List[Path]:
    """Write one Word sample report per row without requiring the Builder UI."""

    log = logger if logger is not None else logging.getLogger(LOGGER_NAME)
    return _export_word_reports(
        dataframe,
        Path(output_dir),
        origin_artifacts or {},
        log,
        microscope_crops=microscope_crops,
        ole_embedding_results=ole_embedding_results,
    )


def _embed_plots_in_excel_openpyxl(
    excel_path: Path,
    dataframe: pd.DataFrame,
    plot_name_to_path: Dict[str, Path],
    plot_dir: Path,
    log: logging.Logger,
    figure_size: Tuple[float, float],
    *,
    microscope_columns: Sequence[str] = (),
    microscope_crops: Optional[Dict[str, Path]] = None,
    highlight_map: Optional[Dict[str, Set[int]]] = None,
) -> None:
    """Insert Matplotlib plot images directly into the Excel export."""

    if not excel_path.exists():
        return
    if dataframe.empty:
        return
    figure_columns = [column for column in FIGURE_COLUMNS if column in dataframe.columns]
    microscope_columns = [
        column for column in microscope_columns if column in dataframe.columns
    ]
    if not figure_columns and not microscope_columns:
        return

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
    except ImportError:  # pragma: no cover - optional dependency guard
        log.warning("openpyxl is required to embed plots into Excel workbooks")
        return

    available: Dict[str, Path] = {}
    for name, path in plot_name_to_path.items():
        if path.exists():
            available[name] = path
    if plot_dir and plot_dir.exists():
        for candidate in plot_dir.glob("*.png"):
            available.setdefault(candidate.name, candidate)
    if not available:
        return

    workbook = load_workbook(excel_path)
    try:
        worksheet = workbook.active

        inserted = False
        reset_df = dataframe.reset_index(drop=True)
        target_width_in, target_height_in = figure_size
        target_width_px, target_height_px = _excel_pixel_limits(figure_size)
        row_height_pts = _excel_row_height(target_height_in)
        column_width_chars = _excel_column_width(target_width_in)
        for row_idx, row in reset_df.iterrows():
            for column in figure_columns:
                name = _single_asset_reference(row.get(column))
                if name is None:
                    continue
                image_path = available.get(name)
                if image_path is None or not image_path.exists():
                    candidate = plot_dir / name if plot_dir else None
                    if candidate is None or not candidate.exists():
                        continue
                    available[name] = candidate
                    image_path = candidate
                width_px, height_px, dpi_x, dpi_y = _image_metrics(image_path)

                try:
                    image = XLImage(str(image_path))
                except Exception:
                    log.exception("Failed to load plot image %s for Excel export", image_path)
                    continue

                if width_px and target_width_px:
                    desired_width_px = int(round(target_width_px))
                    if not math.isclose(width_px, desired_width_px, rel_tol=1e-3):
                        image.width = desired_width_px
                if height_px and target_height_px:
                    desired_height_px = int(round(target_height_px))
                    if not math.isclose(height_px, desired_height_px, rel_tol=1e-3):
                        image.height = desired_height_px

                location = dataframe.columns.get_loc(column)
                if isinstance(location, slice):
                    continue
                if isinstance(location, np.ndarray):
                    continue
                try:
                    column_index = int(location) + 1
                except Exception:
                    continue
                row_number = row_idx + 2  # account for the header row
                column_letter = get_column_letter(column_index)
                cell_reference = f"{column_letter}{row_number}"

                # Clear the textual filename and embed the image anchored at the cell.
                worksheet[cell_reference].value = None
                worksheet.add_image(image, cell_reference)

                if row_height_pts > 0:
                    row_dim = worksheet.row_dimensions[row_number]
                    row_dim.height = row_height_pts
                    row_dim.customHeight = True

                if column_width_chars > 0:
                    col_dim = worksheet.column_dimensions[column_letter]
                    col_dim.width = column_width_chars
                    col_dim.customWidth = True

                inserted = True

            for column in microscope_columns:
                value = row.get(column)
                if not isinstance(value, str):
                    continue
                key = value.strip()
                if not key:
                    continue
                image_path = None
                if microscope_crops:
                    image_path = microscope_crops.get(key)
                if image_path is None:
                    candidate = Path(key)
                    if candidate.exists():
                        image_path = candidate
                if image_path is None or not image_path.exists():
                    continue
                try:
                    image = XLImage(str(image_path))
                except Exception:
                    log.exception(
                        "Failed to load microscope image %s for Excel export", image_path
                    )
                    continue
                location = dataframe.columns.get_loc(column)
                if isinstance(location, slice):
                    continue
                if isinstance(location, np.ndarray):
                    continue
                try:
                    column_index = int(location) + 1
                except Exception:
                    continue
                row_number = row_idx + 2
                column_letter = get_column_letter(column_index)
                cell_reference = f"{column_letter}{row_number}"
                worksheet[cell_reference].value = None
                worksheet.add_image(image, cell_reference)
                width_px, height_px, dpi_x, dpi_y = _image_metrics(image_path)
                if height_px:
                    height_in = height_px / (dpi_y or EXCEL_EMBED_DPI)
                    required_height = _excel_row_height(height_in)
                    row_dim = worksheet.row_dimensions[row_number]
                    existing = row_dim.height or 0.0
                    row_dim.height = max(existing, required_height)
                    row_dim.customHeight = True
                if width_px:
                    width_in = width_px / (dpi_x or EXCEL_EMBED_DPI)
                    required_width = _excel_column_width(width_in)
                    col_dim = worksheet.column_dimensions[column_letter]
                    existing_width = col_dim.width or 0.0
                    col_dim.width = max(existing_width, required_width)
                    col_dim.customWidth = True
                inserted = True

        if highlight_map:
            fill = PatternFill(fill_type="solid", start_color="FFF4B5", end_color="FFF4B5")
            for column, rows in highlight_map.items():
                if column not in dataframe.columns:
                    continue
                location = dataframe.columns.get_loc(column)
                if isinstance(location, slice):
                    continue
                if isinstance(location, np.ndarray):
                    continue
                try:
                    column_index = int(location) + 1
                except Exception:
                    continue
                column_letter = get_column_letter(column_index)
                for row_idx in rows:
                    if row_idx < 0 or row_idx >= len(reset_df):
                        continue
                    row_number = row_idx + 2
                    worksheet[f"{column_letter}{row_number}"].fill = fill

        if inserted:
            workbook.save(excel_path)
            _adjust_drawing_ext_dimensions(excel_path, figure_size, log)
    finally:
        workbook.close()


def _embed_assets_with_xlsxwriter(
    writer: "pd.ExcelWriter",
    dataframe: pd.DataFrame,
    plot_name_to_path: Dict[str, Path],
    plot_dir: Path,
    origin_artifacts: Dict[str, OriginArtifact],
    figure_size: Tuple[float, float],
    log: logging.Logger,
    *,
    microscope_crops: Optional[Dict[str, Path]] = None,
    microscope_columns: Sequence[str] = (),
    highlight_map: Optional[Dict[str, Set[int]]] = None,
) -> None:
    try:
        workbook_obj = getattr(writer, "book")
    except Exception:
        return
    if workbook_obj is None:
        return
    workbook = cast(Any, workbook_obj)
    try:
        sheets_mapping = getattr(writer, "sheets")
    except Exception:
        return
    if not sheets_mapping:
        return
    worksheets = [cast(Any, sheet) for sheet in sheets_mapping.values()]
    if not worksheets:
        return
    worksheet = cast(Any, worksheets[0])

    figure_columns = [column for column in FIGURE_COLUMNS if column in dataframe.columns]
    origin_columns = [
        column for column in ORIGIN_FIGURE_COLUMNS if column in dataframe.columns
    ]
    microscope_columns = [
        column for column in microscope_columns if column in dataframe.columns
    ]
    if not figure_columns and not origin_columns and not microscope_columns:
        return

    reset_df = dataframe.reset_index(drop=True)
    available: Dict[str, Path] = {}
    for name, path in plot_name_to_path.items():
        if path.exists():
            available[name] = path
    if plot_dir and plot_dir.exists():
        for candidate in plot_dir.glob("*.png"):
            available.setdefault(candidate.name, candidate)

    from PIL import Image as PILImage  # local import to avoid hard dependency elsewhere

    target_width_in, target_height_in = figure_size
    target_width_px, target_height_px = _excel_pixel_limits(figure_size)
    row_heights_pts: Dict[int, float] = {}
    column_widths_chars: Dict[int, float] = {}
    row_heights_px: Dict[int, int] = {}
    column_widths_px: Dict[int, int] = {}
    set_row_pixels = getattr(worksheet, "set_row_pixels", None)
    set_column_pixels = getattr(worksheet, "set_column_pixels", None)

    def ensure_row_height(row_idx: int, minimum_px: int, minimum_pts: float) -> None:
        if minimum_px < 0:
            minimum_px = 0
        if minimum_pts < 0:
            minimum_pts = 0.0

        if callable(set_row_pixels):
            try:
                set_row_pixels(row_idx, minimum_px)
            except Exception:
                pass
            else:
                row_heights_px[row_idx] = minimum_px
                row_heights_pts[row_idx] = minimum_pts
        try:
            worksheet.set_row(row_idx, minimum_pts)
        except Exception:
            return
        row_heights_pts[row_idx] = minimum_pts

    def ensure_column_width(col_idx: int, minimum_px: int, minimum_chars: float) -> None:
        if minimum_px < 0:
            minimum_px = 0
        if minimum_chars < 0:
            minimum_chars = 0.0

        if callable(set_column_pixels):
            try:
                set_column_pixels(col_idx, col_idx, minimum_px)
            except Exception:
                pass
            else:
                column_widths_px[col_idx] = minimum_px
                column_widths_chars[col_idx] = minimum_chars
                return

        try:
            worksheet.set_column(col_idx, col_idx, minimum_chars)
        except Exception:
            return
        column_widths_chars[col_idx] = minimum_chars

    def _column_index(column: str) -> Optional[int]:
        location = dataframe.columns.get_loc(column)
        if isinstance(location, slice):
            return None
        if isinstance(location, np.ndarray):
            return None
        if isinstance(location, tuple):
            return None
        if isinstance(location, list):
            return None
        try:
            return int(location)
        except Exception:
            return None

    for row_idx, row in reset_df.iterrows():
        row_number = row_idx + 1
        for column in figure_columns:
            name = _single_asset_reference(row.get(column))
            if name is None:
                continue
            image_path = available.get(name)
            if image_path is None and plot_dir:
                candidate = plot_dir / name
                if candidate.exists():
                    available[name] = candidate
                    image_path = candidate
            if image_path is None or not image_path.exists():
                continue
            width_px, height_px, dpi_x, dpi_y = _image_metrics(image_path)
            x_scale = (
                (target_width_px / float(width_px))
                if width_px
                else 1.0
            )
            y_scale = (
                (target_height_px / float(height_px))
                if height_px
                else 1.0
            )
            options: Dict[str, float] = {}
            if width_px and not math.isclose(x_scale, 1.0, rel_tol=1e-3):
                options["x_scale"] = x_scale
            if height_px and not math.isclose(y_scale, 1.0, rel_tol=1e-3):
                options["y_scale"] = y_scale
            column_index = _column_index(column)
            if column_index is None:
                continue
            worksheet.write_blank(row_number, column_index, None)
            try:
                worksheet.insert_image(row_number, column_index, str(image_path), options)
            except Exception:
                log.exception("Failed to insert plot image %s", image_path)
                continue
            ensure_row_height(
                row_number,
                int(round(target_height_in * EXCEL_EMBED_DPI)),
                _excel_row_height(target_height_in),
            )

            ensure_column_width(
                column_index,
                int(round(target_width_in * EXCEL_EMBED_DPI)),
                _excel_column_width(target_width_in),
            )

        for column in microscope_columns:
            value = row.get(column)
            if not isinstance(value, str):
                continue
            key = value.strip()
            if not key:
                continue
            image_path = None
            if microscope_crops:
                image_path = microscope_crops.get(key)
            if image_path is None:
                candidate = Path(key)
                if candidate.exists():
                    image_path = candidate
            if image_path is None or not image_path.exists():
                continue
            width_px, height_px, dpi_x, dpi_y = _image_metrics(image_path)
            if not width_px:
                width_px = 180
            if not height_px:
                height_px = 140
            width_in = width_px / (dpi_x or EXCEL_EMBED_DPI)
            height_in = height_px / (dpi_y or EXCEL_EMBED_DPI)
            column_index = _column_index(column)
            if column_index is None:
                continue
            worksheet.write_blank(row_number, column_index, None)
            try:
                worksheet.insert_image(row_number, column_index, str(image_path))
            except Exception:
                log.exception("Failed to insert microscope image %s", image_path)
                continue
            ensure_row_height(
                row_number,
                int(round(height_in * EXCEL_EMBED_DPI)),
                _excel_row_height(height_in),
            )
            ensure_column_width(
                column_index,
                int(round(width_in * EXCEL_EMBED_DPI)),
                _excel_column_width(width_in),
            )

        for column in origin_columns:
            descriptor = _single_asset_reference(row.get(column))
            if descriptor is None:
                continue
            artifact = origin_artifacts.get(descriptor)
            if artifact is None or artifact.object_path is None:
                continue
            try:
                if not artifact.object_path.exists():
                    continue
            except OSError:
                continue
            column_index = _column_index(column)
            if column_index is None:
                continue
            worksheet.write_blank(row_number, column_index, None)
            options: Dict[str, object] = {"object_position": 1}
            try:
                worksheet.insert_object(row_number, column_index, str(artifact.object_path), options)
            except Exception:
                log.exception("Failed to insert Origin object %s", artifact.object_path)
                continue
            ensure_row_height(
                row_number,
                int(target_height_px),
                _excel_row_height(target_height_in),
            )
            ensure_column_width(
                column_index,
                int(target_width_px),
                _excel_column_width(target_width_in),
            )

        if highlight_map:
            highlight_format = workbook.add_format({"bg_color": "#FFF4B5"})
            for column, rows in highlight_map.items():
                if column not in dataframe.columns:
                    continue
                column_index = _column_index(column)
                if column_index is None:
                    continue
                for row_idx in rows:
                    if row_idx < 0 or row_idx >= len(reset_df):
                        continue
                    value = reset_df.iloc[row_idx, column_index]
                    excel_row = row_idx + 1
                    if pd.isna(value):
                        worksheet.write_blank(excel_row, column_index, None, highlight_format)
                    elif isinstance(value, (int, float)) and math.isfinite(float(value)):
                        worksheet.write_number(excel_row, column_index, float(value), highlight_format)
                    else:
                        worksheet.write(excel_row, column_index, value, highlight_format)

    excel_path = getattr(writer, "path", None)
    if isinstance(excel_path, str):
        try:
            _adjust_drawing_ext_dimensions(Path(excel_path), figure_size, log)
        except Exception:
            log.exception("Failed to adjust drawing metadata for %s", excel_path)

def _origin_object_name(obj: object) -> Optional[str]:
    if obj is None:
        return None
    for attr in ("lt_name", "name", "lname", "long_name"):
        try:
            value = getattr(obj, attr, None)
        except Exception:
            continue
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _normalise_output_name(name: str) -> str:
    cleaned = "".join("_" if ch in _INVALID_FILENAME_CHARS else ch for ch in name.strip())
    cleaned = cleaned.strip(".")
    if not cleaned:
        return DEFAULT_OUTPUT_NAME
    return cleaned


def _safe_plot_stem(stem: str) -> str:
    normalised = unicodedata.normalize("NFKC", stem)
    cleaned_chars: list[str] = []
    invalid = _INVALID_FILENAME_CHARS | {os.sep}
    if os.altsep:
        invalid.add(os.altsep)
    for ch in normalised:
        if ch in invalid or ord(ch) < 32:
            cleaned_chars.append("_")
        else:
            cleaned_chars.append(ch)
    cleaned = "".join(cleaned_chars).strip("._ ")
    if not cleaned:
        return "measurement"
    return cleaned


def build_database(
    config: BuilderConfig,
    logger: Optional[logging.Logger] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    analysis_progress_callback: Optional[Callable[[int, int], None]] = None,
    analysis_total: Optional[int] = None,
    root_for_relpaths: Optional[Path] = None,
    skip_exports: bool = False,
    *,
    fabrication_index: FabricationIndex | None = None,
    measurement_records: Optional[Iterable[MeasurementRecord]] = None,
    microscope_index: Optional[
        Dict[Tuple[str, int, int, Optional[str]], MicroscopeMeasurements]
    ] = None,
    video_index: Optional[
        Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary]
    ] = None,
    strain_records: Optional[Dict[Tuple[str, int, int, Optional[str]], StrainRecord]] = None,
    strain_entries: Optional[Dict[str, Dict[str, object]]] = None,
    vsm_hysteresis_records: Optional[Iterable[VsmHysteresisRecord]] = None,
    vsm_temperature_scan_records: Optional[Iterable[VsmTemperatureScanRecord]] = None,
    dma_iso_stress_records: Optional[Iterable[DmaIsoStressRecord]] = None,
    mini_dma_records: Optional[Iterable[MiniDmaRecord]] = None,
    shape_memory_stress_strain_records: Optional[
        Iterable[ShapeMemoryStressStrainRecord]
    ] = None,
    shape_memory_entries: Optional[Dict[str, Dict[str, object]]] = None,
    fmr_records: Optional[Iterable[FmrRecord]] = None,
    phase_points: Optional[Dict[str, Dict[str, float]]] = None,
    transition_temps: Optional[Dict[str, Dict[str, object]]] = None,
    current_density_entries: Optional[Dict[str, Dict[str, object]]] = None,
    mini_dma_transition_reviews: Optional[Dict[str, Dict[str, object]]] = None,
    video_overrides: Optional[Dict[str, Dict[str, object]]] = None,
    include_fabrication_draw_siblings: bool = False,
) -> BuildResult:
    log = _logger(logger)
    output_dir = config.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = _normalise_output_name(getattr(config, "output_name", DEFAULT_OUTPUT_NAME))

    include_crops = bool(getattr(config, "include_microscope_crops", True))
    highlight_ocr = bool(getattr(config, "highlight_ocr_values", True))
    output_columns = list(OUTPUT_COLUMNS)
    d_column = DIAMETER_COLUMN
    D_column = GLASS_DIAMETER_COLUMN
    ratio_column = DIAMETER_RATIO_COLUMN
    microscope_image_columns: Tuple[str, ...] = ()
    if include_crops:
        microscope_image_columns = MICROSCOPE_IMAGE_COLUMNS
    if "d (µm)" in output_columns and MICROSCOPE_IMAGE_COLUMNS[0] not in output_columns:
        d_index = output_columns.index("d (µm)")
        output_columns.insert(d_index + 1, MICROSCOPE_IMAGE_COLUMNS[0])
    if "D (µm)" in output_columns and MICROSCOPE_IMAGE_COLUMNS[1] not in output_columns:
        D_index = output_columns.index("D (µm)")
        output_columns.insert(D_index + 1, MICROSCOPE_IMAGE_COLUMNS[1])

    raw_backends = tuple(config.plot_backends) if config.plot_backends else ()
    cleaned_backends = []
    for backend in raw_backends:
        if isinstance(backend, str):
            cleaned_backends.append(backend.lower())
    if config.make_plots and not cleaned_backends:
        cleaned_backends.append("matplotlib")
    deduped_backends = tuple(dict.fromkeys(cleaned_backends))
    recognised_backends = {"matplotlib", "origin"}
    for backend in deduped_backends:
        if backend not in recognised_backends:
            log.warning("Unsupported plot backend '%s'; skipping", backend)
    plot_backends = tuple(b for b in deduped_backends if b in recognised_backends)

    wants_matplotlib = "matplotlib" in plot_backends
    wants_origin_requested = "origin" in plot_backends
    wants_word_reports = not skip_exports and any(
        str(fmt).lower() in {"word", "docx", "word_reports"}
        for fmt in (config.export_formats or ())
    )

    fabrication_index = fabrication_index or build_fabrication_index(
        config.fabrication_files, log
    )
    stats = BuildStats()
    grouped: Dict[Tuple[str, int, int, Optional[str]], List[MeasurementRecord]] = {}
    plot_records: List[str] = []
    plot_cache: Dict[str, Path] = {}
    plot_name_to_path: Dict[str, Path] = {}
    origin_artifacts: Dict[str, OriginArtifact] = {}
    origin_cache: Dict[str, OriginArtifact] = {}
    figure_size = _normalise_figsize(
        getattr(config, "matplotlib_figsize", DEFAULT_FIGSIZE)
    )
    plot_dir = output_dir / config.plot_dir_name
    origin_dir = output_dir / config.origin_dir_name
    origin_enabled = wants_origin_requested
    origin_disabled_reason: Optional[str] = None
    manual_microscope_files: list[Path] = []
    auto_microscope_files: list[Path] = []
    microscope_sources: list[Path] = []
    if microscope_index is None:
        manual_microscope_files = [Path(p) for p in config.microscope_files]
        auto_microscope_files = _auto_discover_microscope_paths(
            config.annealing_files, log
        )
        microscope_sources = list(
            dict.fromkeys(manual_microscope_files + auto_microscope_files)
        )
    else:
        microscope_index = dict(microscope_index)
    video_sources: list[Path] = []
    if video_index is None:
        video_sources = list(dict.fromkeys(Path(p) for p in config.video_files))
    else:
        video_index = dict(video_index)
    if isinstance(video_overrides, dict):
        video_overrides_map: Dict[str, Dict[str, object]] = {
            str(key): dict(payload)
            for key, payload in video_overrides.items()
            if isinstance(payload, dict)
        }
    else:
        video_overrides_map = {}

    cumulative_lengths: Dict[Tuple[str, int, int], Optional[float]] = {}
    if fabrication_index.piece_level:
        by_draw: Dict[Tuple[str, int], List[Tuple[int, Optional[float]]]] = {}
        for (composition, draw_x, piece_y), piece_record in fabrication_index.piece_level.items():
            length_val = _value_for_output(piece_record, "length_m")
            length_num = _parse_numeric(length_val) if length_val is not None else None
            by_draw.setdefault((composition, int(draw_x)), []).append((int(piece_y), length_num))
        for key_str, overrides in video_overrides_map.items():
            if not isinstance(overrides, dict):
                continue
            length_override = overrides.get("Length (m)")
            length_num = _parse_numeric(length_override) if length_override is not None else None
            if length_num is None:
                continue
            parts = _microwire_key_from_string(str(key_str))
            if parts is None:
                continue
            composition, draw_x, piece_y, _suffix = parts
            by_draw.setdefault((composition, int(draw_x)), []).append((int(piece_y), length_num))
        for (composition, draw_x), entries in by_draw.items():
            running: Optional[float] = 0.0
            seen_pieces: Set[int] = set()
            for piece_y, length_num in sorted(entries, key=lambda item: item[0]):
                if piece_y in seen_pieces:
                    continue
                seen_pieces.add(piece_y)
                if running is None or length_num is None:
                    running = None
                    cumulative_lengths[(composition, draw_x, piece_y)] = None
                else:
                    running += length_num
                    cumulative_lengths[(composition, draw_x, piece_y)] = running
    if strain_records is None:
        strain_records = _load_strain_records(getattr(config, "strain_files", []), log)
    else:
        strain_records = dict(strain_records)

    def _normalise_record_key(
        value: object,
    ) -> Optional[Tuple[str, int, int, Optional[str]]]:
        return _split_microwire_key(value)

    def _record_key(
        record: object,
    ) -> Optional[Tuple[str, int, int, Optional[str]]]:
        key = _normalise_record_key(getattr(record, "key", None))
        if key is not None:
            return key
        path_value = getattr(record, "path", None)
        if isinstance(path_value, Path):
            key = _microscope_key(path_value)
        elif isinstance(path_value, str):
            try:
                key = _microscope_key(Path(path_value))
            except Exception:
                key = None
        if key is not None:
            return key
        sample = getattr(record, "sample", None)
        if isinstance(sample, str) and sample.strip():
            try:
                return _microscope_key(Path(sample))
            except Exception:
                return None
        return None

    def _record_label(record: object) -> str:
        label = getattr(record, "label", None)
        if isinstance(label, str) and label.strip():
            return label.strip()
        path_value = getattr(record, "path", None)
        if isinstance(path_value, Path):
            return path_value.name
        if isinstance(path_value, str):
            return Path(path_value).name
        sample = getattr(record, "sample", None)
        if isinstance(sample, str) and sample.strip():
            return sample.strip()
        return ""

    def _group_records(
        records: Optional[Iterable[object]],
    ) -> Dict[Tuple[str, int, int, Optional[str]], List[object]]:
        grouped: Dict[Tuple[str, int, int, Optional[str]], List[object]] = {}
        if not records:
            return grouped
        for record in records:
            key = _record_key(record)
            if key is None:
                continue
            grouped.setdefault(key, []).append(record)
        return grouped

    vsm_hysteresis_groups = _group_records(vsm_hysteresis_records)
    vsm_temperature_groups = _group_records(vsm_temperature_scan_records)
    dma_isostress_groups = _group_records(dma_iso_stress_records)
    mini_dma_groups = _group_records(mini_dma_records)
    shape_memory_stress_strain_groups = _group_records(
        shape_memory_stress_strain_records
    )
    fmr_groups = _group_records(fmr_records)

    def _assign_pyplot_origin_artifacts(
        row: Dict[str, object],
        *,
        records: Iterable[object],
        origin_column: str,
        plugin_name: str,
        display_prefix: str,
        section_token: str,
    ) -> None:
        nonlocal origin_enabled, origin_disabled_reason
        if not (origin_enabled and wants_word_reports):
            return
        paths: List[Path] = []
        for record in records:
            path_value = getattr(record, "path", None)
            if isinstance(path_value, Path):
                path = path_value
            elif isinstance(path_value, str) and path_value.strip():
                path = Path(path_value)
            else:
                continue
            if path.exists() and path not in paths:
                paths.append(path)
        if not paths:
            return
        sample_title = " ".join(
            part
            for part in (
                str(row.get("Composition") or "").strip(),
                str(row.get("Microwire") or "").strip(),
            )
            if part
        ).strip()
        descriptor_prefix = _safe_plot_stem(
            "_".join(
                part
                for part in (
                    sample_title or "sample",
                    section_token,
                )
                if part
            )
        )
        try:
            artifacts = export_pyplot_origin_artifacts_for_paths(
                paths=paths,
                plugin_name=plugin_name,
                origin_dir=origin_dir,
                descriptor_prefix=descriptor_prefix,
                display_prefix=display_prefix,
                log=log,
            )
        except RuntimeError as exc:
            if origin_disabled_reason is None:
                origin_disabled_reason = str(exc) or exc.__class__.__name__
                log.warning("Origin plotting disabled: %s", origin_disabled_reason)
            origin_enabled = False
            return
        except Exception:
            log.exception(
                "Failed to generate %s Origin object(s) for %s",
                display_prefix,
                sample_title or "sample",
            )
            return
        descriptors: List[str] = []
        for artifact in artifacts:
            if not artifact.descriptor:
                continue
            descriptors.append(artifact.descriptor)
            origin_artifacts.setdefault(artifact.descriptor, artifact)
        collapsed = _collapse_asset_references(descriptors)
        if collapsed is not None:
            row[origin_column] = collapsed

    def _pyplot_current_annealing_origin_artifact(
        record: MeasurementRecord,
    ) -> Optional[OriginArtifact]:
        nonlocal origin_enabled, origin_disabled_reason
        path = getattr(record, "path", None)
        if not isinstance(path, Path) or not path.exists():
            return None
        cache_key = f"pyplot-current::{_measurement_cache_key(record)}"
        cached = origin_cache.get(cache_key)
        if cached is not None:
            return cached
        sample_title = " ".join(
            part
            for part in (
                str(record.metadata.composition_token or "").strip(),
                f"{record.metadata.draw_x}/{record.metadata.piece_y}"
                if record.metadata.draw_x is not None and record.metadata.piece_y is not None
                else "",
            )
            if part
        ).strip()
        descriptor_prefix = _safe_plot_stem(
            "_".join(
                part
                for part in (
                    sample_title or "sample",
                    path.stem,
                    "current_annealing",
                )
                if part
            )
        )
        try:
            artifacts = export_pyplot_origin_artifacts_for_paths(
                paths=[path],
                plugin_name="Current Annealing",
                origin_dir=origin_dir,
                descriptor_prefix=descriptor_prefix,
                display_prefix="Current annealing Origin graph",
                log=log,
            )
        except RuntimeError as exc:
            if origin_disabled_reason is None:
                origin_disabled_reason = str(exc) or exc.__class__.__name__
                log.warning("Origin plotting disabled: %s", origin_disabled_reason)
            origin_enabled = False
            return None
        except Exception:
            log.exception("Failed to generate current annealing Origin object for %s", path)
            return None
        artifact = artifacts[0] if artifacts else None
        if artifact is not None:
            origin_cache[cache_key] = artifact
            origin_artifacts.setdefault(artifact.descriptor, artifact)
        return artifact

    phase_points_map: Dict[str, Dict[str, float]] = {}
    if phase_points:
        for key, payload in phase_points.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            key_parts = _microwire_key_from_string(key)
            if key_parts is None:
                continue
            cleaned = {
                label: float(value)
                for label, value in payload.items()
                if isinstance(value, (int, float))
            }
            if cleaned:
                phase_points_map[_microwire_key_to_str(key_parts)] = cleaned
    phase_points_map = dict(phase_points_map)

    transition_temps_map: Dict[str, Dict[str, float]] = {}
    transition_temps_blocked_keys: Set[str] = set()
    transition_temps_status_map: Dict[str, str] = {}
    transition_temps_counts_map: Dict[str, Dict[str, int]] = {}
    if transition_temps:
        for key, payload in transition_temps.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            key_parts = _microwire_key_from_string(key)
            if key_parts is None:
                continue
            key_str = _microwire_key_to_str(key_parts)
            status = str(
                payload.get("__review_status__")
                or payload.get("__status__")
                or payload.get("Review status")
                or payload.get("status")
                or ""
            ).strip()
            counts = _review_counts_from_payload(
                payload.get("__review_counts__")
                or payload.get("__counts__")
                or payload.get("Review counts")
            )
            status_label = _normalise_transition_status(status)
            if status:
                transition_temps_status_map[key_str] = status_label
            if counts.get("total", 0):
                transition_temps_counts_map[key_str] = counts
            included = payload.get("__included__", payload.get("included", None))
            blocked = status_label in {"No transition", "Excluded"} or included is False
            cleaned = {
                label: float(value)
                for label, value in payload.items()
                if label in ("As", "Af", "Ms", "Mf")
                and isinstance(value, (int, float))
                and math.isfinite(float(value))
            }
            if cleaned:
                transition_temps_map[key_str] = cleaned
                transition_temps_status_map.setdefault(key_str, "Manual adjusted")
            elif blocked:
                transition_temps_blocked_keys.add(key_str)
                transition_temps_status_map.setdefault(
                    key_str,
                    "Excluded" if status_label == "Excluded" else "No transition",
                )
                if key_str not in transition_temps_counts_map:
                    counts = _empty_transition_review_counts()
                    counts["total"] = 1
                    if transition_temps_status_map[key_str] == "Excluded":
                        counts["excluded"] = 1
                    else:
                        counts["no_transition"] = 1
                    transition_temps_counts_map[key_str] = counts
    transition_temps_map = dict(transition_temps_map)
    if vsm_temperature_groups:
        for key, records in vsm_temperature_groups.items():
            key_str = _microwire_key_to_str(key)
            if key_str in transition_temps_blocked_keys and key_str not in transition_temps_map:
                continue
            entry = transition_temps_map.setdefault(key_str, {})
            for record in records:
                data = getattr(record, "data", None)
                if not isinstance(data, pd.DataFrame) or data.empty:
                    continue
                try:
                    estimated = estimate_temperature_transition_points(data)
                except Exception:
                    log.exception(
                        "Failed to estimate transition temps from VSM temperature scan %s",
                        getattr(record, "path", ""),
                    )
                    continue
                for label in ("As", "Af", "Ms", "Mf"):
                    if label not in entry and estimated.get(label) is not None:
                        entry[label] = float(estimated[label])
            if not entry:
                transition_temps_map.pop(key_str, None)

    current_density_map: Dict[str, Dict[str, object]] = {}
    if current_density_entries:
        for key, payload in current_density_entries.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            key_parts = _microwire_key_from_string(key)
            if key_parts is None:
                continue
            entry: Dict[str, object] = {}
            for column in CURRENT_DENSITY_EXTRA_COLUMNS:
                if column in payload:
                    entry[column] = payload.get(column)
            if "Notes" in payload:
                entry["Notes"] = payload.get("Notes")
            if entry:
                current_density_map[_microwire_key_to_str(key_parts)] = entry
    current_density_map = dict(current_density_map)
    current_density_phase_columns = {
        "As1 (mA)": "J_As1 (A/mm^2)",
        "Af1 (mA)": "J_Af1 (A/mm^2)",
        "Ms1 (mA)": "J_Ms1 (A/mm^2)",
        "Mf1 (mA)": "J_Mf1 (A/mm^2)",
        "As2 (mA)": "J_As2 (A/mm^2)",
        "Af2 (mA)": "J_Af2 (A/mm^2)",
        "Ms2 (mA)": "J_Ms2 (A/mm^2)",
        "Mf2 (mA)": "J_Mf2 (A/mm^2)",
    }

    mini_dma_review_map: Dict[str, Dict[str, object]] = {}
    if mini_dma_transition_reviews:
        for record_id, payload in mini_dma_transition_reviews.items():
            if isinstance(record_id, str) and record_id.strip() and isinstance(payload, dict):
                mini_dma_review_map[record_id] = dict(payload)

    def _mini_dma_review_record_prefix(record: MiniDmaRecord) -> str:
        path = getattr(record, "path", None)
        if isinstance(path, Path):
            try:
                path_text = str(path.resolve())
            except Exception:
                path_text = str(path)
        else:
            path_text = repr(record)
        return f"{path_text}::"

    def _mini_dma_transition_status_for_records(
        records: Sequence[MiniDmaRecord],
    ) -> Tuple[str, str]:
        if not records:
            return "Not measured", ""
        counts = _empty_transition_review_counts()
        saw_targets = False
        for record in records:
            prefix = _mini_dma_review_record_prefix(record)
            target_labels: List[str] = []
            for line in getattr(record, "transition_summary", ()) or ():
                target = str(line).split(":", 1)[0].strip()
                if target and target not in target_labels:
                    target_labels.append(target)
            matching_reviews: Dict[str, Mapping[str, object]] = {}
            for record_id, payload in mini_dma_review_map.items():
                if not record_id.startswith(prefix):
                    continue
                target = str(payload.get("target_label") or record_id[len(prefix) :]).strip()
                if not target:
                    continue
                if target not in target_labels:
                    target_labels.append(target)
                matching_reviews[target] = payload
            for target in target_labels:
                saw_targets = True
                review = matching_reviews.get(target)
                if review is None:
                    counts["total"] += 1
                    counts["unreviewed"] += 1
                    counts["auto_candidates"] += 1
                    continue
                status = review.get("status")
                manual = _clean_review_values(review.get("manual_values_mA"))
                auto = _clean_review_values(review.get("auto_values_mA"))
                _increment_review_count_for_status(
                    counts,
                    status,
                    has_manual_values=bool(manual),
                    has_auto_values=bool(auto),
                )
        if not saw_targets:
            return "Not measured", ""
        return _aggregate_transition_review_status(counts), _format_transition_review_counts(counts)

    def _backfill_current_densities(row: Dict[str, object]) -> None:
        diameter_um = _parse_numeric(row.get(d_column))
        if diameter_um is None or diameter_um <= 0:
            return
        diameter_mm = diameter_um / 1000.0
        area_mm2 = math.pi * (diameter_mm / 2.0) ** 2
        if area_mm2 <= 0 or not math.isfinite(area_mm2):
            return
        for current_column, density_column in current_density_phase_columns.items():
            if density_column not in output_columns:
                continue
            existing = row.get(density_column)
            if existing not in (None, "") and not _is_nan(existing):
                continue
            current_mA = _parse_numeric(row.get(current_column))
            if current_mA is None or not math.isfinite(current_mA):
                continue
            row[density_column] = (current_mA / 1000.0) / area_mm2
        if (
            "As current density (A/mm^2)" in output_columns
            and (row.get("As current density (A/mm^2)") in (None, "") or _is_nan(row.get("As current density (A/mm^2)")))
            and row.get("J_As1 (A/mm^2)") not in (None, "")
            and not _is_nan(row.get("J_As1 (A/mm^2)"))
        ):
            row["As current density (A/mm^2)"] = row.get("J_As1 (A/mm^2)")
        if (
            "Ms current density (A/mm^2)" in output_columns
            and (row.get("Ms current density (A/mm^2)") in (None, "") or _is_nan(row.get("Ms current density (A/mm^2)")))
            and row.get("J_Ms1 (A/mm^2)") not in (None, "")
            and not _is_nan(row.get("J_Ms1 (A/mm^2)"))
        ):
            row["Ms current density (A/mm^2)"] = row.get("J_Ms1 (A/mm^2)")

    shape_memory_entry_map: Dict[str, Dict[str, object]] = {}
    if shape_memory_entries:
        for key, payload in shape_memory_entries.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            key_parts = _microwire_key_from_string(key)
            if key_parts is None:
                continue
            entry: Dict[str, object] = {}
            for column in SHAPE_MEMORY_VALUE_COLUMNS + SHAPE_MEMORY_FRACTURE_COLUMNS:
                if column in payload:
                    entry[column] = payload.get(column)
                    continue
                for old_name, new_name in SHAPE_MEMORY_ENTRY_ALIASES.items():
                    if new_name == column and old_name in payload:
                        entry[column] = payload.get(old_name)
                        break
            if entry:
                shape_memory_entry_map[_microwire_key_to_str(key_parts)] = entry
    shape_memory_entry_map = dict(shape_memory_entry_map)

    strain_entry_map: Dict[str, Dict[str, object]] = {}
    if strain_entries:
        for key, payload in strain_entries.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            key_parts = _microwire_key_from_string(key)
            if key_parts is None:
                continue
            entry: Dict[str, object] = {}
            for column in list(STRAIN_EXTRA_COLUMNS) + [STRAIN_COLUMN]:
                if column in payload:
                    entry[column] = payload.get(column)
                    continue
                for old_name, new_name in STRAIN_ENTRY_ALIASES.items():
                    if new_name == column and old_name in payload:
                        entry[column] = payload.get(old_name)
                        break
            if entry:
                strain_entry_map[_microwire_key_to_str(key_parts)] = entry
    elif strain_records:
        for key, record in strain_records.items():
            parts = _split_microwire_key(key)
            if parts is None:
                continue
            composition, draw_x, piece_y, suffix = parts
            entry: Dict[str, object] = {}
            if record.m_length is not None:
                entry["M length"] = record.m_length
            if record.a_length is not None:
                entry["A length"] = record.a_length
            entry["Broke"] = bool(record.broke)
            if entry:
                strain_entry_map[_microwire_key_to_str((composition, draw_x, piece_y, suffix))] = entry
    strain_entry_map = dict(strain_entry_map)

    analysis_total_local = analysis_total
    if analysis_total_local is None:
        analysis_total_local = len(microscope_sources) + len(video_sources)
    analysis_done = 0

    def _analysis_notify() -> None:
        if analysis_progress_callback is not None and analysis_total_local > 0:
            total_units = analysis_total_local
            current_units = min(analysis_done, total_units)
            analysis_progress_callback(current_units, total_units)

    last_micro = 0

    def _micro_progress(processed: int, _stage_total: int) -> None:
        nonlocal analysis_done, last_micro
        delta = max(processed - last_micro, 0)
        if delta:
            analysis_done += delta
            _analysis_notify()
        last_micro = processed

    last_video = 0

    def _video_progress(processed: int, _stage_total: int) -> None:
        nonlocal analysis_done, last_video
        delta = max(processed - last_video, 0)
        if delta:
            analysis_done += delta
            _analysis_notify()
        last_video = processed

    micro_callback = _micro_progress if analysis_total_local else None
    video_callback = _video_progress if analysis_total_local else None

    if analysis_total_local:
        _analysis_notify()

    if microscope_index is None:
        microscope_index, _ = _group_microscope_measurements(
            microscope_sources,
            log,
            progress_callback=micro_callback,
        )
    if video_index is None:
        video_index = _collect_video_metrics(
            video_sources,
            log,
            progress_callback=video_callback,
        )
    microscope_crop_dir = output_dir / "microscope_crops"
    microscope_crop_map: Dict[str, Path] = {}
    ocr_highlights: Dict[str, Set[int]] = {}
    if analysis_total_local and analysis_done < analysis_total_local:
        analysis_done = analysis_total_local
        _analysis_notify()
    if measurement_records is not None:
        records_list = [record for record in measurement_records]
        total = len(records_list)
        for idx, record in enumerate(records_list, start=1):
            metadata = record.metadata
            if not record.sanity_ok:
                stats.resistance_checks_failed += 1
            if metadata.draw_x is None or metadata.piece_y is None:
                log.warning(
                    "Skipping %s because the microwire draw/piece identifiers could not be parsed",
                    record.path,
                )
                stats.skipped += 1
            else:
                path = getattr(record, "path", None)
                path_obj: Optional[Path]
                if isinstance(path, Path):
                    path_obj = path
                elif isinstance(path, str) and path:
                    path_obj = Path(path)
                else:
                    path_obj = None
                suffix = _suffix_from_path(path_obj, metadata.draw_x, metadata.piece_y)
                key = (
                    metadata.composition_token,
                    metadata.draw_x,
                    metadata.piece_y,
                    suffix,
                )
                grouped.setdefault(key, []).append(record)
                stats.parsed += 1
            if progress_callback:
                try:
                    progress_callback(idx, total)
                except Exception:
                    pass
    else:
        total = len(config.annealing_files)
        for idx, path in enumerate(config.annealing_files, start=1):
            metadata = _metadata_from_path(path, root_for_relpaths)
            try:
                df = _load_annealing(path, expected_setpoint_mA=metadata.setpoint_mA)
            except Exception:
                log.exception("Failed to parse %s", path)
                stats.skipped += 1
                if progress_callback:
                    progress_callback(idx, total)
                continue
            ok, mean_error = _resistance_sanity_check(df)
            if not ok:
                stats.resistance_checks_failed += 1
                if mean_error is None:
                    log.warning("Râ‰ˆV/I sanity check failed for %s", path)
                else:
                    log.warning(
                        "Râ‰ˆV/I sanity check failed for %s (mean error %.2f%%)",
                        path,
                        mean_error * 100,
                    )
            key = _microscope_key(path)
            if key is None:
                if metadata.draw_x is None or metadata.piece_y is None:
                    log.warning(
                        "Skipping %s because the microwire draw/piece identifiers could not be parsed",
                        path,
                    )
                    stats.skipped += 1
                    if progress_callback:
                        progress_callback(idx, total)
                    continue
                key = (metadata.composition_token, metadata.draw_x, metadata.piece_y, None)
            record = MeasurementRecord(
                path=path,
                metadata=metadata,
                dataframe=df,
                sanity_ok=ok,
                sanity_error=mean_error,
                transition_summary=_annealing_transition_summary(df, label=metadata.file_name),
            )
            grouped.setdefault(key, []).append(record)
            stats.parsed += 1
            if progress_callback:
                progress_callback(idx, total)
    rows: List[Dict[str, object]] = []

    if microscope_index:
        for candidate_key in microscope_index.keys():
            if candidate_key not in grouped:
                grouped[candidate_key] = []
    for collection in (
        vsm_hysteresis_groups,
        vsm_temperature_groups,
        dma_isostress_groups,
        mini_dma_groups,
        shape_memory_stress_strain_groups,
        fmr_groups,
        strain_records,
    ):
        for candidate_key in collection.keys():
            if candidate_key not in grouped:
                grouped[candidate_key] = []

    if include_fabrication_draw_siblings and fabrication_index.piece_level:
        relevant_piece_limits: Dict[Tuple[str, int], Optional[int]] = {}

        def _register_draw(raw_key: object) -> None:
            parts = _split_microwire_key(raw_key)
            if parts is None:
                return
            composition, draw_x, piece_y, _suffix = parts
            key = (composition, int(draw_x))
            if piece_y is None:
                relevant_piece_limits[key] = None
                return
            piece_int = int(piece_y)
            if piece_int <= 0:
                return
            current_limit = relevant_piece_limits.get(key)
            if current_limit is None and key in relevant_piece_limits:
                return
            relevant_piece_limits[key] = max(int(current_limit or 0), piece_int)

        for raw_key in grouped.keys():
            _register_draw(raw_key)
        for collection in (
            vsm_hysteresis_groups,
            vsm_temperature_groups,
            dma_isostress_groups,
            mini_dma_groups,
            shape_memory_stress_strain_groups,
            fmr_groups,
            strain_records,
        ):
            for raw_key in collection.keys():
                _register_draw(raw_key)
        for collection in (
            phase_points_map,
            transition_temps_map,
            current_density_map,
            shape_memory_entry_map,
            strain_entry_map,
            video_overrides_map,
        ):
            for raw_key in collection.keys():
                parsed = _microwire_key_from_string(str(raw_key))
                if parsed is None:
                    continue
                composition, draw_x, _piece_y, _suffix = parsed
                key = (composition, int(draw_x))
                if piece_y is None:
                    relevant_piece_limits[key] = None
                    continue
                piece_int = int(piece_y)
                if piece_int <= 0:
                    continue
                current_limit = relevant_piece_limits.get(key)
                if current_limit is None and key in relevant_piece_limits:
                    continue
                relevant_piece_limits[key] = max(int(current_limit or 0), piece_int)

        sibling_pieces_by_draw: Dict[Tuple[str, int], Set[int]] = {}
        sibling_records_by_draw: Dict[Tuple[str, int], Dict[int, Mapping[str, object]]] = {}
        for (composition, draw_x, piece_y), _piece_record in fabrication_index.piece_level.items():
            sibling_pieces_by_draw.setdefault((composition, int(draw_x)), set()).add(int(piece_y))
            sibling_records_by_draw.setdefault((composition, int(draw_x)), {})[int(piece_y)] = dict(_piece_record)

        for (composition, draw_x), piece_limit in sorted(relevant_piece_limits.items()):
            candidate_pieces = _relevant_sibling_piece_candidates(
                sibling_pieces_by_draw.get((composition, int(draw_x)), ()),
                None if piece_limit is None else [piece_limit],
                sibling_records_by_draw.get((composition, int(draw_x))),
            )
            for piece_y in candidate_pieces:
                grouped.setdefault((composition, int(draw_x), int(piece_y), None), [])

    def _group_sort_key(
        item: Tuple[Tuple[str, int, int, Optional[str]], List[MeasurementRecord]]
    ) -> Tuple[str, int, int, str]:
        key, _records = item
        parts = _split_microwire_key(key)
        if parts is None:
            return ("", 0, 0, "")
        composition, draw_x, piece_y, suffix = parts
        return (composition.lower(), draw_x, piece_y, (suffix or "").lower())

    for key, records in sorted(grouped.items(), key=_group_sort_key):
        parts = _split_microwire_key(key)
        if parts is None:
            continue
        composition, draw_x, piece_y, suffix = parts
        key = parts
        draw_info = fabrication_index.get_draw(composition, draw_x)
        piece_info = fabrication_index.get_piece(composition, draw_x, piece_y)
        row: Dict[str, object] = {column: None for column in output_columns}
        row["Composition"] = composition
        row["Microwire"] = _microwire_label(draw_x, piece_y, suffix)
        ea_value = _compute_ea_from_composition(composition)
        row["e/a"] = ea_value
        row[ESTIMATED_TRANSITION_COLUMN] = _estimate_transition_temp_c(ea_value)
        key_str = _microwire_key_to_str(key)
        row["d (µm)"] = None
        row["D (µm)"] = None
        row[ratio_column] = None
        row[BRITTLE_COLUMN] = None
        row["Length (m)"] = _value_for_output(piece_info, "length_m")
        row["Production datetime"] = _value_for_output(draw_info, "production_datetime")
        row["Mass (g)"] = _value_for_output(draw_info, "mass_g")
        row["Resistance (Ω)"] = _value_for_output(draw_info, "fabrication_resistance_ohm")
        row[CORE_TEMPERATURE_COLUMN] = _value_for_output(
            draw_info,
            "fabrication_temperature_c",
        )
        row[GLASS_TEMPERATURE_COLUMN] = _value_for_output(
            draw_info,
            "fabrication_glass_temperature_c",
        )
        row["Winding speed (m/min)"] = _value_for_output(draw_info, "winding_speed_m_per_min")
        row["Glass feeding (mm/min)"] = _value_for_output(draw_info, "glass_feed_mm_per_min")
        row["Underpressure"] = _value_for_output(draw_info, "underpressure")
        pull_value = _value_for_output(piece_info, "glass_pull_off")
        if pull_value is None:
            pull_value = _value_for_output(draw_info, "glass_pull_off")
        row[GLASS_PULL_COLUMN] = pull_value
        row["Notes"] = _compose_notes(draw_info, piece_info)
        row["Data source"] = (
            "Measured"
            if records
            else "Fabrication only"
            if piece_info or draw_info
            else "Microscope only"
        )
        phase_entry = phase_points_map.get(key_str, {})
        if phase_entry:
            as_value = phase_entry.get("As1")
            if as_value is None:
                as_value = phase_entry.get("As")
            if as_value is not None:
                row["As (mA)"] = as_value
            ms_value = phase_entry.get("Ms1")
            if ms_value is None:
                ms_value = phase_entry.get("Ms")
            if ms_value is not None:
                row["Ms (mA)"] = ms_value
            phase_column_map = {
                "As1": "As1 (mA)",
                "Af1": "Af1 (mA)",
                "Ms1": "Ms1 (mA)",
                "Mf1": "Mf1 (mA)",
                "As2": "As2 (mA)",
                "Af2": "Af2 (mA)",
                "Ms2": "Ms2 (mA)",
                "Mf2": "Mf2 (mA)",
            }
            for label, column in phase_column_map.items():
                if column not in output_columns:
                    continue
                current_value = row.get(column)
                if current_value is not None and not _is_nan(current_value):
                    continue
                value = phase_entry.get(label)
                if value is None and label == "As1":
                    value = phase_entry.get("As")
                if value is None and label == "Ms1":
                    value = phase_entry.get("Ms")
                if value is None:
                    continue
                row[column] = value
        transition_entry = transition_temps_map.get(key_str, {})
        if transition_entry:
            if transition_entry.get("As") is not None:
                row[TRANSITION_TEMP_AS_COLUMN] = transition_entry.get("As")
            if transition_entry.get("Af") is not None:
                row[TRANSITION_TEMP_AF_COLUMN] = transition_entry.get("Af")
            if transition_entry.get("Ms") is not None:
                row[TRANSITION_TEMP_MS_COLUMN] = transition_entry.get("Ms")
            if transition_entry.get("Mf") is not None:
                row[TRANSITION_TEMP_MF_COLUMN] = transition_entry.get("Mf")
        current_density_entry = current_density_map.get(key_str, {})
        if current_density_entry:
            for column in CURRENT_DENSITY_EXTRA_COLUMNS:
                if column not in output_columns:
                    continue
                value = current_density_entry.get(column)
                if value is None or value == "":
                    continue
                if column in {"Setpoints (mA)", "Sources"} and isinstance(
                    value, (list, tuple, set)
                ):
                    value = ", ".join(str(item) for item in value)
                row[column] = value
            if row.get("As (mA)") in (None, "", float("nan")):
                as1_value = current_density_entry.get("As1 (mA)")
                if isinstance(as1_value, (int, float)) and math.isfinite(float(as1_value)):
                    row["As (mA)"] = float(as1_value)
            if row.get("Ms (mA)") in (None, "", float("nan")):
                ms1_value = current_density_entry.get("Ms1 (mA)")
                if isinstance(ms1_value, (int, float)) and math.isfinite(float(ms1_value)):
                    row["Ms (mA)"] = float(ms1_value)
            if (not row.get("Notes")) and current_density_entry.get("Notes"):
                row["Notes"] = current_density_entry.get("Notes")
        current_status_value = row.get(CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN)
        if current_status_value in (None, ""):
            current_counts = _empty_transition_review_counts()
            if records:
                current_counts["total"] = len(records)
                if current_density_entry:
                    current_counts["manual"] = len(records)
                    row[CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN] = "Manual adjusted"
                elif phase_entry:
                    current_counts["auto_candidates"] = len(records)
                    current_counts["unreviewed"] = len(records)
                    row[CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN] = "Auto candidate"
                else:
                    current_counts["unreviewed"] = len(records)
                    row[CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN] = "Unreviewed"
                if row.get(CURRENT_ANNEALING_TRANSITION_COUNTS_COLUMN) in (None, ""):
                    row[CURRENT_ANNEALING_TRANSITION_COUNTS_COLUMN] = _format_transition_review_counts(current_counts)
            else:
                row[CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN] = "Not measured"
        row_highlights: Set[str] = set()
        d_detection: Optional[MicroscopeDetection] = None
        D_detection: Optional[MicroscopeDetection] = None
        d_numeric = _parse_numeric(row["d (µm)"])
        D_numeric = _parse_numeric(row["D (µm)"])
        ratio_numeric = _parse_numeric(row[ratio_column])
        microscope_data = None
        if microscope_index:
            microscope_data = microscope_index.get(key)
            if microscope_data is None:
                microscope_data = microscope_index.get((composition, draw_x, piece_y, None))
            if microscope_data is None:
                microscope_data = microscope_index.get((composition, draw_x, piece_y))
            if microscope_data is None:
                for candidate_key, candidate_value in microscope_index.items():
                    try:
                        if (
                            isinstance(candidate_key, tuple)
                            and len(candidate_key) >= 3
                            and str(candidate_key[0]) == str(composition)
                            and int(candidate_key[1]) == int(draw_x)
                            and int(candidate_key[2]) == int(piece_y)
                        ):
                            microscope_data = candidate_value
                            break
                    except (TypeError, ValueError):
                        continue
        if microscope_data:
            if getattr(microscope_data, "brittle", False):
                row[BRITTLE_COLUMN] = "brittle"
            if d_numeric is None:
                d_detection = microscope_data.best_core_detection()
                if (
                    isinstance(d_detection, MicroscopeDetection)
                    and getattr(d_detection, "category", None) == "core"
                    and isinstance(d_detection.value, (int, float))
                    and math.isfinite(float(d_detection.value))
                ):
                    row[d_column] = float(d_detection.value)
                    d_numeric = float(d_detection.value)
                    if d_detection.source == "ocr":
                        row_highlights.add(d_column)
            if D_numeric is None:
                D_detection = microscope_data.best_glass_detection()
                if (
                    isinstance(D_detection, MicroscopeDetection)
                    and getattr(D_detection, "category", None) == "glass"
                    and isinstance(D_detection.value, (int, float))
                    and math.isfinite(float(D_detection.value))
                ):
                    row[D_column] = float(D_detection.value)
                    D_numeric = float(D_detection.value)
                    if D_detection.source == "ocr":
                        row_highlights.add(D_column)
            if ratio_numeric is None and d_numeric is not None and D_numeric not in (None, 0):
                try:
                    ratio = d_numeric / D_numeric
                except ZeroDivisionError:
                    ratio = None
                if ratio is not None and math.isfinite(ratio):
                    row[ratio_column] = ratio
                    ratio_numeric = ratio
        if include_crops:
            if MICROSCOPE_IMAGE_COLUMNS[0] in row:
                row[MICROSCOPE_IMAGE_COLUMNS[0]] = row.get(MICROSCOPE_IMAGE_COLUMNS[0])
            if MICROSCOPE_IMAGE_COLUMNS[1] in row:
                row[MICROSCOPE_IMAGE_COLUMNS[1]] = row.get(MICROSCOPE_IMAGE_COLUMNS[1])
            if d_detection and d_detection.image_path is not None:
                suffix_text = f"{suffix}" if suffix else ""
                crop_path = d_detection.ensure_crop(
                    microscope_crop_dir,
                    f"{composition}_{draw_x}_{piece_y}{suffix_text}_d",
                )
                if crop_path is not None:
                    try:
                        rel_path = crop_path.relative_to(output_dir)
                        rel_text = str(rel_path).replace(os.sep, "/")
                    except ValueError:
                        rel_text = str(crop_path)
                    row[MICROSCOPE_IMAGE_COLUMNS[0]] = rel_text
                    microscope_crop_map[rel_text] = crop_path
            if D_detection and D_detection.image_path is not None:
                suffix_text = f"{suffix}" if suffix else ""
                crop_path = D_detection.ensure_crop(
                    microscope_crop_dir,
                    f"{composition}_{draw_x}_{piece_y}{suffix_text}_D",
                )
                if crop_path is not None:
                    try:
                        rel_path = crop_path.relative_to(output_dir)
                        rel_text = str(rel_path).replace(os.sep, "/")
                    except ValueError:
                        rel_text = str(crop_path)
                    row[MICROSCOPE_IMAGE_COLUMNS[1]] = rel_text
                    microscope_crop_map[rel_text] = crop_path
        video_data = video_index.get(key) if video_index else None
        if video_data is None:
            video_data = video_index.get((composition, draw_x, piece_y))
        if video_data is None:
            video_data = video_index.get((composition, draw_x, None))
        if video_data:
            temp = video_data.temperature()
            if temp is not None:
                row[CORE_TEMPERATURE_COLUMN] = temp
                row_highlights.add(CORE_TEMPERATURE_COLUMN)
            under_value = video_data.underpressure()
            if under_value is not None:
                row["Underpressure"] = under_value
                row_highlights.add("Underpressure")
            wind = video_data.winding_speed()
            if wind is not None:
                row["Winding speed (m/min)"] = wind
                row_highlights.add("Winding speed (m/min)")
            glass = video_data.glass_feed()
            if glass is not None:
                row["Glass feeding (mm/min)"] = glass
                row_highlights.add("Glass feeding (mm/min)")
        overrides = video_overrides_map.get(key_str, {})
        if overrides:
            for column, value in overrides.items():
                if column in output_columns:
                    row[column] = value
        end_length = row.get(VIDEO_END_LENGTH_COLUMN)
        try:
            end_numeric = float(end_length) if end_length is not None else None
        except (TypeError, ValueError):
            end_numeric = None
        cumulative = cumulative_lengths.get((composition, draw_x, piece_y))
        if (
            end_numeric is not None
            and cumulative is not None
            and math.isfinite(end_numeric)
            and math.isfinite(cumulative)
        ):
            row[VIDEO_MW_LENGTH_COLUMN] = round(end_numeric - cumulative, 3)
        strain_record = (
            strain_records.get(key)
            or strain_records.get((composition, draw_x, piece_y, None))
            or strain_records.get((composition, draw_x, piece_y))
        )
        if strain_record:
            strain_value = _format_strain_value(strain_record)
            if strain_value is not None:
                row[STRAIN_COLUMN] = strain_value
        strain_entry = strain_entry_map.get(key_str, {})
        if strain_entry:
            for column in STRAIN_EXTRA_COLUMNS:
                if column not in output_columns:
                    continue
                value = strain_entry.get(column)
                if value is None or value == "":
                    continue
                row[column] = value
            if STRAIN_COLUMN in strain_entry:
                strain_value = strain_entry.get(STRAIN_COLUMN)
                if row.get(STRAIN_COLUMN) in (None, "") and strain_value not in (None, ""):
                    row[STRAIN_COLUMN] = strain_value
        vsm_records = vsm_hysteresis_groups.get(key, [])
        if not vsm_records:
            vsm_records = vsm_hysteresis_groups.get((composition, draw_x, piece_y, None), [])
        if not vsm_records:
            vsm_records = vsm_hysteresis_groups.get((composition, draw_x, piece_y), [])
        if vsm_records:
            labels = [_record_label(record) for record in vsm_records if _record_label(record)]
            if labels:
                row[VSM_HYSTERESIS_COLUMN] = list(dict.fromkeys(labels))
            _assign_pyplot_origin_artifacts(
                row,
                records=vsm_records,
                origin_column=VSM_HYSTERESIS_ORIGIN_COLUMN,
                plugin_name="VSM Hysteresis Loops",
                display_prefix="VSM hysteresis Origin graph",
                section_token="vsm_hysteresis",
            )
        vsm_scan_records = vsm_temperature_groups.get(key, [])
        if not vsm_scan_records:
            vsm_scan_records = vsm_temperature_groups.get((composition, draw_x, piece_y, None), [])
        if not vsm_scan_records:
            vsm_scan_records = vsm_temperature_groups.get((composition, draw_x, piece_y), [])
        if vsm_scan_records:
            labels = [_record_label(record) for record in vsm_scan_records if _record_label(record)]
            if labels:
                row[VSM_TEMPERATURE_SCAN_COLUMN] = list(dict.fromkeys(labels))
            _assign_pyplot_origin_artifacts(
                row,
                records=vsm_scan_records,
                origin_column=VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN,
                plugin_name="VSM Temperature Scan",
                display_prefix="VSM temperature scan Origin graph",
                section_token="vsm_temperature_scan",
            )
        vsm_status = transition_temps_status_map.get(key_str)
        vsm_counts = transition_temps_counts_map.get(key_str)
        if vsm_counts is None:
            vsm_counts = _empty_transition_review_counts()
            if vsm_scan_records:
                vsm_counts["total"] = len(vsm_scan_records)
                if transition_entry:
                    vsm_counts["accepted"] = len(vsm_scan_records)
                else:
                    vsm_counts["unreviewed"] = len(vsm_scan_records)
                    if vsm_status == "Auto candidate":
                        vsm_counts["auto_candidates"] = len(vsm_scan_records)
        if vsm_status is None:
            if not vsm_scan_records:
                vsm_status = "Not measured"
            elif transition_entry:
                vsm_status = "Auto candidate"
                if vsm_counts.get("total", 0):
                    vsm_counts["accepted"] = 0
                    vsm_counts["unreviewed"] = int(vsm_counts.get("total", 0))
                    vsm_counts["auto_candidates"] = int(vsm_counts.get("total", 0))
            else:
                vsm_status = "Unreviewed"
        row[VSM_TRANSITION_TEMP_STATUS_COLUMN] = vsm_status
        row[VSM_TRANSITION_TEMP_COUNTS_COLUMN] = _format_transition_review_counts(vsm_counts)
        dma_records = dma_isostress_groups.get(key, [])
        if not dma_records:
            dma_records = dma_isostress_groups.get((composition, draw_x, piece_y, None), [])
        if not dma_records:
            dma_records = dma_isostress_groups.get((composition, draw_x, piece_y), [])
        if dma_records:
            labels = [_record_label(record) for record in dma_records if _record_label(record)]
            if labels:
                row[DMA_ISOSTRESS_COLUMN] = list(dict.fromkeys(labels))
            _assign_pyplot_origin_artifacts(
                row,
                records=dma_records,
                origin_column=DMA_ISOSTRESS_ORIGIN_COLUMN,
                plugin_name="DMA Iso-Stress",
                display_prefix="DMA iso-stress Origin graph",
                section_token="dma_iso_stress",
            )
        mini_dma_entries = mini_dma_groups.get(key, [])
        if not mini_dma_entries:
            mini_dma_entries = mini_dma_groups.get((composition, draw_x, piece_y, None), [])
        if not mini_dma_entries:
            mini_dma_entries = mini_dma_groups.get((composition, draw_x, piece_y), [])
        if mini_dma_entries:
            labels = [_record_label(record) for record in mini_dma_entries if _record_label(record)]
            if labels:
                row[MINI_DMA_COLUMN] = list(dict.fromkeys(labels))
            strain_lines: List[str] = []
            transition_lines: List[str] = []
            break_lines: List[str] = []
            for record in mini_dma_entries:
                for line in getattr(record, "strain_summary", ()) or ():
                    if line and line not in strain_lines:
                        strain_lines.append(str(line))
                for line in getattr(record, "transition_summary", ()) or ():
                    if line and line not in transition_lines:
                        transition_lines.append(str(line))
                break_summary = getattr(record, "break_summary", "") or ""
                if break_summary and break_summary not in break_lines:
                    break_lines.append(str(break_summary))
            if strain_lines:
                row[MINI_DMA_STRAIN_COLUMN] = strain_lines
            if transition_lines:
                row[MINI_DMA_TRANSITION_COLUMN] = transition_lines
            (
                row[MINI_DMA_TRANSITION_STATUS_COLUMN],
                row[MINI_DMA_TRANSITION_COUNTS_COLUMN],
            ) = _mini_dma_transition_status_for_records(mini_dma_entries)
            if break_lines:
                row[MINI_DMA_BREAK_COLUMN] = list(dict.fromkeys(break_lines))
            _assign_pyplot_origin_artifacts(
                row,
                records=mini_dma_entries,
                origin_column=MINI_DMA_ORIGIN_COLUMN,
                plugin_name="Mini DMA",
                display_prefix="Mini DMA Origin graph",
                section_token="mini_dma",
            )
        else:
            row[MINI_DMA_TRANSITION_STATUS_COLUMN] = "Not measured"
        shape_memory_records = shape_memory_stress_strain_groups.get(key, [])
        if not shape_memory_records:
            shape_memory_records = shape_memory_stress_strain_groups.get(
                (composition, draw_x, piece_y, None),
                [],
            )
        if not shape_memory_records:
            shape_memory_records = shape_memory_stress_strain_groups.get(
                (composition, draw_x, piece_y),
                [],
            )
        if shape_memory_records:
            labels = [
                _record_label(record)
                for record in shape_memory_records
                if _record_label(record)
            ]
            if labels:
                row[SHAPE_MEMORY_STRESS_STRAIN_COLUMN] = list(dict.fromkeys(labels))
            _assign_pyplot_origin_artifacts(
                row,
                records=shape_memory_records,
                origin_column=SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN,
                plugin_name="Manual Stress/Strain",
                display_prefix="Manual stress/strain Origin graph",
                section_token="shape_memory_stress_strain",
            )
        shape_memory_entry = shape_memory_entry_map.get(key_str, {})
        if shape_memory_entry:
            for column in SHAPE_MEMORY_VALUE_COLUMNS + SHAPE_MEMORY_FRACTURE_COLUMNS:
                if column not in output_columns:
                    continue
                value = shape_memory_entry.get(column)
                if value in (None, ""):
                    continue
                row[column] = value
        fmr_entries = fmr_groups.get(key, [])
        if not fmr_entries:
            fmr_entries = fmr_groups.get((composition, draw_x, piece_y, None), [])
        if not fmr_entries:
            fmr_entries = fmr_groups.get((composition, draw_x, piece_y), [])
        if fmr_entries:
            labels = [_record_label(record) for record in fmr_entries if _record_label(record)]
            if labels:
                row[FMR_COLUMN] = list(dict.fromkeys(labels))
            _assign_pyplot_origin_artifacts(
                row,
                records=fmr_entries,
                origin_column=FMR_ORIGIN_COLUMN,
                plugin_name="FMR",
                display_prefix="FMR Origin graph",
                section_token="fmr",
            )
        ratio_display = _parse_numeric(row["d/D"])
        if ratio_display is not None:
            row["d/D"] = round(ratio_display, 3)
        if not draw_info:
            stats.missing_draw += 1
        if not piece_info:
            stats.missing_piece += 1
        high_record = _select_high_measurement(records)
        other_records = _select_other_measurements(records, high_record)
        if high_record:
            row["File 1000 mA"] = high_record.metadata.file_name
        elif records:
            stats.missing_high_measurement += 1
            log.warning("No 1000 mA measurement found for %s %s", composition, row["Microwire"] or "(unknown)")
        if other_records:
            row["Other annealing files"] = [
                record.metadata.file_name
                for record in other_records
                if getattr(record.metadata, "file_name", None)
            ]
        transition_lines: List[str] = []
        for record in [high_record, *other_records]:
            if record is None:
                continue
            lines = getattr(record, "transition_summary", ()) or ()
            if not lines:
                lines = _annealing_transition_summary(
                    record.dataframe,
                    label=getattr(record.metadata, "file_name", None),
                )
            for line in lines:
                if line and line not in transition_lines:
                    transition_lines.append(str(line))
        if transition_lines:
            row[ANNEALING_TRANSITION_COLUMN] = transition_lines
            if row.get(CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN) == "Unreviewed":
                row[CURRENT_ANNEALING_TRANSITION_STATUS_COLUMN] = "Auto candidate"
                current_counts = _empty_transition_review_counts()
                current_counts["total"] = len([record for record in [high_record, *other_records] if record is not None])
                current_counts["unreviewed"] = current_counts["total"]
                current_counts["auto_candidates"] = current_counts["total"]
                row[CURRENT_ANNEALING_TRANSITION_COUNTS_COLUMN] = _format_transition_review_counts(current_counts)
        if wants_matplotlib:
            if high_record:
                high_cache_key = _measurement_cache_key(high_record)
                cached = plot_cache.get(high_cache_key)
                if cached is None:
                    try:
                        cached = _plot_measurement_matplotlib(
                            high_record.dataframe,
                            high_record.path,
                            plot_dir,
                            figure_size,
                        )
                        plot_cache[high_cache_key] = cached
                    except Exception:
                        log.exception("Failed to generate plot for %s", high_record.path)
                        cached = None
                if cached is not None:
                    figure_name = Path(cached).name
                    row["Figure — 1000 mA"] = figure_name
                    plot_name_to_path.setdefault(figure_name, cached)
                    if figure_name not in plot_records:
                        plot_records.append(figure_name)
            if other_records:
                other_figures: List[str] = []
                for record in other_records:
                    record_cache_key = _measurement_cache_key(record)
                    cached = plot_cache.get(record_cache_key)
                    if cached is None:
                        try:
                            cached = _plot_measurement_matplotlib(
                                record.dataframe,
                                record.path,
                                plot_dir,
                                figure_size,
                            )
                            plot_cache[record_cache_key] = cached
                        except Exception:
                            log.exception("Failed to generate plot for %s", record.path)
                            cached = None
                    if cached is not None:
                        figure_name = Path(cached).name
                        other_figures.append(figure_name)
                        plot_name_to_path.setdefault(figure_name, cached)
                        if figure_name not in plot_records:
                            plot_records.append(figure_name)
                if other_figures:
                    row["Figure — other annealing"] = _collapse_asset_references(other_figures)
        if origin_enabled:
            if high_record:
                high_cache_key = _measurement_cache_key(high_record)
                cached_origin = origin_cache.get(high_cache_key)
                if cached_origin is None:
                    if wants_word_reports:
                        cached_origin = _pyplot_current_annealing_origin_artifact(high_record)
                        if cached_origin is not None:
                            origin_cache[high_cache_key] = cached_origin
                    else:
                        try:
                            cached_origin = _plot_measurement_origin(
                                high_record.dataframe,
                                high_record.path,
                                origin_dir,
                                log,
                            )
                        except RuntimeError as exc:
                            if origin_disabled_reason is None:
                                origin_disabled_reason = str(exc) or exc.__class__.__name__
                                log.warning("Origin plotting disabled: %s", origin_disabled_reason)
                            origin_enabled = False
                            cached_origin = None
                        except Exception:
                            log.exception("Failed to generate Origin plot for %s", high_record.path)
                            cached_origin = None
                        else:
                            if cached_origin is not None:
                                origin_cache[high_cache_key] = cached_origin
                                origin_artifacts.setdefault(cached_origin.descriptor, cached_origin)
                if cached_origin is not None:
                    row["Figure — 1000 mA (Origin)"] = cached_origin.descriptor
            if origin_enabled and other_records:
                other_descriptors: List[str] = []
                for record in other_records:
                    record_cache_key = _measurement_cache_key(record)
                    cached_origin = origin_cache.get(record_cache_key)
                    if cached_origin is None:
                        if wants_word_reports:
                            cached_origin = _pyplot_current_annealing_origin_artifact(record)
                            if cached_origin is not None:
                                origin_cache[record_cache_key] = cached_origin
                        else:
                            try:
                                cached_origin = _plot_measurement_origin(
                                    record.dataframe,
                                    record.path,
                                    origin_dir,
                                    log,
                                )
                            except RuntimeError as exc:
                                if origin_disabled_reason is None:
                                    origin_disabled_reason = str(exc) or exc.__class__.__name__
                                    log.warning("Origin plotting disabled: %s", origin_disabled_reason)
                                origin_enabled = False
                                cached_origin = None
                            except Exception:
                                log.exception("Failed to generate Origin plot for %s", record.path)
                                cached_origin = None
                            else:
                                if cached_origin is not None:
                                    origin_cache[record_cache_key] = cached_origin
                                    origin_artifacts.setdefault(cached_origin.descriptor, cached_origin)
                    if cached_origin is not None:
                        other_descriptors.append(cached_origin.descriptor)
                if other_descriptors:
                    row["Figure — other annealing (Origin)"] = _collapse_asset_references(other_descriptors)
        _backfill_current_densities(row)
        row_index = len(rows)
        rows.append(row)
        if row_highlights:
            for column in row_highlights:
                if column in output_columns:
                    ocr_highlights.setdefault(column, set()).add(row_index)
        stats.rows_built += 1
    if rows:
        df_full = pd.DataFrame(rows, columns=output_columns)
    else:
        df_full = pd.DataFrame(columns=output_columns)
    column_filter = getattr(config, "column_filter", None)
    column_order = getattr(config, "column_order", None)
    sort_spec = getattr(config, "sort_spec", None)
    df_word, _ = _apply_output_preferences(
        df_full,
        column_filter=None,
        column_order=None,
        sort_spec=sort_spec,
        highlight_map=None,
    )
    df_out = df_full
    df_out, ocr_highlights = _apply_output_preferences(
        df_out,
        column_filter=column_filter,
        column_order=column_order,
        sort_spec=sort_spec,
        highlight_map=ocr_highlights if highlight_ocr else None,
    )
    if not highlight_ocr:
        ocr_highlights = {}
    export_columns = list(df_out.columns)
    if microscope_image_columns:
        microscope_image_columns = tuple(
            column for column in microscope_image_columns if column in df_out.columns
        )
    exports: Dict[str, Path] = {}
    if skip_exports:
        requested_formats: Tuple[str, ...] = ()
    else:
        requested_formats = (
            tuple(dict.fromkeys(config.export_formats))
            if config.export_formats
            else ()
        )
    word_reports: List[Path] = []
    behaviours = {
        (key.lower() if isinstance(key, str) else ""): str(value).lower()
        for key, value in (config.export_behaviour or {}).items()
    }
    for fmt in requested_formats:
        fmt_lower = fmt.lower()
        if fmt_lower == "csv":
            csv_path = output_dir / f"{output_name}.csv"
            behaviour = behaviours.get("csv", "replace")
            if behaviour == "update":
                if csv_path.exists():
                    _update_existing_csv_with_strain(
                        csv_path,
                        strain_records,
                        export_columns,
                        log,
                    )
                    exports["csv"] = csv_path
                    continue
                log.warning("CSV export %s does not exist; creating a new file instead", csv_path)
                behaviour = "replace"
            to_write = df_out
            if behaviour == "append" and csv_path.exists():
                try:
                    existing = pd.read_csv(csv_path)
                except Exception:
                    log.exception("Failed to read existing CSV at %s; overwriting", csv_path)
                else:
                    to_write = pd.concat([existing, df_out], ignore_index=True)
                    to_write = to_write.drop_duplicates()
            to_write.to_csv(csv_path, index=False)
            exports["csv"] = csv_path
        elif fmt_lower == "excel":
            excel_path = output_dir / f"{output_name}.xlsx"
            behaviour = behaviours.get("excel", "replace")
            if behaviour == "update":
                if excel_path.exists():
                    try:
                        _update_existing_excel_with_strain(excel_path, strain_records, log)
                    except RuntimeError:
                        log.exception("Unable to update Excel file at %s", excel_path)
                        raise
                    exports["excel"] = excel_path
                    continue
                log.warning("Excel export %s does not exist; creating a new file instead", excel_path)
                behaviour = "replace"
            to_write = df_out
            if behaviour == "append" and excel_path.exists():
                try:
                    existing = pd.read_excel(excel_path)
                except Exception:
                    log.exception("Failed to read existing Excel at %s; overwriting", excel_path)
                else:
                    to_write = pd.concat([existing, df_out], ignore_index=True)
                    to_write = to_write.drop_duplicates()
            excel_frame = to_write.copy()
            columns_to_blank = [
                column for column in (FIGURE_COLUMNS + ORIGIN_FIGURE_COLUMNS) if column in excel_frame.columns
            ]
            if microscope_image_columns:
                columns_to_blank.extend(
                    column for column in microscope_image_columns if column in excel_frame.columns
                )
            for column in columns_to_blank:
                excel_frame[column] = None
            try:
                import xlsxwriter  # noqa: F401

                with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
                    excel_frame.to_excel(writer, index=False, sheet_name="Sheet1")
                    try:
                        _embed_assets_with_xlsxwriter(
                            writer,
                            to_write,
                            plot_name_to_path,
                            plot_dir,
                            origin_artifacts,
                            figure_size,
                            log,
                            microscope_crops=microscope_crop_map if include_crops else None,
                            microscope_columns=microscope_image_columns,
                            highlight_map=ocr_highlights if highlight_ocr else None,
                        )
                    except Exception:
                        log.exception("Failed to embed figures into %s", excel_path)
                try:
                    _adjust_drawing_ext_dimensions(excel_path, figure_size, log)
                except Exception:
                    log.exception("Failed to adjust drawing metadata for %s", excel_path)
            except ImportError:
                excel_frame.to_excel(excel_path, index=False)
                try:
                    _embed_plots_in_excel_openpyxl(
                        excel_path,
                        to_write,
                        plot_name_to_path,
                        plot_dir,
                        log,
                        figure_size,
                        microscope_columns=microscope_image_columns,
                        microscope_crops=microscope_crop_map if include_crops else None,
                        highlight_map=ocr_highlights if highlight_ocr else None,
                    )
                    try:
                        _adjust_drawing_ext_dimensions(excel_path, figure_size, log)
                    except Exception:
                        log.exception("Failed to adjust drawing metadata for %s", excel_path)
                except Exception:
                    log.exception("Failed to embed figure images into %s", excel_path)
            exports["excel"] = excel_path
        elif fmt_lower in {"word", "docx", "word_reports"}:
            report_dir = output_dir / f"{output_name}_{WORD_REPORT_DIR_NAME}"
            word_reports = _export_word_reports(
                df_word,
                report_dir,
                origin_artifacts,
                log,
                microscope_crops=microscope_crop_map if include_crops else None,
            )
            if word_reports:
                exports["word"] = report_dir
        else:
            log.warning("Unsupported export format '%s'; skipping", fmt)
    log.info(
        "Measurements parsed: %s | Skipped: %s | Rows built: %s | Missing draw info: %s | Missing piece info: %s | Missing 1000 mA: %s | Râ‰ˆV/I failures: %s",
        stats.parsed,
        stats.skipped,
        stats.rows_built,
        stats.missing_draw,
        stats.missing_piece,
        stats.missing_high_measurement,
        stats.resistance_checks_failed,
    )

    if wants_matplotlib:
        try:
            shutil.rmtree(plot_dir, ignore_errors=True)
        except Exception:
            pass

    return BuildResult(
        dataframe=df_out,
        exports=exports,
        plot_paths=plot_records,
        origin_artifacts=origin_artifacts,
        stats=stats,
        microscope_crops=microscope_crop_map if include_crops else {},
        ocr_highlights=ocr_highlights if highlight_ocr else {},
        word_reports=word_reports,
    )


__all__ = [
    "BuilderConfig",
    "BuildResult",
    "BuildStats",
    "OriginArtifact",
    "WordGraphSectionEvaluation",
    "WordOleInsertion",
    "WordOleEmbeddingResult",
    "WordPictureInsertion",
    "export_origin_graph_artifact",
    "export_pyplot_origin_artifacts_for_paths",
    "FabricationIndex",
    "StrainRecord",
    "VsmHysteresisRecord",
    "VsmTemperatureScanRecord",
    "DmaIsoStressRecord",
    "MiniDmaRecord",
    "ShapeMemoryStressStrainRecord",
    "FmrRecord",
    "VSM_HYSTERESIS_COLUMN",
    "VSM_TEMPERATURE_SCAN_COLUMN",
    "DMA_ISOSTRESS_COLUMN",
    "MINI_DMA_COLUMN",
    "MINI_DMA_STRAIN_COLUMN",
    "MINI_DMA_TRANSITION_COLUMN",
    "MINI_DMA_BREAK_COLUMN",
    "SHAPE_MEMORY_STRESS_STRAIN_COLUMN",
    "VSM_HYSTERESIS_ORIGIN_COLUMN",
    "VSM_TEMPERATURE_SCAN_ORIGIN_COLUMN",
    "DMA_ISOSTRESS_ORIGIN_COLUMN",
    "MINI_DMA_ORIGIN_COLUMN",
    "SHAPE_MEMORY_STRESS_STRAIN_ORIGIN_COLUMN",
    "SHAPE_MEMORY_DISPLACEMENT_COLUMN",
    "SHAPE_MEMORY_LOAD_COLUMN",
    "SHAPE_MEMORY_STRAIN_COLUMN",
    "SHAPE_MEMORY_STRESS_COLUMN",
    "SHAPE_MEMORY_FRACTURE_LOAD_COLUMN",
    "SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN",
    "SHAPE_MEMORY_FRACTURE_STRESS_COLUMN",
    "FMR_COLUMN",
    "FMR_ORIGIN_COLUMN",
    "RVT_FILE_COLUMN",
    "RVT_GRAPH_COLUMN",
    "RVT_ORIGIN_COLUMN",
    "RVT_RESIDUAL_ORIGIN_COLUMN",
    "RVT_POINT_COUNT_COLUMN",
    "RVT_TEMPERATURE_RANGE_COLUMN",
    "RVT_RESISTANCE_RANGE_COLUMN",
    "build_database",
    "build_fabrication_index",
    "export_word_reports",
    "word_report_section_manifest_for_row",
    "_compute_ea_from_composition",
    "LOGGER_NAME",
    "DEFAULT_OUTPUT_NAME",
    "WORD_REPORT_DIR_NAME",
]
