"""PyQt6 user interface for the microwire database builder."""

from __future__ import annotations

import base64
import faulthandler
import html
import io
import json
import logging
import math
import os
import pickle
import re
import sys
import time
import traceback
import warnings
from datetime import datetime
from dataclasses import dataclass, field
from functools import partial
from pathlib import Path
from tempfile import TemporaryDirectory, gettempdir
from typing import Any, Callable, ClassVar, Dict, Iterable, List, Mapping, Optional, Sequence, Set, Tuple, cast

try:
    from .ocr import ORIGINAL_HOME as OCR_ORIGINAL_HOME
except Exception:
    OCR_ORIGINAL_HOME = None

import pandas as pd

from PyQt6 import QtCore, QtGui, QtWidgets
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg, NavigationToolbar2QT
import matplotlib.pyplot as plt

try:
    import openpyxl
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None  # type: ignore[assignment]

from plotting.plugins.current_annealing.core import plot_one as plot_annealing_curve
from plotting.pyplot.window import _DockSwitcherWidget
from plotting.shared.utils import (
    ensure_app_theme,
    install_standard_menu,
    format_annealing_title,
    developer_options,
)
from plotting.shared.logfiles import append_text_with_rotation, open_rotating_text_log

from .storage import MiniDatabaseData, MiniDatabaseStore

from .core import (
    LOGGER_NAME,
    DEFAULT_FIGSIZE,
    DEFAULT_OUTPUT_NAME,
    BuildResult,
    BuilderConfig,
    BuildCancelledError,
    MicroscopeMeasurements,
    MicroscopeDetection,
    MicroscopeOCRResult,
    MicroscopeCacheEntry,
    MICROSCOPE_IMAGE_COLUMNS,
    MeasurementRecord,
    VideoMetricsSummary,
    FabricationIndex,
    StrainRecord,
    VsmHysteresisRecord,
    VsmTemperatureScanRecord,
    DmaIsoStressRecord,
    ShapeMemoryStressStrainRecord,
    FmrRecord,
    OUTPUT_COLUMNS,
    FIGURE_COLUMNS,
    ORIGIN_FIGURE_COLUMNS,
    VSM_HYSTERESIS_COLUMN,
    VSM_TEMPERATURE_SCAN_COLUMN,
    DMA_ISOSTRESS_COLUMN,
    SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
    SHAPE_MEMORY_DISPLACEMENT_COLUMN,
    SHAPE_MEMORY_LOAD_COLUMN,
    SHAPE_MEMORY_STRAIN_COLUMN,
    SHAPE_MEMORY_STRESS_COLUMN,
    SHAPE_MEMORY_FRACTURE_LOAD_COLUMN,
    SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN,
    SHAPE_MEMORY_FRACTURE_STRESS_COLUMN,
    FMR_COLUMN,
    build_database,
    build_fabrication_index,
    _normalise_output_name,
    _metadata_from_path,
    _microscope_key,
    _microscope_category,
    _draw_key,
    _load_annealing,
    _resistance_sanity_check,
    _group_microscope_measurements,
    _collect_video_metrics,
    _microwire_label,
    _microwire_parts_from_label,
    _microwire_tuple_from_label,
    _microwire_key_to_str,
    _microwire_key_from_string,
    _split_microwire_key,
    MICROWIRE_SORT_RE,
    _parse_strain_float,
    _plot_measurement_matplotlib,
    _select_high_measurement,
    _select_low_measurement,
    _value_for_output,
    _compose_notes,
    _format_dimension_display,
    _clean_str,
    _compute_ea_from_composition,
    _estimate_transition_temp_c,
    CORE_TEMPERATURE_COLUMN,
    GLASS_TEMPERATURE_COLUMN,
    ESTIMATED_TRANSITION_COLUMN,
    GLASS_PULL_COLUMN,
    VIDEO_END_LENGTH_COLUMN,
    VIDEO_MW_LENGTH_COLUMN,
)

try:
    from plotting.plugins.vsm_temperature_scan.core import VSMTemperatureScanProcessor
except Exception:  # pragma: no cover - optional dependency
    VSMTemperatureScanProcessor = None  # type: ignore[assignment]

try:
    from plotting.plugins.dma_iso_stress.parser import parse_dma_txt
except Exception:  # pragma: no cover - optional dependency
    parse_dma_txt = None  # type: ignore[assignment]

try:
    from plotting.plugins.shape_memory_stress_strain.core import (
        load_manual_stress_strain_file,
        make_dual_axis_overlay_figure,
    )
except Exception:  # pragma: no cover - optional dependency
    load_manual_stress_strain_file = None  # type: ignore[assignment]
    make_dual_axis_overlay_figure = None  # type: ignore[assignment]

try:
    from plotting.plugins.vsm_hysteresis.vsm_hysteresis_loops import (
        _read_vsm_file,
        _parse_temperature,
        _parse_angle,
    )
except Exception:  # pragma: no cover - optional dependency
    _read_vsm_file = None  # type: ignore[assignment]
    _parse_temperature = None  # type: ignore[assignment]
    _parse_angle = None  # type: ignore[assignment]

try:
    from plotting.plugins.fmr.core import parse_fmr_csv, select_fmr_axes
except Exception:  # pragma: no cover - optional dependency
    parse_fmr_csv = None  # type: ignore[assignment]
    select_fmr_axes = None  # type: ignore[assignment]


MICROSCOPE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp")
VIDEO_EXTENSIONS = (".mkv", ".mp4", ".avi", ".mov")

MICROSCOPE_D_COLUMN = "d (\u00b5m)"
MICROSCOPE_CAP_D_COLUMN = "D (\u00b5m)"
MICROSCOPE_TABLE_COLUMNS = [
    "Composition",
    "Microwire",
    MICROSCOPE_D_COLUMN,
    MICROSCOPE_CAP_D_COLUMN,
    "d/D",
    MICROSCOPE_IMAGE_COLUMNS[0],
    MICROSCOPE_IMAGE_COLUMNS[1],
    "_key",
    "_core_image",
    "_glass_image",
    "_images",
]

ANNEALING_GRAPH_WIDTH = 480
ANNEALING_GRAPH_HEIGHT = 230
ANNEALING_TITLE_FONT_SIZE = 8
ANNEALING_AXIS_FONT_SIZE = 6
ANNEALING_TICK_FONT_SIZE = 6
ANNEALING_OTHER_GRAPH_COLUMN = "Graph — other mA"
GRAPH_PREVIEW_WIDTH = 720
GRAPH_PREVIEW_HEIGHT = 420
MAX_PLOT_POINTS = 2000
ANNEALING_AS_COLUMN = "As1 (mA)"
ANNEALING_AF1_COLUMN = "Af1 (mA)"
ANNEALING_MS_COLUMN = "Ms1 (mA)"
ANNEALING_MF1_COLUMN = "Mf1 (mA)"
ANNEALING_AS2_COLUMN = "As2 (mA)"
ANNEALING_AF2_COLUMN = "Af2 (mA)"
ANNEALING_MS2_COLUMN = "Ms2 (mA)"
ANNEALING_MF2_COLUMN = "Mf2 (mA)"
CURRENT_DENSITY_AS_DENSITY_COLUMN = "As current density (A/mm^2)"
CURRENT_DENSITY_MS_DENSITY_COLUMN = "Ms current density (A/mm^2)"
CURRENT_DENSITY_AS_DELTA_COLUMN = "As2-As1 (mA)"
CURRENT_DENSITY_AF_DELTA_COLUMN = "Af2-Af1 (mA)"
CURRENT_DENSITY_MS_DELTA_COLUMN = "Ms2-Ms1 (mA)"
CURRENT_DENSITY_MF_DELTA_COLUMN = "Mf2-Mf1 (mA)"
CURRENT_DENSITY_MF_AF1_DELTA_COLUMN = "Mf1-Af1 (mA)"
CURRENT_DENSITY_MF_AF2_DELTA_COLUMN = "Mf2-Af2 (mA)"
PHASE_POINT_LABELS = ("As1", "Af1", "Ms1", "Mf1", "As2", "Af2", "Ms2", "Mf2")
PHASE_POINT_COLUMN_MAP = {
    "As1": ANNEALING_AS_COLUMN,
    "Af1": ANNEALING_AF1_COLUMN,
    "Ms1": ANNEALING_MS_COLUMN,
    "Mf1": ANNEALING_MF1_COLUMN,
    "As2": ANNEALING_AS2_COLUMN,
    "Af2": ANNEALING_AF2_COLUMN,
    "Ms2": ANNEALING_MS2_COLUMN,
    "Mf2": ANNEALING_MF2_COLUMN,
}
TRANSITION_TEMP_AS_COLUMN = "As (°C)"
TRANSITION_TEMP_AF_COLUMN = "Af (°C)"
TRANSITION_TEMP_MS_COLUMN = "Ms (°C)"
TRANSITION_TEMP_MF_COLUMN = "Mf (°C)"
TRANSITION_TEMP_LABELS = ("As", "Af", "Ms", "Mf")
TRANSITION_TEMP_COLUMN_MAP = {
    "As": TRANSITION_TEMP_AS_COLUMN,
    "Af": TRANSITION_TEMP_AF_COLUMN,
    "Ms": TRANSITION_TEMP_MS_COLUMN,
    "Mf": TRANSITION_TEMP_MF_COLUMN,
}
CURRENT_DENSITY_COLUMNS = [
    "Composition",
    "Microwire",
    MICROSCOPE_D_COLUMN,
    ANNEALING_AS_COLUMN,
    ANNEALING_AF1_COLUMN,
    ANNEALING_MS_COLUMN,
    ANNEALING_MF1_COLUMN,
    ANNEALING_AS2_COLUMN,
    ANNEALING_AF2_COLUMN,
    ANNEALING_MS2_COLUMN,
    ANNEALING_MF2_COLUMN,
    CURRENT_DENSITY_AS_DENSITY_COLUMN,
    CURRENT_DENSITY_MS_DENSITY_COLUMN,
    CURRENT_DENSITY_AS_DELTA_COLUMN,
    CURRENT_DENSITY_AF_DELTA_COLUMN,
    CURRENT_DENSITY_MS_DELTA_COLUMN,
    CURRENT_DENSITY_MF_DELTA_COLUMN,
    CURRENT_DENSITY_MF_AF1_DELTA_COLUMN,
    CURRENT_DENSITY_MF_AF2_DELTA_COLUMN,
    "Setpoints (mA)",
    "Sources",
    "Notes",
]
TRANSITION_TEMP_COLUMNS = [
    "Composition",
    "Microwire",
    TRANSITION_TEMP_AS_COLUMN,
    TRANSITION_TEMP_AF_COLUMN,
    TRANSITION_TEMP_MS_COLUMN,
    TRANSITION_TEMP_MF_COLUMN,
]

_SHAPE_MEMORY_COLUMN_ALIASES = {
    "Shape memory displacement (mm)": SHAPE_MEMORY_DISPLACEMENT_COLUMN,
    "Shape memory load (g)": SHAPE_MEMORY_LOAD_COLUMN,
    "Shape memory strain (%)": SHAPE_MEMORY_STRAIN_COLUMN,
    "Shape memory stress (MPa)": SHAPE_MEMORY_STRESS_COLUMN,
    "Shape memory fracture load (g)": SHAPE_MEMORY_FRACTURE_LOAD_COLUMN,
    "Shape memory fracture strain (%)": SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN,
    "Shape memory fracture stress (MPa)": SHAPE_MEMORY_FRACTURE_STRESS_COLUMN,
}


_STAGE_LABELS = {
    "prep": "Preparing support files",
    "analysis": "Analysing microscope/video data",
    "build": "Building database rows",
    "final": "Finalising exports",
}

_TEST_PATH_TOKENS = ("pytest-of-", "pyplot-tests", ".pytest_cache")


def _builder_settings() -> QtCore.QSettings:
    custom_file = os.environ.get("MICROWIRE_BUILDER_SETTINGS_FILE", "").strip()
    if custom_file:
        return QtCore.QSettings(custom_file, QtCore.QSettings.Format.IniFormat)
    platform = os.environ.get("QT_QPA_PLATFORM", "").strip().lower()
    if os.environ.get("PYTEST_CURRENT_TEST") or platform == "offscreen":
        settings_file = (
            Path(gettempdir()) / f"microwire_data_builder_offscreen_{os.getpid()}.ini"
        )
        return QtCore.QSettings(str(settings_file), QtCore.QSettings.Format.IniFormat)
    return QtCore.QSettings("MicrowireLab", "MicrowireDataBuilder")


def _looks_like_test_path(value: object) -> bool:
    if not isinstance(value, (str, Path)):
        return False
    text = str(value).strip().replace("\\", "/").lower()
    if not text:
        return False
    return any(token in text for token in _TEST_PATH_TOKENS)


def _sanitise_existing_directory(value: object) -> Optional[str]:
    if not isinstance(value, str) or not value.strip() or _looks_like_test_path(value):
        return None
    candidate = Path(value).expanduser()
    try:
        if candidate.exists() and candidate.is_dir():
            return str(candidate)
    except Exception:
        return None
    return None


def _sanitise_existing_file(value: object) -> Optional[str]:
    if not isinstance(value, str) or not value.strip() or _looks_like_test_path(value):
        return None
    candidate = Path(value).expanduser()
    try:
        if candidate.exists() and candidate.is_file():
            return str(candidate)
    except Exception:
        return None
    return None


def _dialog_start_directory(preferred: Path | str | None = None) -> Path:
    candidates: List[Path] = []
    if preferred:
        try:
            candidates.append(Path(preferred).expanduser())
        except Exception:
            pass
    if OCR_ORIGINAL_HOME:
        candidates.append(Path(OCR_ORIGINAL_HOME))
    env_home = os.environ.get("MICROWIRE_ORIGINAL_HOME")
    if env_home:
        try:
            candidates.append(Path(env_home))
        except Exception:
            pass
    user_profile = os.environ.get("USERPROFILE")
    if user_profile:
        try:
            candidates.append(Path(user_profile))
        except Exception:
            pass
    try:
        candidates.append(Path.home())
    except Exception:
        pass
    candidates.append(Path.cwd())
    for candidate in candidates:
        try:
            resolved = candidate.expanduser()
        except Exception:
            resolved = candidate
        try:
            resolved = resolved.resolve()
        except Exception:
            pass
        if resolved.exists() and resolved.is_dir():
            return resolved
    return Path.cwd()


def _json_safe(value: Any) -> Any:
    if value is None:
        return None
    if value is pd.NA:  # type: ignore[attr-defined]
        return None
    if getattr(pd, "isna", None):
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
    nat = getattr(pd, "NaT", None)
    if nat is not None and value is nat:
        return None
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return float(value)
    if isinstance(value, (int, str, bool)):
        return value
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _json_safe(val) for key, val in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    item_method = getattr(value, "item", None)
    if callable(item_method):
        try:
            return _json_safe(item_method())
        except Exception:
            pass
    return str(value)


def _normalise_import_header(text: str) -> str:
    cleaned = str(text or "").strip().lower()
    if not cleaned:
        return ""
    return re.sub(r"[^a-z0-9]+", "", cleaned)


def _map_import_header(raw_header: object) -> Optional[str]:
    raw = str(raw_header or "").strip()
    if not raw:
        return None
    lowered = raw.lower().strip()
    if "e/a" in lowered or lowered == "ea":
        return "e/a"
    if re.search(r"d\s*/\s*d", raw, re.IGNORECASE):
        return "d/D"
    cleaned = _normalise_import_header(raw)
    direct_map = {
        "composition": "Composition",
        "microwire": "Microwire",
        "strain": "Strain",
        "stressmpa": "Stress (MPa)",
        "mlength": "M length",
        "alength": "A length",
        "m": "m",
    }
    if cleaned in direct_map:
        return direct_map[cleaned]
    if cleaned in {"d", "dum", "dµm", "diameter"}:
        return "d (µm)"
    if raw.strip().upper().startswith("D"):
        return "D (µm)"
    if cleaned == "dd":
        return "d/D"
    if cleaned in {"as1", "af1", "ms1", "mf1", "as2", "af2", "ms2", "mf2"}:
        prefix = cleaned[:2].capitalize()
        suffix = cleaned[2:]
        return f"{prefix}{suffix} (mA)"
    return raw


def _normalise_import_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned in {"-", "#VALUE!", "#N/A"}:
            return None
        return cleaned
    return value


@dataclass
class WorkerInputs:
    """Configuration passed to :class:`BuildWorker`."""

    annealing_files: list[Path]
    manual_microscope_files: list[Path]
    data_roots: list[Path]
    output_dir: Path
    output_name: str
    export_formats: tuple[str, ...]
    plot_backends: tuple[str, ...]
    export_behaviour: dict[str, str]
    matplotlib_figsize: tuple[float, float]
    include_microscope_crops: bool = True
    highlight_ocr_values: bool = True
    analyse_videos: bool = True
    strain_files: list[Path] = field(default_factory=list)
    phase_points: Dict[str, Dict[str, float]] = field(default_factory=dict)


FIGURE_WIDTH_DEFAULT_MM = round(DEFAULT_FIGSIZE[0] * 25.4, 1)
FIGURE_HEIGHT_DEFAULT_MM = round(DEFAULT_FIGSIZE[1] * 25.4, 1)


def is_microscope_candidate(path: Path) -> bool:
    """Return ``True`` when ``path`` looks like a microscope overlay image."""

    if path.suffix.lower() not in MICROSCOPE_EXTENSIONS:
        return False
    stem = path.stem.lower()
    return "core" in stem or "glass" in stem


def collect_support_files(
    annealing_files: Sequence[Path],
    data_roots: Sequence[Path],
    progress_callback: Optional[Callable[[], None]] = None,
    include_videos: bool = True,
) -> tuple[list[Path], list[Path], list[Path]]:
    """Locate fabrication spreadsheets, microscope images and videos.

    When ``progress_callback`` is supplied it is invoked after each annealing
    file has been analysed so callers can surface responsive progress updates
    while the filesystem scan is running. Set ``include_videos`` to ``False`` to
    skip the slower video discovery path entirely.
    """

    if not annealing_files:
        return [], [], []

    unique_annealing = list(dict.fromkeys(Path(p) for p in annealing_files))
    records: list[tuple[Path, object]] = []
    for path in unique_annealing:
        try:
            meta = _metadata_from_path(path)
        except Exception:
            continue
        composition = getattr(meta, "composition_token", None)
        draw = getattr(meta, "draw_x", None)
        if composition and draw is not None:
            records.append((path, meta))
    if not records:
        return [], [], []

    def _is_relative(path: Path, root: Path) -> bool:
        try:
            path.relative_to(root)
        except ValueError:
            return False
        return True

    candidate_roots: list[Path] = []
    for candidate in dict.fromkeys(data_roots):
        root_path = Path(candidate).expanduser()
        if root_path.is_dir():
            candidate_roots.append(root_path)
    primary_root: Path | None = None
    for root_path in candidate_roots:
        if any(_is_relative(path, root_path) for path, _ in records):
            primary_root = root_path
            break
    if primary_root is None:
        if candidate_roots:
            primary_root = candidate_roots[0]
        else:
            common = Path(records[0][0]).resolve().parent
            primary_root = common

    def _resolve_subdir(root: Path | None, names: tuple[str, ...]) -> Path | None:
        if root is None:
            return None
        for name in names:
            candidate = root / name
            if candidate.is_dir():
                return candidate
        return None

    fabrication_root = (
        _resolve_subdir(primary_root, ("microwire data", "Microwire data"))
        or primary_root
    )
    microscope_root = _resolve_subdir(primary_root, ("microscope", "Microscope"))
    video_root = _resolve_subdir(
        primary_root,
        ("microwire data", "Microwire data", "videos", "Videos", "Microscope", "microscope"),
    )

    fabrication: list[Path] = []
    auto_micro: list[Path] = []
    videos: list[Path] = []
    seen_fabrication: set[Path] = set()
    seen_micro: set[Path] = set()
    seen_video: set[Path] = set()

    def _append_unique(container: list[Path], seen: set[Path], candidate: Path | None) -> None:
        if candidate is None:
            return
        resolved = candidate.expanduser()
        try:
            exists = resolved.exists()
        except OSError:
            exists = False
        if not exists:
            return
        try:
            key = resolved.resolve()
        except OSError:
            key = resolved
        if key in seen:
            return
        seen.add(key)
        container.append(resolved)

    def _composition_dirs(base: Path | None, composition: str) -> list[Path]:
        dirs: list[Path] = []
        if base is None or not base.is_dir():
            return dirs
        exact = base / composition
        if exact.is_dir():
            dirs.append(exact)
        try:
            for child in base.iterdir():
                if child.is_dir() and child.name.lower().startswith(composition.lower()):
                    dirs.append(child)
        except OSError:
            pass
        if not dirs:
            dirs.append(base)
        return dirs

    def _piece_dirs(comp_dir: Path, draw: int) -> list[Path]:
        dirs: list[Path] = []
        draw_token = str(draw)
        try:
            for child in comp_dir.iterdir():
                if child.is_dir() and draw_token in child.name:
                    dirs.append(child)
        except OSError:
            pass
        return dirs

    def _iter_fragment_files(
        root: Path,
        fragment: str,
        extensions: tuple[str, ...],
        *,
        max_depth: int = 4,
        limit: int | None = None,
    ) -> list[Path]:
        if root is None or not root.exists():
            return []
        fragment_lower = fragment.lower()
        stack: list[tuple[Path, int]] = [(root, 0)]
        visited: set[Path] = set()
        matches: list[Path] = []
        keywords = ("microscope", "video", "videos")
        while stack:
            current_root, depth = stack.pop()
            if current_root in visited:
                continue
            visited.add(current_root)
            if depth > max_depth:
                continue
            try:
                entries = list(current_root.iterdir())
            except OSError:
                continue
            for entry in entries:
                try:
                    if entry.is_dir():
                        if entry.name.lower() in keywords:
                            next_depth = depth
                        else:
                            next_depth = depth + 1
                        stack.append((entry, next_depth))
                    elif entry.is_file():
                        if (
                            entry.suffix.lower() in extensions
                            and fragment_lower in entry.name.lower()
                        ):
                            matches.append(entry)
                            if limit is not None and len(matches) >= limit:
                                return matches
                except OSError:
                    continue
        return matches

    for _, meta in records:
        try:
            composition = getattr(meta, "composition_token", None)
            draw = getattr(meta, "draw_x", None)
            piece = getattr(meta, "piece_y", None)
            if composition is None or draw is None:
                continue
            composition_dirs = _composition_dirs(fabrication_root, composition)
            for comp_dir in composition_dirs:
                _append_unique(fabrication, seen_fabrication, comp_dir / f"{composition}.xlsx")
                try:
                    for candidate in comp_dir.glob("*.xlsx"):
                        if candidate.name.lower() == f"{composition.lower()}.xlsx":
                            continue
                        if candidate.stem.startswith(f"{draw}_"):
                            _append_unique(fabrication, seen_fabrication, candidate)
                except OSError:
                    pass
                if piece is not None:
                    for piece_dir in _piece_dirs(comp_dir, draw):
                        try:
                            for candidate in piece_dir.glob("*.xlsx"):
                                _append_unique(fabrication, seen_fabrication, candidate)
                        except OSError:
                            continue
            if piece is None:
                continue
            search_dirs: list[Path] = []
            if microscope_root is not None:
                search_dirs.extend(_composition_dirs(microscope_root, composition))
            if not search_dirs and microscope_root is not None:
                search_dirs.append(microscope_root)
            for comp_dir in composition_dirs:
                for piece_dir in _piece_dirs(comp_dir, draw):
                    if piece_dir not in search_dirs:
                        search_dirs.append(piece_dir)
            fragment_tokens = {
                f"{draw}_{piece}",
                f"{draw}-{piece}",
                f"{draw}{piece}",
                f"{draw} {piece}",
                f"{draw}.{piece}",
            }
            for search_dir in search_dirs:
                if not search_dir.is_dir():
                    continue
                try:
                    candidates: list[Path] = []
                    for fragment in fragment_tokens:
                        candidates.extend(
                            _iter_fragment_files(
                                search_dir,
                                fragment,
                                MICROSCOPE_EXTENSIONS,
                                limit=50,
                            )
                        )
                except Exception:
                    continue
                if not candidates:
                    try:
                        candidates = _iter_fragment_files(
                            search_dir, "", MICROSCOPE_EXTENSIONS, limit=100
                        )
                    except Exception:
                        continue
                for candidate in dict.fromkeys(candidates):
                    if not is_microscope_candidate(candidate):
                        continue
                    key = _microscope_key(candidate)
                    if key is not None and key[:3] == (composition, draw, piece):
                        _append_unique(auto_micro, seen_micro, candidate)
            if include_videos:
                video_dirs: list[Path] = []
                if video_root is not None:
                    video_dirs.extend(_composition_dirs(video_root, composition))
                for comp_dir in composition_dirs:
                    video_dirs.extend(_piece_dirs(comp_dir, draw))
                for search_dir in video_dirs:
                    if not search_dir.is_dir():
                        continue
                    try:
                        video_candidates: list[Path] = []
                        for fragment in fragment_tokens:
                            video_candidates.extend(
                                _iter_fragment_files(
                                    search_dir, fragment, VIDEO_EXTENSIONS, limit=60
                                )
                            )
                    except Exception:
                        continue
                    if not video_candidates:
                        try:
                            video_candidates = _iter_fragment_files(
                                search_dir, "", VIDEO_EXTENSIONS, limit=120
                            )
                        except Exception:
                            continue
                    for candidate in dict.fromkeys(video_candidates):
                        key = _microscope_key(candidate)
                        if key is not None and key[:3] == (composition, draw, piece):
                            _append_unique(videos, seen_video, candidate)
                            continue
                        draw_key = _draw_key(candidate)
                        if draw_key == (composition, draw):
                            _append_unique(videos, seen_video, candidate)
        finally:
            if progress_callback is not None:
                try:
                    progress_callback()
                except BuildCancelledError:
                    raise
                except Exception:
                    pass

    fabrication = list(dict.fromkeys(fabrication))
    auto_micro = list(dict.fromkeys(auto_micro))
    videos = list(dict.fromkeys(videos))
    return fabrication, auto_micro, videos


def _format_duration(seconds: float) -> str:
    if not math.isfinite(seconds):
        return "--"
    total_seconds = max(int(round(seconds)), 0)
    minutes, sec = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours:d}h {minutes:02d}m {sec:02d}s"
    if minutes:
        return f"{minutes:d}m {sec:02d}s"
    return f"{sec:d}s"


class QtLogHandler(logging.Handler):
    """Logging handler that forwards records to a Qt slot."""

    def __init__(self, emit: Callable[[int, str], None]) -> None:
        super().__init__()
        self._emit = emit

    def emit(self, record: logging.LogRecord) -> None:  # pragma: no cover - thin wrapper
        message = self.format(record)
        try:
            self._emit(record.levelno, message)
        except TypeError:
            # Fallback for legacy callbacks that only accept a message.
            self._emit(message)  # type: ignore[misc]


class _ProgressTracker:
    """Translate multi-phase progress updates into a single counter."""

    _ORDER = ("prep", "analysis", "build", "final")

    def __init__(
        self, emit: Callable[[str, int, int, int, int], None]
    ) -> None:
        self._emit_fn = emit
        self._totals = {stage: 0 for stage in self._ORDER}
        self._progress = {stage: 0 for stage in self._ORDER}

    def configure(
        self, prep_units: int, analysis_units: int, build_units: int, final_units: int = 1
    ) -> None:
        self._totals = {
            "prep": max(prep_units, 0),
            "analysis": max(analysis_units, 0),
            "build": max(build_units, 0),
            "final": max(final_units, 0),
        }
        self._progress = {key: 0 for key in self._totals}
        self._emit_all()

    def _total_units(self) -> int:
        total = sum(self._totals.values())
        return total if total > 0 else 1

    def _completed_units(self) -> int:
        total_units = self._total_units()
        completed = 0
        for key, total in self._totals.items():
            progress = self._progress.get(key, 0)
            if total <= 0:
                continue
            completed += min(max(progress, 0), total)
        return min(max(completed, 0), total_units)

    def _emit(self, stage: str) -> None:
        self._emit_fn(
            stage,
            self._progress.get(stage, 0),
            self._totals.get(stage, 0),
            self._completed_units(),
            self._total_units(),
        )

    def _emit_all(self) -> None:
        for stage in self._ORDER:
            self._emit(stage)

    def _set_total(self, key: str, units: int) -> None:
        units = max(units, 0)
        if units == self._totals.get(key, 0):
            return
        self._totals[key] = units
        if self._progress.get(key, 0) > units:
            self._progress[key] = units
        self._emit(key)

    def set_analysis_units(self, units: int) -> None:
        self._set_total("analysis", units)

    def set_final_units(self, units: int) -> None:
        self._set_total("final", units)

    def advance_prepare(self) -> None:
        total = self._totals.get("prep", 0)
        if total <= 0:
            return
        current = self._progress.get("prep", 0)
        if current < total:
            self._progress["prep"] = current + 1
            self._emit("prep")

    def finish_prepare(self) -> None:
        total = self._totals.get("prep", 0)
        if total <= 0:
            return
        if self._progress.get("prep", 0) < total:
            self._progress["prep"] = total
            self._emit("prep")

    def analysis_progress(self, current: int, total: int) -> None:
        total = max(total, 0)
        if total:
            self._set_total("analysis", total)
        mapped_total = self._totals.get("analysis", 0)
        if mapped_total <= 0:
            return
        clamped = min(max(current, 0), mapped_total)
        if clamped > self._progress.get("analysis", 0):
            self._progress["analysis"] = clamped
            self._emit("analysis")

    def finish_analysis(self) -> None:
        total = self._totals.get("analysis", 0)
        if total <= 0:
            return
        if self._progress.get("analysis", 0) < total:
            self._progress["analysis"] = total
            self._emit("analysis")

    def build_progress(self, current: int, total: int) -> None:
        total = max(total, 0)
        if total:
            self._set_total("build", total)
        mapped_total = self._totals.get("build", 0)
        if mapped_total <= 0:
            return
        clamped = min(max(current, 0), mapped_total)
        if clamped > self._progress.get("build", 0):
            self._progress["build"] = clamped
            self._emit("build")

    def finish_build(self) -> None:
        total = self._totals.get("build", 0)
        if total <= 0:
            return
        if self._progress.get("build", 0) < total:
            self._progress["build"] = total
            self._emit("build")

    def advance_final(self, units: int = 1) -> None:
        total = self._totals.get("final", 0)
        if total <= 0:
            return
        current = self._progress.get("final", 0)
        if current < total:
            self._progress["final"] = min(total, current + max(units, 1))
            self._emit("final")

    def finish_final(self) -> None:
        total = self._totals.get("final", 0)
        if total <= 0:
            return
        if self._progress.get("final", 0) < total:
            self._progress["final"] = total
            self._emit("final")

    def finalize(self) -> None:
        self.finish_final()


class BuildWorker(QtCore.QObject):
    """Background worker that runs the database builder."""

    progress = QtCore.pyqtSignal(str, int, int, int, int)
    finished = QtCore.pyqtSignal(object)
    error = QtCore.pyqtSignal(str)
    cancelled = QtCore.pyqtSignal()

    def __init__(self, inputs: WorkerInputs, logger: logging.Logger) -> None:
        super().__init__()
        self.inputs = inputs
        self.logger = logger
        self._tracker = _ProgressTracker(self.progress.emit)
        self._cancelled = False

    @QtCore.pyqtSlot()
    def request_cancel(self) -> None:
        self._cancelled = True

    @QtCore.pyqtSlot()
    def run(self) -> None:  # pragma: no cover - exercised via integration test
        try:
            inputs = self.inputs
            annealing_files = list(dict.fromkeys(inputs.annealing_files))
            manual_microscope = list(dict.fromkeys(inputs.manual_microscope_files))
            prep_units = len(annealing_files)
            build_units = len(annealing_files)
            self._tracker.configure(prep_units, 0, build_units, final_units=1)
            self.logger.info("Preparing support files...")
            
            def _check_cancelled() -> None:
                if self._cancelled:
                    raise BuildCancelledError()

            def _prepare_bridge() -> None:
                _check_cancelled()
                self._tracker.advance_prepare()

            _check_cancelled()
            fabrication_files, auto_microscope, video_files = collect_support_files(
                annealing_files,
                inputs.data_roots,
                progress_callback=_prepare_bridge,
                include_videos=inputs.analyse_videos,
            )
            _check_cancelled()
            self._tracker.finish_prepare()
            _check_cancelled()
            microscope_files = list(dict.fromkeys(manual_microscope + auto_microscope))
            if not inputs.analyse_videos:
                video_files = []
            analysis_units = len(microscope_files) + len(video_files)
            self._tracker.set_analysis_units(analysis_units)
            config = BuilderConfig(
                fabrication_files=fabrication_files,
                annealing_files=annealing_files,
                output_dir=inputs.output_dir,
                microscope_files=microscope_files,
                video_files=video_files,
                strain_files=inputs.strain_files,
                make_plots=bool(inputs.plot_backends),
                export_formats=inputs.export_formats,
                output_name=inputs.output_name,
                plot_backends=inputs.plot_backends,
                export_behaviour=inputs.export_behaviour,
                matplotlib_figsize=inputs.matplotlib_figsize,
                include_microscope_crops=inputs.include_microscope_crops,
                highlight_ocr_values=inputs.highlight_ocr_values,
            )
            config.output_dir.mkdir(parents=True, exist_ok=True)
            self.logger.info(
                "Starting build with %s annealing file(s)", len(config.annealing_files)
            )
            _check_cancelled()
            if fabrication_files:
                self.logger.info(
                    "Using %s fabrication spreadsheet(s)", len(fabrication_files)
                )
            if microscope_files:
                self.logger.info(
                    "Using %s microscope image(s)", len(microscope_files)
                )
            if video_files:
                self.logger.info("Using %s video file(s)", len(video_files))
            elif inputs.analyse_videos:
                self.logger.info("No fabrication videos were found for analysis.")
            else:
                self.logger.info(
                    "Fabrication video analysis disabled; skipping video metrics."
                )

            def _progress_bridge(current: int, total: int) -> None:
                _check_cancelled()
                self._tracker.build_progress(current, total)

            def _analysis_bridge(current: int, total: int) -> None:
                _check_cancelled()
                self._tracker.set_analysis_units(total)
                self._tracker.analysis_progress(current, total)

            _check_cancelled()
            result = build_database(
                config,
                logger=self.logger,
                progress_callback=_progress_bridge,
                analysis_progress_callback=_analysis_bridge,
                analysis_total=analysis_units,
                root_for_relpaths=Path.cwd(),
                phase_points=inputs.phase_points or None,
            )
            final_steps = max(len(result.exports), 1)
            if config.make_plots and (
                "matplotlib" in config.plot_backends or not config.plot_backends
            ):
                final_steps += 1
            if config.make_plots and "origin" in config.plot_backends:
                final_steps += 1
            final_steps += 1  # summary log
            self._tracker.set_final_units(final_steps)
            self._tracker.finish_analysis()
            self._tracker.finish_build()
            if result.exports:
                for fmt, path in result.exports.items():
                    self.logger.info("%s written to %s", fmt.upper(), path)
                    self._tracker.advance_final()
            else:
                self.logger.info("No export files were generated.")
                self._tracker.advance_final()
            if config.make_plots:
                if "matplotlib" in config.plot_backends or not config.plot_backends:
                    self.logger.info("Generated %s Matplotlib plot(s)", len(result.plot_paths))
                    self._tracker.advance_final()
                if "origin" in config.plot_backends:
                    self.logger.info(
                        "Origin plots created: %s",
                        len(result.origin_artifacts),
                    )
                    self._tracker.advance_final()
            stats = result.stats
            self.logger.info(
                "Summary: parsed=%s skipped=%s rows=%s missing_draw=%s missing_piece=%s missing_1000mA=%s missing_low_mA=%s R~=V/I failures=%s",
                stats.parsed,
                stats.skipped,
                stats.rows_built,
                stats.missing_draw,
                stats.missing_piece,
                stats.missing_high_measurement,
                stats.missing_low_measurement,
                stats.resistance_checks_failed,
            )
            self._tracker.advance_final()
            self._tracker.finalize()
            self.finished.emit(result)
        except BuildCancelledError:
            self.logger.info("Build cancelled by user.")
            self.cancelled.emit()
        except Exception as exc:  # pragma: no cover - safety net
            self.logger.exception("Build failed")
            message = str(exc) if str(exc) else exc.__class__.__name__
            self.error.emit(message)


class PreviewWorker(QtCore.QObject):
    """Background worker that builds the Assemble preview dataframe."""

    finished = QtCore.pyqtSignal(object)
    failed = QtCore.pyqtSignal(str)

    def __init__(
        self,
        config: BuilderConfig,
        build_kwargs: Dict[str, Any],
        logger: logging.Logger,
    ) -> None:
        super().__init__()
        self._config = config
        self._build_kwargs = build_kwargs
        self._logger = logger

    @QtCore.pyqtSlot()
    def run(self) -> None:  # pragma: no cover - UI-thread integration
        try:
            result = build_database(self._config, logger=self._logger, **self._build_kwargs)
        except Exception as exc:
            try:
                self._logger.exception("Preview build failed")
            except Exception:
                pass
            self.failed.emit(str(exc))
            return
        try:
            self.finished.emit(result.dataframe)
        except Exception as exc:
            self.failed.emit(str(exc))


class CombineWorker(QtCore.QObject):
    """Background worker that builds the Assemble exports."""

    finished = QtCore.pyqtSignal(object)
    failed = QtCore.pyqtSignal(str)

    def __init__(
        self,
        config: BuilderConfig,
        build_kwargs: Dict[str, Any],
        logger: logging.Logger,
    ) -> None:
        super().__init__()
        self._config = config
        self._build_kwargs = build_kwargs
        self._logger = logger

    @QtCore.pyqtSlot()
    def run(self) -> None:  # pragma: no cover - UI-thread integration
        try:
            result = build_database(self._config, logger=self._logger, **self._build_kwargs)
        except Exception as exc:
            self.failed.emit(str(exc))
            return
        self.finished.emit(result)


class LegacyBuilderWindow(QtWidgets.QMainWindow):
    """Main window that orchestrates the microwire database build."""

    PROJECT_EXTENSION = ".pydpj"
    PROJECT_VERSION = 1
    PROJECT_KIND = "MicrowireDataBuilder"
    log_message = QtCore.pyqtSignal(int, str)

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Microwire Data Builder")
        self.resize(960, 720)

        self.microscope_paths: list[Path] = []
        self.annealing_paths: list[Path] = []
        self.data_roots: list[Path] = []
        self._thread: QtCore.QThread | None = None
        self._worker: BuildWorker | None = None
        self._running = False
        self._progress_start_time: float | None = None
        self._last_progress_value: int = 0
        self._last_progress_timestamp: float | None = None
        self._seconds_per_unit_ema: float | None = None
        self._last_eta_seconds: float | None = None
        self._overall_progress_current: int = 0
        self._overall_progress_total: int = 0
        self._progress_counts_text: str = ""
        self._progress_stage_text: str = ""
        self._last_logged_stage: str | None = None
        self._stage_order: tuple[str, ...] = ("prep", "analysis", "build", "final")
        self._stage_timing_history: dict[str, dict[str, float | int]] = {}
        self._init_stage_tracking()

        self._eta_timer = QtCore.QTimer(self)
        self._eta_timer.setInterval(1000)
        self._eta_timer.timeout.connect(self._on_eta_timer)

        cwd = Path.cwd()
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.exists() and downloads_dir.is_dir():
            self._default_output_dir = downloads_dir
        else:
            self._default_output_dir = cwd / "builder_output"
        self._last_microscope_dir = str(cwd)
        self._last_anneal_dir = str(cwd)
        self._last_root_dir = str(cwd)
        self._last_output_dir = str(self._default_output_dir)
        self._last_strain_dir = str(cwd)
        self.settings = _builder_settings()
        self._project_path: Optional[Path] = None
        self._save_project_action: QtGui.QAction | None = None
        self._save_project_as_action: QtGui.QAction | None = None

        self.log_message.connect(self._append_log)

        self._build_ui()
        self._configure_logging()
        self._load_settings()
        menu_bar = install_standard_menu(
            self,
            help_topic="builder_database",
            console=self.log_group,
            open_file=self._add_microscope_files,
            open_folder=self._add_data_root,
        )
        self._setup_project_actions(menu_bar)
        self._update_project_actions()

    # ------------------------------------------------------------------ setup
    def _init_stage_tracking(self) -> None:
        self._stage_progress: dict[str, int] = {stage: 0 for stage in self._stage_order}
        self._stage_totals: dict[str, int] = {stage: 0 for stage in self._stage_order}
        self._stage_runtime: dict[str, dict[str, object]] = {
            stage: {"start": None, "completed": False} for stage in self._stage_order
        }
        self._stage_metrics: dict[str, dict[str, object]] = {
            stage: {"last_value": 0, "last_timestamp": None, "ema": None}
            for stage in self._stage_order
        }
        self._displayed_percent = 0

    def _reset_progress_tracking(self) -> None:
        now = time.monotonic()
        self._progress_start_time = now
        self._last_progress_value = 0
        self._last_progress_timestamp = now
        self._seconds_per_unit_ema = None
        self._last_eta_seconds = None
        self._overall_progress_current = 0
        self._overall_progress_total = 0
        self._progress_counts_text = ""
        self._progress_stage_text = ""
        self._last_logged_stage = None
        self._init_stage_tracking()

    def _clear_progress_tracking(self) -> None:
        self._progress_start_time = None
        self._last_progress_value = 0
        self._last_progress_timestamp = None
        self._seconds_per_unit_ema = None
        self._last_eta_seconds = None
        self._overall_progress_current = 0
        self._overall_progress_total = 0
        self._progress_counts_text = ""
        self._progress_stage_text = ""
        self._last_logged_stage = None
        self._init_stage_tracking()

    def _reset_stage_metrics(self, stage: str) -> None:
        if stage not in self._stage_order:
            return
        metrics = self._stage_metrics.setdefault(
            stage, {"last_value": 0, "last_timestamp": None, "ema": None}
        )
        metrics["last_value"] = 0
        metrics["last_timestamp"] = None
        metrics["ema"] = None
        runtime = self._stage_runtime.setdefault(
            stage, {"start": None, "completed": False}
        )
        runtime["start"] = None
        runtime["completed"] = False

    def _persist_stage_history(self) -> None:
        if not hasattr(self, "settings"):
            return
        try:
            payload = json.dumps(self._stage_timing_history)
        except TypeError:
            return
        self.settings.setValue("stage_timing", payload)
        self.settings.sync()

    def _record_stage_history(self, stage: str, elapsed: float, units: int) -> None:
        if units <= 0 or elapsed <= 0:
            return
        record = self._stage_timing_history.setdefault(
            stage,
            {"seconds": 0.0, "units": 0, "ema": None, "samples": 0},
        )
        seconds = max(float(record.get("seconds", 0.0)), 0.0) + float(elapsed)
        units_done = max(int(record.get("units", 0)), 0) + int(units)
        record["seconds"] = seconds
        record["units"] = units_done
        average = float(elapsed) / float(units)
        ema = record.get("ema")
        if isinstance(ema, (int, float)) and ema > 0:
            record["ema"] = max((float(ema) * 0.7) + (average * 0.3), 0.0)
        else:
            record["ema"] = max(average, 0.0)
        record["samples"] = int(record.get("samples", 0)) + 1
        self._persist_stage_history()

    def _global_average_rate(self) -> float | None:
        ema_sum = 0.0
        ema_weight = 0.0
        total_seconds = 0.0
        total_units = 0
        for record in self._stage_timing_history.values():
            if not isinstance(record, dict):
                continue
            ema = record.get("ema")
            samples = max(int(record.get("samples", 0)), 0)
            if isinstance(ema, (int, float)) and ema > 0:
                ema_sum += float(ema) * max(samples, 1)
                ema_weight += max(samples, 1)
            seconds = float(record.get("seconds", 0.0))
            units = int(record.get("units", 0))
            total_seconds += max(seconds, 0.0)
            total_units += max(units, 0)
        if ema_weight > 0:
            return ema_sum / ema_weight
        if total_seconds > 0 and total_units > 0:
            return total_seconds / total_units
        return None

    def _estimate_remaining_seconds(
        self,
        now: float,
        active_stage: str | None,
        overall_current: int,
        overall_total: int,
    ) -> float | None:
        remaining = 0.0
        have_estimate = False
        global_rate = self._global_average_rate()

        def _push(rate_list: list[tuple[float, float]], value: object, weight: float) -> None:
            if isinstance(value, (int, float)) and value > 0 and weight > 0:
                rate_list.append((float(value), weight))

        for stage in self._stage_order:
            total_units = max(int(self._stage_totals.get(stage, 0)), 0)
            progress_units = min(max(int(self._stage_progress.get(stage, 0)), 0), total_units)
            remaining_units = total_units - progress_units
            if remaining_units <= 0:
                continue

            rates: list[tuple[float, float]] = []
            metrics = self._stage_metrics.get(stage, {})
            stage_ema = metrics.get("ema") if isinstance(metrics, dict) else None
            _push(rates, stage_ema, 0.3)

            history = self._stage_timing_history.get(stage)
            history_rate: float | None = None
            history_weight = 0.0
            if isinstance(history, dict):
                hist_ema = history.get("ema")
                hist_seconds = float(history.get("seconds", 0.0))
                hist_units = int(history.get("units", 0))
                samples = max(int(history.get("samples", 0)), 0)
                if isinstance(hist_ema, (int, float)) and hist_ema > 0:
                    history_rate = float(hist_ema)
                elif hist_seconds > 0 and hist_units > 0:
                    history_rate = hist_seconds / hist_units
                if history_rate is not None:
                    history_weight = 0.45 + min(samples, 10) * 0.035
                    _push(rates, history_rate, history_weight)

            runtime = self._stage_runtime.get(stage, {})
            if (
                stage == active_stage
                and isinstance(runtime, dict)
                and runtime.get("start") is not None
            ):
                start_time = float(runtime["start"])
                elapsed_stage = max(now - start_time, 0.0)
                completed_units = max(int(self._stage_progress.get(stage, 0)), 0)
                if elapsed_stage > 0 and completed_units > 0:
                    runtime_rate = elapsed_stage / completed_units
                    _push(rates, runtime_rate, 0.5)

            if stage == active_stage and self._seconds_per_unit_ema:
                _push(rates, self._seconds_per_unit_ema, 0.3)

            if not rates and global_rate is not None:
                _push(rates, global_rate, 0.5)

            if not rates:
                continue

            have_estimate = True
            total_weight = sum(weight for _, weight in rates)
            if total_weight <= 0:
                continue
            weighted = sum(rate * weight for rate, weight in rates) / total_weight
            sorted_values = sorted(rate for rate, _ in rates)
            if not sorted_values:
                continue
            mid = len(sorted_values) // 2
            if len(sorted_values) % 2:
                median = sorted_values[mid]
            else:
                median = (sorted_values[mid - 1] + sorted_values[mid]) / 2
            rate = max(weighted, median)
            if history_rate is not None:
                rate = max(rate, history_rate * 0.85)
            remaining += remaining_units * rate

        if have_estimate:
            return remaining if remaining > 0 else 0.0
        start_time = self._progress_start_time
        if start_time is None or overall_current <= 0:
            return None
        elapsed = max(now - start_time, 0.0)
        remaining_units = max(overall_total - overall_current, 0)
        return (elapsed / overall_current) * remaining_units if overall_current else None

    @staticmethod
    def _format_eta(seconds: float) -> str:
        if seconds < 0:
            seconds = 0
        total_seconds = int(round(seconds))
        if total_seconds < 1:
            return "<1s remaining"
        hours, remainder = divmod(total_seconds, 3600)
        minutes, secs = divmod(remainder, 60)
        parts: list[str] = []
        if hours:
            parts.append(f"{hours}h")
        if minutes:
            parts.append(f"{minutes}m")
        if secs or not parts:
            parts.append(f"{secs}s")
        return " ".join(parts) + " remaining"

    def _smooth_eta(self, seconds: float) -> float:
        seconds = max(float(seconds), 0.0)
        previous = self._last_eta_seconds
        if previous is None:
            filtered = seconds
        else:
            if seconds > previous:
                alpha = 0.65
            else:
                alpha = 0.2
            filtered = previous + (seconds - previous) * alpha
        self._last_eta_seconds = max(filtered, 0.0)
        return self._last_eta_seconds

    def _on_eta_timer(self) -> None:
        self._update_eta_display()

    def _resolve_active_stage(self) -> str | None:
        for key in self._stage_order:
            total_units = max(int(self._stage_totals.get(key, 0)), 0)
            progress_units = min(max(int(self._stage_progress.get(key, 0)), 0), total_units)
            if total_units > 0 and progress_units < total_units:
                return key
        return None

    def _update_eta_display(self, now: Optional[float] = None) -> None:
        if not self._running:
            return
        if now is None:
            now = time.monotonic()
        current = max(int(self._overall_progress_current), 0)
        total = max(int(self._overall_progress_total), 1)
        remaining_units = max(total - current, 0)
        eta_text: Optional[str] = None
        if remaining_units <= 0:
            self._last_eta_seconds = None
            if current:
                eta_text = "Finishing..."
        else:
            active_stage = self._resolve_active_stage()
            remaining_seconds = self._estimate_remaining_seconds(now, active_stage, current, total)
            if remaining_seconds is not None and math.isfinite(remaining_seconds):
                smoothed = self._smooth_eta(remaining_seconds)
                eta_text = self._format_eta(smoothed)
            elif current > 0:
                eta_text = "Estimating..."
                self._last_eta_seconds = None

        parts: list[str] = []
        if self._progress_stage_text:
            parts.append(self._progress_stage_text)
        if self._progress_counts_text:
            parts.append(self._progress_counts_text)
        if eta_text:
            parts.append(eta_text)
        label_text = " • ".join(parts) if parts else "Working..."
        if label_text != self.progress_label.text():
            self.progress_label.setText(label_text)

    def _build_ui(self) -> None:
        central = QtWidgets.QWidget(self)
        self.setCentralWidget(central)
        main_layout = QtWidgets.QHBoxLayout(central)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(12)

        left_layout = QtWidgets.QVBoxLayout()
        left_layout.setSpacing(10)
        right_layout = QtWidgets.QVBoxLayout()
        right_layout.setSpacing(10)

        main_layout.addLayout(left_layout, 2)
        main_layout.addLayout(right_layout, 3)

        # Data roots
        self.root_group = QtWidgets.QGroupBox("Microwire data folder")
        root_layout = QtWidgets.QVBoxLayout(self.root_group)
        self.root_list = QtWidgets.QListWidget()
        self.root_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.root_group.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Preferred,
            QtWidgets.QSizePolicy.Policy.Fixed,
        )
        self.root_list.setMaximumHeight(110)
        root_layout.addWidget(self.root_list)
        root_buttons = QtWidgets.QHBoxLayout()
        root_add = QtWidgets.QPushButton("Add folder...")
        root_add.clicked.connect(self._add_data_root)
        root_buttons.addWidget(root_add)
        root_clear = QtWidgets.QPushButton("Clear")
        root_clear.clicked.connect(self._clear_data_roots)
        root_buttons.addWidget(root_clear)
        root_buttons.addStretch(1)
        root_layout.addLayout(root_buttons)
        right_layout.addWidget(self.root_group)

        # Annealing inputs
        self.anneal_group = QtWidgets.QGroupBox("Current-annealing files (.txt)")
        anneal_layout = QtWidgets.QVBoxLayout(self.anneal_group)
        self.anneal_list = QtWidgets.QListWidget()
        self.anneal_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        anneal_layout.addWidget(self.anneal_list)

        anneal_buttons = QtWidgets.QHBoxLayout()
        anneal_add_files = QtWidgets.QPushButton("Add files...")
        anneal_add_files.clicked.connect(self._add_anneal_files)
        anneal_buttons.addWidget(anneal_add_files)
        anneal_add_folder = QtWidgets.QPushButton("Add folder...")
        anneal_add_folder.clicked.connect(self._add_anneal_folder)
        anneal_buttons.addWidget(anneal_add_folder)
        anneal_clear = QtWidgets.QPushButton("Clear")
        anneal_clear.clicked.connect(self._clear_anneal)
        anneal_buttons.addWidget(anneal_clear)
        anneal_buttons.addStretch(1)
        anneal_layout.addLayout(anneal_buttons)

        self.anneal_recursive = QtWidgets.QCheckBox("Recursive scan")
        self.anneal_recursive.setChecked(True)
        anneal_layout.addWidget(self.anneal_recursive)
        right_layout.addWidget(self.anneal_group, 1)

        # Microscope inputs
        self.microscope_group = QtWidgets.QGroupBox("Microscope images")
        micro_layout = QtWidgets.QVBoxLayout(self.microscope_group)
        self.microscope_list = QtWidgets.QListWidget()
        self.microscope_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        micro_layout.addWidget(self.microscope_list)
        micro_buttons = QtWidgets.QHBoxLayout()
        micro_add_files = QtWidgets.QPushButton("Add files...")
        micro_add_files.clicked.connect(self._add_microscope_files)
        micro_buttons.addWidget(micro_add_files)
        micro_add_folder = QtWidgets.QPushButton("Add folder...")
        micro_add_folder.clicked.connect(self._add_microscope_folder)
        micro_buttons.addWidget(micro_add_folder)
        micro_clear = QtWidgets.QPushButton("Clear")
        micro_clear.clicked.connect(self._clear_microscope)
        micro_buttons.addWidget(micro_clear)
        micro_buttons.addStretch(1)
        micro_layout.addLayout(micro_buttons)
        self.microscope_recursive = QtWidgets.QCheckBox("Recursive scan")
        self.microscope_recursive.setChecked(True)
        micro_layout.addWidget(self.microscope_recursive)
        right_layout.addWidget(self.microscope_group, 1)

        # Strain worksheet
        self.strain_group = QtWidgets.QGroupBox("Strain worksheet")
        strain_layout = QtWidgets.QHBoxLayout(self.strain_group)
        strain_layout.setContentsMargins(8, 8, 8, 8)
        self.strain_edit = QtWidgets.QLineEdit()
        self.strain_edit.setPlaceholderText("Select an Excel worksheet with strain data…")
        strain_layout.addWidget(self.strain_edit, 1)
        self.strain_button = QtWidgets.QPushButton("Browse…")
        self.strain_button.clicked.connect(self._select_strain_file)
        strain_layout.addWidget(self.strain_button)
        self.clear_strain_button = QtWidgets.QPushButton("Clear")
        self.clear_strain_button.clicked.connect(self._clear_strain_file)
        strain_layout.addWidget(self.clear_strain_button)
        right_layout.addWidget(self.strain_group)

        # Options
        self.options_group = QtWidgets.QGroupBox("Options")
        options_layout = QtWidgets.QVBoxLayout(self.options_group)
        self.plot_matplotlib_check = QtWidgets.QCheckBox("Matplotlib plots (PNG)")
        self.plot_matplotlib_check.setChecked(True)
        self.plot_matplotlib_check.stateChanged.connect(self._save_settings)
        options_layout.addWidget(self.plot_matplotlib_check)
        self.plot_origin_check = QtWidgets.QCheckBox("Origin plots")
        self.plot_origin_check.stateChanged.connect(self._save_settings)
        options_layout.addWidget(self.plot_origin_check)
        self.export_csv_check = QtWidgets.QCheckBox("Export CSV")
        self.export_csv_check.setChecked(True)
        self.export_csv_check.stateChanged.connect(self._save_settings)
        options_layout.addWidget(self.export_csv_check)
        self.export_excel_check = QtWidgets.QCheckBox("Export Excel")
        self.export_excel_check.stateChanged.connect(self._save_settings)
        options_layout.addWidget(self.export_excel_check)

        microscope_group = QtWidgets.QGroupBox("Microscope review")
        microscope_layout = QtWidgets.QVBoxLayout(microscope_group)
        microscope_layout.setContentsMargins(8, 8, 8, 8)
        self.include_crops_check = QtWidgets.QCheckBox("Attach microscope crops to Excel")
        self.include_crops_check.stateChanged.connect(self._save_settings)
        self.include_crops_check.setToolTip(
            "Add cropped microscope images in new columns next to the d and D values"
        )
        with QtCore.QSignalBlocker(self.include_crops_check):
            self.include_crops_check.setChecked(True)
        microscope_layout.addWidget(self.include_crops_check)
        self.highlight_ocr_check = QtWidgets.QCheckBox("Highlight OCR-sourced values")
        self.highlight_ocr_check.stateChanged.connect(self._save_settings)
        self.highlight_ocr_check.setToolTip(
            "Tint spreadsheet cells where the value was filled from OCR instead of fabrication spreadsheets"
        )
        with QtCore.QSignalBlocker(self.highlight_ocr_check):
            self.highlight_ocr_check.setChecked(True)
        microscope_layout.addWidget(self.highlight_ocr_check)
        options_layout.addWidget(microscope_group)

        self.video_metrics_check = QtWidgets.QCheckBox("Extract fabrication metrics from videos")
        self.video_metrics_check.setChecked(True)
        self.video_metrics_check.stateChanged.connect(self._save_settings)
        self.video_metrics_check.setToolTip(
            "Sample fabrication videos to OCR winding speed, glass feed, temperature, and underpressure values"
        )
        options_layout.addWidget(self.video_metrics_check)

        figure_size_form = QtWidgets.QFormLayout()
        figure_size_form.setHorizontalSpacing(8)
        figure_size_form.setVerticalSpacing(4)
        self.figure_width_spin = QtWidgets.QDoubleSpinBox()
        self.figure_width_spin.setRange(20.0, 400.0)
        self.figure_width_spin.setDecimals(1)
        self.figure_width_spin.setSingleStep(5.0)
        self.figure_width_spin.setValue(FIGURE_WIDTH_DEFAULT_MM)
        self.figure_width_spin.valueChanged.connect(self._save_settings)
        figure_size_form.addRow("Figure width (mm)", self.figure_width_spin)
        self.figure_height_spin = QtWidgets.QDoubleSpinBox()
        self.figure_height_spin.setRange(20.0, 250.0)
        self.figure_height_spin.setDecimals(1)
        self.figure_height_spin.setSingleStep(5.0)
        self.figure_height_spin.setValue(FIGURE_HEIGHT_DEFAULT_MM)
        self.figure_height_spin.valueChanged.connect(self._save_settings)
        figure_size_form.addRow("Figure height (mm)", self.figure_height_spin)
        options_layout.addLayout(figure_size_form)
        left_layout.addWidget(self.options_group)

        # Output directory
        self.output_group = QtWidgets.QGroupBox("Output")
        output_layout = QtWidgets.QGridLayout(self.output_group)
        output_layout.setHorizontalSpacing(8)
        output_layout.setVerticalSpacing(6)
        output_label = QtWidgets.QLabel("Directory:")
        output_layout.addWidget(output_label, 0, 0)
        self.output_edit = QtWidgets.QLineEdit(str(self._default_output_dir))
        self.output_edit.editingFinished.connect(self._save_settings)
        output_layout.addWidget(self.output_edit, 0, 1)
        self.output_button = QtWidgets.QPushButton("Browse...")
        self.output_button.clicked.connect(self._select_output_dir)
        output_layout.addWidget(self.output_button, 0, 2)
        name_label = QtWidgets.QLabel("File name:")
        output_layout.addWidget(name_label, 1, 0)
        self.output_name_edit = QtWidgets.QLineEdit(DEFAULT_OUTPUT_NAME)
        self.output_name_edit.editingFinished.connect(self._save_settings)
        output_layout.addWidget(self.output_name_edit, 1, 1)
        output_layout.setColumnStretch(1, 1)
        left_layout.addWidget(self.output_group)

        # Progress row
        progress_row = QtWidgets.QHBoxLayout()
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        progress_row.addWidget(self.progress_bar, stretch=1)
        self.progress_label = QtWidgets.QLabel("Idle")
        progress_row.addWidget(self.progress_label)
        left_layout.addLayout(progress_row)

        # Log view
        self.log_group = QtWidgets.QGroupBox("Log")
        log_layout = QtWidgets.QVBoxLayout(self.log_group)
        self.log_view = QtWidgets.QPlainTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setMaximumBlockCount(2000)
        log_layout.addWidget(self.log_view)
        left_layout.addWidget(self.log_group, stretch=1)

        # Run button
        run_row = QtWidgets.QHBoxLayout()
        run_row.addStretch(1)
        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.cancel_button.setEnabled(False)
        self.cancel_button.clicked.connect(self.cancel_build)
        run_row.addWidget(self.cancel_button)
        self.run_button = QtWidgets.QPushButton("Run")
        self.run_button.clicked.connect(self.start_build)
        run_row.addWidget(self.run_button)
        left_layout.addLayout(run_row)

    def _configure_logging(self) -> None:
        self.logger = logging.getLogger(LOGGER_NAME)
        self.logger.setLevel(logging.INFO)
        self.logger.propagate = False
        self._log_handler = QtLogHandler(self.log_message.emit)
        self._log_handler.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
        if self._log_handler not in self.logger.handlers:
            self.logger.addHandler(self._log_handler)

    def _load_settings(self) -> None:
        settings_sanitized = False

        def _decode_paths(value: object) -> list[Path]:
            if isinstance(value, list | tuple):
                return [Path(str(item)) for item in value]
            if isinstance(value, str):
                if not value:
                    return []
                try:
                    data = json.loads(value)
                    if isinstance(data, list):
                        return [Path(str(item)) for item in data]
                except json.JSONDecodeError:
                    items = [segment.strip() for segment in value.splitlines() if segment.strip()]
                    return [Path(item) for item in items]
            return []

        def _read_bool(key: str, default: bool) -> bool:
            value = self.settings.value(key, default)
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                lowered = value.strip().lower()
                if lowered in {"true", "1", "yes", "y"}:
                    return True
                if lowered in {"false", "0", "no", "n"}:
                    return False
            return default

        def _read_float(key: str, default: float) -> float:
            value = self.settings.value(key, default)
            try:
                return float(value)
            except (TypeError, ValueError):
                return default

        self.annealing_paths = _decode_paths(self.settings.value("annealing_paths", ""))
        original_anneal_len = len(self.annealing_paths)
        self.annealing_paths = [path for path in self.annealing_paths if not _looks_like_test_path(path)]
        settings_sanitized = settings_sanitized or len(self.annealing_paths) != original_anneal_len
        self._update_list_widget(self.anneal_list, self.annealing_paths)
        self.microscope_paths = _decode_paths(self.settings.value("microscope_paths", ""))
        original_microscope_len = len(self.microscope_paths)
        self.microscope_paths = [
            path for path in self.microscope_paths if not _looks_like_test_path(path)
        ]
        settings_sanitized = settings_sanitized or len(self.microscope_paths) != original_microscope_len
        self._update_list_widget(self.microscope_list, self.microscope_paths)
        self.data_roots = _decode_paths(self.settings.value("data_roots", ""))
        original_root_len = len(self.data_roots)
        self.data_roots = [path for path in self.data_roots if not _looks_like_test_path(path)]
        settings_sanitized = settings_sanitized or len(self.data_roots) != original_root_len
        self._update_list_widget(self.root_list, self.data_roots)

        with QtCore.QSignalBlocker(self.plot_matplotlib_check):
            self.plot_matplotlib_check.setChecked(_read_bool("plot_matplotlib", True))
        with QtCore.QSignalBlocker(self.plot_origin_check):
            self.plot_origin_check.setChecked(_read_bool("plot_origin", False))
        with QtCore.QSignalBlocker(self.export_csv_check):
            self.export_csv_check.setChecked(_read_bool("export_csv", True))
        with QtCore.QSignalBlocker(self.export_excel_check):
            self.export_excel_check.setChecked(_read_bool("export_excel", False))
        with QtCore.QSignalBlocker(self.include_crops_check):
            self.include_crops_check.setChecked(_read_bool("include_microscope_crops", True))
        with QtCore.QSignalBlocker(self.highlight_ocr_check):
            self.highlight_ocr_check.setChecked(_read_bool("highlight_ocr_values", True))
        with QtCore.QSignalBlocker(self.video_metrics_check):
            self.video_metrics_check.setChecked(_read_bool("analyse_videos", True))
        if hasattr(self, "microscope_recursive"):
            with QtCore.QSignalBlocker(self.microscope_recursive):
                self.microscope_recursive.setChecked(_read_bool("microscope_recursive", True))

        width_value = _read_float("figure_width", FIGURE_WIDTH_DEFAULT_MM)
        height_value = _read_float("figure_height", FIGURE_HEIGHT_DEFAULT_MM)
        if width_value <= 20.0 and height_value <= 20.0:
            width_value *= 25.4
            height_value *= 25.4
        with QtCore.QSignalBlocker(self.figure_width_spin):
            self.figure_width_spin.setValue(width_value)
        with QtCore.QSignalBlocker(self.figure_height_spin):
            self.figure_height_spin.setValue(height_value)

        output_dir_value = _sanitise_existing_directory(self.settings.value("output_dir", ""))
        if output_dir_value:
            self.output_edit.setText(output_dir_value)
            self._last_output_dir = output_dir_value
        elif self.settings.value("output_dir", ""):
            settings_sanitized = True

        output_name_value = self.settings.value("output_name", "")
        if isinstance(output_name_value, str) and output_name_value.strip():
            self.output_name_edit.setText(output_name_value)

        strain_path_value = _sanitise_existing_file(self.settings.value("strain_path", ""))
        if strain_path_value:
            self.strain_edit.setText(strain_path_value)
            try:
                self._last_strain_dir = str(Path(strain_path_value).expanduser().parent)
            except Exception:
                pass
        elif self.settings.value("strain_path", ""):
            settings_sanitized = True

        last_microscope = _sanitise_existing_directory(
            self.settings.value("last_microscope_dir", "")
        )
        if last_microscope:
            self._last_microscope_dir = last_microscope
        last_anneal = _sanitise_existing_directory(self.settings.value("last_anneal_dir", ""))
        if last_anneal:
            self._last_anneal_dir = last_anneal
        last_root = _sanitise_existing_directory(self.settings.value("last_root_dir", ""))
        if last_root:
            self._last_root_dir = last_root
        last_output = _sanitise_existing_directory(self.settings.value("last_output_dir", ""))
        if last_output:
            self._last_output_dir = last_output
        last_strain = _sanitise_existing_directory(self.settings.value("last_strain_dir", ""))
        if last_strain:
            self._last_strain_dir = last_strain

        timing_value = self.settings.value("stage_timing", "")
        if isinstance(timing_value, str) and timing_value.strip():
            try:
                stored = json.loads(timing_value)
            except json.JSONDecodeError:
                stored = {}
            if isinstance(stored, dict):
                for key, record in stored.items():
                    if not isinstance(record, dict):
                        continue
                    seconds = float(record.get("seconds", 0.0))
                    units = int(record.get("units", 0))
                    if seconds > 0 and units > 0:
                        ema_value = record.get("ema")
                        samples_value = record.get("samples", 0)
                        if isinstance(ema_value, (int, float)) and ema_value > 0:
                            ema = float(ema_value)
                        else:
                            ema = seconds / units
                        samples = int(samples_value) if isinstance(samples_value, (int, float)) else 0
                        if samples <= 0 and ema > 0:
                            samples = 1
                        self._stage_timing_history[key] = {
                            "seconds": seconds,
                            "units": units,
                            "ema": ema,
                            "samples": max(samples, 0),
                        }
        if settings_sanitized:
            self._save_settings()

    def _save_settings(self) -> None:
        if not hasattr(self, "settings"):
            return
        self.settings.setValue("annealing_paths", json.dumps([str(p) for p in self.annealing_paths]))
        self.settings.setValue("microscope_paths", json.dumps([str(p) for p in self.microscope_paths]))
        self.settings.setValue("data_roots", json.dumps([str(p) for p in self.data_roots]))
        self.settings.setValue("output_dir", self.output_edit.text())
        self.settings.setValue("plot_matplotlib", self.plot_matplotlib_check.isChecked())
        self.settings.setValue("plot_origin", self.plot_origin_check.isChecked())
        self.settings.setValue("export_csv", self.export_csv_check.isChecked())
        self.settings.setValue("export_excel", self.export_excel_check.isChecked())
        self.settings.setValue("include_microscope_crops", self.include_crops_check.isChecked())
        self.settings.setValue("highlight_ocr_values", self.highlight_ocr_check.isChecked())
        self.settings.setValue("analyse_videos", self.video_metrics_check.isChecked())
        self.settings.setValue("figure_width", self.figure_width_spin.value())
        self.settings.setValue("figure_height", self.figure_height_spin.value())
        self.settings.setValue("output_name", self.output_name_edit.text())
        self.settings.setValue("strain_path", self.strain_edit.text())
        self.settings.setValue("last_microscope_dir", self._last_microscope_dir)
        self.settings.setValue("last_anneal_dir", self._last_anneal_dir)
        self.settings.setValue("last_root_dir", self._last_root_dir)
        self.settings.setValue("last_output_dir", self._last_output_dir)
        self.settings.setValue("last_strain_dir", self._last_strain_dir)
        self.settings.setValue("microscope_recursive", self.microscope_recursive.isChecked())
        self.settings.setValue("stage_timing", json.dumps(self._stage_timing_history))
        self.settings.sync()

    # ------------------------------------------------------------------ helpers
    def _extend_paths(self, attr: str, paths: Iterable[Path]) -> None:
        current: list[Path] = getattr(self, attr)
        combined = list(dict.fromkeys(current + [Path(p) for p in paths]))
        setattr(self, attr, combined)

    def _update_list_widget(self, widget: QtWidgets.QListWidget, items: Iterable[Path]) -> None:
        widget.clear()
        for text in sorted({str(Path(p)) for p in items}):
            widget.addItem(text)

    def _is_microscope_candidate(self, path: Path) -> bool:
        return is_microscope_candidate(path)

    def _collect_support_files(
        self, annealing_files: list[Path]
    ) -> tuple[list[Path], list[Path], list[Path]]:
        include_videos = getattr(self, "video_metrics_check", None)
        analyse_videos = True
        if isinstance(include_videos, QtWidgets.QCheckBox):
            analyse_videos = include_videos.isChecked()
        return collect_support_files(
            annealing_files,
            self.data_roots,
            include_videos=analyse_videos,
        )

    def _add_anneal_files(self) -> None:
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            "Select current-annealing files",
            self._last_anneal_dir,
            "Text files (*.txt)",
        )
        if not files:
            return
        self._last_anneal_dir = str(Path(files[0]).parent)
        self._extend_paths("annealing_paths", (Path(f) for f in files))
        self._update_list_widget(self.anneal_list, self.annealing_paths)
        self._save_settings()

    def _add_anneal_folder(self) -> None:
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select folder with current-annealing files",
            self._last_anneal_dir,
        )
        if not folder:
            return
        root = Path(folder)
        iterator = root.rglob("*.txt") if self.anneal_recursive.isChecked() else root.glob("*.txt")
        files = [p for p in iterator if p.is_file()]
        if not files:
            QtWidgets.QMessageBox.information(self, "Microwire Data Builder", "No text files were found in that folder.")
            return
        self._last_anneal_dir = folder
        self._extend_paths("annealing_paths", files)
        self._update_list_widget(self.anneal_list, self.annealing_paths)
        self._save_settings()

    def _clear_anneal(self) -> None:
        self.annealing_paths = []
        self.anneal_list.clear()
        self._save_settings()

    def _add_microscope_files(self) -> None:
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            "Select microscope images",
            self._last_microscope_dir,
            "Image files (*.jpg *.jpeg *.png *.tif *.tiff *.bmp)",
        )
        if not files:
            return
        self._last_microscope_dir = str(Path(files[0]).parent)
        candidates = [Path(f) for f in files if self._is_microscope_candidate(Path(f))]
        if not candidates:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "No matching microscope images were selected.",
            )
            return
        self._extend_paths("microscope_paths", candidates)
        self._update_list_widget(self.microscope_list, self.microscope_paths)
        self._save_settings()

    def _add_microscope_folder(self) -> None:
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select folder with microscope images",
            self._last_microscope_dir,
        )
        if not folder:
            return
        root = Path(folder)
        iterator = root.rglob('*') if self.microscope_recursive.isChecked() else root.glob('*')
        files = [p for p in iterator if p.is_file() and self._is_microscope_candidate(p)]
        if not files:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "No microscope images were found in that folder.",
            )
            return
        self._last_microscope_dir = folder
        self._extend_paths("microscope_paths", files)
        self._update_list_widget(self.microscope_list, self.microscope_paths)
        self._save_settings()

    def _clear_microscope(self) -> None:
        self.microscope_paths = []
        self.microscope_list.clear()
        self._save_settings()

    def _select_strain_file(self) -> None:
        filename, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Select strain worksheet",
            self._last_strain_dir,
            "Excel files (*.xlsx *.xlsm *.xls)",
        )
        if not filename:
            return
        path = Path(filename)
        self._last_strain_dir = str(path.parent)
        self.strain_edit.setText(str(path))
        self._save_settings()

    def _clear_strain_file(self) -> None:
        self.strain_edit.clear()
        self._save_settings()

    def _add_data_root(self) -> None:
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select microwire data folder",
            self._last_root_dir,
        )
        if not folder:
            return
        root = Path(folder)
        if not root.exists():
            QtWidgets.QMessageBox.warning(self, "Microwire Data Builder", "The selected folder does not exist.")
            return
        self._last_root_dir = folder
        self.data_roots = [root]
        self._update_list_widget(self.root_list, self.data_roots)
        self._ingest_data_root(root, announce=True)
        self._save_settings()

    def _clear_data_roots(self) -> None:
        self.data_roots = []
        self.root_list.clear()
        self._save_settings()
    def _clear_data_roots(self) -> None:
        self.data_roots = []
        self.root_list.clear()
        self._save_settings()

    def _ingest_data_root(self, root: Path, announce: bool = False) -> None:
        if announce and hasattr(self, 'logger'):
            self.logger.info('Microwire data folder set to %s', root)
    def _select_output_dir(self) -> None:
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select output directory",
            self.output_edit.text() or self._last_output_dir,
        )
        if directory:
            self._last_output_dir = directory
            self.output_edit.setText(directory)
            self._save_settings()

    def _check_origin_available(self) -> tuple[bool, str]:
        try:
            import originpro  # type: ignore  # noqa: F401
        except Exception as exc:
            message = str(exc) if str(exc) else exc.__class__.__name__
            return False, message
        return True, ""

    def _prompt_overwrite(self, path: Path) -> str | None:
        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("File exists")
        msg.setIcon(QtWidgets.QMessageBox.Icon.Question)
        msg.setText(f"'{path.name}' already exists.")
        msg.setInformativeText("Choose how to continue:")
        replace_btn = msg.addButton("Replace", QtWidgets.QMessageBox.ButtonRole.DestructiveRole)
        update_btn = msg.addButton("Update", QtWidgets.QMessageBox.ButtonRole.AcceptRole)
        continue_btn = msg.addButton("Append", QtWidgets.QMessageBox.ButtonRole.ActionRole)
        cancel_btn = msg.addButton("Cancel", QtWidgets.QMessageBox.ButtonRole.RejectRole)
        msg.exec()
        clicked = msg.clickedButton()
        if clicked is cancel_btn:
            return None
        if clicked is update_btn:
            return "update"
        if clicked is continue_btn:
            return "append"
        return "replace"

    def _preferred_export_path(self, exports: dict[str, Path]) -> Path | None:
        for key in ("excel", "csv"):
            candidate = exports.get(key)
            if isinstance(candidate, Path):
                return candidate
        for candidate in exports.values():
            if isinstance(candidate, Path):
                return candidate
        return None

    def _open_path(self, path: Path) -> None:
        resolved = path.expanduser().resolve()
        if not resolved.exists():
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                f"{resolved} could not be found.",
            )
            return
        url = QtCore.QUrl.fromLocalFile(str(resolved))
        if not QtGui.QDesktopServices.openUrl(url):
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Unable to open the exported file.",
            )

    def _set_running(self, running: bool) -> None:
        self._running = running
        for widget in (
            self.root_group,
            self.anneal_group,
            self.microscope_group,
            self.options_group,
            self.output_group,
        ):
            widget.setEnabled(not running)
        self.run_button.setEnabled(not running)
        self.cancel_button.setEnabled(running)
        if running:
            if not self._eta_timer.isActive():
                self._eta_timer.start()
            self._reset_progress_tracking()
            self.progress_bar.setRange(0, 0)
            self.progress_bar.setValue(0)
            self._progress_stage_text = _STAGE_LABELS.get("prep", "Preparing...")
            self._update_eta_display()
        else:
            self._eta_timer.stop()
            self._clear_progress_tracking()
            if self.progress_bar.maximum() == 0 and self.progress_bar.minimum() == 0:
                self.progress_bar.setRange(0, 100)
                self.progress_bar.setValue(0)
            if self.progress_label.text() not in {"Complete", "Failed"}:
                self.progress_label.setText("Idle")

    # ------------------------------------------------------------------ build orchestration
    def cancel_build(self) -> None:
        if not self._running:
            return
        if self._worker is None:
            self._set_running(False)
            self.progress_bar.setValue(0)
            self.progress_label.setText("Cancelled")
            return
        self.logger.info("Cancellation requested; attempting to stop the build...")
        self.cancel_button.setEnabled(False)
        self.progress_label.setText("Cancelling...")
        QtCore.QMetaObject.invokeMethod(
            self._worker,
            "request_cancel",
            QtCore.Qt.ConnectionType.QueuedConnection,
        )

    def start_build(self) -> None:
        if self._running:
            return
        output_dir_text = self.output_edit.text().strip()
        if not output_dir_text:
            QtWidgets.QMessageBox.warning(self, "Microwire Data Builder", "Please choose an output directory.")
            return
        output_dir = Path(output_dir_text).expanduser()
        export_formats: list[str] = []
        if self.export_csv_check.isChecked():
            export_formats.append("csv")
        if self.export_excel_check.isChecked():
            export_formats.append("excel")
        if not export_formats:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Please select at least one export format.",
            )
            return
        output_name_text = self.output_name_edit.text().strip()
        if not output_name_text:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Please enter a base file name.",
            )
            return
        output_name = _normalise_output_name(output_name_text)
        if output_name != output_name_text:
            self.output_name_edit.setText(output_name)

        plot_backends: list[str] = []
        if self.plot_matplotlib_check.isChecked():
            plot_backends.append("matplotlib")
        if self.plot_origin_check.isChecked():
            origin_ok, origin_error = self._check_origin_available()
            if origin_ok:
                plot_backends.append("origin")
            else:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Microwire Data Builder",
                    "Origin plotting is unavailable.\n\n" + origin_error,
                )
                with QtCore.QSignalBlocker(self.plot_origin_check):
                    self.plot_origin_check.setChecked(False)

        behaviours: dict[str, str] = {}
        for fmt in export_formats:
            extension = "csv" if fmt.lower() == "csv" else "xlsx"
            target_path = output_dir / f"{output_name}.{extension}"
            if target_path.exists():
                action = self._prompt_overwrite(target_path)
                if action is None:
                    return
                behaviours[fmt.lower()] = action
            else:
                behaviours[fmt.lower()] = "replace"

        annealing_files = list(dict.fromkeys(self.annealing_paths))
        requires_measurements = any(action != "update" for action in behaviours.values())
        if requires_measurements and not annealing_files:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Please add at least one annealing file.",
            )
            return

        strain_files: list[Path] = []
        strain_text = self.strain_edit.text().strip()
        if strain_text:
            strain_files.append(Path(strain_text).expanduser())
        if not requires_measurements and not strain_files:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "No strain worksheet selected; nothing to update.",
            )
            return

        phase_points: Dict[str, Dict[str, float]] = {}
        annealing_section = self.sections.get("annealing")
        if isinstance(annealing_section, AnnealingSection):
            phase_points = dict(getattr(annealing_section, "_phase_points", {}))

        worker_inputs = WorkerInputs(
            annealing_files=annealing_files,
            manual_microscope_files=list(dict.fromkeys(self.microscope_paths)),
            data_roots=list(dict.fromkeys(self.data_roots)),
            output_dir=output_dir,
            output_name=output_name,
            export_formats=tuple(export_formats),
            plot_backends=tuple(plot_backends),
            export_behaviour=behaviours,
            matplotlib_figsize=(
                float(self.figure_width_spin.value()) / 25.4,
                float(self.figure_height_spin.value()) / 25.4,
            ),
            include_microscope_crops=self.include_crops_check.isChecked(),
            highlight_ocr_values=self.highlight_ocr_check.isChecked(),
            analyse_videos=self.video_metrics_check.isChecked(),
            strain_files=strain_files,
            phase_points=phase_points,
        )
        self._save_settings()
        self._set_running(True)
        self.log_view.clear()
        self.logger.info(
            "Queued build for %s annealing measurement(s)",
            len(worker_inputs.annealing_files),
        )
        self._start_worker(worker_inputs)

    def _start_worker(self, inputs: WorkerInputs) -> None:
        self._thread = QtCore.QThread(self)
        self._worker = BuildWorker(inputs, self.logger)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self._update_progress)
        self._worker.finished.connect(self._handle_finished)
        self._worker.error.connect(self._handle_failed)
        self._worker.cancelled.connect(self._handle_cancelled)
        self._worker.finished.connect(self._thread.quit)
        self._worker.error.connect(self._thread.quit)
        self._worker.cancelled.connect(self._thread.quit)
        self._thread.finished.connect(self._cleanup_thread)
        self._thread.start()

    def _update_progress(
        self,
        stage: str,
        stage_value: int,
        stage_total: int,
        current: int,
        total: int,
    ) -> None:
        total = max(total, 1)
        if self._progress_start_time is None:
            self._reset_progress_tracking()
        if current < self._last_progress_value:
            self._reset_progress_tracking()
        now = time.monotonic()
        if current > self._last_progress_value:
            delta_units = current - self._last_progress_value
            last_timestamp = self._last_progress_timestamp
            if last_timestamp is not None and delta_units > 0:
                delta_time = max(now - last_timestamp, 0.0)
                if delta_time > 0:
                    instantaneous = delta_time / delta_units
                    ema = self._seconds_per_unit_ema
                    if ema is None:
                        ema = instantaneous
                    elif instantaneous >= ema:
                        ema = (ema * 0.35) + (instantaneous * 0.65)
                    else:
                        ema = (ema * 0.9) + (instantaneous * 0.1)
                    self._seconds_per_unit_ema = max(ema, 0.0)
            self._last_progress_value = current
            self._last_progress_timestamp = now

        if stage not in self._stage_totals:
            self._stage_order += (stage,)
            self._stage_totals[stage] = max(stage_total, 0)
            clamped_value = max(stage_value, 0)
            if self._stage_totals[stage] > 0:
                clamped_value = min(clamped_value, self._stage_totals[stage])
            self._stage_progress[stage] = clamped_value
            self._stage_runtime[stage] = {"start": None, "completed": False}
            self._stage_metrics[stage] = {
                "last_value": 0,
                "last_timestamp": None,
                "ema": None,
            }
        else:
            self._stage_totals[stage] = max(stage_total, 0)
            previous_value = self._stage_progress.get(stage, 0)
            if stage_value < previous_value:
                self._reset_stage_metrics(stage)
            clamped_value = max(stage_value, 0)
            if self._stage_totals[stage] > 0:
                clamped_value = min(clamped_value, self._stage_totals[stage])
            self._stage_progress[stage] = clamped_value
        stage_value = self._stage_progress.get(stage, 0)
        stage_total = self._stage_totals.get(stage, 0)

        active_stage: str | None = None
        active_value: int = 0
        active_total: int = 0
        for key in self._stage_order:
            total_units = max(int(self._stage_totals.get(key, 0)), 0)
            progress_units = min(max(int(self._stage_progress.get(key, 0)), 0), total_units)
            if total_units > 0 and progress_units < total_units:
                active_stage = key
                active_total = total_units
                active_value = progress_units
                break

        metrics = self._stage_metrics.setdefault(
            stage, {"last_value": 0, "last_timestamp": None, "ema": None}
        )
        runtime = self._stage_runtime.setdefault(
            stage, {"start": None, "completed": False}
        )

        if stage_total > 0 and not runtime.get("completed", False):
            if stage == active_stage and runtime.get("start") is None:
                runtime["start"] = now
                metrics["last_timestamp"] = now

        last_stage_value = int(metrics.get("last_value", 0))
        if stage_value > last_stage_value:
            last_timestamp = metrics.get("last_timestamp")
            if isinstance(last_timestamp, (int, float)):
                delta_time = max(now - float(last_timestamp), 0.0)
                delta_units = stage_value - last_stage_value
                if delta_time > 0 and delta_units > 0:
                    instantaneous = delta_time / delta_units
                    ema = metrics.get("ema")
                    if ema is None:
                        ema = instantaneous
                    elif instantaneous >= ema:
                        ema = (ema * 0.35) + (instantaneous * 0.65)
                    else:
                        ema = (ema * 0.9) + (instantaneous * 0.1)
                    metrics["ema"] = max(ema, 0.0)
            else:
                if runtime.get("start") is None:
                    runtime["start"] = now
            metrics["last_timestamp"] = now
            metrics["last_value"] = stage_value
        elif stage_value == 0 and last_stage_value == 0 and stage == active_stage:
            if runtime.get("start") is None:
                runtime["start"] = now
            metrics["last_timestamp"] = now
        elif stage_value < last_stage_value:
            self._reset_stage_metrics(stage)
            metrics = self._stage_metrics[stage]
            runtime = self._stage_runtime[stage]
            if stage == active_stage and stage_total > 0:
                runtime["start"] = now
                metrics["last_timestamp"] = now
            metrics["last_value"] = stage_value

        if (
            stage_total > 0
            and stage_value >= stage_total
            and not runtime.get("completed", False)
        ):
            start = runtime.get("start")
            if isinstance(start, (int, float)):
                elapsed = max(now - float(start), 0.0)
                self._record_stage_history(stage, elapsed, stage_total)
            runtime["completed"] = True
            runtime["start"] = None
            metrics["last_timestamp"] = now
            metrics["last_value"] = max(stage_total, stage_value)

        stage_label: str = ""
        if active_stage is not None:
            label = _STAGE_LABELS.get(active_stage, active_stage.title())
            if active_total > 0:
                stage_label = f"{label} ({active_value}/{active_total})"
            else:
                stage_label = label
        elif current < total:
            stage_label = "Finalising build..."
        if stage_label != self._progress_stage_text:
            self._progress_stage_text = stage_label
        if active_stage != self._last_logged_stage:
            if active_stage is None:
                if current < total:
                    self.logger.info("Stage complete; finalising remaining tasks…")
                else:
                    self.logger.info("All build stages completed.")
            else:
                if active_total > 0:
                    self.logger.info(
                        "Stage: %s (%s/%s)",
                        _STAGE_LABELS.get(active_stage, active_stage.title()),
                        active_value,
                        active_total,
                    )
                else:
                    self.logger.info(
                        "Stage: %s",
                        _STAGE_LABELS.get(active_stage, active_stage.title()),
                    )
            self._last_logged_stage = active_stage

        percent = int(round(100 * current / total))
        if percent < self._displayed_percent:
            percent = self._displayed_percent
        else:
            self._displayed_percent = percent
        if self.progress_bar.maximum() == 0 and self.progress_bar.minimum() == 0:
            self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(max(0, min(100, percent)))
        self._overall_progress_current = current
        self._overall_progress_total = total
        self._progress_counts_text = f"{current}/{total}"
        self._update_eta_display(now=now)

    def _handle_finished(self, result: BuildResult) -> None:
        self._set_running(False)
        self.progress_bar.setValue(100)
        self.progress_label.setText("Complete")
        if result.exports:
            lines = [f"{fmt.upper()}: {path}" for fmt, path in result.exports.items()]
            export_text = "\n".join(lines)
            open_target = self._preferred_export_path(result.exports)
        else:
            export_text = "No export files were created."
            open_target = None

        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Microwire Data Builder")
        msg.setIcon(QtWidgets.QMessageBox.Icon.Information)
        msg.setText("Build finished successfully.")
        msg.setInformativeText(export_text)
        open_button: QtWidgets.QAbstractButton | None = None
        if open_target is not None:
            open_button = msg.addButton(
                "Open", QtWidgets.QMessageBox.ButtonRole.AcceptRole
            )
        close_button = msg.addButton("Close", QtWidgets.QMessageBox.ButtonRole.RejectRole)
        msg.exec()
        if (
            open_target is not None
            and open_button is not None
            and msg.clickedButton() is open_button
        ):
            self._open_path(open_target)

    def _handle_failed(self, message: str) -> None:
        self._set_running(False)
        self.progress_bar.setValue(0)
        self.progress_label.setText("Failed")
        QtWidgets.QMessageBox.critical(
            self,
            "Microwire Data Builder",
            "Build failed.\n\n" + message,
        )

    def _handle_cancelled(self) -> None:
        self._set_running(False)
        self.progress_bar.setValue(0)
        self.progress_label.setText("Cancelled")
        QtWidgets.QMessageBox.information(
            self,
            "Microwire Data Builder",
            "Build cancelled by user.",
        )

    def _cleanup_thread(self) -> None:
        if self._worker is not None:
            self._worker.deleteLater()
            self._worker = None
        if self._thread is not None:
            self._thread.deleteLater()
            self._thread = None

    # ------------------------------------------------------------------ Qt hooks
    def _append_log(self, level: int, message: str) -> None:
        _ = level
        self.log_view.appendPlainText(message)
        scrollbar = self.log_view.verticalScrollBar()
        if scrollbar is not None:
            scrollbar.setValue(scrollbar.maximum())

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:  # type: ignore[override]
        if self._running:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "A build is currently running. Please wait for it to finish before closing.",
            )
            event.ignore()
            return
        if hasattr(self, "_log_handler") and self._log_handler in self.logger.handlers:
            self.logger.removeHandler(self._log_handler)
        self._save_settings()
        super().closeEvent(event)


@dataclass
class SectionProcessResult:
    table: pd.DataFrame
    processed: Dict[str, float]
    payloads: Dict[str, Any] = field(default_factory=dict)
    extra: Dict[str, Any] = field(default_factory=dict)


class DataFrameModel(QtCore.QAbstractTableModel):
    """Expose a pandas DataFrame to Qt view widgets."""

    def __init__(self, frame: pd.DataFrame | None = None, parent: QtCore.QObject | None = None) -> None:
        super().__init__(parent)
        self._frame = frame.copy() if frame is not None else pd.DataFrame()
        self._decoration_provider: Optional[
            Callable[[pd.Series, str], Optional[QtGui.QPixmap | QtGui.QImage]]
        ] = None
        self._background_provider: Optional[
            Callable[[pd.Series, str], Optional[QtGui.QBrush]]
        ] = None
        self._foreground_provider: Optional[
            Callable[[pd.Series, str], Optional[QtGui.QBrush]]
        ] = None
        self._editable_columns: set[str] = set()
        self._text_columns: set[str] = set()

    def set_frame(self, frame: pd.DataFrame | None) -> None:
        self.beginResetModel()
        self._frame = frame.copy() if frame is not None else pd.DataFrame()
        self.endResetModel()
        try:
            self.layoutChanged.emit()
        except Exception:
            pass

    def set_decoration_provider(
        self,
        provider: Optional[Callable[[pd.Series, str], Optional[QtGui.QPixmap | QtGui.QImage]]],
    ) -> None:
        self._decoration_provider = provider
        try:
            self.layoutChanged.emit()
        except Exception:
            pass

    def set_background_provider(
        self,
        provider: Optional[Callable[[pd.Series, str], Optional[QtGui.QBrush]]],
    ) -> None:
        self._background_provider = provider

    def set_foreground_provider(
        self,
        provider: Optional[Callable[[pd.Series, str], Optional[QtGui.QBrush]]],
    ) -> None:
        self._foreground_provider = provider

    def set_editable_columns(self, columns: Iterable[str]) -> None:
        self._editable_columns = {str(column) for column in columns}

    def set_text_columns(self, columns: Iterable[str]) -> None:
        self._text_columns = {str(column) for column in columns}

    def frame(self) -> pd.DataFrame:
        return self._frame

    def rowCount(self, parent: QtCore.QModelIndex = QtCore.QModelIndex()) -> int:  # type: ignore[override]
        if parent.isValid():
            return 0
        return len(self._frame.index)

    def columnCount(self, parent: QtCore.QModelIndex = QtCore.QModelIndex()) -> int:  # type: ignore[override]
        if parent.isValid():
            return 0
        return len(self._frame.columns)

    @staticmethod
    def _sort_value(value: Any) -> Any:
        if isinstance(value, (QtGui.QPixmap, QtGui.QImage)):
            return ""
        if isinstance(value, (pd.Timestamp, datetime)):
            return value
        if isinstance(value, str):
            match = MICROWIRE_SORT_RE.match(value)
            if match:
                try:
                    draw = int(match.group(1))
                    piece = int(match.group(2))
                except (TypeError, ValueError):
                    pass
                else:
                    suffix = match.group(3) or ""
                    suffix_text = suffix.strip().lower()
                    return f"{draw:05d}/{piece:05d}{suffix_text}"
        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(item) for item in value)
        return value

    def data(
        self,
        index: QtCore.QModelIndex,
        role: int = QtCore.Qt.ItemDataRole.DisplayRole,
    ) -> Any:  # type: ignore[override]
        if not index.isValid():
            return None
        try:
            value = self._frame.iat[index.row(), index.column()]
        except Exception:
            return None
        if role == QtCore.Qt.ItemDataRole.DecorationRole:
            provider = getattr(self, "_decoration_provider", None)
            if provider is not None:
                try:
                    column_label = str(self._frame.columns[index.column()])
                    row_series = self._frame.iloc[index.row()]
                except Exception:
                    pass
                else:
                    try:
                        decoration = provider(row_series, column_label)
                    except Exception:
                        decoration = None
                    if isinstance(decoration, (QtGui.QPixmap, QtGui.QImage)):
                        if isinstance(decoration, QtGui.QImage):
                            return QtGui.QPixmap.fromImage(decoration)
                        return decoration
            if isinstance(value, QtGui.QPixmap):
                return value
            if isinstance(value, QtGui.QImage):
                return QtGui.QPixmap.fromImage(value)
            return None
        if role == QtCore.Qt.ItemDataRole.ForegroundRole:
            provider = getattr(self, "_foreground_provider", None)
            if provider is not None:
                try:
                    column_label = str(self._frame.columns[index.column()])
                    row_series = self._frame.iloc[index.row()]
                except Exception:
                    pass
                else:
                    brush = provider(row_series, column_label)
                    if brush is not None:
                        return brush
            try:
                column_label = str(self._frame.columns[index.column()])
            except Exception:
                column_label = ""
            if column_label.lower() == "reviewed":
                ok = bool(value)
                fg = QtGui.QColor("#10b981" if ok else "#ef4444")
                return QtGui.QBrush(fg)
            return None
        if role == QtCore.Qt.ItemDataRole.BackgroundRole:
            provider = getattr(self, "_background_provider", None)
            if provider is not None:
                try:
                    column_label = str(self._frame.columns[index.column()])
                    row_series = self._frame.iloc[index.row()]
                except Exception:
                    pass
                else:
                    brush = provider(row_series, column_label)
                    if brush is not None:
                        return brush
            try:
                column_label = str(self._frame.columns[index.column()])
            except Exception:
                column_label = ""
            if column_label.lower() == "reviewed":
                ok = bool(value)
                bg = QtGui.QColor("#07351f" if ok else "#3a0a0a")
                return QtGui.QBrush(bg)
            return None
        if role not in (
            QtCore.Qt.ItemDataRole.DisplayRole,
            QtCore.Qt.ItemDataRole.EditRole,
        ):
            return None
        if isinstance(value, (QtGui.QPixmap, QtGui.QImage)):
            return ""
        if isinstance(value, float):
            if math.isnan(value):
                return ""
            return f"{value:.4g}"
        return str(value) if value is not None else ""

    def flags(self, index: QtCore.QModelIndex) -> QtCore.Qt.ItemFlags:  # type: ignore[override]
        base_flags = super().flags(index)
        if not index.isValid():
            return base_flags
        try:
            column_label = str(self._frame.columns[index.column()])
        except Exception:
            return base_flags
        if column_label in self._editable_columns:
            return base_flags | QtCore.Qt.ItemFlag.ItemIsEditable
        return base_flags

    def setData(
        self,
        index: QtCore.QModelIndex,
        value: Any,
        role: int = QtCore.Qt.ItemDataRole.EditRole,
    ) -> bool:  # type: ignore[override]
        if role not in (
            QtCore.Qt.ItemDataRole.EditRole,
            QtCore.Qt.ItemDataRole.DisplayRole,
        ):
            return False
        if not index.isValid():
            return False
        try:
            column_label = str(self._frame.columns[index.column()])
        except Exception:
            return False
        if column_label not in self._editable_columns:
            return False
        if isinstance(value, QtCore.QVariant):  # pragma: no cover - PyQt guard
            value = value.value()
        if isinstance(value, str):
            text = value.strip()
            if column_label in self._text_columns:
                coerced = text if text else None
            else:
                numeric_text = text.replace(",", ".")
                if not numeric_text:
                    coerced = None
                else:
                    try:
                        coerced = float(numeric_text)
                    except ValueError:
                        return False
        elif isinstance(value, (int, float)):
            numeric = float(value)
            coerced = numeric if math.isfinite(numeric) else None
        else:
            return False
        try:
            self._frame.iat[index.row(), index.column()] = coerced
        except Exception:
            return False
        try:
            self.dataChanged.emit(
                index,
                index,
                [
                    QtCore.Qt.ItemDataRole.DisplayRole,
                    QtCore.Qt.ItemDataRole.EditRole,
                ],
            )
        except Exception:
            pass
        return True

    def headerData(
        self,
        section: int,
        orientation: QtCore.Qt.Orientation,
        role: int = QtCore.Qt.ItemDataRole.DisplayRole,
    ) -> Any:  # type: ignore[override]
        if role != QtCore.Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == QtCore.Qt.Orientation.Horizontal:
            try:
                return str(self._frame.columns[section])
            except Exception:
                return ""
        try:
            label = self._frame.index[section]
        except Exception:
            return str(section + 1)
        return str(label)

    def sort(self, column: int, order: QtCore.Qt.SortOrder = QtCore.Qt.SortOrder.AscendingOrder) -> None:  # type: ignore[override]
        if self._frame.empty:
            return
        try:
            column_label = self._frame.columns[column]
        except Exception:
            return
        ascending = order != QtCore.Qt.SortOrder.DescendingOrder
        try:
            sorted_frame = self._frame.sort_values(
                by=column_label,
                ascending=ascending,
                kind="mergesort",
                key=lambda col: col.map(self._sort_value) if hasattr(col, "map") else col,
            )
        except Exception:
            return
        self.beginResetModel()
        self._frame = sorted_frame.reset_index(drop=True)
        self.endResetModel()


class _TableSearchProxyModel(QtCore.QSortFilterProxyModel):
    def __init__(
        self,
        parent: QtCore.QObject | None = None,
    ) -> None:
        super().__init__(parent)
        self._search_text = ""
        self._row_predicate: Optional[Callable[[pd.Series], bool]] = None

    def set_search_text(self, text: str) -> None:
        self._search_text = str(text or "").strip().lower()
        self.invalidateFilter()

    def set_row_predicate(
        self,
        predicate: Optional[Callable[[pd.Series], bool]],
    ) -> None:
        self._row_predicate = predicate
        self.invalidateFilter()

    def filterAcceptsRow(
        self,
        source_row: int,
        source_parent: QtCore.QModelIndex,
    ) -> bool:
        model = self.sourceModel()
        if not isinstance(model, DataFrameModel):
            return True
        frame = model.frame()
        if source_row < 0 or source_row >= len(frame.index):
            return False
        try:
            row = frame.iloc[source_row]
        except Exception:
            return False
        predicate = self._row_predicate
        if predicate is not None:
            try:
                if not predicate(row):
                    return False
            except Exception:
                return False
        if not self._search_text:
            return True
        for column in frame.columns:
            name = str(column)
            if name.startswith("_"):
                continue
            value = row.get(column)
            if isinstance(value, (list, tuple, set)):
                text = ", ".join(str(item) for item in value)
            else:
                text = "" if value is None else str(value)
            if self._search_text in text.lower():
                return True
        return False

    def map_row_to_source(self, row: int) -> Optional[int]:
        index = self.index(row, 0)
        if not index.isValid():
            return None
        source_index = self.mapToSource(index)
        if not source_index.isValid():
            return None
        return int(source_index.row())

    def lessThan(
        self,
        left: QtCore.QModelIndex,
        right: QtCore.QModelIndex,
    ) -> bool:
        model = self.sourceModel()
        if not isinstance(model, DataFrameModel):
            return super().lessThan(left, right)
        try:
            left_value = model.frame().iat[left.row(), left.column()]
            right_value = model.frame().iat[right.row(), right.column()]
        except Exception:
            return super().lessThan(left, right)
        return model._sort_value(left_value) < model._sort_value(right_value)


def _dimension_display(field: str, *records: Dict[str, Any]) -> Optional[str]:
    values: List[str] = []
    seen: Set[str] = set()
    for record in records:
        if not isinstance(record, dict):
            continue
        bucket = record.get(f"{field}__display")
        if isinstance(bucket, (list, tuple)):
            for entry in bucket:
                text = _clean_str(entry)
                if not text or text in seen:
                    continue
                values.append(text)
                seen.add(text)
        fallback = _value_for_output(record, field)
        if fallback is None:
            continue
        display = _format_dimension_display(field, fallback)
        if not display:
            continue
        if display not in seen:
            values.append(display)
            seen.add(display)
    if not values:
        return None
    return "\n".join(values)


def _fabrication_index_to_frame(index: FabricationIndex) -> pd.DataFrame:
    columns = [
        "Composition",
        "Data source",
        "e/a",
        ESTIMATED_TRANSITION_COLUMN,
        "Draw",
        "Piece",
        "Length (m)",
        "Piece date",
        MICROSCOPE_D_COLUMN,
        MICROSCOPE_CAP_D_COLUMN,
        "d/D",
        "Resistance (Ω)",
        CORE_TEMPERATURE_COLUMN,
        GLASS_TEMPERATURE_COLUMN,
        "Mass (g)",
        "Winding speed (m/min)",
        "Glass feeding (mm/min)",
        "Underpressure",
        GLASS_PULL_COLUMN,
        "Notes",
        "Production datetime",
        "_source_paths",
    ]
    rows: List[Dict[str, Any]] = []
    for (composition, draw, piece), piece_record in sorted(index.piece_level.items()):
        draw_record = index.get_draw(composition, draw)
        row: Dict[str, Any] = {column: None for column in columns}
        row["Composition"] = composition
        row["Data source"] = None
        ea_value = _compute_ea_from_composition(composition)
        row["e/a"] = ea_value
        row[ESTIMATED_TRANSITION_COLUMN] = _estimate_transition_temp_c(ea_value)
        row["Draw"] = draw
        row["Piece"] = piece
        row["Length (m)"] = _value_for_output(piece_record, "length_m")
        row["Piece date"] = _value_for_output(piece_record, "piece_date")
        row[MICROSCOPE_D_COLUMN] = _dimension_display("d_um", piece_record, draw_record)
        row[MICROSCOPE_CAP_D_COLUMN] = _dimension_display("D_um", piece_record, draw_record)
        row["d/D"] = _dimension_display("d_over_D", piece_record, draw_record)
        piece_resistance = _value_for_output(piece_record, "fabrication_resistance_ohm")
        draw_resistance = _value_for_output(draw_record, "fabrication_resistance_ohm")
        row["Resistance (Ω)"] = piece_resistance if piece_resistance is not None else draw_resistance
        row[CORE_TEMPERATURE_COLUMN] = _value_for_output(
            draw_record,
            "fabrication_temperature_c",
        )
        row[GLASS_TEMPERATURE_COLUMN] = None
        row["Mass (g)"] = _value_for_output(draw_record, "mass_g")
        row["Winding speed (m/min)"] = _value_for_output(draw_record, "winding_speed_m_per_min")
        row["Glass feeding (mm/min)"] = _value_for_output(draw_record, "glass_feed_mm_per_min")
        row["Underpressure"] = _value_for_output(draw_record, "underpressure")
        pull_value = _value_for_output(piece_record, "glass_pull_off")
        if pull_value is None:
            pull_value = _value_for_output(draw_record, "glass_pull_off")
        row[GLASS_PULL_COLUMN] = pull_value
        row["Notes"] = _compose_notes(draw_record, piece_record)
        row["Production datetime"] = _value_for_output(draw_record, "production_datetime")
        sources: List[str] = []
        for record in (piece_record, draw_record):
            path_value = record.get("_source_path") if isinstance(record, dict) else None
            if not path_value:
                continue
            try:
                resolved = str(Path(path_value))
            except Exception:
                resolved = str(path_value)
            if resolved and resolved not in sources:
                sources.append(resolved)
        row["_source_paths"] = sources
        if any("Imported" in source for source in sources):
            row["Data source"] = "Imported"
        else:
            row["Data source"] = "Measured"
        rows.append(row)
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _render_measurement_pixmap(
    record: Optional[MeasurementRecord],
    logger: logging.Logger,
    *,
    width_px: int = ANNEALING_GRAPH_WIDTH,
    height_px: int = ANNEALING_GRAPH_HEIGHT,
) -> Optional[QtGui.QPixmap]:
    if record is None:
        return None
    frame = record.dataframe if isinstance(record.dataframe, pd.DataFrame) else pd.DataFrame()
    if frame.empty:
        return None
    if "I_mA" in frame.columns and "R_ohm" in frame.columns:
        plot_df = pd.DataFrame(
            {
                "I_mA": pd.to_numeric(frame["I_mA"], errors="coerce"),
                "R_Ohm": pd.to_numeric(frame["R_ohm"], errors="coerce"),
            }
        ).dropna()
    elif "I_A" in frame.columns and "R_ohm" in frame.columns:
        currents_A = pd.to_numeric(frame["I_A"], errors="coerce")
        max_abs = float(currents_A.abs().max(skipna=True) or 0.0)
        if max_abs <= 50:
            currents_mA = currents_A * 1e3
        else:
            currents_mA = currents_A
        plot_df = pd.DataFrame(
            {
                "I_mA": currents_mA,
                "R_Ohm": pd.to_numeric(frame["R_ohm"], errors="coerce"),
            }
        ).dropna()
    elif "I_mA" in frame.columns and "R_Ohm" in frame.columns:
        plot_df = pd.DataFrame(
            {
                "I_mA": pd.to_numeric(frame["I_mA"], errors="coerce"),
                "R_Ohm": pd.to_numeric(frame["R_Ohm"], errors="coerce"),
            }
        ).dropna()
    else:
        columns = [str(column) for column in frame.columns]
        logger.debug(
            "Annealing preview missing expected columns: %s",
            ", ".join(columns),
        )
        return None
    if plot_df.empty:
        return None

    metadata = getattr(record, "metadata", None)
    title = ""
    if metadata is not None:
        try:
            title = format_annealing_title(metadata)
        except Exception:
            title = ""
    target_width = max(int(width_px * 2), width_px)
    target_height = max(int(height_px * 2), height_px)
    figsize = (max(target_width / 96.0, 1.0), max(target_height / 96.0, 1.0))
    canvas_agg: FigureCanvasAgg | None = None
    figure = None
    rc_overrides = {
        "axes.titlesize": ANNEALING_TITLE_FONT_SIZE,
        "axes.labelsize": ANNEALING_AXIS_FONT_SIZE,
        "xtick.labelsize": ANNEALING_TICK_FONT_SIZE,
        "ytick.labelsize": ANNEALING_TICK_FONT_SIZE,
        "lines.linewidth": 1.0,
        "lines.markersize": 3.0,
    }
    try:
        with plt.rc_context(rc_overrides):
            figure, _ = plot_annealing_curve(
                plot_df,
                title,
                target_px=(target_width, target_height),
            )
        if figure is not None:
            figure.subplots_adjust(left=0.08, right=0.98, top=0.9, bottom=0.16)
            for ax in figure.axes:
                try:
                    ax.tick_params(labelsize=ANNEALING_TICK_FONT_SIZE)
                except Exception:
                    pass
                try:
                    ax.xaxis.label.set_fontsize(ANNEALING_AXIS_FONT_SIZE)
                    ax.yaxis.label.set_fontsize(ANNEALING_AXIS_FONT_SIZE)
                except Exception:
                    pass
                if ax.get_title():
                    try:
                        ax.set_title(ax.get_title(), fontsize=ANNEALING_TITLE_FONT_SIZE)
                    except Exception:
                        pass
                legend = ax.get_legend()
                if legend is not None:
                    try:
                        legend.remove()
                    except Exception:
                        legend.set_visible(False)
        canvas_agg = FigureCanvasAgg(figure)
        canvas_agg.draw()
        width, height = canvas_agg.get_width_height()
        buffer = canvas_agg.buffer_rgba()
        # PyQt6 exposes formats via the QtGui.QImage.Format enum; older builds fall
        # back to module-level constants. Attempt the modern attribute first and
        # gracefully degrade through sensible alternatives so we never raise an
        # AttributeError (which previously prevented any thumbnails from
        # rendering).
        format_candidates: list[QtGui.QImage.Format | int] = []
        for attr in ("Format_RGBA8888", "Format_RGBA8888_Premultiplied", "Format_ARGB32"):
            candidate = None
            try:
                candidate = getattr(QtGui.QImage.Format, attr)
            except AttributeError:
                candidate = getattr(QtGui.QImage, attr, None)
            if candidate is not None:
                format_candidates.append(candidate)
        if not format_candidates:
            try:
                fallback = QtGui.QImage.Format.Format_ARGB32
            except AttributeError:
                fallback = getattr(QtGui.QImage, "Format_ARGB32", None)
            if fallback is None:
                fallback = QtGui.QImage.Format_ARGB32  # type: ignore[attr-defined]
            format_candidates.append(fallback)

        image: QtGui.QImage | None = None
        for fmt in format_candidates:
            try:
                candidate = QtGui.QImage(buffer, width, height, 4 * width, fmt)
            except TypeError:
                continue
            if not candidate.isNull():
                image = candidate
                break
        if image is None:
            try:
                fallback = QtGui.QImage.Format.Format_ARGB32
            except AttributeError:
                fallback = getattr(QtGui.QImage, "Format_ARGB32", None)
            if fallback is None:
                fallback = QtGui.QImage.Format_ARGB32  # type: ignore[attr-defined]
            image = QtGui.QImage(buffer, width, height, 4 * width, fallback)
        pixmap = QtGui.QPixmap.fromImage(image.copy())
        return pixmap.scaled(
            width_px,
            height_px,
            QtCore.Qt.AspectRatioMode.KeepAspectRatio,
            QtCore.Qt.TransformationMode.SmoothTransformation,
        )
    except Exception:
        logger.exception(
            "Failed to render annealing preview for %s",
            getattr(record, "path", "<unknown>"),
        )
        return None
    finally:
        if figure is not None:
            plt.close(figure)


def _figure_to_pixmap(
    figure: Optional["plt.Figure"],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> Optional[QtGui.QPixmap]:
    if figure is None:
        return None
    canvas_agg: FigureCanvasAgg | None = None
    try:
        canvas_agg = FigureCanvasAgg(figure)
        canvas_agg.draw()
        width, height = canvas_agg.get_width_height()
        buffer = canvas_agg.buffer_rgba()
        format_candidates: list[QtGui.QImage.Format | int] = []
        for attr in ("Format_RGBA8888", "Format_RGBA8888_Premultiplied", "Format_ARGB32"):
            candidate = None
            try:
                candidate = getattr(QtGui.QImage.Format, attr)
            except AttributeError:
                candidate = getattr(QtGui.QImage, attr, None)
            if candidate is not None:
                format_candidates.append(candidate)
        if not format_candidates:
            try:
                fallback = QtGui.QImage.Format.Format_ARGB32
            except AttributeError:
                fallback = getattr(QtGui.QImage, "Format_ARGB32", None)
            if fallback is None:
                fallback = QtGui.QImage.Format_ARGB32  # type: ignore[attr-defined]
            format_candidates.append(fallback)

        image: QtGui.QImage | None = None
        for fmt in format_candidates:
            try:
                candidate = QtGui.QImage(buffer, width, height, 4 * width, fmt)
            except TypeError:
                continue
            if not candidate.isNull():
                image = candidate
                break
        if image is None:
            return None
        pixmap = QtGui.QPixmap.fromImage(image.copy())
        return pixmap.scaled(
            width_px,
            height_px,
            QtCore.Qt.AspectRatioMode.KeepAspectRatio,
            QtCore.Qt.TransformationMode.SmoothTransformation,
        )
    except Exception:
        logger.exception("Failed to render graph preview")
        return None
    finally:
        if figure is not None:
            plt.close(figure)


def _figure_to_data_uri(
    figure: Optional["plt.Figure"],
    logger: logging.Logger,
) -> Optional[str]:
    if figure is None:
        return None
    buffer = io.BytesIO()
    try:
        figure.savefig(buffer, format="png", dpi=96)
        encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
        return f"data:image/png;base64,{encoded}"
    except Exception:
        logger.exception("Failed to render graph preview for HTML export")
        return None
    finally:
        buffer.close()
        plt.close(figure)


def _axis_column_range(
    records: Sequence[object],
    column: str,
) -> Optional[float]:
    best_range: Optional[float] = None
    for entry in records:
        frame = getattr(entry, "data", None)
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            continue
        if column not in frame.columns:
            continue
        series = pd.to_numeric(frame[column], errors="coerce").dropna()
        if series.empty:
            continue
        try:
            span = float(series.max() - series.min())
        except Exception:
            continue
        if best_range is None or span > best_range:
            best_range = span
    return best_range


def _axis_column_bounds(
    records: Sequence[object],
    column: str,
) -> Optional[Tuple[float, float]]:
    minimum: Optional[float] = None
    maximum: Optional[float] = None
    for entry in records:
        frame = getattr(entry, "data", None)
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            continue
        if column not in frame.columns:
            continue
        series = pd.to_numeric(frame[column], errors="coerce").dropna()
        if series.empty:
            continue
        try:
            current_min = float(series.min())
            current_max = float(series.max())
        except Exception:
            continue
        if minimum is None or current_min < minimum:
            minimum = current_min
        if maximum is None or current_max > maximum:
            maximum = current_max
    if minimum is None or maximum is None:
        return None
    return minimum, maximum


def _downsample_series(series: pd.Series, max_points: int = 2000) -> pd.Series:
    if max_points <= 0:
        return series
    length = len(series)
    if length <= max_points:
        return series
    stride = max(int(math.ceil(length / max_points)), 1)
    return series.iloc[::stride]


def _downsample_values(values: Sequence[float], max_points: int = 2000) -> Sequence[float]:
    if max_points <= 0:
        return values
    length = len(values)
    if length <= max_points:
        return values
    stride = max(int(math.ceil(length / max_points)), 1)
    return values[::stride]


def _match_axis_setting(columns: Sequence[str], stored: Optional[str]) -> Optional[str]:
    if not stored:
        return None
    stored_text = str(stored).strip()
    if not stored_text:
        return None
    stored_norm = _normalize_axis_label(stored_text)
    for column in columns:
        if column == stored_text:
            return column
    stored_lower = stored_text.lower()
    for column in columns:
        if column.lower() == stored_lower:
            return column
    for column in columns:
        if _normalize_axis_label(column) == stored_norm:
            return column
    return None


def _normalize_axis_label(label: str) -> str:
    return re.sub(r"\s+", " ", label.strip().lower())


def _choose_axis_column(
    columns: Sequence[str],
    preferences: Sequence[str],
    *,
    records: Optional[Sequence[object]] = None,
    stored: Optional[str] = None,
) -> Optional[str]:
    if not columns:
        return None
    stored_match = _match_axis_setting(columns, stored)
    candidates: List[str] = []
    normalized_columns = [(column, _normalize_axis_label(column)) for column in columns]
    for pref in preferences:
        pref_norm = _normalize_axis_label(pref)
        for column, normalized in normalized_columns:
            if pref_norm in normalized and column not in candidates:
                candidates.append(column)
    if not candidates:
        candidates = list(columns)
    if records:
        best_column = None
        best_range = -1.0
        for column in candidates:
            span = _axis_column_range(records, column)
            if span is None:
                continue
            if span > best_range:
                best_range = span
                best_column = column
        if stored_match:
            stored_range = _axis_column_range(records, stored_match)
            if stored_range is not None:
                if best_range > 0 and stored_range / best_range < 0.01:
                    return best_column or stored_match
                return stored_match
            if best_column:
                return best_column
        if best_column:
            return best_column
    if stored_match:
        return stored_match
    return candidates[0] if candidates else None


def _combine_pixmaps_side_by_side(
    pixmaps: Sequence[QtGui.QPixmap],
    *,
    width_px: int,
    height_px: int,
    spacing: int = 6,
    scale_to_fit: bool = True,
) -> Optional[QtGui.QPixmap]:
    if not pixmaps:
        return None
    count = len(pixmaps)
    width_px = max(int(width_px), 1)
    height_px = max(int(height_px), 1)
    spacing = max(int(spacing), 0)
    if (
        count == 1
        and pixmaps[0].width() == width_px
        and pixmaps[0].height() == height_px
    ):
        return pixmaps[0]
    if scale_to_fit:
        available = max(width_px - spacing * (count - 1), count)
        slot_widths = [max(int(available / count), 1) for _ in pixmaps]
    else:
        slot_widths = [max(int(pixmap.width()), 1) for pixmap in pixmaps]
        total_width = sum(slot_widths) + spacing * (count - 1)
        if total_width > width_px:
            available = max(width_px - spacing * (count - 1), count)
            slot_widths = [max(int(available / count), 1) for _ in pixmaps]
    target = QtGui.QPixmap(width_px, height_px)
    target.fill(QtCore.Qt.GlobalColor.transparent)
    painter = QtGui.QPainter(target)
    try:
        x_pos = 0
        for pixmap, slot_width in zip(pixmaps, slot_widths):
            scaled = pixmap
            if pixmap.width() != slot_width or pixmap.height() != height_px:
                scaled = pixmap.scaled(
                    slot_width,
                    height_px,
                    QtCore.Qt.AspectRatioMode.KeepAspectRatio,
                    QtCore.Qt.TransformationMode.SmoothTransformation,
                )
            y_pos = max((height_px - scaled.height()) // 2, 0)
            painter.drawPixmap(x_pos, y_pos, scaled)
            x_pos += slot_width + spacing
    finally:
        painter.end()
    return target


def _combine_pixmaps_vertical(
    pixmaps: Sequence[QtGui.QPixmap],
    *,
    width_px: int,
    height_px: int,
    spacing: int = 6,
    scale_to_fit: bool = True,
) -> Optional[QtGui.QPixmap]:
    if not pixmaps:
        return None
    count = len(pixmaps)
    width_px = max(int(width_px), 1)
    height_px = max(int(height_px), 1)
    spacing = max(int(spacing), 0)
    if (
        count == 1
        and pixmaps[0].width() == width_px
        and pixmaps[0].height() == height_px
    ):
        return pixmaps[0]
    if scale_to_fit:
        available = max(height_px - spacing * (count - 1), count)
        slot_heights = [max(int(available / count), 1) for _ in pixmaps]
    else:
        slot_heights = [max(int(pixmap.height()), 1) for pixmap in pixmaps]
        total_height = sum(slot_heights) + spacing * (count - 1)
        if total_height > height_px:
            available = max(height_px - spacing * (count - 1), count)
            slot_heights = [max(int(available / count), 1) for _ in pixmaps]
    target = QtGui.QPixmap(width_px, height_px)
    target.fill(QtCore.Qt.GlobalColor.transparent)
    painter = QtGui.QPainter(target)
    try:
        y_pos = 0
        for pixmap, slot_height in zip(pixmaps, slot_heights):
            scaled = pixmap
            if pixmap.width() != width_px or pixmap.height() != slot_height:
                scaled = pixmap.scaled(
                    width_px,
                    slot_height,
                    QtCore.Qt.AspectRatioMode.KeepAspectRatio,
                    QtCore.Qt.TransformationMode.SmoothTransformation,
                )
            x_pos = max((width_px - scaled.width()) // 2, 0)
            painter.drawPixmap(x_pos, y_pos, scaled)
            y_pos += slot_height + spacing
    finally:
        painter.end()
    return target


@dataclass
class _VsmHysteresisPlotGroup:
    label: str
    records: List[VsmHysteresisRecord]
    variant: Optional[str]
    temperature: Optional[float]


def _coerce_finite_float(value: object) -> Optional[float]:
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    return None


def _format_vsm_hysteresis_group_label(
    temperature: Optional[float],
    variant: Optional[str],
) -> str:
    parts: List[str] = []
    if temperature is not None:
        parts.append(f"T{temperature:g}C")
    if variant:
        parts.append(variant)
    return " — ".join(parts)


def _group_vsm_hysteresis_plot_groups(
    records: Sequence[VsmHysteresisRecord],
) -> List[_VsmHysteresisPlotGroup]:
    grouped: Dict[Tuple[Optional[str], Optional[float]], List[VsmHysteresisRecord]] = {}
    temp_tolerance = 1.0
    merge_tolerance = 2.0
    by_variant: Dict[Optional[str], List[Tuple[float, VsmHysteresisRecord]]] = {}

    def _bucket_angle_count(items: Sequence[VsmHysteresisRecord]) -> int:
        angles: Set[float] = set()
        for entry in items:
            angle = _coerce_finite_float(getattr(entry, "angle", None))
            if angle is not None:
                angles.add(round(angle, 3))
        return len(angles) if angles else 1

    for record in records:
        variant = getattr(record, "variant", None)
        if isinstance(variant, str):
            variant = variant.strip() or None
        temp = _coerce_finite_float(getattr(record, "temperature", None))
        if temp is None:
            setattr(record, "_group_temperature", None)
            grouped.setdefault((variant, None), []).append(record)
            continue
        by_variant.setdefault(variant, []).append((temp, record))

    for variant, entries in by_variant.items():
        entries.sort(key=lambda item: item[0])
        buckets: List[Tuple[float, List[VsmHysteresisRecord]]] = []
        for temp, record in entries:
            if not buckets:
                buckets.append((temp, [record]))
                setattr(record, "_group_temperature", temp)
                continue
            last_temp, bucket_records = buckets[-1]
            if abs(temp - last_temp) <= temp_tolerance:
                bucket_records.append(record)
                setattr(record, "_group_temperature", last_temp)
            else:
                buckets.append((temp, [record]))
                setattr(record, "_group_temperature", temp)
        if buckets:
            merged: List[Tuple[float, List[VsmHysteresisRecord]]] = []
            for idx, (bucket_temp, bucket_records) in enumerate(buckets):
                angle_count = _bucket_angle_count(bucket_records)
                merged_into_previous = False
                if merged:
                    prev_temp, prev_records = merged[-1]
                    prev_count = _bucket_angle_count(prev_records)
                    if (
                        angle_count <= 1
                        and prev_count >= angle_count
                        and abs(bucket_temp - prev_temp) <= merge_tolerance
                    ):
                        prev_records.extend(bucket_records)
                        for record in bucket_records:
                            setattr(record, "_group_temperature", prev_temp)
                        merged_into_previous = True
                if merged_into_previous:
                    continue
                next_bucket = buckets[idx + 1] if idx + 1 < len(buckets) else None
                if next_bucket:
                    next_temp, next_records = next_bucket
                    next_count = _bucket_angle_count(next_records)
                    if (
                        angle_count <= 1
                        and next_count >= angle_count
                        and abs(bucket_temp - next_temp) <= merge_tolerance
                    ):
                        next_records.extend(bucket_records)
                        for record in bucket_records:
                            setattr(record, "_group_temperature", next_temp)
                        continue
                merged.append((bucket_temp, bucket_records))
            buckets = merged
        for bucket_temp, bucket_records in buckets:
            grouped.setdefault((variant, bucket_temp), []).extend(bucket_records)

    def _group_sort_key(
        entry: Tuple[Tuple[Optional[str], Optional[float]], List[VsmHysteresisRecord]]
    ) -> Tuple[int, str, int, float]:
        (variant, temp), _records = entry
        variant_key = variant or ""
        temp_value = temp if temp is not None else float("inf")
        return (
            0 if variant is None else 1,
            variant_key.lower(),
            0 if temp is not None else 1,
            temp_value,
        )

    groups: List[_VsmHysteresisPlotGroup] = []
    for (variant, temp), grouped_records in sorted(grouped.items(), key=_group_sort_key):
        grouped_records = sorted(
            grouped_records,
            key=lambda rec: (
                _coerce_finite_float(getattr(rec, "angle", None)) is None,
                _coerce_finite_float(getattr(rec, "angle", None)) or 0.0,
            ),
        )
        label = _format_vsm_hysteresis_group_label(temp, variant)
        groups.append(
            _VsmHysteresisPlotGroup(
                label=label,
                records=grouped_records,
                variant=variant,
                temperature=temp,
            )
        )
    return groups


def _plot_vsm_hysteresis_figure(
    record: VsmHysteresisRecord | Sequence[VsmHysteresisRecord],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> Optional["plt.Figure"]:
    if isinstance(record, VsmHysteresisRecord):
        records = [record]
    else:
        records = [entry for entry in record if isinstance(entry, VsmHysteresisRecord)]
    if not records:
        return None
    columns_set: set[str] = set()
    for entry in records:
        entry_frame = entry.data if isinstance(entry.data, pd.DataFrame) else pd.DataFrame()
        if entry_frame.empty:
            continue
        columns_set.update(str(column) for column in entry_frame.columns)
    if not columns_set:
        return None
    columns = sorted(columns_set)
    x_preferences = [
        "Applied Field For Plot",
        "Raw Applied Field For Plot",
        "Applied Field",
        "Raw Applied Field",
        "Applied Field [Oe]",
        "Field",
    ]
    x_column = _choose_axis_column(
        columns,
        x_preferences,
        records=records,
    )
    if x_column:
        bounds = _axis_column_bounds(records, x_column)
        if bounds is not None and (bounds[0] > 0 or bounds[1] < 0):
            candidates: List[str] = []
            normalized_columns = [(column, _normalize_axis_label(column)) for column in columns]
            for pref in x_preferences:
                pref_norm = _normalize_axis_label(pref)
                for column, normalized in normalized_columns:
                    if pref_norm in normalized and column not in candidates:
                        candidates.append(column)
            if not candidates:
                candidates = list(columns)
            zero_candidates: List[Tuple[float, str]] = []
            for candidate in candidates:
                candidate_bounds = _axis_column_bounds(records, candidate)
                if candidate_bounds is None:
                    continue
                if candidate_bounds[0] <= 0 <= candidate_bounds[1]:
                    zero_candidates.append((candidate_bounds[1] - candidate_bounds[0], candidate))
            if zero_candidates:
                zero_candidates.sort(reverse=True)
                x_column = zero_candidates[0][1]
    y_column = _choose_axis_column(
        columns,
        [
            "Signal X direction",
            "Signal parallel with sample",
            "Signal Magnitude",
            "Moment [emu]",
            "Signal",
        ],
    )
    if y_column == x_column and y_column:
        remaining = [column for column in columns if column != y_column]
        fallback = _choose_axis_column(
            remaining,
            [
                "Signal X direction",
                "Signal parallel with sample",
                "Signal Magnitude",
                "Moment [emu]",
                "Signal",
            ],
            records=records,
        )
        if fallback:
            y_column = fallback
    if not x_column or not y_column:
        logger.debug("VSM hysteresis plot missing axis columns: %s", ", ".join(columns))
        return None
    valid_records: List[Tuple[pd.Series, pd.Series, Optional[float]]] = []
    for entry in records:
        entry_frame = entry.data if isinstance(entry.data, pd.DataFrame) else pd.DataFrame()
        if entry_frame.empty or x_column not in entry_frame.columns or y_column not in entry_frame.columns:
            continue
        subset = entry_frame[[x_column, y_column]].apply(pd.to_numeric, errors="coerce").dropna()
        if subset.empty:
            continue
        x_series = subset[x_column]
        y_series = subset[y_column]
        if len(subset) > MAX_PLOT_POINTS:
            x_series = _downsample_series(x_series, MAX_PLOT_POINTS)
            y_series = y_series.loc[x_series.index]
        angle_value = _coerce_finite_float(getattr(entry, "angle", None))
        valid_records.append((x_series, y_series, angle_value))
    if not valid_records:
        return None
    figsize = (max(width_px / 96.0, 1.0), max(height_px / 96.0, 1.0))
    figure = plt.Figure(figsize=figsize)
    ax = figure.add_subplot(111)
    for x_series, y_series, angle_value in valid_records:
        label = f"{angle_value:g}°" if angle_value is not None else None
        ax.plot(x_series, y_series, linewidth=1.2, label=label)
    base_record = records[0]
    title = base_record.sample or Path(base_record.path).stem
    details: list[str] = []
    group_temp = getattr(base_record, "_group_temperature", None)
    temperature = _coerce_finite_float(group_temp)
    if temperature is None:
        temperature = _coerce_finite_float(getattr(base_record, "temperature", None))
    if temperature is not None:
        details.append(f"T{temperature:g}C")
    variant = getattr(base_record, "variant", None)
    if isinstance(variant, str) and variant.strip():
        details.append(variant.strip())
    if details:
        title = f"{title} ({', '.join(details)})"
    ax.set_title(title)
    ax.set_xlabel(x_column)
    ax.set_ylabel(y_column)
    if len(valid_records) > 1:
        ax.legend(loc="best")
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Tight layout not applied",
                category=UserWarning,
            )
            figure.tight_layout()
    except Exception:
        pass
    try:
        figure.subplots_adjust(left=0.12, right=0.98, top=0.9, bottom=0.16)
    except Exception:
        pass
    return figure


def _plot_vsm_temperature_scan_figure(
    record: VsmTemperatureScanRecord,
    processor: VSMTemperatureScanProcessor,
    *,
    width_px: int,
    height_px: int,
) -> Optional["plt.Figure"]:
    frame = record.data if isinstance(record.data, pd.DataFrame) else pd.DataFrame()
    if frame.empty:
        return None
    series = processor._build_series(frame.copy())
    if not series:
        return None
    color_map = processor.series_color_map(series)
    field_order = processor.field_axis_order([entry.field for entry in series])
    figsize = (max(width_px / 96.0, 1.0), max(height_px / 96.0, 1.0))
    figure = plt.Figure(figsize=figsize)
    ax_left = figure.add_subplot(111)
    ax_right = None
    axes_map: Dict[float, Any] = {}
    legend_handles: List[Any] = []
    legend_labels: List[str] = []
    for idx, entry_series in enumerate(series):
        frame = entry_series.frame
        temps = frame["temperature"]
        signal = frame["signal"]
        if len(temps) > MAX_PLOT_POINTS:
            temps = _downsample_series(temps, MAX_PLOT_POINTS)
            signal = signal.loc[temps.index]
        color_key = (entry_series.field, entry_series.direction, entry_series.segment_index)
        color = color_map.get(color_key, plt.rcParams["axes.prop_cycle"].by_key()["color"][idx % 10])
        label = f"{entry_series.field:.0f} Oe{processor._direction_label(entry_series.direction, entry_series.segment_index)}"
        primary = field_order[0] if field_order else entry_series.field
        secondary = field_order[1] if len(field_order) > 1 else None
        if entry_series.field not in axes_map:
            if entry_series.field == primary or not axes_map:
                axes_map[entry_series.field] = ax_left
            else:
                if ax_right is None:
                    ax_right = ax_left.twinx()
                axes_map[entry_series.field] = ax_right
        if secondary is not None and entry_series.field == secondary and ax_right is None:
            ax_right = ax_left.twinx()
            axes_map[entry_series.field] = ax_right
        axis = axes_map[entry_series.field]
        line = axis.plot(temps, signal, color=color, linewidth=1.4, label=label)[0]
        legend_handles.append(line)
        legend_labels.append(label)
    title = record.sample or Path(record.path).stem
    variant = getattr(record, "variant", None)
    if isinstance(variant, str) and variant.strip():
        title = f"{title} ({variant.strip()})"
    ax_left.set_title(f"{title} - VSM Temperature Scan")
    ax_left.set_xlabel("Temperature (°C)")
    ax_left.set_ylabel("Signal X (emu)")
    if ax_right is not None:
        ax_right.set_ylabel("Signal X (emu) (secondary)")
    if legend_handles:
        ax_left.legend(legend_handles, legend_labels, loc="best")
    figure.tight_layout()
    return figure


def _plot_dma_iso_stress_figure(
    record: DmaIsoStressRecord,
    *,
    width_px: int,
    height_px: int,
) -> Optional["plt.Figure"]:
    datasets = record.datasets or {}
    if not datasets:
        return None
    figsize = (max(width_px / 96.0, 1.0), max(height_px / 96.0, 1.0))
    figure = plt.Figure(figsize=figsize)
    ax = figure.add_subplot(111)
    for stress in sorted(datasets):
        temps, strains = datasets[stress]
        temps = list(_downsample_values(temps, MAX_PLOT_POINTS))
        strains = list(_downsample_values(strains, MAX_PLOT_POINTS))
        ax.plot(temps, strains, linewidth=1.4, label=f"{stress} MPa")
    title = record.sample or Path(record.path).stem
    variant = getattr(record, "variant", None)
    if isinstance(variant, str) and variant.strip():
        title = f"{title} ({variant.strip()})"
    ax.set_title(f"{title} - DMA Iso-Stress")
    ax.set_xlabel("Temperature (°C)")
    ax.set_ylabel("Strain (%)")
    ax.legend(loc="best")
    figure.tight_layout()
    return figure


def _plot_shape_memory_stress_strain_figure(
    record: ShapeMemoryStressStrainRecord,
    *,
    width_px: int,
    height_px: int,
) -> Optional["plt.Figure"]:
    if make_dual_axis_overlay_figure is None:
        return None
    frame = record.data if isinstance(record.data, pd.DataFrame) else pd.DataFrame()
    if frame.empty:
        return None
    title = record.sample or Path(record.path).stem
    variant = getattr(record, "variant", None)
    if isinstance(variant, str) and variant.strip():
        title = f"{title} ({variant.strip()})"
    figure = make_dual_axis_overlay_figure(frame, title=title)
    try:
        figure.set_size_inches(max(width_px / 96.0, 1.0), max(height_px / 96.0, 1.0))
    except Exception:
        pass
    return figure


def _plot_fmr_figure(
    record: FmrRecord,
    *,
    width_px: int,
    height_px: int,
) -> Optional["plt.Figure"]:
    if select_fmr_axes is None:
        return None
    frame = record.data if isinstance(record.data, pd.DataFrame) else pd.DataFrame()
    if frame.empty:
        return None
    columns = [str(col) for col in frame.columns]
    field_col, x_col, y_col = select_fmr_axes(columns)
    if not field_col or not x_col or not y_col:
        return None
    subset = frame[[field_col, x_col, y_col]].apply(pd.to_numeric, errors="coerce").dropna(how="any")
    if subset.empty:
        return None
    if len(subset) > MAX_PLOT_POINTS:
        field_series = _downsample_series(subset[field_col], MAX_PLOT_POINTS)
        subset = subset.loc[field_series.index]
    figsize = (max(width_px / 96.0, 1.0), max(height_px / 96.0, 1.0))
    figure = plt.Figure(figsize=figsize)
    ax = figure.add_subplot(111)
    ax.plot(subset[field_col], subset[x_col], color="#111111", linewidth=1.2, label="X")
    ax.plot(subset[field_col], subset[y_col], color="#dc2626", linewidth=1.2, label="Y")
    title = record.sample or Path(record.path).stem
    variant = getattr(record, "variant", None)
    if isinstance(variant, str) and variant.strip():
        title = f"{title} ({variant.strip()})"
    ax.set_title(title)
    units = getattr(record, "units", {}) if hasattr(record, "units") else {}
    x_unit = units.get(field_col) or "Oe"
    y_unit = units.get(x_col) or units.get(y_col) or "V"
    axis_field = "Field" if "field" in field_col.lower() else field_col
    ax.set_xlabel(f"{axis_field} [{x_unit}]" if x_unit else axis_field)
    ax.set_ylabel(f"X [{y_unit}]" if y_unit else "X")
    ax.legend(loc="best")
    figure.tight_layout()
    return figure


def _annealing_records_to_frame(
    records: List[MeasurementRecord],
    logger: logging.Logger,
) -> pd.DataFrame:
    columns = [
        "Composition",
        "Microwire",
        "Graph — 1000 mA",
        "Graph — low mA",
        ANNEALING_OTHER_GRAPH_COLUMN,
        "_group_key",
        "_sources",
    ]
    grouped: Dict[MicrowireKey, List[MeasurementRecord]] = {}
    for record in records:
        metadata = getattr(record, "metadata", None)
        if metadata is None:
            continue
        composition = getattr(metadata, "composition_token", None)
        draw = getattr(metadata, "draw_x", None)
        piece = getattr(metadata, "piece_y", None)
        if composition is None or draw is None or piece is None:
            continue
        suffix = None
        path = getattr(record, "path", None)
        if isinstance(path, Path):
            parsed_key = _microscope_key(path)
            if parsed_key is not None:
                _, _, _, suffix = parsed_key
        try:
            key = (str(composition), int(draw), int(piece), suffix)
        except (TypeError, ValueError):
            continue
        grouped.setdefault(key, []).append(record)

    rows: List[Dict[str, Any]] = []
    for (composition, draw, piece, suffix), group in sorted(
        grouped.items(),
        key=lambda item: (
            str(item[0][0]).lower(),
            int(item[0][1]),
            int(item[0][2]),
            (str(item[0][3]).lower() if item[0][3] is not None else ""),
        ),
    ):
        high_record, low_record = _select_high_low_pair(group)
        try:
            microwire = _microwire_label(draw, piece, suffix)
        except Exception:
            microwire = f"{draw}/{piece}"

        def _mtime(entry: MeasurementRecord) -> Optional[float]:
            path = getattr(entry, "path", None)
            if not path:
                return None
            try:
                return Path(path).stat().st_mtime
            except Exception:
                return None

        group_key = _microwire_key_to_str((composition, draw, piece, suffix))
        source_paths: List[str] = []
        for entry in (high_record, low_record):
            path = getattr(entry, "path", None)
            if path:
                source_paths.append(str(Path(path)))
        if source_paths:
            source_paths = list(dict.fromkeys(source_paths))
        rows.append(
            {
                "Composition": composition,
                "Microwire": microwire,
                "Graph — 1000 mA": None,
                "Graph — low mA": None,
                ANNEALING_OTHER_GRAPH_COLUMN: None,
                "_group_key": group_key,
                "_sources": source_paths,
            }
        )

    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _extract_setpoint(record: Optional[MeasurementRecord]) -> Optional[float]:
    if record is None:
        return None
    value = getattr(getattr(record, "metadata", object()), "setpoint_mA", None)
    if value is None:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _format_setpoint(value: Optional[float]) -> str:
    if value is None or not math.isfinite(value):
        return ""
    rounded = int(round(value))
    if abs(value - rounded) < 1e-6:
        return f"{rounded}"
    return f"{value:.3f}".rstrip("0").rstrip(".")


def _select_high_low_pair(
    records: List[MeasurementRecord],
) -> Tuple[Optional[MeasurementRecord], Optional[MeasurementRecord]]:
    high_record = _select_high_measurement(records)
    low_record = _select_low_measurement(records)
    if (
        high_record
        and low_record
        and _extract_setpoint(high_record) == _extract_setpoint(low_record)
    ):
        setpoints = {
            _extract_setpoint(record)
            for record in records
            if _extract_setpoint(record) is not None
        }
        if len(setpoints) <= 1:
            low_record = None
    return high_record, low_record


def _select_other_measurements(
    records: Sequence[MeasurementRecord],
    high_record: Optional[MeasurementRecord],
    low_record: Optional[MeasurementRecord],
) -> List[MeasurementRecord]:
    excluded_ids = {id(high_record), id(low_record)}

    def _sort_key(record: MeasurementRecord) -> Tuple[int, str]:
        setpoint = getattr(getattr(record, "metadata", object()), "setpoint_mA", None)
        try:
            setpoint_value = int(setpoint) if setpoint is not None else -1
        except (TypeError, ValueError):
            setpoint_value = -1
        file_name = getattr(getattr(record, "metadata", object()), "file_name", "")
        return (setpoint_value, str(file_name).lower())

    remaining = [record for record in records if id(record) not in excluded_ids]
    return sorted(remaining, key=_sort_key)


class _AnnealingPlotDisplay(QtWidgets.QWidget):
    valuePicked = QtCore.pyqtSignal(float)
    """Render a single annealing plot with contextual details."""

    def __init__(
        self,
        title: str,
        logger: logging.Logger,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._base_title = title
        self._logger = logger
        self._canvas: FigureCanvasQTAgg | None = None
        self._motion_cid: Optional[int] = None
        self._click_cid: Optional[int] = None
        self._cursor_units: str = "mA"

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        self.title_label = QtWidgets.QLabel(title)
        self.title_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        layout.addWidget(self.title_label)

        self.subtitle_label = QtWidgets.QLabel("")
        self.subtitle_label.setObjectName("annealingPlotSubtitle")
        self.subtitle_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        self.subtitle_label.setStyleSheet("color: palette(mid); font-size: 10px;")
        layout.addWidget(self.subtitle_label)

        self._stack = QtWidgets.QStackedLayout()
        self._stack.setContentsMargins(0, 0, 0, 0)
        layout.addLayout(self._stack, 1)

        self._placeholder = QtWidgets.QLabel(
            "Select a row to preview the annealing measurement."
        )
        self._placeholder.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setWordWrap(True)
        self._stack.addWidget(self._placeholder)

        self.cursor_label = QtWidgets.QLabel("Cursor: —")
        self.cursor_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        self.cursor_label.setStyleSheet("color: #f1f3f4; font-size: 11px;")
        layout.addWidget(self.cursor_label)

    def set_record(
        self,
        record: Optional[MeasurementRecord],
        *,
        setpoint: Optional[float],
        description: str,
    ) -> None:
        if record is None:
            self._show_placeholder(description)
            self.title_label.setText(self._base_title)
            return

        try:
            figure = self._build_figure(record)
        except Exception:
            self._logger.exception("Failed to render annealing preview for %s", getattr(record, "path", "?"))
            self._show_placeholder("Failed to render plot for the selected measurement.")
            self.title_label.setText(self._base_title)
            return

        display_title = self._base_title
        formatted_setpoint = _format_setpoint(setpoint)
        if formatted_setpoint:
            display_title = f"{self._base_title} ({formatted_setpoint} mA)"
        self.title_label.setText(display_title)

        details: List[str] = []
        if formatted_setpoint:
            details.append(f"{formatted_setpoint} mA")
        try:
            sample_count = len(record.dataframe.index)
        except Exception:
            sample_count = None
        if sample_count is not None:
            details.append(f"{sample_count} sample(s)")
        file_name = getattr(getattr(record, "metadata", object()), "file_name", "")
        if not file_name:
            path = getattr(record, "path", None)
            if path:
                try:
                    file_name = Path(path).name
                except Exception:
                    file_name = ""
        if file_name:
            details.append(str(file_name))
        self.subtitle_label.setText(" · ".join(details))

        canvas = FigureCanvasQTAgg(figure)
        if self._canvas is not None:
            self._disconnect_motion_handler()
            self._stack.removeWidget(self._canvas)
            self._canvas.setParent(None)
            self._canvas.deleteLater()
        self._stack.insertWidget(0, canvas)
        self._stack.setCurrentWidget(canvas)
        self._canvas = canvas
        try:
            canvas.setMouseTracking(True)
        except Exception:
            pass
        try:
            self._motion_cid = canvas.mpl_connect("motion_notify_event", self._handle_motion)
        except Exception:
            self._motion_cid = None
        try:
            self._click_cid = canvas.mpl_connect("button_press_event", self._handle_click)
        except Exception:
            self._click_cid = None
        self._cursor_units = "mA"
        self._update_cursor_label(None)

    def clear(self, message: str) -> None:
        self._show_placeholder(message)
        self.title_label.setText(self._base_title)

    def _show_placeholder(self, message: str) -> None:
        if self._canvas is not None:
            self._disconnect_motion_handler()
            self._disconnect_click_handler()
            self._stack.removeWidget(self._canvas)
            self._canvas.setParent(None)
            self._canvas.deleteLater()
            self._canvas = None
        self.subtitle_label.setText("")
        self._placeholder.setText(message)
        self._stack.setCurrentWidget(self._placeholder)
        self._update_cursor_label(None)

    def _disconnect_motion_handler(self) -> None:
        if self._canvas is not None and self._motion_cid is not None:
            try:
                self._canvas.mpl_disconnect(self._motion_cid)
            except Exception:
                pass
        self._motion_cid = None

    def _disconnect_click_handler(self) -> None:
        if self._canvas is not None and self._click_cid is not None:
            try:
                self._canvas.mpl_disconnect(self._click_cid)
            except Exception:
                pass
        self._click_cid = None

    def _handle_motion(self, event: Any) -> None:
        if event is None or event.inaxes is None or event.xdata is None:
            self._update_cursor_label(None)
            return
        try:
            value = float(event.xdata)
        except Exception:
            self._update_cursor_label(None)
            return
        self._update_cursor_label(value)

    def _handle_click(self, event: Any) -> None:
        if event is None or not getattr(event, "dblclick", False):
            return
        if event.xdata is None:
            return
        try:
            value = float(event.xdata)
        except Exception:
            return
        self._update_cursor_label(value)
        try:
            self.valuePicked.emit(value)
        except Exception:
            pass

    def _update_cursor_label(self, value: Optional[float]) -> None:
        if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            text = "Cursor: —"
        else:
            formatted = f"{float(value):.3f}".rstrip("0").rstrip(".")
            suffix = f" {self._cursor_units}" if self._cursor_units else ""
            text = f"Cursor: {formatted}{suffix}"
        self.cursor_label.setText(text)

    def _build_figure(self, record: MeasurementRecord):
        frame = record.dataframe if isinstance(record.dataframe, pd.DataFrame) else pd.DataFrame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            raise ValueError("No data to plot")
        if "I_mA" in frame.columns and "R_ohm" in frame.columns:
            plot_df = pd.DataFrame(
                {
                    "I_mA": pd.to_numeric(frame["I_mA"], errors="coerce"),
                    "R_Ohm": pd.to_numeric(frame["R_ohm"], errors="coerce"),
                }
            ).dropna()
        elif "I_A" in frame.columns and "R_ohm" in frame.columns:
            currents_A = pd.to_numeric(frame["I_A"], errors="coerce")
            max_abs = float(currents_A.abs().max(skipna=True) or 0.0)
            if max_abs <= 50:
                currents_mA = currents_A * 1e3
            else:
                currents_mA = currents_A
            plot_df = pd.DataFrame(
                {
                    "I_mA": currents_mA,
                    "R_Ohm": pd.to_numeric(frame["R_ohm"], errors="coerce"),
                }
            ).dropna()
        elif "I_mA" in frame.columns and "R_Ohm" in frame.columns:
            plot_df = pd.DataFrame(
                {
                    "I_mA": pd.to_numeric(frame["I_mA"], errors="coerce"),
                    "R_Ohm": pd.to_numeric(frame["R_Ohm"], errors="coerce"),
                }
            ).dropna()
        else:
            raise ValueError("Current annealing dataframe missing expected columns")
        if plot_df.empty:
            raise ValueError("No valid samples to plot")
        path = getattr(record, "path", None)
        stem = None
        if path:
            try:
                stem = Path(path).stem
            except Exception:
                stem = None
        if not stem:
            stem = getattr(getattr(record, "metadata", object()), "file_name", "")
        title = format_annealing_title(str(stem or "Current annealing"))
        target_width = max(int(ANNEALING_GRAPH_WIDTH * 2), ANNEALING_GRAPH_WIDTH)
        target_height = max(int(ANNEALING_GRAPH_HEIGHT * 2), ANNEALING_GRAPH_HEIGHT)
        figure, _ = plot_annealing_curve(
            plot_df,
            title,
            target_px=(target_width, target_height),
        )
        try:
            axes = figure.axes[0] if figure.axes else None
        except Exception:
            axes = None
        if axes is not None:
            try:
                axes.set_title("")
                axes.set_xlabel("")
                axes.set_ylabel("")
            except Exception:
                pass
            try:
                legend = axes.get_legend()
            except Exception:
                legend = None
            if legend is not None:
                try:
                    legend.remove()
                except Exception:
                    pass
            try:
                axes.tick_params(labelsize=8)
            except Exception:
                pass
        try:
            figure.subplots_adjust(left=0.08, right=0.98, top=0.98, bottom=0.12)
        except Exception:
            pass
        return figure


@dataclass
class _GraphPreviewAction:
    label: str
    callback: Callable[[], None]
    tooltip: Optional[str] = None


@dataclass
class _GraphPreviewItem:
    title: str
    pixmap: QtGui.QPixmap
    actions: Tuple[_GraphPreviewAction, ...] = ()


@dataclass(frozen=True)
class _ShapeMemoryPointSelection:
    index: int
    displacement_mm: float
    load_g: float
    strain_pct: float
    stress_mpa: float


class _GraphGalleryWidget(QtWidgets.QWidget):
    def __init__(
        self,
        placeholder: str,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self._placeholder = QtWidgets.QLabel(placeholder)
        self._placeholder.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignCenter
        )
        self._placeholder.setWordWrap(True)
        layout.addWidget(self._placeholder, 1)

        self._scroll = QtWidgets.QScrollArea(self)
        self._scroll.setWidgetResizable(True)
        self._scroll.setVisible(False)
        container = QtWidgets.QWidget(self._scroll)
        self._scroll.setWidget(container)
        self._items_layout = QtWidgets.QVBoxLayout(container)
        self._items_layout.setContentsMargins(0, 0, 0, 0)
        self._items_layout.setSpacing(10)
        layout.addWidget(self._scroll, 1)

    def clear(self, message: str) -> None:
        self._set_placeholder(message)

    def set_items(self, items: Sequence[_GraphPreviewItem], empty_message: str) -> None:
        if not items:
            self._set_placeholder(empty_message)
            return
        self._clear_items()
        for item in items:
            card = QtWidgets.QWidget(self._scroll)
            card_layout = QtWidgets.QVBoxLayout(card)
            card_layout.setContentsMargins(0, 0, 0, 0)
            card_layout.setSpacing(4)
            label = QtWidgets.QLabel(item.title, card)
            label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
            card_layout.addWidget(label)
            image = QtWidgets.QLabel(card)
            image.setPixmap(item.pixmap)
            image.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
            card_layout.addWidget(image)
            if item.actions:
                action_row = QtWidgets.QHBoxLayout()
                action_row.setContentsMargins(0, 0, 0, 0)
                action_row.setSpacing(6)
                for action in item.actions:
                    button = QtWidgets.QPushButton(action.label, card)
                    if action.tooltip:
                        button.setToolTip(action.tooltip)
                    button.clicked.connect(action.callback)
                    action_row.addWidget(button)
                action_row.addStretch(1)
                card_layout.addLayout(action_row)
            self._items_layout.addWidget(card)
        self._items_layout.addStretch(1)
        self._scroll.setVisible(True)
        self._placeholder.setVisible(False)

    def _clear_items(self) -> None:
        while self._items_layout.count():
            item = self._items_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.setParent(None)
                widget.deleteLater()

    def _set_placeholder(self, message: str) -> None:
        self._clear_items()
        self._placeholder.setText(message)
        self._placeholder.setVisible(True)
        self._scroll.setVisible(False)


class _ShapeMemoryPreviewPanel(QtWidgets.QWidget):
    pointPicked = QtCore.pyqtSignal(str, object)

    def __init__(
        self,
        logger: logging.Logger,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._logger = logger
        self._canvas_records: Dict[FigureCanvasQTAgg, ShapeMemoryStressStrainRecord] = {}
        self._canvas_connections: List[Tuple[FigureCanvasQTAgg, Optional[int], Optional[int]]] = []

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self.header_label = QtWidgets.QLabel("Select a row to preview shape-memory graphs.")
        self.header_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        layout.addWidget(self.header_label)

        self._stack = QtWidgets.QStackedLayout()
        self._placeholder = QtWidgets.QLabel(
            "Select a row to preview shape-memory graphs."
        )
        self._placeholder.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setWordWrap(True)
        self._tab_widget = QtWidgets.QTabWidget(self)
        self._stack.addWidget(self._placeholder)
        self._stack.addWidget(self._tab_widget)
        layout.addLayout(self._stack, 1)

        self.cursor_label = QtWidgets.QLabel(
            "Hover: displacement/load or strain/stress values."
        )
        self.cursor_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        layout.addWidget(self.cursor_label)

        target_row = QtWidgets.QHBoxLayout()
        target_row.setContentsMargins(0, 0, 0, 0)
        target_row.setSpacing(10)
        target_row.addWidget(QtWidgets.QLabel("Double-click target:"))
        self._target_buttons: Dict[str, QtWidgets.QRadioButton] = {}
        normal_button = QtWidgets.QRadioButton("Standard values")
        normal_button.setChecked(True)
        fracture_button = QtWidgets.QRadioButton("Fracture values")
        self._target_buttons["standard"] = normal_button
        self._target_buttons["fracture"] = fracture_button
        target_row.addWidget(normal_button)
        target_row.addWidget(fracture_button)
        target_row.addStretch(1)
        layout.addLayout(target_row)

        picked_row = QtWidgets.QHBoxLayout()
        picked_row.setContentsMargins(0, 0, 0, 0)
        picked_row.setSpacing(10)
        self._picked_labels: Dict[str, QtWidgets.QLabel] = {}
        for key, title in (
            ("displacement_mm", "Displacement"),
            ("load_g", "Load"),
            ("strain_pct", "Strain"),
            ("stress_mpa", "Stress"),
        ):
            picked_row.addWidget(QtWidgets.QLabel(f"{title}:"))
            label = QtWidgets.QLabel("unset")
            self._picked_labels[key] = label
            picked_row.addWidget(label)
        picked_row.addStretch(1)
        layout.addLayout(picked_row)

    def update_selection(
        self,
        title: str,
        records: Sequence[ShapeMemoryStressStrainRecord],
    ) -> None:
        current_index = self._tab_widget.currentIndex() if self._tab_widget.count() else 0
        self.header_label.setText(title or "Shape memory stress/strain")
        self._clear_tabs()
        self._update_picked_labels(None)
        if not records:
            self._placeholder.setText("No shape-memory graphs available for this microwire.")
            self._stack.setCurrentWidget(self._placeholder)
            return
        for record in records:
            figure = _plot_shape_memory_stress_strain_figure(
                record,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            if figure is None:
                continue
            canvas = FigureCanvasQTAgg(figure)
            try:
                canvas.setMouseTracking(True)
            except Exception:
                pass
            motion_cid = None
            click_cid = None
            try:
                motion_cid = canvas.mpl_connect("motion_notify_event", self._handle_motion)
            except Exception:
                motion_cid = None
            try:
                click_cid = canvas.mpl_connect("button_press_event", self._handle_click)
            except Exception:
                click_cid = None
            self._canvas_records[canvas] = record
            self._canvas_connections.append((canvas, motion_cid, click_cid))
            page = QtWidgets.QWidget(self._tab_widget)
            page_layout = QtWidgets.QVBoxLayout(page)
            page_layout.setContentsMargins(0, 0, 0, 0)
            page_layout.addWidget(canvas, 1)
            label = _record_label_for_display(record) or record.sample or "Shape memory"
            self._tab_widget.addTab(page, label)
        if self._tab_widget.count() == 0:
            self._placeholder.setText("No shape-memory graphs available for this microwire.")
            self._stack.setCurrentWidget(self._placeholder)
        else:
            if current_index >= 0:
                self._tab_widget.setCurrentIndex(min(current_index, self._tab_widget.count() - 1))
            self._stack.setCurrentWidget(self._tab_widget)

    def clear(self, message: str) -> None:
        self.header_label.setText("Shape memory stress/strain")
        self._clear_tabs()
        self._update_hover_label(None)
        self._update_picked_labels(None)
        self._placeholder.setText(message)
        self._stack.setCurrentWidget(self._placeholder)

    def _clear_tabs(self) -> None:
        for canvas, motion_cid, click_cid in self._canvas_connections:
            if motion_cid is not None:
                try:
                    canvas.mpl_disconnect(motion_cid)
                except Exception:
                    pass
            if click_cid is not None:
                try:
                    canvas.mpl_disconnect(click_cid)
                except Exception:
                    pass
        self._canvas_connections.clear()
        self._canvas_records.clear()
        while self._tab_widget.count():
            widget = self._tab_widget.widget(0)
            self._tab_widget.removeTab(0)
            if widget is not None:
                widget.setParent(None)
                widget.deleteLater()

    def _selection_from_event(self, event: Any) -> Optional[_ShapeMemoryPointSelection]:
        if event is None or getattr(event, "inaxes", None) is None or getattr(event, "canvas", None) is None:
            return None
        canvas = event.canvas
        if not isinstance(canvas, FigureCanvasQTAgg):
            return None
        record = self._canvas_records.get(canvas)
        if record is None:
            return None
        axis = event.inaxes
        if axis is None:
            return None
        axis_kind = "stress"
        try:
            axis_label = str(axis.get_xlabel() or "").strip().lower()
        except Exception:
            axis_label = ""
        if "displacement" in axis_label:
            axis_kind = "load"
        return _shape_memory_point_selection(
            record.data if isinstance(record.data, pd.DataFrame) else pd.DataFrame(),
            axis_kind=axis_kind,
            x_value=getattr(event, "xdata", None),
            y_value=getattr(event, "ydata", None),
        )

    def _handle_motion(self, event: Any) -> None:
        selection = self._selection_from_event(event)
        self._update_hover_label(selection)

    def _handle_click(self, event: Any) -> None:
        if event is None or not getattr(event, "dblclick", False):
            return
        selection = self._selection_from_event(event)
        self._update_hover_label(selection)
        self._update_picked_labels(selection)
        if selection is not None:
            try:
                self.pointPicked.emit(self._current_target(), selection)
            except Exception:
                pass

    def _current_target(self) -> str:
        fracture = self._target_buttons.get("fracture")
        if isinstance(fracture, QtWidgets.QRadioButton) and fracture.isChecked():
            return "fracture"
        return "standard"

    def _update_hover_label(self, selection: Optional[_ShapeMemoryPointSelection]) -> None:
        if selection is None:
            self.cursor_label.setText("Hover: displacement/load or strain/stress values.")
            return
        self.cursor_label.setText(
            "Hover: "
            f"{self._format_value(selection.displacement_mm, 'mm')} | "
            f"{self._format_value(selection.load_g, 'g')} | "
            f"{self._format_value(selection.strain_pct, '%')} | "
            f"{self._format_value(selection.stress_mpa, 'MPa')}"
        )

    def _update_picked_labels(self, selection: Optional[_ShapeMemoryPointSelection]) -> None:
        mapping = {
            "displacement_mm": (selection.displacement_mm if selection else None, "mm"),
            "load_g": (selection.load_g if selection else None, "g"),
            "strain_pct": (selection.strain_pct if selection else None, "%"),
            "stress_mpa": (selection.stress_mpa if selection else None, "MPa"),
        }
        for key, widget in self._picked_labels.items():
            value, units = mapping[key]
            widget.setText(self._format_value(value, units) if value is not None else "unset")

    @staticmethod
    def _format_value(value: float, units: str) -> str:
        formatted = f"{float(value):.3f}".rstrip("0").rstrip(".")
        return f"{formatted} {units}".strip()


class _GraphGalleryDialog(QtWidgets.QDialog):
    def __init__(
        self,
        title: str,
        items: Sequence[_GraphPreviewItem],
        parent: QtWidgets.QWidget | None = None,
        *,
        empty_message: str = "No graphs available.",
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(980, 720)
        layout = QtWidgets.QVBoxLayout(self)
        gallery = _GraphGalleryWidget(empty_message, self)
        gallery.set_items(items, empty_message)
        layout.addWidget(gallery, 1)
        button_box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Close
        )
        button_box.rejected.connect(self.reject)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)


_VSM_TEMP_PROCESSOR: VSMTemperatureScanProcessor | None = None
_OPEN_PYPLOT_WINDOWS: List[QtWidgets.QWidget] = []


def _open_pyplot_for_paths(
    paths: Sequence[Path],
    plotter_name: str,
    logger: logging.Logger,
    *,
    auto_plot: bool = False,
    open_origin: bool = False,
) -> None:
    filtered = [Path(path) for path in paths if isinstance(path, (Path, str))]
    filtered = [path for path in filtered if path.exists()]
    filtered = list(dict.fromkeys(filtered))
    if not filtered:
        QtWidgets.QMessageBox.information(
            None,
            "PyPlot",
            "No valid source files were found for this graph.",
        )
        return
    try:
        from plotting.pyplot.app import PyPlotWorkbench
    except Exception:
        logger.exception("Failed to import PyPlot workbench.")
        QtWidgets.QMessageBox.warning(
            None,
            "PyPlot",
            "PyPlot is unavailable in this environment.",
        )
        return
    app = QtWidgets.QApplication.instance()
    if app is not None:
        try:
            ensure_app_theme(app)
        except Exception:
            pass
    window = PyPlotWorkbench(initial_plotter=plotter_name)
    window.show()
    _OPEN_PYPLOT_WINDOWS.append(window)
    window.destroyed.connect(
        lambda *_args, ref=window: _OPEN_PYPLOT_WINDOWS.remove(ref)
        if ref in _OPEN_PYPLOT_WINDOWS
        else None
    )

    def _run() -> None:
        try:
            window._import_paths(filtered)
        except Exception:
            logger.exception("Failed to import files into PyPlot.")
            return
        try:
            window._commit_selected_paths(filtered)
        except Exception:
            pass

        def _load_and_plot(attempt: int = 0) -> None:
            plugin = getattr(window, "_current_plugin", None)
            if plugin is None:
                if attempt < 5:
                    QtCore.QTimer.singleShot(100, lambda: _load_and_plot(attempt + 1))
                return
            if hasattr(plugin, "load_data"):
                try:
                    plugin.load_data()
                except Exception:
                    logger.exception("Failed to load data in PyPlot.")
                    return
            measurements = getattr(window, "measurements", None)
            if isinstance(measurements, list) and not measurements:
                if attempt < 3:
                    QtCore.QTimer.singleShot(150, lambda: _load_and_plot(attempt + 1))
                return
            if auto_plot or open_origin:
                try:
                    plugin.generate()
                except Exception:
                    logger.exception("Failed to plot in PyPlot.")
            if open_origin:
                def _open_origin() -> None:
                    try:
                        plugin.open_origin()
                    except Exception:
                        logger.exception("Failed to open Origin from PyPlot.")

                QtCore.QTimer.singleShot(200, _open_origin)

        QtCore.QTimer.singleShot(0, _load_and_plot)

    QtCore.QTimer.singleShot(0, _run)


def _get_vsm_temp_processor(logger: logging.Logger) -> Optional[VSMTemperatureScanProcessor]:
    global _VSM_TEMP_PROCESSOR
    if _VSM_TEMP_PROCESSOR is not None:
        return _VSM_TEMP_PROCESSOR
    if VSMTemperatureScanProcessor is None:
        logger.warning("VSM temperature scan parser is not available.")
        return None
    try:
        _VSM_TEMP_PROCESSOR = VSMTemperatureScanProcessor()
    except Exception:
        logger.exception("Failed to initialize VSM temperature scan processor.")
        _VSM_TEMP_PROCESSOR = None
    return _VSM_TEMP_PROCESSOR


def _vsm_hysteresis_preview_items(
    records: Sequence[VsmHysteresisRecord],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> List[_GraphPreviewItem]:
    items: List[_GraphPreviewItem] = []
    groups = _group_vsm_hysteresis_plot_groups(records)
    for group in groups:
        figure = _plot_vsm_hysteresis_figure(
            group.records,
            logger,
            width_px=width_px,
            height_px=height_px,
        )
        pixmap = _figure_to_pixmap(figure, logger, width_px=width_px, height_px=height_px)
        if pixmap is None:
            continue
        title = group.label or _record_label_for_display(group.records[0]) or group.records[0].sample
        paths = [
            entry.path for entry in group.records if isinstance(entry.path, Path)
        ]
        actions = (
            _GraphPreviewAction(
                "Open in PyPlot",
                partial(
                    _open_pyplot_for_paths,
                    list(dict.fromkeys(paths)),
                    "VSM Hysteresis Loops",
                    logger,
                    auto_plot=True,
                    open_origin=False,
                ),
                tooltip="Open these VSM files in the PyPlot VSM Hysteresis Loops plugin.",
            ),
            _GraphPreviewAction(
                "Open in Origin",
                partial(
                    _open_pyplot_for_paths,
                    list(dict.fromkeys(paths)),
                    "VSM Hysteresis Loops",
                    logger,
                    auto_plot=True,
                    open_origin=True,
                ),
                tooltip="Send the selected VSM files to Origin via PyPlot.",
            ),
        )
        items.append(_GraphPreviewItem(title, pixmap, actions=actions))
    return items


def _vsm_temperature_preview_items(
    records: Sequence[VsmTemperatureScanRecord],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> List[_GraphPreviewItem]:
    processor = _get_vsm_temp_processor(logger)
    if processor is None:
        return []
    preview_width = max(int(width_px * 1.25), width_px)
    preview_height = max(int(height_px * 1.25), height_px)
    items: List[_GraphPreviewItem] = []
    for record in records:
        figure = _plot_vsm_temperature_scan_figure(
            record,
            processor,
            width_px=preview_width,
            height_px=preview_height,
        )
        pixmap = _figure_to_pixmap(
            figure,
            logger,
            width_px=preview_width,
            height_px=preview_height,
        )
        if pixmap is None:
            continue
        title = _record_label_for_display(record) or record.sample
        paths = [record.path] if isinstance(record.path, Path) else []
        actions = (
            _GraphPreviewAction(
                "Open in PyPlot",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "VSM Temperature Scan",
                    logger,
                    auto_plot=True,
                    open_origin=False,
                ),
                tooltip="Open this VSM scan in the PyPlot VSM Temperature Scan plugin.",
            ),
            _GraphPreviewAction(
                "Open in Origin",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "VSM Temperature Scan",
                    logger,
                    auto_plot=True,
                    open_origin=True,
                ),
                tooltip="Send this VSM scan to Origin via PyPlot.",
            ),
        )
        items.append(_GraphPreviewItem(title, pixmap, actions=actions))
    return items


def _dma_iso_stress_preview_items(
    records: Sequence[DmaIsoStressRecord],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> List[_GraphPreviewItem]:
    items: List[_GraphPreviewItem] = []
    for record in records:
        figure = _plot_dma_iso_stress_figure(
            record,
            width_px=width_px,
            height_px=height_px,
        )
        pixmap = _figure_to_pixmap(figure, logger, width_px=width_px, height_px=height_px)
        if pixmap is None:
            continue
        title = _record_label_for_display(record) or record.sample
        paths = [record.path] if isinstance(record.path, Path) else []
        actions = (
            _GraphPreviewAction(
                "Open in PyPlot",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "DMA Iso-Stress",
                    logger,
                    auto_plot=True,
                    open_origin=False,
                ),
                tooltip="Open this DMA iso-stress file in the PyPlot DMA Iso-Stress plugin.",
            ),
            _GraphPreviewAction(
                "Open in Origin",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "DMA Iso-Stress",
                    logger,
                    auto_plot=True,
                    open_origin=True,
                ),
                tooltip="Send this DMA iso-stress file to Origin via PyPlot.",
            ),
        )
        items.append(_GraphPreviewItem(title, pixmap, actions=actions))
    return items


def _shape_memory_stress_strain_preview_items(
    records: Sequence[ShapeMemoryStressStrainRecord],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> List[_GraphPreviewItem]:
    items: List[_GraphPreviewItem] = []
    for record in records:
        figure = _plot_shape_memory_stress_strain_figure(
            record,
            width_px=width_px,
            height_px=height_px,
        )
        pixmap = _figure_to_pixmap(figure, logger, width_px=width_px, height_px=height_px)
        if pixmap is None:
            continue
        title = _record_label_for_display(record) or record.sample
        paths = [record.path] if isinstance(record.path, Path) else []
        actions = (
            _GraphPreviewAction(
                "Open in PyPlot",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "Shape Memory Stress/Strain",
                    logger,
                    auto_plot=True,
                    open_origin=False,
                ),
                tooltip="Open this shape-memory file in the PyPlot Shape Memory Stress/Strain plugin.",
            ),
            _GraphPreviewAction(
                "Open in Origin",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "Shape Memory Stress/Strain",
                    logger,
                    auto_plot=True,
                    open_origin=True,
                ),
                tooltip="Send this shape-memory file to Origin via PyPlot.",
            ),
        )
        items.append(_GraphPreviewItem(title, pixmap, actions=actions))
    return items


def _shape_memory_point_selection(
    frame: pd.DataFrame,
    *,
    axis_kind: str,
    x_value: object,
    y_value: object | None = None,
) -> Optional[_ShapeMemoryPointSelection]:
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        return None
    try:
        x_numeric = float(x_value)
    except Exception:
        return None
    try:
        y_numeric = float(y_value) if y_value is not None else None
    except Exception:
        y_numeric = None

    points: list[tuple[float, int, float, float, float, float]] = []
    for idx, row in frame.iterrows():
        try:
            displacement = float(row["displacement_mm"])
            load = float(row["load_g"])
            strain = float(row["strain_pct"])
            stress = float(row["stress_mpa"])
        except Exception:
            continue
        values = (displacement, load) if axis_kind == "load" else (strain, stress)
        if not all(math.isfinite(value) for value in (displacement, load, strain, stress, *values)):
            continue
        distance = abs(values[0] - x_numeric)
        if y_numeric is not None and math.isfinite(y_numeric):
            distance += abs(values[1] - y_numeric)
        points.append((distance, int(idx), displacement, load, strain, stress))
    if not points:
        return None
    _, index, displacement, load, strain, stress = min(points, key=lambda item: item[0])
    return _ShapeMemoryPointSelection(
        index=index,
        displacement_mm=displacement,
        load_g=load,
        strain_pct=strain,
        stress_mpa=stress,
    )


def _fmr_preview_items(
    records: Sequence[FmrRecord],
    logger: logging.Logger,
    *,
    width_px: int,
    height_px: int,
) -> List[_GraphPreviewItem]:
    items: List[_GraphPreviewItem] = []
    for record in records:
        figure = _plot_fmr_figure(record, width_px=width_px, height_px=height_px)
        pixmap = _figure_to_pixmap(figure, logger, width_px=width_px, height_px=height_px)
        if pixmap is None:
            continue
        title = _record_label_for_display(record) or record.sample
        paths = [record.path] if isinstance(record.path, Path) else []
        actions = (
            _GraphPreviewAction(
                "Open in PyPlot",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "FMR",
                    logger,
                    auto_plot=True,
                    open_origin=False,
                ),
                tooltip="Open this FMR file in the PyPlot FMR plugin.",
            ),
            _GraphPreviewAction(
                "Open in Origin",
                partial(
                    _open_pyplot_for_paths,
                    paths,
                    "FMR",
                    logger,
                    auto_plot=True,
                    open_origin=True,
                ),
                tooltip="Send this FMR file to Origin via PyPlot.",
            ),
        )
        items.append(_GraphPreviewItem(title, pixmap, actions=actions))
    return items


class AnnealingPlotPanel(QtWidgets.QWidget):
    """Display paired annealing plots for the selected microwire."""

    def __init__(
        self,
        logger: logging.Logger,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.header_label = QtWidgets.QLabel("Select a row to preview annealing plots.")
        self.header_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        layout.addWidget(self.header_label)

        splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal, self)
        self._high_display = _AnnealingPlotDisplay("Graph — 1000 mA", logger, splitter)
        self._low_display = _AnnealingPlotDisplay("Graph — low mA", logger, splitter)
        splitter.addWidget(self._high_display)
        splitter.addWidget(self._low_display)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([1, 1])
        layout.addWidget(splitter, 1)

    def update_selection(
        self,
        key: Optional[MicrowireKey],
        high: Optional[MeasurementRecord],
        low: Optional[MeasurementRecord],
    ) -> None:
        if key is None:
            self.header_label.setText("Select a row to preview annealing plots.")
            self._high_display.clear("Select a row to view the 1000 mA measurement.")
            self._low_display.clear("Select a row to view the low-current measurement.")
            return

        composition, draw, piece, suffix = key
        try:
            microwire = _microwire_label(draw, piece, suffix)
        except Exception:
            microwire = f"{draw}/{piece}"
        self.header_label.setText(f"{composition} — {microwire}")

        self._high_display.set_record(
            high,
            setpoint=_extract_setpoint(high),
            description="No 1000 mA measurement available for this microwire.",
        )
        self._low_display.set_record(
            low,
            setpoint=_extract_setpoint(low),
            description="No lower-current measurement available for this microwire.",
        )



class _TransitionPickerDialog(QtWidgets.QDialog):
    def __init__(
        self,
        parent: QtWidgets.QWidget,
        record_pairs: List[Tuple[str, MeasurementRecord]],
        existing: Dict[str, float],
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Interactive annealing plot")
        self.resize(960, 640)
        self._points: Dict[str, Optional[float]] = {
            "As": float(existing.get("As")) if isinstance(existing.get("As"), (int, float)) else None,
            "Ms": float(existing.get("Ms")) if isinstance(existing.get("Ms"), (int, float)) else None,
        }
        self._plots: Dict[str, Dict[str, Any]] = {}
        self._lines: Dict[str, Dict[str, Optional[Any]]] = {}
        self.cursor_value_label = QtWidgets.QLabel("—")
        self.cursor_value_label.setMinimumWidth(80)

        layout = QtWidgets.QVBoxLayout(self)
        self._tab_widget = QtWidgets.QTabWidget(self)
        layout.addWidget(self._tab_widget, 1)

        for label, record in record_pairs:
            frame = record.dataframe if isinstance(record.dataframe, pd.DataFrame) else pd.DataFrame()
            metadata = getattr(record, "metadata", None)
            if metadata is not None:
                title = format_annealing_title(metadata)
            else:
                path_attr = getattr(record, "path", None)
                title = Path(path_attr).name if path_attr else "Current annealing"
            figure, _ = plot_annealing_curve(frame, title, target_px=(1280, 640))
            canvas = FigureCanvasQTAgg(figure)
            toolbar = NavigationToolbar2QT(canvas, self)
            page = QtWidgets.QWidget(self)
            page_layout = QtWidgets.QVBoxLayout(page)
            page_layout.setContentsMargins(0, 0, 0, 0)
            page_layout.addWidget(toolbar)
            page_layout.addWidget(canvas, 1)
            self._tab_widget.addTab(page, label)
            axes = figure.axes[0] if figure.axes else figure.add_subplot(111)
            cid = canvas.mpl_connect("button_press_event", lambda event, lbl=label: self._handle_click(lbl, event))
            motion_cid = canvas.mpl_connect("motion_notify_event", lambda event, lbl=label: self._handle_motion(lbl, event))
            self._plots[label] = {"canvas": canvas, "figure": figure, "axes": axes, "cid": cid, "motion_cid": motion_cid}
            self._lines[label] = {"As": None, "Ms": None}

        controls = QtWidgets.QHBoxLayout()
        controls.addWidget(QtWidgets.QLabel("Cursor:"))
        controls.addWidget(self.cursor_value_label)
        controls.addSpacing(20)
        self.as_radio = QtWidgets.QRadioButton("Set As (mA)")
        self.ms_radio = QtWidgets.QRadioButton("Set Ms (mA)")
        self.as_radio.setChecked(True)
        controls.addWidget(self.as_radio)
        controls.addWidget(self.ms_radio)
        controls.addSpacing(20)
        self.as_value_label = QtWidgets.QLabel(self._format_value(self._points["As"]))
        self.ms_value_label = QtWidgets.QLabel(self._format_value(self._points["Ms"]))
        controls.addWidget(QtWidgets.QLabel("As:"))
        controls.addWidget(self.as_value_label)
        clear_as = QtWidgets.QPushButton("Clear As")
        clear_as.clicked.connect(lambda: self._clear_point("As"))
        controls.addWidget(clear_as)
        controls.addSpacing(10)
        controls.addWidget(QtWidgets.QLabel("Ms:"))
        controls.addWidget(self.ms_value_label)
        clear_ms = QtWidgets.QPushButton("Clear Ms")
        clear_ms.clicked.connect(lambda: self._clear_point("Ms"))
        controls.addWidget(clear_ms)
        controls.addStretch(1)
        layout.addLayout(controls)

        button_box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel,
            parent=self,
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self._refresh_lines()

    @staticmethod
    def _format_value(value: Optional[float]) -> str:
        if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            return "unset"
        return f"{float(value):.3f}".rstrip("0").rstrip(".")

    def _clear_point(self, label: str) -> None:
        self._points[label] = None
        self._refresh_lines()

    def _handle_click(self, label: str, event: Any) -> None:
        if event is None or event.inaxes is None or event.xdata is None:
            return
        if event.button != 1:
            return
        target = "As" if self.as_radio.isChecked() else "Ms"
        self._points[target] = float(event.xdata)
        self._refresh_lines()

    def _handle_motion(self, label: str, event: Any) -> None:
        if event is None or event.inaxes is None or event.xdata is None:
            self.cursor_value_label.setText("—")
            return
        self.cursor_value_label.setText(f"{float(event.xdata):.3f} mA")

    def _refresh_lines(self) -> None:
        for label, plot in self._plots.items():
            axes = plot["axes"]
            canvas = plot["canvas"]
            for point_name, color in (("As", "tab:green"), ("Ms", "tab:purple")):
                value = self._points.get(point_name)
                line = self._lines[label].get(point_name)
                if value is None:
                    if line is not None:
                        line.remove()
                        self._lines[label][point_name] = None
                    continue
                if line is None:
                    line = axes.axvline(float(value), color=color, linestyle="--", linewidth=1.2)
                    self._lines[label][point_name] = line
                else:
                    line.set_xdata([float(value), float(value)])
            canvas.draw_idle()
        self.as_value_label.setText(self._format_value(self._points.get("As")))
        self.ms_value_label.setText(self._format_value(self._points.get("Ms")))

    def result_points(self) -> Dict[str, Optional[float]]:
        return dict(self._points)

class _SectionWorker(QtCore.QObject):
    progress = QtCore.pyqtSignal(int, int, object)
    finished = QtCore.pyqtSignal(object)
    failed = QtCore.pyqtSignal(object)
    cancelled = QtCore.pyqtSignal()

    def __init__(self, section: "MiniDatabaseSection", paths: Iterable[Path]) -> None:
        super().__init__()
        self._section = section
        self._paths = [Path(p) for p in paths]

    def _emit_progress(self, current: int, total: int, message: Optional[str]) -> None:
        try:
            payload = str(message) if message is not None else None
        except Exception:
            payload = None
        try:
            self.progress.emit(int(current), int(total), payload)
        except Exception:
            pass

    @QtCore.pyqtSlot()
    def run(self) -> None:  # pragma: no cover - runs in worker thread
        try:
            result = self._section.process(self._paths, progress=self._emit_progress)
        except BuildCancelledError:
            try:
                self.cancelled.emit()
            except Exception:
                pass
        except Exception as exc:  # pragma: no cover - defensive
            try:
                self.failed.emit(exc)
            except Exception:
                pass
        else:
            try:
                self.finished.emit(result)
            except Exception:
                pass


class _PendingScanWorker(QtCore.QObject):
    """Background worker that counts pending files for a section."""

    finished = QtCore.pyqtSignal(int)
    failed = QtCore.pyqtSignal(object)

    def __init__(
        self,
        sources: list[str],
        processed: dict[str, float],
        supported_suffixes: tuple[str, ...],
        recursive_search: bool,
    ) -> None:
        super().__init__()
        self._sources = [str(source) for source in sources]
        self._processed = dict(processed)
        self._supported_suffixes = tuple(supported_suffixes)
        self._recursive_search = bool(recursive_search)

    @QtCore.pyqtSlot()
    def run(self) -> None:  # pragma: no cover - runs in worker thread
        try:
            count = self._compute_pending_count()
        except Exception as exc:
            try:
                self.failed.emit(exc)
            except Exception:
                pass
        else:
            try:
                self.finished.emit(int(count))
            except Exception:
                pass

    def _compute_pending_count(self) -> int:
        candidates: Dict[str, Path] = {}
        suffixes = {s.lower() for s in self._supported_suffixes}
        thread = QtCore.QThread.currentThread()
        for source in self._sources:
            if thread is not None and thread.isInterruptionRequested():
                return 0
            root = Path(source).expanduser()
            if not root.exists():
                continue
            try:
                iterator: Iterable[Path] = (
                    root.rglob("*") if self._recursive_search else root.glob("*")
                )
            except Exception:
                continue
            for path in iterator:
                if thread is not None and thread.isInterruptionRequested():
                    return 0
                if not path.is_file():
                    continue
                if suffixes and path.suffix.lower() not in suffixes:
                    continue
                try:
                    resolved = str(path.resolve())
                except Exception:
                    resolved = str(path)
                candidates.setdefault(resolved, path)
        pending_count = 0
        processed = self._processed
        for path in candidates.values():
            if thread is not None and thread.isInterruptionRequested():
                return pending_count
            key = str(path)
            try:
                mtime = path.stat().st_mtime
            except OSError:
                continue
            if float(processed.get(key, -1.0)) != float(mtime):
                pending_count += 1
        return pending_count


def _microscope_index_to_frame(
    index: Dict[MicrowireKey, MicroscopeMeasurements],
    overrides: Dict[str, Dict[str, float]],
) -> pd.DataFrame:
    columns = MICROSCOPE_TABLE_COLUMNS.copy()

    def _image_path(entries: Sequence[MicroscopeDetection]) -> Optional[str]:
        for detection in entries:
            crop = getattr(detection, "crop_path", None)
            if crop:
                return str(crop)
            source = getattr(detection, "image_path", None)
            if source:
                return str(source)
        return None

    rows: List[Dict[str, Any]] = []
    for (composition, draw, piece, suffix), measurements in sorted(
        index.items(),
        key=lambda item: (
            str(item[0][0]).lower(),
            int(item[0][1]),
            int(item[0][2]),
            (str(item[0][3]).lower() if item[0][3] is not None else ""),
        ),
    ):
        key = _microwire_key_to_str((composition, draw, piece, suffix))
        override = overrides.get(key, {})
        d_value = override.get("d")
        if d_value is None:
            d_value = measurements.best_core()
        D_value = override.get("D")
        if D_value is None:
            D_value = measurements.best_glass()
        ratio = None
        if isinstance(d_value, (int, float)) and isinstance(D_value, (int, float)) and D_value:
            try:
                ratio = float(d_value) / float(D_value)
            except ZeroDivisionError:
                ratio = None
        if isinstance(ratio, (int, float)):
            ratio = round(float(ratio), 3)

        core_image = _image_path(measurements.core)
        glass_image = _image_path(measurements.glass)
        image_paths: List[str] = []
        for bucket in (measurements.core, measurements.glass, measurements.other):
            for detection in bucket:
                path = getattr(detection, "image_path", None)
                if path:
                    image_paths.append(str(path))
        if image_paths:
            image_paths = list(dict.fromkeys(image_paths))

        rows.append(
            {
                "Composition": composition,
                "Microwire": _microwire_label(draw, piece, suffix),
                MICROSCOPE_D_COLUMN: d_value,
                MICROSCOPE_CAP_D_COLUMN: D_value,
                "d/D": ratio,
                MICROSCOPE_IMAGE_COLUMNS[0]: None,
                MICROSCOPE_IMAGE_COLUMNS[1]: None,
                "_key": key,
                "_core_image": core_image,
                "_glass_image": glass_image,
                "_images": image_paths,
            }
        )
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _video_index_to_frame(
    index: Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary],
    fabrication_frame: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    def _is_missing(value: object) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        try:
            return bool(pd.isna(value))
        except Exception:
            return False

    base_columns: List[str] = []
    if isinstance(fabrication_frame, pd.DataFrame) and not fabrication_frame.empty:
        for column in fabrication_frame.columns:
            text = str(column)
            if text.startswith("_"):
                continue
            if text not in base_columns:
                base_columns.append(text)
    if not base_columns:
        base_columns = [
            "Composition",
            "Data source",
            "e/a",
            ESTIMATED_TRANSITION_COLUMN,
            "Draw",
            "Piece",
            "Length (m)",
            "Piece date",
            MICROSCOPE_D_COLUMN,
            MICROSCOPE_CAP_D_COLUMN,
            "d/D",
            "Resistance (Ω)",
            CORE_TEMPERATURE_COLUMN,
            GLASS_TEMPERATURE_COLUMN,
            "Mass (g)",
            "Winding speed (m/min)",
            "Glass feeding (mm/min)",
            "Underpressure",
            GLASS_PULL_COLUMN,
            "Notes",
            "Production datetime",
        ]
    if "Microwire" not in base_columns:
        base_columns.insert(1, "Microwire")
    if "Draw" not in base_columns:
        base_columns.append("Draw")
    if "Piece" not in base_columns:
        base_columns.append("Piece")
    if VIDEO_END_LENGTH_COLUMN not in base_columns:
        base_columns.append(VIDEO_END_LENGTH_COLUMN)
    if VIDEO_MW_LENGTH_COLUMN not in base_columns:
        base_columns.append(VIDEO_MW_LENGTH_COLUMN)
    columns = list(base_columns) + ["_sources", "_group_key"]
    fabrication_lookup: Dict[Tuple[str, int, int], pd.Series] = {}
    if isinstance(fabrication_frame, pd.DataFrame) and not fabrication_frame.empty:
        for _, row in fabrication_frame.iterrows():
            composition = str(row.get("Composition") or "").strip()
            if not composition or composition == "Imported data:":
                continue
            try:
                draw = int(row.get("Draw"))
                piece = int(row.get("Piece"))
            except (TypeError, ValueError):
                continue
            fabrication_lookup[(composition, draw, piece)] = row
    rows: List[Dict[str, Any]] = []
    seen_keys: Set[Tuple[str, int, int]] = set()
    for (composition, draw, piece), summary in sorted(index.items()):
        row: Dict[str, Any] = {column: None for column in columns}
        row["Composition"] = composition
        row["Draw"] = draw
        row["Piece"] = piece
        if piece is None:
            row["Microwire"] = f"{draw}/?"
        else:
            row["Microwire"] = _microwire_label(draw, piece, None)
        ea_value = _compute_ea_from_composition(composition)
        row["e/a"] = ea_value
        row[ESTIMATED_TRANSITION_COLUMN] = _estimate_transition_temp_c(ea_value)
        fabrication_row = None
        if piece is not None:
            fabrication_row = fabrication_lookup.get((composition, draw, piece))
        if fabrication_row is not None:
            for column in base_columns:
                if column in {"Composition", "Draw", "Piece", "Microwire"}:
                    continue
                candidate = fabrication_row.get(column)
                if not _is_missing(candidate) and _is_missing(row.get(column)):
                    row[column] = candidate
        if piece is not None:
            seen_keys.add((composition, draw, piece))
        if row.get(ESTIMATED_TRANSITION_COLUMN) in (None, ""):
            row[ESTIMATED_TRANSITION_COLUMN] = _estimate_transition_temp_c(row.get("e/a"))
        if row.get(CORE_TEMPERATURE_COLUMN) in (None, ""):
            row[CORE_TEMPERATURE_COLUMN] = summary.temperature()
        if row.get("Underpressure") in (None, ""):
            row["Underpressure"] = summary.underpressure()
        if row.get("Winding speed (m/min)") in (None, ""):
            row["Winding speed (m/min)"] = summary.winding_speed()
        if row.get("Glass feeding (mm/min)") in (None, ""):
            row["Glass feeding (mm/min)"] = summary.glass_feed()
        row["_sources"] = sorted(str(path) for path in getattr(summary, "sources", set()))
        if piece is None:
            row["_group_key"] = ""
        else:
            row["_group_key"] = _microwire_key_to_str((composition, draw, piece, None))
        rows.append(row)
    for (composition, draw, piece), fabrication_row in sorted(fabrication_lookup.items()):
        if (composition, draw, piece) in seen_keys:
            continue
        row: Dict[str, Any] = {column: None for column in columns}
        row["Composition"] = composition
        row["Draw"] = draw
        row["Piece"] = piece
        row["Microwire"] = _microwire_label(draw, piece, None)
        ea_value = fabrication_row.get("e/a")
        if _is_missing(ea_value):
            ea_value = _compute_ea_from_composition(composition)
        row["e/a"] = ea_value
        row[ESTIMATED_TRANSITION_COLUMN] = _estimate_transition_temp_c(ea_value)
        for column in base_columns:
            if column in {"Composition", "Draw", "Piece", "Microwire"}:
                continue
            candidate = fabrication_row.get(column)
            if not _is_missing(candidate):
                row[column] = candidate
        if row.get(ESTIMATED_TRANSITION_COLUMN) in (None, ""):
            row[ESTIMATED_TRANSITION_COLUMN] = _estimate_transition_temp_c(row.get("e/a"))
        row["_sources"] = []
        row["_group_key"] = _microwire_key_to_str((composition, draw, piece, None))
        rows.append(row)
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _strain_records_to_frame(records: Dict[MicrowireKey, StrainRecord]) -> pd.DataFrame:
    columns = [
        "Composition",
        "Draw",
        "Piece",
        "Microwire",
        "Strain (%)",
        "Broke",
    ]
    rows: List[Dict[str, Any]] = []
    for (composition, draw, piece, suffix), record in sorted(records.items()):
        rows.append(
            {
                "Composition": composition,
                "Draw": draw,
                "Piece": piece,
                "Microwire": record.microwire_label or _microwire_label(draw, piece, suffix),
                "Strain (%)": record.percent,
                "Broke": bool(record.broke),
            }
        )
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _apply_microscope_overrides(
    index: Dict[MicrowireKey, MicroscopeMeasurements],
    overrides: Dict[str, Dict[str, float]],
) -> Dict[MicrowireKey, MicroscopeMeasurements]:
    result: Dict[MicrowireKey, MicroscopeMeasurements] = {}
    for key, measurements in index.items():
        clone = MicroscopeMeasurements(
            core=list(measurements.core),
            glass=list(measurements.glass),
            other=list(measurements.other),
        )
        token = _microwire_key_to_text(key) or ""
        override = overrides.get(token, {})
        if not override:
            parts = _split_microwire_key(key)
            if parts is not None:
                composition, draw, piece, _suffix = parts
                base_token = _microwire_key_to_str((composition, draw, piece, None))
                override = overrides.get(base_token, {})
        d_value = override.get("d")
        if isinstance(d_value, (int, float)) and d_value > 0:
            detection = MicroscopeDetection(
                value=float(d_value),
                image_path=None,
                source="manual",
            )
            detection.category = "core"
            clone.core.insert(0, detection)
        D_value = override.get("D")
        if isinstance(D_value, (int, float)) and D_value > 0:
            detection = MicroscopeDetection(
                value=float(D_value),
                image_path=None,
                source="manual",
            )
            detection.category = "glass"
            clone.glass.insert(0, detection)
        result[key] = clone
    return result


def _sample_from_path(path: Path, sources: Sequence[str]) -> str:
    for source in sources:
        root = Path(source).expanduser()
        try:
            rel = path.resolve().relative_to(root.resolve())
        except Exception:
            continue
        parts = rel.parts
        if len(parts) >= 2:
            if parts[0].lower() in {"sweep", "sweeps"}:
                return parts[1]
            return parts[0]
        if len(parts) == 1:
            return path.stem
    parent = path.parent.name
    return parent if parent else path.stem


_SAMPLE_VARIANT_RE = re.compile(
    r"^\s*(?P<comp>[A-Za-z][A-Za-z0-9]*)\s+(?P<draw>\d{1,3})\s*[-_/]\s*(?P<piece>\d{1,3})(?P<suffix>[A-Za-z][A-Za-z0-9]*)?(?P<rest_sep>[\s\-_/.]+(?P<rest>.+))?$"
)


MicrowireKey = Tuple[str, int, int, Optional[str]]


def _microwire_key_to_text(value: object) -> Optional[str]:
    parts = _split_microwire_key(value)
    if parts is None:
        return None
    return _microwire_key_to_str(parts)


def _microwire_parts_from_label_safe(
    label: str,
) -> Optional[Tuple[int, int, Optional[str]]]:
    parts = _microwire_parts_from_label(label)
    if parts is not None:
        return parts
    fallback = _microwire_tuple_from_label(label)
    if fallback is not None:
        return fallback[0], fallback[1], None
    return None


def _split_sample_variant(sample: str) -> Tuple[str, Optional[str]]:
    text = str(sample or "").strip()
    if not text:
        return "", None
    match = _SAMPLE_VARIANT_RE.match(text)
    if not match:
        return text, None
    comp = match.group("comp").strip()
    draw = match.group("draw").strip()
    piece = match.group("piece").strip()
    suffix = match.group("suffix") or ""
    suffix = suffix.strip()
    base = f"{comp} {draw}-{piece}{suffix}"
    rest = (match.group("rest") or "").strip()
    if rest:
        rest = rest.strip(" -_/")
    return base, rest or None


def _microwire_key_from_path(path: Path, sample: str) -> Optional[MicrowireKey]:
    key = _microscope_key(path)
    if key is None and sample:
        try:
            key = _microscope_key(Path(sample))
        except Exception:
            key = None
    return key


def _microwire_info_from_key(
    key: Optional[MicrowireKey],
) -> Tuple[Optional[str], Optional[str]]:
    parts = _split_microwire_key(key) if key is not None else None
    if parts is None:
        return None, None
    composition, draw, piece, suffix = parts
    if not composition:
        return None, None
    try:
        microwire = _microwire_label(int(draw), int(piece), suffix)
    except Exception:
        microwire = f"{draw}/{piece}"
    return str(composition), microwire


def _record_label_for_display(record: object) -> str:
    label = getattr(record, "label", None)
    if isinstance(label, str) and label.strip():
        return label.strip()
    path = getattr(record, "path", None)
    if isinstance(path, Path):
        return path.stem
    if isinstance(path, str):
        return Path(path).stem
    sample = getattr(record, "sample", None)
    return str(sample).strip() if sample else ""


def _record_path_key(record: object) -> Optional[str]:
    path = getattr(record, "path", None)
    if isinstance(path, Path):
        return str(path)
    if isinstance(path, str) and path:
        return path
    return None


def _visibility_items_from_records(records: Sequence[object]) -> List[Tuple[str, str]]:
    items: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    for record in records:
        path = _record_path_key(record)
        if not path or path in seen:
            continue
        seen.add(path)
        label = _record_label_for_display(record)
        sample = getattr(record, "sample", None)
        if isinstance(sample, str) and sample.strip():
            if label and label != sample:
                label = f"{sample} — {label}"
            else:
                label = sample
        if not label:
            label = Path(path).name
        items.append((label, path))
    return items


def _visibility_groups_from_records(
    records: Sequence[object],
) -> Dict[str, List[Tuple[str, str]]]:
    groups: Dict[str, List[Tuple[str, str]]] = {}
    seen: Set[str] = set()
    for record in records:
        path = _record_path_key(record)
        if not path or path in seen:
            continue
        seen.add(path)
        label = _record_label_for_display(record)
        sample = getattr(record, "sample", None)
        group_label = label
        if isinstance(sample, str) and sample.strip():
            if group_label:
                if sample not in group_label:
                    group_label = f"{sample} — {group_label}"
            else:
                group_label = sample
        if not group_label:
            group_label = Path(path).name
        item_label = Path(path).name
        groups.setdefault(group_label, []).append((item_label, path))
    return groups


def _hidden_paths_from_section(section: object) -> Set[str]:
    extra = getattr(getattr(section, "data", None), "extra", None)
    if isinstance(extra, Mapping):
        hidden = extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            return {str(path) for path in hidden if path}
    return set()


def _record_key_for_group(record: object) -> Optional[MicrowireKey]:
    key = getattr(record, "key", None)
    key_parts = _split_microwire_key(key)
    if key_parts is not None:
        return key_parts
    return None


def _row_to_microwire_key(row: pd.Series) -> Optional[str]:
    composition = str(row.get("Composition") or "").strip()
    microwire = str(row.get("Microwire") or "").strip()
    if not composition or not microwire:
        return None
    parsed = _microwire_parts_from_label_safe(microwire)
    if parsed is None:
        return None
    draw, piece, suffix = parsed
    try:
        return _microwire_key_to_str((composition, int(draw), int(piece), suffix))
    except (TypeError, ValueError):
        return None


def _row_sample_value(row: pd.Series) -> Optional[str]:
    for key in ("Sample", "_sample", "sample"):
        value = row.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return _row_to_microwire_key(row)


def _group_graph_records_by_key(records: Sequence[object]) -> Dict[str, List[object]]:
    grouped: Dict[str, List[object]] = {}
    for record in records:
        key = _record_key_for_group(record)
        if key is None:
            sample = getattr(record, "sample", None)
            if isinstance(sample, str) and sample.strip():
                try:
                    key = _microscope_key(Path(sample))
                except Exception:
                    key = None
        if key is None:
            continue
        group_key = _microwire_key_to_text(key)
        if not group_key:
            continue
        grouped.setdefault(group_key, []).append(record)
    return grouped


def _graph_records_to_frame(
    records: Sequence[object],
    graph_column: str,
    *,
    sample_column: str = "Sample",
) -> pd.DataFrame:
    columns = []
    if sample_column:
        columns.append(sample_column)
    columns.extend(
        [
            "Composition",
            "Microwire",
            graph_column,
            "_group_key",
            "_sources",
        ]
    )
    if not records:
        return pd.DataFrame(columns=columns)
    grouped: Dict[str, List[object]] = {}
    for record in records:
        sample = getattr(record, "sample", None)
        if not isinstance(sample, str) or not sample.strip():
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                sample = path.stem
            elif isinstance(path, str):
                sample = Path(path).stem
            else:
                sample = "sample"
        grouped.setdefault(sample, []).append(record)
    rows: List[Dict[str, Any]] = []
    for sample, group in sorted(grouped.items()):
        key: Optional[MicrowireKey] = None
        for record in group:
            key = _record_key_for_group(record)
            if key is not None:
                break
        composition, microwire = _microwire_info_from_key(key)
        labels: List[str] = []
        sources: List[str] = []
        for record in group:
            label = _record_label_for_display(record)
            if label and label not in labels:
                labels.append(label)
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                sources.append(str(path))
            elif isinstance(path, str):
                sources.append(path)
        group_key = _microwire_key_to_text(key) if key is not None else ""
        rows.append(
            {
                sample_column: sample if sample_column else None,
                "Composition": composition,
                "Microwire": microwire,
                graph_column: labels,
                "_group_key": group_key,
                "_sources": list(dict.fromkeys(sources)),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _drop_visible_sample_column(section: "MiniDatabaseSection") -> None:
    frame = section.model.frame() if hasattr(section, "model") else None
    if not isinstance(frame, pd.DataFrame):
        return
    sample_columns: List[str] = []
    for column in frame.columns:
        normalized = str(column).strip().lower()
        if normalized in {"sample", "_sample"}:
            sample_columns.append(str(column))
    if not sample_columns:
        return
    cleaned = frame.copy()
    if "_sample" not in cleaned.columns and sample_columns:
        cleaned["_sample"] = cleaned[sample_columns[0]]
    cleaned = cleaned.drop(columns=sample_columns)
    try:
        section.data.table = cleaned
    except Exception:
        pass
    try:
        section.model.set_frame(cleaned)
    except Exception:
        pass
    try:
        section.store.save(section.data)
    except Exception:
        pass



class MiniDatabaseSection(QtWidgets.QWidget):
    """Base widget for mini-database sections that process a subset of data."""

    section_key = "base"
    section_title = "Base"
    supported_suffixes: tuple[str, ...] = ()
    recursive_search = True

    status_changed = QtCore.pyqtSignal(str)
    sources_changed = QtCore.pyqtSignal(list)
    data_updated = QtCore.pyqtSignal()
    log_emitted = QtCore.pyqtSignal(int, str)
    _processing_owner: ClassVar[Optional["MiniDatabaseSection"]] = None
    _refresh_queue: ClassVar[List["MiniDatabaseSection"]] = []
    _SCROLL_SINGLE_STEP = 12

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.logger = logger
        self._log_callback = log_callback
        self.store = MiniDatabaseStore(self.section_key)
        self.data = self.store.load()
        self.model = DataFrameModel(self.data.table)
        self._search_proxy = _TableSearchProxyModel(self)
        self.table_view: QtWidgets.QTableView | None = None
        self._table_splitter: QtWidgets.QSplitter | None = None
        self.search_edit: QtWidgets.QLineEdit | None = None
        self.search_clear_button: QtWidgets.QPushButton | None = None
        self._cancel_requested = False
        self.log_emitted.connect(
            self._dispatch_log,
            QtCore.Qt.ConnectionType.QueuedConnection,
        )
        self._worker_thread: QtCore.QThread | None = None
        self._worker: Optional["_SectionWorker"] = None
        self._active_candidates: List[Path] = []
        self._pending_count_cache: int | None = None
        self._pending_scan_in_progress = False
        self._pending_scan_generation = 0
        self._pending_scan_thread: QtCore.QThread | None = None
        self._pending_scan_worker: Optional[_PendingScanWorker] = None
        self._progress_dialog: QtWidgets.QProgressDialog | None = None

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)
        self.main_layout = layout

        controls = QtWidgets.QHBoxLayout()
        self.source_button = QtWidgets.QPushButton()
        self.source_button.setText("Connect folder…")
        self.source_button.clicked.connect(self._toggle_source)
        controls.addWidget(self.source_button)

        self.open_sources_button = QtWidgets.QPushButton("Open source file(s)")
        self.open_sources_button.setEnabled(False)
        self.open_sources_button.clicked.connect(self._open_selected_sources)
        controls.addWidget(self.open_sources_button)

        controls.addStretch(1)

        self.refresh_button = QtWidgets.QPushButton("Refresh")
        self.refresh_button.clicked.connect(self.refresh)
        controls.addWidget(self.refresh_button)

        self.stop_button = QtWidgets.QPushButton("Stop")
        self.stop_button.setEnabled(False)
        self.stop_button.clicked.connect(self._request_cancel)
        controls.addWidget(self.stop_button)

        layout.addLayout(controls)
        self.controls_layout = controls

        search_row = QtWidgets.QHBoxLayout()
        search_row.setContentsMargins(0, 0, 0, 0)
        search_row.setSpacing(6)
        search_row.addWidget(QtWidgets.QLabel("Search:"))
        self.search_edit = QtWidgets.QLineEdit(self)
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.setPlaceholderText("Filter rows across visible columns")
        self.search_edit.textChanged.connect(self._handle_search_changed)
        search_row.addWidget(self.search_edit, 1)
        self.search_clear_button = QtWidgets.QPushButton("Clear")
        self.search_clear_button.setEnabled(False)
        self.search_clear_button.clicked.connect(self._clear_search)
        search_row.addWidget(self.search_clear_button)
        layout.addLayout(search_row)

        self.status_label = QtWidgets.QLabel()
        layout.addWidget(self.status_label)

        progress_row = QtWidgets.QHBoxLayout()
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        progress_row.addWidget(self.progress_bar, 1)
        self.progress_label = QtWidgets.QLabel("Idle")
        progress_row.addWidget(self.progress_label)
        self.progress_eta_label = QtWidgets.QLabel("")
        self.progress_eta_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter)
        progress_row.addWidget(self.progress_eta_label)
        layout.addLayout(progress_row)
        self.progress_bar.hide()
        self.progress_label.hide()
        self.progress_eta_label.hide()
        self._progress_total: int = 0
        self._progress_current: int = 0
        self._progress_start: float | None = None

        self.sources_list = QtWidgets.QListWidget(self)
        self.sources_list.hide()

        right_panel = self.create_right_panel(self)
        layout.addWidget(right_panel, 1)
        if isinstance(self.table_view, QtWidgets.QTableView):
            self._search_proxy.setSourceModel(self.model)
            self._search_proxy.set_row_predicate(self._row_visible)
            self.table_view.setModel(self._search_proxy)
        self._configure_table_view()

        self._populate_sources_list()
        self.model.set_frame(self.data.table)
        self._auto_fit_columns()
        self._update_status()
        self._reset_progress_ui()
        self._hook_table_selection()
        self._update_open_sources_enabled()
        try:
            if hasattr(self, "_sanitize_graph_columns"):
                self._sanitize_graph_columns()
        except Exception:
            pass

    # ------------------------------------------------------------------ UI helpers
    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        table = QtWidgets.QTableView(parent)
        table.setModel(self.model)
        table.horizontalHeader().setStretchLastSection(True)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        table.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        table.setSortingEnabled(True)
        self.table_view = table
        container = QtWidgets.QWidget(parent)
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(table, 1)
        return container

    def _configure_table_view(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        table.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Expanding,
            QtWidgets.QSizePolicy.Policy.Expanding,
        )
        try:
            table.setIconSize(
                QtCore.QSize(ANNEALING_GRAPH_WIDTH, ANNEALING_GRAPH_HEIGHT)
            )
        except Exception:
            pass
        header = table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
        table.setVerticalScrollMode(
            QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        table.setHorizontalScrollMode(
            QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        vertical_bar = table.verticalScrollBar()
        if vertical_bar is not None:
            vertical_bar.setSingleStep(self._SCROLL_SINGLE_STEP)

    def _close_active_editor(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        editor = table.focusWidget()
        if editor is None or editor is table:
            return
        try:
            table.closeEditor(
                editor,
                QtWidgets.QAbstractItemDelegate.EndEditHint.NoHint,
            )
        except Exception:
            pass

    def _auto_fit_columns(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        try:
            table.resizeColumnsToContents()
        except Exception:
            return
        model = getattr(table, "model", lambda: None)()
        frame: Optional[pd.DataFrame] = None
        if hasattr(model, "frame"):
            try:
                frame = model.frame()
            except Exception:
                frame = None
        if frame is None or getattr(frame, "empty", False):
            return
        try:
            icon_width = table.iconSize().width()
        except Exception:
            icon_width = 0
        if icon_width <= 0:
            icon_width = ANNEALING_GRAPH_WIDTH
        graph_width = max(int(icon_width), ANNEALING_GRAPH_WIDTH) + 80
        for idx, column_name in enumerate(frame.columns):
            label = str(column_name)
            label_lower = label.lower()
            if (
                "graph" not in label_lower
                and "figure" not in label_lower
                and label not in MICROSCOPE_IMAGE_COLUMNS
            ):
                continue
            current = table.columnWidth(idx)
            target = graph_width if graph_width > current else current
            if target > 0:
                table.setColumnWidth(idx, target)

        header = table.horizontalHeader()
        v_header = table.verticalHeader()
        total_width = 0
        try:
            total_width = header.length()
        except Exception:
            total_width = 0
        if v_header is not None:
            total_width += v_header.width()
        total_width += table.frameWidth() * 2
        screen_rect = None
        try:
            window = self.window()
            if isinstance(window, QtWidgets.QWidget):
                screen = QtGui.QGuiApplication.screenAt(
                    window.mapToGlobal(window.rect().center())
                )
                if screen is None:
                    screen = QtGui.QGuiApplication.primaryScreen()
                if screen is not None:
                    screen_rect = screen.availableGeometry()
        except Exception:
            screen_rect = None
        try:
            table.setSizePolicy(
                QtWidgets.QSizePolicy.Policy.Expanding,
                QtWidgets.QSizePolicy.Policy.Expanding,
            )
        except Exception:
            pass
        if screen_rect is not None:
            max_width = max(640, screen_rect.width() - 200)
        else:
            max_width = 1600
        desired = total_width if total_width > 0 else table.sizeHint().width()
        constrained = max(240, min(desired, max_width))
        table.setMinimumWidth(0)
        table.setMaximumWidth(16777215)
        splitter = self._table_splitter
        if isinstance(splitter, QtWidgets.QSplitter):
            available = screen_rect.width() if screen_rect is not None else constrained * 2
            left = max(320, int(available * 0.7))
            right = max(240, available - left)
            splitter.setSizes([left, right])

    def _update_source_button(self) -> None:
        has_sources = self.sources_list.count() > 0
        text = "Remove folder…" if has_sources else "Connect folder…"
        self.source_button.setText(text)
        if has_sources:
            self.source_button.setToolTip("Disconnect the currently linked folder.")
        else:
            self.source_button.setToolTip("Select a folder to analyse.")

    def has_project_data(self) -> bool:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        return bool(self.data.sources) or bool(self.data.processed) or bool(self.data.extra) or not frame.empty

    def apply_data(self, data: MiniDatabaseData) -> None:
        self.data = data
        try:
            self.store.save(self.data)
        except Exception:
            pass
        self._invalidate_pending_cache()
        self._close_active_editor()
        self.model.set_frame(self.data.table)
        self._populate_sources_list()
        self._auto_fit_columns()
        self._update_status()
        self._reset_progress_ui()
        self._update_open_sources_enabled()
        try:
            self.sources_changed.emit(list(self.data.sources))
        except Exception:
            pass
        try:
            self.status_changed.emit(self.status_label.text())
        except Exception:
            pass

    def _populate_sources_list(self) -> None:
        self.sources_list.clear()
        for source in self.data.sources:
            self.sources_list.addItem(source)
        self._update_source_button()

    def _reset_progress_ui(self) -> None:
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_label.setText("Idle")
        self.progress_eta_label.clear()
        self._progress_total = 0
        self._progress_current = 0
        self._progress_start = None
        self._cancel_requested = False
        self.stop_button.setEnabled(False)

    def _hide_columns(self, names: Sequence[str]) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        model = self.model
        frame: Optional[pd.DataFrame] = None
        if hasattr(model, "frame"):
            try:
                frame = model.frame()
            except Exception:
                frame = None
        if frame is None:
            return
        columns = list(frame.columns)
        normalized_targets = {str(name).strip().lower() for name in names}
        for index, column in enumerate(columns):
            normalized = str(column).strip().lower()
            try:
                self.table_view.setColumnHidden(index, normalized in normalized_targets)
            except Exception:
                continue

    def _hook_table_selection(self) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        selection_model = self.table_view.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(
                lambda *_: self._update_open_sources_enabled()
            )
        model = self.table_view.model()
        if isinstance(model, QtCore.QAbstractItemModel):
            model.modelReset.connect(self._update_open_sources_enabled)
            model.dataChanged.connect(lambda *_: self._update_open_sources_enabled())

    def _selected_rows(self) -> List[int]:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return []
        selection = self.table_view.selectionModel()
        if selection is None:
            return []
        rows: Set[int] = set()
        for index in selection.selectedRows():
            if not index.isValid():
                continue
            source_row = self._search_proxy.map_row_to_source(index.row())
            if source_row is not None:
                rows.add(source_row)
        return sorted(rows)

    def _row_series(self, row: int) -> Optional[pd.Series]:
        frame = self.model.frame()
        if row < 0 or row >= len(frame.index):
            return None
        try:
            return frame.iloc[row]
        except Exception:
            return None

    @staticmethod
    def _path_key(path: Path) -> str:
        try:
            return str(path.resolve())
        except Exception:
            return str(path)

    def _row_sources(self, row: pd.Series) -> List[Path]:
        _ = row
        return []

    def _row_visible(self, row: pd.Series) -> bool:
        _ = row
        return True

    def _handle_search_changed(self, text: str) -> None:
        self._search_proxy.set_search_text(text)
        if isinstance(self.search_clear_button, QtWidgets.QPushButton):
            self.search_clear_button.setEnabled(bool(str(text).strip()))
        self._update_open_sources_enabled()

    def _clear_search(self) -> None:
        if isinstance(self.search_edit, QtWidgets.QLineEdit):
            self.search_edit.clear()

    def _update_open_sources_enabled(self) -> None:
        if not hasattr(self, "open_sources_button"):
            return
        rows = self._selected_rows()
        enabled = False
        if rows:
            for row_index in rows:
                series = self._row_series(row_index)
                if series is None:
                    continue
                if self._row_sources(series):
                    enabled = True
                    break
        self.open_sources_button.setEnabled(enabled)

    def _open_file(self, path: Path) -> bool:
        resolved = path.expanduser()
        if not resolved.exists():
            self.log(
                f"{self.section_title}: source file missing — {resolved}",
                level=logging.WARNING,
            )
            return False
        url = QtCore.QUrl.fromLocalFile(str(resolved))
        opened = QtGui.QDesktopServices.openUrl(url)
        if not opened:
            self.log(
                f"{self.section_title}: could not open {resolved}",
                level=logging.WARNING,
            )
        return opened

    def _open_selected_sources(self) -> None:
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their source files.",
            )
            return
        opened_any = False
        missing: List[Path] = []
        seen: set[Path] = set()
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            for path in self._row_sources(series):
                if path in seen:
                    continue
                seen.add(path)
                if self._open_file(path):
                    opened_any = True
                else:
                    missing.append(path)
        if not opened_any and missing:
            details = "\n".join(str(p) for p in missing[:5])
            if len(missing) > 5:
                details += "\n…"
            QtWidgets.QMessageBox.warning(
                self,
                self.section_title,
                f"None of the selected rows have available source files.\n\n{details}",
            )
    def _start_progress(self, total: int) -> None:
        self._progress_total = max(int(total), 0)
        self._progress_current = 0
        self._progress_start = time.monotonic()
        if self._progress_dialog is not None:
            try:
                self._progress_dialog.close()
            except Exception:
                pass
        if self._progress_total <= 0:
            dialog = QtWidgets.QProgressDialog("Scanning files...", None, 0, 0, self)
            dialog.setWindowTitle(self.section_title)
            dialog.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
            dialog.setCancelButton(None)
            dialog.setMinimumDuration(0)
            dialog.setAutoClose(False)
            dialog.setAutoReset(False)
            dialog.show()
            self._progress_dialog = dialog
            return
        dialog = QtWidgets.QProgressDialog(
            f"Processing 0/{self._progress_total}",
            None,
            0,
            100,
            self,
        )
        dialog.setWindowTitle(self.section_title)
        dialog.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
        dialog.setCancelButton(None)
        dialog.setMinimumDuration(0)
        dialog.setAutoClose(False)
        dialog.setAutoReset(False)
        dialog.setValue(0)
        dialog.show()
        self._progress_dialog = dialog

    def _update_progress(self, current: int, total: Optional[int], message: Optional[str]) -> None:
        if total is not None:
            self._progress_total = max(int(total), 0)
        self._progress_current = max(int(current), 0)
        total_units = self._progress_total
        dialog = self._progress_dialog
        if total_units <= 0:
            if dialog is not None:
                dialog.setRange(0, 0)
        else:
            percent = int(round(min(self._progress_current / total_units, 1.0) * 100)) if total_units else 0
            if dialog is not None:
                dialog.setRange(0, 100)
                dialog.setValue(max(0, min(100, percent)))
        parts: list[str] = []
        if message:
            parts.append(message)
        if total_units > 0:
            parts.append(f"{self._progress_current}/{total_units}")
        else:
            parts.append(f"{self._progress_current} processed")
        label_text = " - ".join(parts)
        start = self._progress_start
        eta_text = ""
        if (
            start is not None
            and self._progress_current > 0
            and total_units > 0
            and self._progress_current < total_units
        ):
            elapsed = time.monotonic() - start
            rate = elapsed / self._progress_current if self._progress_current else None
            if rate and math.isfinite(rate):
                eta = rate * (total_units - self._progress_current)
                eta_text = f"ETA {_format_duration(eta)}"
        elif start is not None and self._progress_current and (
            total_units == 0 or self._progress_current >= total_units
        ):
            elapsed = time.monotonic() - start
            eta_text = f"Elapsed {_format_duration(elapsed)}"
        if dialog is not None:
            dialog.setLabelText("\n".join(part for part in (label_text, eta_text) if part))
        QtWidgets.QApplication.processEvents(
            QtCore.QEventLoop.ProcessEventsFlag.AllEvents
        )

    def _request_cancel(self) -> None:
        if self._cancel_requested:
            return
        self._cancel_requested = True
        self.stop_button.setEnabled(False)
        if self._progress_dialog is not None:
            self._progress_dialog.setRange(0, 0)
            self._progress_dialog.setLabelText("Cancelling...")
        QtWidgets.QApplication.processEvents(
            QtCore.QEventLoop.ProcessEventsFlag.AllEvents
        )

    def is_cancelled(self) -> bool:
        return self._cancel_requested

    def _check_cancelled(self) -> None:
        if self._cancel_requested:
            raise BuildCancelledError()

    def _finish_progress(self) -> None:
        if self._progress_dialog is not None:
            try:
                self._progress_dialog.setRange(0, 100)
                self._progress_dialog.setValue(100)
                text = (
                    f"Complete - {self._progress_total} file(s)"
                    if self._progress_total > 0
                    else "Complete"
                )
                start = self._progress_start
                if start is not None:
                    elapsed = time.monotonic() - start
                    text = f"{text}\nElapsed {_format_duration(elapsed)}"
                self._progress_dialog.setLabelText(text)
                self._progress_dialog.close()
            except Exception:
                pass
        self._progress_dialog = None
        self._progress_start = None
        self._cancel_requested = False
        self.stop_button.setEnabled(False)
        self._release_processing()

    def _fail_progress(self) -> None:
        if self._progress_dialog is not None:
            try:
                self._progress_dialog.setRange(0, 100)
                self._progress_dialog.setValue(0)
                self._progress_dialog.setLabelText("Failed")
                self._progress_dialog.close()
            except Exception:
                pass
        self._progress_dialog = None
        self._progress_start = None
        self._cancel_requested = False
        self.stop_button.setEnabled(False)
        self._release_processing()

    def _cancel_progress(self) -> None:
        if self._progress_dialog is not None:
            try:
                self._progress_dialog.setRange(0, 100)
                self._progress_dialog.setValue(0)
                self._progress_dialog.setLabelText("Cancelled")
                self._progress_dialog.close()
            except Exception:
                pass
        self._progress_dialog = None
        self._progress_start = None
        self._cancel_requested = False
        self.stop_button.setEnabled(False)
        self._release_processing()

    def _release_processing(self) -> None:
        if MiniDatabaseSection._processing_owner is self:
            MiniDatabaseSection._processing_owner = None
            if MiniDatabaseSection._refresh_queue:
                next_section = MiniDatabaseSection._refresh_queue.pop(0)
                if isinstance(next_section, MiniDatabaseSection):
                    QtCore.QTimer.singleShot(0, next_section.refresh)

    def _progress_callback(
        self, current: int, total: int, message: Optional[str] = None
    ) -> None:
        if self._progress_start is None:
            self._start_progress(total)
        self._update_progress(current, total, message)

    def _sync_sources(self) -> None:
        sources: list[str] = []
        for index in range(self.sources_list.count()):
            item = self.sources_list.item(index)
            if item is not None:
                sources.append(item.text())
        self.data.sources = sources
        self.store.save(self.data)
        self._invalidate_pending_cache()
        try:
            self.sources_changed.emit(list(sources))
        except Exception:
            pass
        self._update_status()
        self._update_source_button()

    def set_sources(self, sources: Iterable[str]) -> None:
        unique = []
        seen: Set[str] = set()
        for source in sources:
            normalised = str(Path(source).expanduser())
            if normalised not in seen:
                seen.add(normalised)
                unique.append(normalised)
        self.sources_list.clear()
        for path in unique:
            self.sources_list.addItem(path)
        self._sync_sources()

    def _toggle_source(self) -> None:
        if self.sources_list.count() == 0:
            self._add_source()
        else:
            self._prompt_remove_source()

    def _add_source(self) -> None:
        start_dir = _dialog_start_directory()
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            self.section_title,
            str(start_dir),
        )
        if not directory:
            return
        normalised = str(Path(directory).expanduser())
        existing = {self.sources_list.item(idx).text() for idx in range(self.sources_list.count())}
        if normalised not in existing:
            self.sources_list.addItem(normalised)
            self._sync_sources()

    def _prompt_remove_source(self) -> None:
        sources = [self.sources_list.item(idx).text() for idx in range(self.sources_list.count())]
        if not sources:
            return
        target = sources[0]
        if len(sources) > 1:
            selection, ok = QtWidgets.QInputDialog.getItem(
                self,
                self.section_title,
                "Select a folder to disconnect:",
                sources,
                0,
                False,
            )
            if not ok or not selection:
                return
            target = selection
        message = QtWidgets.QMessageBox(self)
        message.setWindowTitle(self.section_title)
        message.setIcon(QtWidgets.QMessageBox.Icon.Warning)
        message.setText("Remove the connected folder?")
        message.setInformativeText(target)
        remove_btn = message.addButton("Remove folder", QtWidgets.QMessageBox.ButtonRole.AcceptRole)
        message.addButton(QtWidgets.QMessageBox.StandardButton.Cancel)
        message.exec()
        if message.clickedButton() is not remove_btn:
            return
        for idx in range(self.sources_list.count() - 1, -1, -1):
            item = self.sources_list.item(idx)
            if item is not None and item.text() == target:
                self.sources_list.takeItem(idx)
        self._sync_sources()

    # ------------------------------------------------------------------ data handling
    def _collect_candidates(self) -> List[Path]:
        candidates: Dict[str, Path] = {}
        for source in self.data.sources:
            root = Path(source).expanduser()
            if not root.exists():
                continue
            iterator: Iterable[Path]
            try:
                iterator = root.rglob("*") if self.recursive_search else root.glob("*")
            except Exception:
                continue
            for path in iterator:
                if not path.is_file():
                    continue
                if self.supported_suffixes and path.suffix.lower() not in self.supported_suffixes:
                    continue
                try:
                    resolved = str(path.resolve())
                except Exception:
                    resolved = str(path)
                candidates.setdefault(resolved, path)
        return sorted(candidates.values())

    def _pending_paths(self) -> List[Path]:
        pending: List[Path] = []
        processed = self.data.processed
        for path in self._collect_candidates():
            key = str(path)
            try:
                mtime = path.stat().st_mtime
            except OSError:
                continue
            if float(processed.get(key, -1.0)) != float(mtime):
                pending.append(path)
        return pending

    def _invalidate_pending_cache(self) -> None:
        self._stop_pending_scan(wait_ms=200)
        self._pending_count_cache = None
        self._pending_scan_in_progress = False
        self._pending_scan_generation += 1

    def _stop_pending_scan(self, *, wait_ms: int = 500) -> None:
        thread = self._pending_scan_thread
        self._pending_scan_thread = None
        self._pending_scan_worker = None
        self._pending_scan_in_progress = False
        if thread is None:
            return
        try:
            thread.requestInterruption()
        except Exception:
            pass
        try:
            thread.quit()
        except Exception:
            pass
        try:
            if thread.isRunning():
                thread.wait(max(0, int(wait_ms)))
        except Exception:
            pass
        try:
            if thread.isRunning():
                thread.terminate()
                thread.wait(200)
        except Exception:
            pass

    def _request_pending_scan(self) -> None:
        if self._pending_scan_in_progress:
            return
        if not self.data.sources:
            self._pending_count_cache = 0
            return
        self._stop_pending_scan(wait_ms=100)
        self._pending_scan_generation += 1
        generation = self._pending_scan_generation
        worker = _PendingScanWorker(
            list(self.data.sources),
            dict(self.data.processed),
            tuple(self.supported_suffixes),
            bool(self.recursive_search),
        )
        thread = QtCore.QThread(self)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(partial(self._handle_pending_scan_finished, generation, thread, worker))
        worker.failed.connect(partial(self._handle_pending_scan_failed, generation, thread, worker))
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        self._pending_scan_thread = thread
        self._pending_scan_worker = worker
        self._pending_scan_in_progress = True
        thread.start()

    def _handle_pending_scan_finished(
        self,
        generation: int,
        thread: QtCore.QThread,
        worker: _PendingScanWorker,
        count: int,
    ) -> None:
        if generation != self._pending_scan_generation:
            return
        self._pending_scan_in_progress = False
        self._pending_count_cache = int(count)
        self._pending_scan_thread = None
        self._pending_scan_worker = None
        self._update_status()

    def _handle_pending_scan_failed(
        self,
        generation: int,
        thread: QtCore.QThread,
        worker: _PendingScanWorker,
        exc: object,
    ) -> None:
        if generation != self._pending_scan_generation:
            return
        self._pending_scan_in_progress = False
        self._pending_scan_thread = None
        self._pending_scan_worker = None
        self.logger.debug("%s pending scan failed", self.section_key, exc_info=exc)
        self._update_status()

    def _update_status(self) -> None:
        sources_count = len(self.data.sources)
        pending_count = self._pending_count_cache
        if sources_count == 0:
            message = "Connect one or more folders to begin."
            self.refresh_button.setEnabled(False)
            self._pending_count_cache = 0
        else:
            self.refresh_button.setEnabled(True)
            if pending_count is None:
                message = "Scanning for new or updated files…"
                if not self._pending_scan_in_progress:
                    self._request_pending_scan()
            elif pending_count:
                message = f"⚠️ {pending_count} new or updated file(s) pending processing."
            elif not self.data.table.empty:
                message = f"Up to date ({len(self.data.table)} record(s))."
            else:
                message = "No processed data available yet."
        self.status_label.setText(message)
        try:
            self.status_changed.emit(message)
        except Exception:
            pass

    def _dispatch_log(self, level: int, message: str) -> None:
        try:
            self._log_callback(level, message)
        except Exception:
            self.logger.log(level, message)

    def log(self, message: str, level: int = logging.INFO) -> None:
        text = str(message)
        try:
            self.log_emitted.emit(int(level), text)
        except Exception:
            self._dispatch_log(level, text)

    def reset_to_blank(self) -> None:
        """Clear all processed data and disconnect sources for a fresh start."""

        self.data = MiniDatabaseData()
        self.model.set_frame(pd.DataFrame())
        self.sources_list.clear()
        self._sync_sources()
        self.store.clear_table()
        self.store.save(self.data)
        self._auto_fit_columns()
        self._update_status()
        self._reset_progress_ui()
        self._update_open_sources_enabled()
        try:
            self.data_updated.emit()
        except Exception:
            pass
        try:
            if hasattr(self, "_sanitize_graph_columns"):
                self._sanitize_graph_columns()
        except Exception:
            pass

    def export_project_payload(self) -> Dict[str, Any]:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        columns = [str(col) for col in getattr(frame, "columns", [])]
        rows: List[Dict[str, Any]] = []
        if isinstance(frame, pd.DataFrame) and not frame.empty:
            for record in frame.to_dict(orient="records"):
                payload: Dict[str, Any] = {}
                for column in columns:
                    payload[column] = _json_safe(record.get(column))
                rows.append(payload)
        index_payload: List[Any] = []
        if isinstance(frame, pd.DataFrame) and not frame.empty:
            for entry in frame.index.tolist():
                index_payload.append(_json_safe(entry))
        extra_payload = _json_safe(self.data.extra)
        if not isinstance(extra_payload, (dict, list, tuple, str, int, float, bool)) and extra_payload is not None:
            extra_payload = str(extra_payload)
        return {
            "section": self.section_key,
            "title": self.section_title,
            "columns": columns,
            "rows": rows,
            "index": index_payload,
            "extra": extra_payload,
            "sources": list(self.data.sources),
            "processed": dict(self.data.processed),
        }

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:
        """Restore section state from a project payload."""

        if not isinstance(payload, Mapping):
            self.reset_to_blank()
            return

        columns_payload = payload.get("columns")
        if isinstance(columns_payload, (list, tuple)):
            column_names = [str(column) for column in columns_payload]
        else:
            column_names = []
        rows_payload = payload.get("rows")
        if isinstance(rows_payload, (list, tuple)):
            frame = pd.DataFrame(list(rows_payload), columns=column_names or None)
        else:
            frame = pd.DataFrame(columns=column_names)

        index_payload = payload.get("index")
        if isinstance(index_payload, list) and len(index_payload) == len(frame.index):
            try:
                frame.index = pd.Index(index_payload)
            except Exception:
                pass

        extra_payload = payload.get("extra")
        extra: Dict[str, Any]
        if isinstance(extra_payload, Mapping):
            extra = dict(extra_payload)
        else:
            extra = {}

        sources_payload = payload.get("sources")
        sources = (
            [str(source) for source in sources_payload]
            if isinstance(sources_payload, (list, tuple))
            else []
        )

        processed_payload = payload.get("processed")
        processed: Dict[str, float] = {}
        if isinstance(processed_payload, Mapping):
            for key, value in processed_payload.items():
                try:
                    processed[str(key)] = float(value)
                except (TypeError, ValueError):
                    continue

        self.data = MiniDatabaseData(sources=sources, processed=processed, table=frame, extra=extra)
        self.model.set_frame(frame)
        self.store.save(self.data)
        self._populate_sources_list()
        self._auto_fit_columns()
        self._update_status()
        self._update_open_sources_enabled()
        self._reset_progress_ui()
        try:
            self.sources_changed.emit(list(sources))
        except Exception:
            pass
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _new_project(self) -> None:
        if self._dirty:
            box = QtWidgets.QMessageBox(self)
            box.setWindowTitle("Unsaved project")
            box.setText("Save changes to this Microwire Data Builder project before starting a new one?")
            save_btn = box.addButton(QtWidgets.QMessageBox.StandardButton.Save)
            discard_btn = box.addButton("Discard", QtWidgets.QMessageBox.ButtonRole.DestructiveRole)
            cancel_btn = box.addButton(QtWidgets.QMessageBox.StandardButton.Cancel)
            box.setDefaultButton(save_btn)
            box.exec()
            clicked = box.clickedButton()
            if clicked is cancel_btn:
                return
            if clicked is save_btn:
                self._save_project()
                if self._dirty:
                    return
        self._suppress_dirty = True
        for section in self.sections.values():
            if isinstance(section, MiniDatabaseSection):
                section.reset_to_blank()
        self._project_path = None
        self._dirty = False
        self._suppress_dirty = False
        self._update_project_title()
        self._update_project_actions()
        self._refresh_sections_after_project_load()

    # Compatibility alias used by menu wiring in some launch contexts
    def new_project(self) -> None:
        self._new_project()

    def has_project_data(self) -> bool:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if isinstance(frame, pd.DataFrame) and not frame.empty:
            return True
        return bool(self.data.extra)

    def refresh(self) -> None:
        candidates = self._collect_candidates()
        if not candidates:
            self.log(f"{self.section_title}: no files found in connected folders.")
            self.data.processed = {}
            self.data.table = pd.DataFrame()
            self.store.save(self.data)
            self.model.set_frame(self.data.table)
            self._auto_fit_columns()
            self._update_status()
            self._reset_progress_ui()
            try:
                self.data_updated.emit()
            except Exception:
                pass
            return
        self._cancel_requested = False
        self.stop_button.setEnabled(True)
        owner = MiniDatabaseSection._processing_owner
        if owner is not None and owner is not self:
            other_title = getattr(owner, "section_title", "Another section")
            if self not in MiniDatabaseSection._refresh_queue:
                MiniDatabaseSection._refresh_queue.append(self)
            status_message = f"Queued - waiting for {other_title}"
            self.status_label.setText(status_message)
            try:
                self.status_changed.emit(status_message)
            except Exception:
                pass
            self.log(
                f"{self.section_title}: queued refresh while {other_title} completes."
            )
            self._reset_progress_ui()
            self.stop_button.setEnabled(False)
            return
        MiniDatabaseSection._processing_owner = self
        status_message = f"Processing {len(candidates)} file(s)."
        self.status_label.setText(status_message)
        try:
            self.status_changed.emit(status_message)
        except Exception:
            pass
        self._start_progress(len(candidates))
        self._start_section_worker(candidates)

    def _start_section_worker(self, candidates: List[Path]) -> None:
        self._active_candidates = list(candidates)
        if self._worker_thread is not None and self._worker_thread.isRunning():
            try:
                self._worker_thread.quit()
                self._worker_thread.wait(100)
            except Exception:
                pass
        self._cleanup_worker_thread()
        thread = QtCore.QThread(self)
        worker = _SectionWorker(self, candidates)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.progress.connect(self._handle_worker_progress)
        worker.finished.connect(self._handle_worker_finished)
        worker.failed.connect(self._handle_worker_failed)
        worker.cancelled.connect(self._handle_worker_cancelled)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        worker.cancelled.connect(thread.quit)
        thread.finished.connect(self._cleanup_worker_thread)
        self._worker_thread = thread
        self._worker = worker
        thread.start()

    def _cleanup_worker_thread(self) -> None:
        if self._worker is not None:
            try:
                self._worker.deleteLater()
            except Exception:
                pass
            self._worker = None
        if self._worker_thread is not None:
            try:
                self._worker_thread.deleteLater()
            except Exception:
                pass
            self._worker_thread = None

    def _shutdown_background_threads(self) -> None:
        self._cancel_requested = True
        self._stop_pending_scan(wait_ms=800)
        thread = self._worker_thread
        if thread is not None:
            try:
                thread.quit()
            except Exception:
                pass
            try:
                if thread.isRunning():
                    thread.wait(800)
            except Exception:
                pass
            try:
                if thread.isRunning():
                    thread.terminate()
                    thread.wait(200)
            except Exception:
                pass
        self._cleanup_worker_thread()

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:  # type: ignore[override]
        self._shutdown_background_threads()
        super().closeEvent(event)

    def _handle_worker_progress(self, current: int, total: int, message: object) -> None:
        text: Optional[str] = None
        if isinstance(message, str) and message:
            text = message
        self._progress_callback(int(current), int(total), text)

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        self._finish_progress()
        existing_payloads = set(self.data.extra.get("payloads", {}).keys())
        new_payloads = set(result.payloads.keys())
        for name in existing_payloads - new_payloads:
            self.store.clear_payload(name)

        payload_map: Dict[str, str] = {}
        for name, payload in result.payloads.items():
            self.store.save_payload(name, payload)
            payload_map[name] = name
        if payload_map:
            self.data.extra["payloads"] = payload_map
        if result.extra:
            self.data.extra.update(result.extra)
        self.data.processed = result.processed
        self.data.table = result.table
        self.store.save(self.data)
        self._close_active_editor()
        self.model.set_frame(result.table)
        self._auto_fit_columns()
        self._update_status()
        processed_count = len(self._active_candidates)
        self.log(
            f"{self.section_title}: processed {processed_count} file(s)."
        )
        self._active_candidates = []
        self._update_open_sources_enabled()
        try:
            self.sources_changed.emit(list(self.data.sources))
        except Exception:
            pass
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _handle_worker_failed(self, exc: object) -> None:
        self._active_candidates = []
        self._fail_progress()
        if isinstance(exc, Exception):
            self.logger.exception("%s processing failed", self.section_title, exc_info=exc)
            detail = str(exc) if str(exc) else exc.__class__.__name__
        else:
            self.logger.error("%s processing failed: %s", self.section_title, exc)
            detail = str(exc)
        QtWidgets.QMessageBox.critical(
            self,
            self.section_title,
            f"Failed to process data:\\n{detail}",
        )

    def _handle_worker_cancelled(self) -> None:
        self._active_candidates = []
        self.log(f"{self.section_title}: processing cancelled by user.")
        self._cancel_progress()

    # ------------------------------------------------------------------ hooks for subclasses
    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        raise NotImplementedError

class FabricationSection(MiniDatabaseSection):
    section_key = "fabrication"
    section_title = "Fabrication data"
    supported_suffixes = (".xlsx", ".xls", ".xlsm")

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._table_splitter: QtWidgets.QSplitter | None = None
        self._separate_imported = False
        super().__init__(logger, log_callback, parent)
        self._normalize_temperature_columns()
        self._hide_columns(["_source_paths", "_source_path"])
        self.model.set_editable_columns(self._editable_columns())
        self.model.set_text_columns({GLASS_PULL_COLUMN, "Notes"})
        try:
            self.model.dataChanged.connect(self._handle_cell_edited)
        except Exception:
            pass

    def _normalize_temperature_columns(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else None
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        updated = frame.copy()
        if "Temperature (°C)" in updated.columns:
            if CORE_TEMPERATURE_COLUMN not in updated.columns:
                updated = updated.rename(columns={"Temperature (°C)": CORE_TEMPERATURE_COLUMN})
            else:
                legacy = updated["Temperature (°C)"]
                target = updated[CORE_TEMPERATURE_COLUMN]
                updated[CORE_TEMPERATURE_COLUMN] = target.where(
                    ~(target.isna() | (target == "")),
                    legacy,
                )
                updated = updated.drop(columns=["Temperature (°C)"])
        if GLASS_TEMPERATURE_COLUMN not in updated.columns:
            updated[GLASS_TEMPERATURE_COLUMN] = None
        if ESTIMATED_TRANSITION_COLUMN not in updated.columns:
            if "e/a" in updated.columns:
                updated[ESTIMATED_TRANSITION_COLUMN] = updated["e/a"].map(
                    _estimate_transition_temp_c
                )
            else:
                updated[ESTIMATED_TRANSITION_COLUMN] = None
        elif "e/a" in updated.columns:
            series = updated[ESTIMATED_TRANSITION_COLUMN]
            computed = updated["e/a"].map(_estimate_transition_temp_c)
            mask = series.isna() | (series == "")
            updated[ESTIMATED_TRANSITION_COLUMN] = series.where(~mask, computed)
        if GLASS_PULL_COLUMN not in updated.columns:
            updated[GLASS_PULL_COLUMN] = None
        if updated is not frame:
            self.data.table = updated
            self.model.set_frame(updated)
            try:
                self.store.save(self.data)
            except Exception:
                pass

    def _collect_candidates(self) -> List[Path]:
        _, relevant_compositions = self._load_relevant_map()
        tokens = {
            self._normalise_token(comp)
            for comp in relevant_compositions
            if self._normalise_token(comp)
        }
        if not tokens:
            return super()._collect_candidates()

        candidates: Dict[str, Path] = {}

        def _should_consider(path: Path, root: Path) -> bool:
            try:
                relative = path.relative_to(root)
                text = self._normalise_token(str(relative))
            except Exception:
                text = self._normalise_token(path.name)
            if not text:
                return False
            return any(token in text for token in tokens)

        for source in self.data.sources:
            root = Path(source).expanduser()
            if not root.exists():
                continue
            try:
                resolved_root = root.resolve()
            except Exception:
                resolved_root = root
            if self.recursive_search:
                stack: List[Tuple[Path, bool]] = [(resolved_root, False)]
                while stack:
                    current, matched = stack.pop()
                    try:
                        entries = list(current.iterdir())
                    except Exception:
                        continue
                    for entry in entries:
                        if entry.is_dir():
                            next_matched = matched or _should_consider(entry, resolved_root)
                            if current is resolved_root or next_matched:
                                stack.append((entry, next_matched))
                            continue
                        if not entry.is_file():
                            continue
                        if self.supported_suffixes and entry.suffix.lower() not in self.supported_suffixes:
                            continue
                        if not (matched or _should_consider(entry, resolved_root)):
                            continue
                        try:
                            resolved = str(entry.resolve())
                        except Exception:
                            resolved = str(entry)
                        candidates.setdefault(resolved, entry)
            else:
                try:
                    entries = list(resolved_root.iterdir())
                except Exception:
                    entries = []
                for entry in entries:
                    if not entry.is_file():
                        continue
                    if self.supported_suffixes and entry.suffix.lower() not in self.supported_suffixes:
                        continue
                    if not _should_consider(entry, resolved_root):
                        continue
                    try:
                        resolved = str(entry.resolve())
                    except Exception:
                        resolved = str(entry)
                    candidates.setdefault(resolved, entry)
        if not candidates:
            return super()._collect_candidates()
        return sorted(candidates.values())

    def _update_status(self) -> None:  # type: ignore[override]
        super()._update_status()
        missing = self._missing_microwires()
        if not missing:
            return
        base_message = self.status_label.text()
        preview = ", ".join(missing[:3])
        if len(missing) > 3:
            preview += ", …"
        message = f"{base_message} — Missing {len(missing)} annealing record(s): {preview}"
        self.status_label.setText(message)
        try:
            self.status_changed.emit(message)
        except Exception:
            pass

    def _missing_microwires(self) -> List[str]:
        expected_map, _ = self._load_relevant_map()
        expected: Set[Tuple[str, int, int]] = set()
        for composition, draws in expected_map.items():
            for draw, pieces in draws.items():
                for piece in pieces:
                    if piece is None:
                        continue
                    try:
                        expected.add((str(composition), int(draw), int(piece)))
                    except (TypeError, ValueError):
                        continue
        if not expected:
            return []
        present: Set[Tuple[str, int, int]] = set()
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if isinstance(frame, pd.DataFrame) and not frame.empty:
            if "_group_key" in frame.columns:
                for value in frame["_group_key"]:
                    if value in (None, "", "None"):
                        continue
                    key_parts = _microwire_key_from_string(str(value))
                    if key_parts is None:
                        continue
                    comp, draw_val, piece_val, _suffix = key_parts
                    present.add((comp, draw_val, piece_val))
            else:
                for _, row in frame.iterrows():
                    comp = str(row.get("Composition", "")).strip()
                    microwire = str(row.get("Microwire", "")).strip()
                    parsed = _microwire_parts_from_label_safe(microwire)
                    if parsed is None:
                        continue
                    draw_val, piece_val, _suffix = parsed
                    present.add((comp, int(draw_val), int(piece_val)))
        missing = [
            f"{composition} {draw}/{piece}"
            for composition, draw, piece in sorted(expected)
            if (composition, draw, piece) not in present
        ]
        return missing

    @staticmethod
    def _normalise_int(value: object) -> Optional[int]:
        if value is None:
            return None
        if isinstance(value, (int,)):
            return int(value)
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return None
        if math.isnan(numeric):
            return None
        return int(numeric)

    def _load_relevant_map(
        self,
    ) -> Tuple[Dict[str, Dict[Optional[int], Set[Optional[int]]]], Set[str]]:
        try:
            store = MiniDatabaseStore("annealing")
            records = store.load_payload("annealing_records")
        except Exception:
            records = None
        relevant: Dict[str, Dict[Optional[int], Set[Optional[int]]]] = {}
        if isinstance(records, list):
            for record in records:
                metadata = getattr(record, "metadata", None)
                if metadata is None:
                    continue
                composition = getattr(metadata, "composition_token", None)
                if not composition:
                    continue
                composition_key = str(composition).strip()
                if not composition_key:
                    continue
                draw_value = self._normalise_int(getattr(metadata, "draw_x", None))
                piece_value = self._normalise_int(getattr(metadata, "piece_y", None))
                bucket = relevant.setdefault(composition_key, {})
                piece_bucket = bucket.setdefault(draw_value, set())
                piece_bucket.add(piece_value)
        return relevant, set(relevant.keys())

    @staticmethod
    def _allow_draw(
        relevant_map: Dict[str, Dict[Optional[int], Set[Optional[int]]]],
        composition: str,
        draw: int,
    ) -> bool:
        draw_map = relevant_map.get(composition)
        if not draw_map:
            return True
        if draw in draw_map:
            return True
        if None in draw_map:
            return True
        return False

    @staticmethod
    def _allow_piece(
        relevant_map: Dict[str, Dict[Optional[int], Set[Optional[int]]]],
        composition: str,
        draw: int,
        piece: int,
    ) -> bool:
        draw_map = relevant_map.get(composition)
        if not draw_map:
            return True
        allowed: Set[Optional[int]] = set()
        direct = draw_map.get(draw)
        if direct:
            allowed.update(direct)
        fallback = draw_map.get(None)
        if fallback:
            allowed.update(fallback)
        if not allowed:
            return False
        if None in allowed:
            return True
        return piece in allowed

    def _filter_index(
        self,
        index: FabricationIndex,
        relevant_map: Dict[str, Dict[Optional[int], Set[Optional[int]]]],
        relevant_compositions: Set[str],
    ) -> FabricationIndex:
        filtered = FabricationIndex()
        for (composition, draw), draw_data in index.draw_level.items():
            comp_key = str(composition).strip()
            if comp_key not in relevant_compositions:
                continue
            if self._allow_draw(relevant_map, comp_key, int(draw)):
                filtered.set_draw(comp_key, int(draw), dict(draw_data))
        for (composition, draw, piece), piece_data in index.piece_level.items():
            comp_key = str(composition).strip()
            if comp_key not in relevant_compositions:
                continue
            draw_int = int(draw)
            piece_int = int(piece)
            if not self._allow_draw(relevant_map, comp_key, draw_int):
                continue
            if not self._allow_piece(relevant_map, comp_key, draw_int, piece_int):
                continue
            filtered.set_piece(comp_key, draw_int, piece_int, dict(piece_data))
        return filtered

    @staticmethod
    def _normalise_token(text: str) -> str:
        return re.sub(r"[^a-z0-9]", "", str(text).lower())

    def _filter_candidates_for_relevance(
        self,
        candidates: List[Path],
        relevant_map: Dict[str, Dict[Optional[int], Set[Optional[int]]]],
        relevant_compositions: Set[str],
    ) -> Tuple[List[Path], int, bool]:
        if not relevant_compositions:
            return candidates, 0, False
        composition_tokens = {
            comp: self._normalise_token(comp)
            for comp in relevant_compositions
            if self._normalise_token(comp)
        }
        if not composition_tokens:
            return candidates, 0, False
        draw_tokens: Dict[str, Set[str]] = {}
        for comp, draw_map in relevant_map.items():
            comp_key = self._normalise_token(comp)
            if not comp_key:
                continue
            bucket = draw_tokens.setdefault(comp_key, set())
            for draw, pieces in draw_map.items():
                if draw is not None:
                    bucket.add(self._normalise_token(draw))
                for piece in pieces:
                    if piece is not None:
                        bucket.add(self._normalise_token(f"{draw}{piece}"))
        filtered: List[Path] = []
        skipped = 0
        for path in candidates:
            text = self._normalise_token(path)
            if not text:
                filtered.append(path)
                continue
            matched = False
            for _, token in composition_tokens.items():
                if token and token in text:
                    matched = True
                    break
                draw_set = draw_tokens.get(token)
                if draw_set and any(draw_token in text for draw_token in draw_set if draw_token):
                    matched = True
                    break
            if matched:
                filtered.append(path)
            else:
                skipped += 1
        if not filtered:
            return candidates, 0, True
        return filtered, skipped, False

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        unique_paths = list(dict.fromkeys(Path(p) for p in paths))

        relevant_map, relevant_compositions = self._load_relevant_map()
        filtered_paths, skipped, reverted = self._filter_candidates_for_relevance(
            unique_paths, relevant_map, relevant_compositions
        )
        if reverted:
            self.log(
                "Fabrication data: relevance filter could not match any files; falling back to full file list."
            )
        elif skipped:
            self.log(
                f"Fabrication data: skipped {skipped} file(s) without matching current annealing composition."
            )

        def _progress(idx: int, total: int) -> None:
            self._check_cancelled()
            if progress is None:
                return
            message: Optional[str] = None
            if 0 < idx <= len(filtered_paths):
                message = f"Parsing {filtered_paths[idx - 1].name}"
            try:
                progress(idx, total, message)
            except Exception:
                pass

        index = build_fabrication_index(
            filtered_paths,
            self.logger,
            progress_callback=_progress,
            cancel_callback=self.is_cancelled,
        )
        self._check_cancelled()
        if relevant_compositions:
            original_draws = len(index.draw_level)
            original_pieces = len(index.piece_level)
            index = self._filter_index(index, relevant_map, relevant_compositions)
            removed_draws = original_draws - len(index.draw_level)
            removed_pieces = original_pieces - len(index.piece_level)
            if removed_draws > 0 or removed_pieces > 0:
                self.log(
                    "Fabrication data: skipped %d draw(s) and %d piece(s) without matching current annealing records."
                    % (removed_draws, removed_pieces)
                )
        table = _fabrication_index_to_frame(index)
        table = self._apply_imported_separation(table)
        processed: Dict[str, float] = {}
        for path in filtered_paths:
            try:
                processed[str(path)] = float(path.stat().st_mtime)
            except OSError:
                continue
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"fabrication_index": index},
        )

    def refresh(self) -> None:
        super().refresh()
        table = self.data.table
        if isinstance(table, pd.DataFrame):
            if "Temperature (°C)" in table.columns and CORE_TEMPERATURE_COLUMN not in table.columns:
                table = table.rename(columns={"Temperature (°C)": CORE_TEMPERATURE_COLUMN})
                table[GLASS_TEMPERATURE_COLUMN] = None
            if ESTIMATED_TRANSITION_COLUMN not in table.columns:
                if "e/a" in table.columns:
                    table[ESTIMATED_TRANSITION_COLUMN] = table["e/a"].map(
                        _estimate_transition_temp_c
                    )
                else:
                    table[ESTIMATED_TRANSITION_COLUMN] = None
            elif "e/a" in table.columns:
                series = table[ESTIMATED_TRANSITION_COLUMN]
                computed = table["e/a"].map(_estimate_transition_temp_c)
                mask = series.isna() | (series == "")
                table[ESTIMATED_TRANSITION_COLUMN] = series.where(~mask, computed)
            if GLASS_PULL_COLUMN not in table.columns:
                table[GLASS_PULL_COLUMN] = None
            self.data.table = self._apply_imported_separation(table)
            self.model.set_frame(self.data.table)
        self._hide_columns(["_source_paths", "_source_path"])

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._normalize_temperature_columns()
        table = self.data.table
        if isinstance(table, pd.DataFrame):
            self.data.table = self._apply_imported_separation(table)
            self.model.set_frame(self.data.table)
        self._hide_columns(["_source_paths", "_source_path"])

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw_paths: List[str] = []
        path_values = row.get("_source_paths")
        if isinstance(path_values, (list, tuple, set)):
            raw_paths.extend(str(value) for value in path_values if value)
        fallback = row.get("_source_path")
        if fallback:
            raw_paths.append(str(fallback))
        for entry in dict.fromkeys(raw_paths):
            try:
                sources.append(Path(entry))
            except Exception:
                continue
        return sources

    def set_import_separation(self, enabled: bool) -> None:
        self._separate_imported = bool(enabled)
        try:
            index = self.store.load_payload("fabrication_index")
        except Exception:
            index = None
        if not isinstance(index, FabricationIndex):
            return
        table = _fabrication_index_to_frame(index)
        table = self._apply_imported_separation(table)
        self.data.table = table
        self.model.set_frame(table)
        self._auto_fit_columns()
        self._update_status()

    def _apply_imported_separation(self, table: pd.DataFrame) -> pd.DataFrame:
        if not self._separate_imported:
            return table
        if not isinstance(table, pd.DataFrame) or table.empty:
            return table
        if "Data source" not in table.columns:
            return table
        imported_mask = table["Data source"].astype(str).str.contains("Imported", na=False)
        if not imported_mask.any():
            return table
        normal = table.loc[~imported_mask]
        imported = table.loc[imported_mask]
        separator = {column: None for column in table.columns}
        separator["Composition"] = "Imported data:"
        combined = pd.concat([normal, pd.DataFrame([separator]), imported], ignore_index=True)
        return combined

    @staticmethod
    def _coerce_text(value: Any) -> Optional[str]:
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            if isinstance(value, float) and math.isnan(value):
                return None
        text = str(value).strip()
        return text if text else None

    def _editable_columns(self) -> Set[str]:
        return {GLASS_PULL_COLUMN, "Notes"}

    def _handle_cell_edited(
        self,
        top_left: QtCore.QModelIndex,
        bottom_right: QtCore.QModelIndex,
        roles: Tuple[QtCore.Qt.ItemDataRole, ...] = (),
    ) -> None:
        if roles and QtCore.Qt.ItemDataRole.EditRole not in roles:
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        columns = list(frame.columns[top_left.column() : bottom_right.column() + 1])
        relevant = self._editable_columns()
        if not any(column in relevant for column in columns):
            return
        try:
            index = self.store.load_payload("fabrication_index")
        except Exception:
            index = None
        if not isinstance(index, FabricationIndex):
            index = FabricationIndex()
        updated_any = False
        for row_idx in range(top_left.row(), bottom_right.row() + 1):
            if row_idx < 0 or row_idx >= len(frame.index):
                continue
            series = frame.iloc[row_idx]
            composition = str(series.get("Composition") or "").strip()
            if not composition:
                continue
            try:
                draw = int(series.get("Draw"))
                piece = int(series.get("Piece"))
            except (TypeError, ValueError):
                continue
            key = (composition, draw, piece)
            piece_data = dict(index.piece_level.get(key, {}))
            row_updated = False
            for column in columns:
                if column not in relevant:
                    continue
                field = "notes" if column == "Notes" else "glass_pull_off"
                value = self._coerce_text(series.get(column))
                if value is None:
                    piece_data.pop(field, None)
                else:
                    piece_data[field] = value
                row_updated = True
                updated_any = True
            if row_updated:
                if piece_data:
                    piece_data.setdefault("_source_path", piece_data.get("_source_path") or "Manual")
                index.piece_level[key] = piece_data
        if updated_any:
            self.data.table = frame
            self.store.save_payload("fabrication_index", index)
            payload_map = dict(self.data.extra.get("payloads", {}))
            payload_map["fabrication_index"] = "fabrication_index"
            self.data.extra["payloads"] = payload_map
            self.store.save(self.data)

    def apply_imported_samples(self, records: Iterable[Dict[str, Any]]) -> int:
        index = self.store.load_payload("fabrication_index")
        if not isinstance(index, FabricationIndex):
            index = FabricationIndex()
        added = 0
        for record in records:
            composition = str(record.get("Composition") or "").strip()
            microwire = str(record.get("Microwire") or "").strip()
            if not composition or not microwire:
                continue
            parts = _microwire_parts_from_label_safe(microwire)
            if parts is None:
                continue
            draw_x, piece_y, _suffix = parts
            piece_data: Dict[str, object] = {
                "length_m": record.get("Length (m)"),
                "piece_date": record.get("Piece date"),
                "fabrication_resistance_ohm": record.get("Resistance (Ω)"),
                "glass_pull_off": record.get(GLASS_PULL_COLUMN),
                "notes": record.get("Notes"),
                "_source_path": record.get("Data source") or "Imported",
            }
            temperature_value = record.get(CORE_TEMPERATURE_COLUMN)
            if temperature_value in (None, ""):
                temperature_value = record.get("Temperature (°C)")
            glass_temperature_value = record.get(GLASS_TEMPERATURE_COLUMN)
            draw_data: Dict[str, object] = {
                "fabrication_temperature_c": temperature_value,
                "fabrication_glass_temperature_c": glass_temperature_value,
                "mass_g": record.get("Mass (g)"),
                "winding_speed_m_per_min": record.get("Winding speed (m/min)"),
                "glass_feed_mm_per_min": record.get("Glass feeding (mm/min)"),
                "underpressure": record.get("Underpressure"),
                "production_datetime": record.get("Production datetime"),
                "fabrication_resistance_ohm": record.get("Resistance (Ω)"),
                "notes": record.get("Notes"),
                "_source_path": record.get("Data source") or "Imported",
            }
            before = bool(index.get_piece(composition, draw_x, piece_y))
            index.set_piece(composition, int(draw_x), int(piece_y), piece_data)
            index.set_draw(composition, int(draw_x), draw_data)
            if not before:
                added += 1
        table = _fabrication_index_to_frame(index)
        table = self._apply_imported_separation(table)
        self.data.table = table
        self.store.save_payload("fabrication_index", index)
        payload_map = dict(self.data.extra.get("payloads", {}))
        payload_map["fabrication_index"] = "fabrication_index"
        self.data.extra["payloads"] = payload_map
        self.store.save(self.data)
        self.model.set_frame(table)
        self._auto_fit_columns()
        self._update_status()
        try:
            self.data_updated.emit()
        except Exception:
            pass
        return added


class AnnealingSection(MiniDatabaseSection):
    section_key = "annealing"
    section_title = "Current annealing"
    supported_suffixes = (".txt", ".csv", ".tsv")

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._table_splitter: QtWidgets.QSplitter | None = None
        self._preview_other_count = 1
        self._preview_spacing = 6
        super().__init__(logger, log_callback, parent)
        self._hidden_paths: Set[str] = set()
        self._all_records: List[MeasurementRecord] = []
        self._pixmap_cache: Dict[Tuple[object, ...], Optional[QtGui.QPixmap]] = {}
        self._phase_points: Dict[str, Dict[str, float]] = {}
        stored_phase_points = self.data.extra.get("phase_points")
        if isinstance(stored_phase_points, dict):
            cleaned: Dict[str, Dict[str, float]] = {}
            for key, payload in stored_phase_points.items():
                if not isinstance(key, str) or not isinstance(payload, dict):
                    continue
                entry = self._clean_phase_points_payload(payload)
                if entry:
                    cleaned[key] = entry
            self._phase_points = cleaned
        self._load_hidden_paths()
        if isinstance(self.model, DataFrameModel):
            self.model.set_decoration_provider(self._preview_decoration)
        self._sanitize_graph_columns()
        self._record_groups: Dict[str, List[MeasurementRecord]] = {}
        self.export_button = QtWidgets.QPushButton("Export worksheet…")
        self.export_button.clicked.connect(self._export_worksheet)
        self.controls_layout.addWidget(self.export_button)
        self.open_pyplot_button = QtWidgets.QPushButton("Open in PyPlot")
        self.open_pyplot_button.setToolTip("Open the selected annealing files in PyPlot.")
        self.open_pyplot_button.clicked.connect(self._open_selected_in_pyplot)
        self.controls_layout.addWidget(self.open_pyplot_button)
        self.open_origin_button = QtWidgets.QPushButton("Open in Origin")
        self.open_origin_button.setToolTip("Send the selected annealing files to Origin via PyPlot.")
        self.open_origin_button.clicked.connect(self._open_selected_in_origin)
        self.controls_layout.addWidget(self.open_origin_button)
        self.visibility_button = QtWidgets.QPushButton("Visibility...")
        self.visibility_button.setToolTip("Show or hide specific annealing graphs.")
        self.visibility_button.clicked.connect(self._open_visibility_dialog)
        self.controls_layout.addWidget(self.visibility_button)
        self._update_export_enabled()
        self._refresh_record_groups()
        self._hide_columns(["_group_key", "_sources"])

    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        table = QtWidgets.QTableView(parent)
        table.setModel(self.model)
        table.horizontalHeader().setStretchLastSection(True)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        table.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        table.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.SelectedClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed
        )
        table.setSortingEnabled(True)
        table.setIconSize(
            QtCore.QSize(self._preview_icon_width(), ANNEALING_GRAPH_HEIGHT)
        )
        header = table.verticalHeader()
        if header is not None:
            default_height = ANNEALING_GRAPH_HEIGHT + 24
            header.setDefaultSectionSize(default_height)
            header.setMinimumSectionSize(default_height)
        self.table_view = table
        return table

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        records: List[MeasurementRecord] = []
        processed: Dict[str, float] = {}
        total = len(paths)
        sanity_failures: List[Tuple[Path, Optional[float]]] = []
        for idx, path in enumerate(paths, start=1):
            self._check_cancelled()
            metadata = _metadata_from_path(path)
            try:
                df = _load_annealing(path, expected_setpoint_mA=metadata.setpoint_mA)
            except Exception:
                self.logger.exception("Failed to parse %s", path)
                if progress is not None:
                    try:
                        progress(idx, total, f"Failed: {path.name}")
                    except Exception:
                        pass
                continue
            ok, mean_error = _resistance_sanity_check(df)
            if not ok:
                sanity_failures.append((path, mean_error))
            record = MeasurementRecord(
                path=path,
                metadata=metadata,
                dataframe=df,
                sanity_ok=ok,
                sanity_error=mean_error,
            )
            records.append(record)
            try:
                processed[str(path)] = float(path.stat().st_mtime)
            except OSError:
                processed[str(path)] = 0.0
            if progress is not None:
                try:
                    progress(idx, total, f"Parsed {path.name}")
                except Exception:
                    pass
        table = _annealing_records_to_frame(records, self.logger)
        column_keys = table.get("_group_key") if isinstance(table, pd.DataFrame) else None
        valid_keys = {
            str(value)
            for value in (column_keys.tolist() if column_keys is not None else [])
            if value not in (None, "", "None") and not (isinstance(value, float) and math.isnan(value))
        }
        self._prune_phase_points(valid_keys, store=False)
        if sanity_failures:
            preview = ", ".join(p.name for p, _ in sanity_failures[:5])
            if len(sanity_failures) > 5:
                preview += ", …"
            errors = [err for _, err in sanity_failures if isinstance(err, (int, float))]
            worst = f" (worst error {max(errors) * 100:.1f}%)" if errors else ""
            summary = (
                f"{self.section_title}: R≈V/I sanity check failed for {len(sanity_failures)} file(s){worst}"
            )
            if preview:
                summary += f": {preview}"
            self.log(summary, level=logging.WARNING)
            self.logger.warning(summary)
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"annealing_records": records},
            extra={"phase_points": dict(self._phase_points)},
        )

    def refresh(self) -> None:
        super().refresh()
        self._sanitize_graph_columns()
        self._hide_columns(["_group_key", "_sources"])
        self._refresh_record_groups()
        self._prune_phase_points()
        self._update_export_enabled()

    def _update_export_enabled(self) -> None:
        has_rows = isinstance(self.data.table, pd.DataFrame) and not self.data.table.empty
        if hasattr(self, "export_button"):
            self.export_button.setEnabled(has_rows)

    def _load_hidden_paths(self) -> None:
        hidden = self.data.extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            self._hidden_paths = {str(path) for path in hidden if path}
        else:
            self._hidden_paths = set()

    def _store_hidden_paths(self) -> None:
        self.data.extra["hidden_paths"] = sorted(self._hidden_paths)
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist annealing visibility settings")

    def _visible_records(
        self, records: Sequence[MeasurementRecord]
    ) -> List[MeasurementRecord]:
        if not self._hidden_paths:
            return list(records)
        return [
            record
            for record in records
            if _record_path_key(record) not in self._hidden_paths
        ]

    def _open_visibility_dialog(self) -> None:
        items = _visibility_items_from_records(self._all_records)
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No annealing graphs are available yet.",
            )
            return
        groups = _visibility_groups_from_records(self._all_records)
        dialog = _GraphVisibilityDialog(
            "Annealing graph visibility",
            items,
            self._hidden_paths,
            groups=groups,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._hidden_paths = dialog.hidden_paths()
            self._store_hidden_paths()
            self._refresh_record_groups()

    def _clean_phase_points_payload(self, payload: Dict[str, Any]) -> Dict[str, float]:
        entry: Dict[str, float] = {}
        for label in PHASE_POINT_LABELS + ("As", "Ms"):
            value = payload.get(label)
            if isinstance(value, (int, float)) and math.isfinite(float(value)):
                entry[label] = float(value)
        if "As1" not in entry and "As" in entry:
            entry["As1"] = entry["As"]
        if "Ms1" not in entry and "Ms" in entry:
            entry["Ms1"] = entry["Ms"]
        if "As1" in entry:
            entry["As"] = entry["As1"]
        if "Ms1" in entry:
            entry["Ms"] = entry["Ms1"]
        return entry

    def _store_phase_points(self) -> None:
        snapshot: Dict[str, Dict[str, float]] = {}
        for key, payload in self._phase_points.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            cleaned = self._clean_phase_points_payload(payload)
            if cleaned:
                snapshot[key] = cleaned
        self.data.extra["phase_points"] = snapshot
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist phase transition points")

    def phase_points_snapshot(self) -> Dict[str, Dict[str, float]]:
        snapshot: Dict[str, Dict[str, float]] = {}
        for key, payload in self._phase_points.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            cleaned = self._clean_phase_points_payload(payload)
            if cleaned:
                snapshot[key] = cleaned
        return snapshot

    def set_phase_points_for_key(
        self,
        key: str,
        *,
        as_value: Optional[float] = None,
        ms_value: Optional[float] = None,
        phase_values: Optional[Dict[str, Optional[float]]] = None,
    ) -> None:
        key_text = str(key).strip()
        if not key_text:
            return
        entry: Dict[str, float] = {}

        def _clean(value: Optional[float]) -> Optional[float]:
            if value is None:
                return None
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                return None
            return numeric if math.isfinite(numeric) else None

        phase_values = phase_values or {}
        for label in PHASE_POINT_LABELS:
            cleaned = _clean(phase_values.get(label))
            if cleaned is not None:
                entry[label] = cleaned
        if "As1" not in entry:
            as_clean = _clean(as_value)
            if as_clean is not None:
                entry["As1"] = as_clean
        if "Ms1" not in entry:
            ms_clean = _clean(ms_value)
            if ms_clean is not None:
                entry["Ms1"] = ms_clean
        if "As1" in entry:
            entry["As"] = entry["As1"]
        if "Ms1" in entry:
            entry["Ms"] = entry["Ms1"]
        if entry:
            self._phase_points[key_text] = entry
        elif key_text in self._phase_points:
            self._phase_points.pop(key_text, None)
        self._store_phase_points()
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _prune_phase_points(self, valid_keys: Optional[Iterable[str]] = None, *, store: bool = True) -> None:
        if valid_keys is None:
            frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
            if frame.empty or "_group_key" not in frame.columns:
                valid_set: Set[str] = set()
            else:
                valid_set = {str(value) for value in frame["_group_key"].dropna().astype(str)}
        else:
            valid_set = {str(value) for value in valid_keys if value}
        removed = False
        for key in list(self._phase_points.keys()):
            if key not in valid_set:
                self._phase_points.pop(key, None)
                removed = True
        if removed and store:
            self._store_phase_points()

    def _selected_group_key(self) -> Optional[str]:
        rows = self._selected_rows()
        if not rows:
            return None
        series = self._row_series(rows[0])
        if series is None:
            return None
        key = series.get("_group_key")
        if key in (None, "", float("nan")):
            return None
        return str(key)

    def _selected_records(self) -> List[MeasurementRecord]:
        rows = self._selected_rows()
        records: List[MeasurementRecord] = []
        if not rows:
            return records
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            key = series.get("_group_key")
            if key in (None, "", float("nan")):
                continue
            records.extend(self._record_groups.get(str(key), []))
        return records

    def _open_selected_in_pyplot(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [
            record.path for record in records if isinstance(record.path, Path)
        ]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "Current Annealing",
            self.logger,
            auto_plot=True,
            open_origin=False,
        )

    def _open_selected_in_origin(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [
            record.path for record in records if isinstance(record.path, Path)
        ]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "Current Annealing",
            self.logger,
            auto_plot=True,
            open_origin=True,
        )


    def _update_open_sources_enabled(self) -> None:
        super()._update_open_sources_enabled()

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._refresh_record_groups()
        self._update_export_enabled()

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[MeasurementRecord]] = {}
        try:
            payload = self.store.load_payload("annealing_records")
        except Exception:
            payload = None
        all_records = list(payload) if isinstance(payload, list) else []
        self._all_records = list(all_records)
        visible_records = self._visible_records(all_records)
        if visible_records:
            for record in visible_records:
                metadata = getattr(record, "metadata", None)
                if metadata is None:
                    continue
                composition = getattr(metadata, "composition_token", None)
                draw = getattr(metadata, "draw_x", None)
                piece = getattr(metadata, "piece_y", None)
                if composition is None or draw is None or piece is None:
                    continue
                suffix = None
                path_value = getattr(record, "path", None)
                if isinstance(path_value, Path):
                    parsed_key = _microscope_key(path_value)
                    if parsed_key is not None:
                        _, _, _, suffix = parsed_key
                try:
                    key = _microwire_key_to_str((composition, int(draw), int(piece), suffix))
                except (TypeError, ValueError):
                    continue
                grouped.setdefault(key, []).append(record)
        self._record_groups = grouped
        max_other = 1
        for records in grouped.values():
            high_record, low_record = _select_high_low_pair(records)
            other_records = _select_other_measurements(records, high_record, low_record)
            if other_records:
                max_other = max(max_other, len(other_records))
        self._preview_other_count = max_other
        self._invalidate_previews()
        self._update_preview_icon_size()

    def _sanitize_graph_columns(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            return
        changed = False
        for column in ("Graph — 1000 mA", "Graph — low mA", ANNEALING_OTHER_GRAPH_COLUMN):
            if column not in frame.columns:
                continue
            series = frame[column]
            if not hasattr(series, "apply"):
                continue
            cleaned = series.apply(
                lambda value: None
                if isinstance(value, (QtGui.QPixmap, QtGui.QImage))
                else value
            )
            if not cleaned.equals(series):
                frame[column] = cleaned
                changed = True
        drop_columns = {
            ANNEALING_AS_COLUMN,
            ANNEALING_MS_COLUMN,
            "Low current setpoint",
            "Updated",
        }
        if any(column in frame.columns for column in drop_columns):
            frame = frame.drop(columns=[column for column in drop_columns if column in frame.columns])
            changed = True

        desired_order = [
            "Composition",
            "Microwire",
            "Graph — 1000 mA",
            "Graph — low mA",
            ANNEALING_OTHER_GRAPH_COLUMN,
            "_group_key",
            "_sources",
        ]
        current_columns = list(frame.columns)
        preferred = [col for col in desired_order if col in frame.columns]
        remainder = [col for col in current_columns if col not in desired_order]
        new_order = preferred + remainder
        if new_order != current_columns and new_order:
            frame = frame.loc[:, new_order]
            changed = True
        if changed:
            self.data.table = frame
            if isinstance(self.model, DataFrameModel):
                self.model.set_frame(frame)
                self._auto_fit_columns()
            try:
                self.store.save(self.data)
            except Exception:
                pass
            try:
                self.data_updated.emit()
            except Exception:
                pass
        self._invalidate_previews()

    def _invalidate_previews(self) -> None:
        self._pixmap_cache.clear()
        if isinstance(self.model, DataFrameModel):
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass

    def _preview_icon_width(self) -> int:
        count = max(int(getattr(self, "_preview_other_count", 1)), 1)
        return ANNEALING_GRAPH_WIDTH * count + self._preview_spacing * (count - 1)

    def _update_preview_icon_size(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        width = self._preview_icon_width()
        try:
            table.setIconSize(
                QtCore.QSize(max(width, ANNEALING_GRAPH_WIDTH), ANNEALING_GRAPH_HEIGHT)
            )
        except Exception:
            pass
        header = table.verticalHeader()
        if header is not None:
            try:
                header.setDefaultSectionSize(ANNEALING_GRAPH_HEIGHT + 24)
            except Exception:
                pass
        self._auto_fit_columns()

    def _preview_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column not in {"Graph — 1000 mA", "Graph — low mA", ANNEALING_OTHER_GRAPH_COLUMN}:
            return None
        key = row.get("_group_key")
        if not isinstance(key, str) or not key:
            return None
        cache_key: Tuple[object, ...] = (key, column)
        if cache_key in self._pixmap_cache:
            return self._pixmap_cache[cache_key]
        records = self._record_groups.get(key)
        pixmap: Optional[QtGui.QPixmap] = None
        if records:
            high_record, low_record = _select_high_low_pair(records)
            if column == "Graph — 1000 mA":
                target = high_record
                pixmap = _render_measurement_pixmap(target, self.logger)
            elif column == "Graph — low mA":
                target = low_record
                pixmap = _render_measurement_pixmap(target, self.logger)
            else:
                other_records = _select_other_measurements(records, high_record, low_record)
                if other_records:
                    signature = tuple(
                        str(getattr(record, "path", "")) for record in other_records
                    )
                    cache_key = (key, column, signature)
                    cached = self._pixmap_cache.get(cache_key)
                    if cached is not None:
                        return cached
                    pixmaps: List[QtGui.QPixmap] = []
                    for record in other_records:
                        preview = _render_measurement_pixmap(record, self.logger)
                        if preview is not None:
                            pixmaps.append(preview)
                    pixmap = _combine_pixmaps_side_by_side(
                        pixmaps,
                        width_px=self._preview_icon_width(),
                        height_px=ANNEALING_GRAPH_HEIGHT,
                        spacing=self._preview_spacing,
                        scale_to_fit=False,
                    )
                    self._pixmap_cache[cache_key] = pixmap
                    return pixmap
        if pixmap is None:
            path_value = row.get(column)
            if isinstance(path_value, str) and path_value:
                candidate = Path(path_value)
                if candidate.exists():
                    loaded = QtGui.QPixmap(str(candidate))
                    if not loaded.isNull():
                        pixmap = loaded
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw_sources = row.get("_sources")
        if isinstance(raw_sources, (list, tuple)):
            for entry in raw_sources:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        key = row.get("_group_key")
        if isinstance(key, str):
            for record in self._record_groups.get(key, []):
                path = getattr(record, "path", None)
                if path:
                    try:
                        candidate = Path(path)
                    except Exception:
                        continue
                    if candidate not in sources:
                        sources.append(candidate)
        return sources

    def _export_worksheet(self) -> None:
        records = self.store.load_payload("annealing_records")
        if not isinstance(records, list) or not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Process current annealing files before exporting.",
            )
            return

        path_str, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save annealing worksheet",
            str(Path.cwd() / "annealing_summary.xlsx"),
            "Excel files (*.xlsx)",
        )
        if not path_str:
            return
        output_path = Path(path_str)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        try:
            import xlsxwriter  # type: ignore[import-not-found]
        except ImportError:
            QtWidgets.QMessageBox.warning(
                self,
                self.section_title,
                "xlsxwriter is required to export worksheets.",
            )
            return

        grouped: Dict[MicrowireKey, List[MeasurementRecord]] = {}
        for record in records:
            metadata = getattr(record, "metadata", None)
            if metadata is None:
                continue
            draw_x = getattr(metadata, "draw_x", None)
            piece_y = getattr(metadata, "piece_y", None)
            composition = getattr(metadata, "composition_token", None)
            if composition is None or draw_x is None or piece_y is None:
                continue
            suffix = None
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                parsed_key = _microscope_key(path)
                if parsed_key is not None:
                    _, _, _, suffix = parsed_key
            grouped.setdefault((str(composition), int(draw_x), int(piece_y), suffix), []).append(record)

        if not grouped:
            QtWidgets.QMessageBox.warning(
                self,
                self.section_title,
                "No complete microwire records available for export.",
            )
            return

        try:
            with TemporaryDirectory() as tmpdir:
                workbook = xlsxwriter.Workbook(str(output_path))
                try:
                    worksheet = workbook.add_worksheet("Summary")
                    header_format = workbook.add_format({"bold": True})
                    headers = [
                        "Composition",
                        "Microwire",
                        "1000 mA file",
                        "Low mA (mA)",
                        "Low mA file",
                        "Graph — 1000 mA",
                        "Graph — low mA",
                    ]
                    worksheet.write_row(0, 0, headers, header_format)
                    worksheet.set_column(0, 1, 18)
                    worksheet.set_column(2, 4, 22)
                    worksheet.set_column(5, 6, 42)

                    plot_dir = Path(tmpdir)
                    row_idx = 1
                    for (composition, draw_x, piece_y, suffix), recs in sorted(
                        grouped.items(),
                        key=lambda item: (
                            str(item[0][0]).lower(),
                            int(item[0][1]),
                            int(item[0][2]),
                            (str(item[0][3]).lower() if item[0][3] is not None else ""),
                        ),
                    ):
                        high_record, low_record = _select_high_low_pair(recs)

                        microwire_label = _microwire_label(draw_x, piece_y, suffix)
                        low_setpoint = (
                            low_record.metadata.setpoint_mA if low_record else None
                        )
                        high_file = (
                            high_record.metadata.file_name if high_record else ""
                        )
                        low_file = (
                            low_record.metadata.file_name if low_record else ""
                        )
                        formatted_low = _format_setpoint(low_setpoint)

                        worksheet.write_row(
                            row_idx,
                            0,
                            [
                                composition,
                                microwire_label,
                                high_file,
                                formatted_low,
                                low_file,
                                "",
                                "",
                            ],
                        )

                        worksheet.set_row(row_idx, 160)

                        def _plot(record: MeasurementRecord) -> Optional[Path]:
                            if record is None:
                                return None
                            try:
                                plot_path = _plot_measurement_matplotlib(
                                    record.dataframe,
                                    Path(record.path),
                                    plot_dir,
                                    DEFAULT_FIGSIZE,
                                )
                            except Exception:
                                self.logger.exception(
                                    "Failed to render plot for %s", record.path
                                )
                                return None
                            return plot_path

                        high_plot = _plot(high_record) if high_record else None
                        low_plot = _plot(low_record) if low_record else None
                        image_options = {"x_scale": 0.6, "y_scale": 0.6}
                        if high_plot is not None:
                            worksheet.insert_image(
                                row_idx,
                                5,
                                str(high_plot),
                                image_options,
                            )
                        if low_plot is not None:
                            worksheet.insert_image(
                                row_idx,
                                6,
                                str(low_plot),
                                image_options,
                            )
                        row_idx += 1
                finally:
                    workbook.close()
        except Exception as exc:
            try:
                if output_path.exists():
                    output_path.unlink()
            except OSError:
                pass
            QtWidgets.QMessageBox.critical(
                self,
                self.section_title,
                f"Failed to export worksheet:\n{exc}",
            )
            return

        self.log(f"Current annealing worksheet saved to {output_path}")
        QtWidgets.QMessageBox.information(
            self,
            self.section_title,
            f"Worksheet saved to {output_path}",
        )


class _MicroscopePreviewLabel(QtWidgets.QLabel):
    def __init__(self, placeholder: str, parent: Optional[QtWidgets.QWidget] = None) -> None:
        super().__init__(placeholder, parent)
        self._placeholder = placeholder
        self._pixmap: Optional[QtGui.QPixmap] = None
        self._scale_timer = QtCore.QTimer(self)
        self._scale_timer.setSingleShot(True)
        self._scale_timer.timeout.connect(self._update_scaled_pixmap)
        self.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.setWordWrap(True)

    def set_placeholder(self) -> None:
        try:
            self._scale_timer.stop()
        except Exception:
            pass
        self._pixmap = None
        super().setPixmap(QtGui.QPixmap())
        super().setText(self._placeholder)

    def set_preview(self, pixmap: Optional[QtGui.QPixmap]) -> None:
        if pixmap is None or pixmap.isNull():
            self.set_placeholder()
            return
        self._pixmap = QtGui.QPixmap(pixmap)
        self._schedule_scaled_pixmap()

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:  # pragma: no cover - Qt callback
        super().resizeEvent(event)
        self._schedule_scaled_pixmap()

    def _schedule_scaled_pixmap(self) -> None:
        if self._pixmap is None:
            return
        try:
            self._scale_timer.start(0)
        except Exception:
            self._update_scaled_pixmap()

    def _update_scaled_pixmap(self) -> None:
        if self._pixmap is None:
            return
        target_size = self.size()
        if target_size.width() <= 0 or target_size.height() <= 0:
            return
        scaled = self._pixmap.scaled(
            target_size,
            QtCore.Qt.AspectRatioMode.KeepAspectRatio,
            QtCore.Qt.TransformationMode.SmoothTransformation,
        )
        super().setPixmap(scaled)
        super().setText("")


class MicroscopeSection(MiniDatabaseSection):
    section_key = "microscope"
    section_title = "Microscope OCR"
    supported_suffixes = MICROSCOPE_EXTENSIONS
    partial_row_ready = QtCore.pyqtSignal(dict)

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._overrides: Dict[str, Dict[str, float]] = {}
        self._ocr_cache: Dict[str, MicroscopeCacheEntry] = {}
        self._validated: Dict[str, Dict[str, Any]] = {}
        self._selected_key: str | None = None
        self._ocr_debug_enabled = False
        self._pixmap_cache: Dict[Tuple[str, str], Optional[QtGui.QPixmap]] = {}
        self._expected_keys_current: Set[MicrowireKey] = set()
        self._prepopulated_keys: Set[str] = set()
        self._table_splitter: QtWidgets.QSplitter | None = None
        self._force_ocr_next = False
        self._active_column: str = ""
        self._pending_advance_key: str | None = None
        self._pending_advance_column: str | None = None
        self._pending_advance_review: bool = False
        self._pending_partial_rows: List[dict] = []
        self._pending_partial_flush = False
        super().__init__(logger, log_callback, parent)
        self._show_other_ends = bool(self.data.extra.get("show_other_ends", True))

        # Removed the missing-items list UI; missing values are visible in the table.
        self._missing_summary_label = None  # type: ignore[assignment]
        self._missing_list = None  # type: ignore[assignment]
        if hasattr(self, "controls_layout"):
            try:
                self.other_end_checkbox = QtWidgets.QCheckBox("Show other ends")
                self.other_end_checkbox.setChecked(self._show_other_ends)
                self.other_end_checkbox.toggled.connect(self._toggle_other_ends)
                self.controls_layout.addWidget(self.other_end_checkbox)
                self.defer_ocr_checkbox = QtWidgets.QCheckBox("Defer OCR")
                self.defer_ocr_checkbox.setChecked(True)
                self.controls_layout.addWidget(self.defer_ocr_checkbox)
                self.run_ocr_button = QtWidgets.QPushButton("Run OCR now")
                self.run_ocr_button.clicked.connect(self._trigger_ocr_run)
                self.controls_layout.addWidget(self.run_ocr_button)
            except Exception:
                pass

        self._load_extra_state()

        # Always normalise the table after load so legacy "Reviewed" columns
        # are removed even when there are no overrides/validations stored.
        self._apply_overrides_to_table()
        self._update_hidden_columns()
        self._update_missing_summary()
        self.partial_row_ready.connect(
            self._apply_partial_row,
            QtCore.Qt.ConnectionType.QueuedConnection,
        )
        self._update_review_buttons()
        QtCore.QTimer.singleShot(0, self._ensure_table_autosized)
        self._install_diameter_handlers()
        self.model.set_editable_columns({MICROSCOPE_D_COLUMN, MICROSCOPE_CAP_D_COLUMN})
        self.model.set_background_provider(self._background_brush_for_cell)
        self.model.set_foreground_provider(self._foreground_brush_for_cell)
        try:
            self.model.dataChanged.connect(self._handle_cell_edited)
        except Exception:
            pass

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._load_extra_state()
        self._apply_overrides_to_table()
        self._show_other_ends = bool(self.data.extra.get("show_other_ends", True))
        if hasattr(self, "other_end_checkbox"):
            self.other_end_checkbox.setChecked(self._show_other_ends)
        self._search_proxy.set_row_predicate(self._row_visible)
        self._update_hidden_columns()
        self._update_missing_summary()
        self._update_review_buttons()
        QtCore.QTimer.singleShot(0, self._ensure_table_autosized)

    def apply_data(self, data: MiniDatabaseData) -> None:  # type: ignore[override]
        super().apply_data(data)
        self._show_other_ends = bool(self.data.extra.get("show_other_ends", True))
        if hasattr(self, "other_end_checkbox"):
            self.other_end_checkbox.setChecked(self._show_other_ends)
        self._search_proxy.set_row_predicate(self._row_visible)

    def _load_extra_state(self) -> None:
        stored_overrides = self.data.extra.get("overrides")
        if isinstance(stored_overrides, dict):
            self._overrides = {
                str(key): {k: float(v) for k, v in value.items() if isinstance(v, (int, float))}
                for key, value in stored_overrides.items()
                if isinstance(value, dict)
            }

        stored_cache = self.data.extra.get("ocr_cache")
        if isinstance(stored_cache, dict):
            cache: Dict[str, MicroscopeCacheEntry] = {}
            for key, payload in stored_cache.items():
                entry: Optional[MicroscopeCacheEntry]
                if isinstance(payload, MicroscopeCacheEntry):
                    entry = payload
                elif isinstance(payload, dict):
                    entry = MicroscopeCacheEntry.from_dict(payload)
                else:
                    entry = None
                if entry is None:
                    continue
                cache[str(key)] = entry
            self._ocr_cache = cache

        stored_validated = self.data.extra.get("validated")
        if isinstance(stored_validated, dict):
            cleaned: Dict[str, Dict[str, Any]] = {}
            for key, payload in stored_validated.items():
                if not isinstance(payload, dict):
                    continue
                cleaned[str(key)] = dict(payload)
            self._validated = cleaned

    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal, parent)
        splitter.setChildrenCollapsible(False)
        splitter.setOpaqueResize(False)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)
        self._table_splitter = splitter

        table = QtWidgets.QTableView(splitter)
        table.setModel(self.model)
        header = table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
            header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
            header.setMinimumSectionSize(60)
        table.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Expanding,
            QtWidgets.QSizePolicy.Policy.Expanding,
        )
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems
        )
        table.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        table.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed
            | QtWidgets.QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        table.setSortingEnabled(True)
        v_header = table.verticalHeader()
        if v_header is not None:
            default_height = 40
            v_header.setDefaultSectionSize(default_height)
            v_header.setMinimumSectionSize(default_height)
        self.table_view = table
        selection_model = table.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._handle_selection_changed)
            selection_model.currentChanged.connect(self._handle_current_changed)
        table.installEventFilter(self)
        table.setTabKeyNavigation(False)
        self._tab_forward_shortcut = QtGui.QShortcut(
            QtGui.QKeySequence(QtCore.Qt.Key.Key_Tab),
            table,
        )
        self._tab_forward_shortcut.setContext(
            QtCore.Qt.ShortcutContext.WidgetWithChildrenShortcut
        )
        self._tab_forward_shortcut.activated.connect(
            lambda: self._advance_d_column(True)
        )
        self._tab_backward_shortcut = QtGui.QShortcut(
            QtGui.QKeySequence(QtCore.Qt.Key.Key_Backtab),
            table,
        )
        self._tab_backward_shortcut.setContext(
            QtCore.Qt.ShortcutContext.WidgetWithChildrenShortcut
        )
        self._tab_backward_shortcut.activated.connect(
            lambda: self._advance_d_column(False)
        )

        preview_container = QtWidgets.QWidget(splitter)
        preview_container.setMinimumWidth(360)
        preview_layout = QtWidgets.QVBoxLayout(preview_container)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_layout.setSpacing(6)

        scroll = QtWidgets.QScrollArea(preview_container)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Expanding,
            QtWidgets.QSizePolicy.Policy.Ignored,
        )
        scroll.setMinimumHeight(0)
        preview_layout.addWidget(scroll, 1)

        stack_widget = QtWidgets.QWidget()
        stack_layout = QtWidgets.QVBoxLayout(stack_widget)
        stack_layout.setContentsMargins(0, 0, 0, 0)
        stack_layout.setSpacing(8)

        def _make_preview_panel(title: str) -> tuple[QtWidgets.QWidget, _MicroscopePreviewLabel]:
            panel = QtWidgets.QWidget(stack_widget)
            panel.setSizePolicy(
                QtWidgets.QSizePolicy.Policy.Expanding,
                QtWidgets.QSizePolicy.Policy.Expanding,
            )
            column_layout = QtWidgets.QVBoxLayout(panel)
            column_layout.setContentsMargins(0, 0, 0, 0)
            column_layout.setSpacing(4)
            caption = QtWidgets.QLabel(title, panel)
            caption.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            caption.setStyleSheet("font-weight: 600;")
            column_layout.addWidget(caption)
            label = _MicroscopePreviewLabel("Select a row to preview.", panel)
            label.setMinimumSize(320, 240)
            label.setSizePolicy(
                QtWidgets.QSizePolicy.Policy.Expanding,
                QtWidgets.QSizePolicy.Policy.Expanding,
            )
            label.setScaledContents(False)
            column_layout.addWidget(label, 1)
            stack_layout.addWidget(panel, 1)
            return panel, label

        self.core_preview_panel, self.core_preview_label = _make_preview_panel("Core image")
        self.glass_preview_panel, self.glass_preview_label = _make_preview_panel("Glass image")
        scroll.setWidget(stack_widget)

        form = QtWidgets.QFormLayout()
        form.setContentsMargins(0, 0, 0, 0)
        form.setSpacing(4)

        self.d_edit = QtWidgets.QLineEdit()
        self.d_edit.setPlaceholderText("auto")
        self.D_edit = QtWidgets.QLineEdit()
        self.D_edit.setPlaceholderText("auto")
        self.d_edit.returnPressed.connect(partial(self._apply_override, MICROSCOPE_D_COLUMN))
        self.D_edit.returnPressed.connect(partial(self._apply_override, MICROSCOPE_CAP_D_COLUMN))
        form.addRow(MICROSCOPE_D_COLUMN, self.d_edit)
        form.addRow(MICROSCOPE_CAP_D_COLUMN, self.D_edit)
        preview_layout.addLayout(form)

        button_row = QtWidgets.QHBoxLayout()
        self.apply_override_button = QtWidgets.QPushButton("Apply override")
        self.apply_override_button.clicked.connect(self._apply_override)
        button_row.addWidget(self.apply_override_button)
        self.clear_override_button = QtWidgets.QPushButton("Clear override")
        self.clear_override_button.clicked.connect(self._clear_override)
        button_row.addWidget(self.clear_override_button)
        self.mark_reviewed_button = QtWidgets.QPushButton("Mark reviewed")
        self.mark_reviewed_button.clicked.connect(self._mark_reviewed)
        button_row.addWidget(self.mark_reviewed_button)
        self.clear_review_button = QtWidgets.QPushButton("Clear review")
        self.clear_review_button.clicked.connect(self._clear_review)
        button_row.addWidget(self.clear_review_button)
        preview_layout.addLayout(button_row)

        return splitter

    def reset_to_blank(self) -> None:  # type: ignore[override]
        super().reset_to_blank()
        self._overrides.clear()
        self._ocr_cache.clear()
        self._validated.clear()
        self._prepopulated_keys.clear()
        self._expected_keys_current = set()
        self._pixmap_cache.clear()
        self._show_other_ends = True
        if hasattr(self, "other_end_checkbox"):
            self.other_end_checkbox.setChecked(True)
        self._search_proxy.set_row_predicate(self._row_visible)
        self._update_missing_summary()
        self._update_review_buttons()

    def _toggle_other_ends(self, checked: bool) -> None:
        self._show_other_ends = bool(checked)
        self.data.extra["show_other_ends"] = self._show_other_ends
        try:
            self.store.save(self.data)
        except Exception:
            pass
        self._search_proxy.set_row_predicate(self._row_visible)

    def _row_visible(self, row: pd.Series) -> bool:  # type: ignore[override]
        if self._show_other_ends:
            return True
        microwire = str(row.get("Microwire") or "").strip()
        parsed = _microwire_parts_from_label_safe(microwire)
        if parsed is None:
            return True
        suffix = str(parsed[2] or "").strip().lower()
        return suffix != "oe"

    def _collect_candidates(self) -> List[Path]:  # type: ignore[override]
        base = MiniDatabaseSection._collect_candidates(self)
        pending: List[Path] = []
        processed = self.data.processed
        for path in base:
            key = str(path)
            try:
                mtime = path.stat().st_mtime
            except OSError:
                continue
            if float(processed.get(key, -1.0)) != float(mtime):
                pending.append(path)
        if pending:
            return pending
        return base

    def _auto_fit_columns(self) -> None:  # type: ignore[override]
        super()._auto_fit_columns()
        QtCore.QTimer.singleShot(0, self._ensure_table_autosized)

    def _ensure_table_autosized(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        try:
            table.resizeColumnsToContents()
        except Exception:
            return
        header = table.horizontalHeader()
        column_count = header.count() if header is not None else 0
        total_width = table.frameWidth() * 2
        v_header = table.verticalHeader()
        if v_header is not None:
            try:
                total_width += v_header.sizeHint().width()
            except Exception:
                total_width += v_header.width()
        for index in range(column_count):
            total_width += table.columnWidth(index)
        max_width = None
        try:
            window = self.window()
            if isinstance(window, QtWidgets.QWidget):
                screen = QtGui.QGuiApplication.screenAt(
                    window.mapToGlobal(window.rect().center())
                )
                if screen is None:
                    screen = QtGui.QGuiApplication.primaryScreen()
                if screen is not None:
                    max_width = max(640, screen.availableGeometry().width() - 200)
        except Exception:
            max_width = None
        if total_width > 0:
            preview_min = 360
            table_min = 480
            available_width = (
                max_width
                if max_width is not None
                else total_width + preview_min + 200
            )
            available_width = max(available_width, table_min + preview_min + 80)
            table_target = min(max(total_width, table_min), available_width - preview_min)
            preview_target = max(preview_min, available_width - table_target)
            if preview_target < preview_min and available_width > preview_min:
                preview_target = preview_min
                table_target = max(table_min, available_width - preview_target)
            table.setMinimumWidth(table_min)
            table.setMaximumWidth(max(table_min, available_width - preview_min))
        splitter = self._table_splitter
        if isinstance(splitter, QtWidgets.QSplitter) and column_count:
            sizes = splitter.sizes()
            total = sum(sizes) if sizes and any(sizes) else (total_width + 360)
            if max_width is not None:
                total = max(total, max_width)
            table_target = min(max(total_width, 480), total - 360)
            preview_target = total - table_target
            preview_target = max(preview_target, 360)
            splitter.setSizes([table_target, preview_target])

    def _expected_microwire_keys(self) -> Set[MicrowireKey]:
        keys: Set[MicrowireKey] = set()
        try:
            annealing_records = MiniDatabaseStore("annealing").load_payload(
                "annealing_records"
            )
        except Exception:
            annealing_records = None
        if not isinstance(annealing_records, list):
            return keys
        for record in annealing_records:
            metadata = getattr(record, "metadata", None)
            if metadata is None:
                continue
            composition = getattr(metadata, "composition_token", None)
            draw = getattr(metadata, "draw_x", None)
            piece = getattr(metadata, "piece_y", None)
            if not composition or draw is None or piece is None:
                continue
            suffix = None
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                parsed_key = _microscope_key(path)
                if parsed_key is not None:
                    _, _, _, suffix = parsed_key
            try:
                keys.add((str(composition), int(draw), int(piece), suffix))
            except (TypeError, ValueError):
                continue
        keys.update(self._extra_expected_keys())
        return keys

    def _extra_expected_keys(self) -> Set[MicrowireKey]:
        extra: Set[MicrowireKey] = set()
        tokens = set(self._overrides.keys()) | set(self._validated.keys())
        for token in tokens:
            if not token:
                continue
            parts = _microwire_key_from_string(str(token))
            if parts is not None:
                extra.add(parts)
        return extra

    def _protected_key_tokens(self) -> Set[str]:
        return {str(key) for key in self._overrides.keys()} | {str(key) for key in self._validated.keys()}

    def _prepare_initial_table(self, expected_keys: Set[MicrowireKey]) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            frame = pd.DataFrame(columns=MICROSCOPE_TABLE_COLUMNS)
        else:
            frame = frame.copy()
        if expected_keys:
            allowed = {
                _microwire_key_to_str((composition, draw, piece, suffix))
                for composition, draw, piece, suffix in expected_keys
            }
            allowed.update(self._protected_key_tokens())
            if "_key" in frame.columns:
                try:
                    mask = frame["_key"].astype(str).isin(allowed)
                    frame = frame.loc[mask].reset_index(drop=True)
                except Exception:
                    pass
            else:
                filtered_rows: List[int] = []
                for idx, row in frame.iterrows():
                    composition = str(row.get("Composition", "")).strip()
                    microwire_label = str(row.get("Microwire", "")).strip()
                    wire_tuple = _microwire_parts_from_label_safe(microwire_label)
                    if wire_tuple is None:
                        continue
                    draw, piece, suffix = wire_tuple
                    key_str = _microwire_key_to_str(
                        (composition, int(draw), int(piece), suffix)
                    )
                    if key_str in allowed:
                        filtered_rows.append(idx)
                if filtered_rows:
                    frame = frame.loc[filtered_rows].reset_index(drop=True)
        for column in MICROSCOPE_TABLE_COLUMNS:
            if column not in frame.columns:
                frame[column] = pd.Series([None] * len(frame))
        existing_keys = set(str(key) for key in frame.get("_key", []))
        new_rows: List[Dict[str, object]] = []
        for composition, draw, piece, suffix in sorted(
            expected_keys,
            key=lambda key: (
                str(key[0]).lower(),
                int(key[1]),
                int(key[2]),
                (str(key[3]).lower() if key[3] is not None else ""),
            ),
        ):
            key_str = _microwire_key_to_str((composition, draw, piece, suffix))
            if key_str in existing_keys:
                continue
            new_rows.append(
                {
                    "Composition": composition,
                    "Microwire": _microwire_label(draw, piece, suffix),
                    MICROSCOPE_D_COLUMN: None,
                    MICROSCOPE_CAP_D_COLUMN: None,
                    "d/D": None,
                    MICROSCOPE_IMAGE_COLUMNS[0]: None,
                    MICROSCOPE_IMAGE_COLUMNS[1]: None,
                    "_key": key_str,
                    "_core_image": None,
                    "_glass_image": None,
                    "_images": [],
                }
            )
        if new_rows:
            if frame.empty:
                frame = pd.DataFrame(new_rows)
            else:
                try:
                    base_frame = frame.dropna(how="all")
                except Exception:
                    base_frame = frame
                existing_rows = (
                    base_frame.to_dict(orient="records")
                    if not base_frame.empty
                    else []
                )
                frame = pd.DataFrame(existing_rows + new_rows)
        frame = frame.loc[:, MICROSCOPE_TABLE_COLUMNS]
        frame = frame.sort_values(["Composition", "Microwire"]).reset_index(drop=True)
        self.data.table = frame
        self.model.set_frame(frame)
        self._auto_fit_columns()
        self._update_missing_summary()

    @staticmethod
    def _is_valid_diameter(value: object) -> bool:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return False
        return math.isfinite(numeric) and numeric > 0

    def _validated_value(self, key: str, column: str) -> Optional[float]:
        entry = self._validated.get(key)
        if not isinstance(entry, dict):
            return None
        has_flags = "d_reviewed" in entry or "D_reviewed" in entry
        if column == MICROSCOPE_D_COLUMN:
            if has_flags and not entry.get("d_reviewed"):
                return None
            value = entry.get("d")
        elif column == MICROSCOPE_CAP_D_COLUMN:
            if has_flags and not entry.get("D_reviewed"):
                return None
            value = entry.get("D")
        else:
            return None
        if not self._is_valid_diameter(value):
            return None
        return float(value)

    def _record_to_row(
        self,
        key: MicrowireKey,
        measurement: MicroscopeMeasurements,
    ) -> Dict[str, object]:
        composition, draw, piece, suffix = key
        key_str = _microwire_key_to_str((composition, draw, piece, suffix))
        override = self._overrides.get(key_str, {})

        d_value = override.get("d")
        if d_value is None:
            core_detection = measurement.best_core_detection()
            if (
                isinstance(core_detection, MicroscopeDetection)
                and getattr(core_detection, "category", None) == "core"
                and self._is_valid_diameter(getattr(core_detection, "value", None))
            ):
                try:
                    d_value = float(core_detection.value)
                except (TypeError, ValueError):
                    d_value = None
        if d_value is None:
            validated = self._validated_value(key_str, MICROSCOPE_D_COLUMN)
            if validated is not None:
                d_value = validated

        D_value = override.get("D")
        if D_value is None:
            glass_detection = measurement.best_glass_detection()
            if (
                isinstance(glass_detection, MicroscopeDetection)
                and getattr(glass_detection, "category", None) == "glass"
                and self._is_valid_diameter(getattr(glass_detection, "value", None))
            ):
                try:
                    D_value = float(glass_detection.value)
                except (TypeError, ValueError):
                    D_value = None
        if D_value is None:
            validated = self._validated_value(key_str, MICROSCOPE_CAP_D_COLUMN)
            if validated is not None:
                D_value = validated

        ratio = None
        if isinstance(d_value, (int, float)) and isinstance(D_value, (int, float)) and D_value:
            try:
                ratio = round(float(d_value) / float(D_value), 3)
            except ZeroDivisionError:
                ratio = None

        def _first_image(entries: Sequence[MicroscopeDetection]) -> Optional[str]:
            for detection in entries:
                crop = getattr(detection, "crop_path", None)
                if crop:
                    return str(crop)
                source = getattr(detection, "image_path", None)
                if source:
                    return str(source)
            return None

        core_image = _first_image(measurement.core)
        glass_image = _first_image(measurement.glass)
        image_paths: List[str] = []
        for bucket in (measurement.core, measurement.glass, measurement.other):
            for detection in bucket:
                path = getattr(detection, "image_path", None)
                if path:
                    image_paths.append(str(path))
        if image_paths:
            image_paths = list(dict.fromkeys(image_paths))

        return {
            "Composition": composition,
            "Microwire": _microwire_label(draw, piece, suffix),
            MICROSCOPE_D_COLUMN: d_value,
            MICROSCOPE_CAP_D_COLUMN: D_value,
            "d/D": ratio,
            MICROSCOPE_IMAGE_COLUMNS[0]: None,
            MICROSCOPE_IMAGE_COLUMNS[1]: None,
            "_key": key_str,
            "_core_image": core_image,
            "_glass_image": glass_image,
            "_images": image_paths,
        }

    def _apply_partial_row(self, row: dict) -> None:
        table = self.table_view
        if (
            isinstance(table, QtWidgets.QTableView)
            and table.state() == QtWidgets.QAbstractItemView.State.EditingState
        ):
            self._pending_partial_rows.append(row)
            if not self._pending_partial_flush:
                self._pending_partial_flush = True
                QtCore.QTimer.singleShot(0, self._flush_pending_partial_rows)
            return
        self._apply_partial_row_now(row)

    def _flush_pending_partial_rows(self) -> None:
        self._pending_partial_flush = False
        table = self.table_view
        if (
            isinstance(table, QtWidgets.QTableView)
            and table.state() == QtWidgets.QAbstractItemView.State.EditingState
        ):
            self._pending_partial_flush = True
            QtCore.QTimer.singleShot(50, self._flush_pending_partial_rows)
            return
        if not self._pending_partial_rows:
            return
        pending = list(self._pending_partial_rows)
        self._pending_partial_rows.clear()
        for row in pending:
            self._apply_partial_row_now(row)

    def _apply_partial_row_now(self, row: dict) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            frame = pd.DataFrame(columns=row.keys())
        else:
            frame = frame.copy()
        key = str(row.get("_key", ""))
        if "_key" not in frame.columns:
            frame["_key"] = pd.Series([None] * len(frame))
        existing_idx = frame.index[frame["_key"] == key].tolist()
        if existing_idx:
            idx = existing_idx[0]
            for column, value in row.items():
                if column not in frame.columns:
                    frame[column] = pd.Series([None] * len(frame))
                frame.at[idx, column] = value
        else:
            if frame.empty:
                frame = pd.DataFrame([row])
            else:
                frame = pd.concat([frame, pd.DataFrame([row])], ignore_index=True, sort=False)
        for column in MICROSCOPE_TABLE_COLUMNS:
            if column not in frame.columns:
                frame[column] = pd.Series([None] * len(frame))
        frame = frame.loc[:, MICROSCOPE_TABLE_COLUMNS]
        self.data.table = frame
        self.model.set_frame(frame)
        self._auto_fit_columns()
        self._update_missing_summary()

    def _update_missing_summary(self) -> None:
        if getattr(self, "_missing_list", None) is None:
            return
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        missing_entries: List[Tuple[str, str]] = []
        missing_d = 0
        missing_D = 0
        missing_core_images = 0
        missing_glass_images = 0
        if not frame.empty and "_key" in frame.columns:
            for _, row in frame.iterrows():
                key = str(row.get("_key", ""))
                composition = str(row.get("Composition", "") or "")
                microwire = str(row.get("Microwire", "") or "")
                needs: List[str] = []
                d_value = row.get(MICROSCOPE_D_COLUMN)
                D_value = row.get(MICROSCOPE_CAP_D_COLUMN)
                if not self._is_valid_diameter(d_value):
                    missing_d += 1
                    needs.append("d")
                if not self._is_valid_diameter(D_value):
                    missing_D += 1
                    needs.append("D")
                if not row.get("_core_image"):
                    missing_core_images += 1
                if not row.get("_glass_image"):
                    missing_glass_images += 1
                if needs:
                    label = f"{composition} {microwire} (missing {', '.join(needs)})".strip()
                    missing_entries.append((key, label))
        summary_parts = []
        summary_parts.append(f"Missing d: {missing_d}")
        summary_parts.append(f"Missing D: {missing_D}")
        summary_parts.append(f"Missing core images: {missing_core_images}")
        summary_parts.append(f"Missing glass images: {missing_glass_images}")
        summary_text = " | ".join(summary_parts)
        self._missing_summary_label.setText(summary_text)
        self._missing_summary_label.setVisible(True)

        self._missing_list.clear()
        if missing_entries:
            for key, label in missing_entries:
                item = QtWidgets.QListWidgetItem(label)
                item.setData(QtCore.Qt.ItemDataRole.UserRole, key)
                self._missing_list.addItem(item)
            self._missing_list.setVisible(True)
        else:
            self._missing_list.setVisible(False)

    def _install_diameter_handlers(self) -> None:
        for edit in (self.d_edit, self.D_edit):
            try:
                edit.textEdited.connect(partial(self._normalize_decimal_input, edit))
            except Exception:
                pass
            for key, handler in (
                (QtCore.Qt.Key.Key_Up, self._select_previous_row),
                (QtCore.Qt.Key.Key_Down, self._select_next_row),
            ):
                shortcut = QtGui.QShortcut(QtGui.QKeySequence(key), edit)
                shortcut.setContext(QtCore.Qt.ShortcutContext.WidgetShortcut)
                shortcut.activated.connect(handler)

    def _normalize_decimal_input(self, edit: QtWidgets.QLineEdit, text: str) -> None:
        normalized = text.replace(",", ".")
        if normalized == text:
            return
        cursor = edit.cursorPosition()
        delta = len(normalized) - len(text)
        blocker = QtCore.QSignalBlocker(edit)
        try:
            edit.setText(normalized)
        finally:
            del blocker
        try:
            edit.setCursorPosition(max(0, cursor + delta))
        except Exception:
            pass

    def _normalized_decimal_text(self, edit: QtWidgets.QLineEdit) -> str:
        return edit.text().replace(",", ".").strip()

    def _select_previous_row(self) -> None:
        self._move_selection(-1)

    def _select_next_row(self) -> None:
        self._move_selection(1)

    def _move_selection(self, offset: int) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        model = self.table_view.model()
        selection = self.table_view.selectionModel()
        if model is None or selection is None:
            return
        row_count = model.rowCount()
        if row_count <= 0:
            return
        current_index = selection.currentIndex()
        current_row = current_index.row() if current_index.isValid() else 0
        current_col = current_index.column() if current_index.isValid() else 0
        new_row = max(0, min(row_count - 1, current_row + offset))
        target_index = model.index(new_row, max(0, current_col))
        if target_index.isValid():
            selection.setCurrentIndex(
                target_index,
                QtCore.QItemSelectionModel.SelectionFlag.ClearAndSelect,
            )
        try:
            self.table_view.scrollTo(
                model.index(new_row, max(0, current_col)),
            QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter,
        )
        except Exception:
            pass
        self._focus_d_input(select_all=True)

    def eventFilter(self, obj: QtCore.QObject, event: QtCore.QEvent) -> bool:
        if obj is self.table_view and event.type() == QtCore.QEvent.Type.KeyPress:
            key_event = cast(QtGui.QKeyEvent, event)
            key = key_event.key()
            if key in (QtCore.Qt.Key.Key_Return, QtCore.Qt.Key.Key_Enter):
                index = self._current_index()
                if index.isValid():
                    try:
                        column_label = str(self.model.frame().columns[index.column()])
                    except Exception:
                        column_label = ""
                    if column_label in {MICROSCOPE_D_COLUMN, MICROSCOPE_CAP_D_COLUMN}:
                        if (
                            isinstance(self.table_view, QtWidgets.QTableView)
                            and self.table_view.state()
                            == QtWidgets.QAbstractItemView.State.EditingState
                        ):
                            key_value = self._selected_key
                            if not key_value:
                                row = self._selected_row()
                                raw_key = row.get("_key") if row is not None else None
                                if raw_key is not None:
                                    key_value = str(raw_key)
                            self._queue_advance_after_restore(
                                key_value,
                                column_label,
                                mark_review=True,
                            )
                            return False
                        self._mark_reviewed_and_advance(column_label)
                        return True
        return super().eventFilter(obj, event)

    def _mark_reviewed_and_advance_for_key(self, key: str, column_label: str) -> None:
        self._select_row_for_key(key, column_label=column_label)
        self._mark_reviewed_and_advance(column_label)

    def _mark_reviewed_and_advance(self, column_label: str) -> None:
        index = self._current_index()
        if not index.isValid():
            return
        self._mark_reviewed(auto=True, columns={column_label})
        self._advance_after_review(index, column_label)

    def _advance_after_review(self, index: QtCore.QModelIndex, column_label: str) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        model = self.table_view.model()
        selection = self.table_view.selectionModel()
        if model is None or selection is None:
            return
        columns = list(self.model.frame().columns)
        try:
            d_col = columns.index(MICROSCOPE_D_COLUMN)
            D_col = columns.index(MICROSCOPE_CAP_D_COLUMN)
        except ValueError:
            return
        row_count = model.rowCount()
        if row_count <= 0:
            return
        target_row = index.row()
        target_col = index.column()
        if column_label == MICROSCOPE_D_COLUMN:
            target_col = D_col
        elif column_label == MICROSCOPE_CAP_D_COLUMN and target_row + 1 < row_count:
            target_row += 1
            target_col = d_col
        target = model.index(target_row, target_col)
        if not target.isValid():
            return
        selection.setCurrentIndex(
            target,
            QtCore.QItemSelectionModel.SelectionFlag.ClearAndSelect,
        )
        try:
            self.table_view.scrollTo(
                target,
                QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter,
            )
        except Exception:
            pass
        try:
            self.table_view.setFocus(QtCore.Qt.FocusReason.OtherFocusReason)
        except Exception:
            pass

    def _close_table_editor(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        editor = table.focusWidget()
        if editor is None or editor is table:
            return
        try:
            table.closeEditor(
                editor,
                QtWidgets.QAbstractItemDelegate.EndEditHint.NoHint,
            )
        except Exception:
            pass

    def _advance_d_column(self, forward: bool) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        model = self.table_view.model()
        selection = self.table_view.selectionModel()
        if model is None or selection is None:
            return
        frame = self.model.frame()
        columns = list(frame.columns) if hasattr(frame, "columns") else []
        try:
            d_col = columns.index(MICROSCOPE_D_COLUMN)
            D_col = columns.index(MICROSCOPE_CAP_D_COLUMN)
        except ValueError:
            return
        row_count = model.rowCount()
        if row_count <= 0:
            return
        current = selection.currentIndex()
        row = current.row() if current.isValid() else 0
        col = current.column() if current.isValid() else d_col
        if forward:
            if col == d_col:
                target_row, target_col = row, D_col
            elif col == D_col:
                target_row = min(row + 1, row_count - 1)
                target_col = d_col
            else:
                target_row, target_col = row, d_col
        else:
            if col == D_col:
                target_row, target_col = row, d_col
            elif col == d_col:
                target_row = max(row - 1, 0)
                target_col = D_col if row > 0 else d_col
            else:
                target_row, target_col = row, d_col
        target = model.index(target_row, target_col)
        if not target.isValid():
            return
        self._close_table_editor()
        selection.setCurrentIndex(
            target,
            QtCore.QItemSelectionModel.SelectionFlag.ClearAndSelect,
        )
        try:
            self.table_view.scrollTo(
                target,
                QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter,
            )
        except Exception:
            pass
        try:
            self.table_view.edit(target)
        except Exception:
            pass

    def _background_brush_for_cell(self, row: pd.Series, column: str) -> Optional[QtGui.QBrush]:
        missing_images = self._row_missing_images(row)
        if missing_images:
            return QtGui.QBrush(QtGui.QColor("#3a0a0a"))
        key = str(row.get("_key", ""))
        if column in {MICROSCOPE_D_COLUMN, MICROSCOPE_CAP_D_COLUMN}:
            value = row.get(column)
            if not self._is_valid_diameter(value):
                return QtGui.QBrush(QtGui.QColor("#3a0a0a"))
            reviewed = self._is_cell_reviewed(key, column)
            return QtGui.QBrush(QtGui.QColor("#0f3b26" if reviewed else "#3a0a0a"))
        return None

    def _foreground_brush_for_cell(self, row: pd.Series, column: str) -> Optional[QtGui.QBrush]:
        missing_images = self._row_missing_images(row)
        if missing_images:
            return QtGui.QBrush(QtGui.QColor("#ffd6d6"))
        key = str(row.get("_key", ""))
        if column in {MICROSCOPE_D_COLUMN, MICROSCOPE_CAP_D_COLUMN}:
            value = row.get(column)
            if not self._is_valid_diameter(value):
                return QtGui.QBrush(QtGui.QColor("#ef4444"))
            reviewed = self._is_cell_reviewed(key, column)
            return QtGui.QBrush(QtGui.QColor("#22c55e" if reviewed else "#ef4444"))
        return None

    def _is_cell_reviewed(self, key: str, column: str) -> bool:
        entry = self._validated.get(key)
        if not isinstance(entry, dict):
            return False
        has_flags = "d_reviewed" in entry or "D_reviewed" in entry
        if column == MICROSCOPE_D_COLUMN:
            reviewed = bool(entry.get("d_reviewed")) if has_flags else True
        elif column == MICROSCOPE_CAP_D_COLUMN:
            reviewed = bool(entry.get("D_reviewed")) if has_flags else True
        else:
            return False
        if not reviewed:
            return False
        row = self._row_for_key(key)
        if row is not None and not self._is_valid_diameter(row.get(column)):
            return False
        return True

    def _row_missing_images(self, row: pd.Series) -> bool:
        core_present = bool(row.get("_core_image"))
        glass_present = bool(row.get("_glass_image"))
        extras = row.get("_images")
        if not core_present and isinstance(extras, (list, tuple)) and extras:
            core_present = True
        if not glass_present and isinstance(extras, (list, tuple)) and extras:
            glass_present = True
        return not (core_present and glass_present)

    def _row_for_key(self, key: str) -> Optional[pd.Series]:
        frame = self.model.frame()
        if frame.empty or "_key" not in frame.columns:
            return None
        try:
            matches = frame.index[frame["_key"] == key].tolist()
        except Exception:
            return None
        if not matches:
            return None
        try:
            return frame.iloc[matches[0]]
        except Exception:
            return None

    def _sync_review_flags_for_row(self, key: str, row: pd.Series) -> bool:
        entry = self._validated.get(key)
        if not isinstance(entry, dict):
            return False
        changed = False
        if not self._is_valid_diameter(row.get(MICROSCOPE_D_COLUMN)):
            if "d_reviewed" in entry:
                entry.pop("d_reviewed", None)
                changed = True
            if "d" in entry:
                entry.pop("d", None)
                changed = True
        if not self._is_valid_diameter(row.get(MICROSCOPE_CAP_D_COLUMN)):
            if "D_reviewed" in entry:
                entry.pop("D_reviewed", None)
                changed = True
            if "D" in entry:
                entry.pop("D", None)
                changed = True
        if not changed:
            return False
        if not entry.get("d_reviewed") and not entry.get("D_reviewed"):
            self._validated.pop(key, None)
        else:
            self._validated[key] = entry
        return True

    def _handle_cell_edited(
        self,
        top_left: QtCore.QModelIndex,
        bottom_right: QtCore.QModelIndex,
        roles: list[int] | None = None,
    ) -> None:
        if roles and QtCore.Qt.ItemDataRole.EditRole not in roles and QtCore.Qt.ItemDataRole.DisplayRole not in roles:
            return
        frame = self.model.frame()
        self.data.table = frame.copy()
        validation_changed = False
        edited_rows = range(top_left.row(), bottom_right.row() + 1)
        for row_idx in edited_rows:
            try:
                row = frame.iloc[row_idx]
            except Exception:
                continue
            key = str(row.get("_key", ""))
            override: Dict[str, float] = {}
            d_val = row.get(MICROSCOPE_D_COLUMN)
            D_val = row.get(MICROSCOPE_CAP_D_COLUMN)
            if isinstance(d_val, (int, float)) and math.isfinite(float(d_val)):
                override["d"] = float(d_val)
            if isinstance(D_val, (int, float)) and math.isfinite(float(D_val)):
                override["D"] = float(D_val)
            if override:
                self._overrides[key] = override
            elif key in self._overrides:
                self._overrides.pop(key, None)
            if self._sync_review_flags_for_row(key, row):
                validation_changed = True
        self._store_overrides()
        if validation_changed:
            self._store_validation()

    def _queue_advance_after_restore(
        self,
        key: Optional[str],
        column_label: str,
        *,
        mark_review: bool,
    ) -> None:
        if not column_label:
            return
        self._pending_advance_key = str(key) if key else None
        self._pending_advance_column = column_label
        self._pending_advance_review = bool(mark_review)

    def _restore_selection(self) -> None:
        key = self._pending_advance_key or self._selected_key
        if not key:
            self._pending_advance_key = None
            self._pending_advance_column = None
            self._pending_advance_review = False
            return
        active_column = self._pending_advance_column or self._active_column

        def _restore() -> None:
            self._select_row_for_key(key, active_column)
            self._selected_key = key
            if self._pending_advance_column:
                column_label = self._pending_advance_column
                if self._pending_advance_review:
                    self._mark_reviewed(auto=True, columns={column_label})
                index = self._current_index()
                if index.isValid():
                    self._advance_after_review(index, column_label)
                self._pending_advance_key = None
                self._pending_advance_column = None
                self._pending_advance_review = False

        QtCore.QTimer.singleShot(0, _restore)

    def _handle_missing_item_activated(self, item: QtWidgets.QListWidgetItem) -> None:
        key = item.data(QtCore.Qt.ItemDataRole.UserRole)
        if isinstance(key, str):
            self._select_row_for_key(key)

    def _select_row_for_key(self, key: str, column_label: str | None = None) -> None:
        frame = self.model.frame()
        if frame.empty or "_key" not in frame.columns:
            return
        row_idx = None
        try:
            mask = frame["_key"] == key
            row_idx = next(
                idx for idx, matched in enumerate(mask.tolist()) if bool(matched)
            )
        except Exception:
            row_idx = None
        if row_idx is None:
            return
        if isinstance(self.table_view, QtWidgets.QTableView):
            model = self.table_view.model()
            col_idx = 0
            if column_label and hasattr(frame, "columns") and column_label in frame.columns:
                col_idx = list(frame.columns).index(column_label)
            target = model.index(row_idx, col_idx) if model is not None else None
            if target is not None and target.isValid():
                selection = self.table_view.selectionModel()
                if selection is not None:
                    selection.setCurrentIndex(
                        target,
                        QtCore.QItemSelectionModel.SelectionFlag.ClearAndSelect,
                    )
            self.table_view.scrollTo(
                self.table_view.model().index(row_idx, col_idx),
                QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter,
            )

    def set_ocr_debug_enabled(self, enabled: bool) -> None:
        self._ocr_debug_enabled = bool(enabled)

    def _ocr_debug_callback(self, path: Path, result: MicroscopeOCRResult) -> None:
        if not self._ocr_debug_enabled:
            return
        values = [
            f"{float(value):.3f}"
            for value in result.values
            if isinstance(value, (int, float)) and math.isfinite(float(value))
        ]
        value_text = ", ".join(values) if values else "—"
        sample_texts: List[str] = []
        for detection in result.detections:
            raw = getattr(detection, "text", None)
            if raw:
                cleaned = str(raw).replace("\n", " ").strip()
                if cleaned:
                    sample_texts.append(cleaned)
        for text in result.texts:
            cleaned = str(text).replace("\n", " ").strip()
            if cleaned:
                sample_texts.append(cleaned)
        text_preview = " | ".join(sample_texts) if sample_texts else "—"
        message = f"OCR debug {Path(path).name}: values={value_text}"
        message += f"; text={text_preview}"
        self.log(message, level=logging.INFO)

    def _update_hidden_columns(self) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        model = self.table_view.model()
        if model is None:
            return
        hidden_columns = ["_key", "_images", "_core_image", "_glass_image", "Reviewed"]
        hidden_columns.extend(MICROSCOPE_IMAGE_COLUMNS)
        for column_name in hidden_columns:
            try:
                column_index = list(model.frame().columns).index(column_name)  # type: ignore[arg-type]
            except Exception:
                continue
            self.table_view.setColumnHidden(column_index, True)

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        images = row.get("_images")
        if isinstance(images, (list, tuple, set)):
            for entry in images:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        for column_name in ("_core_image", "_glass_image"):
            path_value = row.get(column_name)
            if not path_value:
                continue
            try:
                candidate = Path(path_value)
            except Exception:
                continue
            if candidate not in sources:
                sources.append(candidate)
        return sources

    def _image_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column not in MICROSCOPE_IMAGE_COLUMNS:
            return None
        key = row.get("_key")
        if not isinstance(key, str) or not key:
            return None
        cache_key = (key, column)
        cached = self._pixmap_cache.get(cache_key)
        if cached is not None or cache_key in self._pixmap_cache:
            return cached
        hidden = "_core_image" if column == MICROSCOPE_IMAGE_COLUMNS[0] else "_glass_image"
        path_value = row.get(hidden)
        pixmap: Optional[QtGui.QPixmap] = None
        candidates: List[Path] = []
        if path_value:
            try:
                candidates.append(Path(path_value))
            except Exception:
                pass
        extras = row.get("_images")
        if isinstance(extras, (list, tuple)):
            for entry in extras:
                try:
                    candidates.append(Path(entry))
                except Exception:
                    continue
        if not candidates:
            candidates = self._row_sources(row)
        for candidate in candidates:
            if candidate and candidate.exists():
                reader = QtGui.QImageReader(str(candidate))
                reader.setAutoTransform(True)
                reader.setQuality(100)
                image = reader.read()
                if not image.isNull():
                    pixmap = QtGui.QPixmap.fromImage(image)
                    break
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _current_index(self) -> QtCore.QModelIndex:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return QtCore.QModelIndex()
        selection = self.table_view.selectionModel()
        if selection is None:
            return QtCore.QModelIndex()
        index = selection.currentIndex()
        return index if index.isValid() else QtCore.QModelIndex()

    def _selected_row(self) -> Optional[pd.Series]:
        index = self._current_index()
        if not index.isValid():
            return None
        try:
            return self.model.frame().iloc[index.row()]
        except Exception:
            return None

    def _focus_d_input(self, select_all: bool = False) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        try:
            table.setFocus(QtCore.Qt.FocusReason.OtherFocusReason)
            if select_all:
                index = self._current_index()
                if index.isValid():
                    table.setCurrentIndex(index)
        except Exception:
            pass

    def _handle_current_changed(self, current: QtCore.QModelIndex, _: QtCore.QModelIndex) -> None:
        self._handle_selection_changed()

    def _handle_selection_changed(self, *_: Any) -> None:
        index = self._current_index()
        row = self._selected_row()
        if row is None:
            self._selected_key = None
            self._active_column = ""
            for label in (self.core_preview_label, self.glass_preview_label):
                label.set_placeholder()
                label.show()
            if hasattr(self, "core_preview_panel") and hasattr(self, "glass_preview_panel"):
                self.core_preview_panel.show()
                self.glass_preview_panel.show()
            self.d_edit.clear()
            self.D_edit.clear()
            self._update_review_buttons()
            return
        key = row.get("_key")
        self._selected_key = str(key) if key is not None else None
        d_value = row.get(MICROSCOPE_D_COLUMN)
        D_value = row.get(MICROSCOPE_CAP_D_COLUMN)
        self.d_edit.setText("" if d_value is None or (isinstance(d_value, float) and math.isnan(d_value)) else f"{float(d_value):.3f}")
        self.D_edit.setText("" if D_value is None or (isinstance(D_value, float) and math.isnan(D_value)) else f"{float(D_value):.3f}")

        active_column = ""
        if index.isValid():
            try:
                active_column = str(self.model.frame().columns[index.column()])
            except Exception:
                active_column = ""
        self._active_column = active_column

        if active_column == MICROSCOPE_D_COLUMN:
            show_core, show_glass = True, False
        elif active_column == MICROSCOPE_CAP_D_COLUMN:
            show_core, show_glass = False, True
        else:
            show_core = show_glass = True

        for column_name, label, should_show in (
            ("_core_image", self.core_preview_label, show_core),
            ("_glass_image", self.glass_preview_label, show_glass),
        ):
            if not should_show:
                label.set_placeholder()
                continue
            path_value = row.get(column_name)
            candidate = None
            if path_value:
                try:
                    candidate = Path(path_value)
                except Exception:
                    candidate = None
            if (candidate is None or not candidate.exists()) and row.get("_images"):
                try:
                    fallback = Path(row["_images"][0])
                    if fallback.exists():
                        candidate = fallback
                except Exception:
                    candidate = None
            if candidate and candidate.exists():
                reader = QtGui.QImageReader(str(candidate))
                reader.setAutoTransform(True)
                reader.setQuality(100)
                image = reader.read()
                pixmap = QtGui.QPixmap.fromImage(image) if not image.isNull() else QtGui.QPixmap()
                if not pixmap.isNull():
                    label.set_preview(pixmap)
                    continue
            label.set_placeholder()
        if hasattr(self, "core_preview_panel") and hasattr(self, "glass_preview_panel"):
            self.core_preview_panel.setVisible(show_core)
            self.glass_preview_panel.setVisible(show_glass)
        else:
            self.core_preview_label.setVisible(show_core)
            self.glass_preview_label.setVisible(show_glass)

        self._update_review_buttons()

    def _apply_override(self, advance_column: str | None = None) -> None:
        if not self._selected_key:
            return
        if advance_column:
            self._queue_advance_after_restore(
                self._selected_key,
                advance_column,
                mark_review=False,
            )
        d_text = self._normalized_decimal_text(self.d_edit)
        D_text = self._normalized_decimal_text(self.D_edit)
        override: Dict[str, float] = {}
        if d_text:
            try:
                override["d"] = float(d_text)
            except ValueError:
                QtWidgets.QMessageBox.warning(self, self.section_title, "Invalid d value.")
                return
        if D_text:
            try:
                override["D"] = float(D_text)
            except ValueError:
                QtWidgets.QMessageBox.warning(self, self.section_title, "Invalid D value.")
                return
        if override:
            self._overrides[self._selected_key] = override
        else:
            self._overrides.pop(self._selected_key, None)
        self._store_overrides()
        columns: set[str] = set()
        if "d" in override:
            columns.add(MICROSCOPE_D_COLUMN)
        if "D" in override:
            columns.add(MICROSCOPE_CAP_D_COLUMN)
        if columns:
            self._mark_reviewed(auto=True, columns=columns)

    def _clear_override(self) -> None:
        if not self._selected_key:
            return
        if self._selected_key in self._overrides:
            self._overrides.pop(self._selected_key, None)
            self._store_overrides()
        row = self._row_for_key(self._selected_key)
        if row is not None and self._sync_review_flags_for_row(self._selected_key, row):
            self._store_validation()
        self.d_edit.clear()
        self.D_edit.clear()

    def _mark_reviewed(
        self,
        *,
        auto: bool = False,
        columns: set[str] | None = None,
        allow_without_sources: bool = True,
    ) -> None:
        if not self._selected_key:
            return
        row = self._selected_row()
        if row is None:
            return
        key = self._selected_key
        sources = self._row_sources(row)
        if not sources and not allow_without_sources:
            if auto:
                self.logger.warning("Auto-review skipped for %s due to missing sources", key)
            else:
                QtWidgets.QMessageBox.warning(
                    self,
                    self.section_title,
                    "No source images are associated with the selected row; cannot mark as reviewed.",
                )
            return

        metadata: List[Dict[str, Any]] = []
        for source in sources:
            try:
                stat = source.stat()
            except OSError:
                continue
            metadata.append(
                {
                    "path": str(source),
                    "key": self._path_key(source),
                    "mtime": float(stat.st_mtime),
                    "size": int(stat.st_size),
                }
            )
        if not metadata and not allow_without_sources:
            QtWidgets.QMessageBox.warning(
                self,
                self.section_title,
                "The source files for this row are not accessible; review status was not updated.",
            )
            return

        existing = self._validated.get(key, {})
        entry: Dict[str, Any] = dict(existing) if isinstance(existing, dict) else {}
        if metadata:
            entry["sources"] = metadata
        else:
            entry.setdefault("sources", [])
        if allow_without_sources and not sources:
            entry["allow_without_sources"] = True

        d_value = row.get(MICROSCOPE_D_COLUMN)
        if isinstance(d_value, (int, float)) and math.isfinite(float(d_value)):
            entry["d"] = float(d_value)
        D_value = row.get(MICROSCOPE_CAP_D_COLUMN)
        if isinstance(D_value, (int, float)) and math.isfinite(float(D_value)):
            entry["D"] = float(D_value)

        columns_to_mark = columns or {MICROSCOPE_D_COLUMN, MICROSCOPE_CAP_D_COLUMN}
        if MICROSCOPE_D_COLUMN in columns_to_mark:
            entry["d_reviewed"] = True
        if MICROSCOPE_CAP_D_COLUMN in columns_to_mark:
            entry["D_reviewed"] = True

        entry["timestamp"] = datetime.utcnow().isoformat() + "Z"
        self._validated[key] = entry
        self._store_validation()
        override_changed = False
        override_entry = dict(self._overrides.get(key, {}))
        if MICROSCOPE_D_COLUMN in columns_to_mark and self._is_valid_diameter(d_value):
            if override_entry.get("d") != float(d_value):
                override_entry["d"] = float(d_value)
                override_changed = True
        if MICROSCOPE_CAP_D_COLUMN in columns_to_mark and self._is_valid_diameter(D_value):
            if override_entry.get("D") != float(D_value):
                override_entry["D"] = float(D_value)
                override_changed = True
        if override_changed:
            self._overrides[key] = override_entry
            self._store_overrides()

    def _clear_review(self) -> None:
        if not self._selected_key:
            return
        if self._selected_key in self._validated:
            self._validated.pop(self._selected_key, None)
            self._store_validation()
        self._update_review_buttons()

    def _store_validation(self, persist: bool = True) -> None:
        self.data.extra["validated"] = self._validated
        if persist:
            self.store.save(self.data)
        self._refresh_review_display()
        self._update_review_buttons()
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _refresh_review_display(self) -> None:
        try:
            rows = self.model.rowCount()
            cols = self.model.columnCount()
        except Exception:
            rows = cols = 0
        if rows <= 0 or cols <= 0:
            return
        try:
            self.model.dataChanged.emit(
                self.model.index(0, 0),
                self.model.index(rows - 1, cols - 1),
                [
                    QtCore.Qt.ItemDataRole.BackgroundRole,
                    QtCore.Qt.ItemDataRole.ForegroundRole,
                ],
            )
        except Exception:
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass
        if isinstance(self.table_view, QtWidgets.QTableView):
            try:
                self.table_view.viewport().update()
            except Exception:
                pass

    def _refresh_validations(self) -> None:
        changed = False
        for key, payload in list(self._validated.items()):
            entry_changed = False
            sources = payload.get("sources")
            allow_without_sources = bool(payload.get("allow_without_sources"))
            if not isinstance(sources, list) or (not sources and not allow_without_sources):
                continue
            else:
                for source in sources:
                    path_text = source.get("path")
                    if not path_text:
                        continue
                    path_obj = Path(path_text)
                    try:
                        stat = path_obj.stat()
                    except OSError:
                        continue
                    if float(source.get("mtime", -1.0)) != float(stat.st_mtime):
                        source["mtime"] = float(stat.st_mtime)
                        entry_changed = True
                    if int(source.get("size", -1)) != int(stat.st_size):
                        source["size"] = int(stat.st_size)
                        entry_changed = True
                    source.setdefault("key", self._path_key(path_obj))
            if entry_changed:
                self._validated[key] = payload
                changed = True
        if changed:
            self._store_validation()

    def _update_review_buttons(self) -> None:
        if not hasattr(self, "mark_reviewed_button"):
            return
        has_selection = bool(self._selected_key)
        key = self._selected_key or ""
        has_any_review = bool(has_selection and self._validated.get(key))
        if has_selection:
            d_reviewed = self._is_cell_reviewed(key, MICROSCOPE_D_COLUMN)
            D_reviewed = self._is_cell_reviewed(key, MICROSCOPE_CAP_D_COLUMN)
            is_reviewed = d_reviewed and D_reviewed
        else:
            is_reviewed = False
        self.mark_reviewed_button.setEnabled(has_selection)
        self.mark_reviewed_button.setText("Update review" if is_reviewed else "Mark reviewed")
        self.clear_review_button.setEnabled(has_any_review)

    def _trigger_ocr_run(self) -> None:
        self._force_ocr_next = True
        self.refresh()

    def _store_overrides(self) -> None:
        self.data.extra["overrides"] = self._overrides
        self.store.save(self.data)
        self._apply_overrides_to_table()
        self._update_hidden_columns()
        self._update_missing_summary()
        try:
            self.data_updated.emit()
        except Exception:
            pass
        self._restore_selection()
        self._ensure_table_autosized()

    def _apply_overrides_to_table(self) -> None:
        frame = self.data.table.copy()
        frame = frame.drop(columns=["Reviewed"], errors="ignore")
        if frame.empty:
            self.model.set_frame(frame)
            return
        self._pixmap_cache.clear()
        for index, row in frame.iterrows():
            key = str(row.get("_key"))
            override = self._overrides.get(key)
            d_value = row.get(MICROSCOPE_D_COLUMN)
            D_value = row.get(MICROSCOPE_CAP_D_COLUMN)
            if override:
                if "d" in override:
                    d_value = override.get("d")
                if "D" in override:
                    D_value = override.get("D")
            if not self._is_valid_diameter(d_value):
                validated = self._validated_value(key, MICROSCOPE_D_COLUMN)
                if validated is not None:
                    d_value = validated
            if not self._is_valid_diameter(D_value):
                validated = self._validated_value(key, MICROSCOPE_CAP_D_COLUMN)
                if validated is not None:
                    D_value = validated
            ratio = None
            if isinstance(d_value, (int, float)) and isinstance(D_value, (int, float)) and D_value:
                try:
                    ratio = float(d_value) / float(D_value)
                except ZeroDivisionError:
                    ratio = None
            frame.at[index, MICROSCOPE_D_COLUMN] = d_value
            frame.at[index, MICROSCOPE_CAP_D_COLUMN] = D_value
            frame.at[index, "d/D"] = round(ratio, 3) if ratio is not None else None
        self.data.table = frame
        self.model.set_frame(frame)
        self._auto_fit_columns()
        self._update_missing_summary()
        self._update_review_buttons()
        self._restore_selection()

    def refresh(self) -> None:
        self._refresh_validations()
        self._expected_keys_current = self._expected_microwire_keys()
        if self._expected_keys_current:
            self._prepare_initial_table(self._expected_keys_current)
        super().refresh()
        if self.data.table.empty and self._expected_keys_current:
            self._prepare_initial_table(self._expected_keys_current)
        self._apply_overrides_to_table()
        self._update_hidden_columns()
        self._update_missing_summary()
        self._update_review_buttons()

    def _prepopulate_image_refs(self, candidates: Iterable[Path]) -> None:
        expected = self._expected_keys_current or self._expected_microwire_keys()
        allowed: Set[str] | None = None
        if expected:
            allowed = set()
            for item in expected:
                composition: object
                draw: object
                piece: object
                suffix: object | None
                try:
                    composition, draw, piece, suffix = item  # type: ignore[misc]
                except Exception:
                    try:
                        composition, draw, piece = item  # type: ignore[misc]
                    except Exception:
                        continue
                    suffix = None
                try:
                    key_token = _microwire_key_to_str(
                        (str(composition), int(draw), int(piece), suffix)
                    )
                except Exception:
                    continue
                allowed.add(key_token)
        grouped: Dict[str, Dict[str, Any]] = {}
        for path in candidates:
            key_tuple = _microscope_key(path)
            if key_tuple is None:
                continue
            composition, draw, piece, suffix = key_tuple
            key = _microwire_key_to_str((composition, draw, piece, suffix))
            if allowed is not None and key not in allowed:
                continue
            entry = grouped.setdefault(
                key,
                {
                    "_key": key,
                    "Composition": composition,
                    "Microwire": _microwire_label(draw, piece, suffix),
                    "_images": [],
                },
            )
            images: List[str] = entry.setdefault("_images", [])  # type: ignore[assignment]
            image_path = str(path)
            if image_path not in images:
                images.append(image_path)
            category = _microscope_category(path)
            if category == "core" and not entry.get("_core_image"):
                entry["_core_image"] = image_path
            elif category == "glass" and not entry.get("_glass_image"):
                entry["_glass_image"] = image_path

        for key, payload in grouped.items():
            images_list = payload.get("_images", [])
            if isinstance(images_list, list):
                payload["_images"] = list(dict.fromkeys(images_list))
            row = {
                "Composition": payload.get("Composition", ""),
                "Microwire": payload.get("Microwire", ""),
                MICROSCOPE_D_COLUMN: None,
                MICROSCOPE_CAP_D_COLUMN: None,
                "d/D": None,
                MICROSCOPE_IMAGE_COLUMNS[0]: None,
                MICROSCOPE_IMAGE_COLUMNS[1]: None,
                "_key": key,
                "_core_image": payload.get("_core_image"),
                "_glass_image": payload.get("_glass_image"),
                "_images": payload.get("_images", []),
            }
            self._pixmap_cache.pop((key, MICROSCOPE_IMAGE_COLUMNS[0]), None)
            self._pixmap_cache.pop((key, MICROSCOPE_IMAGE_COLUMNS[1]), None)
            self._apply_partial_row(row)
            self._prepopulated_keys.add(key)

        if grouped:
            self._update_missing_summary()

    def _start_section_worker(self, candidates: List[Path]) -> None:  # type: ignore[override]
        if candidates:
            self._prepopulate_image_refs(candidates)
        super()._start_section_worker(candidates)

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        self._refresh_validations()
        unique_paths = list(dict.fromkeys(Path(p) for p in paths))
        run_ocr = self._force_ocr_next or not getattr(self, "defer_ocr_checkbox", QtWidgets.QCheckBox()).isChecked()
        self._force_ocr_next = False

        def _progress(idx: int, total: int) -> None:
            self._check_cancelled()
            if progress is None:
                return
            message = None
            if 0 < idx <= len(unique_paths):
                try:
                    message = f"Grouping {unique_paths[idx - 1].name}"
                except Exception:
                    message = None
            try:
                progress(idx, total, message)
            except Exception:
                pass

        debug_cb = self._ocr_debug_callback if self._ocr_debug_enabled else None
        expected_keys = self._expected_keys_current or self._expected_microwire_keys()
        discovered_keys: Set[MicrowireKey] = set()
        for candidate in unique_paths:
            try:
                parsed_key = _microscope_key(Path(candidate))
            except Exception:
                parsed_key = None
            if parsed_key is not None:
                discovered_keys.add(parsed_key)
        if discovered_keys:
            expected_keys = set(expected_keys) | discovered_keys

        def _emit_partial(key: MicrowireKey, measurement: MicroscopeMeasurements) -> None:
            try:
                row = self._record_to_row(key, measurement)
            except Exception:
                return
            try:
                self.partial_row_ready.emit(row)
            except Exception:
                pass

        cache_lookup: Dict[str, MicroscopeCacheEntry] = {}
        for candidate in unique_paths:
            path_obj = Path(candidate)
            key_tuple = _microscope_key(path_obj)
            if key_tuple is None:
                continue
            comp, draw, piece, suffix = key_tuple
            if expected_keys and (comp, draw, piece, suffix) not in expected_keys:
                if (comp, draw, piece, None) not in expected_keys:
                    continue
            key = _microwire_key_to_str((comp, draw, piece, suffix))
            validated_entry = self._validated.get(key)
            if not isinstance(validated_entry, dict):
                continue
            sources = validated_entry.get("sources")
            if not isinstance(sources, list):
                continue
            path_key = self._path_key(path_obj)
            if not any(
                isinstance(source, dict)
                and source.get("path")
                and self._path_key(Path(source.get("path"))) == path_key
                for source in sources
            ):
                continue
            cache_entry = self._ocr_cache.get(path_key)
            if cache_entry is None:
                continue
            cache_lookup[path_key] = cache_entry

        if not run_ocr:
            self._prepare_initial_table(expected_keys)
            processed: Dict[str, float] = {}
            for path in unique_paths:
                try:
                    processed[str(path)] = float(path.stat().st_mtime)
                except OSError:
                    continue
            return SectionProcessResult(
                table=self.data.table,
                processed=processed,
                payloads=self.data.extra.get("payloads", {}),
                extra=self.data.extra,
            )

        index, cache_map = _group_microscope_measurements(
            unique_paths,
            self.logger,
            progress_callback=_progress if progress is not None else None,
            debug_callback=debug_cb,
            update_callback=_emit_partial,
            cache=cache_lookup,
        )
        self._ocr_cache = dict(cache_map)
        if expected_keys:
            for key in expected_keys:
                index.setdefault(key, MicroscopeMeasurements())
        self._check_cancelled()
        filtered_overrides = {
            key: value
            for key, value in self._overrides.items()
            if any(
                key == _microwire_key_to_str((comp, draw, piece, suffix))
                for comp, draw, piece, suffix in index.keys()
            )
        }
        self._overrides = filtered_overrides
        table = _microscope_index_to_frame(index, filtered_overrides)

        processed: Dict[str, float] = {}
        for path in unique_paths:
            try:
                processed[str(path)] = float(path.stat().st_mtime)
            except OSError:
                continue

        def _count_measurements(entries: Sequence[MicroscopeDetection]) -> int:
            count = 0
            for entry in entries:
                value = getattr(entry, "value", None)
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    numeric = math.nan
                if math.isfinite(numeric) and numeric > 0:
                    count += 1
            return count

        total_records = len(index)
        total_core = sum(_count_measurements(m.core) for m in index.values())
        total_glass = sum(_count_measurements(m.glass) for m in index.values())
        if total_core or total_glass:
            self.log(
                f"Microscope OCR detected {total_core} core and {total_glass} glass diameter(s) across {total_records} microwire(s).",
                level=logging.INFO,
            )
        else:
            self.log(
                "Microscope OCR completed but no diameters were detected. Ensure the PaddleOCR models are installed and the microscope captures contain visible annotations.",
                level=logging.WARNING,
            )
        cache_payload = {key: entry.as_dict() for key, entry in self._ocr_cache.items()}
        extra_payload = {
            "overrides": filtered_overrides,
            "ocr_cache": cache_payload,
            "validated": self._validated,
        }
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"microscope_index": index},
            extra=extra_payload,
        )

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._update_missing_summary()
        self._update_review_buttons()

    @property
    def overrides(self) -> Dict[str, Dict[str, float]]:
        return dict(self._overrides)


class _CurrentDensityPreviewPanel(QtWidgets.QWidget):
    valuePicked = QtCore.pyqtSignal(str, float)
    def __init__(self, logger: logging.Logger, parent: QtWidgets.QWidget | None = None) -> None:
        super().__init__(parent)
        self._logger = logger
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.header_label = QtWidgets.QLabel("Select a row to preview annealing plots.", self)
        self.header_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        layout.addWidget(self.header_label)

        self._high_display = _AnnealingPlotDisplay("Graph — 1000 mA", logger, self)
        self._low_display = _AnnealingPlotDisplay("Graph — low mA", logger, self)
        layout.addWidget(self._high_display, 1)
        layout.addWidget(self._low_display, 1)
        layout.setStretch(1, 1)
        layout.setStretch(2, 1)
        self._high_display.valuePicked.connect(lambda value: self.valuePicked.emit("As", value))
        self._low_display.valuePicked.connect(lambda value: self.valuePicked.emit("Ms", value))

    def update_selection(
        self,
        key: Optional[MicrowireKey],
        high: Optional[MeasurementRecord],
        low: Optional[MeasurementRecord],
    ) -> None:
        if key is None:
            self.header_label.setText("Select a row to preview annealing plots.")
            self._high_display.clear("Select a row to view the 1000 mA measurement.")
            self._low_display.clear("Select a row to view the low-current measurement.")
            return
        composition, draw, piece, suffix = key
        try:
            microwire = _microwire_label(draw, piece, suffix)
        except Exception:
            microwire = f"{draw}/{piece}"
        self.header_label.setText(f"{composition} — {microwire}")

        self._high_display.set_record(
            high,
            setpoint=_extract_setpoint(high),
            description="No 1000 mA measurement available for this microwire.",
        )
        self._low_display.set_record(
            low,
            setpoint=_extract_setpoint(low),
            description="No lower-current measurement available for this microwire.",
        )


class CurrentDensitySection(QtWidgets.QWidget):
    section_key = "current_density"
    section_title = "Current density"

    status_changed = QtCore.pyqtSignal(str)
    sources_changed = QtCore.pyqtSignal(list)
    data_updated = QtCore.pyqtSignal()

    def __init__(
        self,
        annealing_section: AnnealingSection,
        microscope_section: MicroscopeSection,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.logger = logger
        self._log_callback = log_callback
        self._annealing_section = annealing_section
        self._microscope_section = microscope_section
        self._current_frame = pd.DataFrame(columns=CURRENT_DENSITY_COLUMNS)
        self._last_sources: List[str] = []
        self._table_splitter: QtWidgets.QSplitter | None = None
        self._preview_panel: _CurrentDensityPreviewPanel | None = None
        self._search_proxy = _TableSearchProxyModel(self)
        self.search_edit: QtWidgets.QLineEdit | None = None
        self.search_clear_button: QtWidgets.QPushButton | None = None

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        controls = QtWidgets.QHBoxLayout()
        self.refresh_button = QtWidgets.QPushButton("Recalculate")
        self.refresh_button.clicked.connect(self.refresh_data)
        controls.addWidget(self.refresh_button)
        self.export_button = QtWidgets.QPushButton("Export worksheet...")
        self.export_button.clicked.connect(self._export_worksheet)
        self.export_button.setEnabled(False)
        controls.addWidget(self.export_button)
        controls.addStretch(1)
        layout.addLayout(controls)

        search_row = QtWidgets.QHBoxLayout()
        search_row.addWidget(QtWidgets.QLabel("Search:"))
        self.search_edit = QtWidgets.QLineEdit(self)
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.setPlaceholderText("Filter rows across visible columns")
        self.search_edit.textChanged.connect(self._handle_search_changed)
        search_row.addWidget(self.search_edit, 1)
        self.search_clear_button = QtWidgets.QPushButton("Clear")
        self.search_clear_button.setEnabled(False)
        self.search_clear_button.clicked.connect(lambda: self.search_edit.clear())
        search_row.addWidget(self.search_clear_button)
        layout.addLayout(search_row)

        self.status_label = QtWidgets.QLabel("Waiting for data.", self)
        layout.addWidget(self.status_label)

        self.model = DataFrameModel(self._current_frame)
        self.model.set_editable_columns(set(PHASE_POINT_COLUMN_MAP.values()))
        self.model.dataChanged.connect(self._handle_model_data_changed)
        self._search_proxy.setSourceModel(self.model)

        splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal, self)
        splitter.setChildrenCollapsible(False)
        splitter.setOpaqueResize(False)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
        layout.addWidget(splitter, 1)
        self._table_splitter = splitter

        table = QtWidgets.QTableView(splitter)
        table.setModel(self.model)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setSortingEnabled(True)
        table.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel)
        table.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel)
        table.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.SelectedClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed
        )
        header = table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
            header.setSectionsMovable(True)
            header.setSectionsClickable(True)
        vertical_bar = table.verticalScrollBar()
        if vertical_bar is not None:
            vertical_bar.setSingleStep(MiniDatabaseSection._SCROLL_SINGLE_STEP)
        table.setModel(self._search_proxy)
        splitter.addWidget(table)
        self.table_view = table

        selection_model = table.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._handle_selection_changed)

        preview_panel = _CurrentDensityPreviewPanel(logger, splitter)
        splitter.addWidget(preview_panel)
        self._preview_panel = preview_panel
        preview_panel.valuePicked.connect(self._apply_picked_value)

        if hasattr(self._annealing_section, "data_updated"):
            try:
                self._annealing_section.data_updated.connect(self.refresh_data)
            except Exception:
                pass
        if hasattr(self._microscope_section, "data_updated"):
            try:
                self._microscope_section.data_updated.connect(self.refresh_data)
            except Exception:
                pass
        QtCore.QTimer.singleShot(0, self.refresh_data)

    def log(self, message: str, level: int = logging.INFO) -> None:
        try:
            self._log_callback(level, message)
        except Exception:
            self.logger.log(level, message)

    def refresh_data(self) -> None:
        previous_order = self._current_column_order()
        selected_key = self._current_selection_key()
        try:
            frame = self._calculate_frame()
        except Exception:
            self.logger.exception("Failed to calculate current density table")
            self.status_label.setText("Failed to calculate current density.")
            self.export_button.setEnabled(False)
            return
        self._current_frame = frame
        self.model.set_frame(frame)
        if previous_order:
            QtCore.QTimer.singleShot(
                0,
                lambda: (
                    self._apply_column_order(previous_order),
                    self._hide_internal_columns(),
                ),
            )
        else:
            self._hide_internal_columns()
        try:
            self.table_view.resizeColumnsToContents()
        except Exception:
            pass
        self._restore_selection(selected_key)
        self._update_preview()
        total = len(frame.index) if isinstance(frame, pd.DataFrame) else 0
        annotated = 0
        if (
            total
            and isinstance(frame, pd.DataFrame)
            and ANNEALING_AS_COLUMN in frame.columns
            and ANNEALING_MS_COLUMN in frame.columns
        ):
            as_series = pd.to_numeric(frame[ANNEALING_AS_COLUMN], errors="coerce")
            ms_series = pd.to_numeric(frame[ANNEALING_MS_COLUMN], errors="coerce")
            annotated = int((as_series.notna() & ms_series.notna()).sum())
        status_text = (
            f"{annotated} of {total} microwire(s) have As1/Ms1 annotated."
            if total
            else "No overlapping microscope and annealing data yet."
        )
        self.status_label.setText(status_text)
        self.export_button.setEnabled(total > 0)
        try:
            self.status_changed.emit(status_text)
        except Exception:
            pass
        try:
            self.sources_changed.emit(list(self._last_sources))
        except Exception:
            pass
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def current_density_snapshot(self) -> Dict[str, Dict[str, Any]]:
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return {}
        snapshot: Dict[str, Dict[str, Any]] = {}
        for _, row in frame.iterrows():
            key_text = str(row.get("_group_key") or "").strip()
            if not key_text:
                key_tuple = self._extract_key(
                    row.get("Composition"),
                    row.get("Microwire"),
                    None,
                )
                if key_tuple:
                    key_text = _microwire_key_to_str(key_tuple)
            if not key_text:
                continue
            entry: Dict[str, Any] = {}
            for column in frame.columns:
                if column == "_group_key":
                    continue
                entry[column] = row.get(column)
            if entry:
                snapshot[key_text] = entry
        return snapshot

    def _hide_internal_columns(self) -> None:
        table = self.table_view
        frame = self.model.frame()
        if not isinstance(table, QtWidgets.QTableView):
            return
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        try:
            columns = list(frame.columns)
        except Exception:
            return
        for idx, column in enumerate(columns):
            try:
                table.setColumnHidden(idx, str(column).startswith("_"))
            except Exception:
                continue

    def _current_selection_key(self) -> Optional[MicrowireKey]:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return None
        selection_model = table.selectionModel()
        if selection_model is None:
            return None
        current_index = selection_model.currentIndex()
        if current_index.isValid():
            row_index = self._source_row(current_index.row())
        else:
            rows = selection_model.selectedRows()
            if not rows:
                return None
            row_index = self._source_row(rows[0].row())
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or row_index < 0 or row_index >= len(frame.index):
            return None
        key_value = frame.iloc[row_index].get("_group_key")
        return self._parse_group_key(key_value)

    def _restore_selection(self, key: Optional[MicrowireKey]) -> None:
        if key is None:
            return
        table = self.table_view
        frame = self.model.frame()
        if not isinstance(table, QtWidgets.QTableView):
            return
        if not isinstance(frame, pd.DataFrame) or "_group_key" not in frame.columns:
            return
        target = _microwire_key_to_str(key)
        try:
            matches = frame.index[frame["_group_key"] == target].tolist()
        except Exception:
            matches = []
        if not matches:
            return
        row = matches[0]
        try:
            source_index = self.model.index(int(row), 0)
            index = self._search_proxy.mapFromSource(source_index)
            if index.isValid():
                table.selectRow(index.row())
                table.scrollTo(index, QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter)
        except Exception:
            pass

    def _update_preview(self) -> None:
        panel = self._preview_panel
        if panel is None:
            return
        key = self._current_selection_key()
        if key is None:
            panel.update_selection(None, None, None)
            return
        high, low = self._fetch_records_for_key(key)
        panel.update_selection(key, high, low)

    def _handle_selection_changed(self, *_args: Any) -> None:
        self._update_preview()

    def _handle_model_data_changed(
        self,
        top_left: QtCore.QModelIndex,
        bottom_right: QtCore.QModelIndex,
        roles: Tuple[QtCore.Qt.ItemDataRole, ...] = (),
    ) -> None:
        if roles and QtCore.Qt.ItemDataRole.EditRole not in roles:
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        try:
            columns_slice = frame.columns[top_left.column() : bottom_right.column() + 1]
        except Exception:
            columns_slice = []
        relevant = set(PHASE_POINT_COLUMN_MAP.values())
        if not any(column in relevant for column in columns_slice):
            return
        setter = getattr(self._annealing_section, "set_phase_points_for_key", None)
        updated = False
        setter_used = False
        for row in range(top_left.row(), bottom_right.row() + 1):
            if row < 0 or row >= len(frame.index):
                continue
            series = frame.iloc[row]
            key_raw = series.get("_group_key")
            key = str(key_raw).strip() if key_raw not in (None, "", float("nan")) else ""
            if not key:
                continue
            phase_values = {
                label: self._coerce_phase_value(series.get(column))
                for label, column in PHASE_POINT_COLUMN_MAP.items()
            }
            try:
                if callable(setter):
                    setter(key, phase_values=phase_values)
                    setter_used = True
                else:
                    phase_points = getattr(self._annealing_section, "_phase_points", {})
                    if isinstance(phase_points, dict):
                        entry: Dict[str, float] = {}
                        for label, value in phase_values.items():
                            if value is not None:
                                entry[label] = value
                        if "As1" in entry:
                            entry["As"] = entry["As1"]
                        if "Ms1" in entry:
                            entry["Ms"] = entry["Ms1"]
                        if entry:
                            phase_points[key] = entry
                        elif key in phase_points:
                            phase_points.pop(key, None)
                        store = getattr(self._annealing_section, "_store_phase_points", None)
                        if callable(store):
                            store()
                        updated_signal = getattr(self._annealing_section, "data_updated", None)
                        if hasattr(updated_signal, "emit"):
                            try:
                                updated_signal.emit()
                            except Exception:
                                pass
                updated = True
            except Exception:
                self.logger.exception("Failed to persist phase transition points for %s", key)
        if updated and not setter_used:
            QtCore.QTimer.singleShot(0, self.refresh_data)

    def _column_index_for_kind(self, kind: str) -> Optional[int]:
        target_column = ANNEALING_AS_COLUMN if kind == "As" else ANNEALING_MS_COLUMN
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame):
            return None
        try:
            return int(frame.columns.get_loc(target_column))
        except Exception:
            return None

    def _apply_picked_value(self, kind: str, value: float) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        selection_model = table.selectionModel()
        if selection_model is None:
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        current_index = selection_model.currentIndex()
        column_index = None
        if current_index.isValid():
            source_row = self._source_row(current_index.row())
            try:
                current_label = str(frame.columns[current_index.column()])
            except Exception:
                current_label = ""
            if current_label in PHASE_POINT_COLUMN_MAP.values():
                column_index = current_index.column()
        if column_index is None:
            column_index = self._column_index_for_kind(kind)
        if column_index is None:
            return
        row = source_row if current_index.isValid() else None
        if row is None:
            rows = selection_model.selectedRows()
            if rows:
                row = self._source_row(rows[0].row())
        if row is None or row < 0 or row >= len(frame.index):
            return
        target_index = current_index if (current_index.isValid() and current_index.column() == column_index) else self.model.index(row, column_index)
        if not target_index.isValid():
            return
        if not self.model.setData(target_index, float(value)):
            return
        try:
            table.setCurrentIndex(target_index)
            table.scrollTo(target_index, QtWidgets.QAbstractItemView.ScrollHint.EnsureVisible)
        except Exception:
            pass
        self._update_preview()

    def _source_row(self, proxy_row: int) -> Optional[int]:
        return self._search_proxy.map_row_to_source(proxy_row)

    def _handle_search_changed(self, text: str) -> None:
        self._search_proxy.set_search_text(text)
        if isinstance(self.search_clear_button, QtWidgets.QPushButton):
            self.search_clear_button.setEnabled(bool(str(text).strip()))

    @staticmethod
    def _coerce_phase_value(value: Any) -> Optional[float]:
        if value is None:
            return None
        if hasattr(pd, "isna"):
            try:
                if pd.isna(value):
                    return None
            except Exception:
                pass
        if isinstance(value, (int, float)):
            numeric = float(value)
            return numeric if math.isfinite(numeric) else None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                numeric = float(text)
            except ValueError:
                return None
            return numeric if math.isfinite(numeric) else None
        return None

    def _fetch_records_for_key(
        self,
        key: MicrowireKey,
    ) -> Tuple[Optional[MeasurementRecord], Optional[MeasurementRecord]]:
        key_str = _microwire_key_to_str(key)
        groups = getattr(self._annealing_section, "_record_groups", {})
        if not isinstance(groups, dict):
            return None, None
        records = groups.get(key_str, [])
        if not records:
            parts = _split_microwire_key(key)
            if parts is not None:
                composition, draw, piece, _suffix = parts
                base_key = _microwire_key_to_str((composition, draw, piece, None))
                records = groups.get(base_key, [])
        if not records:
            return None, None
        return _select_high_low_pair(records)

    @staticmethod
    def _parse_group_key(value: object) -> Optional[MicrowireKey]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return _microwire_key_from_string(text)

    def _calculate_frame(self) -> pd.DataFrame:
        diameter_map = self._collect_microscope_data()
        setpoint_map = self._collect_setpoint_data()
        phase_map = self._collect_phase_points()
        keys = sorted(
            set(setpoint_map.keys()) | set(phase_map.keys()),
            key=lambda item: (
                str(item[0]).lower(),
                int(item[1]),
                int(item[2]),
                (str(item[3]).lower() if item[3] is not None else ""),
            ),
        )
        rows: List[Dict[str, Any]] = []
        all_sources: Set[str] = set()
        if not keys:
            columns = CURRENT_DENSITY_COLUMNS + ["_group_key"]
            return pd.DataFrame(columns=columns)
        for composition, draw, piece, suffix in keys:
            key = (composition, draw, piece, suffix)
            base_key = (composition, draw, piece, None)
            micro_info = diameter_map.get(key) or diameter_map.get(base_key, {})
            setpoint_info = setpoint_map.get(key) or setpoint_map.get(base_key, {})
            phase_info = phase_map.get(key) or phase_map.get(base_key, {})
            diameter_um = micro_info.get("diameter")
            area_mm2 = self._diameter_to_area(diameter_um)
            setpoints = setpoint_info.get("setpoints", [])
            sources = setpoint_info.get("sources", [])
            all_sources.update(str(source) for source in sources)
            notes: List[str] = []
            if diameter_um is None or area_mm2 is None:
                notes.append("Missing diameter")
            composition_label = micro_info.get("composition") or setpoint_info.get("composition") or composition
            try:
                microwire_label = micro_info.get("label") or _microwire_label(draw, piece, suffix)
            except Exception:
                microwire_label = micro_info.get("label") or f"{draw}/{piece}"
            as1_value = phase_info.get("As1")
            if as1_value is None:
                as1_value = phase_info.get("As")
            af1_value = phase_info.get("Af1")
            ms1_value = phase_info.get("Ms1")
            if ms1_value is None:
                ms1_value = phase_info.get("Ms")
            mf1_value = phase_info.get("Mf1")
            as2_value = phase_info.get("As2")
            af2_value = phase_info.get("Af2")
            ms2_value = phase_info.get("Ms2")
            mf2_value = phase_info.get("Mf2")
            as_density = self._compute_density(as1_value, area_mm2)
            ms_density = self._compute_density(ms1_value, area_mm2)
            as_delta = self._compute_delta(as2_value, as1_value)
            af_delta = self._compute_delta(af2_value, af1_value)
            ms_delta = self._compute_delta(ms2_value, ms1_value)
            mf_delta = self._compute_delta(mf2_value, mf1_value)
            mf1_af1 = self._compute_delta(mf1_value, af1_value)
            mf2_af2 = self._compute_delta(mf2_value, af2_value)
            if as1_value is None:
                notes.append("As1 missing")
            if ms1_value is None:
                notes.append("Ms1 missing")
            if not setpoints:
                notes.append("No setpoint data")
            rows.append(
                {
                    "Composition": composition_label,
                    "Microwire": microwire_label,
                    MICROSCOPE_D_COLUMN: diameter_um,
                    ANNEALING_AS_COLUMN: as1_value,
                    ANNEALING_AF1_COLUMN: af1_value,
                    ANNEALING_MS_COLUMN: ms1_value,
                    ANNEALING_MF1_COLUMN: mf1_value,
                    ANNEALING_AS2_COLUMN: as2_value,
                    ANNEALING_AF2_COLUMN: af2_value,
                    ANNEALING_MS2_COLUMN: ms2_value,
                    ANNEALING_MF2_COLUMN: mf2_value,
                    CURRENT_DENSITY_AS_DENSITY_COLUMN: as_density,
                    CURRENT_DENSITY_MS_DENSITY_COLUMN: ms_density,
                    CURRENT_DENSITY_AS_DELTA_COLUMN: as_delta,
                    CURRENT_DENSITY_AF_DELTA_COLUMN: af_delta,
                    CURRENT_DENSITY_MS_DELTA_COLUMN: ms_delta,
                    CURRENT_DENSITY_MF_DELTA_COLUMN: mf_delta,
                    CURRENT_DENSITY_MF_AF1_DELTA_COLUMN: mf1_af1,
                    CURRENT_DENSITY_MF_AF2_DELTA_COLUMN: mf2_af2,
                    "Setpoints (mA)": self._format_setpoints(setpoints),
                    "Sources": self._summarise_sources(sources),
                    "Notes": "; ".join(notes) if notes else "",
                    "_group_key": _microwire_key_to_str(key),
                }
            )
        self._last_sources = sorted(all_sources)
        if not rows:
            columns = CURRENT_DENSITY_COLUMNS + ["_group_key"]
            return pd.DataFrame(columns=columns)
        frame = pd.DataFrame(rows)
        desired_order = [column for column in CURRENT_DENSITY_COLUMNS + ["_group_key"] if column in frame.columns]
        return frame.loc[:, desired_order]

    def _collect_microscope_data(self) -> Dict[MicrowireKey, Dict[str, Any]]:
        result: Dict[MicrowireKey, Dict[str, Any]] = {}
        table = getattr(getattr(self._microscope_section, "data", None), "table", None)
        if not isinstance(table, pd.DataFrame) or table.empty:
            return result
        for _, row in table.iterrows():
            key = self._extract_key(row.get("Composition"), row.get("Microwire"), row.get("_key"))
            if key is None:
                continue
            composition, draw, piece, suffix = key
            diameter = self._to_positive_float(row.get(MICROSCOPE_D_COLUMN))
            composition_label = self._normalise_text(row.get("Composition")) or composition
            label = self._normalise_text(row.get("Microwire"))
            if not label:
                try:
                    label = _microwire_label(draw, piece, suffix)
                except Exception:
                    label = f"{draw}/{piece}"
            result[key] = {
                "composition": composition_label,
                "label": label,
                "diameter": diameter,
            }
        return result

    def _collect_setpoint_data(self) -> Dict[MicrowireKey, Dict[str, Any]]:
        result: Dict[MicrowireKey, Dict[str, Any]] = {}
        groups = getattr(self._annealing_section, "_record_groups", {})
        if not isinstance(groups, dict) or not groups:
            return result
        for records in groups.values():
            for record in records:
                metadata = getattr(record, "metadata", None)
                if metadata is None:
                    continue
                composition = getattr(metadata, "composition_token", None)
                draw = getattr(metadata, "draw_x", None)
                piece = getattr(metadata, "piece_y", None)
                setpoint = getattr(metadata, "setpoint_mA", None)
                if (
                    composition is None
                    or draw is None
                    or piece is None
                    or setpoint is None
                ):
                    continue
                suffix = None
                path = getattr(record, "path", None)
                if isinstance(path, Path):
                    parsed_key = _microscope_key(path)
                    if parsed_key is not None:
                        _, _, _, suffix = parsed_key
                try:
                    key = (
                        str(composition),
                        int(draw),
                        int(piece),
                        str(suffix).strip() or None,
                    )
                    setpoint_value = float(setpoint)
                except (TypeError, ValueError):
                    continue
                if not math.isfinite(setpoint_value):
                    continue
                entry = result.setdefault(
                    key,
                    {
                        "setpoints": [],
                        "sources": set(),
                        "composition": str(composition),
                    },
                )
                entry["setpoints"].append(setpoint_value)
                path = getattr(record, "path", None)
                if path:
                    try:
                        entry["sources"].add(str(Path(path)))
                    except Exception:
                        pass
        for entry in result.values():
            entry["setpoints"] = sorted(dict.fromkeys(entry["setpoints"]))
            entry["sources"] = sorted(entry["sources"])
        return result

    def _collect_phase_points(self) -> Dict[MicrowireKey, Dict[str, float]]:
        result: Dict[MicrowireKey, Dict[str, float]] = {}
        snapshot_provider = getattr(self._annealing_section, "phase_points_snapshot", None)
        if callable(snapshot_provider):
            raw = snapshot_provider()
        else:
            raw = getattr(self._annealing_section, "_phase_points", {})
        if not isinstance(raw, dict):
            return result
        for key, payload in raw.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            parts = _microwire_key_from_string(key)
            if parts is None:
                continue
            composition, draw, piece, suffix = parts
            cleaned: Dict[str, float] = {}
            for label in PHASE_POINT_LABELS:
                try:
                    numeric = float(payload.get(label))
                except (TypeError, ValueError):
                    continue
                if math.isfinite(numeric):
                    cleaned[label] = numeric
            if "As1" not in cleaned:
                try:
                    numeric = float(payload.get("As"))
                except (TypeError, ValueError):
                    numeric = None
                if isinstance(numeric, (int, float)) and math.isfinite(float(numeric)):
                    cleaned["As1"] = float(numeric)
            if "Ms1" not in cleaned:
                try:
                    numeric = float(payload.get("Ms"))
                except (TypeError, ValueError):
                    numeric = None
                if isinstance(numeric, (int, float)) and math.isfinite(float(numeric)):
                    cleaned["Ms1"] = float(numeric)
            if cleaned:
                result[(composition, draw, piece, suffix)] = cleaned
        return result

    @staticmethod
    def _normalise_text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, (float, int)):
            try:
                if math.isnan(float(value)):
                    return ""
            except Exception:
                pass
        text = str(value).strip()
        if not text or text.lower() in {"nan", "none"}:
            return ""
        return text

    def _extract_key(
        self,
        composition: object,
        microwire: object,
        stored_key: object,
    ) -> Optional[MicrowireKey]:
        comp_text = self._normalise_text(composition)
        if isinstance(stored_key, str):
            parts = _microwire_key_from_string(stored_key)
            if parts is not None:
                base_comp, draw, piece, suffix = parts
                comp_value = comp_text or base_comp
                if comp_value:
                    return (comp_value, draw, piece, suffix)
        label_text = self._normalise_text(microwire)
        if label_text and comp_text:
            draw_piece = _microwire_parts_from_label_safe(label_text)
            if draw_piece is not None:
                draw, piece, suffix = draw_piece
                return (comp_text, int(draw), int(piece), suffix)
        return None

    @staticmethod
    def _to_positive_float(value: object) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            numeric = float(value)
        else:
            text = str(value).strip().replace(",", ".")
            if not text:
                return None
            try:
                numeric = float(text)
            except ValueError:
                return None
        if not math.isfinite(numeric) or numeric <= 0:
            return None
        return numeric

    @staticmethod
    def _diameter_to_area(diameter_um: Optional[float]) -> Optional[float]:
        if diameter_um is None or not math.isfinite(diameter_um) or diameter_um <= 0:
            return None
        radius_mm = (diameter_um * 1e-3) / 2.0
        if radius_mm <= 0:
            return None
        return math.pi * radius_mm * radius_mm

    @staticmethod
    def _compute_density(setpoint_mA: Optional[float], area_mm2: Optional[float]) -> Optional[float]:
        if setpoint_mA is None or area_mm2 is None or area_mm2 <= 0:
            return None
        current_A = setpoint_mA / 1000.0
        return current_A / area_mm2

    @staticmethod
    def _compute_delta(value_2: Optional[float], value_1: Optional[float]) -> Optional[float]:
        if value_2 is None or value_1 is None:
            return None
        if not (math.isfinite(value_2) and math.isfinite(value_1)):
            return None
        return value_2 - value_1

    @staticmethod
    def _format_setpoints(values: Sequence[float]) -> str:
        if not values:
            return ""
        formatted = [f"{value:.3f}".rstrip("0").rstrip(".") if isinstance(value, float) else str(value) for value in values]
        return ", ".join(formatted)

    @staticmethod
    def _summarise_sources(sources: Iterable[str]) -> str:
        unique = []
        for entry in sources:
            text = str(entry)
            if text and text not in unique:
                unique.append(text)
        names = [Path(text).name for text in unique]
        return ", ".join(names)

    def _current_column_order(self) -> List[str]:
        header = self.table_view.horizontalHeader()
        frame = self.model.frame()
        if header is None or not isinstance(frame, pd.DataFrame):
            return []
        order: List[str] = []
        for visual_index in range(header.count()):
            logical = header.logicalIndex(visual_index)
            try:
                column = str(frame.columns[logical])
            except Exception:
                continue
            order.append(column)
        return order

    def _apply_column_order(self, order: Sequence[str]) -> None:
        if not order:
            return
        header = self.table_view.horizontalHeader()
        frame = self.model.frame()
        if header is None or not isinstance(frame, pd.DataFrame):
            return
        mapping = {str(column): idx for idx, column in enumerate(frame.columns)}
        for target_visual, column_name in enumerate(order):
            logical = mapping.get(column_name)
            if logical is None:
                continue
            current_visual = header.visualIndex(logical)
            if current_visual == target_visual:
                continue
            header.moveSection(current_visual, target_visual)

    def _ordered_export_frame(self) -> pd.DataFrame:
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame):
            return pd.DataFrame(columns=CURRENT_DENSITY_COLUMNS)
        order = self._current_column_order()
        if order:
            missing = [column for column in frame.columns if column not in order]
            order = list(order) + missing
            export_frame = frame.loc[:, order].copy()
        else:
            export_frame = frame.copy()
        export_frame = export_frame.loc[
            :,
            [column for column in export_frame.columns if not str(column).startswith("_")],
        ]
        return export_frame

    def _export_worksheet(self) -> None:
        frame = self._ordered_export_frame()
        if frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Export worksheet",
                "There is no current density data to export yet.",
            )
            return
        start_dir = _dialog_start_directory()
        suggested = start_dir / "current_density.xlsx"
        path_str, selected_filter = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export worksheet",
            str(suggested),
            "Excel files (*.xlsx);;CSV files (*.csv)",
        )
        if not path_str:
            return
        path = Path(path_str)
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".csv"}:
            if "Excel" in selected_filter:
                path = path.with_suffix(".xlsx")
                suffix = ".xlsx"
            else:
                path = path.with_suffix(".csv")
                suffix = ".csv"
        export_frame = frame.copy()
        for column in export_frame.columns:
            series = export_frame[column]
            if getattr(series, "dtype", None) == object:
                export_frame[column] = series.map(
                    lambda value: "" if isinstance(value, (QtGui.QPixmap, QtGui.QImage)) else value
                )
        try:
            if suffix == ".xlsx":
                export_frame.to_excel(path, index=False)
            else:
                export_frame.to_csv(path, index=False)
        except Exception as exc:
            self.logger.exception("Failed to export current density worksheet")
            QtWidgets.QMessageBox.critical(
                self,
                "Export worksheet",
                f"Failed to export worksheet:\n{exc}",
            )
            return
        self.log(f"Current density worksheet exported to {path}")
        QtWidgets.QMessageBox.information(
            self,
            "Export worksheet",
            f"Worksheet exported to:\n{path}",
        )


class _TransitionTempPreviewPanel(QtWidgets.QWidget):
    valuePicked = QtCore.pyqtSignal(str, float)

    def __init__(
        self,
        logger: logging.Logger,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._logger = logger
        self._processor = _get_vsm_temp_processor(logger)
        self._canvas_connections: List[Tuple[FigureCanvasQTAgg, Optional[int], Optional[int]]] = []
        self._cursor_units = "°C"
        self._target_buttons: Dict[str, QtWidgets.QRadioButton] = {}
        self._value_labels: Dict[str, QtWidgets.QLabel] = {}

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self.header_label = QtWidgets.QLabel("Select a row to preview VSM temperature scans.")
        self.header_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        layout.addWidget(self.header_label)

        self._stack = QtWidgets.QStackedLayout()
        self._placeholder = QtWidgets.QLabel(
            "Select a row to preview VSM temperature scans."
        )
        self._placeholder.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setWordWrap(True)
        self._tab_widget = QtWidgets.QTabWidget(self)
        self._stack.addWidget(self._placeholder)
        self._stack.addWidget(self._tab_widget)
        layout.addLayout(self._stack, 1)

        controls = QtWidgets.QHBoxLayout()
        self.cursor_label = QtWidgets.QLabel("Cursor: —")
        self.cursor_label.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        controls.addWidget(self.cursor_label)
        controls.addSpacing(20)

        for label in TRANSITION_TEMP_LABELS:
            radio = QtWidgets.QRadioButton(f"Set {label}")
            if label == "As":
                radio.setChecked(True)
            self._target_buttons[label] = radio
            controls.addWidget(radio)
        controls.addSpacing(20)

        for label in TRANSITION_TEMP_LABELS:
            value_label = QtWidgets.QLabel("unset")
            self._value_labels[label] = value_label
            controls.addWidget(QtWidgets.QLabel(f"{label}:"))
            controls.addWidget(value_label)
        controls.addStretch(1)
        layout.addLayout(controls)

    def set_target(self, label: Optional[str]) -> None:
        if not label:
            return
        button = self._target_buttons.get(label)
        if button is not None:
            button.setChecked(True)

    def update_selection(
        self,
        title: str,
        records: Sequence[VsmTemperatureScanRecord],
        values: Mapping[str, float],
    ) -> None:
        current_index = self._tab_widget.currentIndex() if self._tab_widget.count() else 0
        self.header_label.setText(title or "Transition temps")
        self._update_value_labels(values)
        self._clear_tabs()

        if self._processor is None:
            self._placeholder.setText("VSM temperature scan parser is unavailable.")
            self._stack.setCurrentWidget(self._placeholder)
            return
        if not records:
            self._placeholder.setText("No VSM temperature scans available for this microwire.")
            self._stack.setCurrentWidget(self._placeholder)
            return
        for record in records:
            figure = _plot_vsm_temperature_scan_figure(
                record,
                self._processor,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            if figure is None:
                continue
            canvas = FigureCanvasQTAgg(figure)
            try:
                canvas.setMouseTracking(True)
            except Exception:
                pass
            click_cid = None
            motion_cid = None
            try:
                click_cid = canvas.mpl_connect("button_press_event", self._handle_click)
            except Exception:
                click_cid = None
            try:
                motion_cid = canvas.mpl_connect("motion_notify_event", self._handle_motion)
            except Exception:
                motion_cid = None
            self._canvas_connections.append((canvas, motion_cid, click_cid))
            label = _record_label_for_display(record) or record.sample or "Scan"
            page = QtWidgets.QWidget(self._tab_widget)
            page_layout = QtWidgets.QVBoxLayout(page)
            page_layout.setContentsMargins(0, 0, 0, 0)
            page_layout.addWidget(canvas, 1)
            self._tab_widget.addTab(page, label)
        if self._tab_widget.count() == 0:
            self._placeholder.setText("No VSM temperature scans available for this microwire.")
            self._stack.setCurrentWidget(self._placeholder)
        else:
            if current_index >= 0:
                self._tab_widget.setCurrentIndex(
                    min(current_index, self._tab_widget.count() - 1)
                )
            self._stack.setCurrentWidget(self._tab_widget)

    def _current_target(self) -> str:
        for label, button in self._target_buttons.items():
            if button.isChecked():
                return label
        return "As"

    def _clear_tabs(self) -> None:
        for canvas, motion_cid, click_cid in self._canvas_connections:
            if motion_cid is not None:
                try:
                    canvas.mpl_disconnect(motion_cid)
                except Exception:
                    pass
            if click_cid is not None:
                try:
                    canvas.mpl_disconnect(click_cid)
                except Exception:
                    pass
        self._canvas_connections.clear()
        while self._tab_widget.count():
            widget = self._tab_widget.widget(0)
            self._tab_widget.removeTab(0)
            if widget is not None:
                widget.setParent(None)
                widget.deleteLater()
        self._update_cursor_label(None)

    def _handle_motion(self, event: Any) -> None:
        if event is None or event.inaxes is None or event.xdata is None:
            self._update_cursor_label(None)
            return
        try:
            value = float(event.xdata)
        except Exception:
            self._update_cursor_label(None)
            return
        self._update_cursor_label(value)

    def _handle_click(self, event: Any) -> None:
        if event is None or not getattr(event, "dblclick", False):
            return
        if event.xdata is None:
            return
        try:
            value = float(event.xdata)
        except Exception:
            return
        target = self._current_target()
        self._update_cursor_label(value)
        try:
            self.valuePicked.emit(target, value)
        except Exception:
            pass

    def _update_cursor_label(self, value: Optional[float]) -> None:
        if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            text = "Cursor: —"
        else:
            formatted = f"{float(value):.3f}".rstrip("0").rstrip(".")
            suffix = f" {self._cursor_units}" if self._cursor_units else ""
            text = f"Cursor: {formatted}{suffix}"
        self.cursor_label.setText(text)

    @staticmethod
    def _format_value(value: Optional[float]) -> str:
        if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            return "unset"
        return f"{float(value):.3f}".rstrip("0").rstrip(".")

    def _update_value_labels(self, values: Mapping[str, float]) -> None:
        for label, widget in self._value_labels.items():
            widget.setText(self._format_value(values.get(label)))


class TransitionTempsSection(QtWidgets.QWidget):
    section_key = "transition_temps"
    section_title = "Transition temps"

    status_changed = QtCore.pyqtSignal(str)
    sources_changed = QtCore.pyqtSignal(list)
    data_updated = QtCore.pyqtSignal()

    def __init__(
        self,
        vsm_temperature_section: VsmTemperatureScanSection,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.logger = logger
        self._log_callback = log_callback
        self._vsm_temperature_section = vsm_temperature_section
        self.store = MiniDatabaseStore(self.section_key)
        self.data = self.store.load()
        self._transition_points = self._load_transition_points()
        self._record_groups: Dict[str, List[VsmTemperatureScanRecord]] = {}
        self._last_sources: List[str] = []
        self._current_frame = pd.DataFrame(columns=TRANSITION_TEMP_COLUMNS + ["_group_key"])
        self.model = DataFrameModel(self._current_frame)
        self.model.set_editable_columns(set(TRANSITION_TEMP_COLUMN_MAP.values()))
        self.model.dataChanged.connect(self._handle_model_data_changed)
        self._preview_panel: _TransitionTempPreviewPanel | None = None
        self._search_proxy = _TableSearchProxyModel(self)
        self.search_edit: QtWidgets.QLineEdit | None = None
        self.search_clear_button: QtWidgets.QPushButton | None = None
        self._search_proxy.setSourceModel(self.model)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        controls = QtWidgets.QHBoxLayout()
        self.refresh_button = QtWidgets.QPushButton("Recalculate")
        self.refresh_button.clicked.connect(self.refresh_data)
        controls.addWidget(self.refresh_button)
        self.export_button = QtWidgets.QPushButton("Export worksheet...")
        self.export_button.clicked.connect(self._export_worksheet)
        self.export_button.setEnabled(False)
        controls.addWidget(self.export_button)
        controls.addStretch(1)
        layout.addLayout(controls)

        search_row = QtWidgets.QHBoxLayout()
        search_row.addWidget(QtWidgets.QLabel("Search:"))
        self.search_edit = QtWidgets.QLineEdit(self)
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.setPlaceholderText("Filter rows across visible columns")
        self.search_edit.textChanged.connect(self._handle_search_changed)
        search_row.addWidget(self.search_edit, 1)
        self.search_clear_button = QtWidgets.QPushButton("Clear")
        self.search_clear_button.setEnabled(False)
        self.search_clear_button.clicked.connect(lambda: self.search_edit.clear())
        search_row.addWidget(self.search_clear_button)
        layout.addLayout(search_row)

        self.status_label = QtWidgets.QLabel("Waiting for VSM temperature scan data.", self)
        layout.addWidget(self.status_label)

        splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal, self)
        splitter.setChildrenCollapsible(False)
        splitter.setOpaqueResize(False)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
        layout.addWidget(splitter, 1)

        table = QtWidgets.QTableView(splitter)
        table.setModel(self._search_proxy)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setSortingEnabled(True)
        table.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel)
        table.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel)
        table.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.SelectedClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed
        )
        header = table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
            header.setSectionsMovable(True)
            header.setSectionsClickable(True)
        vertical_bar = table.verticalScrollBar()
        if vertical_bar is not None:
            vertical_bar.setSingleStep(MiniDatabaseSection._SCROLL_SINGLE_STEP)
        splitter.addWidget(table)
        self.table_view = table

        selection_model = table.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._handle_selection_changed)

        preview_panel = _TransitionTempPreviewPanel(logger, splitter)
        splitter.addWidget(preview_panel)
        self._preview_panel = preview_panel
        preview_panel.valuePicked.connect(self._apply_picked_value)

        if hasattr(self._vsm_temperature_section, "data_updated"):
            try:
                self._vsm_temperature_section.data_updated.connect(self.refresh_data)
            except Exception:
                pass
        QtCore.QTimer.singleShot(0, self.refresh_data)

    def log(self, message: str, level: int = logging.INFO) -> None:
        try:
            self._log_callback(level, message)
        except Exception:
            self.logger.log(level, message)

    def _load_transition_points(self) -> Dict[str, Dict[str, float]]:
        stored = self.data.extra.get("transition_temps")
        if not isinstance(stored, dict):
            return {}
        cleaned: Dict[str, Dict[str, float]] = {}
        for key, payload in stored.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            entry: Dict[str, float] = {}
            for label in TRANSITION_TEMP_LABELS:
                value = payload.get(label)
                if isinstance(value, (int, float)) and math.isfinite(float(value)):
                    entry[label] = float(value)
            if entry:
                cleaned[key] = entry
        return cleaned

    def _store_transition_points(self) -> None:
        snapshot: Dict[str, Dict[str, float]] = {}
        for key, payload in self._transition_points.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            entry: Dict[str, float] = {}
            for label in TRANSITION_TEMP_LABELS:
                value = payload.get(label)
                if isinstance(value, (int, float)) and math.isfinite(float(value)):
                    entry[label] = float(value)
            if entry:
                snapshot[key] = entry
        self.data.extra["transition_temps"] = snapshot
        self.data.table = self.model.frame()
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist transition temps")

    def transition_points_snapshot(self) -> Dict[str, Dict[str, float]]:
        snapshot: Dict[str, Dict[str, float]] = {}
        for key, payload in self._transition_points.items():
            if not isinstance(key, str) or not isinstance(payload, dict):
                continue
            cleaned: Dict[str, float] = {}
            for label in TRANSITION_TEMP_LABELS:
                value = payload.get(label)
                if isinstance(value, (int, float)) and math.isfinite(float(value)):
                    cleaned[label] = float(value)
            if cleaned:
                snapshot[key] = cleaned
        return snapshot

    def refresh_data(self) -> None:
        previous_order = self._current_column_order()
        selected_key = self._current_selection_key()
        self._refresh_record_groups()
        valid_keys = set(self._record_groups.keys())
        if self._prune_transition_points(valid_keys):
            self._store_transition_points()
        frame = self._build_frame()
        self._current_frame = frame
        self.model.set_frame(frame)
        if previous_order:
            QtCore.QTimer.singleShot(
                0,
                lambda: (
                    self._apply_column_order(previous_order),
                    self._hide_internal_columns(),
                ),
            )
        else:
            self._hide_internal_columns()
        try:
            self.table_view.resizeColumnsToContents()
        except Exception:
            pass
        self._restore_selection(selected_key)
        self._update_preview()

        total = len(frame.index) if isinstance(frame, pd.DataFrame) else 0
        annotated = 0
        if total and isinstance(frame, pd.DataFrame):
            series = []
            for column in TRANSITION_TEMP_COLUMN_MAP.values():
                series.append(pd.to_numeric(frame.get(column), errors="coerce"))
            if series:
                stacked = pd.concat(series, axis=1)
                annotated = int(stacked.notna().any(axis=1).sum())
        status_text = (
            f"{annotated} of {total} sample(s) have transition temps annotated."
            if total
            else "No VSM temperature scan data yet."
        )
        self.status_label.setText(status_text)
        self.export_button.setEnabled(total > 0)
        try:
            self.status_changed.emit(status_text)
        except Exception:
            pass
        try:
            self.sources_changed.emit(list(self._last_sources))
        except Exception:
            pass
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[VsmTemperatureScanRecord]] = {}
        payload = None
        try:
            payload = self._vsm_temperature_section.store.load_payload(
                "vsm_temperature_scan_records"
            )
        except Exception:
            payload = None
        if isinstance(payload, list):
            hidden = _hidden_paths_from_section(self._vsm_temperature_section)
            visible_records = [
                record
                for record in payload
                if _record_path_key(record) not in hidden
            ]
            grouped = _group_graph_records_by_key(visible_records)
            for records in grouped.values():
                records.sort(key=_record_label_for_display)
        self._record_groups = grouped
        sources: List[str] = []
        for records in grouped.values():
            for record in records:
                path = getattr(record, "path", None)
                if isinstance(path, Path):
                    sources.append(str(path))
        self._last_sources = list(dict.fromkeys(sources))

    def _build_frame(self) -> pd.DataFrame:
        rows: List[Dict[str, Any]] = []
        for key in sorted(self._record_groups.keys()):
            key_tuple = self._parse_group_key(key)
            if key_tuple is None:
                continue
            composition, microwire = _microwire_info_from_key(key_tuple)
            values = self._transition_points.get(key, {})
            rows.append(
                {
                    "Composition": composition,
                    "Microwire": microwire,
                    TRANSITION_TEMP_AS_COLUMN: values.get("As"),
                    TRANSITION_TEMP_AF_COLUMN: values.get("Af"),
                    TRANSITION_TEMP_MS_COLUMN: values.get("Ms"),
                    TRANSITION_TEMP_MF_COLUMN: values.get("Mf"),
                    "_group_key": key,
                }
            )
        if not rows:
            return pd.DataFrame(columns=TRANSITION_TEMP_COLUMNS + ["_group_key"])
        frame = pd.DataFrame(rows)
        desired = [col for col in TRANSITION_TEMP_COLUMNS + ["_group_key"] if col in frame.columns]
        return frame.loc[:, desired]

    def _hide_internal_columns(self) -> None:
        table = self.table_view
        frame = self.model.frame()
        if not isinstance(table, QtWidgets.QTableView):
            return
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        try:
            columns = list(frame.columns)
        except Exception:
            return
        for idx, column in enumerate(columns):
            try:
                table.setColumnHidden(idx, str(column).startswith("_"))
            except Exception:
                continue

    def _current_selection_key(self) -> Optional[str]:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return None
        selection_model = table.selectionModel()
        if selection_model is None:
            return None
        current_index = selection_model.currentIndex()
        if current_index.isValid():
            row_index = self._source_row(current_index.row())
        else:
            rows = selection_model.selectedRows()
            if not rows:
                return None
            row_index = self._source_row(rows[0].row())
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or row_index < 0 or row_index >= len(frame.index):
            return None
        key_value = frame.iloc[row_index].get("_group_key")
        if key_value in (None, "", float("nan")):
            return None
        return str(key_value)

    def _restore_selection(self, key: Optional[str]) -> None:
        if not key:
            return
        table = self.table_view
        frame = self.model.frame()
        if not isinstance(table, QtWidgets.QTableView):
            return
        if not isinstance(frame, pd.DataFrame) or "_group_key" not in frame.columns:
            return
        try:
            matches = frame.index[frame["_group_key"] == key].tolist()
        except Exception:
            matches = []
        if not matches:
            return
        row = matches[0]
        try:
            source_index = self.model.index(int(row), 0)
            index = self._search_proxy.mapFromSource(source_index)
            if index.isValid():
                table.selectRow(index.row())
                table.scrollTo(index, QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter)
        except Exception:
            pass

    def _update_preview(self) -> None:
        panel = self._preview_panel
        if panel is None:
            return
        key = self._current_selection_key()
        if not key:
            panel.update_selection(
                "Select a row to preview VSM temperature scans.",
                [],
                {},
            )
            return
        key_tuple = self._parse_group_key(key)
        if key_tuple is None:
            panel.update_selection("Transition temps", [], {})
            return
        composition, microwire = _microwire_info_from_key(key_tuple)
        title = f"{composition} — {microwire}" if composition and microwire else "Transition temps"
        records = self._record_groups.get(key, [])
        values = self._transition_points.get(key, {})
        panel.update_selection(title, records, values)

    def _handle_selection_changed(self, *_args: Any) -> None:
        self._update_preview()
        self._sync_preview_target()

    def _sync_preview_target(self) -> None:
        panel = self._preview_panel
        if panel is None:
            return
        table = self.table_view
        frame = self.model.frame()
        if not isinstance(table, QtWidgets.QTableView) or not isinstance(frame, pd.DataFrame):
            return
        index = table.currentIndex()
        if not index.isValid():
            return
        try:
            column_label = str(frame.columns[index.column()])
        except Exception:
            return
        for label, column in TRANSITION_TEMP_COLUMN_MAP.items():
            if column == column_label:
                panel.set_target(label)
                break

    def _handle_model_data_changed(
        self,
        top_left: QtCore.QModelIndex,
        bottom_right: QtCore.QModelIndex,
        roles: Tuple[QtCore.Qt.ItemDataRole, ...] = (),
    ) -> None:
        if roles and QtCore.Qt.ItemDataRole.EditRole not in roles:
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        try:
            columns_slice = frame.columns[top_left.column() : bottom_right.column() + 1]
        except Exception:
            columns_slice = []
        relevant = set(TRANSITION_TEMP_COLUMN_MAP.values())
        if not any(column in relevant for column in columns_slice):
            return
        updated = False
        for row in range(top_left.row(), bottom_right.row() + 1):
            if row < 0 or row >= len(frame.index):
                continue
            series = frame.iloc[row]
            key_raw = series.get("_group_key")
            key = str(key_raw).strip() if key_raw not in (None, "", float("nan")) else ""
            if not key:
                continue
            entry: Dict[str, float] = {}
            for label, column in TRANSITION_TEMP_COLUMN_MAP.items():
                value = self._coerce_transition_value(series.get(column))
                if value is not None:
                    entry[label] = value
            if entry:
                self._transition_points[key] = entry
            elif key in self._transition_points:
                self._transition_points.pop(key, None)
            updated = True
        if updated:
            self._store_transition_points()
            self._update_preview()

    @staticmethod
    def _coerce_transition_value(value: Any) -> Optional[float]:
        if value is None:
            return None
        if hasattr(pd, "isna"):
            try:
                if pd.isna(value):
                    return None
            except Exception:
                pass
        if isinstance(value, (int, float)):
            numeric = float(value)
            return numeric if math.isfinite(numeric) else None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                numeric = float(text)
            except ValueError:
                return None
            return numeric if math.isfinite(numeric) else None
        return None

    def _apply_picked_value(self, label: str, value: float) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        selection_model = table.selectionModel()
        if selection_model is None:
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        column_label = TRANSITION_TEMP_COLUMN_MAP.get(label)
        if not column_label:
            return
        try:
            column_index = int(frame.columns.get_loc(column_label))
        except Exception:
            return
        current_index = selection_model.currentIndex()
        row = self._source_row(current_index.row()) if current_index.isValid() else None
        if row is None:
            rows = selection_model.selectedRows()
            if rows:
                row = self._source_row(rows[0].row())
        if row is None or row < 0 or row >= len(frame.index):
            return
        target_index = (
            current_index
            if current_index.isValid() and current_index.column() == column_index
            else self.model.index(row, column_index)
        )
        if not target_index.isValid():
            return
        if not self.model.setData(target_index, float(value)):
            return
        try:
            table.setCurrentIndex(target_index)
            table.scrollTo(target_index, QtWidgets.QAbstractItemView.ScrollHint.EnsureVisible)
        except Exception:
            pass
        self._update_preview()

    def _source_row(self, proxy_row: int) -> Optional[int]:
        return self._search_proxy.map_row_to_source(proxy_row)

    def _handle_search_changed(self, text: str) -> None:
        self._search_proxy.set_search_text(text)
        if isinstance(self.search_clear_button, QtWidgets.QPushButton):
            self.search_clear_button.setEnabled(bool(str(text).strip()))

    @staticmethod
    def _parse_group_key(value: object) -> Optional[MicrowireKey]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return _microwire_key_from_string(text)

    def _prune_transition_points(self, valid_keys: Iterable[str]) -> bool:
        valid_set = {str(value) for value in valid_keys if value}
        removed = False
        for key in list(self._transition_points.keys()):
            if key not in valid_set:
                self._transition_points.pop(key, None)
                removed = True
        return removed

    def _current_column_order(self) -> List[str]:
        header = self.table_view.horizontalHeader()
        frame = self.model.frame()
        if header is None or not isinstance(frame, pd.DataFrame):
            return []
        order: List[str] = []
        for visual_index in range(header.count()):
            logical = header.logicalIndex(visual_index)
            try:
                column = str(frame.columns[logical])
            except Exception:
                continue
            order.append(column)
        return order

    def _apply_column_order(self, order: Sequence[str]) -> None:
        if not order:
            return
        header = self.table_view.horizontalHeader()
        frame = self.model.frame()
        if header is None or not isinstance(frame, pd.DataFrame):
            return
        mapping = {str(column): idx for idx, column in enumerate(frame.columns)}
        for target_visual, column_name in enumerate(order):
            logical = mapping.get(column_name)
            if logical is None:
                continue
            current_visual = header.visualIndex(logical)
            if current_visual == target_visual:
                continue
            header.moveSection(current_visual, target_visual)

    def _ordered_export_frame(self) -> pd.DataFrame:
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame):
            return pd.DataFrame(columns=TRANSITION_TEMP_COLUMNS)
        order = self._current_column_order()
        if order:
            missing = [column for column in frame.columns if column not in order]
            order = list(order) + missing
            export_frame = frame.loc[:, order].copy()
        else:
            export_frame = frame.copy()
        export_frame = export_frame.loc[
            :,
            [column for column in export_frame.columns if not str(column).startswith("_")],
        ]
        return export_frame

    def _export_worksheet(self) -> None:
        frame = self._ordered_export_frame()
        if frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Export worksheet",
                "There is no transition temperature data to export yet.",
            )
            return
        start_dir = _dialog_start_directory()
        suggested = start_dir / "transition_temps.xlsx"
        path_str, selected_filter = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export worksheet",
            str(suggested),
            "Excel files (*.xlsx);;CSV files (*.csv)",
        )
        if not path_str:
            return
        path = Path(path_str)
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".csv"}:
            if "Excel" in selected_filter:
                path = path.with_suffix(".xlsx")
                suffix = ".xlsx"
            else:
                path = path.with_suffix(".csv")
                suffix = ".csv"
        export_frame = frame.copy()
        for column in export_frame.columns:
            series = export_frame[column]
            if getattr(series, "dtype", None) == object:
                export_frame[column] = series.map(
                    lambda value: "" if isinstance(value, (QtGui.QPixmap, QtGui.QImage)) else value
                )
        try:
            if suffix == ".xlsx":
                export_frame.to_excel(path, index=False)
            else:
                export_frame.to_csv(path, index=False)
        except Exception as exc:
            self.logger.exception("Failed to export transition temps worksheet")
            QtWidgets.QMessageBox.critical(
                self,
                "Export worksheet",
                f"Failed to export worksheet:\n{exc}",
            )
            return
        self.log(f"Transition temps worksheet exported to {path}")
        QtWidgets.QMessageBox.information(
            self,
            "Export worksheet",
            f"Worksheet exported to:\n{path}",
        )


class VideoSection(MiniDatabaseSection):
    section_key = "videos"
    section_title = "Fabrication videos"
    supported_suffixes = VIDEO_EXTENSIONS

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._overrides: Dict[str, Dict[str, Any]] = {}
        super().__init__(logger, log_callback, parent)
        self.source_button.hide()
        self.refresh_button.setText("Start video OCR")
        self.open_sources_button.setText("Open video(s)")
        self.open_sources_button.setToolTip("Open the selected video files.")
        self._hide_columns(["_sources", "_group_key", "_cumulative_length_m"])
        self._load_overrides()
        self._normalize_temperature_columns()
        self._apply_overrides_to_model()
        self.model.set_editable_columns(self._editable_columns())
        self.model.set_text_columns({"Notes", "Piece date", "Production datetime"})
        try:
            self.model.dataChanged.connect(self._handle_cell_edited)
        except Exception:
            pass

    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        table = QtWidgets.QTableView(parent)
        table.setModel(self.model)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setSortingEnabled(True)
        table.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.SelectedClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed
            | QtWidgets.QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        header = table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
            header.setSectionsMovable(True)
            header.setSectionsClickable(True)
        self.table_view = table
        container = QtWidgets.QWidget(parent)
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(table, 1)
        return container

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        unique_paths = list(dict.fromkeys(Path(p) for p in paths))

        def _progress(idx: int, total: int) -> None:
            self._check_cancelled()
            if progress is None:
                return
            message: Optional[str] = None
            if 0 < idx <= len(unique_paths):
                message = f"Analysing {unique_paths[idx - 1].name}"
            try:
                progress(idx, total, message)
            except Exception:
                pass

        index = _collect_video_metrics(
            unique_paths,
            self.logger,
            progress_callback=_progress if progress is not None else None,
        )
        self._check_cancelled()
        table = _video_index_to_frame(index, self._fabrication_table())
        table = self._apply_overrides_to_table(table)
        processed: Dict[str, float] = {}
        for path in unique_paths:
            try:
                processed[str(path)] = float(path.stat().st_mtime)
            except OSError:
                continue
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"video_index": index},
        )

    def refresh(self) -> None:
        super().refresh()
        self._hide_columns(["_sources", "_group_key", "_cumulative_length_m"])

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw = row.get("_sources")
        if isinstance(raw, (list, tuple, set)):
            for entry in raw:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        return sources

    def _selected_rows(self) -> List[int]:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return []
        selection = self.table_view.selectionModel()
        if selection is None:
            return []
        indexes = selection.selectedIndexes()
        if not indexes:
            return []
        rows = {index.row() for index in indexes if index.isValid()}
        return sorted(rows)

    def _update_open_sources_enabled(self) -> None:
        if not hasattr(self, "open_sources_button"):
            return
        rows = self._selected_rows()
        self.open_sources_button.setEnabled(bool(rows))

    def _open_selected_sources(self) -> None:
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their video files.",
            )
            return
        opened_any = False
        missing_paths: List[Path] = []
        missing_labels: List[str] = []
        seen: set[Path] = set()
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sources = self._row_sources(series)
            if not sources:
                composition = str(series.get("Composition") or "").strip()
                microwire = str(series.get("Microwire") or "").strip()
                label = f"{composition} {microwire}".strip()
                missing_labels.append(label or f"Row {row_index + 1}")
                continue
            for path in sources:
                if path in seen:
                    continue
                seen.add(path)
                if self._open_file(path):
                    opened_any = True
                else:
                    missing_paths.append(path)
        if opened_any:
            return
        details: List[str] = []
        if missing_labels:
            preview = ", ".join(missing_labels[:5])
            if len(missing_labels) > 5:
                preview += ", …"
            details.append(f"No video found for: {preview}")
        if missing_paths:
            path_preview = "\n".join(str(p) for p in missing_paths[:5])
            if len(missing_paths) > 5:
                path_preview += "\n…"
            details.append(f"Missing files:\n{path_preview}")
        message = "No video files are available for the selected row(s)."
        if details:
            message = f"{message}\n\n" + "\n".join(details)
        QtWidgets.QMessageBox.warning(self, self.section_title, message)

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._load_overrides()
        self._normalize_temperature_columns()
        self._apply_overrides_to_model()
        self._hide_columns(["_sources", "_group_key", "_cumulative_length_m"])

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._normalize_temperature_columns()
        self._apply_overrides_to_model()
        self._hide_columns(["_sources", "_group_key", "_cumulative_length_m"])

    def sync_with_fabrication(self) -> None:
        self._apply_overrides_to_model()
        self._hide_columns(["_sources", "_group_key", "_cumulative_length_m"])

    @staticmethod
    def _is_missing(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        try:
            return bool(pd.isna(value))
        except Exception:
            return False

    def _coerce_float(self, value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            numeric = float(value)
        else:
            text = str(value).strip().replace(",", ".")
            if not text:
                return None
            try:
                numeric = float(text)
            except ValueError:
                return None
        if not math.isfinite(numeric):
            return None
        return numeric

    def _compute_video_length(self, row: pd.Series) -> Optional[float]:
        end_length = self._coerce_float(row.get(VIDEO_END_LENGTH_COLUMN))
        cumulative = self._coerce_float(row.get("_cumulative_length_m"))
        if end_length is None or cumulative is None:
            return None
        return round(end_length - cumulative, 3)

    def _build_cumulative_lengths(
        self,
        base_frame: pd.DataFrame,
        fabrication_frame: Optional[pd.DataFrame],
    ) -> Dict[Tuple[str, int, int], Optional[float]]:
        lengths: Dict[Tuple[str, int, int], Optional[float]] = {}
        if isinstance(fabrication_frame, pd.DataFrame) and not fabrication_frame.empty:
            for _, row in fabrication_frame.iterrows():
                composition = str(row.get("Composition") or "").strip()
                if not composition or composition == "Imported data:":
                    continue
                try:
                    draw = int(row.get("Draw"))
                    piece = int(row.get("Piece"))
                except (TypeError, ValueError):
                    continue
                length_val = self._coerce_float(row.get("Length (m)"))
                lengths[(composition, draw, piece)] = length_val
        for _, row in base_frame.iterrows():
            composition = str(row.get("Composition") or "").strip()
            if not composition or composition == "Imported data:":
                continue
            try:
                draw = int(row.get("Draw"))
                piece = int(row.get("Piece"))
            except (TypeError, ValueError):
                continue
            length_val = self._coerce_float(row.get("Length (m)"))
            lengths[(composition, draw, piece)] = length_val

        cumulative_map: Dict[Tuple[str, int, int], Optional[float]] = {}
        grouped: Dict[Tuple[str, int], List[Tuple[int, Optional[float]]]] = {}
        for (composition, draw, piece), length_val in lengths.items():
            grouped.setdefault((composition, draw), []).append((piece, length_val))
        for (composition, draw), entries in grouped.items():
            running: Optional[float] = 0.0
            for piece, length_val in sorted(entries, key=lambda item: item[0]):
                if running is None or length_val is None:
                    running = None
                    cumulative_map[(composition, draw, piece)] = None
                else:
                    running += length_val
                    cumulative_map[(composition, draw, piece)] = running
        return cumulative_map

    def _load_overrides(self) -> None:
        stored = self.data.extra.get("overrides")
        if isinstance(stored, dict):
            cleaned: Dict[str, Dict[str, Any]] = {}
            for key, payload in stored.items():
                if not isinstance(key, str) or not isinstance(payload, dict):
                    continue
                cleaned[key] = dict(payload)
            self._overrides = cleaned
        else:
            self._overrides = {}

    def _load_video_index(self) -> Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary]:
        try:
            payload = self.store.load_payload("video_index")
        except Exception:
            payload = None
        if isinstance(payload, dict):
            return payload
        return {}

    def _store_overrides(self) -> None:
        self.data.extra["overrides"] = self._overrides
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist video overrides")

    def overrides_snapshot(self) -> Dict[str, Dict[str, Any]]:
        return {
            str(key): dict(payload)
            for key, payload in self._overrides.items()
            if isinstance(payload, dict)
        }

    def _apply_overrides_to_table(self, table: pd.DataFrame) -> pd.DataFrame:
        if not isinstance(table, pd.DataFrame) or table.empty:
            return table
        updated = self._ensure_core_columns(table.copy())
        fabrication_frame = self._fabrication_table()
        cumulative_map = self._build_cumulative_lengths(updated, fabrication_frame)
        for idx, row in updated.iterrows():
            key_raw = row.get("_group_key")
            key = str(key_raw).strip() if key_raw not in (None, "") else ""
            overrides = self._overrides.get(key) if key else None
            if isinstance(overrides, dict):
                for column, value in overrides.items():
                    if column in updated.columns:
                        updated.at[idx, column] = value
            try:
                composition = str(updated.at[idx, "Composition"]).strip()
                draw = int(updated.at[idx, "Draw"])
                piece = int(updated.at[idx, "Piece"])
            except (TypeError, ValueError):
                cumulative = None
            else:
                cumulative = cumulative_map.get((composition, draw, piece))
            updated.at[idx, "_cumulative_length_m"] = cumulative
            updated.at[idx, VIDEO_MW_LENGTH_COLUMN] = self._compute_video_length(updated.loc[idx])
        return updated

    def _apply_overrides_to_model(self) -> None:
        frame = self.model.frame()
        fabrication_frame = self._fabrication_table()
        index = self._load_video_index()
        if (
            (isinstance(fabrication_frame, pd.DataFrame) and not fabrication_frame.empty)
            or index
            or not isinstance(frame, pd.DataFrame)
            or frame.empty
        ):
            frame = _video_index_to_frame(index, fabrication_frame)
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        updated = self._ensure_core_columns(frame.copy())
        if CORE_TEMPERATURE_COLUMN not in updated.columns and "Temperature (°C)" in updated.columns:
            updated = updated.rename(columns={"Temperature (°C)": CORE_TEMPERATURE_COLUMN})
        if GLASS_TEMPERATURE_COLUMN not in updated.columns:
            updated[GLASS_TEMPERATURE_COLUMN] = None
        if VIDEO_END_LENGTH_COLUMN not in updated.columns:
            updated[VIDEO_END_LENGTH_COLUMN] = None
        if VIDEO_MW_LENGTH_COLUMN not in updated.columns:
            updated[VIDEO_MW_LENGTH_COLUMN] = None
        if "_group_key" not in updated.columns:
            keys: List[str] = []
            for _, row in updated.iterrows():
                composition = str(row.get("Composition") or "").strip()
                try:
                    draw = int(row.get("Draw"))
                    piece = int(row.get("Piece"))
                except (TypeError, ValueError):
                    keys.append("")
                    continue
                if composition:
                    keys.append(_microwire_key_to_str((composition, draw, piece, None)))
                else:
                    keys.append("")
            updated["_group_key"] = keys
        frame = updated
        updated = self._apply_overrides_to_table(frame)
        self.data.table = updated
        self.model.set_frame(updated)

    def _ensure_core_columns(self, frame: pd.DataFrame) -> pd.DataFrame:
        updated = frame.copy()
        if "Composition" not in updated.columns:
            updated["Composition"] = ""
        if "Draw" not in updated.columns:
            updated["Draw"] = None
        if "Piece" not in updated.columns:
            updated["Piece"] = None

        draw_values = list(updated["Draw"]) if "Draw" in updated.columns else [None] * len(updated.index)
        piece_values = list(updated["Piece"]) if "Piece" in updated.columns else [None] * len(updated.index)
        microwire_values = (
            list(updated["Microwire"])
            if "Microwire" in updated.columns
            else [None] * len(updated.index)
        )
        for row_idx, value in enumerate(microwire_values):
            try:
                has_draw = not self._is_missing(draw_values[row_idx])
                has_piece = not self._is_missing(piece_values[row_idx])
            except IndexError:
                continue
            if has_draw and has_piece:
                continue
            parts = _microwire_parts_from_label(str(value or ""))
            if not parts:
                continue
            draw, piece, _suffix = parts
            if not has_draw:
                draw_values[row_idx] = draw
            if not has_piece:
                piece_values[row_idx] = piece
        updated["Draw"] = draw_values
        updated["Piece"] = piece_values

        if "Microwire" not in updated.columns:
            labels: List[str] = []
            for draw, piece in zip(draw_values, piece_values):
                try:
                    draw_int = int(draw)
                except (TypeError, ValueError):
                    labels.append("")
                    continue
                try:
                    piece_int = int(piece)
                except (TypeError, ValueError):
                    labels.append(f"{draw_int}/?")
                    continue
                labels.append(_microwire_label(draw_int, piece_int, None))
            updated["Microwire"] = labels
        return updated

    def _normalize_temperature_columns(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else None
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        updated = frame.copy()
        if "Temperature (°C)" in updated.columns:
            if CORE_TEMPERATURE_COLUMN not in updated.columns:
                updated = updated.rename(columns={"Temperature (°C)": CORE_TEMPERATURE_COLUMN})
            else:
                legacy = updated["Temperature (°C)"]
                target = updated[CORE_TEMPERATURE_COLUMN]
                updated[CORE_TEMPERATURE_COLUMN] = target.where(
                    ~(target.isna() | (target == "")),
                    legacy,
                )
                updated = updated.drop(columns=["Temperature (°C)"])
        if GLASS_TEMPERATURE_COLUMN not in updated.columns:
            updated[GLASS_TEMPERATURE_COLUMN] = None
        self.data.table = updated
        self.model.set_frame(updated)

    def _editable_columns(self) -> Set[str]:
        return {
            "Length (m)",
            "Piece date",
            "Resistance (Ω)",
            CORE_TEMPERATURE_COLUMN,
            GLASS_TEMPERATURE_COLUMN,
            "Mass (g)",
            "Winding speed (m/min)",
            "Glass feeding (mm/min)",
            "Underpressure",
            "Notes",
            "Production datetime",
            VIDEO_END_LENGTH_COLUMN,
        }

    def _fabrication_table(self) -> Optional[pd.DataFrame]:
        try:
            store = MiniDatabaseStore("fabrication")
            data = store.load()
        except Exception:
            return None
        table = data.table if isinstance(data.table, pd.DataFrame) else None
        return table

    def _handle_cell_edited(
        self,
        top_left: QtCore.QModelIndex,
        bottom_right: QtCore.QModelIndex,
        roles: Tuple[QtCore.Qt.ItemDataRole, ...] = (),
    ) -> None:
        if roles and QtCore.Qt.ItemDataRole.EditRole not in roles:
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        columns = list(frame.columns[top_left.column() : bottom_right.column() + 1])
        relevant = self._editable_columns()
        if not any(column in relevant for column in columns):
            return
        updated_any = False
        changed_draws: Set[Tuple[str, int]] = set()
        for row_idx in range(top_left.row(), bottom_right.row() + 1):
            if row_idx < 0 or row_idx >= len(frame.index):
                continue
            series = frame.iloc[row_idx]
            key_raw = series.get("_group_key")
            key = str(key_raw).strip() if key_raw not in (None, "") else ""
            if not key:
                continue
            bucket = self._overrides.setdefault(key, {})
            for column in columns:
                if column not in relevant:
                    continue
                value = series.get(column)
                if self._is_missing(value):
                    bucket[column] = None
                else:
                    bucket[column] = value
                updated_any = True
                if column in {"Length (m)", VIDEO_END_LENGTH_COLUMN}:
                    try:
                        composition = str(series.get("Composition") or "").strip()
                        draw = int(series.get("Draw"))
                    except (TypeError, ValueError):
                        continue
                    if composition:
                        changed_draws.add((composition, draw))
            if VIDEO_END_LENGTH_COLUMN in columns or "Length (m)" in columns:
                computed = self._compute_video_length(series)
                frame.at[row_idx, VIDEO_MW_LENGTH_COLUMN] = computed
                bucket[VIDEO_MW_LENGTH_COLUMN] = computed
        if updated_any:
            if changed_draws:
                fabrication_frame = self._fabrication_table()
                cumulative_map = self._build_cumulative_lengths(frame, fabrication_frame)
                for idx, row in frame.iterrows():
                    try:
                        composition = str(row.get("Composition") or "").strip()
                        draw = int(row.get("Draw"))
                        piece = int(row.get("Piece"))
                    except (TypeError, ValueError):
                        continue
                    if (composition, draw) not in changed_draws:
                        continue
                    cumulative = cumulative_map.get((composition, draw, piece))
                    frame.at[idx, "_cumulative_length_m"] = cumulative
                    frame.at[idx, VIDEO_MW_LENGTH_COLUMN] = self._compute_video_length(
                        frame.loc[idx]
                    )
            self.data.table = frame
            self._store_overrides()
            if VIDEO_MW_LENGTH_COLUMN in frame.columns:
                try:
                    col_idx = frame.columns.get_loc(VIDEO_MW_LENGTH_COLUMN)
                except Exception:
                    col_idx = None
                if col_idx is not None:
                    top = self.model.index(top_left.row(), col_idx)
                    bottom = self.model.index(bottom_right.row(), col_idx)
                    try:
                        self.model.dataChanged.emit(
                            top,
                            bottom,
                            [QtCore.Qt.ItemDataRole.DisplayRole],
                        )
                    except Exception:
                        pass


class VsmHysteresisSection(MiniDatabaseSection):
    section_key = "vsm_hysteresis"
    section_title = "VSM hysteresis loops"
    supported_suffixes = (".dat", ".vsm-hys-data", ".txt", ".csv")

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._pixmap_cache: Dict[str, Optional[QtGui.QPixmap]] = {}
        self._record_groups: Dict[str, List[VsmHysteresisRecord]] = {}
        self._hidden_paths: Set[str] = set()
        self._all_records: List[VsmHysteresisRecord] = []
        self._preview_group_count = 1
        self._preview_spacing = 6
        self._table_splitter: QtWidgets.QSplitter | None = None
        super().__init__(logger, log_callback, parent)
        self._load_hidden_paths()
        if isinstance(self.model, DataFrameModel):
            self.model.set_decoration_provider(self._preview_decoration)
        self.open_graphs_button = QtWidgets.QPushButton("Open graphs")
        self.open_graphs_button.clicked.connect(self._open_selected_graphs)
        self.controls_layout.addWidget(self.open_graphs_button)
        self.open_pyplot_button = QtWidgets.QPushButton("Open in PyPlot")
        self.open_pyplot_button.setToolTip("Open the selected VSM files in PyPlot.")
        self.open_pyplot_button.clicked.connect(self._open_selected_in_pyplot)
        self.controls_layout.addWidget(self.open_pyplot_button)
        self.open_origin_button = QtWidgets.QPushButton("Open in Origin")
        self.open_origin_button.setToolTip("Send the selected VSM files to Origin via PyPlot.")
        self.open_origin_button.clicked.connect(self._open_selected_in_origin)
        self.controls_layout.addWidget(self.open_origin_button)
        self.visibility_button = QtWidgets.QPushButton("Visibility...")
        self.visibility_button.setToolTip("Show or hide specific VSM hysteresis graphs.")
        self.visibility_button.clicked.connect(self._open_visibility_dialog)
        self.controls_layout.addWidget(self.visibility_button)
        header = self.table_view.verticalHeader() if self.table_view is not None else None
        if header is not None:
            default_height = ANNEALING_GRAPH_HEIGHT + 24
            header.setDefaultSectionSize(default_height)
            header.setMinimumSectionSize(default_height)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        if _read_vsm_file is None:
            raise RuntimeError("VSM hysteresis parser is not available.")
        records: List[VsmHysteresisRecord] = []
        processed: Dict[str, float] = {}
        total = len(paths)
        for idx, path in enumerate(paths, start=1):
            self._check_cancelled()
            try:
                frame = _read_vsm_file(Path(path))
            except ValueError as exc:
                if "No data rows detected" in str(exc):
                    message = f"Skipped {Path(path).name}: no data rows detected."
                    self.logger.warning(message)
                    self.log(message, level=logging.WARNING)
                    frame = pd.DataFrame()
                else:
                    self.logger.exception("Failed to parse %s", path)
                    frame = pd.DataFrame()
            except Exception:
                self.logger.exception("Failed to parse %s", path)
                frame = pd.DataFrame()
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                if progress is not None:
                    try:
                        progress(idx, total, f"Skipped {Path(path).name}")
                    except Exception:
                        pass
                continue
            columns = [str(column) for column in frame.columns]
            x_column = _choose_axis_column(
                columns,
                [
                    "Applied Field For Plot",
                    "Raw Applied Field For Plot",
                    "Applied Field",
                    "Raw Applied Field",
                    "Applied Field [Oe]",
                    "Field",
                ],
            )
            y_column = _choose_axis_column(
                columns,
                [
                    "Signal X direction",
                    "Signal parallel with sample",
                    "Signal Magnitude",
                    "Moment [emu]",
                    "Signal",
                ],
            )
            if not x_column or not y_column:
                message = f"Skipped {Path(path).name}: missing field/signal columns."
                self.logger.warning(message)
                self.log(message, level=logging.WARNING)
                if progress is not None:
                    try:
                        progress(idx, total, message)
                    except Exception:
                        pass
                continue
            raw_sample = _sample_from_path(Path(path), self.data.sources)
            sample, variant = _split_sample_variant(raw_sample)
            key = _microwire_key_from_path(Path(path), sample or raw_sample)
            temperature = _parse_temperature(Path(path)) if callable(_parse_temperature) else None
            angle = _parse_angle(Path(path)) if callable(_parse_angle) else None
            label = _format_vsm_hysteresis_group_label(
                _coerce_finite_float(temperature),
                variant,
            )
            if not label:
                label = Path(path).stem
            record = VsmHysteresisRecord(
                path=Path(path),
                sample=sample,
                data=frame,
                temperature=temperature,
                angle=angle,
                key=key,
                label=label,
            )
            setattr(record, "variant", variant)
            records.append(record)
            try:
                processed[str(path)] = float(Path(path).stat().st_mtime)
            except OSError:
                processed[str(path)] = 0.0
            if progress is not None:
                try:
                    progress(idx, total, f"Parsed {Path(path).name}")
                except Exception:
                    pass
        table = _graph_records_to_frame(
            records,
            VSM_HYSTERESIS_COLUMN,
            sample_column="_sample",
        )
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"vsm_hysteresis_records": records},
        )

    def refresh(self) -> None:
        super().refresh()
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _load_hidden_paths(self) -> None:
        hidden = self.data.extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            self._hidden_paths = {str(path) for path in hidden if path}
        else:
            self._hidden_paths = set()

    def _store_hidden_paths(self) -> None:
        self.data.extra["hidden_paths"] = sorted(self._hidden_paths)
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist VSM hysteresis visibility settings")

    def _visible_records(
        self, records: Sequence[VsmHysteresisRecord]
    ) -> List[VsmHysteresisRecord]:
        if not self._hidden_paths:
            return list(records)
        return [
            record
            for record in records
            if _record_path_key(record) not in self._hidden_paths
        ]

    def _open_visibility_dialog(self) -> None:
        items = _visibility_items_from_records(self._all_records)
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No VSM hysteresis graphs are available yet.",
            )
            return
        groups = _visibility_groups_from_records(self._all_records)
        dialog = _GraphVisibilityDialog(
            "VSM hysteresis visibility",
            items,
            self._hidden_paths,
            groups=groups,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._hidden_paths = dialog.hidden_paths()
            self._store_hidden_paths()
            self._refresh_record_groups()

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._load_hidden_paths()
        _drop_visible_sample_column(self)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[VsmHysteresisRecord]] = {}
        try:
            payload = self.store.load_payload("vsm_hysteresis_records")
        except Exception:
            payload = None
        all_records = list(payload) if isinstance(payload, list) else []
        self._all_records = list(all_records)
        visible_records = self._visible_records(all_records)
        if visible_records:
            for record in visible_records:
                sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    existing_variant = getattr(record, "variant", None)
                    variant: Optional[str] = None
                    if isinstance(existing_variant, str) and existing_variant.strip():
                        variant = existing_variant.strip()
                    else:
                        base_sample, parsed_variant = _split_sample_variant(sample)
                        if base_sample:
                            try:
                                record.sample = base_sample
                            except Exception:
                                pass
                            sample = base_sample
                        if parsed_variant:
                            variant = parsed_variant
                        setattr(record, "variant", variant)
                    label = _format_vsm_hysteresis_group_label(
                        _coerce_finite_float(getattr(record, "temperature", None)),
                        variant,
                    )
                    if label:
                        try:
                            record.label = label
                        except Exception:
                            pass
                    sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    grouped.setdefault(sample, []).append(record)
        self._record_groups = grouped
        self._record_groups_by_key = _group_graph_records_by_key(visible_records)
        max_groups = 1
        for records in grouped.values():
            groups = _group_vsm_hysteresis_plot_groups(records)
            group_count = len(groups)
            if group_count > max_groups:
                max_groups = group_count
            for record in records:
                record_temp = getattr(record, "_group_temperature", None)
                label = _format_vsm_hysteresis_group_label(
                    _coerce_finite_float(record_temp)
                    if record_temp is not None
                    else _coerce_finite_float(getattr(record, "temperature", None)),
                    getattr(record, "variant", None),
                )
                if label:
                    try:
                        record.label = label
                    except Exception:
                        pass
        self._preview_group_count = max_groups
        self._update_preview_icon_size()
        self._pixmap_cache.clear()
        if isinstance(self.model, DataFrameModel):
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass

    def _preview_icon_width(self) -> int:
        count = max(int(self._preview_group_count), 1)
        return ANNEALING_GRAPH_WIDTH * count + self._preview_spacing * (count - 1)

    def _preview_icon_height(self) -> int:
        return ANNEALING_GRAPH_HEIGHT

    def _update_preview_icon_size(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        width = self._preview_icon_width()
        height = self._preview_icon_height()
        try:
            table.setIconSize(
                QtCore.QSize(max(width, ANNEALING_GRAPH_WIDTH), max(height, ANNEALING_GRAPH_HEIGHT))
            )
        except Exception:
            pass
        header = table.verticalHeader()
        if header is not None:
            try:
                header.setDefaultSectionSize(max(height + 24, ANNEALING_GRAPH_HEIGHT + 24))
            except Exception:
                pass
        self._auto_fit_columns()

    def _preview_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column != VSM_HYSTERESIS_COLUMN:
            return None
        sample = _row_sample_value(row)
        if not sample:
            return None
        cache_key = f"{sample}|{column}"
        if cache_key in self._pixmap_cache:
            return self._pixmap_cache[cache_key]
        records = self._record_groups.get(sample, [])
        if not records:
            row_key = _row_to_microwire_key(row)
            if row_key:
                records = self._record_groups_by_key.get(row_key, [])
        pixmap: Optional[QtGui.QPixmap] = None
        if records:
            groups = _group_vsm_hysteresis_plot_groups(records)
            pixmaps: List[QtGui.QPixmap] = []
            for group in groups:
                figure = _plot_vsm_hysteresis_figure(
                    group.records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                preview = _figure_to_pixmap(
                    figure,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                if preview is not None:
                    pixmaps.append(preview)
            icon_width = self._preview_icon_width()
            pixmap = _combine_pixmaps_side_by_side(
                pixmaps,
                width_px=icon_width,
                height_px=self._preview_icon_height(),
                spacing=self._preview_spacing,
                scale_to_fit=False,
            )
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _selected_records(self) -> List[VsmHysteresisRecord]:
        rows = self._selected_rows()
        records: List[VsmHysteresisRecord] = []
        if not rows:
            return records
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            matched = self._record_groups.get(sample, [])
            if not matched:
                row_key = _row_to_microwire_key(series)
                if row_key:
                    matched = self._record_groups_by_key.get(row_key, [])
            records.extend(matched)
        return records

    def _open_selected_graphs(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        items = _vsm_hysteresis_preview_items(
            records,
            self.logger,
            width_px=GRAPH_PREVIEW_WIDTH,
            height_px=GRAPH_PREVIEW_HEIGHT,
        )
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No graphs are available for the selected rows.",
            )
            return
        dialog = _GraphGalleryDialog(
            "VSM hysteresis graphs",
            items,
            parent=self,
            empty_message="No VSM hysteresis graphs available.",
        )
        dialog.exec()

    def _open_selected_in_pyplot(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [
            record.path for record in records if isinstance(record.path, Path)
        ]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "VSM Hysteresis Loops",
            self.logger,
            auto_plot=True,
            open_origin=False,
        )

    def _open_selected_in_origin(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [
            record.path for record in records if isinstance(record.path, Path)
        ]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "VSM Hysteresis Loops",
            self.logger,
            auto_plot=True,
            open_origin=True,
        )

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw = row.get("_sources")
        if isinstance(raw, (list, tuple, set)):
            for entry in raw:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        return sources


class VsmTemperatureScanSection(MiniDatabaseSection):
    section_key = "vsm_temperature_scan"
    section_title = "VSM temperature scan"
    supported_suffixes = (".vsm-tscn-data", ".txt", ".csv")

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._pixmap_cache: Dict[str, Optional[QtGui.QPixmap]] = {}
        self._record_groups: Dict[str, List[VsmTemperatureScanRecord]] = {}
        self._record_groups_by_key: Dict[str, List[VsmTemperatureScanRecord]] = {}
        self._hidden_paths: Set[str] = set()
        self._all_records: List[VsmTemperatureScanRecord] = []
        self._preview_group_count = 1
        self._preview_spacing = 6
        self._table_splitter: QtWidgets.QSplitter | None = None
        super().__init__(logger, log_callback, parent)
        self._load_hidden_paths()
        if isinstance(self.model, DataFrameModel):
            self.model.set_decoration_provider(self._preview_decoration)
        self.open_graphs_button = QtWidgets.QPushButton("Open graphs")
        self.open_graphs_button.clicked.connect(self._open_selected_graphs)
        self.controls_layout.addWidget(self.open_graphs_button)
        self.open_pyplot_button = QtWidgets.QPushButton("Open in PyPlot")
        self.open_pyplot_button.setToolTip("Open the selected VSM files in PyPlot.")
        self.open_pyplot_button.clicked.connect(self._open_selected_in_pyplot)
        self.controls_layout.addWidget(self.open_pyplot_button)
        self.open_origin_button = QtWidgets.QPushButton("Open in Origin")
        self.open_origin_button.setToolTip("Send the selected VSM files to Origin via PyPlot.")
        self.open_origin_button.clicked.connect(self._open_selected_in_origin)
        self.controls_layout.addWidget(self.open_origin_button)
        self.visibility_button = QtWidgets.QPushButton("Visibility...")
        self.visibility_button.setToolTip("Show or hide specific VSM temperature scan graphs.")
        self.visibility_button.clicked.connect(self._open_visibility_dialog)
        self.controls_layout.addWidget(self.visibility_button)
        header = self.table_view.verticalHeader() if self.table_view is not None else None
        if header is not None:
            default_height = ANNEALING_GRAPH_HEIGHT + 24
            header.setDefaultSectionSize(default_height)
            header.setMinimumSectionSize(default_height)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        processor = _get_vsm_temp_processor(self.logger)
        if processor is None:
            raise RuntimeError("VSM temperature scan parser is not available.")
        records: List[VsmTemperatureScanRecord] = []
        processed: Dict[str, float] = {}
        total = len(paths)
        for idx, path in enumerate(paths, start=1):
            self._check_cancelled()
            try:
                frame, parsed_sample = processor._parse_file(Path(path))
            except Exception:
                self.logger.exception("Failed to parse %s", path)
                frame = pd.DataFrame()
                parsed_sample = ""
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                if progress is not None:
                    try:
                        progress(idx, total, f"Skipped {Path(path).name}")
                    except Exception:
                        pass
                continue
            raw_sample = str(parsed_sample or "").strip() or _sample_from_path(Path(path), self.data.sources)
            sample = raw_sample
            key = _microwire_key_from_path(Path(path), sample or raw_sample)
            label = Path(path).stem
            record = VsmTemperatureScanRecord(
                path=Path(path),
                sample=sample,
                data=frame,
                key=key,
                label=label,
            )
            records.append(record)
            try:
                processed[str(path)] = float(Path(path).stat().st_mtime)
            except OSError:
                processed[str(path)] = 0.0
            if progress is not None:
                try:
                    progress(idx, total, f"Parsed {Path(path).name}")
                except Exception:
                    pass
        table = _graph_records_to_frame(
            records,
            VSM_TEMPERATURE_SCAN_COLUMN,
            sample_column="_sample",
        )
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"vsm_temperature_scan_records": records},
        )

    def refresh(self) -> None:
        super().refresh()
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _load_hidden_paths(self) -> None:
        hidden = self.data.extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            self._hidden_paths = {str(path) for path in hidden if path}
        else:
            self._hidden_paths = set()

    def _store_hidden_paths(self) -> None:
        self.data.extra["hidden_paths"] = sorted(self._hidden_paths)
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception(
                "Failed to persist VSM temperature scan visibility settings"
            )

    def _visible_records(
        self, records: Sequence[VsmTemperatureScanRecord]
    ) -> List[VsmTemperatureScanRecord]:
        if not self._hidden_paths:
            return list(records)
        return [
            record
            for record in records
            if _record_path_key(record) not in self._hidden_paths
        ]

    def _open_visibility_dialog(self) -> None:
        items = _visibility_items_from_records(self._all_records)
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No VSM temperature scan graphs are available yet.",
            )
            return
        groups = _visibility_groups_from_records(self._all_records)
        dialog = _GraphVisibilityDialog(
            "VSM temperature scan visibility",
            items,
            self._hidden_paths,
            groups=groups,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._hidden_paths = dialog.hidden_paths()
            self._store_hidden_paths()
            self._refresh_record_groups()

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._load_hidden_paths()
        _drop_visible_sample_column(self)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[VsmTemperatureScanRecord]] = {}
        try:
            payload = self.store.load_payload("vsm_temperature_scan_records")
        except Exception:
            payload = None
        all_records = list(payload) if isinstance(payload, list) else []
        self._all_records = list(all_records)
        visible_records = self._visible_records(all_records)
        if visible_records:
            for record in visible_records:
                sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    grouped.setdefault(sample, []).append(record)
        self._record_groups = grouped
        self._record_groups_by_key = _group_graph_records_by_key(visible_records)
        max_groups = 1
        for records in grouped.values():
            if len(records) > max_groups:
                max_groups = len(records)
        self._preview_group_count = max_groups
        self._update_preview_icon_size()
        self._pixmap_cache.clear()
        if isinstance(self.model, DataFrameModel):
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass

    def _preview_icon_width(self) -> int:
        count = max(int(self._preview_group_count), 1)
        return ANNEALING_GRAPH_WIDTH * count + self._preview_spacing * (count - 1)

    def _preview_icon_height(self) -> int:
        return ANNEALING_GRAPH_HEIGHT

    def _update_preview_icon_size(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        width = self._preview_icon_width()
        height = self._preview_icon_height()
        try:
            table.setIconSize(
                QtCore.QSize(max(width, ANNEALING_GRAPH_WIDTH), max(height, ANNEALING_GRAPH_HEIGHT))
            )
        except Exception:
            pass
        header = table.verticalHeader()
        if header is not None:
            try:
                header.setDefaultSectionSize(max(height + 24, ANNEALING_GRAPH_HEIGHT + 24))
            except Exception:
                pass
        self._auto_fit_columns()

    def _preview_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column != VSM_TEMPERATURE_SCAN_COLUMN:
            return None
        sample = _row_sample_value(row)
        if not sample:
            return None
        cache_key = f"{sample}|{column}"
        if cache_key in self._pixmap_cache:
            return self._pixmap_cache[cache_key]
        records = self._record_groups.get(sample, [])
        if not records:
            row_key = _row_to_microwire_key(row)
            if row_key:
                records = self._record_groups_by_key.get(row_key, [])
        pixmap: Optional[QtGui.QPixmap] = None
        if records:
            items = _vsm_temperature_preview_items(
                records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            pixmaps = [item.pixmap for item in items if item.pixmap is not None]
            if pixmaps:
                pixmap = _combine_pixmaps_side_by_side(
                    pixmaps,
                    width_px=self._preview_icon_width(),
                    height_px=self._preview_icon_height(),
                    spacing=self._preview_spacing,
                    scale_to_fit=False,
                )
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _open_selected_graphs(self) -> None:
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        records: List[VsmTemperatureScanRecord] = []
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            records.extend(self._record_groups.get(sample, []))
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No graphs are available for the selected rows.",
            )
            return
        items = _vsm_temperature_preview_items(
            records,
            self.logger,
            width_px=GRAPH_PREVIEW_WIDTH,
            height_px=GRAPH_PREVIEW_HEIGHT,
        )
        dialog = _GraphGalleryDialog(
            "VSM temperature scan graphs",
            items,
            parent=self,
            empty_message="No VSM temperature scan graphs available.",
        )
        dialog.exec()

    def _selected_records(self) -> List[VsmTemperatureScanRecord]:
        rows = self._selected_rows()
        records: List[VsmTemperatureScanRecord] = []
        if not rows:
            return records
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            matched = self._record_groups.get(sample, [])
            if not matched:
                row_key = _row_to_microwire_key(series)
                if row_key:
                    matched = self._record_groups_by_key.get(row_key, [])
            records.extend(matched)
        return records

    def _open_selected_in_pyplot(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "VSM Temperature Scan",
            self.logger,
            auto_plot=True,
            open_origin=False,
        )

    def _open_selected_in_origin(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "VSM Temperature Scan",
            self.logger,
            auto_plot=True,
            open_origin=True,
        )

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw = row.get("_sources")
        if isinstance(raw, (list, tuple, set)):
            for entry in raw:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        return sources


class DmaIsoStressSection(MiniDatabaseSection):
    section_key = "dma_iso_stress"
    section_title = "DMA iso-stress"
    supported_suffixes = (".txt",)

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._pixmap_cache: Dict[str, Optional[QtGui.QPixmap]] = {}
        self._record_groups: Dict[str, List[DmaIsoStressRecord]] = {}
        self._record_groups_by_key: Dict[str, List[DmaIsoStressRecord]] = {}
        self._hidden_paths: Set[str] = set()
        self._all_records: List[DmaIsoStressRecord] = []
        self._preview_group_count = 1
        self._preview_spacing = 6
        self._table_splitter: QtWidgets.QSplitter | None = None
        super().__init__(logger, log_callback, parent)
        self._load_hidden_paths()
        if isinstance(self.model, DataFrameModel):
            self.model.set_decoration_provider(self._preview_decoration)
        self.open_graphs_button = QtWidgets.QPushButton("Open graphs")
        self.open_graphs_button.clicked.connect(self._open_selected_graphs)
        self.controls_layout.addWidget(self.open_graphs_button)
        self.open_pyplot_button = QtWidgets.QPushButton("Open in PyPlot")
        self.open_pyplot_button.setToolTip("Open the selected DMA files in PyPlot.")
        self.open_pyplot_button.clicked.connect(self._open_selected_in_pyplot)
        self.controls_layout.addWidget(self.open_pyplot_button)
        self.open_origin_button = QtWidgets.QPushButton("Open in Origin")
        self.open_origin_button.setToolTip("Send the selected DMA files to Origin via PyPlot.")
        self.open_origin_button.clicked.connect(self._open_selected_in_origin)
        self.controls_layout.addWidget(self.open_origin_button)
        self.visibility_button = QtWidgets.QPushButton("Visibility...")
        self.visibility_button.setToolTip("Show or hide specific DMA iso-stress graphs.")
        self.visibility_button.clicked.connect(self._open_visibility_dialog)
        self.controls_layout.addWidget(self.visibility_button)
        header = self.table_view.verticalHeader() if self.table_view is not None else None
        if header is not None:
            default_height = ANNEALING_GRAPH_HEIGHT + 24
            header.setDefaultSectionSize(default_height)
            header.setMinimumSectionSize(default_height)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        if parse_dma_txt is None:
            raise RuntimeError("DMA iso-stress parser is not available.")
        records: List[DmaIsoStressRecord] = []
        processed: Dict[str, float] = {}
        total = len(paths)
        for idx, path in enumerate(paths, start=1):
            self._check_cancelled()
            try:
                datasets = parse_dma_txt(Path(path))
            except Exception:
                self.logger.exception("Failed to parse %s", path)
                datasets = {}
            if not datasets:
                if progress is not None:
                    try:
                        progress(idx, total, f"Skipped {Path(path).name}")
                    except Exception:
                        pass
                continue
            raw_sample = _sample_from_path(Path(path), self.data.sources)
            sample, variant = _split_sample_variant(raw_sample)
            key = _microwire_key_from_path(Path(path), sample or raw_sample)
            label = Path(path).stem
            if variant:
                label = f"{variant} — {label}"
            records.append(
                DmaIsoStressRecord(
                    path=Path(path),
                    sample=sample or raw_sample,
                    datasets=datasets,
                    key=key,
                    label=label,
                )
            )
            if variant:
                setattr(records[-1], "variant", variant)
            try:
                processed[str(path)] = float(Path(path).stat().st_mtime)
            except OSError:
                processed[str(path)] = 0.0
            if progress is not None:
                try:
                    progress(idx, total, f"Parsed {Path(path).name}")
                except Exception:
                    pass
        table = _graph_records_to_frame(
            records,
            DMA_ISOSTRESS_COLUMN,
            sample_column="_sample",
        )
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"dma_iso_stress_records": records},
        )

    def refresh(self) -> None:
        super().refresh()
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _load_hidden_paths(self) -> None:
        hidden = self.data.extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            self._hidden_paths = {str(path) for path in hidden if path}
        else:
            self._hidden_paths = set()

    def _store_hidden_paths(self) -> None:
        self.data.extra["hidden_paths"] = sorted(self._hidden_paths)
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist DMA iso-stress visibility settings")

    def _visible_records(
        self, records: Sequence[DmaIsoStressRecord]
    ) -> List[DmaIsoStressRecord]:
        if not self._hidden_paths:
            return list(records)
        return [
            record
            for record in records
            if _record_path_key(record) not in self._hidden_paths
        ]

    def _open_visibility_dialog(self) -> None:
        items = _visibility_items_from_records(self._all_records)
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No DMA iso-stress graphs are available yet.",
            )
            return
        groups = _visibility_groups_from_records(self._all_records)
        dialog = _GraphVisibilityDialog(
            "DMA iso-stress visibility",
            items,
            self._hidden_paths,
            groups=groups,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._hidden_paths = dialog.hidden_paths()
            self._store_hidden_paths()
            self._refresh_record_groups()

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._load_hidden_paths()
        _drop_visible_sample_column(self)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[DmaIsoStressRecord]] = {}
        try:
            payload = self.store.load_payload("dma_iso_stress_records")
        except Exception:
            payload = None
        all_records = list(payload) if isinstance(payload, list) else []
        self._all_records = list(all_records)
        visible_records = self._visible_records(all_records)
        if visible_records:
            for record in visible_records:
                sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    existing_variant = getattr(record, "variant", None)
                    variant: Optional[str] = None
                    if isinstance(existing_variant, str) and existing_variant.strip():
                        variant = existing_variant.strip()
                    else:
                        base_sample, parsed_variant = _split_sample_variant(sample)
                        if base_sample:
                            try:
                                record.sample = base_sample
                            except Exception:
                                pass
                            sample = base_sample
                        if parsed_variant:
                            variant = parsed_variant
                        setattr(record, "variant", variant)
                    if variant:
                        label = getattr(record, "label", None)
                        if isinstance(label, str) and label.strip():
                            if variant not in label:
                                try:
                                    record.label = f"{variant} — {label}"
                                except Exception:
                                    pass
                    sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    grouped.setdefault(sample, []).append(record)
        self._record_groups = grouped
        self._record_groups_by_key = _group_graph_records_by_key(visible_records)
        max_groups = 1
        for records in grouped.values():
            if len(records) > max_groups:
                max_groups = len(records)
        self._preview_group_count = max_groups
        self._update_preview_icon_size()
        self._pixmap_cache.clear()
        if isinstance(self.model, DataFrameModel):
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass

    def _preview_icon_width(self) -> int:
        count = max(int(self._preview_group_count), 1)
        return ANNEALING_GRAPH_WIDTH * count + self._preview_spacing * (count - 1)

    def _preview_icon_height(self) -> int:
        return ANNEALING_GRAPH_HEIGHT

    def _update_preview_icon_size(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        width = self._preview_icon_width()
        height = self._preview_icon_height()
        try:
            table.setIconSize(
                QtCore.QSize(max(width, ANNEALING_GRAPH_WIDTH), max(height, ANNEALING_GRAPH_HEIGHT))
            )
        except Exception:
            pass
        header = table.verticalHeader()
        if header is not None:
            try:
                header.setDefaultSectionSize(max(height + 24, ANNEALING_GRAPH_HEIGHT + 24))
            except Exception:
                pass
        self._auto_fit_columns()

    def _preview_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column != DMA_ISOSTRESS_COLUMN:
            return None
        sample = _row_sample_value(row)
        if not sample:
            return None
        cache_key = f"{sample}|{column}"
        if cache_key in self._pixmap_cache:
            return self._pixmap_cache[cache_key]
        records = self._record_groups.get(sample, [])
        if not records:
            row_key = _row_to_microwire_key(row)
            if row_key:
                records = self._record_groups_by_key.get(row_key, [])
        pixmap: Optional[QtGui.QPixmap] = None
        if records:
            items = _dma_iso_stress_preview_items(
                records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            pixmaps = [item.pixmap for item in items if item.pixmap is not None]
            if pixmaps:
                pixmap = _combine_pixmaps_side_by_side(
                    pixmaps,
                    width_px=self._preview_icon_width(),
                    height_px=self._preview_icon_height(),
                    spacing=self._preview_spacing,
                    scale_to_fit=False,
                )
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _open_selected_graphs(self) -> None:
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        records: List[DmaIsoStressRecord] = []
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            records.extend(self._record_groups.get(sample, []))
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No graphs are available for the selected rows.",
            )
            return
        items = _dma_iso_stress_preview_items(
            records,
            self.logger,
            width_px=GRAPH_PREVIEW_WIDTH,
            height_px=GRAPH_PREVIEW_HEIGHT,
        )
        dialog = _GraphGalleryDialog(
            "DMA iso-stress graphs",
            items,
            parent=self,
            empty_message="No DMA iso-stress graphs available.",
        )
        dialog.exec()

    def _selected_records(self) -> List[DmaIsoStressRecord]:
        rows = self._selected_rows()
        records: List[DmaIsoStressRecord] = []
        if not rows:
            return records
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            matched = self._record_groups.get(sample, [])
            if not matched:
                row_key = _row_to_microwire_key(series)
                if row_key:
                    matched = self._record_groups_by_key.get(row_key, [])
            records.extend(matched)
        return records

    def _open_selected_in_pyplot(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "DMA Iso-Stress",
            self.logger,
            auto_plot=True,
            open_origin=False,
        )

    def _open_selected_in_origin(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "DMA Iso-Stress",
            self.logger,
            auto_plot=True,
            open_origin=True,
        )

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw = row.get("_sources")
        if isinstance(raw, (list, tuple, set)):
            for entry in raw:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        return sources

class ShapeMemoryStressStrainSection(MiniDatabaseSection):
    section_key = "shape_memory_stress_strain"
    section_title = "Shape memory stress/strain"
    supported_suffixes = (".txt",)
    VALUE_COLUMNS = [
        SHAPE_MEMORY_DISPLACEMENT_COLUMN,
        SHAPE_MEMORY_LOAD_COLUMN,
        SHAPE_MEMORY_STRAIN_COLUMN,
        SHAPE_MEMORY_STRESS_COLUMN,
    ]
    FRACTURE_COLUMNS = [
        SHAPE_MEMORY_FRACTURE_LOAD_COLUMN,
        SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN,
        SHAPE_MEMORY_FRACTURE_STRESS_COLUMN,
    ]

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._pixmap_cache: Dict[str, Optional[QtGui.QPixmap]] = {}
        self._record_groups: Dict[str, List[ShapeMemoryStressStrainRecord]] = {}
        self._record_groups_by_key: Dict[str, List[ShapeMemoryStressStrainRecord]] = {}
        self._hidden_paths: Set[str] = set()
        self._all_records: List[ShapeMemoryStressStrainRecord] = []
        self._preview_group_count = 1
        self._preview_spacing = 6
        self._table_splitter: QtWidgets.QSplitter | None = None
        self._preview_panel: _ShapeMemoryPreviewPanel | None = None
        self._preview_toggle: QtWidgets.QCheckBox | None = None
        super().__init__(logger, log_callback, parent)
        self._normalise_value_columns()
        self._load_hidden_paths()
        if isinstance(self.model, DataFrameModel):
            self.model.set_decoration_provider(self._preview_decoration)
        self.open_graphs_button = QtWidgets.QPushButton("Open graphs")
        self.open_graphs_button.clicked.connect(self._open_selected_graphs)
        self.controls_layout.addWidget(self.open_graphs_button)
        self.open_pyplot_button = QtWidgets.QPushButton("Open in PyPlot")
        self.open_pyplot_button.setToolTip(
            "Open the selected shape-memory files in PyPlot."
        )
        self.open_pyplot_button.clicked.connect(self._open_selected_in_pyplot)
        self.controls_layout.addWidget(self.open_pyplot_button)
        self.open_origin_button = QtWidgets.QPushButton("Open in Origin")
        self.open_origin_button.setToolTip(
            "Send the selected shape-memory files to Origin via PyPlot."
        )
        self.open_origin_button.clicked.connect(self._open_selected_in_origin)
        self.controls_layout.addWidget(self.open_origin_button)
        self.visibility_button = QtWidgets.QPushButton("Visibility...")
        self.visibility_button.setToolTip(
            "Show or hide specific shape-memory graphs."
        )
        self.visibility_button.clicked.connect(self._open_visibility_dialog)
        self.controls_layout.addWidget(self.visibility_button)
        self._preview_toggle = QtWidgets.QCheckBox("Show graph preview panel")
        self._preview_toggle.setChecked(self._preview_panel_visible())
        self._preview_toggle.toggled.connect(self._toggle_preview_panel)
        self.controls_layout.addWidget(self._preview_toggle)
        header = self.table_view.verticalHeader() if self.table_view is not None else None
        if header is not None:
            default_height = ANNEALING_GRAPH_HEIGHT + 24
            header.setDefaultSectionSize(default_height)
            header.setMinimumSectionSize(default_height)
        selection_model = self.table_view.selectionModel() if self.table_view is not None else None
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._handle_selection_changed)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])
        self._toggle_preview_panel(self._preview_panel_visible())
        self._update_preview()

    def _source_row(self, proxy_row: int) -> Optional[int]:
        return self._search_proxy.map_row_to_source(proxy_row)

    def _handle_search_changed(self, text: str) -> None:
        self._search_proxy.set_search_text(text)
        if isinstance(self.search_clear_button, QtWidgets.QPushButton):
            self.search_clear_button.setEnabled(bool(str(text).strip()))

    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        table = QtWidgets.QTableView(parent)
        table.setModel(self.model)
        table.horizontalHeader().setStretchLastSection(True)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        table.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        table.setSortingEnabled(True)
        self.table_view = table

        splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal, parent)
        splitter.setChildrenCollapsible(False)
        splitter.setOpaqueResize(False)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
        splitter.addWidget(table)
        preview_panel = _ShapeMemoryPreviewPanel(self.logger, splitter)
        preview_panel.pointPicked.connect(self._apply_picked_selection)
        splitter.addWidget(preview_panel)
        self._preview_panel = preview_panel
        self._table_splitter = splitter

        container = QtWidgets.QWidget(parent)
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(splitter, 1)
        return container

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        if load_manual_stress_strain_file is None:
            raise RuntimeError("Shape-memory stress/strain parser is not available.")
        records: List[ShapeMemoryStressStrainRecord] = []
        processed: Dict[str, float] = {}
        total = len(paths)
        for idx, path in enumerate(paths, start=1):
            self._check_cancelled()
            try:
                frame = load_manual_stress_strain_file(Path(path))
            except Exception:
                self.logger.exception("Failed to parse %s", path)
                frame = pd.DataFrame()
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                if progress is not None:
                    try:
                        progress(idx, total, f"Skipped {Path(path).name}")
                    except Exception:
                        pass
                continue
            raw_sample = _sample_from_path(Path(path), self.data.sources)
            sample, variant = _split_sample_variant(raw_sample)
            stem_sample, stem_variant = _split_sample_variant(Path(path).stem)
            key = _microwire_key_from_path(Path(path), sample or raw_sample)
            if key is None:
                key = _microwire_key_from_path(Path(path), stem_sample or Path(path).stem)
            if key is not None:
                composition, microwire = _microwire_info_from_key(key)
                if (
                    composition
                    and microwire
                    and (
                        not sample
                        or sample == Path(path).parent.name
                        or _microwire_key_from_path(Path(path), sample) is None
                    )
                ):
                    sample = f"{composition} {microwire}"
                    if variant is None:
                        variant = stem_variant
            elif stem_sample:
                sample = stem_sample
                if variant is None:
                    variant = stem_variant
            label = Path(path).stem
            if variant:
                label = f"{variant} - {label}"
            record = ShapeMemoryStressStrainRecord(
                path=Path(path),
                sample=sample or raw_sample,
                data=frame,
                key=key,
                label=label,
            )
            if variant:
                setattr(record, "variant", variant)
            records.append(record)
            try:
                processed[str(path)] = float(Path(path).stat().st_mtime)
            except OSError:
                processed[str(path)] = 0.0
            if progress is not None:
                try:
                    progress(idx, total, f"Parsed {Path(path).name}")
                except Exception:
                    pass
        table = _graph_records_to_frame(
            records,
            SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
            sample_column="_sample",
        )
        existing_entries = self.entries_snapshot()
        if existing_entries:
            for row_index, row in table.iterrows():
                row_key = _row_to_microwire_key(row)
                if not row_key:
                    continue
                payload = existing_entries.get(row_key, {})
                if not isinstance(payload, dict):
                    continue
                for column in self.VALUE_COLUMNS + self.FRACTURE_COLUMNS:
                    if column not in table.columns:
                        table[column] = None
                    if column in payload:
                        table.at[row_index, column] = payload.get(column)
        for column in self.VALUE_COLUMNS + self.FRACTURE_COLUMNS:
            if column not in table.columns:
                table[column] = None
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"shape_memory_stress_strain_records": records},
        )

    def refresh(self) -> None:
        super().refresh()
        self._normalise_value_columns()
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])
        self._toggle_preview_panel(self._preview_panel_visible())
        self._update_preview()

    def _load_hidden_paths(self) -> None:
        hidden = self.data.extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            self._hidden_paths = {str(path) for path in hidden if path}
        else:
            self._hidden_paths = set()

    def _store_hidden_paths(self) -> None:
        self.data.extra["hidden_paths"] = sorted(self._hidden_paths)
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception(
                "Failed to persist shape-memory visibility settings"
            )

    def _visible_records(
        self, records: Sequence[ShapeMemoryStressStrainRecord]
    ) -> List[ShapeMemoryStressStrainRecord]:
        if not self._hidden_paths:
            return list(records)
        return [
            record
            for record in records
            if _record_path_key(record) not in self._hidden_paths
        ]

    def _open_visibility_dialog(self) -> None:
        items = _visibility_items_from_records(self._all_records)
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No shape-memory graphs are available yet.",
            )
            return
        groups = _visibility_groups_from_records(self._all_records)
        dialog = _GraphVisibilityDialog(
            "Shape-memory visibility",
            items,
            self._hidden_paths,
            groups=groups,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._hidden_paths = dialog.hidden_paths()
            self._store_hidden_paths()
            self._refresh_record_groups()

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._normalise_value_columns()
        self._load_hidden_paths()
        _drop_visible_sample_column(self)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])
        self._toggle_preview_panel(self._preview_panel_visible())
        self._update_preview()

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._normalise_value_columns()
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])
        self._toggle_preview_panel(self._preview_panel_visible())
        self._update_preview()

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[ShapeMemoryStressStrainRecord]] = {}
        try:
            payload = self.store.load_payload("shape_memory_stress_strain_records")
        except Exception:
            payload = None
        all_records = list(payload) if isinstance(payload, list) else []
        self._all_records = list(all_records)
        visible_records = self._visible_records(all_records)
        if visible_records:
            for record in visible_records:
                sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    existing_variant = getattr(record, "variant", None)
                    variant: Optional[str] = None
                    if isinstance(existing_variant, str) and existing_variant.strip():
                        variant = existing_variant.strip()
                    else:
                        base_sample, parsed_variant = _split_sample_variant(sample)
                        if base_sample:
                            try:
                                record.sample = base_sample
                            except Exception:
                                pass
                            sample = base_sample
                        if parsed_variant:
                            variant = parsed_variant
                        setattr(record, "variant", variant)
                    if variant:
                        label = getattr(record, "label", None)
                        if isinstance(label, str) and label.strip():
                            if variant not in label:
                                try:
                                    record.label = f"{variant} - {label}"
                                except Exception:
                                    pass
                    sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    grouped.setdefault(sample, []).append(record)
        self._record_groups = grouped
        self._record_groups_by_key = _group_graph_records_by_key(visible_records)
        max_groups = 1
        for records in grouped.values():
            group_count = len(records)
            if group_count > max_groups:
                max_groups = group_count
        self._preview_group_count = max_groups
        self._update_preview_icon_size()
        self._pixmap_cache.clear()
        if isinstance(self.model, DataFrameModel):
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass

    def _preview_icon_width(self) -> int:
        count = max(int(self._preview_group_count), 1)
        return ANNEALING_GRAPH_WIDTH * count + self._preview_spacing * (count - 1)

    def _preview_icon_height(self) -> int:
        return ANNEALING_GRAPH_HEIGHT

    def _update_preview_icon_size(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        width = self._preview_icon_width()
        height = self._preview_icon_height()
        try:
            table.setIconSize(
                QtCore.QSize(max(width, ANNEALING_GRAPH_WIDTH), max(height, ANNEALING_GRAPH_HEIGHT))
            )
        except Exception:
            pass
        header = table.verticalHeader()
        if header is not None:
            try:
                header.setDefaultSectionSize(max(height + 24, ANNEALING_GRAPH_HEIGHT + 24))
            except Exception:
                pass
        self._auto_fit_columns()

    def _preview_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column != SHAPE_MEMORY_STRESS_STRAIN_COLUMN:
            return None
        sample = _row_sample_value(row)
        if not sample:
            return None
        cache_key = f"{sample}|{column}"
        if cache_key in self._pixmap_cache:
            return self._pixmap_cache[cache_key]
        records = self._record_groups.get(sample, [])
        if not records:
            row_key = _row_to_microwire_key(row)
            if row_key:
                records = self._record_groups_by_key.get(row_key, [])
        pixmap: Optional[QtGui.QPixmap] = None
        if records:
            items = _shape_memory_stress_strain_preview_items(
                records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            pixmaps = [item.pixmap for item in items if item.pixmap is not None]
            if pixmaps:
                pixmap = _combine_pixmaps_side_by_side(
                    pixmaps,
                    width_px=self._preview_icon_width(),
                    height_px=self._preview_icon_height(),
                    spacing=self._preview_spacing,
                    scale_to_fit=False,
                )
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _selected_records(self) -> List[ShapeMemoryStressStrainRecord]:
        rows = self._selected_rows()
        records: List[ShapeMemoryStressStrainRecord] = []
        if not rows:
            return records
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            matched = self._record_groups.get(sample, [])
            if not matched:
                row_key = _row_to_microwire_key(series)
                if row_key:
                    matched = self._record_groups_by_key.get(row_key, [])
            records.extend(matched)
        return records

    def _selected_preview_records(self) -> List[ShapeMemoryStressStrainRecord]:
        rows = self._selected_rows()
        if not rows:
            return []
        series = self._row_series(rows[0])
        if series is None:
            return []
        sample = _row_sample_value(series)
        matched = self._record_groups.get(sample, []) if sample else []
        if not matched:
            row_key = _row_to_microwire_key(series)
            if row_key:
                matched = self._record_groups_by_key.get(row_key, [])
        return list(matched)

    def _update_preview(self) -> None:
        panel = self._preview_panel
        if panel is None:
            return
        records = self._selected_preview_records()
        if not records:
            panel.clear("Select a row to preview shape-memory graphs.")
            return
        first = records[0]
        title = getattr(first, "sample", None) or self.section_title
        panel.update_selection(str(title), records)

    def _handle_selection_changed(self, *_args: Any) -> None:
        self._update_preview()

    def entries_snapshot(self) -> Dict[str, Dict[str, Any]]:
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return {}
        snapshot: Dict[str, Dict[str, Any]] = {}
        for _, row in frame.iterrows():
            key = _row_to_microwire_key(row)
            if not key:
                continue
            entry: Dict[str, Any] = {}
            for column in self.VALUE_COLUMNS + self.FRACTURE_COLUMNS:
                value = row.get(column)
                if value in (None, ""):
                    continue
                entry[column] = value
            if entry:
                snapshot[key] = entry
        return snapshot

    def _normalise_value_columns(self) -> None:
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame):
            return
        updated = frame.copy()
        renamed = False
        for old_name, new_name in _SHAPE_MEMORY_COLUMN_ALIASES.items():
            if old_name not in updated.columns:
                continue
            if new_name not in updated.columns:
                updated = updated.rename(columns={old_name: new_name})
            else:
                mask = updated[new_name].isna() | (updated[new_name] == "")
                updated.loc[mask, new_name] = updated.loc[mask, old_name]
                updated = updated.drop(columns=[old_name])
            renamed = True
        for column in self.VALUE_COLUMNS + self.FRACTURE_COLUMNS:
            if column not in updated.columns:
                updated[column] = None
        if renamed or not frame.columns.equals(updated.columns):
            self.data.table = updated
            self.model.set_frame(updated)
            try:
                self.store.save(self.data)
            except Exception:
                pass

    def _preview_panel_visible(self) -> bool:
        extra = self.data.extra if isinstance(self.data.extra, dict) else {}
        visible = extra.get("preview_panel_visible", True)
        return bool(visible)

    def _toggle_preview_panel(self, checked: bool) -> None:
        extra = self.data.extra if isinstance(self.data.extra, dict) else {}
        extra = dict(extra)
        extra["preview_panel_visible"] = bool(checked)
        self.data.extra = extra
        panel = self._preview_panel
        splitter = self._table_splitter
        if isinstance(panel, QtWidgets.QWidget):
            panel.setVisible(bool(checked))
        if isinstance(splitter, QtWidgets.QSplitter):
            if checked:
                sizes = splitter.sizes()
                if sizes and sizes[1] == 0:
                    total = max(sum(sizes), 1)
                    splitter.setSizes([int(total * 0.55), int(total * 0.45)])
            else:
                splitter.setSizes([1, 0])
        try:
            self.store.save(self.data)
        except Exception:
            pass

    def _apply_picked_selection(self, target: str, selection: object) -> None:
        if not isinstance(selection, _ShapeMemoryPointSelection):
            return
        rows = self._selected_rows()
        if not rows:
            return
        row_index = rows[0]
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or row_index < 0 or row_index >= len(frame.index):
            return
        if target == "fracture":
            updates = {
                SHAPE_MEMORY_FRACTURE_LOAD_COLUMN: round(selection.load_g, 6),
                SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN: round(selection.strain_pct, 6),
                SHAPE_MEMORY_FRACTURE_STRESS_COLUMN: round(selection.stress_mpa, 6),
            }
        else:
            updates = {
                SHAPE_MEMORY_DISPLACEMENT_COLUMN: round(selection.displacement_mm, 6),
                SHAPE_MEMORY_LOAD_COLUMN: round(selection.load_g, 6),
                SHAPE_MEMORY_STRAIN_COLUMN: round(selection.strain_pct, 6),
                SHAPE_MEMORY_STRESS_COLUMN: round(selection.stress_mpa, 6),
            }
        updated = frame.copy()
        for column, value in updates.items():
            if column not in updated.columns:
                updated[column] = None
            updated.at[row_index, column] = value
        self.data.table = updated
        self.model.set_frame(updated)
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])
        try:
            if self.table_view is not None:
                self.table_view.selectRow(row_index)
        except Exception:
            pass
        try:
            self.store.save(self.data)
        except Exception:
            pass
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _open_selected_graphs(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        items = _shape_memory_stress_strain_preview_items(
            records,
            self.logger,
            width_px=GRAPH_PREVIEW_WIDTH,
            height_px=GRAPH_PREVIEW_HEIGHT,
        )
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No graphs are available for the selected rows.",
            )
            return
        dialog = _GraphGalleryDialog(
            "Shape memory stress/strain graphs",
            items,
            parent=self,
            empty_message="No shape-memory graphs available.",
        )
        dialog.exec()

    def _open_selected_in_pyplot(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "Shape Memory Stress/Strain",
            self.logger,
            auto_plot=True,
            open_origin=False,
        )

    def _open_selected_in_origin(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "Shape Memory Stress/Strain",
            self.logger,
            auto_plot=True,
            open_origin=True,
        )

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw = row.get("_sources")
        if isinstance(raw, (list, tuple, set)):
            for entry in raw:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        return sources


class FmrSection(MiniDatabaseSection):
    section_key = "fmr"
    section_title = "FMR"
    supported_suffixes = (".csv",)

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._pixmap_cache: Dict[str, Optional[QtGui.QPixmap]] = {}
        self._record_groups: Dict[str, List[FmrRecord]] = {}
        self._record_groups_by_key: Dict[str, List[FmrRecord]] = {}
        self._hidden_paths: Set[str] = set()
        self._all_records: List[FmrRecord] = []
        self._preview_group_count = 1
        self._preview_spacing = 6
        self._table_splitter: QtWidgets.QSplitter | None = None
        super().__init__(logger, log_callback, parent)
        self._load_hidden_paths()
        if isinstance(self.model, DataFrameModel):
            self.model.set_decoration_provider(self._preview_decoration)
        self.open_graphs_button = QtWidgets.QPushButton("Open graphs")
        self.open_graphs_button.clicked.connect(self._open_selected_graphs)
        self.controls_layout.addWidget(self.open_graphs_button)
        self.open_pyplot_button = QtWidgets.QPushButton("Open in PyPlot")
        self.open_pyplot_button.setToolTip("Open the selected FMR files in PyPlot.")
        self.open_pyplot_button.clicked.connect(self._open_selected_in_pyplot)
        self.controls_layout.addWidget(self.open_pyplot_button)
        self.open_origin_button = QtWidgets.QPushButton("Open in Origin")
        self.open_origin_button.setToolTip("Send the selected FMR files to Origin via PyPlot.")
        self.open_origin_button.clicked.connect(self._open_selected_in_origin)
        self.controls_layout.addWidget(self.open_origin_button)
        self.visibility_button = QtWidgets.QPushButton("Visibility...")
        self.visibility_button.setToolTip("Show or hide specific FMR graphs.")
        self.visibility_button.clicked.connect(self._open_visibility_dialog)
        self.controls_layout.addWidget(self.visibility_button)
        header = self.table_view.verticalHeader() if self.table_view is not None else None
        if header is not None:
            default_height = ANNEALING_GRAPH_HEIGHT + 24
            header.setDefaultSectionSize(default_height)
            header.setMinimumSectionSize(default_height)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        if parse_fmr_csv is None:
            raise RuntimeError("FMR parser is not available.")
        records: List[FmrRecord] = []
        processed: Dict[str, float] = {}
        total = len(paths)
        for idx, path in enumerate(paths, start=1):
            self._check_cancelled()
            try:
                parsed = parse_fmr_csv(Path(path))
                frame = parsed.frame
            except Exception:
                self.logger.exception("Failed to parse %s", path)
                frame = pd.DataFrame()
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                if progress is not None:
                    try:
                        progress(idx, total, f"Skipped {Path(path).name}")
                    except Exception:
                        pass
                continue
            raw_sample = _sample_from_path(Path(path), self.data.sources)
            sample, variant = _split_sample_variant(raw_sample)
            key = _microwire_key_from_path(Path(path), sample or raw_sample)
            label = Path(path).stem
            if variant:
                label = f"{variant} — {label}"
            record = FmrRecord(
                path=Path(path),
                sample=sample,
                data=frame,
                key=key,
                label=label,
            )
            try:
                setattr(record, "units", parsed.units)
            except Exception:
                pass
            setattr(record, "variant", variant)
            records.append(record)
            try:
                processed[str(path)] = float(Path(path).stat().st_mtime)
            except OSError:
                processed[str(path)] = 0.0
            if progress is not None:
                try:
                    progress(idx, total, f"Parsed {Path(path).name}")
                except Exception:
                    pass
        table = _graph_records_to_frame(
            records,
            FMR_COLUMN,
            sample_column="_sample",
        )
        return SectionProcessResult(
            table=table,
            processed=processed,
            payloads={"fmr_records": records},
        )

    def refresh(self) -> None:
        super().refresh()
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _load_hidden_paths(self) -> None:
        hidden = self.data.extra.get("hidden_paths")
        if isinstance(hidden, (list, tuple, set)):
            self._hidden_paths = {str(path) for path in hidden if path}
        else:
            self._hidden_paths = set()

    def _store_hidden_paths(self) -> None:
        self.data.extra["hidden_paths"] = sorted(self._hidden_paths)
        try:
            self.store.save(self.data)
        except Exception:
            self.logger.exception("Failed to persist FMR visibility settings")

    def _visible_records(self, records: Sequence[FmrRecord]) -> List[FmrRecord]:
        if not self._hidden_paths:
            return list(records)
        return [
            record
            for record in records
            if _record_path_key(record) not in self._hidden_paths
        ]

    def _open_visibility_dialog(self) -> None:
        items = _visibility_items_from_records(self._all_records)
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No FMR graphs are available yet.",
            )
            return
        groups = _visibility_groups_from_records(self._all_records)
        dialog = _GraphVisibilityDialog(
            "FMR visibility",
            items,
            self._hidden_paths,
            groups=groups,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._hidden_paths = dialog.hidden_paths()
            self._store_hidden_paths()
            self._refresh_record_groups()

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._load_hidden_paths()
        _drop_visible_sample_column(self)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _handle_worker_finished(self, result: SectionProcessResult) -> None:
        super()._handle_worker_finished(result)
        self._refresh_record_groups()
        self._hide_columns(["Sample", "_sample", "_group_key", "_sources"])

    def _refresh_record_groups(self) -> None:
        grouped: Dict[str, List[FmrRecord]] = {}
        try:
            payload = self.store.load_payload("fmr_records")
        except Exception:
            payload = None
        all_records = list(payload) if isinstance(payload, list) else []
        self._all_records = list(all_records)
        visible_records = self._visible_records(all_records)
        if visible_records:
            for record in visible_records:
                sample = getattr(record, "sample", None)
                if isinstance(sample, str) and sample.strip():
                    existing_variant = getattr(record, "variant", None)
                    variant: Optional[str] = None
                    if isinstance(existing_variant, str) and existing_variant.strip():
                        variant = existing_variant.strip()
                    else:
                        base_sample, parsed_variant = _split_sample_variant(sample)
                        if base_sample:
                            try:
                                record.sample = base_sample
                            except Exception:
                                pass
                            sample = base_sample
                        if parsed_variant:
                            variant = parsed_variant
                        setattr(record, "variant", variant)
                if isinstance(sample, str) and sample.strip():
                    grouped.setdefault(sample, []).append(record)
        self._record_groups = grouped
        self._record_groups_by_key = _group_graph_records_by_key(visible_records)
        max_groups = 1
        for records in grouped.values():
            group_count = len(records)
            if group_count > max_groups:
                max_groups = group_count
        self._preview_group_count = max_groups
        self._update_preview_icon_size()
        self._pixmap_cache.clear()
        if isinstance(self.model, DataFrameModel):
            try:
                self.model.layoutChanged.emit()
            except Exception:
                pass

    def _preview_icon_width(self) -> int:
        count = max(int(self._preview_group_count), 1)
        return ANNEALING_GRAPH_WIDTH * count + self._preview_spacing * (count - 1)

    def _preview_icon_height(self) -> int:
        return ANNEALING_GRAPH_HEIGHT

    def _update_preview_icon_size(self) -> None:
        table = self.table_view
        if not isinstance(table, QtWidgets.QTableView):
            return
        width = self._preview_icon_width()
        height = self._preview_icon_height()
        try:
            table.setIconSize(
                QtCore.QSize(max(width, ANNEALING_GRAPH_WIDTH), max(height, ANNEALING_GRAPH_HEIGHT))
            )
        except Exception:
            pass
        header = table.verticalHeader()
        if header is not None:
            try:
                header.setDefaultSectionSize(max(height + 24, ANNEALING_GRAPH_HEIGHT + 24))
            except Exception:
                pass
        self._auto_fit_columns()

    def _preview_decoration(
        self,
        row: pd.Series,
        column: str,
    ) -> Optional[QtGui.QPixmap]:
        if column != FMR_COLUMN:
            return None
        sample = _row_sample_value(row)
        if not sample:
            return None
        cache_key = f"{sample}|{column}"
        if cache_key in self._pixmap_cache:
            return self._pixmap_cache[cache_key]
        records = self._record_groups.get(sample, [])
        if not records:
            row_key = _row_to_microwire_key(row)
            if row_key:
                records = self._record_groups_by_key.get(row_key, [])
        pixmap: Optional[QtGui.QPixmap] = None
        if records:
            pixmaps: List[QtGui.QPixmap] = []
            for record in records:
                figure = _plot_fmr_figure(
                    record,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                preview = _figure_to_pixmap(
                    figure,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                if preview is not None:
                    pixmaps.append(preview)
            pixmap = _combine_pixmaps_side_by_side(
                pixmaps,
                width_px=self._preview_icon_width(),
                height_px=self._preview_icon_height(),
                spacing=self._preview_spacing,
                scale_to_fit=False,
            )
        self._pixmap_cache[cache_key] = pixmap
        return pixmap

    def _selected_records(self) -> List[FmrRecord]:
        rows = self._selected_rows()
        records: List[FmrRecord] = []
        if not rows:
            return records
        for row_index in rows:
            series = self._row_series(row_index)
            if series is None:
                continue
            sample = _row_sample_value(series)
            if not sample:
                continue
            matched = self._record_groups.get(sample, [])
            if not matched:
                row_key = _row_to_microwire_key(series)
                if row_key:
                    matched = self._record_groups_by_key.get(row_key, [])
            records.extend(matched)
        return records

    def _open_selected_graphs(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        items = _fmr_preview_items(
            records,
            self.logger,
            width_px=GRAPH_PREVIEW_WIDTH,
            height_px=GRAPH_PREVIEW_HEIGHT,
        )
        if not items:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No graphs are available for the selected rows.",
            )
            return
        dialog = _GraphGalleryDialog(
            "FMR graphs",
            items,
            parent=self,
            empty_message="No FMR graphs available.",
        )
        dialog.exec()

    def _open_selected_in_pyplot(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "FMR",
            self.logger,
            auto_plot=True,
            open_origin=False,
        )

    def _open_selected_in_origin(self) -> None:
        records = self._selected_records()
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to open their graphs.",
            )
            return
        paths = [record.path for record in records if isinstance(record.path, Path)]
        paths = list(dict.fromkeys(paths))
        if not paths:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No files are available for the selected rows.",
            )
            return
        _open_pyplot_for_paths(
            paths,
            "FMR",
            self.logger,
            auto_plot=True,
            open_origin=True,
        )

    def _row_sources(self, row: pd.Series) -> List[Path]:
        sources: List[Path] = []
        raw = row.get("_sources")
        if isinstance(raw, (list, tuple, set)):
            for entry in raw:
                if not entry:
                    continue
                try:
                    sources.append(Path(entry))
                except Exception:
                    continue
        return sources


class StrainSection(MiniDatabaseSection):
    section_key = "strain"
    section_title = "Strain data"
    supported_suffixes: tuple[str, ...] = ()
    recursive_search = False

    COLUMN_COMPOSITION = "Composition"
    COLUMN_MICROWIRE = "Microwire"
    COLUMN_DRAW = "Draw"
    COLUMN_PIECE = "Piece"
    COLUMN_D = MICROSCOPE_D_COLUMN
    COLUMN_MODE = "Calc mode"
    COLUMN_CLAMP_SPAN = "Clamp span (mm)"
    COLUMN_MASS = "m"
    COLUMN_TARGET_STRESS = "Legacy stress (MPa)"
    COLUMN_M_LENGTH = "M length"
    COLUMN_A_LENGTH = "A length"
    COLUMN_STRAIN = "Legacy strain"
    COLUMN_BROKE = "Broke"
    TABLE_COLUMNS = [
        COLUMN_COMPOSITION,
        COLUMN_MICROWIRE,
        COLUMN_DRAW,
        COLUMN_PIECE,
        COLUMN_D,
        COLUMN_MODE,
        COLUMN_CLAMP_SPAN,
        COLUMN_MASS,
        COLUMN_TARGET_STRESS,
        COLUMN_M_LENGTH,
        COLUMN_A_LENGTH,
        COLUMN_STRAIN,
        COLUMN_BROKE,
    ]
    HIDDEN_COLUMNS = (COLUMN_DRAW, COLUMN_PIECE, COLUMN_BROKE)
    STRAIN_MODE_LINEAR = "linear"
    STRAIN_MODE_DUAL_SUPPORT = "dual_support"
    COLUMN_ALIASES = {
        "Strain": COLUMN_STRAIN,
        "Stress (MPa)": COLUMN_TARGET_STRESS,
    }

    def __init__(
        self,
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self._wire_choices: Dict[str, Dict[str, tuple[int, int, Optional[str]]]] = {}
        self._d_lookup: Dict[MicrowireKey, float] = {}
        self._suspend_auto_fill = False
        self._editing_index: Optional[int] = None
        self._editing_key: Optional[MicrowireKey] = None
        self._selected_wire_key: Optional[MicrowireKey] = None
        self._mass_override_active = False
        self._suppress_table_selection = False
        self._strain_offsets: Dict[str, float] = {
            self.STRAIN_MODE_LINEAR: 0.0,
            self.STRAIN_MODE_DUAL_SUPPORT: 0.0,
        }
        self._strain_mode: str = self.STRAIN_MODE_LINEAR
        self._clamp_span_mm: float = 0.0
        super().__init__(logger, log_callback, parent)
        self._normalise_legacy_columns()
        self.source_button.hide()
        self.refresh_button.hide()
        if hasattr(self, "open_sources_button"):
            self.open_sources_button.hide()
        self.sources_list.hide()
        self.sources_list.setMaximumWidth(0)
        self.status_label.setWordWrap(True)
        self._reload_strain_settings_from_extra()
        self._ensure_table_structure()
        self._refresh_table_view()
        self._load_reference_data()
        self._sync_payload()
        if hasattr(self, "composition_combo"):
            self._update_composition_suggestions()
        self._update_status()

    def _reload_strain_settings_from_extra(self) -> None:
        if not isinstance(self.data.extra, dict):
            self.data.extra = {}
        stored_offsets = self.data.extra.get("strain_offsets")
        if isinstance(stored_offsets, dict):
            for mode, value in stored_offsets.items():
                if isinstance(value, (int, float)) and mode in self._strain_offsets:
                    self._strain_offsets[mode] = float(value)
        stored_offset = self.data.extra.get("strain_offset")
        if isinstance(stored_offset, (int, float)):
            self._strain_offsets[self.STRAIN_MODE_LINEAR] = float(stored_offset)
        stored_mode = self.data.extra.get("strain_mode")
        stored_span = self.data.extra.get("clamp_span_mm")
        self._strain_mode = stored_mode if isinstance(stored_mode, str) and stored_mode else self.STRAIN_MODE_LINEAR
        self._clamp_span_mm = float(stored_span) if isinstance(stored_span, (int, float)) else 0.0
        self.data.extra["strain_offsets"] = dict(self._strain_offsets)
        self.data.extra["strain_mode"] = self._strain_mode
        self.data.extra["clamp_span_mm"] = self._clamp_span_mm
        try:
            self.store.save(self.data)
        except Exception:
            pass
        if hasattr(self, "strain_offset_spin"):
            blocked = self.strain_offset_spin.blockSignals(True)
            self.strain_offset_spin.setValue(self._current_offset())
            self.strain_offset_spin.blockSignals(blocked)
        if hasattr(self, "strain_mode_combo"):
            blocked = self.strain_mode_combo.blockSignals(True)
            idx = max(0, self.strain_mode_combo.findData(self._strain_mode))
            self.strain_mode_combo.setCurrentIndex(idx)
            self.strain_mode_combo.blockSignals(blocked)
        if hasattr(self, "clamp_span_spin"):
            blocked = self.clamp_span_spin.blockSignals(True)
            self.clamp_span_spin.setValue(max(0.0, self._clamp_span_mm))
            self.clamp_span_spin.blockSignals(blocked)
        self._update_strain_mode_visibility()

    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        container = QtWidgets.QWidget(parent)
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        form_container = QtWidgets.QWidget(container)
        form_layout = QtWidgets.QHBoxLayout(form_container)
        form_layout.setContentsMargins(0, 0, 0, 0)
        form_layout.setSpacing(12)

        def _add_field(label_text: str, widget: QtWidgets.QWidget) -> QtWidgets.QLabel:
            field_box = QtWidgets.QWidget(form_container)
            column = QtWidgets.QVBoxLayout(field_box)
            column.setContentsMargins(0, 0, 0, 0)
            column.setSpacing(4)
            label = QtWidgets.QLabel(label_text, field_box)
            label.setAlignment(
                QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter
            )
            column.addWidget(label)
            column.addWidget(widget)
            form_layout.addWidget(field_box)
            return label

        self.composition_combo = QtWidgets.QComboBox(form_container)
        self.composition_combo.setEditable(True)
        self.composition_combo.setInsertPolicy(QtWidgets.QComboBox.InsertPolicy.NoInsert)
        self.composition_combo.setSizeAdjustPolicy(
            QtWidgets.QComboBox.SizeAdjustPolicy.AdjustToContents
        )
        line_edit = self.composition_combo.lineEdit()
        if line_edit is not None:
            line_edit.textEdited.connect(self._composition_text_edited)
        self.composition_combo.currentTextChanged.connect(self._composition_changed)
        _add_field("Composition", self.composition_combo)

        self.microwire_combo = QtWidgets.QComboBox(form_container)
        self.microwire_combo.currentIndexChanged.connect(self._microwire_changed)
        _add_field("Microwire", self.microwire_combo)

        self.strain_mode_combo = QtWidgets.QComboBox(form_container)
        self.strain_mode_combo.addItem("Single span (straight pull)", self.STRAIN_MODE_LINEAR)
        self.strain_mode_combo.addItem(
            "Dual-point load (A/B supports)",
            self.STRAIN_MODE_DUAL_SUPPORT,
        )
        self.strain_mode_combo.currentIndexChanged.connect(self._strain_mode_changed)
        _add_field(self.COLUMN_MODE, self.strain_mode_combo)

        self.target_stress_spin = QtWidgets.QDoubleSpinBox(form_container)
        self.target_stress_spin.setDecimals(2)
        self.target_stress_spin.setRange(0.0, 1_000_000.0)
        self.target_stress_spin.setSingleStep(5.0)
        self.target_stress_spin.setSuffix(" MPa")
        self.target_stress_spin.valueChanged.connect(self._handle_stress_changed)
        _add_field(self.COLUMN_TARGET_STRESS, self.target_stress_spin)

        self.d_edit = QtWidgets.QLineEdit(form_container)
        self.d_edit.setPlaceholderText("auto")
        self.d_edit.textChanged.connect(self._handle_d_changed)
        _add_field(self.COLUMN_D, self.d_edit)

        self.mass_display = QtWidgets.QLineEdit(form_container)
        self.mass_display.setPlaceholderText("auto")
        self.mass_display.editingFinished.connect(self._update_stress_from_mass)
        _add_field(self.COLUMN_MASS, self.mass_display)

        self.M_length_edit = QtWidgets.QLineEdit(form_container)
        self.M_length_edit.setPlaceholderText("mm")
        self.M_length_edit.textChanged.connect(self._update_strain_display)
        _add_field(self.COLUMN_M_LENGTH, self.M_length_edit)

        self.A_length_edit = QtWidgets.QLineEdit(form_container)
        self.A_length_edit.setPlaceholderText("mm or '-' if broke")
        self.A_length_edit.textChanged.connect(self._update_strain_display)
        _add_field(self.COLUMN_A_LENGTH, self.A_length_edit)

        self.strain_offset_spin = QtWidgets.QDoubleSpinBox(form_container)
        self.strain_offset_spin.setDecimals(6)
        self.strain_offset_spin.setRange(-1000.0, 1000.0)
        self.strain_offset_spin.setSingleStep(0.1)
        self.strain_offset_spin.valueChanged.connect(self._strain_offset_changed)
        _add_field("C offset", self.strain_offset_spin)

        self.clamp_span_spin = QtWidgets.QDoubleSpinBox(form_container)
        self.clamp_span_spin.setDecimals(3)
        self.clamp_span_spin.setRange(0.0, 1_000_000.0)
        self.clamp_span_spin.setSingleStep(1.0)
        self.clamp_span_spin.setSuffix(" mm")
        self.clamp_span_spin.valueChanged.connect(self._clamp_span_changed)
        self.clamp_span_label = _add_field(self.COLUMN_CLAMP_SPAN, self.clamp_span_spin)
        self.clamp_span_container = self.clamp_span_spin.parent()

        self.strain_display = QtWidgets.QLineEdit(form_container)
        self.strain_display.setReadOnly(True)
        _add_field(f"{self.COLUMN_STRAIN} (%)", self.strain_display)

        form_layout.addStretch(1)

        layout.addWidget(form_container)

        self.setTabOrder(self.composition_combo, self.microwire_combo)
        self.setTabOrder(self.microwire_combo, self.strain_mode_combo)
        self.setTabOrder(self.strain_mode_combo, self.target_stress_spin)
        self.setTabOrder(self.target_stress_spin, self.d_edit)
        self.setTabOrder(self.d_edit, self.mass_display)
        self.setTabOrder(self.mass_display, self.M_length_edit)
        self.setTabOrder(self.M_length_edit, self.A_length_edit)
        self.setTabOrder(self.A_length_edit, self.strain_offset_spin)
        self.setTabOrder(self.strain_offset_spin, self.clamp_span_spin)
        self.setTabOrder(self.clamp_span_spin, self.strain_display)

        enter_shortcut = QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key.Key_Return), form_container)
        enter_shortcut.setContext(QtCore.Qt.ShortcutContext.WidgetWithChildrenShortcut)
        enter_shortcut.activated.connect(self._save_entry)
        enter_keypad_shortcut = QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key.Key_Enter), form_container)
        enter_keypad_shortcut.setContext(QtCore.Qt.ShortcutContext.WidgetWithChildrenShortcut)
        enter_keypad_shortcut.activated.connect(self._save_entry)

        button_row = QtWidgets.QHBoxLayout()
        self.add_update_button = QtWidgets.QPushButton("Add entry")
        self.add_update_button.clicked.connect(self._save_entry)
        button_row.addWidget(self.add_update_button)

        self.clear_button = QtWidgets.QPushButton("Clear")
        self.clear_button.clicked.connect(self._clear_form)
        button_row.addWidget(self.clear_button)

        self.delete_button = QtWidgets.QPushButton("Remove entry")
        self.delete_button.clicked.connect(self._delete_selected)
        self.delete_button.setEnabled(False)
        button_row.addWidget(self.delete_button)

        button_row.addStretch(1)

        self.refresh_sources_button = QtWidgets.QPushButton("Reload data")
        self.refresh_sources_button.clicked.connect(self._reload_references_clicked)
        button_row.addWidget(self.refresh_sources_button)

        self.export_button = QtWidgets.QPushButton("Export to Excel…")
        self.export_button.clicked.connect(self._export_to_excel)
        self.export_button.setEnabled(False)
        button_row.addWidget(self.export_button)

        layout.addLayout(button_row)

        self.table_view = QtWidgets.QTableView(container)
        self.table_view.setModel(self.model)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.table_view.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        self.table_view.setSortingEnabled(True)
        selection_model = self.table_view.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._handle_table_selection)
        layout.addWidget(self.table_view, 1)

        self._configure_combo_popup(self.composition_combo)
        self._configure_combo_popup(self.microwire_combo)

        return container

    def reset_to_blank(self) -> None:  # type: ignore[override]
        super().reset_to_blank()
        self._reload_strain_settings_from_extra()
        self._recompute_table_metrics()
        self._refresh_table_view()
        self._update_status()

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        self._normalise_legacy_columns()
        self._reload_strain_settings_from_extra()
        self._recompute_table_metrics()
        self._refresh_table_view()
        self._update_status()

    def _update_status(self) -> None:
        entries = len(self.data.table.index) if isinstance(self.data.table, pd.DataFrame) else 0
        entry_word = "entry" if entries == 1 else "entries"
        available = self._available_wire_count()
        if not self._wire_choices:
            suffix = "Process current annealing data to populate suggestions."
        elif available:
            suffix = f"{available} microwire(s) awaiting strain logging."
        else:
            suffix = "All processed microwires are represented."
        warning = ""
        if self._strain_mode == self.STRAIN_MODE_DUAL_SUPPORT and not self._clamp_span():
            warning = " Set clamp span for dual-point stress mode."
        self.status_label.setText(f"{entries} strain {entry_word} stored. {suffix}{warning}")
        if hasattr(self, "export_button"):
            self.export_button.setEnabled(entries > 0)
        if hasattr(self, "delete_button"):
            has_selection = self._editing_index is not None and entries > 0
            self.delete_button.setEnabled(has_selection)

    def refresh(self) -> None:
        self._load_reference_data()
        if hasattr(self, "composition_combo"):
            self._update_composition_suggestions()
        self._update_status()

    def _composition_text_edited(self, _: str) -> None:
        if self._suspend_auto_fill:
            return
        self.composition_combo.showPopup()
        self._update_microwire_options()

    def _composition_changed(self, _: str) -> None:
        if self._suspend_auto_fill:
            return
        self._update_microwire_options()

    def _microwire_changed(self) -> None:
        comp = self.composition_combo.currentText().strip()
        data = self.microwire_combo.currentData()
        key: Optional[tuple[int, int, Optional[str]]] = None
        if isinstance(data, tuple):
            suffix = None
            if len(data) >= 3 and data[2]:
                suffix = str(data[2]).strip() or None
            key = (int(data[0]), int(data[1]), suffix)
        else:
            parsed = _microwire_parts_from_label_safe(self.microwire_combo.currentText())
            if parsed:
                key = (int(parsed[0]), int(parsed[1]), parsed[2])
        if key is None:
            self._selected_wire_key = None
            if not self._suspend_auto_fill:
                self.d_edit.clear()
                self._refresh_mass_or_stress()
            return
        self._selected_wire_key = (comp, key[0], key[1], key[2])
        if self._suspend_auto_fill:
            return
        d_value = self._d_lookup.get(self._selected_wire_key)
        if d_value is None:
            d_value = self._d_lookup.get((comp, key[0], key[1], None))
        if d_value is not None:
            self.d_edit.setText(f"{d_value:.4f}")
        self._refresh_mass_or_stress()

    def _handle_stress_changed(self, _: float) -> None:
        self._mass_override_active = False
        self._update_mass_display()

    def _handle_d_changed(self, _: str) -> None:
        self._refresh_mass_or_stress()

    def _refresh_mass_or_stress(self) -> None:
        if self._mass_override_active:
            self._update_stress_from_mass()
        else:
            self._update_mass_display()

    def _update_mass_display(self) -> None:
        self._mass_override_active = False
        value = _parse_strain_float(self.d_edit.text())
        target = _parse_strain_float(self.target_stress_spin.value() if hasattr(self, "target_stress_spin") else None)
        mass = self._calculate_mass(
            value,
            area_multiplier=self._cross_section_multiplier(),
            target_stress_mpa=target,
        )
        blocker = QtCore.QSignalBlocker(self.mass_display)
        if mass is None:
            self.mass_display.setText("")
        else:
            self.mass_display.setText(f"{mass:.6f}")
        del blocker

    def _update_stress_from_mass(self) -> None:
        mass_value = _parse_strain_float(self.mass_display.text())
        if mass_value is None:
            self._mass_override_active = False
            self._update_mass_display()
            return
        d_value = _parse_strain_float(self.d_edit.text())
        stress = self._calculate_stress(
            d_value,
            area_multiplier=self._cross_section_multiplier(),
            mass_g=mass_value,
        )
        if stress is None:
            return
        self._mass_override_active = True
        if hasattr(self, "target_stress_spin"):
            blocker = QtCore.QSignalBlocker(self.target_stress_spin)
            self.target_stress_spin.setValue(max(0.0, float(stress)))
            del blocker

    def _update_strain_display(self) -> None:
        text = self.A_length_edit.text().strip()
        if text == "-" or text.lower() == "broke":
            self.strain_display.setText("broke")
            return
        m_length = _parse_strain_float(self.M_length_edit.text())
        a_length = _parse_strain_float(text)
        if m_length in (None, 0) or a_length is None:
            self.strain_display.setText("")
            return
        percent = self._compute_strain_percent(m_length, a_length)
        if percent is None:
            self.strain_display.setText("")
            return
        self.strain_display.setText(f"{percent:.3f}")

    def _handle_table_selection(self, *_: Any) -> None:
        if getattr(self, "_suppress_table_selection", False):
            return
        self._load_row(self._selected_row_index())

    def _selected_row_index(self) -> Optional[int]:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return None
        selection = self.table_view.selectionModel()
        if selection is None:
            return None
        rows = selection.selectedRows()
        if not rows:
            return None
        return rows[0].row()

    def _load_row(self, row_index: Optional[int]) -> None:
        if row_index is None or row_index < 0 or row_index >= len(self.data.table.index):
            self._editing_index = None
            self._editing_key = None
            self._suspend_auto_fill = True
            self.composition_combo.setEditText("")
            self.microwire_combo.clear()
            self.d_edit.clear()
            self.mass_display.clear()
            self.M_length_edit.clear()
            self.A_length_edit.clear()
            self.strain_display.clear()
            self._suspend_auto_fill = False
            self.add_update_button.setText("Add entry")
            self._update_status()
            return

        row = self.data.table.iloc[row_index]
        composition = str(row.get(self.COLUMN_COMPOSITION) or "").strip()
        microwire = str(row.get(self.COLUMN_MICROWIRE) or "").strip()
        draw = row.get(self.COLUMN_DRAW)
        piece = row.get(self.COLUMN_PIECE)
        key: Optional[MicrowireKey] = None
        suffix: Optional[str] = None
        parsed_label: Optional[Tuple[int, int, Optional[str]]] = None
        if microwire:
            parsed_label = _microwire_parts_from_label_safe(microwire)
            if parsed_label is not None:
                suffix = parsed_label[2]
        if pd.notna(draw) and pd.notna(piece):
            try:
                key = (composition, int(float(draw)), int(float(piece)), suffix)
            except (TypeError, ValueError):
                key = None
        if key is None and parsed_label is not None:
            key = (composition, int(parsed_label[0]), int(parsed_label[1]), suffix)
        self._editing_index = row_index
        self._editing_key = key

        self._suspend_auto_fill = True
        self.composition_combo.setEditText(composition)
        self._update_composition_suggestions()
        self._update_microwire_options()
        if microwire:
            idx = self.microwire_combo.findText(microwire)
            if idx >= 0:
                self.microwire_combo.setCurrentIndex(idx)
            elif key is not None:
                self.microwire_combo.insertItem(0, microwire, (key[1], key[2], key[3]))
                self.microwire_combo.setCurrentIndex(0)
        d_value = _parse_strain_float(row.get(self.COLUMN_D))
        d_block = QtCore.QSignalBlocker(self.d_edit)
        self.d_edit.setText("" if d_value is None else f"{d_value:.4f}")
        del d_block
        mass_value = _parse_strain_float(row.get(self.COLUMN_MASS))
        mass_block = QtCore.QSignalBlocker(self.mass_display)
        self.mass_display.setText("" if mass_value is None else f"{mass_value:.6f}")
        del mass_block
        mode_value = str(row.get(self.COLUMN_MODE) or "").strip()
        if hasattr(self, "strain_mode_combo") and mode_value:
            idx = self.strain_mode_combo.findData(mode_value)
            if idx >= 0:
                self.strain_mode_combo.setCurrentIndex(idx)
        span_value = _parse_strain_float(row.get(self.COLUMN_CLAMP_SPAN))
        if hasattr(self, "clamp_span_spin"):
            self.clamp_span_spin.setValue(span_value if span_value is not None else 0.0)
        target_stress = _parse_strain_float(row.get(self.COLUMN_TARGET_STRESS))
        if hasattr(self, "target_stress_spin"):
            stress_block = QtCore.QSignalBlocker(self.target_stress_spin)
            self.target_stress_spin.setValue(target_stress if target_stress is not None else 0.0)
            del stress_block
        m_length = _parse_strain_float(row.get(self.COLUMN_M_LENGTH))
        self.M_length_edit.setText("" if m_length is None else f"{m_length:.4f}")
        a_entry = row.get(self.COLUMN_A_LENGTH)
        if isinstance(a_entry, str) and a_entry.strip():
            self.A_length_edit.setText(a_entry)
        else:
            a_length = _parse_strain_float(a_entry)
            self.A_length_edit.setText("" if a_length is None else f"{a_length:.4f}")
        strain_value = row.get(self.COLUMN_STRAIN)
        if isinstance(strain_value, str) and strain_value.strip().lower() == "broke":
            self.strain_display.setText("broke")
        else:
            strain_float = _parse_strain_float(strain_value)
            self.strain_display.setText("" if strain_float is None else f"{strain_float:.3f}")
        self._suspend_auto_fill = False
        self.add_update_button.setText("Update entry")
        self._update_status()
        self._mass_override_active = mass_value is not None
        self._refresh_mass_or_stress()
        self._update_strain_mode_visibility()

    def _clear_form(self) -> None:
        self._editing_index = None
        self._editing_key = None
        self._selected_wire_key = None
        if isinstance(self.table_view, QtWidgets.QTableView):
            self.table_view.clearSelection()
        self._suspend_auto_fill = True
        self.composition_combo.setEditText("")
        self.microwire_combo.clear()
        self.d_edit.clear()
        self.mass_display.clear()
        if hasattr(self, "target_stress_spin"):
            stress_block = QtCore.QSignalBlocker(self.target_stress_spin)
            self.target_stress_spin.setValue(0.0)
            del stress_block
        self._mass_override_active = False
        self.M_length_edit.clear()
        self.A_length_edit.clear()
        self.strain_display.clear()
        self._suspend_auto_fill = False
        self.add_update_button.setText("Add entry")
        if hasattr(self, "delete_button"):
            self.delete_button.setEnabled(False)
        self._update_composition_suggestions()
        self._update_status()

    def _save_entry(self) -> None:
        composition = self.composition_combo.currentText().strip()
        if not composition:
            QtWidgets.QMessageBox.warning(self, self.section_title, "Enter a composition.")
            return
        if self.microwire_combo.count() == 0:
            QtWidgets.QMessageBox.warning(self, self.section_title, "Select a microwire.")
            return
        label = self.microwire_combo.currentText().strip()
        data = self.microwire_combo.currentData()
        key: Optional[MicrowireKey] = None
        if isinstance(data, tuple):
            suffix = None
            if len(data) >= 3 and data[2]:
                suffix = str(data[2]).strip() or None
            key = (composition, int(data[0]), int(data[1]), suffix)
        elif self._selected_wire_key and self._selected_wire_key[0] == composition:
            key = self._selected_wire_key
        else:
            parsed = _microwire_parts_from_label_safe(label)
            if parsed:
                key = (composition, int(parsed[0]), int(parsed[1]), parsed[2])
        if key is None:
            QtWidgets.QMessageBox.warning(
                self,
                self.section_title,
                "Unable to determine the draw/piece for the selected microwire.",
            )
            return
        used = self._used_wire_keys()
        if self._editing_key in used:
            used.discard(self._editing_key)
        if key in used:
            QtWidgets.QMessageBox.warning(
                self,
                self.section_title,
                "This microwire already has a strain entry.",
            )
            return

        d_value = _parse_strain_float(self.d_edit.text())
        if d_value is None or d_value <= 0:
            QtWidgets.QMessageBox.warning(self, self.section_title, "Enter a valid diameter.")
            return
        stress_value = float(self.target_stress_spin.value()) if hasattr(self, "target_stress_spin") else None
        mass_value = _parse_strain_float(self.mass_display.text())
        if mass_value is not None and mass_value > 0:
            computed_stress = self._calculate_stress(
                d_value,
                area_multiplier=self._cross_section_multiplier(),
                mass_g=mass_value,
            )
            if computed_stress is None or computed_stress <= 0:
                QtWidgets.QMessageBox.warning(
                    self, self.section_title, "Enter a valid weight to compute stress."
                )
                return
            stress_value = float(computed_stress)
            self._mass_override_active = True
            if hasattr(self, "target_stress_spin"):
                blocked = QtCore.QSignalBlocker(self.target_stress_spin)
                self.target_stress_spin.setValue(stress_value)
                del blocked
        else:
            if stress_value is None or stress_value <= 0:
                QtWidgets.QMessageBox.warning(self, self.section_title, "Set stress (MPa) before saving.")
                return
            mass_value = self._calculate_mass(
                d_value,
                area_multiplier=self._cross_section_multiplier(),
                target_stress_mpa=stress_value,
            )
            self._mass_override_active = False

        m_length = _parse_strain_float(self.M_length_edit.text())
        a_text = self.A_length_edit.text().strip()
        broke = a_text == "-" or a_text.lower() == "broke"
        a_length = None if broke else _parse_strain_float(a_text)
        if not broke and a_text and a_length is None:
            QtWidgets.QMessageBox.warning(self, self.section_title, "Enter a valid A length or '-' if the wire broke.")
            return
        strain_percent = None
        if broke:
            strain_display = "broke"
            a_display: object = "-"
        else:
            a_display = a_length
            if m_length not in (None, 0) and a_length is not None:
                strain_percent = self._compute_strain_percent(m_length, a_length)
            strain_display = "" if strain_percent is None else f"{strain_percent:.3f}"

        row_data = {
            self.COLUMN_COMPOSITION: composition,
            self.COLUMN_MICROWIRE: label,
            self.COLUMN_DRAW: key[1],
            self.COLUMN_PIECE: key[2],
            self.COLUMN_D: d_value,
            self.COLUMN_MODE: self._strain_mode,
            self.COLUMN_CLAMP_SPAN: self._clamp_span_mm if self._strain_mode == self.STRAIN_MODE_DUAL_SUPPORT else None,
            self.COLUMN_MASS: mass_value,
            self.COLUMN_TARGET_STRESS: stress_value,
            self.COLUMN_M_LENGTH: m_length,
            self.COLUMN_A_LENGTH: a_display,
            self.COLUMN_STRAIN: strain_display if broke else (None if strain_percent is None else round(strain_percent, 3)),
            self.COLUMN_BROKE: broke,
        }

        updating_existing = self._editing_index is not None and 0 <= self._editing_index < len(self.data.table.index)
        if updating_existing:
            idx = self._editing_index
        else:
            idx = len(self.data.table.index)
        self.data.table.loc[idx, :] = row_data
        self._editing_index = idx if updating_existing else None
        self._editing_key = key if updating_existing else None
        suppress_selection = not updating_existing
        if suppress_selection:
            self._suppress_table_selection = True
        try:
            self._save_table()
        finally:
            if suppress_selection:
                QtCore.QTimer.singleShot(0, self._resume_table_selection)
        if updating_existing:
            self._select_key(key)
        else:
            self._clear_form()
            QtCore.QTimer.singleShot(0, self._focus_entry_selector)
        self.log(
            f"Strain: recorded {composition} {label}"
        )

    def _resume_table_selection(self) -> None:
        self._suppress_table_selection = False

    def _delete_selected(self) -> None:
        row_index = self._selected_row_index()
        if row_index is None:
            return
        self.data.table = self.data.table.drop(index=row_index).reset_index(drop=True)
        self._editing_index = None
        self._editing_key = None
        self._selected_wire_key = None
        self._save_table()
        self._clear_form()
        self.log("Strain: removed selected entry.")

    def _reload_references_clicked(self) -> None:
        before = self._available_wire_count()
        self._load_reference_data()
        if hasattr(self, "composition_combo"):
            self._update_composition_suggestions()
        after = self._available_wire_count()
        self.log(f"Strain: reloaded reference data ({after} microwire option(s) available).")
        self._update_status()

    def _export_to_excel(self) -> None:
        if self.data.table.empty:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No strain entries to export.",
            )
            return
        suggested = Path.cwd() / "strain_data.xlsx"
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export strain worksheet",
            str(suggested),
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return
        export_path = Path(path)
        if export_path.suffix.lower() != ".xlsx":
            export_path = export_path.with_suffix(".xlsx")
        frame = self.data.table[
            [
                self.COLUMN_COMPOSITION,
                self.COLUMN_MICROWIRE,
                self.COLUMN_MODE,
                self.COLUMN_CLAMP_SPAN,
                self.COLUMN_D,
                self.COLUMN_MASS,
                self.COLUMN_TARGET_STRESS,
                self.COLUMN_M_LENGTH,
                self.COLUMN_A_LENGTH,
                self.COLUMN_STRAIN,
            ]
        ].copy()
        mode_labels = {
            self.STRAIN_MODE_LINEAR: "Single span",
            self.STRAIN_MODE_DUAL_SUPPORT: "Dual span",
        }
        if self.COLUMN_MODE in frame.columns:
            def _label_mode(value: object) -> object:
                if value is None:
                    return ""
                text = str(value).strip()
                return mode_labels.get(text, value)

            frame[self.COLUMN_MODE] = frame[self.COLUMN_MODE].map(_label_mode)
        try:
            frame.to_excel(export_path, index=False)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(
                self,
                self.section_title,
                f"Failed to export worksheet:\n{exc}",
            )
            return
        self.log(f"Strain: exported worksheet to {export_path}")

    def _update_composition_suggestions(self) -> None:
        if not hasattr(self, "composition_combo"):
            return
        current_text = self.composition_combo.currentText().strip()
        available: List[str] = []
        used = self._used_wire_keys()
        if self._editing_key in used:
            used.discard(self._editing_key)
        for composition, wires in self._wire_choices.items():
            if any(
                (composition, draw, piece, suffix) not in used
                for draw, piece, suffix in wires.values()
            ):
                available.append(composition)
        available.sort(key=lambda value: value.lower())
        was_blocked = self.composition_combo.blockSignals(True)
        self.composition_combo.clear()
        for composition in available:
            self.composition_combo.addItem(composition)
        self.composition_combo.setEditText(current_text)
        self.composition_combo.blockSignals(was_blocked)
        completer = self.composition_combo.completer()
        if completer is not None:
            completer.setCaseSensitivity(QtCore.Qt.CaseSensitivity.CaseInsensitive)
            completer.setModel(QtCore.QStringListModel(available, completer))
        self._update_microwire_options()
        self._configure_combo_popup(self.composition_combo)

    def _update_microwire_options(self) -> None:
        if not hasattr(self, "microwire_combo"):
            return
        composition = self.composition_combo.currentText().strip()
        was_blocked = self.microwire_combo.blockSignals(True)
        current_label = self.microwire_combo.currentText().strip()
        self.microwire_combo.clear()
        used = self._used_wire_keys()
        if self._editing_key in used:
            used.discard(self._editing_key)
        options = []
        for label, key in self._wire_choices.get(composition, {}).items():
            if (composition, key[0], key[1], key[2]) in used:
                continue
            options.append((label, key))
        options.sort(
            key=lambda item: (
                item[1][0],
                item[1][1],
                str(item[1][2] or ""),
                item[0],
            )
        )
        for label, key in options:
            self.microwire_combo.addItem(label, key)
        if self._editing_key and self._editing_key[0] == composition:
            draw, piece, suffix = self._editing_key[1:]
            label = _microwire_label(draw, piece, suffix)
            if label and self.microwire_combo.findText(label) == -1:
                self.microwire_combo.insertItem(0, label, (draw, piece, suffix))
        if current_label:
            idx = self.microwire_combo.findText(current_label)
            if idx >= 0:
                self.microwire_combo.setCurrentIndex(idx)
        if self.microwire_combo.count() > 0 and self.microwire_combo.currentIndex() < 0:
            self.microwire_combo.setCurrentIndex(0)
        self.microwire_combo.blockSignals(was_blocked)
        self._microwire_changed()
        self._configure_combo_popup(self.microwire_combo)

    def _focus_entry_selector(self) -> None:
        if not hasattr(self, "composition_combo"):
            return
        try:
            self.composition_combo.setFocus(QtCore.Qt.FocusReason.OtherFocusReason)
        except Exception:
            pass
        self._configure_combo_popup(self.composition_combo)
        try:
            self.composition_combo.showPopup()
        except Exception:
            pass

    def _configure_combo_popup(self, combo: QtWidgets.QComboBox) -> None:
        view = combo.view()
        if view is None:
            return
        try:
            screen = QtGui.QGuiApplication.screenAt(
                combo.mapToGlobal(combo.rect().center())
            )
            if screen is None:
                screen = QtGui.QGuiApplication.primaryScreen()
        except Exception:
            screen = QtGui.QGuiApplication.primaryScreen()
        available_height = screen.availableGeometry().height() if screen is not None else 900
        max_height = max(200, available_height - 240)
        model = view.model()
        row_count = model.rowCount() if model is not None else 0
        row_height = view.sizeHintForRow(0)
        if row_height <= 0:
            row_height = view.fontMetrics().height() + 8
        max_items = max(6, max(1, min(row_count or 12, max_height // row_height)))
        combo.setMaxVisibleItems(max_items)
        target_height = min(max_height, row_height * max_items + view.frameWidth() * 2)
        view.setMinimumHeight(max(120, int(target_height)))

    def _refresh_table_view(self) -> None:
        self.model.set_frame(self.data.table)
        if isinstance(self.table_view, QtWidgets.QTableView):
            self._update_hidden_columns()
        self._auto_fit_columns()

    def _update_hidden_columns(self) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        columns = list(self.data.table.columns)
        for name in self.HIDDEN_COLUMNS:
            if name in columns:
                index = columns.index(name)
                self.table_view.setColumnHidden(index, True)

    def _available_wire_count(self) -> int:
        count = 0
        used = self._used_wire_keys()
        for composition, wires in self._wire_choices.items():
            for draw, piece, suffix in wires.values():
                if (composition, draw, piece, suffix) not in used:
                    count += 1
        return count

    def _used_wire_keys(self) -> set[MicrowireKey]:
        keys: set[MicrowireKey] = set()
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            return keys
        for _, row in frame.iterrows():
            composition = str(row.get(self.COLUMN_COMPOSITION) or "").strip()
            if not composition:
                continue
            draw = row.get(self.COLUMN_DRAW)
            piece = row.get(self.COLUMN_PIECE)
            suffix: Optional[str] = None
            parsed = _microwire_parts_from_label_safe(str(row.get(self.COLUMN_MICROWIRE) or ""))
            if parsed:
                suffix = parsed[2]
            if pd.notna(draw) and pd.notna(piece):
                try:
                    draw_int = int(float(draw))
                    piece_int = int(float(piece))
                except (TypeError, ValueError):
                    continue
                keys.add((composition, draw_int, piece_int, suffix))
                continue
            if parsed:
                keys.add((composition, int(parsed[0]), int(parsed[1]), parsed[2]))
        return keys

    def _select_key(self, key: MicrowireKey) -> None:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return
        index = self._find_row_by_key(key)
        if index is None:
            return
        selection_model = self.table_view.selectionModel()
        if selection_model is None:
            return
        model_index = self.model.index(index, 0)
        selection_model.select(
            model_index,
            QtCore.QItemSelectionModel.SelectionFlag.ClearAndSelect
            | QtCore.QItemSelectionModel.SelectionFlag.Rows,
        )
        self.table_view.scrollTo(model_index, QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter)

    def _find_row_by_key(self, key: MicrowireKey) -> Optional[int]:
        if not isinstance(self.data.table, pd.DataFrame) or self.data.table.empty:
            return None
        composition, draw, piece, suffix = key
        for index, row in self.data.table.iterrows():
            row_comp = str(row.get(self.COLUMN_COMPOSITION) or "").strip()
            if row_comp != composition:
                continue
            row_draw = row.get(self.COLUMN_DRAW)
            row_piece = row.get(self.COLUMN_PIECE)
            parsed = _microwire_parts_from_label_safe(str(row.get(self.COLUMN_MICROWIRE) or ""))
            parsed_suffix = parsed[2] if parsed else None
            try:
                row_draw_int = int(float(row_draw))
                row_piece_int = int(float(row_piece))
            except (TypeError, ValueError):
                if parsed is None:
                    continue
                row_draw_int = int(parsed[0])
                row_piece_int = int(parsed[1])
            if row_draw_int != draw or row_piece_int != piece:
                continue
            if (parsed_suffix or None) != (suffix or None):
                continue
            return index
        return None

    def _load_reference_data(self) -> None:
        wire_choices: Dict[str, Dict[str, tuple[int, int, Optional[str]]]] = {}
        try:
            annealing_store = MiniDatabaseStore("annealing")
            records = annealing_store.load_payload("annealing_records")
        except Exception:
            records = None
        if isinstance(records, list):
            for record in records:
                metadata = getattr(record, "metadata", None)
                if metadata is None:
                    continue
                composition = getattr(metadata, "composition_token", None)
                draw = getattr(metadata, "draw_x", None)
                piece = getattr(metadata, "piece_y", None)
                if not composition or draw is None or piece is None:
                    continue
                label = _microwire_label(int(draw), int(piece), None)
                bucket = wire_choices.setdefault(composition, {})
                bucket[label] = (int(draw), int(piece), None)
        self._wire_choices = {
            composition: dict(
                sorted(
                    bucket.items(),
                    key=lambda item: (item[1][0], item[1][1], str(item[1][2] or ""), item[0]),
                )
            )
            for composition, bucket in wire_choices.items()
        }

        d_lookup: Dict[MicrowireKey, float] = {}
        try:
            microscope_data = MiniDatabaseStore("microscope").load()
            frame = microscope_data.table if isinstance(microscope_data.table, pd.DataFrame) else pd.DataFrame()
        except Exception:
            frame = pd.DataFrame()
        if isinstance(frame, pd.DataFrame) and not frame.empty:
            for _, row in frame.iterrows():
                composition = str(row.get("Composition") or "").strip()
                draw_int: Optional[int] = None
                piece_int: Optional[int] = None
                suffix: Optional[str] = None
                key_text = row.get("_key")
                if isinstance(key_text, str) and "|" in key_text:
                    parts = _microwire_key_from_string(key_text)
                    if parts is not None:
                        composition, draw_int, piece_int, suffix = parts
                if draw_int is None or piece_int is None:
                    draw = row.get("Draw")
                    piece = row.get("Piece")
                    if draw is not None and piece is not None:
                        try:
                            draw_int = int(float(draw))
                            piece_int = int(float(piece))
                        except (TypeError, ValueError):
                            draw_int = None
                            piece_int = None
                if (draw_int is None or piece_int is None) and row.get("Microwire"):
                    parsed = _microwire_parts_from_label_safe(str(row.get("Microwire") or ""))
                    if parsed:
                        draw_int, piece_int, suffix = int(parsed[0]), int(parsed[1]), parsed[2]
                if not composition or draw_int is None or piece_int is None:
                    continue
                d_value = _parse_strain_float(row.get(MICROSCOPE_D_COLUMN))
                if d_value is None:
                    continue
                d_lookup[(composition, draw_int, piece_int, suffix)] = d_value
        self._d_lookup = d_lookup

    def _ensure_table_structure(self) -> None:
        frame = self.data.table
        if not isinstance(frame, pd.DataFrame):
            frame = pd.DataFrame(columns=self.TABLE_COLUMNS)
        else:
            frame = frame.copy()
            legacy_map = {
                "Strain (%)": self.COLUMN_STRAIN,
                "Target stress (MPa)": self.COLUMN_TARGET_STRESS,
            }
            for old, new in legacy_map.items():
                if old in frame.columns and new not in frame.columns:
                    frame[new] = frame[old]
            for column in self.TABLE_COLUMNS:
                if column not in frame.columns:
                    frame[column] = pd.Series([None] * len(frame))
            if self.COLUMN_MODE in frame.columns and frame[self.COLUMN_MODE].isna().all():
                frame[self.COLUMN_MODE] = self._strain_mode
            if self.COLUMN_CLAMP_SPAN in frame.columns and frame[self.COLUMN_CLAMP_SPAN].isna().all():
                frame[self.COLUMN_CLAMP_SPAN] = None
            frame = frame[self.TABLE_COLUMNS]
        self.data.table = frame.reset_index(drop=True)
        self._migrate_rows()

    def _migrate_rows(self) -> None:
        frame = self.data.table
        if frame.empty:
            return
        for index, row in frame.iterrows():
            microwire = str(row.get(self.COLUMN_MICROWIRE) or "").strip()
            draw = row.get(self.COLUMN_DRAW)
            piece = row.get(self.COLUMN_PIECE)
            if microwire and (pd.isna(draw) or pd.isna(piece)):
                parsed = _microwire_parts_from_label_safe(microwire)
                if parsed:
                    frame.at[index, self.COLUMN_DRAW] = int(parsed[0])
                    frame.at[index, self.COLUMN_PIECE] = int(parsed[1])
            a_value = row.get(self.COLUMN_A_LENGTH)
            broke = bool(row.get(self.COLUMN_BROKE))
            if isinstance(a_value, str) and a_value.strip() in {"-", "broke"}:
                frame.at[index, self.COLUMN_A_LENGTH] = "-"
                frame.at[index, self.COLUMN_BROKE] = True
            elif broke and a_value != "-":
                frame.at[index, self.COLUMN_A_LENGTH] = "-"
            mode_value = str(row.get(self.COLUMN_MODE) or "").strip()
            if not mode_value:
                frame.at[index, self.COLUMN_MODE] = self._strain_mode
            clamp_span_value = row.get(self.COLUMN_CLAMP_SPAN)
            if pd.isna(clamp_span_value):
                frame.at[index, self.COLUMN_CLAMP_SPAN] = None
        self._recompute_table_metrics()

    def _recompute_table_metrics(self) -> None:
        frame = self.data.table
        for index, row in frame.iterrows():
            mode_value = str(row.get(self.COLUMN_MODE) or "").strip() or self.STRAIN_MODE_LINEAR
            clamp_span_value = _parse_strain_float(row.get(self.COLUMN_CLAMP_SPAN))
            multiplier = 2.0 if mode_value == self.STRAIN_MODE_DUAL_SUPPORT else 1.0
            d_value = _parse_strain_float(row.get(self.COLUMN_D))
            mass_value = _parse_strain_float(row.get(self.COLUMN_MASS))
            stress_value = _parse_strain_float(row.get(self.COLUMN_TARGET_STRESS))
            if mass_value is not None:
                stress_value = self._calculate_stress(
                    d_value,
                    area_multiplier=multiplier,
                    mass_g=mass_value,
                )
                frame.at[index, self.COLUMN_TARGET_STRESS] = (
                    None if stress_value is None else round(stress_value, 3)
                )
                frame.at[index, self.COLUMN_MASS] = round(mass_value, 6)
            else:
                mass = self._calculate_mass(
                    d_value,
                    area_multiplier=multiplier,
                    target_stress_mpa=stress_value,
                )
                frame.at[index, self.COLUMN_MASS] = None if mass is None else round(mass, 6)
            broke = bool(row.get(self.COLUMN_BROKE))
            a_value = row.get(self.COLUMN_A_LENGTH)
            if isinstance(a_value, str) and a_value.strip() in {"-", "broke"}:
                broke = True
                frame.at[index, self.COLUMN_A_LENGTH] = "-"
            if broke:
                frame.at[index, self.COLUMN_BROKE] = True
                frame.at[index, self.COLUMN_STRAIN] = "broke"
                continue
            frame.at[index, self.COLUMN_BROKE] = False
            m_length = _parse_strain_float(row.get(self.COLUMN_M_LENGTH))
            a_length = _parse_strain_float(a_value)
            if m_length in (None, 0) or a_length is None:
                frame.at[index, self.COLUMN_STRAIN] = None
                continue
            percent = self._compute_strain_percent(
                m_length,
                a_length,
                mode=mode_value,
                clamp_span=clamp_span_value,
            )
            frame.at[index, self.COLUMN_STRAIN] = None if percent is None else round(percent, 3)

    def _save_table(self) -> None:
        self._ensure_table_structure()
        self._recompute_table_metrics()
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if not isinstance(frame, pd.DataFrame):
            frame = pd.DataFrame(columns=self.TABLE_COLUMNS)
        missing = [col for col in self.TABLE_COLUMNS if col not in frame.columns]
        for col in missing:
            frame[col] = None
        frame = frame.reindex(columns=self.TABLE_COLUMNS).copy()
        frame.reset_index(drop=True, inplace=True)
        self.data.table = frame
        self._sync_payload()
        self.store.save(self.data)
        self._refresh_table_view()
        if hasattr(self, "composition_combo"):
            self._update_composition_suggestions()
        self._update_status()
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _build_records_from_table(self) -> Dict[MicrowireKey, StrainRecord]:
        records: Dict[MicrowireKey, StrainRecord] = {}
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            return records
        for _, row in frame.iterrows():
            composition = str(row.get(self.COLUMN_COMPOSITION) or "").strip()
            if not composition:
                continue
            draw = row.get(self.COLUMN_DRAW)
            piece = row.get(self.COLUMN_PIECE)
            draw_int: Optional[int] = None
            piece_int: Optional[int] = None
            suffix: Optional[str] = None
            parsed = _microwire_parts_from_label_safe(str(row.get(self.COLUMN_MICROWIRE) or ""))
            if parsed is not None:
                suffix = parsed[2]
            if pd.notna(draw) and pd.notna(piece):
                try:
                    draw_int = int(float(draw))
                    piece_int = int(float(piece))
                except (TypeError, ValueError):
                    draw_int = None
                    piece_int = None
            if draw_int is None or piece_int is None:
                if parsed:
                    draw_int = int(parsed[0])
                    piece_int = int(parsed[1])
            if draw_int is None or piece_int is None:
                continue
            label = _microwire_label(draw_int, piece_int, suffix)
            m_length = _parse_strain_float(row.get(self.COLUMN_M_LENGTH))
            a_value = row.get(self.COLUMN_A_LENGTH)
            broke = bool(row.get(self.COLUMN_BROKE))
            if isinstance(a_value, str) and a_value.strip() in {"-", "broke"}:
                broke = True
            a_length = None
        else:
            a_length = _parse_strain_float(a_value)
        strain_value = row.get(self.COLUMN_STRAIN)
        if broke:
            percent = None
        else:
            percent = _parse_strain_float(strain_value)
            if percent is None and m_length not in (None, 0) and a_length is not None:
                percent = self._compute_strain_percent(m_length, a_length)
        records[(composition, draw_int, piece_int, suffix)] = StrainRecord(
            composition=composition,
            draw=draw_int,
            piece=piece_int,
            microwire_label=label,
            m_length=m_length,
            a_length=a_length,
            percent=percent,
            broke=broke,
            source=Path("manual_entry"),
        )
        return records

    def records_snapshot(self) -> Dict[MicrowireKey, StrainRecord]:
        return self._build_records_from_table()

    def entries_snapshot(self) -> Dict[str, Dict[str, Any]]:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            return {}
        snapshot: Dict[str, Dict[str, Any]] = {}
        for _, row in frame.iterrows():
            composition = str(row.get(self.COLUMN_COMPOSITION) or "").strip()
            if not composition:
                continue
            draw = row.get(self.COLUMN_DRAW)
            piece = row.get(self.COLUMN_PIECE)
            draw_int: Optional[int] = None
            piece_int: Optional[int] = None
            suffix: Optional[str] = None
            parsed = _microwire_parts_from_label_safe(str(row.get(self.COLUMN_MICROWIRE) or ""))
            if parsed:
                suffix = parsed[2]
            if pd.notna(draw) and pd.notna(piece):
                try:
                    draw_int = int(float(draw))
                    piece_int = int(float(piece))
                except (TypeError, ValueError):
                    draw_int = None
                    piece_int = None
            if draw_int is None or piece_int is None:
                if parsed:
                    draw_int = int(parsed[0])
                    piece_int = int(parsed[1])
            if draw_int is None or piece_int is None:
                continue
            key = _microwire_key_to_str((composition, draw_int, piece_int, suffix))
            entry = {str(col): row.get(col) for col in frame.columns}
            snapshot[key] = entry
        return snapshot

    def _sync_payload(self) -> None:
        records = self._build_records_from_table()
        payloads = self.data.extra.get("payloads")
        if not isinstance(payloads, dict):
            payloads = {}
        payloads["strain_records"] = "strain_records"
        self.data.extra["payloads"] = payloads
        self.store.save_payload("strain_records", records)

    def _strain_offset_changed(self, value: float) -> None:
        self._strain_offsets[self._strain_mode] = float(value)
        if not isinstance(self.data.extra, dict):
            self.data.extra = {}
        self.data.extra["strain_offsets"] = dict(self._strain_offsets)
        self.data.extra["strain_offset"] = self._strain_offsets.get(self.STRAIN_MODE_LINEAR, value)
        self._recompute_table_metrics()
        self._refresh_table_view()
        self._sync_payload()
        self.store.save(self.data)
        self._update_status()
        self._update_strain_mode_visibility()
        self._update_strain_display()
        self._refresh_mass_or_stress()
        self._update_strain_mode_visibility()

    def _strain_mode_changed(self, _: int) -> None:
        mode = self.strain_mode_combo.currentData()
        if not isinstance(mode, str) or not mode:
            mode = self.STRAIN_MODE_LINEAR
        self._strain_mode = mode
        if not isinstance(self.data.extra, dict):
            self.data.extra = {}
        self.data.extra["strain_mode"] = self._strain_mode
        self.data.extra["strain_offsets"] = dict(self._strain_offsets)
        self._recompute_table_metrics()
        self._refresh_table_view()
        self._sync_payload()
        self.store.save(self.data)
        self._update_status()
        self._update_strain_mode_visibility()
        if hasattr(self, "strain_offset_spin"):
            blocked = self.strain_offset_spin.blockSignals(True)
            self.strain_offset_spin.setValue(self._current_offset())
            self.strain_offset_spin.blockSignals(blocked)
        self._update_strain_display()
        self._refresh_mass_or_stress()

    def _clamp_span_changed(self, value: float) -> None:
        self._clamp_span_mm = float(value)
        if not isinstance(self.data.extra, dict):
            self.data.extra = {}
        self.data.extra["clamp_span_mm"] = self._clamp_span_mm
        self._recompute_table_metrics()
        self._refresh_table_view()
        self._sync_payload()
        self.store.save(self.data)
        self._update_status()
        self._update_strain_display()

    def _clamp_span(self) -> Optional[float]:
        try:
            span = float(self._clamp_span_mm)
        except (TypeError, ValueError):
            return None
        return span if span > 0 else None

    def _cross_section_multiplier(self) -> float:
        return 2.0 if self._strain_mode == self.STRAIN_MODE_DUAL_SUPPORT else 1.0

    def _current_offset(self) -> float:
        return float(self._strain_offsets.get(self._strain_mode, 0.0))

    def _normalise_legacy_columns(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else None
        if not isinstance(frame, pd.DataFrame):
            return
        updated = frame.copy()
        changed = False
        for old_name, new_name in self.COLUMN_ALIASES.items():
            if old_name not in updated.columns:
                continue
            if new_name not in updated.columns:
                updated = updated.rename(columns={old_name: new_name})
            else:
                mask = updated[new_name].isna() | (updated[new_name] == "")
                updated.loc[mask, new_name] = updated.loc[mask, old_name]
                updated = updated.drop(columns=[old_name])
            changed = True
        for column in self.TABLE_COLUMNS:
            if column not in updated.columns:
                updated[column] = None
        if changed or not frame.columns.equals(updated.columns):
            updated = updated.reindex(columns=self.TABLE_COLUMNS).copy()
            self.data.table = updated
            self.model.set_frame(updated)
            try:
                self.store.save(self.data)
            except Exception:
                pass

    def _update_strain_mode_visibility(self) -> None:
        if hasattr(self, "clamp_span_spin"):
            visible = self._strain_mode == self.STRAIN_MODE_DUAL_SUPPORT
            self.clamp_span_spin.setVisible(visible)
            label = getattr(self, "clamp_span_label", None)
            if isinstance(label, QtWidgets.QWidget):
                label.setVisible(visible)
            parent = getattr(self, "clamp_span_container", None)
            if not isinstance(parent, QtWidgets.QWidget):
                parent = self.clamp_span_spin.parent()
            if isinstance(parent, QtWidgets.QWidget):
                parent.setVisible(visible)
        if hasattr(self, "A_length_edit"):
            placeholder = "mm or '-' if broke"
            if self._strain_mode == self.STRAIN_MODE_DUAL_SUPPORT and self._clamp_span() is None:
                placeholder = "mm or '-' (set clamp span)"
            self.A_length_edit.setPlaceholderText(placeholder)

    def _compute_strain_percent(
        self,
        m_length: Optional[float],
        a_length: Optional[float],
        *,
        mode: Optional[str] = None,
        clamp_span: Optional[float] = None,
    ) -> Optional[float]:
        if m_length in (None, 0) or a_length is None:
            return None
        active_mode = mode or self._strain_mode
        if clamp_span is None and active_mode == self.STRAIN_MODE_DUAL_SUPPORT:
            clamp_span = self._clamp_span()
        try:
            offset = float(self._current_offset())
        except (TypeError, ValueError):
            offset = 0.0
        m_eff = m_length + offset
        a_eff = a_length + offset
        if clamp_span:
            half_span = clamp_span / 2.0
            try:
                initial = math.hypot(half_span, m_eff)
                current = math.hypot(half_span, a_eff)
            except Exception:
                return None
            if current <= 0:
                return None
            base_ratio = (initial - current) / current
        else:
            try:
                base_ratio = (m_eff - a_eff) / a_eff
            except ZeroDivisionError:
                return None
        return base_ratio * 100

    @staticmethod
    def _calculate_mass(
        d_um: Optional[float],
        *,
        area_multiplier: float = 1.0,
        target_stress_mpa: Optional[float] = None,
    ) -> Optional[float]:
        if d_um is None or d_um <= 0:
            return None
        radius_m = (d_um * 1e-6) / 2.0
        if radius_m <= 0:
            return None
        area = math.pi * radius_m * radius_m * max(1.0, float(area_multiplier))
        if not isinstance(target_stress_mpa, (int, float)) or target_stress_mpa <= 0:
            return None
        stress_pa = float(target_stress_mpa) * 1e6
        return stress_pa * area / 9.80665 * 1000.0

    @staticmethod
    def _calculate_stress(
        d_um: Optional[float],
        *,
        area_multiplier: float = 1.0,
        mass_g: Optional[float] = None,
    ) -> Optional[float]:
        if d_um is None or d_um <= 0:
            return None
        if not isinstance(mass_g, (int, float)) or mass_g <= 0:
            return None
        radius_m = (d_um * 1e-6) / 2.0
        if radius_m <= 0:
            return None
        area = math.pi * radius_m * radius_m * max(1.0, float(area_multiplier))
        mass_kg = float(mass_g) / 1000.0
        stress_pa = mass_kg * 9.80665 / area
        return stress_pa / 1e6


class _ColumnSelectionDialog(QtWidgets.QDialog):
    def __init__(
        self,
        column_groups: Mapping[str, Sequence[str]],
        selected: Set[str],
        mandatory: Set[str],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Select columns")
        self._mandatory = set(mandatory)
        self._column_items: Dict[str, List[QtWidgets.QTreeWidgetItem]] = {}
        self._updating = False

        layout = QtWidgets.QVBoxLayout(self)
        intro = QtWidgets.QLabel(
            "Choose which columns to include in the assembled output."
        )
        intro.setWordWrap(True)
        layout.addWidget(intro)

        self.tree = QtWidgets.QTreeWidget()
        self.tree.setHeaderLabel("Columns")
        layout.addWidget(self.tree, 1)

        selected_columns = set(selected) | self._mandatory
        for group_label, columns in column_groups.items():
            if not columns:
                continue
            group_item = QtWidgets.QTreeWidgetItem([group_label])
            group_item.setFlags(
                group_item.flags()
                | QtCore.Qt.ItemFlag.ItemIsUserCheckable
                | QtCore.Qt.ItemFlag.ItemIsAutoTristate
            )
            self.tree.addTopLevelItem(group_item)
            for column in columns:
                item = QtWidgets.QTreeWidgetItem([column])
                item.setData(0, QtCore.Qt.ItemDataRole.UserRole, column)
                item.setFlags(item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
                state = (
                    QtCore.Qt.CheckState.Checked
                    if column in selected_columns
                    else QtCore.Qt.CheckState.Unchecked
                )
                item.setCheckState(0, state)
                group_item.addChild(item)
                self._column_items.setdefault(column, []).append(item)
            self._sync_group_state(group_item)

        self.tree.expandAll()
        self.tree.itemChanged.connect(self._handle_item_changed)

        button_row = QtWidgets.QHBoxLayout()
        select_all_button = QtWidgets.QPushButton("Select all")
        select_all_button.clicked.connect(self._select_all)
        button_row.addWidget(select_all_button)
        select_none_button = QtWidgets.QPushButton("Select none")
        select_none_button.clicked.connect(self._select_none)
        button_row.addWidget(select_none_button)
        button_row.addStretch(1)
        layout.addLayout(button_row)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok
            | QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_columns(self) -> Set[str]:
        selected: Set[str] = set(self._mandatory)
        for column, items in self._column_items.items():
            if any(item.checkState(0) == QtCore.Qt.CheckState.Checked for item in items):
                selected.add(column)
        return selected

    def _select_all(self) -> None:
        self._set_all(QtCore.Qt.CheckState.Checked)

    def _select_none(self) -> None:
        self._set_all(QtCore.Qt.CheckState.Unchecked)

    def _set_all(self, state: QtCore.Qt.CheckState) -> None:
        self._updating = True
        for column, items in self._column_items.items():
            for item in items:
                if column in self._mandatory:
                    item.setCheckState(0, QtCore.Qt.CheckState.Checked)
                else:
                    item.setCheckState(0, state)
        for index in range(self.tree.topLevelItemCount()):
            group_item = self.tree.topLevelItem(index)
            if group_item is not None:
                self._sync_group_state(group_item)
        self._updating = False

    def _handle_item_changed(self, item: QtWidgets.QTreeWidgetItem, _: int) -> None:
        if self._updating:
            return
        column = item.data(0, QtCore.Qt.ItemDataRole.UserRole)
        if column is None:
            state = item.checkState(0)
            if state == QtCore.Qt.CheckState.PartiallyChecked:
                return
            self._updating = True
            for idx in range(item.childCount()):
                child = item.child(idx)
                if child is None:
                    continue
                column_name = child.data(0, QtCore.Qt.ItemDataRole.UserRole)
                if column_name in self._mandatory:
                    child.setCheckState(0, QtCore.Qt.CheckState.Checked)
                else:
                    child.setCheckState(0, state)
                if column_name is not None:
                    for peer in self._column_items.get(column_name, []):
                        if peer is child:
                            continue
                        peer.setCheckState(0, child.checkState(0))
                        parent = peer.parent()
                        if parent is not None:
                            self._sync_group_state(parent)
            self._sync_group_state(item)
            self._updating = False
            return
        if column in self._mandatory and item.checkState(0) != QtCore.Qt.CheckState.Checked:
            self._updating = True
            for peer in self._column_items.get(column, [item]):
                peer.setCheckState(0, QtCore.Qt.CheckState.Checked)
                parent = peer.parent()
                if parent is not None:
                    self._sync_group_state(parent)
            self._updating = False
            return
        state = item.checkState(0)
        peers = self._column_items.get(column, [])
        if len(peers) > 1:
            self._updating = True
            for peer in peers:
                if peer is item:
                    continue
                peer.setCheckState(0, state)
                parent = peer.parent()
                if parent is not None:
                    self._sync_group_state(parent)
            self._updating = False
        parent = item.parent()
        if parent is not None:
            self._updating = True
            self._sync_group_state(parent)
            self._updating = False

    def _sync_group_state(self, group_item: QtWidgets.QTreeWidgetItem) -> None:
        total = group_item.childCount()
        if total == 0:
            group_item.setCheckState(0, QtCore.Qt.CheckState.Unchecked)
            return
        checked = 0
        for idx in range(total):
            child = group_item.child(idx)
            if child is None:
                continue
            if child.checkState(0) == QtCore.Qt.CheckState.Checked:
                checked += 1
        if checked == 0:
            state = QtCore.Qt.CheckState.Unchecked
        elif checked == total:
            state = QtCore.Qt.CheckState.Checked
        else:
            state = QtCore.Qt.CheckState.PartiallyChecked
        group_item.setCheckState(0, state)


class _GraphVisibilityDialog(QtWidgets.QDialog):
    def __init__(
        self,
        title: str,
        items: Sequence[Tuple[str, str]],
        hidden_paths: Set[str],
        *,
        groups: Optional[Mapping[str, Sequence[Tuple[str, str]]]] = None,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self._hidden_paths = set(hidden_paths)
        self._updating_tree = False

        layout = QtWidgets.QVBoxLayout(self)
        intro = QtWidgets.QLabel("Toggle which graphs should be visible.")
        intro.setWordWrap(True)
        layout.addWidget(intro)

        self.tree = QtWidgets.QTreeWidget()
        self.tree.setHeaderHidden(True)
        layout.addWidget(self.tree, 1)
        self.tree.itemChanged.connect(self._handle_item_changed)

        if groups:
            for group_label, group_items in groups.items():
                parent_item = QtWidgets.QTreeWidgetItem([group_label])
                parent_item.setFlags(
                    parent_item.flags()
                    | QtCore.Qt.ItemFlag.ItemIsUserCheckable
                    | QtCore.Qt.ItemFlag.ItemIsAutoTristate
                )
                child_states: List[QtCore.Qt.CheckState] = []
                for label, path in group_items:
                    child = QtWidgets.QTreeWidgetItem([label])
                    child.setData(0, QtCore.Qt.ItemDataRole.UserRole, path)
                    child.setFlags(child.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
                    state = (
                        QtCore.Qt.CheckState.Unchecked
                        if path in self._hidden_paths
                        else QtCore.Qt.CheckState.Checked
                    )
                    child.setCheckState(0, state)
                    child_states.append(state)
                    parent_item.addChild(child)
                if child_states:
                    if all(state == QtCore.Qt.CheckState.Checked for state in child_states):
                        parent_item.setCheckState(0, QtCore.Qt.CheckState.Checked)
                    elif all(state == QtCore.Qt.CheckState.Unchecked for state in child_states):
                        parent_item.setCheckState(0, QtCore.Qt.CheckState.Unchecked)
                    else:
                        parent_item.setCheckState(0, QtCore.Qt.CheckState.PartiallyChecked)
                self.tree.addTopLevelItem(parent_item)
        else:
            for label, path in items:
                item = QtWidgets.QTreeWidgetItem([label])
                item.setData(0, QtCore.Qt.ItemDataRole.UserRole, path)
                item.setFlags(item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
                state = (
                    QtCore.Qt.CheckState.Unchecked
                    if path in self._hidden_paths
                    else QtCore.Qt.CheckState.Checked
                )
                item.setCheckState(0, state)
                self.tree.addTopLevelItem(item)

        button_row = QtWidgets.QHBoxLayout()
        show_all = QtWidgets.QPushButton("Show all")
        show_all.clicked.connect(lambda: self._set_all(QtCore.Qt.CheckState.Checked))
        button_row.addWidget(show_all)
        hide_all = QtWidgets.QPushButton("Hide all")
        hide_all.clicked.connect(lambda: self._set_all(QtCore.Qt.CheckState.Unchecked))
        button_row.addWidget(hide_all)
        button_row.addStretch(1)
        layout.addLayout(button_row)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok
            | QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def hidden_paths(self) -> Set[str]:
        hidden: Set[str] = set()
        for idx in range(self.tree.topLevelItemCount()):
            parent_item = self.tree.topLevelItem(idx)
            if parent_item is None:
                continue
            if parent_item.childCount() == 0:
                path = parent_item.data(0, QtCore.Qt.ItemDataRole.UserRole)
                if isinstance(path, str) and parent_item.checkState(0) != QtCore.Qt.CheckState.Checked:
                    hidden.add(path)
                continue
            for child_idx in range(parent_item.childCount()):
                child = parent_item.child(child_idx)
                if child is None:
                    continue
                path = child.data(0, QtCore.Qt.ItemDataRole.UserRole)
                if not isinstance(path, str):
                    continue
                if child.checkState(0) != QtCore.Qt.CheckState.Checked:
                    hidden.add(path)
        return hidden

    def _set_all(self, state: QtCore.Qt.CheckState) -> None:
        self._updating_tree = True
        try:
            for idx in range(self.tree.topLevelItemCount()):
                item = self.tree.topLevelItem(idx)
                if item is None:
                    continue
                item.setCheckState(0, state)
                for child_idx in range(item.childCount()):
                    child = item.child(child_idx)
                    if child is not None:
                        child.setCheckState(0, state)
        finally:
            self._updating_tree = False

    def _handle_item_changed(self, item: QtWidgets.QTreeWidgetItem, *_: Any) -> None:
        if self._updating_tree:
            return
        self._updating_tree = True
        try:
            if item.childCount() > 0:
                state = item.checkState(0)
                for idx in range(item.childCount()):
                    child = item.child(idx)
                    if child is not None:
                        child.setCheckState(0, state)
            else:
                parent = item.parent()
                if parent is not None:
                    states = [parent.child(idx).checkState(0) for idx in range(parent.childCount())]
                    if all(state == QtCore.Qt.CheckState.Checked for state in states):
                        parent.setCheckState(0, QtCore.Qt.CheckState.Checked)
                    elif all(state == QtCore.Qt.CheckState.Unchecked for state in states):
                        parent.setCheckState(0, QtCore.Qt.CheckState.Unchecked)
                    else:
                        parent.setCheckState(0, QtCore.Qt.CheckState.PartiallyChecked)
        finally:
            self._updating_tree = False


class _ColumnOrderDialog(QtWidgets.QDialog):
    def __init__(
        self,
        columns: Sequence[str],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Reorder columns")
        self._columns = [str(column) for column in columns]

        layout = QtWidgets.QVBoxLayout(self)
        intro = QtWidgets.QLabel("Drag or move columns into the preferred order.")
        intro.setWordWrap(True)
        layout.addWidget(intro)

        self.list = QtWidgets.QListWidget()
        self.list.addItems(self._columns)
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.list.setDragEnabled(True)
        self.list.setDragDropMode(QtWidgets.QAbstractItemView.DragDropMode.InternalMove)
        self.list.setDefaultDropAction(QtCore.Qt.DropAction.MoveAction)
        layout.addWidget(self.list, 1)

        move_row = QtWidgets.QHBoxLayout()
        up_button = QtWidgets.QPushButton("Move up")
        up_button.clicked.connect(lambda: self._move_item(-1))
        move_row.addWidget(up_button)
        down_button = QtWidgets.QPushButton("Move down")
        down_button.clicked.connect(lambda: self._move_item(1))
        move_row.addWidget(down_button)
        move_row.addStretch(1)
        layout.addLayout(move_row)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok
            | QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def ordered_columns(self) -> List[str]:
        return [self.list.item(idx).text() for idx in range(self.list.count())]

    def _move_item(self, delta: int) -> None:
        current = self.list.currentRow()
        if current < 0:
            return
        target = current + delta
        if target < 0 or target >= self.list.count():
            return
        item = self.list.takeItem(current)
        if item is None:
            return
        self.list.insertItem(target, item)
        self.list.setCurrentRow(target)


class _SortSpecDialog(QtWidgets.QDialog):
    def __init__(
        self,
        columns: Sequence[str],
        sort_spec: Sequence[Tuple[str, bool]],
        parent: QtWidgets.QWidget | None = None,
        *,
        default_column: Optional[str] = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Sort columns")
        self._columns = [str(column) for column in columns]
        self._rows: List[Tuple[QtWidgets.QWidget, QtWidgets.QComboBox, QtWidgets.QComboBox]] = []

        layout = QtWidgets.QVBoxLayout(self)
        intro = QtWidgets.QLabel("Define the sort priority for the preview/export.")
        intro.setWordWrap(True)
        layout.addWidget(intro)

        self._rows_layout = QtWidgets.QVBoxLayout()
        layout.addLayout(self._rows_layout)

        if sort_spec:
            for column, ascending in sort_spec:
                self._add_row(column, ascending)
        else:
            fallback = (
                default_column
                if default_column in self._columns
                else (self._columns[0] if self._columns else "")
            )
            if fallback:
                self._add_row(fallback, True)

        control_row = QtWidgets.QHBoxLayout()
        add_button = QtWidgets.QPushButton("Add level")
        add_button.clicked.connect(self._add_default_row)
        control_row.addWidget(add_button)
        clear_button = QtWidgets.QPushButton("Clear")
        clear_button.clicked.connect(self._clear_rows)
        control_row.addWidget(clear_button)
        control_row.addStretch(1)
        layout.addLayout(control_row)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok
            | QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def sort_spec(self) -> List[Tuple[str, bool]]:
        spec: List[Tuple[str, bool]] = []
        seen: Set[str] = set()
        for _, column_combo, order_combo in self._rows:
            column = column_combo.currentText().strip()
            if not column or column in seen:
                continue
            ascending = order_combo.currentIndex() == 0
            spec.append((column, ascending))
            seen.add(column)
        return spec

    def _add_default_row(self) -> None:
        if not self._columns:
            return
        self._add_row(self._columns[0], True)

    def _add_row(self, column: str, ascending: bool) -> None:
        row_widget = QtWidgets.QWidget(self)
        row_layout = QtWidgets.QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)

        column_combo = QtWidgets.QComboBox()
        column_combo.addItems(self._columns)
        if column in self._columns:
            column_combo.setCurrentText(column)
        row_layout.addWidget(column_combo, 2)

        order_combo = QtWidgets.QComboBox()
        order_combo.addItems(["Ascending", "Descending"])
        order_combo.setCurrentIndex(0 if ascending else 1)
        row_layout.addWidget(order_combo, 1)

        remove_button = QtWidgets.QPushButton("Remove")
        remove_button.clicked.connect(lambda: self._remove_row(row_widget))
        row_layout.addWidget(remove_button)

        self._rows_layout.addWidget(row_widget)
        self._rows.append((row_widget, column_combo, order_combo))

    def _remove_row(self, row_widget: QtWidgets.QWidget) -> None:
        self._rows = [row for row in self._rows if row[0] is not row_widget]
        row_widget.setParent(None)
        row_widget.deleteLater()

    def _clear_rows(self) -> None:
        for row_widget, _, _ in list(self._rows):
            self._remove_row(row_widget)


class _AssemblyExportDialog(QtWidgets.QDialog):
    def __init__(
        self,
        *,
        output_dir: str,
        output_name: str,
        export_csv: bool,
        export_excel: bool,
        export_html: bool,
        export_matplotlib: bool,
        export_origin: bool,
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Export settings")

        layout = QtWidgets.QVBoxLayout(self)

        form = QtWidgets.QFormLayout()
        form.setContentsMargins(0, 0, 0, 0)
        form.setSpacing(6)
        self.output_dir_edit = QtWidgets.QLineEdit(output_dir)
        browse_button = QtWidgets.QPushButton("Browse...")
        browse_button.clicked.connect(self._choose_output_dir)
        dir_row = QtWidgets.QHBoxLayout()
        dir_row.addWidget(self.output_dir_edit)
        dir_row.addWidget(browse_button)
        dir_container = QtWidgets.QWidget()
        dir_container.setLayout(dir_row)
        form.addRow("Output directory", dir_container)

        self.output_name_edit = QtWidgets.QLineEdit(output_name)
        form.addRow("Output name", self.output_name_edit)
        layout.addLayout(form)

        formats_box = QtWidgets.QGroupBox("Export formats")
        formats_layout = QtWidgets.QHBoxLayout(formats_box)
        self.csv_checkbox = QtWidgets.QCheckBox("CSV")
        self.csv_checkbox.setChecked(export_csv)
        formats_layout.addWidget(self.csv_checkbox)
        self.excel_checkbox = QtWidgets.QCheckBox("Excel")
        self.excel_checkbox.setChecked(export_excel)
        formats_layout.addWidget(self.excel_checkbox)
        self.html_checkbox = QtWidgets.QCheckBox("HTML (self-contained)")
        self.html_checkbox.setChecked(export_html)
        formats_layout.addWidget(self.html_checkbox)
        formats_layout.addStretch(1)
        layout.addWidget(formats_box)

        plot_box = QtWidgets.QGroupBox("Plot outputs")
        plot_layout = QtWidgets.QHBoxLayout(plot_box)
        self.matplotlib_checkbox = QtWidgets.QCheckBox("Matplotlib images")
        self.matplotlib_checkbox.setChecked(export_matplotlib)
        plot_layout.addWidget(self.matplotlib_checkbox)
        self.origin_checkbox = QtWidgets.QCheckBox("Origin workbooks")
        self.origin_checkbox.setChecked(export_origin)
        plot_layout.addWidget(self.origin_checkbox)
        plot_layout.addStretch(1)
        layout.addWidget(plot_box)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok
            | QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        buttons.button(QtWidgets.QDialogButtonBox.StandardButton.Ok).setText("Export")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def export_settings(self) -> Dict[str, Any]:
        return {
            "output_dir": self.output_dir_edit.text().strip(),
            "output_name": self.output_name_edit.text().strip(),
            "export_csv": self.csv_checkbox.isChecked(),
            "export_excel": self.excel_checkbox.isChecked(),
            "export_html": self.html_checkbox.isChecked(),
            "export_matplotlib": self.matplotlib_checkbox.isChecked(),
            "export_origin": self.origin_checkbox.isChecked(),
        }

    def _choose_output_dir(self) -> None:
        start_dir = _dialog_start_directory(self.output_dir_edit.text().strip())
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select output directory",
            str(start_dir),
        )
        if directory:
            self.output_dir_edit.setText(directory)


class CompareSection(MiniDatabaseSection):
    section_key = "compare"
    section_title = "Compare"
    supported_suffixes: tuple[str, ...] = ()
    recursive_search = False

    def __init__(
        self,
        sections: Dict[str, MiniDatabaseSection],
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        self.sections = sections
        self._cached_annealing_records: List[MeasurementRecord] = []
        self._cached_annealing_groups: Dict[str, List[MeasurementRecord]] = {}
        self._cached_vsm_hysteresis_records: List[VsmHysteresisRecord] = []
        self._cached_vsm_hysteresis_groups: Dict[str, List[VsmHysteresisRecord]] = {}
        self._cached_vsm_temperature_records: List[VsmTemperatureScanRecord] = []
        self._cached_vsm_temperature_groups: Dict[str, List[VsmTemperatureScanRecord]] = {}
        self._cached_dma_isostress_records: List[DmaIsoStressRecord] = []
        self._cached_dma_isostress_groups: Dict[str, List[DmaIsoStressRecord]] = {}
        self._cached_shape_memory_stress_strain_records: List[ShapeMemoryStressStrainRecord] = []
        self._cached_shape_memory_stress_strain_groups: Dict[str, List[ShapeMemoryStressStrainRecord]] = {}
        self._cached_fmr_records: List[FmrRecord] = []
        self._cached_fmr_groups: Dict[str, List[FmrRecord]] = {}
        self._row_keys: Set[str] = set()
        self._compare_columns: List[str] = []
        self._compare_view_mode = "matrix"
        self._compare_fields: Set[str] = set()
        self._compare_field_order: List[str] = []
        self._matrix_column_keys: Dict[str, str] = {}
        self._matrix_graph_rows: Set[str] = set()
        self._matrix_pixmap_cache: Dict[Tuple[object, ...], QtGui.QPixmap] = {}
        self._matrix_inline_graph_columns: Set[str] = set(FIGURE_COLUMNS) | {
            VSM_HYSTERESIS_COLUMN,
            VSM_TEMPERATURE_SCAN_COLUMN,
            DMA_ISOSTRESS_COLUMN,
            SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
            FMR_COLUMN,
        }
        super().__init__(logger, log_callback, parent)
        self.source_button.hide()
        self.open_sources_button.hide()
        self.refresh_button.hide()
        self.stop_button.hide()
        self.progress_bar.hide()
        self.progress_label.hide()
        self.progress_eta_label.hide()

        self.remove_button = QtWidgets.QPushButton("Remove selected")
        self.remove_button.clicked.connect(self._remove_selected)
        self.controls_layout.addWidget(self.remove_button)
        self.clear_button = QtWidgets.QPushButton("Clear")
        self.clear_button.clicked.connect(self._clear_compare)
        self.controls_layout.addWidget(self.clear_button)
        self.view_mode_label = QtWidgets.QLabel("View:")
        self.controls_layout.addWidget(self.view_mode_label)
        self.view_mode_combo = QtWidgets.QComboBox()
        self.view_mode_combo.addItems(["Samples as columns", "Rows"])
        self.view_mode_combo.currentIndexChanged.connect(self._handle_view_mode_changed)
        self.controls_layout.addWidget(self.view_mode_combo)
        self.fields_button = QtWidgets.QPushButton("Fields...")
        self.fields_button.clicked.connect(self._open_compare_field_selector)
        self.controls_layout.addWidget(self.fields_button)
        self.field_order_button = QtWidgets.QPushButton("Order...")
        self.field_order_button.clicked.connect(self._open_compare_field_order_dialog)
        self.controls_layout.addWidget(self.field_order_button)

        graph_row = QtWidgets.QHBoxLayout()
        self.graph_panel_checkbox = QtWidgets.QCheckBox("Show graph preview panel")
        self.graph_panel_checkbox.setChecked(False)
        self.graph_panel_checkbox.toggled.connect(self._toggle_graph_preview_panel)
        graph_row.addWidget(self.graph_panel_checkbox)
        self.open_high_plot_button = QtWidgets.QPushButton("Open 1000 mA graph")
        self.open_high_plot_button.clicked.connect(lambda: self._open_preview_graph("high"))
        self.open_high_plot_button.setEnabled(False)
        graph_row.addWidget(self.open_high_plot_button)
        self.open_low_plot_button = QtWidgets.QPushButton("Open low mA graph")
        self.open_low_plot_button.clicked.connect(lambda: self._open_preview_graph("low"))
        self.open_low_plot_button.setEnabled(False)
        graph_row.addWidget(self.open_low_plot_button)
        self.open_vsm_hysteresis_button = QtWidgets.QPushButton("Open VSM hyst graphs")
        self.open_vsm_hysteresis_button.clicked.connect(
            lambda: self._open_preview_graph("vsm_hysteresis")
        )
        self.open_vsm_hysteresis_button.setEnabled(False)
        graph_row.addWidget(self.open_vsm_hysteresis_button)
        self.open_vsm_temperature_button = QtWidgets.QPushButton("Open VSM temp graphs")
        self.open_vsm_temperature_button.clicked.connect(
            lambda: self._open_preview_graph("vsm_temperature")
        )
        self.open_vsm_temperature_button.setEnabled(False)
        graph_row.addWidget(self.open_vsm_temperature_button)
        self.open_dma_button = QtWidgets.QPushButton("Open DMA iso-stress graphs")
        self.open_dma_button.clicked.connect(lambda: self._open_preview_graph("dma_iso_stress"))
        self.open_dma_button.setEnabled(False)
        graph_row.addWidget(self.open_dma_button)
        self.open_shape_memory_button = QtWidgets.QPushButton("Open shape-memory graphs")
        self.open_shape_memory_button.clicked.connect(
            lambda: self._open_preview_graph("shape_memory_stress_strain")
        )
        self.open_shape_memory_button.setEnabled(False)
        graph_row.addWidget(self.open_shape_memory_button)
        self.open_fmr_button = QtWidgets.QPushButton("Open FMR graphs")
        self.open_fmr_button.clicked.connect(lambda: self._open_preview_graph("fmr"))
        self.open_fmr_button.setEnabled(False)
        graph_row.addWidget(self.open_fmr_button)
        graph_row.addStretch(1)
        self.main_layout.insertLayout(3, graph_row)

        selection_model = self.table_view.selectionModel() if self.table_view is not None else None
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._update_preview_graph_buttons)
            selection_model.selectionChanged.connect(self._update_graph_preview_panel)
        matrix_selection = (
            self.matrix_view.selectionModel() if hasattr(self, "matrix_view") else None
        )
        if matrix_selection is not None:
            matrix_selection.selectionChanged.connect(self._update_preview_graph_buttons)
        self._set_compare_view_mode(self._compare_view_mode)
        self._update_preview_graph_buttons()
        self._update_status()

    def create_right_panel(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        table = QtWidgets.QTableView(parent)
        table.setModel(self.model)
        table.horizontalHeader().setStretchLastSection(True)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        table.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        table.setSortingEnabled(True)
        self.table_view = table

        self.graph_preview_panel = QtWidgets.QWidget(parent)
        graph_layout = QtWidgets.QVBoxLayout(self.graph_preview_panel)
        graph_layout.setContentsMargins(0, 0, 0, 0)
        graph_layout.setSpacing(6)
        self.graph_preview_tabs = QtWidgets.QTabWidget(self.graph_preview_panel)
        graph_layout.addWidget(self.graph_preview_tabs, 1)

        annealing_tab = QtWidgets.QWidget(self.graph_preview_panel)
        annealing_layout = QtWidgets.QVBoxLayout(annealing_tab)
        annealing_layout.setContentsMargins(0, 0, 0, 0)
        annealing_layout.setSpacing(6)
        self.high_preview_display = _AnnealingPlotDisplay(
            "Graph — 1000 mA", self.logger, annealing_tab
        )
        annealing_layout.addWidget(self.high_preview_display, 1)
        self.low_preview_display = _AnnealingPlotDisplay(
            "Graph — low mA", self.logger, annealing_tab
        )
        annealing_layout.addWidget(self.low_preview_display, 1)
        self.graph_preview_tabs.addTab(annealing_tab, "Annealing")

        self.vsm_hysteresis_gallery = _GraphGalleryWidget(
            "Select a row to preview VSM hysteresis graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.vsm_hysteresis_gallery, "VSM hyst")
        self.vsm_temperature_gallery = _GraphGalleryWidget(
            "Select a row to preview VSM temperature scans.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.vsm_temperature_gallery, "VSM temp")
        self.dma_iso_gallery = _GraphGalleryWidget(
            "Select a row to preview DMA iso-stress graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.dma_iso_gallery, "DMA iso-stress")
        self.shape_memory_gallery = _GraphGalleryWidget(
            "Select a row to preview shape-memory graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.shape_memory_gallery, "Shape memory")
        self.fmr_gallery = _GraphGalleryWidget(
            "Select a row to preview FMR graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.fmr_gallery, "FMR")
        self.graph_preview_panel.setVisible(False)

        self.preview_splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal)
        self.preview_splitter.addWidget(table)
        self.preview_splitter.addWidget(self.graph_preview_panel)
        self.preview_splitter.setStretchFactor(0, 3)
        self.preview_splitter.setStretchFactor(1, 2)
        self._table_splitter = self.preview_splitter
        preview_container = QtWidgets.QWidget(parent)
        preview_layout = QtWidgets.QVBoxLayout(preview_container)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_layout.addWidget(self.preview_splitter, 1)

        self.matrix_model = DataFrameModel()
        self.matrix_model.set_decoration_provider(self._matrix_decoration)
        self.matrix_view = QtWidgets.QTableView(parent)
        self.matrix_view.setModel(self.matrix_model)
        self.matrix_view.horizontalHeader().setStretchLastSection(True)
        self.matrix_view.setAlternatingRowColors(True)
        self.matrix_view.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.matrix_view.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        self.matrix_view.setSortingEnabled(False)
        self.matrix_view.setVerticalScrollMode(
            QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.matrix_view.setHorizontalScrollMode(
            QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        try:
            self.matrix_view.setViewportUpdateMode(
                QtWidgets.QAbstractItemView.ViewportUpdateMode.MinimalViewportUpdate
            )
        except Exception:
            pass
        try:
            self.matrix_view.setIconSize(
                QtCore.QSize(ANNEALING_GRAPH_WIDTH, ANNEALING_GRAPH_HEIGHT)
            )
        except Exception:
            pass
        header = self.matrix_view.horizontalHeader()
        if header is not None:
            header.setSectionsMovable(True)
            header.setSectionsClickable(True)
        v_header = self.matrix_view.verticalHeader()
        if v_header is not None:
            v_header.setVisible(False)

        matrix_container = QtWidgets.QWidget(parent)
        matrix_layout = QtWidgets.QVBoxLayout(matrix_container)
        matrix_layout.setContentsMargins(0, 0, 0, 0)
        matrix_layout.addWidget(self.matrix_view, 1)

        self.compare_view_stack = QtWidgets.QStackedWidget(parent)
        self.compare_view_stack.addWidget(preview_container)
        self.compare_view_stack.addWidget(matrix_container)

        container = QtWidgets.QWidget(parent)
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.compare_view_stack, 1)
        return container

    def _load_payload(self, section_key: str, name: str) -> Any:
        section = self.sections.get(section_key)
        if section is None:
            return None
        return section.store.load_payload(name)

    def _row_key(self, row: pd.Series) -> str:
        key = _row_to_microwire_key(row)
        if key:
            return key
        payload = {str(k): str(v) for k, v in row.to_dict().items()}
        return json.dumps(payload, sort_keys=True)

    def _rebuild_row_keys(self, frame: pd.DataFrame) -> None:
        self._row_keys = {self._row_key(row) for _, row in frame.iterrows()}

    def _apply_compare_frame(self, frame: pd.DataFrame) -> None:
        extra = {
            "compare_keys": list(self._row_keys),
            "compare_columns": list(self._compare_columns),
            "compare_view_mode": self._compare_view_mode,
            "compare_fields": list(self._compare_fields),
            "compare_field_order": list(self._compare_field_order),
        }
        self.apply_data(MiniDatabaseData(table=frame, extra=extra))

    def _update_status(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        count = len(frame.index) if isinstance(frame, pd.DataFrame) else 0
        if count:
            message = f"{count} microwire(s) selected for comparison."
        else:
            message = "Select rows in Assemble to compare."
        self.status_label.setText(message)
        try:
            self.status_changed.emit(message)
        except Exception:
            pass

    def _handle_view_mode_changed(self, index: int) -> None:
        mode = "matrix" if index == 0 else "rows"
        self._set_compare_view_mode(mode)
        self._persist_compare_settings()

    def _set_compare_view_mode(self, mode: str) -> None:
        if mode not in {"matrix", "rows"}:
            return
        self._compare_view_mode = mode
        stack = getattr(self, "compare_view_stack", None)
        if isinstance(stack, QtWidgets.QStackedWidget):
            stack.setCurrentIndex(0 if mode == "rows" else 1)
        self._update_compare_view_controls()
        self._update_preview_graph_buttons()
        self._update_graph_preview_panel()

    def _update_compare_view_controls(self) -> None:
        is_matrix = self._compare_view_mode == "matrix"
        if hasattr(self, "fields_button"):
            self.fields_button.setEnabled(is_matrix)
        if hasattr(self, "field_order_button"):
            self.field_order_button.setEnabled(is_matrix)
        if hasattr(self, "graph_panel_checkbox"):
            self.graph_panel_checkbox.setEnabled(not is_matrix)
            if is_matrix:
                try:
                    self.graph_preview_panel.setVisible(False)
                except Exception:
                    pass
                splitter = getattr(self, "preview_splitter", None)
                if isinstance(splitter, QtWidgets.QSplitter):
                    splitter.setSizes([1, 0])
            else:
                self._toggle_graph_preview_panel(self.graph_panel_checkbox.isChecked())

    def _persist_compare_settings(self) -> None:
        extra = self.data.extra if isinstance(self.data.extra, dict) else {}
        extra = dict(extra)
        extra.update(
            {
                "compare_keys": list(self._row_keys),
                "compare_columns": list(self._compare_columns),
                "compare_view_mode": self._compare_view_mode,
                "compare_fields": list(self._compare_fields),
                "compare_field_order": list(self._compare_field_order),
            }
        )
        self.data.extra = extra
        try:
            self.store.save(self.data)
        except Exception:
            pass

    def _active_compare_key(self) -> Optional[str]:
        if self._compare_view_mode == "matrix":
            return self._matrix_selected_key()
        row = self._selected_preview_row()
        return _row_to_microwire_key(row) if row is not None else None

    def _matrix_selected_key(self) -> Optional[str]:
        if not isinstance(getattr(self, "matrix_view", None), QtWidgets.QTableView):
            return None
        selection = self.matrix_view.selectionModel()
        if selection is None:
            return None
        indexes = selection.selectedIndexes()
        if not indexes:
            current = selection.currentIndex()
            if current.isValid():
                indexes = [current]
        if not indexes:
            return None
        frame = self.matrix_model.frame()
        if not isinstance(frame, pd.DataFrame):
            return None
        for index in indexes:
            column_index = index.column()
            if column_index >= len(frame.columns):
                continue
            column_label = str(frame.columns[column_index])
            if column_label == "Field":
                continue
            key = self._matrix_column_keys.get(column_label)
            if key:
                return key
        return None

    def _available_compare_fields(self, frame: pd.DataFrame) -> List[str]:
        fields: List[str] = []
        for column in frame.columns:
            name = str(column)
            if name.startswith("_"):
                continue
            if name not in fields:
                fields.append(name)
        return fields

    def _default_compare_fields(self, fields: Sequence[str]) -> List[str]:
        filtered = [field for field in fields if field not in {"Composition", "Microwire"}]
        return filtered or list(fields)

    def _resolve_compare_field_order(self, fields: Sequence[str]) -> List[str]:
        order = [field for field in self._compare_field_order if field in fields]
        for field in fields:
            if field not in order:
                order.append(field)
        return order

    def _open_compare_field_selector(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Add rows to Compare before selecting fields.",
            )
            return
        fields = self._available_compare_fields(frame)
        if not fields:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "No fields are available for comparison yet.",
            )
            return
        selected = set(self._compare_fields) or set(self._default_compare_fields(fields))
        dialog = _ColumnSelectionDialog({"Fields": fields}, selected, set(), parent=self)
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            selected = dialog.selected_columns()
            if not selected:
                selected = set(self._default_compare_fields(fields))
            self._compare_fields = set(selected)
            self._compare_field_order = [
                field for field in self._compare_field_order if field in self._compare_fields
            ]
            self._update_matrix_view()
            self._persist_compare_settings()

    def _open_compare_field_order_dialog(self) -> None:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        if frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Add rows to Compare before reordering fields.",
            )
            return
        fields = self._available_compare_fields(frame)
        if not fields:
            return
        selected = set(self._compare_fields) or set(self._default_compare_fields(fields))
        ordered = [field for field in self._resolve_compare_field_order(fields) if field in selected]
        dialog = _ColumnOrderDialog(ordered, parent=self)
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            ordered = dialog.ordered_columns()
            for field in selected:
                if field not in ordered:
                    ordered.append(field)
            self._compare_field_order = ordered
            self._compare_fields = set(selected)
            self._update_matrix_view()
            self._persist_compare_settings()

    def _handle_search_changed(self, text: str) -> None:  # type: ignore[override]
        super()._handle_search_changed(text)
        self._update_matrix_view()

    def _build_matrix_frame(self) -> pd.DataFrame:
        frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
        query = ""
        if isinstance(getattr(self, "search_edit", None), QtWidgets.QLineEdit):
            query = self.search_edit.text().strip().lower()
        if query and not frame.empty:
            filtered_rows: List[int] = []
            for idx, row in frame.iterrows():
                for column in frame.columns:
                    label = str(column)
                    if label.startswith("_"):
                        continue
                    value = row.get(column)
                    text = (
                        ", ".join(str(item) for item in value)
                        if isinstance(value, (list, tuple, set))
                        else ("" if value is None else str(value))
                    )
                    if query in text.lower():
                        filtered_rows.append(int(idx))
                        break
            frame = frame.loc[filtered_rows].reset_index(drop=True) if filtered_rows else pd.DataFrame(columns=frame.columns)
        if frame.empty:
            self._matrix_column_keys = {}
            self._matrix_graph_rows = set()
            return pd.DataFrame(columns=["Field"])
        fields = self._available_compare_fields(frame)
        selected = set(self._compare_fields) or set(self._default_compare_fields(fields))
        selected = {field for field in selected if field in fields}
        if not selected:
            selected = set(self._default_compare_fields(fields))
        order = [field for field in self._resolve_compare_field_order(fields) if field in selected]
        for field in selected:
            if field not in order:
                order.append(field)
        self._compare_fields = set(order)
        self._compare_field_order = list(order)

        sample_labels: List[str] = []
        sample_rows: List[pd.Series] = []
        column_keys: Dict[str, str] = {}
        for _, row in frame.iterrows():
            key = _row_to_microwire_key(row)
            if not key:
                continue
            composition = str(row.get("Composition") or "").strip()
            microwire = str(row.get("Microwire") or "").strip()
            label = f"{composition} {microwire}".strip()
            if not label:
                label = key
            base_label = label
            counter = 1
            while label in column_keys:
                counter += 1
                label = f"{base_label} ({counter})"
            column_keys[label] = key
            sample_labels.append(label)
            sample_rows.append(row)

        graph_fields = {field for field in order if field in self._matrix_inline_graph_columns}
        self._matrix_graph_rows = graph_fields
        rows: List[Dict[str, Any]] = []
        for field in order:
            row_data: Dict[str, Any] = {"Field": field}
            for label, row in zip(sample_labels, sample_rows):
                if field in graph_fields:
                    row_data[label] = ""
                else:
                    row_data[label] = row.get(field)
            rows.append(row_data)
        columns = ["Field"] + sample_labels
        self._matrix_column_keys = column_keys
        return pd.DataFrame(rows, columns=columns)

    def _update_matrix_row_heights(self) -> None:
        if not isinstance(getattr(self, "matrix_view", None), QtWidgets.QTableView):
            return
        frame = self.matrix_model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        default_height = self.matrix_view.verticalHeader().defaultSectionSize()
        spacing = 6
        padding = 32
        max_graph_count = 1
        for row_idx in range(len(frame.index)):
            field = str(frame.iloc[row_idx].get("Field") or "")
            if field in self._matrix_graph_rows:
                row_max = 1
                for column in frame.columns:
                    if str(column) == "Field":
                        continue
                    key = self._matrix_column_keys.get(str(column))
                    if not key:
                        continue
                    count = self._matrix_graph_stack_count(field, key)
                    if count > row_max:
                        row_max = count
                max_graph_count = max(max_graph_count, row_max)
                target = ANNEALING_GRAPH_HEIGHT * row_max + spacing * (row_max - 1) + padding
            else:
                target = default_height
            try:
                self.matrix_view.setRowHeight(row_idx, target)
            except Exception:
                continue
        try:
            self.matrix_view.setIconSize(
                QtCore.QSize(
                    ANNEALING_GRAPH_WIDTH,
                    ANNEALING_GRAPH_HEIGHT * max_graph_count + spacing * (max_graph_count - 1),
                )
            )
        except Exception:
            pass

    def _matrix_graph_stack_count(self, field: str, key: str) -> int:
        if field in FIGURE_COLUMNS:
            if field == ANNEALING_OTHER_GRAPH_COLUMN:
                records = self._ensure_annealing_groups().get(key, [])
                if not records:
                    return 1
                high_record, low_record = _select_high_low_pair(records)
                other_records = _select_other_measurements(records, high_record, low_record)
                return max(len(other_records), 1)
            return 1
        if field == VSM_HYSTERESIS_COLUMN:
            records = self._ensure_vsm_hysteresis_groups().get(key, [])
            return len(_group_vsm_hysteresis_plot_groups(records))
        if field == VSM_TEMPERATURE_SCAN_COLUMN:
            records = self._ensure_vsm_temperature_groups().get(key, [])
            return len(records)
        if field == DMA_ISOSTRESS_COLUMN:
            records = self._ensure_dma_isostress_groups().get(key, [])
            return len(records)
        if field == FMR_COLUMN:
            records = self._ensure_fmr_groups().get(key, [])
            return len(records)
        return 1

    def _update_matrix_column_widths(self) -> None:
        if not isinstance(getattr(self, "matrix_view", None), QtWidgets.QTableView):
            return
        frame = self.matrix_model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        try:
            self.matrix_view.resizeColumnsToContents()
        except Exception:
            return
        if not self._matrix_graph_rows:
            return
        graph_width = ANNEALING_GRAPH_WIDTH + 80
        for idx, column in enumerate(frame.columns):
            if str(column) == "Field":
                continue
            current = self.matrix_view.columnWidth(idx)
            if graph_width > current:
                self.matrix_view.setColumnWidth(idx, graph_width)

    def _update_matrix_view(self) -> None:
        if not hasattr(self, "matrix_model"):
            return
        self._matrix_pixmap_cache.clear()
        matrix_frame = self._build_matrix_frame()
        self.matrix_model.set_frame(matrix_frame)
        self._update_matrix_row_heights()
        self._update_matrix_column_widths()

    def _record_signature(self, records: Sequence[object]) -> Tuple[str, ...]:
        signature: List[str] = []
        for record in records:
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                signature.append(str(path))
                continue
            if isinstance(path, str) and path:
                signature.append(path)
                continue
            label = _record_label_for_display(record)
            if label:
                signature.append(label)
                continue
            sample = getattr(record, "sample", None)
            if isinstance(sample, str) and sample:
                signature.append(sample)
                continue
            signature.append(repr(record))
        return tuple(signature)

    def _combined_graph_pixmap(
        self,
        cache_key: Tuple[object, ...],
        items: Sequence[_GraphPreviewItem],
        *,
        stack_vertical: bool = False,
    ) -> Optional[QtGui.QPixmap]:
        cached = self._matrix_pixmap_cache.get(cache_key)
        if cached is not None:
            return cached
        pixmaps = [item.pixmap for item in items if item.pixmap is not None]
        count = max(len(pixmaps), 1)
        spacing = 6
        if stack_vertical:
            combined = _combine_pixmaps_vertical(
                pixmaps,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT * count + spacing * (count - 1),
                spacing=spacing,
                scale_to_fit=False,
            )
        else:
            combined = _combine_pixmaps_side_by_side(
                pixmaps,
                width_px=ANNEALING_GRAPH_WIDTH * count + spacing * (count - 1),
                height_px=ANNEALING_GRAPH_HEIGHT,
                spacing=spacing,
                scale_to_fit=False,
            )
        if combined is not None:
            self._matrix_pixmap_cache[cache_key] = combined
        return combined

    def _matrix_decoration(
        self, row: pd.Series, column_label: str
    ) -> Optional[QtGui.QPixmap]:
        field = str(row.get("Field") or "")
        if field not in self._matrix_inline_graph_columns:
            return None
        if column_label == "Field":
            return None
        matrix_view = getattr(self, "matrix_view", None)
        if isinstance(matrix_view, QtWidgets.QTableView):
            try:
                row_idx = int(row.name)
            except Exception:
                row_idx = None
            if row_idx is None:
                frame = self.matrix_model.frame()
                if isinstance(frame, pd.DataFrame):
                    try:
                        row_idx = int(frame.index.get_loc(row.name))
                    except Exception:
                        row_idx = None
            if row_idx is not None:
                top = matrix_view.rowAt(0)
                bottom = matrix_view.rowAt(matrix_view.viewport().height() - 1)
                if bottom == -1:
                    bottom = top
                if top != -1 and bottom != -1 and (row_idx < top or row_idx > bottom):
                    return None
        key = self._matrix_column_keys.get(column_label)
        if not key:
            return None
        try:
            if field in FIGURE_COLUMNS:
                records = self._ensure_annealing_groups().get(key, [])
                if not records:
                    return None
                high_record, low_record = _select_high_low_pair(records)
                if field == FIGURE_COLUMNS[0]:
                    target = high_record
                elif field == FIGURE_COLUMNS[1]:
                    target = low_record
                else:
                    target = None
                if target is not None:
                    measurement_id = getattr(getattr(target, "metadata", None), "measurement_id", None)
                    cache_key = (
                        "annealing",
                        field,
                        measurement_id or str(getattr(target, "path", "")),
                    )
                    cached = self._matrix_pixmap_cache.get(cache_key)
                    if cached is not None:
                        return cached
                    pixmap = _render_measurement_pixmap(
                        target,
                        self.logger,
                        width_px=ANNEALING_GRAPH_WIDTH,
                        height_px=ANNEALING_GRAPH_HEIGHT,
                    )
                    if pixmap is None:
                        return None
                    self._matrix_pixmap_cache[cache_key] = pixmap
                    return pixmap
                other_records = _select_other_measurements(records, high_record, low_record)
                if not other_records:
                    return None
                signature = tuple(
                    str(getattr(record, "path", "")) for record in other_records
                )
                cache_key = ("annealing_other", key, signature)
                cached = self._matrix_pixmap_cache.get(cache_key)
                if cached is not None:
                    return cached
                pixmaps: List[QtGui.QPixmap] = []
                for record in other_records:
                    preview = _render_measurement_pixmap(
                        record,
                        self.logger,
                        width_px=ANNEALING_GRAPH_WIDTH,
                        height_px=ANNEALING_GRAPH_HEIGHT,
                    )
                    if preview is not None:
                        pixmaps.append(preview)
                combined = _combine_pixmaps_vertical(
                    pixmaps,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT * max(len(pixmaps), 1)
                    + 6 * (max(len(pixmaps), 1) - 1),
                    spacing=6,
                    scale_to_fit=False,
                )
                if combined is not None:
                    self._matrix_pixmap_cache[cache_key] = combined
                return combined
            if field == VSM_HYSTERESIS_COLUMN:
                records = self._ensure_vsm_hysteresis_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("vsm_hysteresis", key, signature)
                items = _vsm_hysteresis_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=True)
            if field == VSM_TEMPERATURE_SCAN_COLUMN:
                records = self._ensure_vsm_temperature_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("vsm_temperature", key, signature)
                items = _vsm_temperature_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=True)
            if field == DMA_ISOSTRESS_COLUMN:
                records = self._ensure_dma_isostress_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("dma_iso_stress", key, signature)
                items = _dma_iso_stress_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=True)
            if field == SHAPE_MEMORY_STRESS_STRAIN_COLUMN:
                records = self._ensure_shape_memory_stress_strain_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("shape_memory_stress_strain", key, signature)
                items = _shape_memory_stress_strain_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=True)
            if field == FMR_COLUMN:
                records = self._ensure_fmr_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("fmr", key, signature)
                items = _fmr_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=True)
        except Exception:
            self.logger.exception("Failed to render compare matrix preview")
        return None

    def process(
        self,
        paths: List[Path],
        progress: Optional[Callable[[int, int, Optional[str]], None]] = None,
    ) -> SectionProcessResult:
        _ = paths, progress
        return SectionProcessResult(table=self.model.frame(), processed={}, payloads={})

    def apply_data(self, data: MiniDatabaseData) -> None:
        super().apply_data(data)
        self._update_status()
        self._update_matrix_view()
        self._update_preview_graph_buttons()

    def refresh(self) -> None:
        return

    def add_rows_from_frame(self, frame: pd.DataFrame, row_indices: Sequence[int]) -> int:
        if not isinstance(frame, pd.DataFrame) or frame.empty or not row_indices:
            return 0
        current = self.model.frame()
        if not isinstance(current, pd.DataFrame):
            current = pd.DataFrame()
        new_rows: List[Dict[str, Any]] = []
        for idx in row_indices:
            if idx < 0 or idx >= len(frame.index):
                continue
            row = frame.iloc[idx]
            key = self._row_key(row)
            if key in self._row_keys:
                continue
            self._row_keys.add(key)
            new_rows.append(row.to_dict())
        if not new_rows:
            return 0
        new_frame = pd.DataFrame(new_rows)
        existing_rows: List[Dict[str, Any]] = []
        if not current.empty:
            try:
                cleaned_current = current.dropna(how="all")
            except Exception:
                cleaned_current = current
            if not cleaned_current.empty:
                existing_rows = [
                    dict(row)
                    for row in cleaned_current.to_dict(orient="records")
                ]
        combined = pd.DataFrame(existing_rows + new_rows) if existing_rows else new_frame
        columns = list(self._compare_columns or combined.columns)
        for column in combined.columns:
            if column not in columns:
                columns.append(column)
        self._compare_columns = columns
        combined = combined.reindex(columns=columns)
        self._apply_compare_frame(combined)
        try:
            if self.table_view is not None:
                selection = self.table_view.selectionModel()
                if selection is not None:
                    selection.clearSelection()
                start = max(len(combined.index) - len(new_rows), 0)
                for row_idx in range(start, len(combined.index)):
                    self.table_view.selectRow(row_idx)
        except Exception:
            pass
        return len(new_rows)

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:  # type: ignore[override]
        super().import_project_payload(payload)
        extra = self.data.extra if isinstance(self.data.extra, dict) else {}
        keys = extra.get("compare_keys")
        if isinstance(keys, list):
            self._row_keys = {str(key) for key in keys}
        else:
            frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
            self._rebuild_row_keys(frame)
        columns = extra.get("compare_columns")
        if isinstance(columns, list) and columns:
            self._compare_columns = [str(col) for col in columns]
        else:
            frame = self.data.table if isinstance(self.data.table, pd.DataFrame) else pd.DataFrame()
            self._compare_columns = [str(col) for col in frame.columns]
        fields = extra.get("compare_fields")
        if isinstance(fields, list):
            self._compare_fields = {str(field) for field in fields}
        else:
            self._compare_fields = set()
        field_order = extra.get("compare_field_order")
        if isinstance(field_order, list):
            self._compare_field_order = [str(field) for field in field_order]
        else:
            self._compare_field_order = []
        view_mode = extra.get("compare_view_mode")
        if isinstance(view_mode, str) and view_mode in {"matrix", "rows"}:
            self._compare_view_mode = view_mode
        else:
            self._compare_view_mode = "matrix"
        if hasattr(self, "view_mode_combo"):
            try:
                self.view_mode_combo.blockSignals(True)
                self.view_mode_combo.setCurrentIndex(0 if self._compare_view_mode == "matrix" else 1)
                self.view_mode_combo.blockSignals(False)
            except Exception:
                pass
        self._set_compare_view_mode(self._compare_view_mode)
        self._update_matrix_view()
        self._update_status()
        self._update_preview_graph_buttons()
        self._update_graph_preview_panel()

    def _selected_preview_row_index(self) -> Optional[int]:
        if not isinstance(self.table_view, QtWidgets.QTableView):
            return None
        selection = self.table_view.selectionModel()
        if selection is None:
            return None
        rows = selection.selectedRows()
        if not rows:
            return None
        return rows[0].row()

    def _selected_preview_row(self) -> Optional[pd.Series]:
        row_index = self._selected_preview_row_index()
        if row_index is None:
            return None
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return None
        if row_index >= len(frame.index):
            return None
        try:
            return frame.iloc[row_index]
        except Exception:
            return None

    def _update_preview_graph_buttons(self, *_: Any) -> None:
        key = self._active_compare_key()
        enabled = key is not None
        self.open_high_plot_button.setEnabled(enabled)
        self.open_low_plot_button.setEnabled(enabled)
        self.open_vsm_hysteresis_button.setEnabled(
            bool(enabled and self._ensure_vsm_hysteresis_groups().get(key or "", []))
        )
        self.open_vsm_temperature_button.setEnabled(
            bool(enabled and self._ensure_vsm_temperature_groups().get(key or "", []))
        )
        self.open_dma_button.setEnabled(
            bool(enabled and self._ensure_dma_isostress_groups().get(key or "", []))
        )
        self.open_shape_memory_button.setEnabled(
            bool(
                enabled
                and self._ensure_shape_memory_stress_strain_groups().get(key or "", [])
            )
        )
        self.open_fmr_button.setEnabled(
            bool(enabled and self._ensure_fmr_groups().get(key or "", []))
        )

    def _toggle_graph_preview_panel(self, checked: bool) -> None:
        if self._compare_view_mode != "rows":
            return
        if not hasattr(self, "graph_preview_panel"):
            return
        self.graph_preview_panel.setVisible(bool(checked))
        splitter = getattr(self, "preview_splitter", None)
        if isinstance(splitter, QtWidgets.QSplitter):
            if checked:
                sizes = splitter.sizes()
                if sizes and sizes[1] == 0:
                    total = sum(sizes) or 1
                    splitter.setSizes([int(total * 0.6), int(total * 0.4)])
            else:
                splitter.setSizes([1, 0])
        self._update_graph_preview_panel()

    def _update_graph_preview_panel(self, *_: Any) -> None:
        try:
            if self._compare_view_mode != "rows":
                return
            if not getattr(self, "graph_preview_panel", None):
                return
            if not self.graph_preview_panel.isVisible():
                return
            row = self._selected_preview_row()
            if row is None:
                self.high_preview_display.set_record(
                    None,
                    setpoint=None,
                    description="Select a row to preview the annealing measurement.",
                )
                self.low_preview_display.set_record(
                    None,
                    setpoint=None,
                    description="Select a row to preview the annealing measurement.",
                )
                self.vsm_hysteresis_gallery.clear(
                    "Select a row to preview VSM hysteresis graphs."
                )
                self.vsm_temperature_gallery.clear(
                    "Select a row to preview VSM temperature scans."
                )
                self.dma_iso_gallery.clear(
                    "Select a row to preview DMA iso-stress graphs."
                )
                self.shape_memory_gallery.clear(
                    "Select a row to preview shape-memory graphs."
                )
                self.fmr_gallery.clear("Select a row to preview FMR graphs.")
                return
            key = _row_to_microwire_key(row)
            if not key:
                message = "Select a microwire row with annealing data first."
                self.high_preview_display.set_record(
                    None, setpoint=None, description=message
                )
                self.low_preview_display.set_record(
                    None, setpoint=None, description=message
                )
                self.vsm_hysteresis_gallery.clear(message)
                self.vsm_temperature_gallery.clear(message)
                self.dma_iso_gallery.clear(message)
                self.shape_memory_gallery.clear(message)
                self.fmr_gallery.clear(message)
                return

            records = self._ensure_annealing_groups().get(key, [])
            if records:
                high_record, low_record = _select_high_low_pair(records)
                self.high_preview_display.set_record(
                    high_record,
                    setpoint=_extract_setpoint(high_record),
                    description="No 1000 mA measurement available for this microwire.",
                )
                self.low_preview_display.set_record(
                    low_record,
                    setpoint=_extract_setpoint(low_record),
                    description="No low mA measurement available for this microwire.",
                )
            else:
                message = "No annealing records found for this microwire."
                self.high_preview_display.set_record(
                    None, setpoint=None, description=message
                )
                self.low_preview_display.set_record(
                    None, setpoint=None, description=message
                )

            vsm_records = self._ensure_vsm_hysteresis_groups().get(key, [])
            vsm_items = _vsm_hysteresis_preview_items(
                vsm_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.vsm_hysteresis_gallery.set_items(
                vsm_items, "No VSM hysteresis graphs available for this microwire."
            )

            vsm_temp_records = self._ensure_vsm_temperature_groups().get(key, [])
            vsm_temp_items = _vsm_temperature_preview_items(
                vsm_temp_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.vsm_temperature_gallery.set_items(
                vsm_temp_items, "No VSM temperature scans available for this microwire."
            )

            dma_records = self._ensure_dma_isostress_groups().get(key, [])
            dma_items = _dma_iso_stress_preview_items(
                dma_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.dma_iso_gallery.set_items(
                dma_items, "No DMA iso-stress graphs available for this microwire."
            )

            shape_memory_records = self._ensure_shape_memory_stress_strain_groups().get(key, [])
            shape_memory_items = _shape_memory_stress_strain_preview_items(
                shape_memory_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.shape_memory_gallery.set_items(
                shape_memory_items,
                "No shape-memory graphs available for this microwire.",
            )

            fmr_records = self._ensure_fmr_groups().get(key, [])
            fmr_items = _fmr_preview_items(
                fmr_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.fmr_gallery.set_items(
                fmr_items, "No FMR graphs available for this microwire."
            )
        except Exception as exc:
            self.logger.exception("Failed to update compare graph preview panel")
            self.log(
                f"Failed to update compare graph preview panel: {exc}",
                level=logging.ERROR,
            )

    def _open_preview_graph(self, kind: str) -> None:
        key = self._active_compare_key()
        if not key:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "Select a microwire row with annealing data first.",
            )
            return
        if kind in {"high", "low"}:
            records = self._ensure_annealing_groups().get(key, [])
            if not records:
                QtWidgets.QMessageBox.information(
                    self,
                    "Microwire Data Builder",
                    "No annealing records found for the selected microwire.",
                )
                return
            high_record, low_record = _select_high_low_pair(records)
            record = high_record if kind == "high" else low_record
            label = "1000 mA" if kind == "high" else "low mA"
            if record is None:
                QtWidgets.QMessageBox.information(
                    self,
                    "Microwire Data Builder",
                    f"No {label} measurement available for this microwire.",
                )
                return
            self._show_annealing_record(record, label)
            return

        if kind == "vsm_hysteresis":
            records = self._ensure_vsm_hysteresis_groups().get(key, [])
            title = "VSM hysteresis graphs"
            items = _vsm_hysteresis_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No VSM hysteresis graphs available."
        elif kind == "vsm_temperature":
            records = self._ensure_vsm_temperature_groups().get(key, [])
            title = "VSM temperature scan graphs"
            items = _vsm_temperature_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No VSM temperature scan graphs available."
        elif kind == "dma_iso_stress":
            records = self._ensure_dma_isostress_groups().get(key, [])
            title = "DMA iso-stress graphs"
            items = _dma_iso_stress_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No DMA iso-stress graphs available."
        elif kind == "shape_memory_stress_strain":
            records = self._ensure_shape_memory_stress_strain_groups().get(key, [])
            title = "Shape memory stress/strain graphs"
            items = _shape_memory_stress_strain_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No shape-memory graphs available."
        else:
            records = self._ensure_fmr_groups().get(key, [])
            title = "FMR graphs"
            items = _fmr_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No FMR graphs available."
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                empty_message,
            )
            return
        dialog = _GraphGalleryDialog(
            title,
            items,
            parent=self,
            empty_message=empty_message,
        )
        dialog.exec()

    def _show_annealing_record(self, record: MeasurementRecord, label: str) -> None:
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"Annealing graph — {label}")
        dialog.resize(960, 640)
        layout = QtWidgets.QVBoxLayout(dialog)
        display = _AnnealingPlotDisplay(f"Graph — {label}", self.logger, dialog)
        display.set_record(
            record,
            setpoint=_extract_setpoint(record),
            description="No annealing measurement available for this microwire.",
        )
        layout.addWidget(display, 1)
        dialog.exec()

    def _ensure_annealing_groups(self) -> Dict[str, List[MeasurementRecord]]:
        groups = self._cached_annealing_groups
        if not groups:
            payload = self._load_payload("annealing", "annealing_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "annealing")
                self._cached_annealing_records = list(records)
                self._cached_annealing_groups = self._group_annealing_records(records)
                groups = self._cached_annealing_groups
        return groups

    def _filter_hidden_records(
        self, records: Sequence[object], section_key: str
    ) -> List[object]:
        hidden = _hidden_paths_from_section(self.sections.get(section_key))
        if not hidden:
            return list(records)
        filtered: List[object] = []
        for record in records:
            path_key = _record_path_key(record)
            if path_key and path_key in hidden:
                continue
            filtered.append(record)
        return filtered

    def _ensure_vsm_hysteresis_groups(self) -> Dict[str, List[VsmHysteresisRecord]]:
        groups = self._cached_vsm_hysteresis_groups
        if not groups:
            payload = self._load_payload("vsm_hysteresis", "vsm_hysteresis_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "vsm_hysteresis")
                self._cached_vsm_hysteresis_records = list(records)
                self._cached_vsm_hysteresis_groups = _group_graph_records_by_key(records)
                groups = self._cached_vsm_hysteresis_groups
        return groups

    def _ensure_vsm_temperature_groups(self) -> Dict[str, List[VsmTemperatureScanRecord]]:
        groups = self._cached_vsm_temperature_groups
        if not groups:
            payload = self._load_payload("vsm_temperature_scan", "vsm_temperature_scan_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "vsm_temperature_scan")
                self._cached_vsm_temperature_records = list(records)
                self._cached_vsm_temperature_groups = _group_graph_records_by_key(records)
                groups = self._cached_vsm_temperature_groups
        return groups

    def _ensure_dma_isostress_groups(self) -> Dict[str, List[DmaIsoStressRecord]]:
        groups = self._cached_dma_isostress_groups
        if not groups:
            payload = self._load_payload("dma_iso_stress", "dma_iso_stress_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "dma_iso_stress")
                self._cached_dma_isostress_records = list(records)
                self._cached_dma_isostress_groups = _group_graph_records_by_key(records)
                groups = self._cached_dma_isostress_groups
        return groups

    def _ensure_shape_memory_stress_strain_groups(
        self,
    ) -> Dict[str, List[ShapeMemoryStressStrainRecord]]:
        groups = self._cached_shape_memory_stress_strain_groups
        if not groups:
            payload = self._load_payload(
                "shape_memory_stress_strain",
                "shape_memory_stress_strain_records",
            )
            if isinstance(payload, list):
                records = self._filter_hidden_records(
                    payload,
                    "shape_memory_stress_strain",
                )
                self._cached_shape_memory_stress_strain_records = list(records)
                self._cached_shape_memory_stress_strain_groups = (
                    _group_graph_records_by_key(records)
                )
                groups = self._cached_shape_memory_stress_strain_groups
        return groups

    def _ensure_fmr_groups(self) -> Dict[str, List[FmrRecord]]:
        groups = self._cached_fmr_groups
        if not groups:
            payload = self._load_payload("fmr", "fmr_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "fmr")
                self._cached_fmr_records = list(records)
                self._cached_fmr_groups = _group_graph_records_by_key(records)
                groups = self._cached_fmr_groups
            if not groups:
                section = self.sections.get("fmr")
                if isinstance(section, FmrSection):
                    fallback = getattr(section, "_record_groups_by_key", None)
                    if isinstance(fallback, dict) and fallback:
                        self._cached_fmr_records = list(
                            getattr(section, "_all_records", []) or []
                        )
                        self._cached_fmr_groups = dict(fallback)
                        groups = self._cached_fmr_groups
        return groups

    def _group_annealing_records(
        self,
        records: Sequence[MeasurementRecord],
    ) -> Dict[str, List[MeasurementRecord]]:
        grouped: Dict[str, List[MeasurementRecord]] = {}
        for record in records:
            metadata = getattr(record, "metadata", None)
            if metadata is None:
                continue
            composition = getattr(metadata, "composition_token", None)
            draw = getattr(metadata, "draw_x", None)
            piece = getattr(metadata, "piece_y", None)
            if composition is None or draw is None or piece is None:
                continue
            suffix = None
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                parsed_key = _microscope_key(path)
                if parsed_key is not None:
                    _, _, _, suffix = parsed_key
            try:
                key = _microwire_key_to_str((str(composition), int(draw), int(piece), suffix))
            except (TypeError, ValueError):
                continue
            grouped.setdefault(key, []).append(record)
        return grouped

    def _remove_selected(self) -> None:
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.information(
                self,
                self.section_title,
                "Select one or more rows to remove.",
            )
            return
        frame = self.model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return
        drop_index = [frame.index[idx] for idx in rows if idx < len(frame.index)]
        updated = frame.drop(index=drop_index).reset_index(drop=True)
        self._rebuild_row_keys(updated)
        self._compare_columns = [str(col) for col in updated.columns]
        self._apply_compare_frame(updated)
        self._update_preview_graph_buttons()
        self._update_graph_preview_panel()

    def _clear_compare(self) -> None:
        self._row_keys = set()
        self._compare_columns = []
        self._apply_compare_frame(pd.DataFrame())
        self._update_preview_graph_buttons()
        self._update_graph_preview_panel()


class AssemblySection(QtWidgets.QWidget):
    """Final step that merges prepared mini-databases into a spreadsheet."""

    data_updated = QtCore.pyqtSignal()

    def __init__(
        self,
        sections: Dict[str, MiniDatabaseSection],
        logger: logging.Logger,
        log_callback: Callable[[int, str], None],
        parent: QtWidgets.QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.sections = sections
        self.logger = logger
        self._log_callback = log_callback
        self._cached_annealing_records: List[MeasurementRecord] = []
        self._cached_annealing_groups: Dict[str, List[MeasurementRecord]] = {}
        self._cached_vsm_hysteresis_records: List[VsmHysteresisRecord] = []
        self._cached_vsm_hysteresis_groups: Dict[str, List[VsmHysteresisRecord]] = {}
        self._cached_vsm_temperature_records: List[VsmTemperatureScanRecord] = []
        self._cached_vsm_temperature_groups: Dict[str, List[VsmTemperatureScanRecord]] = {}
        self._cached_dma_isostress_records: List[DmaIsoStressRecord] = []
        self._cached_dma_isostress_groups: Dict[str, List[DmaIsoStressRecord]] = {}
        self._cached_shape_memory_stress_strain_records: List[ShapeMemoryStressStrainRecord] = []
        self._cached_shape_memory_stress_strain_groups: Dict[str, List[ShapeMemoryStressStrainRecord]] = {}
        self._cached_fmr_records: List[FmrRecord] = []
        self._cached_fmr_groups: Dict[str, List[FmrRecord]] = {}
        self._compare_section: Optional["CompareSection"] = None
        self._raw_preview_frame: Optional[pd.DataFrame] = None
        self._measured_preview_frame: Optional[pd.DataFrame] = None
        self._preview_row_index_map: List[int] = []
        self._selected_columns: Optional[Set[str]] = None
        self._column_order: List[str] = []
        self._sort_spec: List[Tuple[str, bool]] = []
        self._mandatory_columns: Set[str] = {"Composition", "Microwire"}
        self._known_columns: Set[str] = set()
        self._graph_columns: Set[str] = set(FIGURE_COLUMNS) | set(ORIGIN_FIGURE_COLUMNS) | {
            VSM_HYSTERESIS_COLUMN,
            VSM_TEMPERATURE_SCAN_COLUMN,
            DMA_ISOSTRESS_COLUMN,
            SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
            FMR_COLUMN,
        }
        self._inline_graph_columns: Set[str] = set(FIGURE_COLUMNS) | {
            VSM_HYSTERESIS_COLUMN,
            VSM_TEMPERATURE_SCAN_COLUMN,
            DMA_ISOSTRESS_COLUMN,
            SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
            FMR_COLUMN,
        }
        self._graph_pixmap_cache: Dict[Tuple[object, ...], QtGui.QPixmap] = {}
        self._preview_thread: QtCore.QThread | None = None
        self._preview_worker: PreviewWorker | None = None
        self._preview_dialog: QtWidgets.QProgressDialog | None = None
        self._combine_thread: QtCore.QThread | None = None
        self._combine_worker: CombineWorker | None = None
        self._combine_dialog: QtWidgets.QProgressDialog | None = None
        self._combine_output_dir: Optional[Path] = None
        self._combine_output_name: str = ""
        self._imported_rows: Dict[str, Dict[str, Any]] = {}
        self._imported_sources: List[str] = []
        self._show_imported = True
        self._preview_search_text: str = ""

        self._output_dir = str(Path.cwd())
        self._output_name = DEFAULT_OUTPUT_NAME
        self._export_csv = True
        self._export_excel = False
        self._export_matplotlib = False
        self._export_origin = False
        self._export_html = False
        self._section_choices = [
            ("fabrication", "Fabrication"),
            ("annealing", "Current annealing"),
            ("microscope", "Microscope"),
            ("current_density", "Current density"),
            ("videos", "Videos"),
            ("vsm_hysteresis", "VSM hysteresis"),
            ("vsm_temperature_scan", "VSM temperature scan"),
            ("transition_temps", "Transition temps"),
            ("dma_iso_stress", "DMA iso-stress"),
            ("shape_memory_stress_strain", "Shape memory stress/strain"),
            ("fmr", "FMR"),
            ("strain", "Strain"),
        ]
        self._section_states = {key: True for key, _ in self._section_choices}

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        settings_row = QtWidgets.QHBoxLayout()
        self.export_button = QtWidgets.QPushButton("Export...")
        self.export_button.clicked.connect(self._open_export_dialog)
        settings_row.addWidget(self.export_button)
        self.export_summary_label = QtWidgets.QLabel("")
        self.export_summary_label.setWordWrap(True)
        settings_row.addWidget(self.export_summary_label, 1)
        settings_row.addStretch(1)
        layout.addLayout(settings_row)
        self._update_export_summary()

        self.status_label = QtWidgets.QLabel(
            "Ready to assemble once all sections are processed."
        )
        layout.addWidget(self.status_label)

        graph_row = QtWidgets.QHBoxLayout()
        self.graph_panel_checkbox = QtWidgets.QCheckBox("Show graph preview panel")
        self.graph_panel_checkbox.setChecked(False)
        self.graph_panel_checkbox.toggled.connect(self._toggle_graph_preview_panel)
        graph_row.addWidget(self.graph_panel_checkbox)
        self.open_high_plot_button = QtWidgets.QPushButton("Open 1000 mA graph")
        self.open_high_plot_button.clicked.connect(lambda: self._open_preview_graph("high"))
        self.open_high_plot_button.setEnabled(False)
        graph_row.addWidget(self.open_high_plot_button)
        self.open_low_plot_button = QtWidgets.QPushButton("Open low mA graph")
        self.open_low_plot_button.clicked.connect(lambda: self._open_preview_graph("low"))
        self.open_low_plot_button.setEnabled(False)
        graph_row.addWidget(self.open_low_plot_button)
        self.open_vsm_hysteresis_button = QtWidgets.QPushButton("Open VSM hyst graphs")
        self.open_vsm_hysteresis_button.clicked.connect(
            lambda: self._open_preview_graph("vsm_hysteresis")
        )
        self.open_vsm_hysteresis_button.setEnabled(False)
        graph_row.addWidget(self.open_vsm_hysteresis_button)
        self.open_vsm_temperature_button = QtWidgets.QPushButton("Open VSM temp graphs")
        self.open_vsm_temperature_button.clicked.connect(
            lambda: self._open_preview_graph("vsm_temperature")
        )
        self.open_vsm_temperature_button.setEnabled(False)
        graph_row.addWidget(self.open_vsm_temperature_button)
        self.open_dma_button = QtWidgets.QPushButton("Open DMA iso-stress graphs")
        self.open_dma_button.clicked.connect(lambda: self._open_preview_graph("dma_iso_stress"))
        self.open_dma_button.setEnabled(False)
        graph_row.addWidget(self.open_dma_button)
        self.open_shape_memory_button = QtWidgets.QPushButton("Open shape-memory graphs")
        self.open_shape_memory_button.clicked.connect(
            lambda: self._open_preview_graph("shape_memory_stress_strain")
        )
        self.open_shape_memory_button.setEnabled(False)
        graph_row.addWidget(self.open_shape_memory_button)
        self.open_fmr_button = QtWidgets.QPushButton("Open FMR graphs")
        self.open_fmr_button.clicked.connect(lambda: self._open_preview_graph("fmr"))
        self.open_fmr_button.setEnabled(False)
        graph_row.addWidget(self.open_fmr_button)
        graph_row.addStretch(1)
        layout.addLayout(graph_row)

        tools_row = QtWidgets.QHBoxLayout()
        self.columns_button = QtWidgets.QPushButton("Columns...")
        self.columns_button.clicked.connect(self._open_column_selector)
        tools_row.addWidget(self.columns_button)
        self.order_button = QtWidgets.QPushButton("Order...")
        self.order_button.clicked.connect(self._open_column_order_dialog)
        tools_row.addWidget(self.order_button)
        self.sort_button = QtWidgets.QPushButton("Sort...")
        self.sort_button.clicked.connect(self._open_sort_dialog)
        tools_row.addWidget(self.sort_button)
        self.clear_sort_button = QtWidgets.QPushButton("Clear sort")
        self.clear_sort_button.clicked.connect(self._clear_sort)
        self.clear_sort_button.setEnabled(False)
        tools_row.addWidget(self.clear_sort_button)
        self.reset_order_button = QtWidgets.QPushButton("Reset order")
        self.reset_order_button.clicked.connect(self._reset_column_order)
        tools_row.addWidget(self.reset_order_button)
        self.add_to_compare_button = QtWidgets.QPushButton("Add to compare")
        self.add_to_compare_button.clicked.connect(self._add_selected_to_compare)
        self.add_to_compare_button.setEnabled(False)
        tools_row.addWidget(self.add_to_compare_button)
        tools_row.addStretch(1)
        layout.addLayout(tools_row)

        search_row = QtWidgets.QHBoxLayout()
        search_label = QtWidgets.QLabel("Search:")
        search_row.addWidget(search_label)
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.setPlaceholderText(
            "Filter rows across visible columns"
        )
        self.search_edit.textChanged.connect(self._handle_preview_search_changed)
        search_row.addWidget(self.search_edit, 1)
        self.search_clear_button = QtWidgets.QPushButton("Clear")
        self.search_clear_button.setEnabled(False)
        self.search_clear_button.clicked.connect(self._clear_preview_search)
        search_row.addWidget(self.search_clear_button)
        layout.addLayout(search_row)

        self.preview_model = DataFrameModel()
        self.preview_model.set_decoration_provider(self._preview_decoration)
        self.preview_table = QtWidgets.QTableView()
        self.preview_table.setModel(self.preview_model)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.preview_table.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        header = self.preview_table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
            header.setSectionsMovable(True)
            header.setSectionsClickable(True)
        self.preview_table.setVerticalScrollMode(
            QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.preview_table.setHorizontalScrollMode(
            QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        preview_bar = self.preview_table.verticalScrollBar()
        if preview_bar is not None:
            preview_bar.setSingleStep(MiniDatabaseSection._SCROLL_SINGLE_STEP)
        header = self.preview_table.horizontalHeader()
        if header is not None:
            try:
                header.sectionMoved.connect(self._handle_preview_column_moved)
            except Exception:
                pass
        self.graph_preview_panel = QtWidgets.QWidget()
        graph_layout = QtWidgets.QVBoxLayout(self.graph_preview_panel)
        graph_layout.setContentsMargins(0, 0, 0, 0)
        graph_layout.setSpacing(6)
        self.graph_preview_tabs = QtWidgets.QTabWidget(self.graph_preview_panel)
        graph_layout.addWidget(self.graph_preview_tabs, 1)

        annealing_tab = QtWidgets.QWidget(self.graph_preview_panel)
        annealing_layout = QtWidgets.QVBoxLayout(annealing_tab)
        annealing_layout.setContentsMargins(0, 0, 0, 0)
        annealing_layout.setSpacing(6)
        self.high_preview_display = _AnnealingPlotDisplay(
            "Graph — 1000 mA", self.logger, annealing_tab
        )
        annealing_layout.addWidget(self.high_preview_display, 1)
        self.low_preview_display = _AnnealingPlotDisplay(
            "Graph — low mA", self.logger, annealing_tab
        )
        annealing_layout.addWidget(self.low_preview_display, 1)
        self.graph_preview_tabs.addTab(annealing_tab, "Annealing")

        self.vsm_hysteresis_gallery = _GraphGalleryWidget(
            "Select a row to preview VSM hysteresis graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.vsm_hysteresis_gallery, "VSM hyst")
        self.vsm_temperature_gallery = _GraphGalleryWidget(
            "Select a row to preview VSM temperature scans.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.vsm_temperature_gallery, "VSM temp")
        self.dma_iso_gallery = _GraphGalleryWidget(
            "Select a row to preview DMA iso-stress graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.dma_iso_gallery, "DMA iso-stress")
        self.shape_memory_gallery = _GraphGalleryWidget(
            "Select a row to preview shape-memory graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.shape_memory_gallery, "Shape memory")
        self.fmr_gallery = _GraphGalleryWidget(
            "Select a row to preview FMR graphs.",
            self.graph_preview_panel,
        )
        self.graph_preview_tabs.addTab(self.fmr_gallery, "FMR")
        self.graph_preview_panel.setVisible(False)

        self.preview_splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal)
        self.preview_splitter.addWidget(self.preview_table)
        self.preview_splitter.addWidget(self.graph_preview_panel)
        self.preview_splitter.setStretchFactor(0, 3)
        self.preview_splitter.setStretchFactor(1, 2)
        layout.addWidget(self.preview_splitter, 1)
        self.preview_table.show()
        selection_model = self.preview_table.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(self._update_preview_graph_buttons)
            selection_model.selectionChanged.connect(self._update_graph_preview_panel)

        button_row = QtWidgets.QHBoxLayout()
        button_row.addStretch(1)
        self.export_preview_button = QtWidgets.QPushButton("Export worksheet...")
        self.export_preview_button.clicked.connect(self._export_preview_worksheet)
        self.export_preview_button.setEnabled(False)
        button_row.addWidget(self.export_preview_button)
        self.preview_button = QtWidgets.QPushButton("Preview database")
        self.preview_button.clicked.connect(self._preview)
        button_row.addWidget(self.preview_button)
        layout.addLayout(button_row)

    def log(self, message: str, level: int = logging.INFO) -> None:
        try:
            self._log_callback(level, message)
        except Exception:
            self.logger.log(level, message)

    def attach_compare_section(self, compare_section: "CompareSection") -> None:
        self._compare_section = compare_section
        self._update_preview_graph_buttons()

    def has_project_data(self) -> bool:
        frame = self._raw_preview_frame
        return isinstance(frame, pd.DataFrame) and not frame.empty

    def export_project_payload(self) -> Dict[str, Any]:
        frame = self._raw_preview_frame if isinstance(self._raw_preview_frame, pd.DataFrame) else pd.DataFrame()
        columns = [str(col) for col in getattr(frame, "columns", [])]
        rows: List[Dict[str, Any]] = []
        if not frame.empty:
            for record in frame.to_dict(orient="records"):
                payload: Dict[str, Any] = {}
                for column in columns:
                    payload[column] = _json_safe(record.get(column))
                rows.append(payload)
        index_payload: List[Any] = []
        if not frame.empty:
            for entry in frame.index.tolist():
                index_payload.append(_json_safe(entry))
        return {
            "section": "assemble",
            "title": "Assemble",
            "columns": columns,
            "rows": rows,
            "index": index_payload,
            "selected_columns": list(self._selected_columns or []),
            "column_order": list(self._column_order),
            "sort_spec": list(self._sort_spec),
            "search_query": self._preview_search_text,
            "export_settings": self._export_settings_payload(),
            "graph_preview": bool(self.graph_panel_checkbox.isChecked()),
            "imported_rows": [
                {str(key): _json_safe(val) for key, val in row.items()}
                for row in self._imported_rows.values()
            ],
            "imported_sources": list(self._imported_sources),
            "show_imported": bool(self._show_imported),
        }

    def import_project_payload(self, payload: Mapping[str, Any]) -> None:
        if not isinstance(payload, Mapping):
            return
        self._apply_export_settings(payload.get("export_settings", {}))
        graph_preview = payload.get("graph_preview")
        if isinstance(graph_preview, bool):
            try:
                self.graph_panel_checkbox.setChecked(graph_preview)
            except Exception:
                pass
        selected_columns = payload.get("selected_columns")
        if isinstance(selected_columns, (list, tuple)):
            mapped = []
            for col in selected_columns:
                if not col:
                    continue
                text = str(col)
                if text == "Temperature (°C)":
                    text = CORE_TEMPERATURE_COLUMN
                mapped.append(text)
            self._selected_columns = set(mapped)
        else:
            self._selected_columns = None
        column_order = payload.get("column_order")
        if isinstance(column_order, (list, tuple)):
            mapped = []
            for col in column_order:
                if not col:
                    continue
                text = str(col)
                if text == "Temperature (°C)":
                    text = CORE_TEMPERATURE_COLUMN
                mapped.append(text)
            self._column_order = mapped
        else:
            self._column_order = []
        sort_spec = payload.get("sort_spec")
        restored_sort: List[Tuple[str, bool]] = []
        if isinstance(sort_spec, (list, tuple)):
            for entry in sort_spec:
                if not entry:
                    continue
                if isinstance(entry, (list, tuple)) and entry:
                    column = entry[0]
                    ascending = entry[1] if len(entry) > 1 else True
                elif isinstance(entry, dict):
                    column = entry.get("column")
                    ascending = entry.get("ascending", True)
                else:
                    continue
                if isinstance(column, str):
                    if column == "Temperature (°C)":
                        column = CORE_TEMPERATURE_COLUMN
                    restored_sort.append((column, bool(ascending)))
        self._sort_spec = restored_sort
        search_query = payload.get("search_query")
        if isinstance(search_query, str):
            self._preview_search_text = self._normalise_search_text(search_query)
        else:
            self._preview_search_text = ""
        if hasattr(self, "search_edit"):
            self.search_edit.blockSignals(True)
            self.search_edit.setText(self._preview_search_text)
            self.search_edit.blockSignals(False)
        if hasattr(self, "search_clear_button"):
            self.search_clear_button.setEnabled(bool(self._preview_search_text))

        columns_payload = payload.get("columns")
        if isinstance(columns_payload, (list, tuple)):
            column_names = []
            for column in columns_payload:
                text = str(column)
                if text == "Temperature (°C)":
                    text = CORE_TEMPERATURE_COLUMN
                column_names.append(text)
        else:
            column_names = []
        rows_payload = payload.get("rows")
        if isinstance(rows_payload, (list, tuple)):
            frame = pd.DataFrame(list(rows_payload), columns=column_names or None)
        else:
            frame = pd.DataFrame(columns=column_names)
        index_payload = payload.get("index")
        if isinstance(index_payload, list) and len(index_payload) == len(frame.index):
            try:
                frame.index = pd.Index(index_payload)
            except Exception:
                pass
        if isinstance(frame, pd.DataFrame):
            frame = self._apply_column_universe(frame)
        if self._selected_columns is not None:
            self._sync_section_states_from_columns(self._selected_columns, frame.columns)
        imported_rows = payload.get("imported_rows")
        if isinstance(imported_rows, list):
            cleaned_rows: Dict[str, Dict[str, Any]] = {}
            for row in imported_rows:
                if isinstance(row, Mapping):
                    record = {str(k): v for k, v in row.items()}
                    key = _row_to_microwire_key(pd.Series(record))
                    if key:
                        cleaned_rows[key] = record
            self._imported_rows = cleaned_rows
        imported_sources = payload.get("imported_sources")
        if isinstance(imported_sources, list):
            self._imported_sources = [str(entry) for entry in imported_sources if entry]
        show_imported = payload.get("show_imported")
        if isinstance(show_imported, bool):
            self._show_imported = show_imported
        self._measured_preview_frame = None
        self._raw_preview_frame = frame
        self._refresh_preview_frame()

    def _open_export_dialog(self) -> None:
        dialog = _AssemblyExportDialog(
            output_dir=self._output_dir,
            output_name=self._output_name,
            export_csv=self._export_csv,
            export_excel=self._export_excel,
            export_html=self._export_html,
            export_matplotlib=self._export_matplotlib,
            export_origin=self._export_origin,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._apply_export_settings(dialog.export_settings())
            self._combine()

    def _export_settings_payload(self) -> Dict[str, Any]:
        return {
            "output_dir": self._output_dir,
            "output_name": self._output_name,
            "export_csv": self._export_csv,
            "export_excel": self._export_excel,
            "export_html": self._export_html,
            "export_matplotlib": self._export_matplotlib,
            "export_origin": self._export_origin,
            "sections": dict(self._section_states),
        }

    def _apply_export_settings(self, settings: Mapping[str, Any]) -> None:
        if not isinstance(settings, Mapping):
            return
        output_dir = settings.get("output_dir")
        if isinstance(output_dir, str) and output_dir.strip():
            self._output_dir = output_dir.strip()
        output_name = settings.get("output_name")
        if isinstance(output_name, str) and output_name.strip():
            self._output_name = output_name.strip()
        self._export_csv = bool(settings.get("export_csv", self._export_csv))
        self._export_excel = bool(settings.get("export_excel", self._export_excel))
        self._export_html = bool(settings.get("export_html", self._export_html))
        self._export_matplotlib = bool(settings.get("export_matplotlib", self._export_matplotlib))
        self._export_origin = bool(settings.get("export_origin", self._export_origin))
        sections = settings.get("sections")
        if isinstance(sections, Mapping):
            for key, _label in self._section_choices:
                if key in sections:
                    self._section_states[key] = bool(sections.get(key))
        self._update_export_summary()

    def _open_import_dialog(self) -> None:
        if openpyxl is None:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Importing workbooks requires openpyxl.",
            )
            return
        start_dir = _dialog_start_directory("sample_data")
        path_text, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Import data workbook",
            str(start_dir),
            "Excel files (*.xlsx *.xlsm *.xltx *.xltm);;All files (*.*)",
        )
        if not path_text:
            return
        path = Path(path_text)
        fabrication_index = self._load_payload("fabrication", "fabrication_index")
        if not isinstance(fabrication_index, FabricationIndex):
            fabrication_index = FabricationIndex()
        new_rows = self._load_import_rows(path, fabrication_index)
        if not new_rows:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "No usable rows were found in that workbook.",
            )
            return
        stats = self._merge_imported_payload(new_rows)
        if not stats["new_samples"] and not stats["updated_samples"] and not self._imported_rows:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "No new rows were added from that workbook.",
            )
        if str(path) not in self._imported_sources:
            self._imported_sources.append(str(path))
        base_frame = self._measured_preview_frame or self._raw_preview_frame
        if isinstance(base_frame, pd.DataFrame) and not base_frame.empty:
            merged = self._merge_imported_rows(base_frame)
            self._update_preview(merged)
        else:
            self._update_preview(pd.DataFrame(list(self._imported_rows.values())))
        self._sync_imports_to_fabrication()
        summary_lines = [
            f"Imported workbook: {path.name}",
            f"New samples added: {stats['new_samples']}",
            f"Existing samples updated: {stats['updated_samples']}",
        ]
        if stats["added_fields"]:
            summary_lines.append(f"Fields filled: {stats['added_fields']}")
        if stats["new_labels"]:
            summary_lines.append("New samples: " + ", ".join(stats["new_labels"][:8]))
        if stats["updated_labels"]:
            summary_lines.append("Updated samples: " + ", ".join(stats["updated_labels"][:8]))
        QtWidgets.QMessageBox.information(
            self,
            "Microwire Data Builder",
            "\n".join(summary_lines),
        )
        self.log(f"Imported {len(new_rows)} row(s) from {path.name}.")
        self._mark_dirty()

    def open_import_dialog(self) -> None:
        self._open_import_dialog()

    def _load_import_rows(
        self, path: Path, fabrication_index: FabricationIndex
    ) -> Dict[str, Dict[str, Any]]:
        if openpyxl is None:
            return {}
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            self.logger.exception("Failed to open import workbook %s", path)
            return {}
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return {}
        mapped_headers: List[Optional[str]] = []
        seen: Dict[str, int] = {}
        for header in header_row:
            column = _map_import_header(header)
            if column is None:
                mapped_headers.append(None)
                continue
            count = seen.get(column, 0) + 1
            seen[column] = count
            if count > 1:
                column = f"{column} ({count})"
            mapped_headers.append(column)
        rows: Dict[str, Dict[str, Any]] = {}
        source_label = f"Imported ({path.name})"
        for row in rows_iter:
            record: Dict[str, Any] = {}
            for idx, value in enumerate(row):
                if idx >= len(mapped_headers):
                    break
                column = mapped_headers[idx]
                if not column:
                    continue
                record[column] = _normalise_import_value(value)
            if "Temperature (°C)" in record and CORE_TEMPERATURE_COLUMN not in record:
                record[CORE_TEMPERATURE_COLUMN] = record.pop("Temperature (°C)")
            composition = str(record.get("Composition") or "").strip()
            microwire = str(record.get("Microwire") or "").strip()
            if not composition or not microwire:
                continue
            normalised_wire = self._normalise_import_microwire(microwire)
            record["Composition"] = composition
            record["Microwire"] = normalised_wire or microwire
            import_key = _row_to_microwire_key(pd.Series(record))
            if not import_key:
                continue
            if not record.get("e/a"):
                ea_value = _compute_ea_from_composition(composition)
                if ea_value is not None:
                    record["e/a"] = ea_value
            record.setdefault("Data source", source_label)
            self._apply_fabrication_defaults(record, fabrication_index)
            rows[import_key] = record
        return rows

    def _normalise_import_microwire(self, microwire: str) -> Optional[str]:
        parts = _microwire_parts_from_label_safe(str(microwire))
        if parts is None:
            return None
        draw, piece, suffix = parts
        try:
            return _microwire_label(int(draw), int(piece), suffix)
        except Exception:
            return None

    @staticmethod
    def _is_missing_import_value(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        try:
            return bool(pd.isna(value))
        except Exception:
            return False

    @classmethod
    def _should_fill_import_value(cls, existing: Any, incoming: Any) -> bool:
        return cls._is_missing_import_value(existing) and not cls._is_missing_import_value(incoming)

    def _apply_fabrication_defaults(
        self, record: Dict[str, Any], fabrication_index: FabricationIndex
    ) -> None:
        composition = str(record.get("Composition") or "").strip()
        microwire = str(record.get("Microwire") or "").strip()
        if not composition or not microwire:
            return
        parts = _microwire_parts_from_label_safe(microwire)
        if parts is None:
            return
        draw_x, piece_y, _suffix = parts
        draw_info = fabrication_index.get_draw(composition, draw_x)
        piece_info = fabrication_index.get_piece(composition, draw_x, piece_y)

        def fill(column: str, value: object) -> None:
            if self._should_fill_import_value(record.get(column), value):
                record[column] = value

        fill("Length (m)", _value_for_output(piece_info, "length_m"))
        fill("Production datetime", _value_for_output(draw_info, "production_datetime"))
        fill("Mass (g)", _value_for_output(draw_info, "mass_g"))
        piece_resistance = _value_for_output(piece_info, "fabrication_resistance_ohm")
        draw_resistance = _value_for_output(draw_info, "fabrication_resistance_ohm")
        fill("Resistance (Ω)", piece_resistance if piece_resistance is not None else draw_resistance)
        fill(CORE_TEMPERATURE_COLUMN, _value_for_output(draw_info, "fabrication_temperature_c"))
        fill(
            GLASS_TEMPERATURE_COLUMN,
            _value_for_output(draw_info, "fabrication_glass_temperature_c"),
        )
        fill("Winding speed (m/min)", _value_for_output(draw_info, "winding_speed_m_per_min"))
        fill("Glass feeding (mm/min)", _value_for_output(draw_info, "glass_feed_mm_per_min"))
        fill("Underpressure", _value_for_output(draw_info, "underpressure"))
        pull_value = _value_for_output(piece_info, "glass_pull_off")
        if pull_value is None:
            pull_value = _value_for_output(draw_info, "glass_pull_off")
        fill(GLASS_PULL_COLUMN, pull_value)
        notes = _compose_notes(draw_info, piece_info)
        fill("Notes", notes)
        ea_value = record.get("e/a")
        if ea_value in (None, ""):
            ea_value = _compute_ea_from_composition(composition)
        fill(ESTIMATED_TRANSITION_COLUMN, _estimate_transition_temp_c(ea_value))

    def _merge_imported_rows(self, frame: pd.DataFrame) -> pd.DataFrame:
        if not self._imported_rows or not self._show_imported:
            return frame
        imported = pd.DataFrame(list(self._imported_rows.values()))
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return imported
        merged_frame = frame.copy()
        key_index: Dict[str, int] = {}
        for idx, row in merged_frame.iterrows():
            key = _row_to_microwire_key(row)
            if key and key not in key_index:
                key_index[key] = idx
        append_rows: List[Dict[str, Any]] = []
        for key, record in self._imported_rows.items():
            if key in key_index:
                row_idx = key_index[key]
                row = merged_frame.loc[row_idx]
                updated = False
                for column, value in record.items():
                    if column in {"Composition", "Microwire"}:
                        continue
                    if column not in merged_frame.columns:
                        merged_frame[column] = None
                    existing = row.get(column)
                    if self._should_fill_import_value(existing, value):
                        merged_frame.at[row_idx, column] = value
                        updated = True
                if updated:
                    current_source = str(row.get("Data source") or "").strip()
                    if current_source and "Imported" not in current_source:
                        merged_frame.at[row_idx, "Data source"] = "Measured + Imported"
            else:
                append_rows.append(record)
        if append_rows:
            merged_frame = pd.concat(
                [merged_frame, pd.DataFrame(append_rows)], ignore_index=True, sort=False
            )
        return merged_frame

    def _merge_imported_payload(self, incoming: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        added_fields = 0
        new_samples = 0
        updated_samples = 0
        new_labels: List[str] = []
        updated_labels: List[str] = []
        for key, record in incoming.items():
            existing = self._imported_rows.get(key)
            if existing is None:
                self._imported_rows[key] = dict(record)
                new_samples += 1
                label = f"{record.get('Composition', '')} {record.get('Microwire', '')}".strip()
                if label:
                    new_labels.append(label)
                continue
            updated = False
            for column, value in record.items():
                if self._should_fill_import_value(existing.get(column), value):
                    existing[column] = value
                    added_fields += 1
                    updated = True
            if updated:
                updated_samples += 1
                label = f"{record.get('Composition', '')} {record.get('Microwire', '')}".strip()
                if label:
                    updated_labels.append(label)
        return {
            "new_samples": new_samples,
            "updated_samples": updated_samples,
            "added_fields": added_fields,
            "new_labels": new_labels,
            "updated_labels": updated_labels,
        }

    def set_show_imported(self, enabled: bool) -> None:
        self._show_imported = bool(enabled)
        base_frame = self._measured_preview_frame or self._raw_preview_frame
        if isinstance(base_frame, pd.DataFrame) and not base_frame.empty:
            merged = self._merge_imported_rows(base_frame)
            self._update_preview(merged)
        elif self._show_imported:
            self._update_preview(pd.DataFrame(list(self._imported_rows.values())))
        else:
            self._update_preview(pd.DataFrame())

    def clear_imported_data(self) -> None:
        if not self._imported_rows and not self._imported_sources:
            return
        self._imported_rows = {}
        self._imported_sources = []
        base_frame = self._measured_preview_frame or self._raw_preview_frame
        if isinstance(base_frame, pd.DataFrame):
            self._update_preview(base_frame)
        else:
            self._update_preview(pd.DataFrame())
        self._mark_dirty()

    def imported_sources(self) -> List[str]:
        return list(self._imported_sources)

    def _mark_dirty(self) -> None:
        try:
            self.data_updated.emit()
        except Exception:
            pass

    def _sync_imports_to_fabrication(self) -> None:
        fabrication = self.sections.get("fabrication")
        if not isinstance(fabrication, FabricationSection):
            return
        if not self._imported_rows:
            return
        try:
            added = fabrication.apply_imported_samples(self._imported_rows.values())
        except Exception:
            self.logger.exception("Failed to sync imported samples into fabrication section")
            return
        if added:
            self.log(f"Added {added} imported sample(s) to Fabrication.")

    def _update_export_summary(self) -> None:
        output_dir = self._output_dir or str(Path.cwd())
        output_name = self._output_name or DEFAULT_OUTPUT_NAME
        formats: List[str] = []
        if self._export_csv:
            formats.append("CSV")
        if self._export_excel:
            formats.append("Excel")
        if self._export_html:
            formats.append("HTML")
        if self._export_matplotlib:
            formats.append("Matplotlib")
        if self._export_origin:
            formats.append("Origin")
        if not formats:
            formats.append("No exports")
        selected_sections = [
            label for key, label in self._section_choices if self._section_states.get(key)
        ]
        section_summary = f"{len(selected_sections)}/{len(self._section_choices)} sections"
        self.export_summary_label.setText(
            f"{output_name} → {output_dir} | {', '.join(formats)} | {section_summary}"
        )

    def _choose_output_dir(self) -> None:
        start_dir = _dialog_start_directory(self._output_dir)
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select output directory",
            str(start_dir),
        )
        if directory:
            self._output_dir = directory
            self._update_export_summary()

    def _load_payload(self, section_key: str, name: str) -> Any:
        section = self.sections.get(section_key)
        if section is None:
            return None
        return section.store.load_payload(name)

    def _selected_sections(self) -> set[str]:
        return {key for key, enabled in self._section_states.items() if enabled}

    def _prepare_builder_inputs(
        self,
        selected: set[str],
        *,
        require_payloads: bool = True,
    ) -> Optional[
        tuple[
            FabricationIndex,
            List[MeasurementRecord],
            List[VsmHysteresisRecord],
            List[VsmTemperatureScanRecord],
            List[DmaIsoStressRecord],
            List[ShapeMemoryStressStrainRecord],
            Dict[str, Dict[str, Any]],
            List[FmrRecord],
            Dict[MicrowireKey, MicroscopeMeasurements],
            Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary],
            Dict[MicrowireKey, StrainRecord],
            Dict[str, Dict[str, Any]],
            Dict[str, Dict[str, Any]],
            Dict[str, Dict[str, float]],
            Dict[str, Dict[str, float]],
            Dict[str, Dict[str, float]],
            Dict[str, Dict[str, Any]],
        ]
    ]:
        if "annealing" not in selected:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Current annealing data must be included to assemble a database.",
            )
            return None

        missing: list[str] = []

        def _mark_missing(label: str) -> None:
            if require_payloads:
                missing.append(label)
            else:
                self.log(f"Skipping {label} because no data were found.", level=logging.WARNING)

        fabrication_index = FabricationIndex()
        if "fabrication" in selected:
            payload = self._load_payload("fabrication", "fabrication_index")
            if isinstance(payload, FabricationIndex):
                fabrication_index = payload
            else:
                _mark_missing("fabrication")

        annealing_records_payload = self._load_payload("annealing", "annealing_records")
        annealing_records: List[MeasurementRecord] = []
        if isinstance(annealing_records_payload, list):
            annealing_records = list(
                self._filter_hidden_records(annealing_records_payload, "annealing")
            )
        if not annealing_records:
            _mark_missing("annealing")
        self._cached_annealing_records = list(annealing_records)
        self._cached_annealing_groups = self._group_annealing_records(annealing_records)

        microscope_index: Dict[MicrowireKey, MicroscopeMeasurements] = {}
        overrides: Dict[str, Dict[str, float]] = {}
        if "microscope" in selected:
            payload = self._load_payload("microscope", "microscope_index")
            microscope_section = self.sections.get("microscope")
            if isinstance(microscope_section, MicroscopeSection):
                overrides = microscope_section.overrides
                table = microscope_section.data.table
            else:
                table = None
            if isinstance(payload, dict) and payload:
                microscope_index = payload
            elif isinstance(table, pd.DataFrame) and not table.empty:
                microscope_index = self._build_microscope_index_from_table(table)
            else:
                _mark_missing("microscope")

        video_index: Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary] = {}
        if "videos" in selected:
            payload = self._load_payload("videos", "video_index")
            if isinstance(payload, dict):
                video_index = payload
            else:
                if require_payloads:
                    self.log("Skipping videos because no data were found.", level=logging.WARNING)
                else:
                    _mark_missing("videos")
        video_overrides: Dict[str, Dict[str, Any]] = {}
        video_section = self.sections.get("videos")
        if isinstance(video_section, VideoSection):
            video_overrides = video_section.overrides_snapshot()

        strain_records: Dict[MicrowireKey, StrainRecord] = {}
        if "strain" in selected:
            payload = self._load_payload("strain", "strain_records")
            if isinstance(payload, dict):
                strain_records = payload
            if not strain_records:
                section = self.sections.get("strain")
                if isinstance(section, StrainSection):
                    strain_records = section.records_snapshot()
            if not strain_records:
                _mark_missing("strain")

        strain_entries: Dict[str, Dict[str, Any]] = {}
        if "strain" in selected:
            section = self.sections.get("strain")
            if isinstance(section, StrainSection):
                strain_entries = section.entries_snapshot()

        vsm_hysteresis_records: List[VsmHysteresisRecord] = []
        if "vsm_hysteresis" in selected:
            payload = self._load_payload("vsm_hysteresis", "vsm_hysteresis_records")
            if isinstance(payload, list):
                vsm_hysteresis_records = list(
                    self._filter_hidden_records(payload, "vsm_hysteresis")
                )
            else:
                _mark_missing("VSM hysteresis")
        self._cached_vsm_hysteresis_records = list(vsm_hysteresis_records)
        self._cached_vsm_hysteresis_groups = _group_graph_records_by_key(vsm_hysteresis_records)

        vsm_temperature_records: List[VsmTemperatureScanRecord] = []
        if "vsm_temperature_scan" in selected:
            payload = self._load_payload("vsm_temperature_scan", "vsm_temperature_scan_records")
            if isinstance(payload, list):
                vsm_temperature_records = list(
                    self._filter_hidden_records(payload, "vsm_temperature_scan")
                )
            else:
                _mark_missing("VSM temperature scan")
        self._cached_vsm_temperature_records = list(vsm_temperature_records)
        self._cached_vsm_temperature_groups = _group_graph_records_by_key(vsm_temperature_records)

        dma_isostress_records: List[DmaIsoStressRecord] = []
        if "dma_iso_stress" in selected:
            payload = self._load_payload("dma_iso_stress", "dma_iso_stress_records")
            if isinstance(payload, list):
                dma_isostress_records = list(
                    self._filter_hidden_records(payload, "dma_iso_stress")
                )
            else:
                _mark_missing("DMA iso-stress")
        self._cached_dma_isostress_records = list(dma_isostress_records)
        self._cached_dma_isostress_groups = _group_graph_records_by_key(dma_isostress_records)

        shape_memory_stress_strain_records: List[ShapeMemoryStressStrainRecord] = []
        if "shape_memory_stress_strain" in selected:
            payload = self._load_payload(
                "shape_memory_stress_strain",
                "shape_memory_stress_strain_records",
            )
            if isinstance(payload, list):
                shape_memory_stress_strain_records = list(
                    self._filter_hidden_records(payload, "shape_memory_stress_strain")
                )
            else:
                _mark_missing("Shape memory stress/strain")
        self._cached_shape_memory_stress_strain_records = list(
            shape_memory_stress_strain_records
        )
        self._cached_shape_memory_stress_strain_groups = _group_graph_records_by_key(
            shape_memory_stress_strain_records
        )

        shape_memory_entries: Dict[str, Dict[str, Any]] = {}
        if "shape_memory_stress_strain" in selected:
            section = self.sections.get("shape_memory_stress_strain")
            if isinstance(section, ShapeMemoryStressStrainSection):
                shape_memory_entries = section.entries_snapshot()

        fmr_records: List[FmrRecord] = []
        if "fmr" in selected:
            payload = self._load_payload("fmr", "fmr_records")
            if isinstance(payload, list):
                fmr_records = list(self._filter_hidden_records(payload, "fmr"))
            else:
                _mark_missing("FMR")
        self._cached_fmr_records = list(fmr_records)
        self._cached_fmr_groups = _group_graph_records_by_key(fmr_records)

        current_density_entries: Dict[str, Dict[str, Any]] = {}
        if "current_density" in selected:
            current_section = self.sections.get("current_density")
            if isinstance(current_section, CurrentDensitySection):
                snapshot_provider = getattr(
                    current_section, "current_density_snapshot", None
                )
                if not callable(snapshot_provider):
                    self.logger.warning(
                        "Current density snapshot method is missing; skipping current density values."
                    )
                else:
                    try:
                        current_density_entries = snapshot_provider()
                    except Exception:
                        self.logger.exception("Failed to snapshot current density")
                    if not current_density_entries:
                        try:
                            current_section.refresh_data()
                        except Exception:
                            self.logger.exception("Failed to refresh current density data")
                        try:
                            current_density_entries = snapshot_provider()
                        except Exception:
                            self.logger.exception(
                                "Failed to snapshot current density after refresh"
                            )
            if not current_density_entries:
                _mark_missing("current density")

        phase_points: Dict[str, Dict[str, float]] = {}
        if "current_density" in selected:
            annealing_section = self.sections.get("annealing")
            if isinstance(annealing_section, AnnealingSection):
                phase_points = dict(getattr(annealing_section, "_phase_points", {}))
        transition_points: Dict[str, Dict[str, float]] = {}
        if "transition_temps" in selected:
            transition_section = self.sections.get("transition_temps")
            if isinstance(transition_section, TransitionTempsSection):
                transition_points = transition_section.transition_points_snapshot()
        if missing and require_payloads:
            QtWidgets.QMessageBox.warning(
                self,
                "Microwire Data Builder",
                "Process the following sections first: " + ", ".join(sorted(missing)),
            )
            return None

        return (
            fabrication_index,
            annealing_records,
            vsm_hysteresis_records,
            vsm_temperature_records,
            dma_isostress_records,
            shape_memory_stress_strain_records,
            shape_memory_entries,
            fmr_records,
            microscope_index,
            video_index,
            strain_records,
            strain_entries,
            current_density_entries,
            overrides,
            phase_points,
            transition_points,
            video_overrides,
        )

    @staticmethod
    def _parse_microscope_key_from_row(row: pd.Series) -> Optional[MicrowireKey]:
        key_value = row.get("_key")
        if isinstance(key_value, str):
            parsed = _microwire_key_from_string(key_value)
            if parsed is not None:
                return parsed
        composition = str(row.get("Composition") or "").strip()
        microwire = str(row.get("Microwire") or "").strip()
        if not composition or not microwire:
            return None
        parsed = _microwire_parts_from_label_safe(microwire)
        if parsed is None:
            return None
        try:
            draw = int(parsed[0])
            piece = int(parsed[1])
            suffix = parsed[2]
        except (TypeError, ValueError):
            return None
        return (composition, draw, piece, suffix)

    @staticmethod
    def _parse_positive_float(value: object) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            numeric = float(value)
        else:
            text = str(value).strip().replace(",", ".")
            if not text:
                return None
            try:
                numeric = float(text)
            except ValueError:
                return None
        if not math.isfinite(numeric) or numeric <= 0:
            return None
        return numeric

    @staticmethod
    def _coerce_path(value: object) -> Optional[Path]:
        if value is None:
            return None
        if isinstance(value, Path):
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                return Path(text)
            except Exception:
                return None
        return None

    def _image_path_from_row(self, row: pd.Series, kind: str) -> Optional[Path]:
        candidates: List[Path] = []
        if kind == "core":
            image_column = MICROSCOPE_IMAGE_COLUMNS[0]
            mapped = "_core_image"
        else:
            image_column = MICROSCOPE_IMAGE_COLUMNS[1]
            mapped = "_glass_image"
        direct_path = self._coerce_path(row.get(image_column))
        if direct_path is not None:
            candidates.append(direct_path)
        mapped_path = self._coerce_path(row.get(mapped))
        if mapped_path is not None:
            candidates.append(mapped_path)
        images = row.get("_images")
        if isinstance(images, (list, tuple)):
            for entry in images:
                path = self._coerce_path(entry)
                if path is not None:
                    candidates.append(path)
                    break
        for candidate in candidates:
            if candidate is not None:
                return candidate
        return None

    def _build_microscope_index_from_table(
        self, table: pd.DataFrame
    ) -> Dict[MicrowireKey, MicroscopeMeasurements]:
        index: Dict[MicrowireKey, MicroscopeMeasurements] = {}
        for _, row in table.iterrows():
            key = self._parse_microscope_key_from_row(row)
            if key is None:
                continue
            measurements = index.setdefault(key, MicroscopeMeasurements())
            d_value = self._parse_positive_float(row.get(MICROSCOPE_D_COLUMN))
            if d_value is not None:
                d_path = self._image_path_from_row(row, "core")
                detection = MicroscopeDetection(
                    value=float(d_value),
                    image_path=d_path,
                    source="manual",
                )
                detection.category = "core"
                measurements.core.append(detection)
            else:
                d_path = self._image_path_from_row(row, "core")
                if d_path is not None:
                    measurements.add_placeholder("core", d_path)
            D_value = self._parse_positive_float(row.get(MICROSCOPE_CAP_D_COLUMN))
            if D_value is not None:
                D_path = self._image_path_from_row(row, "glass")
                detection = MicroscopeDetection(
                    value=float(D_value),
                    image_path=D_path,
                    source="manual",
                )
                detection.category = "glass"
                measurements.glass.append(detection)
            else:
                D_path = self._image_path_from_row(row, "glass")
                if D_path is not None:
                    measurements.add_placeholder("glass", D_path)
        return index

    def _update_preview(self, frame: pd.DataFrame) -> None:
        previous_order = self._column_order or self._current_preview_column_order()
        if previous_order:
            self._column_order = list(previous_order)
        self._graph_pixmap_cache.clear()
        if isinstance(frame, pd.DataFrame):
            frame = self._apply_column_universe(frame)
            self._raw_preview_frame = frame.copy()
        else:
            self._raw_preview_frame = pd.DataFrame()
        self._refresh_preview_frame()

    def _column_universe(self) -> List[str]:
        universe: List[str] = []

        def add(column: str) -> None:
            if column and column not in universe:
                universe.append(column)

        for column in OUTPUT_COLUMNS:
            add(str(column))
        for column in MICROSCOPE_IMAGE_COLUMNS:
            add(str(column))
        for column in FIGURE_COLUMNS:
            add(str(column))
        for column in ORIGIN_FIGURE_COLUMNS:
            add(str(column))
        for column in (
            VSM_HYSTERESIS_COLUMN,
            VSM_TEMPERATURE_SCAN_COLUMN,
            DMA_ISOSTRESS_COLUMN,
        ):
            add(str(column))
        for column in CURRENT_DENSITY_COLUMNS:
            add(str(column))
        for column in TRANSITION_TEMP_COLUMNS:
            add(str(column))
        for column in getattr(StrainSection, "TABLE_COLUMNS", ()):
            add(str(column))
        return universe

    def _apply_column_universe(self, frame: pd.DataFrame) -> pd.DataFrame:
        universe = self._column_universe()
        if not universe:
            return frame
        updated = frame.copy()
        for column in universe:
            if column not in updated.columns:
                updated[column] = None
        ordered = [column for column in universe if column in updated.columns]
        for column in updated.columns:
            if column not in ordered:
                ordered.append(column)
        return updated.loc[:, ordered]

    def _refresh_preview_frame(self) -> None:
        raw_frame = self._raw_preview_frame
        total_rows = 0
        if not isinstance(raw_frame, pd.DataFrame) or raw_frame.empty:
            display_frame = pd.DataFrame()
            self._preview_row_index_map = []
        else:
            sorted_frame, row_map = self._apply_sort_spec(raw_frame)
            selected_columns = self._resolve_selected_columns(sorted_frame.columns)
            display_frame = sorted_frame.loc[:, selected_columns] if selected_columns else sorted_frame.loc[:, []]
            total_rows = len(display_frame.index)
            display_frame, row_map = self._apply_search_filter(display_frame, row_map)
            self._preview_row_index_map = row_map
        self.preview_model.set_frame(display_frame)
        if self._column_order:
            QtCore.QTimer.singleShot(0, lambda: self._apply_preview_column_order(self._column_order))
        self.preview_table.setVisible(True)
        display_columns = (
            [str(column) for column in display_frame.columns]
            if isinstance(display_frame, pd.DataFrame)
            else []
        )
        show_graphs = any(
            column in self._inline_graph_columns for column in display_columns
        )
        try:
            if show_graphs:
                max_count = self._preview_graph_max_count(display_columns)
                spacing = 6
                icon_width = ANNEALING_GRAPH_WIDTH * max_count + spacing * (max_count - 1)
                self.preview_table.setIconSize(
                    QtCore.QSize(max(icon_width, ANNEALING_GRAPH_WIDTH), ANNEALING_GRAPH_HEIGHT)
                )
                header = self.preview_table.verticalHeader()
                if header is not None:
                    header.setDefaultSectionSize(ANNEALING_GRAPH_HEIGHT + 12)
            else:
                icon_extent = self.preview_table.style().pixelMetric(
                    QtWidgets.QStyle.PixelMetric.PM_SmallIconSize
                )
                if icon_extent <= 0:
                    icon_extent = 16
                self.preview_table.setIconSize(QtCore.QSize(icon_extent, icon_extent))
                header = self.preview_table.verticalHeader()
                if header is not None:
                    header.setDefaultSectionSize(self.preview_table.fontMetrics().height() + 8)
        except Exception:
            pass
        try:
            self.preview_table.resizeColumnsToContents()
        except Exception:
            pass
        self._update_preview_graph_buttons()
        self._update_graph_preview_panel()
        row_count = len(display_frame.index) if isinstance(display_frame, pd.DataFrame) else 0
        if row_count:
            if self._preview_search_text and total_rows != row_count:
                self.status_label.setText(
                    f"Preview ready - {row_count} of {total_rows} row(s) shown."
                )
            else:
                self.status_label.setText(f"Preview ready - {row_count} row(s).")
        else:
            if self._preview_search_text and total_rows:
                self.status_label.setText("No preview rows match the current search.")
            else:
                self.status_label.setText("Preview is empty.")
        if hasattr(self, "export_preview_button"):
            self.export_preview_button.setEnabled(row_count > 0)
        if hasattr(self, "clear_sort_button"):
            self.clear_sort_button.setEnabled(bool(self._sort_spec))

    @staticmethod
    def _normalise_search_text(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _handle_preview_search_changed(self, text: str) -> None:
        query = self._normalise_search_text(text)
        self._preview_search_text = query
        if hasattr(self, "search_clear_button"):
            self.search_clear_button.setEnabled(bool(query))
        self._refresh_preview_frame()

    def _clear_preview_search(self) -> None:
        if hasattr(self, "search_edit"):
            self.search_edit.clear()

    def _apply_search_filter(
        self,
        frame: pd.DataFrame,
        row_map: Sequence[int],
    ) -> Tuple[pd.DataFrame, List[int]]:
        query = self._preview_search_text
        if not query:
            return frame, list(row_map)
        lowered = query.casefold()
        keep_rows: List[int] = []
        for idx, row in frame.iterrows():
            for value in row.values:
                if value is None:
                    continue
                if isinstance(value, float) and math.isnan(value):
                    continue
                text = self._serialise_preview_value(value)
                if text is None:
                    continue
                if lowered in str(text).casefold():
                    keep_rows.append(int(idx))
                    break
        if not keep_rows:
            return frame.iloc[0:0].copy(), []
        filtered = frame.iloc[keep_rows].reset_index(drop=True)
        mapped_rows = [int(row_map[idx]) for idx in keep_rows if idx < len(row_map)]
        return filtered, mapped_rows

    def _preview_graph_max_count(self, columns: Sequence[str]) -> int:
        max_count = 1
        column_set = {str(column) for column in columns}
        if ANNEALING_OTHER_GRAPH_COLUMN in column_set:
            for records in self._ensure_annealing_groups().values():
                if not records:
                    continue
                high_record, low_record = _select_high_low_pair(records)
                other_records = _select_other_measurements(records, high_record, low_record)
                max_count = max(max_count, len(other_records))
        if VSM_HYSTERESIS_COLUMN in column_set:
            for records in self._ensure_vsm_hysteresis_groups().values():
                max_count = max(max_count, len(_group_vsm_hysteresis_plot_groups(records)))
        if VSM_TEMPERATURE_SCAN_COLUMN in column_set:
            for records in self._ensure_vsm_temperature_groups().values():
                max_count = max(max_count, len(records))
        if DMA_ISOSTRESS_COLUMN in column_set:
            for records in self._ensure_dma_isostress_groups().values():
                max_count = max(max_count, len(records))
        if FMR_COLUMN in column_set:
            for records in self._ensure_fmr_groups().values():
                max_count = max(max_count, len(records))
        return max(max_count, 1)

    def _resolve_selected_columns(self, columns: Sequence[str]) -> List[str]:
        column_names = [str(column) for column in columns]
        if self._selected_columns is None:
            self._selected_columns = {
                column for column in column_names if column not in self._graph_columns
            }
        else:
            new_columns = [col for col in column_names if col not in self._known_columns]
            self._selected_columns.update(
                column for column in new_columns if column not in self._graph_columns
            )
        self._known_columns.update(column_names)
        self._selected_columns.update(self._mandatory_columns)
        self._sync_section_states_from_columns(self._selected_columns, column_names)
        return [column for column in column_names if column in self._selected_columns]

    def _apply_sort_spec(self, frame: pd.DataFrame) -> Tuple[pd.DataFrame, List[int]]:
        def _index_position(index: pd.Index, value: Any, fallback: int) -> int:
            try:
                return int(value)
            except Exception:
                pass
            try:
                return int(index.get_loc(value))
            except Exception:
                return fallback

        if not self._sort_spec:
            return frame.copy(), [
                _index_position(frame.index, idx, fallback)
                for fallback, idx in enumerate(frame.index)
            ]
        sort_columns: List[str] = []
        ascending: List[bool] = []
        for column, is_ascending in self._sort_spec:
            if column in frame.columns and column not in sort_columns:
                sort_columns.append(column)
                ascending.append(bool(is_ascending))
        if not sort_columns:
            return frame.copy(), [
                _index_position(frame.index, idx, fallback)
                for fallback, idx in enumerate(frame.index)
            ]
        try:
            sorted_frame = frame.sort_values(
                by=sort_columns,
                ascending=ascending,
                kind="mergesort",
                key=lambda col: col.map(DataFrameModel._sort_value) if hasattr(col, "map") else col,
            )
        except Exception:
            return frame.copy(), [
                _index_position(frame.index, idx, fallback)
                for fallback, idx in enumerate(frame.index)
            ]
        row_map = [
            _index_position(frame.index, idx, fallback)
            for fallback, idx in enumerate(sorted_frame.index)
        ]
        return sorted_frame.reset_index(drop=True), row_map

    def _record_signature(self, records: Sequence[object]) -> Tuple[str, ...]:
        signature: List[str] = []
        for record in records:
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                signature.append(str(path))
                continue
            if isinstance(path, str) and path:
                signature.append(path)
                continue
            label = _record_label_for_display(record)
            if label:
                signature.append(label)
                continue
            sample = getattr(record, "sample", None)
            if isinstance(sample, str) and sample:
                signature.append(sample)
                continue
            signature.append(repr(record))
        return tuple(signature)

    def _combined_graph_pixmap(
        self,
        cache_key: Tuple[object, ...],
        items: Sequence[_GraphPreviewItem],
        *,
        stack_vertical: bool = False,
    ) -> Optional[QtGui.QPixmap]:
        cached = self._graph_pixmap_cache.get(cache_key)
        if cached is not None:
            return cached
        pixmaps = [item.pixmap for item in items if item.pixmap is not None]
        count = max(len(pixmaps), 1)
        spacing = 6
        if stack_vertical:
            combined = _combine_pixmaps_vertical(
                pixmaps,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT * count + spacing * (count - 1),
                spacing=spacing,
                scale_to_fit=False,
            )
        else:
            combined = _combine_pixmaps_side_by_side(
                pixmaps,
                width_px=ANNEALING_GRAPH_WIDTH * count + spacing * (count - 1),
                height_px=ANNEALING_GRAPH_HEIGHT,
                spacing=spacing,
                scale_to_fit=False,
            )
        if combined is not None:
            self._graph_pixmap_cache[cache_key] = combined
        return combined

    def _preview_decoration(
        self, row: pd.Series, column_label: str
    ) -> Optional[QtGui.QPixmap]:
        if column_label not in self._inline_graph_columns:
            return None
        key = _row_to_microwire_key(row)
        if not key:
            return None
        try:
            if column_label in FIGURE_COLUMNS:
                records = self._ensure_annealing_groups().get(key, [])
                if not records:
                    return None
                high_record, low_record = _select_high_low_pair(records)
                if column_label == FIGURE_COLUMNS[0]:
                    target = high_record
                elif column_label == FIGURE_COLUMNS[1]:
                    target = low_record
                else:
                    target = None
                if target is not None:
                    measurement_id = getattr(getattr(target, "metadata", None), "measurement_id", None)
                    cache_key = (
                        "annealing",
                        column_label,
                        measurement_id or str(getattr(target, "path", "")),
                    )
                    cached = self._graph_pixmap_cache.get(cache_key)
                    if cached is not None:
                        return cached
                    pixmap = _render_measurement_pixmap(
                        target,
                        self.logger,
                        width_px=ANNEALING_GRAPH_WIDTH,
                        height_px=ANNEALING_GRAPH_HEIGHT,
                    )
                    if pixmap is None:
                        return None
                    self._graph_pixmap_cache[cache_key] = pixmap
                    return pixmap
                other_records = _select_other_measurements(records, high_record, low_record)
                if not other_records:
                    return None
                signature = tuple(
                    str(getattr(record, "path", "")) for record in other_records
                )
                cache_key = ("annealing_other", key, signature)
                cached = self._graph_pixmap_cache.get(cache_key)
                if cached is not None:
                    return cached
                pixmap_stack: List[QtGui.QPixmap] = []
                for record in other_records:
                    preview = _render_measurement_pixmap(
                        record,
                        self.logger,
                        width_px=ANNEALING_GRAPH_WIDTH,
                        height_px=ANNEALING_GRAPH_HEIGHT,
                    )
                    if preview is not None:
                        pixmap_stack.append(preview)
                combined = _combine_pixmaps_side_by_side(
                    pixmap_stack,
                    width_px=ANNEALING_GRAPH_WIDTH * max(len(pixmap_stack), 1)
                    + 6 * (max(len(pixmap_stack), 1) - 1),
                    height_px=ANNEALING_GRAPH_HEIGHT,
                    spacing=6,
                    scale_to_fit=False,
                )
                if combined is not None:
                    self._graph_pixmap_cache[cache_key] = combined
                return combined
            if column_label == VSM_HYSTERESIS_COLUMN:
                records = self._ensure_vsm_hysteresis_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("vsm_hysteresis", key, signature)
                items = _vsm_hysteresis_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=False)
            if column_label == VSM_TEMPERATURE_SCAN_COLUMN:
                records = self._ensure_vsm_temperature_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("vsm_temperature", key, signature)
                items = _vsm_temperature_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=False)
            if column_label == DMA_ISOSTRESS_COLUMN:
                records = self._ensure_dma_isostress_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("dma_iso_stress", key, signature)
                items = _dma_iso_stress_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=False)
            if column_label == SHAPE_MEMORY_STRESS_STRAIN_COLUMN:
                records = self._ensure_shape_memory_stress_strain_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("shape_memory_stress_strain", key, signature)
                items = _shape_memory_stress_strain_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=False)
            if column_label == FMR_COLUMN:
                records = self._ensure_fmr_groups().get(key, [])
                if not records:
                    return None
                signature = self._record_signature(records)
                cache_key = ("fmr", key, signature)
                items = _fmr_preview_items(
                    records,
                    self.logger,
                    width_px=ANNEALING_GRAPH_WIDTH,
                    height_px=ANNEALING_GRAPH_HEIGHT,
                )
                return self._combined_graph_pixmap(cache_key, items, stack_vertical=False)
        except Exception:
            self.logger.exception("Failed to render assemble graph preview")
        return None

    def _group_annealing_records(
        self,
        records: Sequence[MeasurementRecord],
    ) -> Dict[str, List[MeasurementRecord]]:
        grouped: Dict[str, List[MeasurementRecord]] = {}
        for record in records:
            metadata = getattr(record, "metadata", None)
            if metadata is None:
                continue
            composition = getattr(metadata, "composition_token", None)
            draw = getattr(metadata, "draw_x", None)
            piece = getattr(metadata, "piece_y", None)
            if composition is None or draw is None or piece is None:
                continue
            suffix = None
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                parsed_key = _microscope_key(path)
                if parsed_key is not None:
                    _, _, _, suffix = parsed_key
            try:
                key = _microwire_key_to_str((str(composition), int(draw), int(piece), suffix))
            except (TypeError, ValueError):
                continue
            grouped.setdefault(key, []).append(record)
        return grouped

    def _ensure_annealing_groups(self) -> Dict[str, List[MeasurementRecord]]:
        groups = self._cached_annealing_groups
        if not groups:
            payload = self._load_payload("annealing", "annealing_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "annealing")
                self._cached_annealing_records = list(records)
                self._cached_annealing_groups = self._group_annealing_records(records)
                groups = self._cached_annealing_groups
        return groups

    def _filter_hidden_records(
        self, records: Sequence[object], section_key: str
    ) -> List[object]:
        hidden = _hidden_paths_from_section(self.sections.get(section_key))
        if not hidden:
            return list(records)
        filtered: List[object] = []
        for record in records:
            path_key = _record_path_key(record)
            if path_key and path_key in hidden:
                continue
            filtered.append(record)
        return filtered

    def _ensure_vsm_hysteresis_groups(self) -> Dict[str, List[VsmHysteresisRecord]]:
        groups = self._cached_vsm_hysteresis_groups
        if not groups:
            payload = self._load_payload("vsm_hysteresis", "vsm_hysteresis_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "vsm_hysteresis")
                self._cached_vsm_hysteresis_records = list(records)
                self._cached_vsm_hysteresis_groups = _group_graph_records_by_key(records)
                groups = self._cached_vsm_hysteresis_groups
        return groups

    def _ensure_vsm_temperature_groups(self) -> Dict[str, List[VsmTemperatureScanRecord]]:
        groups = self._cached_vsm_temperature_groups
        if not groups:
            payload = self._load_payload("vsm_temperature_scan", "vsm_temperature_scan_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "vsm_temperature_scan")
                self._cached_vsm_temperature_records = list(records)
                self._cached_vsm_temperature_groups = _group_graph_records_by_key(records)
                groups = self._cached_vsm_temperature_groups
        return groups

    def _ensure_dma_isostress_groups(self) -> Dict[str, List[DmaIsoStressRecord]]:
        groups = self._cached_dma_isostress_groups
        if not groups:
            payload = self._load_payload("dma_iso_stress", "dma_iso_stress_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "dma_iso_stress")
                self._cached_dma_isostress_records = list(records)
                self._cached_dma_isostress_groups = _group_graph_records_by_key(records)
                groups = self._cached_dma_isostress_groups
        return groups

    def _ensure_shape_memory_stress_strain_groups(
        self,
    ) -> Dict[str, List[ShapeMemoryStressStrainRecord]]:
        groups = self._cached_shape_memory_stress_strain_groups
        if not groups:
            payload = self._load_payload(
                "shape_memory_stress_strain",
                "shape_memory_stress_strain_records",
            )
            if isinstance(payload, list):
                records = self._filter_hidden_records(
                    payload,
                    "shape_memory_stress_strain",
                )
                self._cached_shape_memory_stress_strain_records = list(records)
                self._cached_shape_memory_stress_strain_groups = (
                    _group_graph_records_by_key(records)
                )
                groups = self._cached_shape_memory_stress_strain_groups
        return groups

    def _ensure_fmr_groups(self) -> Dict[str, List[FmrRecord]]:
        groups = self._cached_fmr_groups
        if not groups:
            payload = self._load_payload("fmr", "fmr_records")
            if isinstance(payload, list):
                records = self._filter_hidden_records(payload, "fmr")
                self._cached_fmr_records = list(records)
                self._cached_fmr_groups = _group_graph_records_by_key(records)
                groups = self._cached_fmr_groups
            if not groups:
                section = self.sections.get("fmr")
                if isinstance(section, FmrSection):
                    fallback = getattr(section, "_record_groups_by_key", None)
                    if isinstance(fallback, dict) and fallback:
                        self._cached_fmr_records = list(
                            getattr(section, "_all_records", []) or []
                        )
                        self._cached_fmr_groups = dict(fallback)
                        groups = self._cached_fmr_groups
        return groups

    def _selected_preview_row_index(self) -> Optional[int]:
        if not isinstance(self.preview_table, QtWidgets.QTableView):
            return None
        selection = self.preview_table.selectionModel()
        if selection is None:
            return None
        rows = selection.selectedRows()
        if rows:
            return rows[0].row()
        current = selection.currentIndex()
        if current.isValid():
            return current.row()
        return None

    def _selected_preview_rows(self) -> List[int]:
        if not isinstance(self.preview_table, QtWidgets.QTableView):
            return []
        selection = self.preview_table.selectionModel()
        if selection is None:
            return []
        rows = selection.selectedRows()
        if rows:
            return sorted({index.row() for index in rows})
        current = selection.currentIndex()
        if current.isValid():
            return [current.row()]
        return sorted({index.row() for index in selection.selectedIndexes()})

    def _selected_preview_row(self, *, raw: bool = False) -> Optional[pd.Series]:
        row_index = self._selected_preview_row_index()
        if row_index is None or row_index < 0:
            return None
        if raw:
            frame = self._raw_preview_frame
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                return None
            source_index = row_index
            if self._preview_row_index_map and row_index < len(self._preview_row_index_map):
                source_index = self._preview_row_index_map[row_index]
            if source_index < 0 or source_index >= len(frame.index):
                return None
            try:
                return frame.iloc[source_index]
            except Exception:
                return None
        frame = self.preview_model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return None
        if row_index >= len(frame.index):
            return None
        try:
            return frame.iloc[row_index]
        except Exception:
            return None

    def _update_preview_graph_buttons(self, *_: Any) -> None:
        try:
            row_index = self._selected_preview_row_index()
            row = self._selected_preview_row(raw=True)
            key = _row_to_microwire_key(row) if row is not None else None
            enabled = row_index is not None and key is not None
            if hasattr(self, "open_high_plot_button"):
                self.open_high_plot_button.setEnabled(enabled)
            if hasattr(self, "open_low_plot_button"):
                self.open_low_plot_button.setEnabled(enabled)
            if hasattr(self, "open_vsm_hysteresis_button"):
                self.open_vsm_hysteresis_button.setEnabled(
                    bool(enabled and self._ensure_vsm_hysteresis_groups().get(key or "", []))
                )
            if hasattr(self, "open_vsm_temperature_button"):
                self.open_vsm_temperature_button.setEnabled(
                    bool(enabled and self._ensure_vsm_temperature_groups().get(key or "", []))
                )
            if hasattr(self, "open_dma_button"):
                self.open_dma_button.setEnabled(
                    bool(enabled and self._ensure_dma_isostress_groups().get(key or "", []))
                )
            if hasattr(self, "open_shape_memory_button"):
                self.open_shape_memory_button.setEnabled(
                    bool(
                        enabled
                        and self._ensure_shape_memory_stress_strain_groups().get(
                            key or "",
                            [],
                        )
                    )
                )
            if hasattr(self, "open_fmr_button"):
                self.open_fmr_button.setEnabled(
                    bool(enabled and self._ensure_fmr_groups().get(key or "", []))
                )
            if hasattr(self, "add_to_compare_button"):
                self.add_to_compare_button.setEnabled(bool(row_index is not None))
        except Exception as exc:
            self.logger.exception("Failed to update assemble preview buttons")
            self.log(
                f"Failed to update assemble preview buttons: {exc}",
                level=logging.ERROR,
            )

    def _toggle_graph_preview_panel(self, checked: bool) -> None:
        if not hasattr(self, "graph_preview_panel"):
            return
        self.graph_preview_panel.setVisible(bool(checked))
        splitter = getattr(self, "preview_splitter", None)
        if isinstance(splitter, QtWidgets.QSplitter):
            if checked:
                sizes = splitter.sizes()
                if sizes and sizes[1] == 0:
                    total = sum(sizes) or 1
                    splitter.setSizes([int(total * 0.6), int(total * 0.4)])
            else:
                splitter.setSizes([1, 0])
        self._update_graph_preview_panel()

    def _update_graph_preview_panel(self, *_: Any) -> None:
        try:
            if not getattr(self, "graph_preview_panel", None):
                return
            if not self.graph_preview_panel.isVisible():
                return
            row = self._selected_preview_row(raw=True)
            if row is None:
                self.high_preview_display.set_record(
                    None,
                    setpoint=None,
                    description="Select a row to preview the annealing measurement.",
                )
                self.low_preview_display.set_record(
                    None,
                    setpoint=None,
                    description="Select a row to preview the annealing measurement.",
                )
                self.vsm_hysteresis_gallery.clear("Select a row to preview VSM hysteresis graphs.")
                self.vsm_temperature_gallery.clear("Select a row to preview VSM temperature scans.")
                self.dma_iso_gallery.clear("Select a row to preview DMA iso-stress graphs.")
                self.shape_memory_gallery.clear("Select a row to preview shape-memory graphs.")
                self.fmr_gallery.clear("Select a row to preview FMR graphs.")
                return
            key = _row_to_microwire_key(row)
            if not key:
                message = "Select a microwire row with annealing data first."
                self.high_preview_display.set_record(
                    None, setpoint=None, description=message
                )
                self.low_preview_display.set_record(
                    None, setpoint=None, description=message
                )
                self.vsm_hysteresis_gallery.clear(message)
                self.vsm_temperature_gallery.clear(message)
                self.dma_iso_gallery.clear(message)
                self.shape_memory_gallery.clear(message)
                self.fmr_gallery.clear(message)
                return
            records = self._ensure_annealing_groups().get(key, [])
            if records:
                high_record, low_record = _select_high_low_pair(records)
                self.high_preview_display.set_record(
                    high_record,
                    setpoint=_extract_setpoint(high_record),
                    description="No 1000 mA measurement available for this microwire.",
                )
                self.low_preview_display.set_record(
                    low_record,
                    setpoint=_extract_setpoint(low_record),
                    description="No low mA measurement available for this microwire.",
                )
            else:
                message = "No annealing records found for this microwire."
                self.high_preview_display.set_record(
                    None, setpoint=None, description=message
                )
                self.low_preview_display.set_record(
                    None, setpoint=None, description=message
                )

            vsm_records = self._ensure_vsm_hysteresis_groups().get(key, [])
            vsm_items = _vsm_hysteresis_preview_items(
                vsm_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.vsm_hysteresis_gallery.set_items(
                vsm_items, "No VSM hysteresis graphs available for this microwire."
            )

            vsm_temp_records = self._ensure_vsm_temperature_groups().get(key, [])
            vsm_temp_items = _vsm_temperature_preview_items(
                vsm_temp_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.vsm_temperature_gallery.set_items(
                vsm_temp_items, "No VSM temperature scans available for this microwire."
            )

            dma_records = self._ensure_dma_isostress_groups().get(key, [])
            dma_items = _dma_iso_stress_preview_items(
                dma_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.dma_iso_gallery.set_items(
                dma_items, "No DMA iso-stress graphs available for this microwire."
            )

            shape_memory_records = self._ensure_shape_memory_stress_strain_groups().get(key, [])
            shape_memory_items = _shape_memory_stress_strain_preview_items(
                shape_memory_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.shape_memory_gallery.set_items(
                shape_memory_items,
                "No shape-memory graphs available for this microwire.",
            )

            fmr_records = self._ensure_fmr_groups().get(key, [])
            fmr_items = _fmr_preview_items(
                fmr_records,
                self.logger,
                width_px=ANNEALING_GRAPH_WIDTH,
                height_px=ANNEALING_GRAPH_HEIGHT,
            )
            self.fmr_gallery.set_items(
                fmr_items, "No FMR graphs available for this microwire."
            )
        except Exception as exc:
            self.logger.exception("Failed to update compare graph preview panel")
            self.log(
                f"Failed to update compare graph preview panel: {exc}",
                level=logging.ERROR,
            )

    def _section_column_map(self, columns: Sequence[str]) -> Dict[str, List[str]]:
        available = {str(column) for column in columns}
        mapping: Dict[str, List[str]] = {key: [] for key, _ in self._section_choices}

        def add(key: str, members: Sequence[str]) -> None:
            if key not in mapping:
                return
            for column in members:
                if column in available and column not in mapping[key]:
                    mapping[key].append(column)

        add(
            "microscope",
            [
                MICROSCOPE_D_COLUMN,
                MICROSCOPE_CAP_D_COLUMN,
                "d/D",
                *MICROSCOPE_IMAGE_COLUMNS,
            ],
        )
        add(
            "annealing",
            [
                "Low mA value (mA)",
                "File 1000 mA",
                "File low mA",
                *FIGURE_COLUMNS,
                *ORIGIN_FIGURE_COLUMNS,
            ],
        )
        add("current_density", CURRENT_DENSITY_COLUMNS)
        add(
            "transition_temps",
            [
                TRANSITION_TEMP_AS_COLUMN,
                TRANSITION_TEMP_AF_COLUMN,
                TRANSITION_TEMP_MS_COLUMN,
                TRANSITION_TEMP_MF_COLUMN,
            ],
        )
        add(
            "fabrication",
            [
                "Length (m)",
                "Production datetime",
                "Mass (g)",
                "Resistance (Ω)",
                CORE_TEMPERATURE_COLUMN,
                GLASS_TEMPERATURE_COLUMN,
                ESTIMATED_TRANSITION_COLUMN,
                "Notes",
                GLASS_PULL_COLUMN,
            ],
        )
        add(
            "videos",
            [
                CORE_TEMPERATURE_COLUMN,
                GLASS_TEMPERATURE_COLUMN,
                "Winding speed (m/min)",
                "Glass feeding (mm/min)",
                "Underpressure",
                VIDEO_END_LENGTH_COLUMN,
                VIDEO_MW_LENGTH_COLUMN,
            ],
        )
        add(
            "strain",
            [
                StrainSection.COLUMN_STRAIN,
                "Calc mode",
                "Clamp span (mm)",
                "m",
                StrainSection.COLUMN_TARGET_STRESS,
                "M length",
                "A length",
                "Broke",
            ],
        )
        add("vsm_hysteresis", [VSM_HYSTERESIS_COLUMN])
        add("vsm_temperature_scan", [VSM_TEMPERATURE_SCAN_COLUMN])
        add("dma_iso_stress", [DMA_ISOSTRESS_COLUMN])
        add(
            "shape_memory_stress_strain",
            [
                SHAPE_MEMORY_STRESS_STRAIN_COLUMN,
                SHAPE_MEMORY_DISPLACEMENT_COLUMN,
                SHAPE_MEMORY_LOAD_COLUMN,
                SHAPE_MEMORY_STRAIN_COLUMN,
                SHAPE_MEMORY_STRESS_COLUMN,
                SHAPE_MEMORY_FRACTURE_LOAD_COLUMN,
                SHAPE_MEMORY_FRACTURE_STRAIN_COLUMN,
                SHAPE_MEMORY_FRACTURE_STRESS_COLUMN,
            ],
        )
        add("fmr", [FMR_COLUMN])
        return mapping

    def _sync_section_states_from_columns(
        self,
        selected_columns: Set[str],
        available_columns: Sequence[str],
    ) -> None:
        mapping = self._section_column_map(available_columns)
        for key, _label in self._section_choices:
            section_columns = mapping.get(key, [])
            if not section_columns:
                continue
            self._section_states[key] = any(
                column in selected_columns for column in section_columns
            )
        self._update_export_summary()

    def _column_groups(self, columns: Sequence[str]) -> Dict[str, List[str]]:
        available = [str(column) for column in columns]
        available_set = {str(column) for column in columns}
        included: Set[str] = set()
        groups: List[Tuple[str, List[str]]] = []

        def add_group(label: str, members: Sequence[str]) -> None:
            subset = [col for col in members if col in available_set]
            if not subset:
                return
            groups.append((label, subset))
            included.update(subset)

        section_map = self._section_column_map(columns)
        add_group("Core", ["Composition", "Microwire"])
        add_group("Microscope", section_map.get("microscope", []))
        add_group("Current annealing", section_map.get("annealing", []))
        add_group("Current density", section_map.get("current_density", []))
        add_group("VSM hysteresis", section_map.get("vsm_hysteresis", []))
        add_group("VSM temperature scan", section_map.get("vsm_temperature_scan", []))
        add_group("Transition temps", section_map.get("transition_temps", []))
        add_group("DMA iso-stress", section_map.get("dma_iso_stress", []))
        add_group(
            "Shape memory stress/strain",
            section_map.get("shape_memory_stress_strain", []),
        )
        add_group("FMR", section_map.get("fmr", []))
        add_group("Fabrication", section_map.get("fabrication", []))
        add_group("Videos", section_map.get("videos", []))
        add_group("Strain", section_map.get("strain", []))
        remaining = [col for col in available if col not in included]
        if remaining:
            groups.append(("Other", remaining))
        return dict(groups)

    def _open_column_selector(self) -> None:
        frame = self._raw_preview_frame
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Select columns",
                "Preview the database first to load the available columns.",
            )
            return
        columns = self._column_universe()
        for column in frame.columns:
            text = str(column)
            if text not in columns:
                columns.append(text)
        groups = self._column_groups(columns)
        selected = self._selected_columns or {
            column for column in columns if column not in self._graph_columns
        }
        dialog = _ColumnSelectionDialog(
            groups,
            selected,
            self._mandatory_columns,
            parent=self,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._selected_columns = dialog.selected_columns()
            self._sync_section_states_from_columns(self._selected_columns, columns)
            self._refresh_preview_frame()

    def _open_column_order_dialog(self) -> None:
        frame = self.preview_model.frame()
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Reorder columns",
                "Preview the database first to load the available columns.",
            )
            return
        current_order = self._current_preview_column_order()
        columns = current_order or [str(column) for column in frame.columns]
        dialog = _ColumnOrderDialog(columns, parent=self)
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._column_order = dialog.ordered_columns()
            self._apply_preview_column_order(self._column_order)

    def _open_sort_dialog(self) -> None:
        frame = self._raw_preview_frame
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Sort preview",
                "Preview the database first to load the sortable columns.",
            )
            return
        default_column = None
        current_index = self.preview_table.currentIndex()
        if current_index.isValid():
            display_frame = self.preview_model.frame()
            if isinstance(display_frame, pd.DataFrame):
                try:
                    default_column = str(display_frame.columns[current_index.column()])
                except Exception:
                    default_column = None
        dialog = _SortSpecDialog(
            [str(column) for column in frame.columns],
            self._sort_spec,
            parent=self,
            default_column=default_column,
        )
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self._sort_spec = dialog.sort_spec()
            self._refresh_preview_frame()

    def _clear_sort(self) -> None:
        if not self._sort_spec:
            return
        self._sort_spec = []
        self._refresh_preview_frame()

    def _reset_column_order(self) -> None:
        frame = self.preview_model.frame()
        if not isinstance(frame, pd.DataFrame):
            return
        self._column_order = [str(column) for column in frame.columns]
        self._apply_preview_column_order(self._column_order)

    def _add_selected_to_compare(self) -> None:
        compare_section = self._compare_section
        if compare_section is None:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "Compare section is not available yet.",
            )
            return
        raw_frame = self._raw_preview_frame
        if not isinstance(raw_frame, pd.DataFrame) or raw_frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "Preview the database first to add rows to Compare.",
            )
            return
        rows = self._selected_preview_rows()
        if not rows:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "Select one or more rows to compare first.",
            )
            return
        mapped_rows: List[int] = []
        for row in rows:
            if row < 0:
                continue
            if self._preview_row_index_map and row < len(self._preview_row_index_map):
                mapped_rows.append(self._preview_row_index_map[row])
            else:
                mapped_rows.append(row)
        added = compare_section.add_rows_from_frame(raw_frame, mapped_rows)
        if added:
            self.log(f"Compare: added {added} row(s).")
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                f"Added {added} row(s) to Compare.",
            )
        else:
            self.log("Compare: no new rows added.")
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "No new rows were added to Compare (they may already be present).",
            )

    def _open_preview_graph(self, kind: str) -> None:
        row = self._selected_preview_row(raw=True)
        if row is None:
            return
        key = _row_to_microwire_key(row)
        if not key:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                "Select a microwire row with annealing data first.",
            )
            return
        if kind in {"high", "low"}:
            records = self._ensure_annealing_groups().get(key, [])
            if not records:
                QtWidgets.QMessageBox.information(
                    self,
                    "Microwire Data Builder",
                    "No annealing records found for the selected microwire.",
                )
                return
            high_record, low_record = _select_high_low_pair(records)
            record = high_record if kind == "high" else low_record
            label = "1000 mA" if kind == "high" else "low mA"
            if record is None:
                QtWidgets.QMessageBox.information(
                    self,
                    "Microwire Data Builder",
                    f"No {label} measurement available for this microwire.",
                )
                return
            self._show_annealing_record(record, label)
            return

        if kind == "vsm_hysteresis":
            records = self._ensure_vsm_hysteresis_groups().get(key, [])
            title = "VSM hysteresis graphs"
            items = _vsm_hysteresis_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No VSM hysteresis graphs available."
        elif kind == "vsm_temperature":
            records = self._ensure_vsm_temperature_groups().get(key, [])
            title = "VSM temperature scan graphs"
            items = _vsm_temperature_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No VSM temperature scan graphs available."
        elif kind == "dma_iso_stress":
            records = self._ensure_dma_isostress_groups().get(key, [])
            title = "DMA iso-stress graphs"
            items = _dma_iso_stress_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No DMA iso-stress graphs available."
        elif kind == "shape_memory_stress_strain":
            records = self._ensure_shape_memory_stress_strain_groups().get(key, [])
            title = "Shape memory stress/strain graphs"
            items = _shape_memory_stress_strain_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No shape-memory graphs available."
        else:
            records = self._ensure_fmr_groups().get(key, [])
            title = "FMR graphs"
            items = _fmr_preview_items(
                records,
                self.logger,
                width_px=GRAPH_PREVIEW_WIDTH,
                height_px=GRAPH_PREVIEW_HEIGHT,
            )
            empty_message = "No FMR graphs available."
        if not records:
            QtWidgets.QMessageBox.information(
                self,
                "Microwire Data Builder",
                empty_message,
            )
            return
        dialog = _GraphGalleryDialog(
            title,
            items,
            parent=self,
            empty_message=empty_message,
        )
        dialog.exec()

    def _show_annealing_record(self, record: MeasurementRecord, label: str) -> None:
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"Annealing graph — {label}")
        dialog.resize(960, 640)
        layout = QtWidgets.QVBoxLayout(dialog)
        display = _AnnealingPlotDisplay(f"Graph — {label}", self.logger, dialog)
        display.set_record(
            record,
            setpoint=_extract_setpoint(record),
            description="No annealing measurement available for this microwire.",
        )
        layout.addWidget(display, 1)
        dialog.exec()

    def _current_preview_column_order(self) -> List[str]:
        header = self.preview_table.horizontalHeader()
        frame = self.preview_model.frame()
        if header is None or not isinstance(frame, pd.DataFrame):
            return []
        order: List[str] = []
        for visual_index in range(header.count()):
            logical = header.logicalIndex(visual_index)
            try:
                column = str(frame.columns[logical])
            except Exception:
                continue
            order.append(column)
        return order

    def _handle_preview_column_moved(self, *_: Any) -> None:
        self._column_order = self._current_preview_column_order()

    def _apply_preview_column_order(self, order: Sequence[str]) -> None:
        if not order:
            return
        header = self.preview_table.horizontalHeader()
        frame = self.preview_model.frame()
        if header is None or not isinstance(frame, pd.DataFrame):
            return
        mapping = {str(column): idx for idx, column in enumerate(frame.columns)}
        for target_visual, column_name in enumerate(order):
            logical = mapping.get(column_name)
            if logical is None:
                continue
            current_visual = header.visualIndex(logical)
            if current_visual == target_visual:
                continue
            header.moveSection(current_visual, target_visual)

    def _ordered_preview_frame(self) -> pd.DataFrame:
        frame = self.preview_model.frame()
        if not isinstance(frame, pd.DataFrame):
            return pd.DataFrame()
        order = self._current_preview_column_order()
        if order:
            remaining = [column for column in frame.columns if column not in order]
            return frame.loc[:, list(order) + remaining].copy()
        return frame.copy()

    @staticmethod
    def _serialise_preview_value(value: Any) -> Any:
        if isinstance(value, (QtGui.QPixmap, QtGui.QImage)):
            return ""
        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(item) for item in value)
        return value

    def _current_column_filter(self) -> Optional[Tuple[str, ...]]:
        if not self._selected_columns:
            return None
        selected = set(self._selected_columns) | self._mandatory_columns
        frame = self._raw_preview_frame
        if isinstance(frame, pd.DataFrame):
            ordered = [str(column) for column in frame.columns if str(column) in selected]
        else:
            ordered = sorted(selected)
        return tuple(ordered) if ordered else None

    def _current_column_order_for_export(self) -> Optional[Tuple[str, ...]]:
        order = self._column_order or self._current_preview_column_order()
        return tuple(order) if order else None

    def _current_sort_spec(self) -> Optional[Tuple[Tuple[str, bool], ...]]:
        if not self._sort_spec:
            return None
        return tuple((str(column), bool(ascending)) for column, ascending in self._sort_spec)

    def _export_preview_worksheet(self) -> None:
        frame = self._preview_export_frame()
        if frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Export worksheet",
                "There is no preview data to export.",
            )
            return
        preferred_dir = self._output_dir
        start_dir = _dialog_start_directory(preferred_dir) if preferred_dir else _dialog_start_directory()
        default_name = _normalise_output_name(self._output_name or DEFAULT_OUTPUT_NAME)
        suggested = start_dir / f"{default_name}_worksheet.xlsx"
        path_str, selected_filter = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export worksheet",
            str(suggested),
            "Excel files (*.xlsx);;CSV files (*.csv)",
        )
        if not path_str:
            return
        path = Path(path_str)
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".csv"}:
            if "Excel" in selected_filter:
                path = path.with_suffix(".xlsx")
                suffix = ".xlsx"
            else:
                path = path.with_suffix(".csv")
                suffix = ".csv"
        try:
            if suffix == ".xlsx":
                frame.to_excel(path, index=False)
            else:
                frame.to_csv(path, index=False)
        except Exception as exc:
            self.logger.exception("Failed to export preview worksheet")
            QtWidgets.QMessageBox.critical(
                self,
                "Export worksheet",
                f"Failed to export worksheet:\n{exc}",
            )
            return
        self.log(f"Preview worksheet exported to {path}")
        QtWidgets.QMessageBox.information(
            self,
            "Export worksheet",
            f"Worksheet exported to:\n{path}",
        )

    def _preview_export_frame(self) -> pd.DataFrame:
        export_frame = self._ordered_preview_frame().copy()
        for column in export_frame.columns:
            series = export_frame[column]
            if getattr(series, "dtype", None) == object:
                export_frame[column] = series.map(self._serialise_preview_value)
        return export_frame

    def _html_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and math.isnan(value):
            return ""
        if isinstance(value, (pd.Timestamp, datetime)):
            return value.isoformat(sep=" ", timespec="seconds")
        text = self._serialise_preview_value(value)
        if text is None:
            return ""
        text_str = str(text)
        return "" if text_str.lower() == "nan" else text_str

    def _resolve_image_path_for_html(self, value: Any, output_dir: Path) -> Optional[Path]:
        path = self._coerce_path(value)
        if path is None:
            return None
        candidate = path if path.is_absolute() else output_dir / path
        if candidate.exists():
            return candidate
        if path.exists():
            return path
        return None

    def _image_path_to_data_uri(
        self,
        path: Path,
        cache: Dict[Path, str],
    ) -> Optional[str]:
        cached = cache.get(path)
        if cached is not None:
            return cached
        try:
            data = path.read_bytes()
        except Exception:
            cache[path] = ""
            return None
        ext = path.suffix.lower()
        if ext in {".jpg", ".jpeg"}:
            mime = "image/jpeg"
        elif ext == ".gif":
            mime = "image/gif"
        else:
            mime = "image/png"
        encoded = base64.b64encode(data).decode("ascii")
        uri = f"data:{mime};base64,{encoded}"
        cache[path] = uri
        return uri

    def _graph_data_uri(
        self,
        record: Optional[MeasurementRecord],
        plot_dir: Path,
        cache: Dict[str, str],
        image_cache: Dict[Path, str],
    ) -> str:
        if record is None:
            return ""
        measurement_id = getattr(getattr(record, "metadata", object()), "measurement_id", None)
        if isinstance(measurement_id, str) and measurement_id in cache:
            return cache[measurement_id]
        try:
            plot_path = _plot_measurement_matplotlib(
                record.dataframe,
                record.path,
                plot_dir,
                DEFAULT_FIGSIZE,
            )
        except Exception:
            if isinstance(measurement_id, str):
                cache[measurement_id] = ""
            return ""
        uri = self._image_path_to_data_uri(plot_path, image_cache)
        if isinstance(measurement_id, str):
            cache[measurement_id] = uri or ""
        return uri or ""

    def _export_html_file(
        self,
        frame: pd.DataFrame,
        output_dir: Path,
        output_name: str,
    ) -> Optional[Path]:
        if frame.empty:
            QtWidgets.QMessageBox.information(
                self,
                "Export HTML",
                "There is no preview data to export.",
            )
            return None
        html_path = output_dir / f"{output_name}.html"
        groups = self._ensure_annealing_groups()
        vsm_hyst_groups = self._ensure_vsm_hysteresis_groups()
        vsm_temp_groups = self._ensure_vsm_temperature_groups()
        dma_groups = self._ensure_dma_isostress_groups()
        shape_memory_groups = self._ensure_shape_memory_stress_strain_groups()
        fmr_groups = self._ensure_fmr_groups()
        image_cache: Dict[Path, str] = {}
        graph_cache: Dict[str, str] = {}
        vsm_hyst_cache: Dict[str, str] = {}
        vsm_temp_cache: Dict[str, str] = {}
        dma_cache: Dict[str, str] = {}
        shape_memory_cache: Dict[str, str] = {}
        fmr_cache: Dict[str, str] = {}
        rows_html: List[str] = []
        has_graphs = False
        has_microscope = False
        has_vsm_hyst = False
        has_vsm_temp = False
        has_dma = False
        has_shape_memory = False
        has_fmr = False
        vsm_temp_processor = _get_vsm_temp_processor(self.logger) if vsm_temp_groups else None

        def _record_cache_key(record: object) -> str:
            if isinstance(record, _VsmHysteresisPlotGroup):
                paths = [
                    str(entry.path)
                    for entry in record.records
                    if isinstance(getattr(entry, "path", None), Path)
                ]
                return "|".join(paths) or record.label or repr(record)
            path = getattr(record, "path", None)
            if isinstance(path, Path):
                return str(path)
            if isinstance(path, str) and path:
                return path
            label = _record_label_for_display(record)
            return label or repr(record)

        def _record_to_data_uri(
            record: object,
            cache: Dict[str, str],
            builder: Callable[[object], Optional["plt.Figure"]],
        ) -> str:
            key = _record_cache_key(record)
            if key in cache:
                return cache[key]
            try:
                figure = builder(record)
            except Exception:
                self.logger.exception("Failed to render graph for HTML export")
                cache[key] = ""
                return ""
            uri = _figure_to_data_uri(figure, self.logger)
            cache[key] = uri or ""
            return uri or ""

        with TemporaryDirectory() as tmp_dir:
            plot_dir = Path(tmp_dir)
            for _, row in frame.iterrows():
                composition = str(row.get("Composition") or "").strip()
                microwire = str(row.get("Microwire") or "").strip()
                high_uri = ""
                low_uri = ""
                key = None
                if composition and microwire:
                    parsed = _microwire_parts_from_label_safe(microwire)
                    if parsed is not None:
                        draw, piece, suffix = parsed
                        key = _microwire_key_to_str((composition, int(draw), int(piece), suffix))
                        records = groups.get(key, [])
                        if records:
                            high_record, low_record = _select_high_low_pair(records)
                            high_uri = self._graph_data_uri(
                                high_record, plot_dir, graph_cache, image_cache
                            )
                            low_uri = self._graph_data_uri(
                                low_record, plot_dir, graph_cache, image_cache
                            )
                            other_records = _select_other_measurements(
                                records, high_record, low_record
                            )
                            other_uris = [
                                self._graph_data_uri(
                                    record, plot_dir, graph_cache, image_cache
                                )
                                for record in other_records
                            ]
                        else:
                            other_uris = []
                    else:
                        other_uris = []
                else:
                    other_uris = []
                if high_uri or low_uri or any(other_uris):
                    has_graphs = True
                vsm_hyst_uris: List[str] = []
                vsm_temp_uris: List[str] = []
                dma_uris: List[str] = []
                shape_memory_uris: List[str] = []
                fmr_uris: List[str] = []
                if key:
                    vsm_records = vsm_hyst_groups.get(key, [])
                    vsm_groups = _group_vsm_hysteresis_plot_groups(vsm_records)
                    for group in vsm_groups:
                        uri = _record_to_data_uri(
                            group,
                            vsm_hyst_cache,
                            lambda grp: _plot_vsm_hysteresis_figure(
                                grp.records,
                                self.logger,
                                width_px=GRAPH_PREVIEW_WIDTH,
                                height_px=GRAPH_PREVIEW_HEIGHT,
                            ),
                        )
                        if uri:
                            vsm_hyst_uris.append(uri)
                    if vsm_hyst_uris:
                        has_vsm_hyst = True

                    vsm_temp_records = vsm_temp_groups.get(key, [])
                    if vsm_temp_processor is not None:
                        for record in vsm_temp_records:
                            uri = _record_to_data_uri(
                                record,
                                vsm_temp_cache,
                                lambda rec: _plot_vsm_temperature_scan_figure(
                                    rec,
                                    vsm_temp_processor,
                                    width_px=GRAPH_PREVIEW_WIDTH,
                                    height_px=GRAPH_PREVIEW_HEIGHT,
                                ),
                            )
                            if uri:
                                vsm_temp_uris.append(uri)
                    if vsm_temp_uris:
                        has_vsm_temp = True

                    dma_records = dma_groups.get(key, [])
                    for record in dma_records:
                        uri = _record_to_data_uri(
                            record,
                            dma_cache,
                            lambda rec: _plot_dma_iso_stress_figure(
                                rec, width_px=GRAPH_PREVIEW_WIDTH, height_px=GRAPH_PREVIEW_HEIGHT
                            ),
                        )
                        if uri:
                            dma_uris.append(uri)
                    if dma_uris:
                        has_dma = True
                    shape_memory_records = shape_memory_groups.get(key, [])
                    for record in shape_memory_records:
                        uri = _record_to_data_uri(
                            record,
                            shape_memory_cache,
                            lambda rec: _plot_shape_memory_stress_strain_figure(
                                rec,
                                width_px=GRAPH_PREVIEW_WIDTH,
                                height_px=GRAPH_PREVIEW_HEIGHT,
                            ),
                        )
                        if uri:
                            shape_memory_uris.append(uri)
                    if shape_memory_uris:
                        has_shape_memory = True
                    fmr_records = fmr_groups.get(key, [])
                    for record in fmr_records:
                        uri = _record_to_data_uri(
                            record,
                            fmr_cache,
                            lambda rec: _plot_fmr_figure(
                                rec,
                                self.logger,
                                width_px=GRAPH_PREVIEW_WIDTH,
                                height_px=GRAPH_PREVIEW_HEIGHT,
                            ),
                        )
                        if uri:
                            fmr_uris.append(uri)
                    if fmr_uris:
                        has_fmr = True
                core_uri = ""
                glass_uri = ""
                if MICROSCOPE_IMAGE_COLUMNS[0] in frame.columns:
                    core_path = self._resolve_image_path_for_html(
                        row.get(MICROSCOPE_IMAGE_COLUMNS[0]), output_dir
                    )
                    if core_path is not None:
                        core_uri = self._image_path_to_data_uri(core_path, image_cache) or ""
                if MICROSCOPE_IMAGE_COLUMNS[1] in frame.columns:
                    glass_path = self._resolve_image_path_for_html(
                        row.get(MICROSCOPE_IMAGE_COLUMNS[1]), output_dir
                    )
                    if glass_path is not None:
                        glass_uri = self._image_path_to_data_uri(glass_path, image_cache) or ""
                if core_uri or glass_uri:
                    has_microscope = True

                cells: List[str] = []
                for column in frame.columns:
                    cell_text = self._html_cell_value(row.get(column))
                    cells.append(f"<td>{html.escape(cell_text)}</td>")
                vsm_hyst_blob = "|".join(vsm_hyst_uris)
                vsm_temp_blob = "|".join(vsm_temp_uris)
                dma_blob = "|".join(dma_uris)
                shape_memory_blob = "|".join(shape_memory_uris)
                other_blob = "|".join(uri for uri in other_uris if uri)
                fmr_blob = "|".join(fmr_uris)
                attrs = [
                    f'data-high="{html.escape(high_uri)}"',
                    f'data-low="{html.escape(low_uri)}"',
                    f'data-other="{html.escape(other_blob)}"',
                    f'data-core="{html.escape(core_uri)}"',
                    f'data-glass="{html.escape(glass_uri)}"',
                    f'data-vsm-hyst="{html.escape(vsm_hyst_blob)}"',
                    f'data-vsm-temp="{html.escape(vsm_temp_blob)}"',
                    f'data-dma="{html.escape(dma_blob)}"',
                    f'data-shape-memory="{html.escape(shape_memory_blob)}"',
                    f'data-fmr="{html.escape(fmr_blob)}"',
                ]
                rows_html.append(f"<tr {' '.join(attrs)}>{''.join(cells)}</tr>")

        header_cells = "".join(
            f"<th>{html.escape(str(column))}</th>" for column in frame.columns
        )
        title_text = html.escape(output_name or "Microwire database")
        preview_class = "preview-panel"
        preview_classes = preview_class
        microscope_section = ""
        if has_microscope:
            microscope_section = """
            <div class="preview-section">
              <div class="preview-title">Microscope images</div>
              <div class="preview-grid">
                <div class="preview-card">
                  <div class="preview-label">d (core)</div>
                  <img id="preview-core" class="preview-image" alt="Core image" />
                  <div id="preview-core-empty" class="preview-empty">No core image</div>
                </div>
                <div class="preview-card">
                  <div class="preview-label">D (glass)</div>
                  <img id="preview-glass" class="preview-image" alt="Glass image" />
                  <div id="preview-glass-empty" class="preview-empty">No glass image</div>
                </div>
              </div>
            </div>
            """
        graph_section = ""
        if has_graphs:
            graph_section = """
            <div class="preview-section">
              <div class="preview-title">Annealing graphs</div>
              <div class="preview-grid">
                <div class="preview-card">
                  <div class="preview-label">1000 mA</div>
                  <img id="preview-high" class="preview-image" alt="1000 mA graph" />
                  <div id="preview-high-empty" class="preview-empty">No 1000 mA graph</div>
                </div>
                <div class="preview-card">
                  <div class="preview-label">Low mA</div>
                  <img id="preview-low" class="preview-image" alt="Low mA graph" />
                  <div id="preview-low-empty" class="preview-empty">No low mA graph</div>
                </div>
                <div class="preview-card">
                  <div class="preview-label">Other mA</div>
                  <div id="preview-other" class="preview-stack"></div>
                  <div id="preview-other-empty" class="preview-empty">No other mA graphs</div>
                </div>
              </div>
            </div>
            """
        vsm_hyst_section = ""
        if has_vsm_hyst:
            vsm_hyst_section = """
            <div class="preview-section">
              <div class="preview-title">VSM hysteresis</div>
              <div id="preview-vsm-hyst" class="preview-stack"></div>
              <div id="preview-vsm-hyst-empty" class="preview-empty">No VSM hysteresis graphs</div>
            </div>
            """
        vsm_temp_section = ""
        if has_vsm_temp:
            vsm_temp_section = """
            <div class="preview-section">
              <div class="preview-title">VSM temperature scan</div>
              <div id="preview-vsm-temp" class="preview-stack"></div>
              <div id="preview-vsm-temp-empty" class="preview-empty">No VSM temperature scans</div>
            </div>
            """
        dma_section = ""
        if has_dma:
            dma_section = """
            <div class="preview-section">
              <div class="preview-title">DMA iso-stress</div>
              <div id="preview-dma" class="preview-stack"></div>
              <div id="preview-dma-empty" class="preview-empty">No DMA iso-stress graphs</div>
            </div>
            """
        shape_memory_section = ""
        if has_shape_memory:
            shape_memory_section = """
            <div class="preview-section">
              <div class="preview-title">Shape memory stress/strain</div>
              <div id="preview-shape-memory" class="preview-stack"></div>
              <div id="preview-shape-memory-empty" class="preview-empty">No shape-memory graphs</div>
            </div>
            """
        fmr_section = ""
        if has_fmr:
            fmr_section = """
            <div class="preview-section">
              <div class="preview-title">FMR</div>
              <div id="preview-fmr" class="preview-stack"></div>
              <div id="preview-fmr-empty" class="preview-empty">No FMR graphs</div>
            </div>
            """
        if (
            not graph_section
            and not microscope_section
            and not vsm_hyst_section
            and not vsm_temp_section
            and not dma_section
            and not shape_memory_section
            and not fmr_section
        ):
            preview_classes = f"{preview_class} empty"

        html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title_text}</title>
  <style>
    :root {{
      color-scheme: light dark;
    }}
    body {{
      font-family: "Segoe UI", "Calibri", "Arial", sans-serif;
      margin: 0;
      background: #141414;
      color: #f5f5f5;
    }}
    .toolbar {{
      padding: 16px 24px;
      border-bottom: 1px solid #2a2a2a;
      background: linear-gradient(135deg, #1f1f1f, #0f0f0f);
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
    }}
    .toolbar-actions {{
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }}
    .toolbar-actions input[type="search"] {{
      min-width: 220px;
      background: #0f0f0f;
      color: #f5f5f5;
      border: 1px solid #333;
      border-radius: 6px;
      padding: 6px 10px;
      font-size: 12px;
    }}
    .toolbar-actions input[type="search"]::placeholder {{
      color: #8f8f8f;
    }}
    .toolbar button {{
      background: #222;
      color: #f5f5f5;
      border: 1px solid #333;
      border-radius: 6px;
      padding: 6px 10px;
      font-size: 12px;
      cursor: pointer;
    }}
    .toolbar button:hover {{
      border-color: #4a4a4a;
    }}
    .title {{
      font-size: 20px;
      font-weight: 600;
    }}
    .meta {{
      font-size: 12px;
      color: #b0b0b0;
    }}
    .content {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) 420px;
      gap: 16px;
      padding: 16px 24px 24px;
    }}
    .compare-panel {{
      margin: 0 24px 24px;
      border: 1px solid #2a2a2a;
      border-radius: 8px;
      background: #101010;
      padding: 12px;
      display: flex;
      flex-direction: column;
      gap: 10px;
    }}
    .compare-panel.hidden {{
      display: none;
    }}
    .compare-header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
    }}
    .compare-title {{
      font-size: 14px;
      font-weight: 600;
    }}
    .compare-help {{
      font-size: 12px;
      color: #9b9b9b;
    }}
    .compare-table-container {{
      overflow: auto;
      border: 1px solid #2a2a2a;
      border-radius: 8px;
      background: #0f0f0f;
    }}
    .compare-table {{
      width: 100%;
      border-collapse: collapse;
    }}
    .compare-table th,
    .compare-table td {{
      padding: 8px 10px;
      border-bottom: 1px solid #222;
      font-size: 12px;
      text-align: left;
      vertical-align: top;
      white-space: nowrap;
    }}
    .table-container {{
      overflow: auto;
      border: 1px solid #2a2a2a;
      border-radius: 8px;
      background: #0f0f0f;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
    }}
    th, td {{
      padding: 8px 10px;
      border-bottom: 1px solid #222;
      font-size: 12px;
      text-align: left;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #1c1c1c;
      cursor: pointer;
    }}
    tr:hover {{
      background: rgba(255, 255, 255, 0.05);
    }}
    tr.active {{
      background: rgba(0, 145, 255, 0.2);
    }}
    tr.compare {{
      background: rgba(0, 145, 255, 0.35);
    }}
    .preview-panel {{
      border: 1px solid #2a2a2a;
      border-radius: 8px;
      padding: 12px;
      background: #101010;
      display: flex;
      flex-direction: column;
      gap: 16px;
    }}
    .preview-panel.empty {{
      color: #9b9b9b;
    }}
    .preview-title {{
      font-size: 14px;
      font-weight: 600;
      margin-bottom: 6px;
    }}
    .preview-grid {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 12px;
    }}
    .preview-stack {{
      display: flex;
      flex-direction: column;
      gap: 10px;
    }}
    .preview-card {{
      background: #161616;
      border: 1px solid #242424;
      border-radius: 6px;
      padding: 8px;
    }}
    .preview-label {{
      font-size: 12px;
      color: #b5b5b5;
      margin-bottom: 6px;
    }}
    .preview-image {{
      width: 100%;
      height: auto;
      display: none;
    }}
    .preview-empty {{
      font-size: 12px;
      color: #6f6f6f;
      display: none;
    }}
    @media (max-width: 1100px) {{
      .content {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <div class="toolbar">
    <div>
      <div class="title">{title_text}</div>
      <div id="row-meta" class="meta">{len(frame)} row(s)</div>
    </div>
    <div class="toolbar-actions">
      <input id="table-search" type="search" placeholder="Search rows..." />
      <button id="search-clear" type="button">Clear search</button>
      <button id="compare-clear" type="button">Clear compare</button>
      <div id="compare-count" class="meta">0 selected</div>
    </div>
  </div>
    <div class="content">
    <div class="table-container">
      <table id="data-table">
        <thead>
          <tr>{header_cells}</tr>
        </thead>
        <tbody>
          {''.join(rows_html)}
        </tbody>
      </table>
    </div>
    <div class="{preview_classes}">
      {graph_section}
      {microscope_section}
      {vsm_hyst_section}
      {vsm_temp_section}
      {dma_section}
      {shape_memory_section}
      {fmr_section}
    </div>
  </div>
  <div id="compare-panel" class="compare-panel hidden">
    <div class="compare-header">
      <div class="compare-title">Compare</div>
      <div class="compare-help">Ctrl/Cmd + click rows to compare</div>
    </div>
    <div class="compare-table-container" id="compare-table"></div>
  </div>
  <script>
    const table = document.getElementById('data-table');
    const rows = Array.from(table.querySelectorAll('tbody tr'));
    const preview = {{
      high: document.getElementById('preview-high'),
      highEmpty: document.getElementById('preview-high-empty'),
      low: document.getElementById('preview-low'),
      lowEmpty: document.getElementById('preview-low-empty'),
      other: document.getElementById('preview-other'),
      otherEmpty: document.getElementById('preview-other-empty'),
      core: document.getElementById('preview-core'),
      coreEmpty: document.getElementById('preview-core-empty'),
      glass: document.getElementById('preview-glass'),
      glassEmpty: document.getElementById('preview-glass-empty'),
      vsmHyst: document.getElementById('preview-vsm-hyst'),
      vsmHystEmpty: document.getElementById('preview-vsm-hyst-empty'),
      vsmTemp: document.getElementById('preview-vsm-temp'),
      vsmTempEmpty: document.getElementById('preview-vsm-temp-empty'),
      dma: document.getElementById('preview-dma'),
      dmaEmpty: document.getElementById('preview-dma-empty'),
      shapeMemory: document.getElementById('preview-shape-memory'),
      shapeMemoryEmpty: document.getElementById('preview-shape-memory-empty'),
      fmr: document.getElementById('preview-fmr'),
      fmrEmpty: document.getElementById('preview-fmr-empty'),
    }};
    const comparePanel = document.getElementById('compare-panel');
    const compareTable = document.getElementById('compare-table');
    const compareCount = document.getElementById('compare-count');
    const compareClear = document.getElementById('compare-clear');
    const rowMeta = document.getElementById('row-meta');
    const searchInput = document.getElementById('table-search');
    const searchClear = document.getElementById('search-clear');

    function isRowVisible(row) {{
      return row.style.display !== 'none';
    }}

    function updateRowMeta(visibleCount, totalCount, hasFilter) {{
      if (!rowMeta) {{
        return;
      }}
      if (hasFilter) {{
        rowMeta.textContent = `${{visibleCount}} of ${{totalCount}} row(s)`;
      }} else {{
        rowMeta.textContent = `${{totalCount}} row(s)`;
      }}
    }}

    function setImage(imgEl, emptyEl, data) {{
      if (!imgEl || !emptyEl) {{
        return;
      }}
      if (data) {{
        imgEl.src = data;
        imgEl.style.display = 'block';
        emptyEl.style.display = 'none';
      }} else {{
        imgEl.removeAttribute('src');
        imgEl.style.display = 'none';
        emptyEl.style.display = 'block';
      }}
    }}

    function setImageList(container, emptyEl, data) {{
      if (!container || !emptyEl) {{
        return;
      }}
      const items = data ? data.split('|').filter(Boolean) : [];
      container.innerHTML = '';
      if (items.length) {{
        items.forEach((uri) => {{
          const img = document.createElement('img');
          img.src = uri;
          img.className = 'preview-image';
          img.style.display = 'block';
          img.alt = 'Graph preview';
          container.appendChild(img);
        }});
        container.style.display = 'flex';
        emptyEl.style.display = 'none';
      }} else {{
        container.style.display = 'none';
        emptyEl.style.display = 'block';
      }}
    }}

    function updatePreview(row) {{
      setImage(preview.high, preview.highEmpty, row ? row.dataset.high : '');
      setImage(preview.low, preview.lowEmpty, row ? row.dataset.low : '');
      setImageList(preview.other, preview.otherEmpty, row ? row.dataset.other : '');
      setImage(preview.core, preview.coreEmpty, row ? row.dataset.core : '');
      setImage(preview.glass, preview.glassEmpty, row ? row.dataset.glass : '');
      setImageList(preview.vsmHyst, preview.vsmHystEmpty, row ? row.dataset.vsmHyst : '');
      setImageList(preview.vsmTemp, preview.vsmTempEmpty, row ? row.dataset.vsmTemp : '');
      setImageList(preview.dma, preview.dmaEmpty, row ? row.dataset.dma : '');
      setImageList(preview.shapeMemory, preview.shapeMemoryEmpty, row ? row.dataset.shapeMemory : '');
      setImageList(preview.fmr, preview.fmrEmpty, row ? row.dataset.fmr : '');
    }}

    function rowLabel(row, headers) {{
      const compositionIndex = headers.indexOf('Composition');
      const microwireIndex = headers.indexOf('Microwire');
      const cells = row.children;
      const composition = compositionIndex >= 0 && cells[compositionIndex]
        ? cells[compositionIndex].innerText.trim()
        : '';
      const microwire = microwireIndex >= 0 && cells[microwireIndex]
        ? cells[microwireIndex].innerText.trim()
        : '';
      const combined = `${{composition}} ${{microwire}}`.trim();
      return combined || `Row ${{row.rowIndex}}`;
    }}

    function graphDataForField(field, row) {{
      if (field === 'Figure — 1000 mA') {{
        return row.dataset.high || '';
      }}
      if (field === 'Figure — low mA') {{
        return row.dataset.low || '';
      }}
      if (field === 'Figure — other mA') {{
        return row.dataset.other || '';
      }}
      if (field === 'VSM hysteresis graphs') {{
        return row.dataset.vsmHyst || '';
      }}
      if (field === 'VSM temperature scan graphs') {{
        return row.dataset.vsmTemp || '';
      }}
      if (field === 'DMA iso-stress graphs') {{
        return row.dataset.dma || '';
      }}
      if (field === 'Shape memory stress/strain graphs') {{
        return row.dataset.shapeMemory || '';
      }}
      if (field === 'FMR graphs') {{
        return row.dataset.fmr || '';
      }}
      return '';
    }}

    function renderGraphCell(cell, data) {{
      const items = data ? data.split('|').filter(Boolean) : [];
      if (!items.length) {{
        cell.textContent = '';
        return;
      }}
      const stack = document.createElement('div');
      stack.className = 'preview-stack';
      items.forEach((uri) => {{
        const img = document.createElement('img');
        img.src = uri;
        img.className = 'preview-image';
        img.style.display = 'block';
        img.alt = 'Graph preview';
        stack.appendChild(img);
      }});
      cell.appendChild(stack);
    }}

    function updateCompare() {{
      if (!comparePanel || !compareTable) {{
        return;
      }}
      const selected = rows.filter(
        (row) => row.classList.contains('compare') && isRowVisible(row)
      );
      if (compareCount) {{
        compareCount.textContent = `${{selected.length}} selected`;
      }}
      if (selected.length < 2) {{
        comparePanel.classList.add('hidden');
        compareTable.innerHTML = '';
        return;
      }}
      const headers = Array.from(table.querySelectorAll('thead th')).map((th) =>
        th.innerText.trim()
      );
      const graphColumns = new Set([
        'Figure — 1000 mA',
        'Figure — low mA',
        'Figure — other mA',
        'VSM hysteresis graphs',
        'VSM temperature scan graphs',
        'DMA iso-stress graphs',
        'Shape memory stress/strain graphs',
        'FMR graphs',
      ]);
      const compare = document.createElement('table');
      compare.className = 'compare-table';
      const headerRow = document.createElement('tr');
      headerRow.appendChild(document.createElement('th')).innerText = 'Field';
      selected.forEach((row) => {{
        const th = document.createElement('th');
        th.innerText = rowLabel(row, headers);
        headerRow.appendChild(th);
      }});
      const thead = document.createElement('thead');
      thead.appendChild(headerRow);
      compare.appendChild(thead);
      const tbody = document.createElement('tbody');
      headers.forEach((field, index) => {{
        if (!field) {{
          return;
        }}
        const rowEl = document.createElement('tr');
        const fieldCell = document.createElement('td');
        fieldCell.innerText = field;
        rowEl.appendChild(fieldCell);
        selected.forEach((row) => {{
          const cell = document.createElement('td');
          if (graphColumns.has(field)) {{
            const data = graphDataForField(field, row);
            renderGraphCell(cell, data);
          }} else {{
            const cellValue = row.children[index]
              ? row.children[index].innerText.trim()
              : '';
            cell.innerText = cellValue;
          }}
          rowEl.appendChild(cell);
        }});
        tbody.appendChild(rowEl);
      }});
      compare.appendChild(tbody);
      compareTable.innerHTML = '';
      compareTable.appendChild(compare);
      comparePanel.classList.remove('hidden');
    }}

    function ensureActiveVisibleRow() {{
      const active = rows.find(
        (row) => row.classList.contains('active') && isRowVisible(row)
      );
      if (active) {{
        updatePreview(active);
        return;
      }}
      rows.forEach((row) => row.classList.remove('active'));
      const firstVisible = rows.find((row) => isRowVisible(row));
      if (firstVisible) {{
        firstVisible.classList.add('active');
        updatePreview(firstVisible);
        return;
      }}
      updatePreview(null);
    }}

    function applySearchFilter() {{
      const query = searchInput ? searchInput.value.trim().toLowerCase() : '';
      let visibleCount = 0;
      rows.forEach((row) => {{
        const text = row.innerText.toLowerCase();
        const visible = !query || text.includes(query);
        row.style.display = visible ? '' : 'none';
        if (!visible) {{
          row.classList.remove('compare');
        }} else {{
          visibleCount += 1;
        }}
      }});
      ensureActiveVisibleRow();
      updateRowMeta(visibleCount, rows.length, Boolean(query));
      updateCompare();
    }}

    rows.forEach((row) => {{
      row.addEventListener('click', (event) => {{
        rows.forEach((item) => item.classList.remove('active'));
        row.classList.add('active');
        if (event.ctrlKey || event.metaKey) {{
          row.classList.toggle('compare');
        }}
        updatePreview(row);
        updateCompare();
      }});
    }});

    if (rows.length > 0) {{
      rows[0].classList.add('active');
      updatePreview(rows[0]);
    }} else {{
      updatePreview(null);
    }}
    updateCompare();
    applySearchFilter();

    if (compareClear) {{
      compareClear.addEventListener('click', () => {{
        rows.forEach((row) => row.classList.remove('compare'));
        updateCompare();
      }});
    }}
    if (searchInput) {{
      searchInput.addEventListener('input', () => {{
        applySearchFilter();
      }});
    }}
    if (searchClear) {{
      searchClear.addEventListener('click', () => {{
        if (searchInput) {{
          searchInput.value = '';
          searchInput.focus();
        }}
        applySearchFilter();
      }});
    }}

    const headers = Array.from(table.querySelectorAll('th'));
    let sortState = {{}};
    headers.forEach((header, index) => {{
      header.addEventListener('click', () => {{
        const dir = sortState[index] === 'asc' ? 'desc' : 'asc';
        sortState = {{ [index]: dir }};
        const body = table.tBodies[0];
        const sorted = rows.slice().sort((a, b) => {{
          const aText = a.children[index].innerText.trim();
          const bText = b.children[index].innerText.trim();
          const aNum = parseFloat(aText.replace(/,/g, ''));
          const bNum = parseFloat(bText.replace(/,/g, ''));
          if (!Number.isNaN(aNum) && !Number.isNaN(bNum)) {{
            return dir === 'asc' ? aNum - bNum : bNum - aNum;
          }}
          return dir === 'asc'
            ? aText.localeCompare(bText)
            : bText.localeCompare(aText);
        }});
        sorted.forEach((row) => body.appendChild(row));
        applySearchFilter();
      }});
    }});
  </script>
</body>
</html>
"""
        try:
            html_path.write_text(html_text, encoding="utf-8")
        except Exception as exc:
            self.logger.exception("Failed to export HTML")
            QtWidgets.QMessageBox.critical(
                self,
                "Export HTML",
                f"Failed to export HTML:\n{exc}",
            )
            return None
        self.log(f"HTML export saved to {html_path}")
        return html_path

    def _combine(self) -> None:
        if self._combine_thread is not None and self._combine_thread.isRunning():
            return
        selected = self._selected_sections()
        inputs = self._prepare_builder_inputs(selected, require_payloads=False)
        if inputs is None:
            return

        (
            fabrication_index,
            annealing_records,
            vsm_hysteresis_records,
            vsm_temperature_records,
            dma_isostress_records,
            shape_memory_stress_strain_records,
            shape_memory_entries,
            fmr_records,
            microscope_index,
            video_index,
            strain_records,
            strain_entries,
            current_density_entries,
            overrides,
            phase_points,
            transition_points,
            video_overrides,
        ) = inputs

        if "microscope" in selected and overrides:
            microscope_index = _apply_microscope_overrides(microscope_index, overrides)
        else:
            microscope_index = {} if "microscope" not in selected else microscope_index

        output_dir = Path(self._output_dir or Path.cwd())
        output_dir.mkdir(parents=True, exist_ok=True)
        output_name = _normalise_output_name(self._output_name or DEFAULT_OUTPUT_NAME)

        formats: List[str] = []
        if self._export_csv:
            formats.append("csv")
        if self._export_excel:
            formats.append("excel")
        backends: List[str] = []
        if self._export_matplotlib:
            backends.append("matplotlib")
        if self._export_origin:
            backends.append("origin")
        column_filter = self._current_column_filter()
        column_order = self._current_column_order_for_export()
        sort_spec = self._current_sort_spec()

        config = BuilderConfig(
            annealing_files=[],
            fabrication_files=[],
            output_dir=output_dir,
            microscope_files=[],
            video_files=[],
            strain_files=[],
            make_plots=bool(self._export_matplotlib or self._export_origin),
            export_formats=tuple(formats),
            plot_backends=tuple(backends),
            output_name=output_name,
            column_filter=column_filter,
            column_order=column_order,
            sort_spec=sort_spec,
        )
        build_kwargs = {
            "fabrication_index": fabrication_index,
            "measurement_records": annealing_records,
            "vsm_hysteresis_records": (
                vsm_hysteresis_records if "vsm_hysteresis" in selected else []
            ),
            "vsm_temperature_scan_records": (
                vsm_temperature_records if "vsm_temperature_scan" in selected else []
            ),
            "dma_iso_stress_records": (
                dma_isostress_records if "dma_iso_stress" in selected else []
            ),
            "shape_memory_stress_strain_records": (
                shape_memory_stress_strain_records
                if "shape_memory_stress_strain" in selected
                else []
            ),
            "shape_memory_entries": (
                shape_memory_entries
                if "shape_memory_stress_strain" in selected
                else {}
            ),
            "fmr_records": fmr_records if "fmr" in selected else [],
            "microscope_index": microscope_index if "microscope" in selected else {},
            "video_index": video_index if "videos" in selected else {},
            "video_overrides": video_overrides,
            "strain_records": strain_records if "strain" in selected else {},
            "strain_entries": strain_entries if "strain" in selected else {},
            "current_density_entries": (
                current_density_entries if "current_density" in selected else {}
            ),
            "phase_points": phase_points,
            "transition_temps": transition_points,
        }

        self._combine_output_dir = output_dir
        self._combine_output_name = output_name
        self._open_combine_progress()
        self.export_button.setEnabled(False)

        self._combine_thread = QtCore.QThread(self)
        self._combine_worker = CombineWorker(config, build_kwargs, self.logger)
        self._combine_worker.moveToThread(self._combine_thread)
        self._combine_thread.started.connect(self._combine_worker.run)
        self._combine_worker.finished.connect(self._handle_combine_finished)
        self._combine_worker.failed.connect(self._handle_combine_failed)
        self._combine_worker.finished.connect(self._combine_thread.quit)
        self._combine_worker.failed.connect(self._combine_thread.quit)
        self._combine_thread.finished.connect(self._cleanup_combine_thread)
        self._combine_thread.start()
        self.log("Combine build started.")

    def _preview(self) -> None:
        if self._preview_thread is not None and self._preview_thread.isRunning():
            return
        if hasattr(self, "preview_button"):
            self.preview_button.setEnabled(False)
        self._open_preview_progress()
        selected = self._selected_sections()
        try:
            inputs = self._prepare_builder_inputs(selected, require_payloads=False)
        except Exception as exc:
            self.logger.exception("Preview input preparation failed")
            self._close_preview_progress()
            QtWidgets.QMessageBox.critical(
                self,
                "Microwire Data Builder",
                f"Failed to prepare preview inputs:\n{exc}",
            )
            return
        if inputs is None:
            self._close_preview_progress()
            return

        (
            fabrication_index,
            annealing_records,
            vsm_hysteresis_records,
            vsm_temperature_records,
            dma_isostress_records,
            shape_memory_stress_strain_records,
            shape_memory_entries,
            fmr_records,
            microscope_index,
            video_index,
            strain_records,
            strain_entries,
            current_density_entries,
            overrides,
            phase_points,
            transition_points,
            video_overrides,
        ) = inputs

        if "microscope" in selected and overrides:
            microscope_index = _apply_microscope_overrides(microscope_index, overrides)
        else:
            microscope_index = {} if "microscope" not in selected else microscope_index

        output_dir = Path(self._output_dir or Path.cwd())
        output_dir.mkdir(parents=True, exist_ok=True)
        output_name = _normalise_output_name(self._output_name or DEFAULT_OUTPUT_NAME)

        config = BuilderConfig(
            annealing_files=[],
            fabrication_files=[],
            output_dir=output_dir,
            microscope_files=[],
            video_files=[],
            strain_files=[],
            make_plots=False,
            export_formats=(),
            plot_backends=(),
            output_name=output_name,
        )

        build_kwargs = {
            "fabrication_index": fabrication_index,
            "measurement_records": annealing_records,
            "vsm_hysteresis_records": (
                vsm_hysteresis_records if "vsm_hysteresis" in selected else []
            ),
            "vsm_temperature_scan_records": (
                vsm_temperature_records if "vsm_temperature_scan" in selected else []
            ),
            "dma_iso_stress_records": (
                dma_isostress_records if "dma_iso_stress" in selected else []
            ),
            "shape_memory_stress_strain_records": (
                shape_memory_stress_strain_records
                if "shape_memory_stress_strain" in selected
                else []
            ),
            "shape_memory_entries": (
                shape_memory_entries
                if "shape_memory_stress_strain" in selected
                else {}
            ),
            "fmr_records": fmr_records if "fmr" in selected else [],
            "microscope_index": microscope_index if "microscope" in selected else {},
            "video_index": video_index if "videos" in selected else {},
            "video_overrides": video_overrides,
            "strain_records": strain_records if "strain" in selected else {},
            "strain_entries": strain_entries if "strain" in selected else {},
            "current_density_entries": (
                current_density_entries if "current_density" in selected else {}
            ),
            "phase_points": phase_points,
            "transition_temps": transition_points,
            "skip_exports": True,
        }
        self._preview_thread = QtCore.QThread(self)
        self._preview_worker = PreviewWorker(config, build_kwargs, self.logger)
        self._preview_worker.moveToThread(self._preview_thread)
        self._preview_thread.started.connect(self._preview_worker.run)
        self._preview_worker.finished.connect(self._handle_preview_finished)
        self._preview_worker.failed.connect(self._handle_preview_failed)
        self._preview_worker.finished.connect(self._preview_thread.quit)
        self._preview_worker.failed.connect(self._preview_thread.quit)
        self._preview_thread.finished.connect(self._cleanup_preview_thread)
        self._preview_thread.start()
        self.log("Preview build started.")

    def _open_preview_progress(self) -> None:
        if self._preview_dialog is not None:
            try:
                self._preview_dialog.close()
            except Exception:
                pass
        try:
            dialog = QtWidgets.QProgressDialog(
                "Building preview database...",
                None,
                0,
                0,
                self,
            )
            dialog.setWindowTitle("Preview database")
            dialog.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
            dialog.setCancelButton(None)
            dialog.setMinimumDuration(0)
            dialog.setAutoClose(False)
            dialog.setAutoReset(False)
            dialog.show()
            QtWidgets.QApplication.processEvents(
                QtCore.QEventLoop.ProcessEventsFlag.AllEvents
            )
            self._preview_dialog = dialog
        except Exception:
            self._preview_dialog = None

    def _close_preview_progress(self) -> None:
        if hasattr(self, "preview_button"):
            self.preview_button.setEnabled(True)
        dialog = self._preview_dialog
        self._preview_dialog = None
        if dialog is not None:
            try:
                dialog.close()
            except Exception:
                try:
                    dialog.cancel()
                except Exception:
                    pass

    def _handle_preview_finished(self, dataframe: object) -> None:
        self._close_preview_progress()
        if isinstance(dataframe, pd.DataFrame):
            self._measured_preview_frame = dataframe.copy()
            merged = self._merge_imported_rows(dataframe)
            self._update_preview(merged)
        else:
            self._update_preview(pd.DataFrame())
        self.log("Preview updated.")

    def _handle_preview_failed(self, message: str) -> None:
        self._close_preview_progress()
        self.logger.error("Preview failed: %s", message)
        QtWidgets.QMessageBox.critical(
            self,
            "Microwire Data Builder",
            f"Failed to preview database:\n{message}",
        )

    def _cleanup_preview_thread(self) -> None:
        if self._preview_worker is not None:
            try:
                self._preview_worker.deleteLater()
            except Exception:
                pass
        if self._preview_thread is not None:
            try:
                self._preview_thread.deleteLater()
            except Exception:
                pass
        self._preview_worker = None
        self._preview_thread = None

    def _open_combine_progress(self) -> None:
        if self._combine_dialog is not None:
            try:
                self._combine_dialog.close()
            except Exception:
                pass
        try:
            dialog = QtWidgets.QProgressDialog(
                "Exporting database...",
                None,
                0,
                0,
                self,
            )
            dialog.setWindowTitle("Export database")
            dialog.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
            dialog.setCancelButton(None)
            dialog.setMinimumDuration(0)
            dialog.setAutoClose(False)
            dialog.setAutoReset(False)
            dialog.show()
            QtWidgets.QApplication.processEvents(
                QtCore.QEventLoop.ProcessEventsFlag.AllEvents
            )
            self._combine_dialog = dialog
        except Exception:
            self._combine_dialog = None

    def _close_combine_progress(self) -> None:
        self.export_button.setEnabled(True)
        dialog = self._combine_dialog
        self._combine_dialog = None
        if dialog is not None:
            try:
                dialog.close()
            except Exception:
                try:
                    dialog.cancel()
                except Exception:
                    pass

    def _handle_combine_finished(self, result: object) -> None:
        self._close_combine_progress()
        if not isinstance(result, BuildResult):
            self.logger.error("Combine finished with unexpected result type: %s", type(result))
            return
        exports: Dict[str, Path] = dict(result.exports or {})
        self._update_preview(result.dataframe)
        export_frame = self._preview_export_frame()
        output_dir = self._combine_output_dir or Path(self._output_dir or Path.cwd())
        output_name = self._combine_output_name or _normalise_output_name(
            self._output_name or DEFAULT_OUTPUT_NAME
        )
        csv_path = exports.get("csv")
        if csv_path is not None:
            try:
                export_frame.to_csv(csv_path, index=False)
            except Exception:
                self.logger.exception("Failed to rewrite CSV export from Assemble preview")
        excel_path = exports.get("excel")
        if excel_path is not None:
            try:
                export_frame.to_excel(excel_path, index=False)
            except Exception:
                self.logger.exception("Failed to rewrite Excel export from Assemble preview")
        if self._export_html:
            html_path = self._export_html_file(export_frame, output_dir, output_name)
            if html_path is not None:
                exports["html"] = html_path
        if exports:
            lines = [f"{fmt.upper()}: {path}" for fmt, path in exports.items()]
            export_text = "\n".join(lines)
        else:
            export_text = "No export files were created."
        QtWidgets.QMessageBox.information(
            self,
            "Microwire Data Builder",
            f"Export finished successfully.\n\n{export_text}",
        )
        self.log("Combine export finished.")

    def _handle_combine_failed(self, message: str) -> None:
        self._close_combine_progress()
        self.logger.error("Combine failed: %s", message)
        QtWidgets.QMessageBox.critical(
            self,
            "Microwire Data Builder",
            f"Failed to export database:\n{message}",
        )

    def _cleanup_combine_thread(self) -> None:
        if self._combine_worker is not None:
            try:
                self._combine_worker.deleteLater()
            except Exception:
                pass
        if self._combine_thread is not None:
            try:
                self._combine_thread.deleteLater()
            except Exception:
                pass
        self._combine_worker = None
        self._combine_thread = None


class BuilderWindow(QtWidgets.QMainWindow):
    """New workbench for preparing and assembling microwire databases."""

    PROJECT_EXTENSION = ".pydpj"
    PROJECT_VERSION = 1
    PROJECT_KIND = "MicrowireDataBuilder"

    def __init__(self) -> None:
        super().__init__()
        self.logger = logging.getLogger(LOGGER_NAME)
        self._base_title = "Microwire Data Builder"
        self.setWindowTitle(self._base_title)
        self.resize(1100, 720)

        self._project_path: Optional[Path] = None
        self._save_project_action: QtGui.QAction | None = None
        self._save_project_as_action: QtGui.QAction | None = None
        downloads_dir = Path.home() / "Downloads"
        self._default_output_dir = (
            downloads_dir if downloads_dir.exists() and downloads_dir.is_dir() else Path.cwd()
        )
        self._last_output_dir = str(self._default_output_dir)
        self.settings = _builder_settings()
        self._clamp_active = False
        self._recent_projects: List[str] = []
        self._recent_projects_menu: QtWidgets.QMenu | None = None
        self._load_recent_projects_setting()
        raw_auto = self.settings.value(self._project_settings_key("auto_open_last"), False)
        self._auto_open_last: bool = bool(raw_auto)
        self._auto_open_last_action: QtGui.QAction | None = None
        self._data_menu: QtWidgets.QMenu | None = None
        self._show_imported_action: QtGui.QAction | None = None
        self._separate_imported_action: QtGui.QAction | None = None
        self._remove_imported_action: QtGui.QAction | None = None
        self._imported_item: QtWidgets.QTreeWidgetItem | None = None
        self._fullscreen_snap_pending = False

        central = QtWidgets.QWidget(self)
        layout = QtWidgets.QVBoxLayout(central)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.tab_widget = QtWidgets.QTabWidget(central)
        layout.addWidget(self.tab_widget, 1)

        self.setCentralWidget(central)

        self._primary_dock_widths: Dict[QtWidgets.QDockWidget, int] = {}
        self._retabbing_docks = False
        self._retabify_pending = False
        self._dirty = False
        self._suppress_dirty = False

        self.log_view = QtWidgets.QPlainTextEdit(self)
        self.log_view.setReadOnly(True)
        self.log_view.setMaximumBlockCount(2000)
        self.log_view.setObjectName("builderMessageLogView")

        self.sections: Dict[str, QtWidgets.QWidget] = {}
        self._log_has_unread_errors = False
        self._log_highlight_active = False
        self._log_capture_enabled = False
        self._log_capture_path: Optional[Path] = None
        self._log_capture_faulted = False
        self._crash_log_handle: Optional[io.TextIOWrapper] = None
        self._crash_handlers_installed = False

        def _append_log(level: int, message: str) -> None:
            self.log_view.appendPlainText(message)
            scrollbar = self.log_view.verticalScrollBar()
            if scrollbar is not None:
                scrollbar.setValue(scrollbar.maximum())
            self._append_log_to_file(level, message)
            if level >= logging.ERROR:
                self._log_has_unread_errors = True
            self._update_log_highlight()

        self._log_handler = QtLogHandler(_append_log)
        self._log_handler.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
        if self._log_handler not in self.logger.handlers:
            self.logger.addHandler(self._log_handler)
        self._install_crash_handlers()

        def _pump_events() -> None:
            try:
                QtWidgets.QApplication.processEvents(
                    QtCore.QEventLoop.ProcessEventsFlag.AllEvents
                )
            except Exception:
                pass

        self.annealing_section = AnnealingSection(self.logger, _append_log)
        self.tab_widget.addTab(self.annealing_section, "Current annealing")
        self.sections["annealing"] = self.annealing_section
        _pump_events()

        self.fabrication_section = FabricationSection(self.logger, _append_log)
        self.tab_widget.addTab(self.fabrication_section, "Fabrication")
        self.sections["fabrication"] = self.fabrication_section
        _pump_events()

        self.microscope_section = MicroscopeSection(self.logger, _append_log)
        self.tab_widget.addTab(self.microscope_section, "Microscope")
        self.sections["microscope"] = self.microscope_section
        _pump_events()

        self.current_density_section = CurrentDensitySection(
            self.annealing_section,
            self.microscope_section,
            self.logger,
            _append_log,
        )
        self.tab_widget.addTab(self.current_density_section, "Current density")
        self.sections["current_density"] = self.current_density_section
        _pump_events()

        self.video_section = VideoSection(self.logger, _append_log)
        self.tab_widget.addTab(self.video_section, "Videos")
        self.sections["videos"] = self.video_section
        _pump_events()

        self.vsm_hysteresis_section = VsmHysteresisSection(self.logger, _append_log)
        self.tab_widget.addTab(self.vsm_hysteresis_section, "VSM hysteresis")
        self.sections["vsm_hysteresis"] = self.vsm_hysteresis_section
        _pump_events()

        self.vsm_temperature_section = VsmTemperatureScanSection(self.logger, _append_log)
        self.tab_widget.addTab(self.vsm_temperature_section, "VSM temp scan")
        self.sections["vsm_temperature_scan"] = self.vsm_temperature_section
        _pump_events()

        self.transition_temps_section = TransitionTempsSection(
            self.vsm_temperature_section,
            self.logger,
            _append_log,
        )
        self.tab_widget.addTab(self.transition_temps_section, "Transition temps")
        self.sections["transition_temps"] = self.transition_temps_section
        _pump_events()

        self.dma_iso_stress_section = DmaIsoStressSection(self.logger, _append_log)
        self.tab_widget.addTab(self.dma_iso_stress_section, "DMA iso-stress")
        self.sections["dma_iso_stress"] = self.dma_iso_stress_section
        _pump_events()

        self.shape_memory_stress_strain_section = ShapeMemoryStressStrainSection(
            self.logger, _append_log
        )
        self.tab_widget.addTab(
            self.shape_memory_stress_strain_section,
            "Shape memory stress/strain",
        )
        self.sections["shape_memory_stress_strain"] = (
            self.shape_memory_stress_strain_section
        )
        _pump_events()

        self.fmr_section = FmrSection(self.logger, _append_log)
        self.tab_widget.addTab(self.fmr_section, "FMR")
        self.sections["fmr"] = self.fmr_section
        _pump_events()

        self._developer_options = developer_options()
        self._ocr_debug_supported = all(
            hasattr(self._developer_options, attr)
            for attr in ("ocr_debug", "ocr_debug_changed")
        )
        if self._ocr_debug_supported:
            try:
                self._developer_options.ocr_debug_changed.connect(
                    self._handle_ocr_debug_changed
                )
            except Exception:
                self._ocr_debug_supported = False
        initial_debug = False
        if self._ocr_debug_supported:
            try:
                initial_debug = bool(self._developer_options.ocr_debug())
            except Exception:
                initial_debug = False
        self._handle_ocr_debug_changed(initial_debug)
        if hasattr(self._developer_options, "message_log_capture_changed"):
            try:
                self._developer_options.message_log_capture_changed.connect(
                    self._handle_log_capture_changed
                )
            except Exception:
                pass
            try:
                initial_capture = bool(self._developer_options.capture_message_log())
            except Exception:
                initial_capture = False
            self._handle_log_capture_changed(initial_capture)

        self.strain_section = StrainSection(self.logger, _append_log)
        self.tab_widget.addTab(self.strain_section, "Strain")
        self.sections["strain"] = self.strain_section
        _pump_events()

        assembly = AssemblySection(
            self.sections,
            self.logger,
            _append_log,
        )
        self.assembly_section = assembly
        self.tab_widget.addTab(assembly, "Assemble")
        try:
            assembly.data_updated.connect(self._handle_section_data_updated)
        except Exception:
            pass
        _pump_events()

        self.compare_section = CompareSection(
            self.sections,
            self.logger,
            _append_log,
        )
        self.sections["compare"] = self.compare_section
        self.tab_widget.addTab(self.compare_section, "Compare")
        _pump_events()
        assembly.attach_compare_section(self.compare_section)

        self.fabrication_section.sources_changed.connect(
            self._handle_fabrication_sources_changed
        )
        self._handle_fabrication_sources_changed(self.fabrication_section.data.sources)
        try:
            self.fabrication_section.data_updated.connect(self.video_section.sync_with_fabrication)
        except Exception:
            pass
        try:
            if hasattr(self.annealing_section, "_sanitize_graph_columns"):
                self.annealing_section._sanitize_graph_columns()
            if hasattr(self.annealing_section, "_refresh_record_groups"):
                self.annealing_section._refresh_record_groups()
        except Exception:
            pass

        self.project_tree = QtWidgets.QTreeWidget()
        self.project_tree.setHeaderLabels(["Section", "Status / Source"])
        self.project_tree.header().setStretchLastSection(True)
        self._project_items: Dict[str, QtWidgets.QTreeWidgetItem] = {}
        for key, section in self.sections.items():
            status_text = section.status_label.text() if hasattr(section, "status_label") else ""
            item = QtWidgets.QTreeWidgetItem([section.section_title, status_text])
            self.project_tree.addTopLevelItem(item)
            self._project_items[key] = item
            section.status_changed.connect(partial(self._handle_section_status_changed, key))
            section.sources_changed.connect(partial(self._handle_section_sources_changed, key))
            try:
                section.data_updated.connect(self._handle_section_data_updated)
            except Exception:
                pass
            initial_sources: Iterable[str] = []
            if isinstance(section, MiniDatabaseSection):
                initial_sources = section.data.sources
            self._handle_section_sources_changed(key, initial_sources)

        self._imported_item = QtWidgets.QTreeWidgetItem(["Imported data", ""])
        self.project_tree.addTopLevelItem(self._imported_item)
        self._update_imported_data_item()

        self.project_dock = QtWidgets.QDockWidget("Project Explorer", self)
        self.project_dock.setObjectName("builderProjectExplorerDock")
        self.project_dock.setWidget(self.project_tree)
        self.addDockWidget(QtCore.Qt.DockWidgetArea.LeftDockWidgetArea, self.project_dock)
        self.project_dock.setMinimumWidth(260)
        self._primary_dock_widths[self.project_dock] = 320

        self.log_dock = QtWidgets.QDockWidget("Message Log", self)
        self.log_dock.setObjectName("builderMessageLogDock")
        self.log_dock.setWidget(self.log_view)
        self.addDockWidget(QtCore.Qt.DockWidgetArea.LeftDockWidgetArea, self.log_dock)
        self.log_dock.setMinimumWidth(260)
        self.log_dock.visibilityChanged.connect(self._handle_log_visibility)
        self.log_dock.hide()
        self._update_log_highlight()

        for dock in (self.project_dock, self.log_dock):
            if dock is None:
                continue
            dock.setAllowedAreas(QtCore.Qt.DockWidgetArea.AllDockWidgetAreas)
            dock.setFeatures(
                QtWidgets.QDockWidget.DockWidgetFeature.DockWidgetClosable
                | QtWidgets.QDockWidget.DockWidgetFeature.DockWidgetMovable
                | QtWidgets.QDockWidget.DockWidgetFeature.DockWidgetFloatable
            )
            dock.hide()

        self._dock_switcher_panels: List[QtWidgets.QDockWidget | None] = []
        if self._dock_switcher_supported():
            self._dock_switcher_panels.append(
                self._create_dock_switcher(
                    [self.project_dock, self.log_dock],
                    side="left",
                    initial_visible=(),
                )
            )
        else:
            self._dock_switcher_panels.append(None)

        for dock in (self.project_dock, self.log_dock):
            if dock is None:
                continue
            try:
                dock.dockLocationChanged.connect(
                    lambda area, d=dock: self._handle_primary_dock_location_change(d, area)
                )
            except Exception:
                pass
            try:
                dock.visibilityChanged.connect(
                    lambda _visible, d=dock: self._handle_primary_dock_visibility_changed(d)
                )
            except Exception:
                pass

        menu_bar = install_standard_menu(self, help_topic="builder_database", console=self.log_view)
        self._setup_project_actions(menu_bar)
        self._setup_settings_menu(menu_bar)
        self._setup_data_menu(menu_bar)
        self._update_project_actions()
        self._suppress_dirty = True
        for section in self.sections.values():
            if isinstance(section, MiniDatabaseSection):
                section.reset_to_blank()
        self._suppress_dirty = False
        self._dirty = False
        self._update_project_title()
        self._set_initial_geometry()
        self._retabify_primary_docks()
        QtCore.QTimer.singleShot(0, self._maybe_auto_open_last_project)

    def _dock_switcher_supported(self) -> bool:
        override = os.environ.get("MW_DISABLE_DOCK_SWITCHER", "")
        return override.strip().lower() not in {"1", "true", "yes", "on"}

    def _set_initial_geometry(self) -> None:
        """Resize and position the window within the visible screen."""

        try:
            screen = QtGui.QGuiApplication.screenAt(self.mapToGlobal(self.rect().center()))
            if screen is None:
                screen = QtGui.QGuiApplication.primaryScreen()
        except Exception:
            screen = QtGui.QGuiApplication.primaryScreen()
        available = screen.availableGeometry() if screen is not None else QtCore.QRect(0, 0, 1600, 900)
        max_w = max(1100, available.width() - 120)
        max_h = max(760, available.height() - 180)
        width = min(max_w, int(available.width() * 0.88))
        height = min(max_h, int(available.height() * 0.86))
        self.setMinimumSize(960, 640)
        self.resize(width, height)
        target_x = available.left() + max(0, (available.width() - width) // 2)
        target_y = available.top() + max(0, (available.height() - height) // 2)
        self.move(target_x, target_y)

    def _create_dock_switcher(
        self,
        docks: Sequence[QtWidgets.QDockWidget],
        *,
        side: str,
        initial_visible: Iterable[int] | None = None,
    ) -> QtWidgets.QDockWidget | None:
        if not docks:
            return None
        panel = QtWidgets.QDockWidget("", self)
        panel.setObjectName(f"builder_{side}_dock_switcher")
        panel.setAllowedAreas(
            QtCore.Qt.DockWidgetArea.LeftDockWidgetArea
            if side == "left"
            else QtCore.Qt.DockWidgetArea.RightDockWidgetArea
        )
        panel.setFeatures(QtWidgets.QDockWidget.DockWidgetFeature.NoDockWidgetFeatures)
        panel.setTitleBarWidget(QtWidgets.QWidget(panel))

        switcher = _DockSwitcherWidget(docks, side=side, parent=panel)
        if initial_visible is not None:
            try:
                switcher.set_initial_visible(tuple(initial_visible))
            except Exception:
                pass
        panel.setWidget(switcher)
        width = switcher.sizeHint().width()
        panel.setMinimumWidth(width)
        panel.setMaximumWidth(width)

        area = (
            QtCore.Qt.DockWidgetArea.LeftDockWidgetArea
            if side == "left"
            else QtCore.Qt.DockWidgetArea.RightDockWidgetArea
        )
        self.addDockWidget(area, panel)
        reference = docks[0]
        try:
            self.splitDockWidget(panel, reference, QtCore.Qt.Orientation.Horizontal)
        except Exception:
            pass
        return panel

    def _handle_log_visibility(self, visible: bool) -> None:
        if visible and getattr(self, "_log_has_unread_errors", False):
            self._log_has_unread_errors = False
            self._update_log_highlight()

    def _update_log_highlight(self) -> None:
        dock = getattr(self, "log_dock", None)
        if not isinstance(dock, QtWidgets.QDockWidget):
            return
        highlight = bool(getattr(self, "_log_has_unread_errors", False))
        if getattr(self, "_log_highlight_active", False) == highlight:
            return
        self._log_highlight_active = highlight
        if highlight:
            dock.setStyleSheet(
                """
                QDockWidget#builderMessageLogDock {
                    border: 1px solid #c62828;
                }
                QDockWidget#builderMessageLogDock::title {
                    background: #c62828;
                    color: #ffffff;
                    padding-left: 8px;
                }
                """
            )
            self.log_view.setStyleSheet(
                "background-color: #2b0b0b; color: #ffeaea; border: 1px solid #c62828;"
            )
        else:
            dock.setStyleSheet("")
            self.log_view.setStyleSheet("")
        for panel in getattr(self, "_dock_switcher_panels", []):
            if not isinstance(panel, QtWidgets.QDockWidget):
                continue
            switcher = panel.widget()
            if isinstance(switcher, _DockSwitcherWidget):
                switcher.set_tab_alert(dock, highlight)

    def _handle_fabrication_sources_changed(self, sources: Iterable[str]) -> None:
        video = getattr(self, "video_section", None)
        if isinstance(video, MiniDatabaseSection):
            video.set_sources(sources)

    def _collect_primary_docks(self) -> List[QtWidgets.QDockWidget]:
        docks: List[QtWidgets.QDockWidget] = []
        for candidate in (getattr(self, "project_dock", None), getattr(self, "log_dock", None)):
            if isinstance(candidate, QtWidgets.QDockWidget):
                docks.append(candidate)
        return docks

    def _handle_primary_dock_location_change(
        self,
        dock: QtWidgets.QDockWidget,
        area: QtCore.Qt.DockWidgetArea,
    ) -> None:
        if getattr(self, "_retabbing_docks", False):
            return
        if area == QtCore.Qt.DockWidgetArea.LeftDockWidgetArea:
            self._queue_retabify_primary_docks()

    def _handle_primary_dock_visibility_changed(self, dock: QtWidgets.QDockWidget) -> None:
        if getattr(self, "_retabbing_docks", False):
            return
        _ = dock
        self._queue_retabify_primary_docks()

    def _queue_retabify_primary_docks(self) -> None:
        if getattr(self, "_retabify_pending", False):
            return
        self._retabify_pending = True
        QtCore.QTimer.singleShot(0, self._run_queued_retabify)

    def _run_queued_retabify(self) -> None:
        self._retabify_pending = False
        self._retabify_primary_docks()

    def _retabify_primary_docks(self) -> None:
        if getattr(self, "_retabbing_docks", False):
            return
        self._retabbing_docks = True
        try:
            docks = [dock for dock in self._collect_primary_docks() if not dock.isFloating()]
            if not docks:
                return
            primary = docks[0]
            for dock in docks[1:]:
                try:
                    self.tabifyDockWidget(primary, dock)
                except Exception:
                    continue
            for dock in docks:
                width = max(dock.width(), self._primary_dock_widths.get(dock, 240))
                self._apply_dock_width(dock, width)
        finally:
            self._retabbing_docks = False

    def _apply_dock_width(self, dock: QtWidgets.QDockWidget, width: int) -> None:
        if dock.isFloating() or width <= 0:
            return
        try:
            screen = QtGui.QGuiApplication.screenAt(self.mapToGlobal(self.rect().center()))
            if screen is None:
                screen = QtGui.QGuiApplication.primaryScreen()
            available = screen.availableGeometry() if screen is not None else None
        except Exception:
            available = None
        if available is not None:
            width = max(160, min(width, available.width()))
        try:
            self.resizeDocks([dock], [width], QtCore.Qt.Orientation.Horizontal)
        except Exception:
            try:
                dock.resize(width, dock.height())
            except Exception:
                pass
        self._primary_dock_widths[dock] = width

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        self._clamp_to_available_geometry()

    def changeEvent(self, event: QtCore.QEvent) -> None:  # type: ignore[override]
        super().changeEvent(event)
        if event.type() == QtCore.QEvent.Type.WindowStateChange:
            if self.windowState() & (
                QtCore.Qt.WindowState.WindowMaximized | QtCore.Qt.WindowState.WindowFullScreen
            ):
                self._queue_fullscreen_snap()

    def showEvent(self, event: QtGui.QShowEvent) -> None:  # type: ignore[override]
        super().showEvent(event)
        QtCore.QTimer.singleShot(0, self._clamp_to_available_geometry)

    def _queue_fullscreen_snap(self) -> None:
        if self._fullscreen_snap_pending:
            return
        self._fullscreen_snap_pending = True
        QtCore.QTimer.singleShot(0, self._apply_fullscreen_geometry)

    def _apply_fullscreen_geometry(self) -> None:
        self._fullscreen_snap_pending = False
        if not (self.isMaximized() or self.isFullScreen()):
            return
        want_fullscreen = self.isFullScreen()
        want_maximized = self.isMaximized()
        if want_maximized and not want_fullscreen and sys.platform.startswith("win"):
            return
        screen = None
        try:
            handle = self.windowHandle()
            if handle is not None:
                screen = handle.screen()
        except Exception:
            screen = None
        if screen is None:
            try:
                screen = QtGui.QGuiApplication.screenAt(self.mapToGlobal(self.rect().center()))
            except Exception:
                screen = None
        if screen is None:
            screen = QtGui.QGuiApplication.primaryScreen()
        if screen is None:
            return
        rect = screen.geometry() if want_fullscreen else screen.availableGeometry()
        geom = self.geometry()
        frame = self.frameGeometry()
        tolerance = 2
        if (
            abs(frame.left() - rect.left()) <= tolerance
            and abs(frame.top() - rect.top()) <= tolerance
            and abs(frame.width() - rect.width()) <= tolerance
            and abs(frame.height() - rect.height()) <= tolerance
        ):
            return
        frame_margin_w = max(0, frame.width() - geom.width())
        frame_margin_h = max(0, frame.height() - geom.height())
        target_w = max(self.minimumWidth(), rect.width() - frame_margin_w)
        target_h = max(self.minimumHeight(), rect.height() - frame_margin_h)
        try:
            self.resize(max(1, int(target_w)), max(1, int(target_h)))
            self.move(rect.topLeft())
        except Exception:
            pass
        if want_fullscreen and not self.isFullScreen():
            try:
                self.setWindowState(self.windowState() | QtCore.Qt.WindowState.WindowFullScreen)
            except Exception:
                pass
        elif want_maximized and not self.isMaximized():
            try:
                self.setWindowState(self.windowState() | QtCore.Qt.WindowState.WindowMaximized)
            except Exception:
                pass

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:  # type: ignore[override]
        if self._dirty:
            box = QtWidgets.QMessageBox(self)
            box.setWindowTitle("Unsaved project")
            box.setText("Save changes to this Microwire Data Builder project before closing?")
            save_btn = box.addButton(QtWidgets.QMessageBox.StandardButton.Save)
            discard_btn = box.addButton("Discard", QtWidgets.QMessageBox.ButtonRole.DestructiveRole)
            cancel_btn = box.addButton(QtWidgets.QMessageBox.StandardButton.Cancel)
            box.setDefaultButton(save_btn)
            box.exec()
            clicked = box.clickedButton()
            if clicked is cancel_btn:
                event.ignore()
                return
            if clicked is save_btn:
                self._save_project()
                if self._dirty:
                    event.ignore()
                    return
        for section in self.sections.values():
            if isinstance(section, MiniDatabaseSection):
                try:
                    section._shutdown_background_threads()
                except Exception:
                    pass
        super().closeEvent(event)

    def _clamp_to_available_geometry(self) -> None:
        if getattr(self, "_clamp_active", False):
            return
        if self.isMaximized() or self.isFullScreen():
            return
        if self.windowState() & (
            QtCore.Qt.WindowState.WindowMaximized | QtCore.Qt.WindowState.WindowFullScreen
        ):
            return
        try:
            screen = QtGui.QGuiApplication.screenAt(self.mapToGlobal(self.rect().center()))
            if screen is None:
                screen = QtGui.QGuiApplication.primaryScreen()
        except Exception:
            screen = QtGui.QGuiApplication.primaryScreen()
        if screen is None:
            return
        available = screen.availableGeometry()
        self._clamp_active = True
        try:
            geom = self.geometry()
            frame = self.frameGeometry()
            if (
                abs(frame.width() - available.width()) <= 2
                and abs(frame.height() - available.height()) <= 2
            ):
                return
            frame_margin_w = max(0, frame.width() - geom.width())
            frame_margin_h = max(0, frame.height() - geom.height())
            max_width = max(1, available.width() - frame_margin_w)
            max_height = max(1, available.height() - frame_margin_h)

            new_width = max(self.minimumWidth(), min(geom.width(), max_width))
            new_height = max(self.minimumHeight(), min(geom.height(), max_height))
            if new_width != geom.width() or new_height != geom.height():
                self.resize(new_width, new_height)
                frame = self.frameGeometry()

            if (
                frame.width() >= available.width() - 2
                and frame.height() >= available.height() - 2
            ):
                self.move(available.topLeft())
                return

            max_left = available.right() - frame.width() + 1
            max_top = available.bottom() - frame.height() + 1
            bounded_left = min(max(frame.left(), available.left()), max_left)
            bounded_top = min(max(frame.top(), available.top()), max_top)
            if bounded_left != frame.left() or bounded_top != frame.top():
                self.move(bounded_left, bounded_top)
        finally:
            self._clamp_active = False

    def _handle_ocr_debug_changed(self, enabled: bool) -> None:
        section = getattr(self, "microscope_section", None)
        if isinstance(section, MicroscopeSection):
            try:
                section.set_ocr_debug_enabled(bool(enabled))
            except Exception:
                pass

    def _handle_log_capture_changed(self, enabled: bool) -> None:
        self._log_capture_enabled = bool(enabled)
        if self._log_capture_enabled:
            self._log_capture_path = self._resolve_log_capture_path()
            self._log_capture_faulted = False
            try:
                self.log(f"Message log capture enabled: {self._log_capture_path}")
            except Exception:
                pass
        else:
            try:
                self.log("Message log capture disabled.")
            except Exception:
                pass

    def _resolve_log_capture_path(self) -> Path:
        try:
            repo_root = Path(__file__).resolve().parents[1]
        except Exception:
            repo_root = Path.cwd()
        return repo_root / "logs" / "message_log.txt"

    def _append_log_to_file(self, level: int, message: str) -> None:
        if not self._log_capture_enabled or self._log_capture_faulted:
            return
        if self._log_capture_path is None:
            self._log_capture_path = self._resolve_log_capture_path()
        path = self._log_capture_path
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            level_name = logging.getLevelName(int(level))
            append_text_with_rotation(path, f"{timestamp} [{level_name}] {message}\n")
        except Exception as exc:
            self._log_capture_faulted = True
            self._log_capture_enabled = False
            try:
                self.log_view.appendPlainText(
                    f"WARNING: Failed to write message log capture file: {exc}"
                )
            except Exception:
                pass

    def _resolve_crash_log_path(self) -> Path:
        try:
            repo_root = Path(__file__).resolve().parents[1]
        except Exception:
            repo_root = Path.cwd()
        return repo_root / "logs" / "crash_log.txt"

    def _install_crash_handlers(self) -> None:
        if getattr(self, "_crash_handlers_installed", False):
            return
        self._crash_handlers_installed = True
        crash_path = self._resolve_crash_log_path()
        try:
            self._crash_log_handle = open_rotating_text_log(crash_path)
        except Exception:
            self._crash_log_handle = None
            return
        try:
            faulthandler.enable(file=self._crash_log_handle, all_threads=True)
        except Exception:
            pass

        previous_hook = sys.excepthook

        def _exception_hook(exc_type: type[BaseException], exc: BaseException, tb: Any) -> None:
            timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            trace = "".join(traceback.format_exception(exc_type, exc, tb))
            try:
                self._crash_log_handle.write(
                    f"{timestamp} [EXCEPTION] {exc_type.__name__}: {exc}\n{trace}\n"
                )
                self._crash_log_handle.flush()
            except Exception:
                pass
            try:
                self.logger.error("Unhandled exception: %s", exc)
            except Exception:
                pass
            try:
                previous_hook(exc_type, exc, tb)
            except Exception:
                pass

        sys.excepthook = _exception_hook

    def _handle_section_status_changed(self, key: str, status: str) -> None:
        item = self._project_items.get(key)
        if item is not None:
            item.setText(1, status)

    def _handle_section_sources_changed(self, key: str, sources: Iterable[str]) -> None:
        item = self._project_items.get(key)
        if item is None:
            return
        item.takeChildren()
        section = self.sections.get(key)
        processed: Dict[str, float] = {}
        if isinstance(section, MiniDatabaseSection):
            processed = dict(getattr(section.data, "processed", {}))
        for source in sources:
            source_text = str(source)
            child = QtWidgets.QTreeWidgetItem(["", source_text])
            child.setToolTip(1, source_text)
            source_path = Path(source_text)
            related_files: List[str] = []
            for path_text in sorted(processed.keys()):
                try:
                    path_obj = Path(path_text)
                except Exception:
                    continue
                try:
                    rel = path_obj.relative_to(source_path)
                    related_files.append(str(rel))
                except Exception:
                    if path_text.startswith(source_text):
                        related_files.append(path_text[len(source_text) :].lstrip(os.sep))
            related_files = list(dict.fromkeys(related_files))
            for rel_path in related_files:
                if not rel_path:
                    continue
                leaf = QtWidgets.QTreeWidgetItem(["", rel_path])
                leaf.setToolTip(1, rel_path)
                child.addChild(leaf)
            item.addChild(child)
        if sources:
            item.setExpanded(False)
        self._update_project_actions()
        self._mark_dirty()

    def _handle_section_data_updated(self) -> None:
        self._mark_dirty()
        self._update_project_actions()

    def _project_settings_key(self, name: str) -> str:
        return f"project/{name}"

    def _setup_project_actions(self, menu_bar: QtWidgets.QMenuBar) -> None:
        file_menu = menu_bar.findChild(QtWidgets.QMenu, "mw_shared_file")
        if file_menu is None:
            return
        new_action = QtGui.QAction("New Project", self)
        try:
            new_action.setShortcut(QtGui.QKeySequence(QtGui.QKeySequence.StandardKey.New))
        except Exception:
            pass
        new_action.triggered.connect(self._new_project)
        open_action = QtGui.QAction("Open Project…", self)
        try:
            open_action.setShortcut(QtGui.QKeySequence(QtGui.QKeySequence.StandardKey.Open))
        except Exception:
            pass
        open_action.triggered.connect(self._open_project)

        recent_menu = QtWidgets.QMenu("Recent Projects", file_menu)
        recent_menu.setObjectName("mw_builder_recent_projects")
        self._recent_projects_menu = recent_menu
        self._update_recent_projects_menu()

        save_action = QtGui.QAction("Save Project", self)
        try:
            save_action.setShortcut(QtGui.QKeySequence(QtGui.QKeySequence.StandardKey.Save))
        except Exception:
            pass
        save_action.triggered.connect(self._save_project)
        save_action.setEnabled(False)

        save_as_action = QtGui.QAction("Save Project &As…", self)
        try:
            save_as_action.setShortcut(QtGui.QKeySequence("Ctrl+Shift+S"))
        except Exception:
            pass
        save_as_action.triggered.connect(self._save_project_as)
        save_as_action.setEnabled(False)

        insert_before: Optional[QtGui.QAction] = None
        for action in file_menu.actions():
            text = action.text() or ""
            if "Close" in text:
                insert_before = action
                break
        if insert_before is not None:
            file_menu.insertAction(insert_before, new_action)
            file_menu.insertAction(insert_before, save_as_action)
            file_menu.insertAction(insert_before, save_action)
            file_menu.insertSeparator(insert_before)
            file_menu.insertMenu(insert_before, recent_menu)
            file_menu.insertAction(insert_before, open_action)
        else:
            file_menu.addAction(new_action)
            file_menu.addAction(open_action)
            file_menu.addMenu(recent_menu)
            file_menu.addSeparator()
            file_menu.addAction(save_action)
            file_menu.addAction(save_as_action)

        self._save_project_action = save_action
        self._save_project_as_action = save_as_action

    def _update_project_actions(self, *_: object) -> None:
        has_data = self._has_project_data_to_save()
        if self._save_project_action is not None:
            self._save_project_action.setEnabled(has_data)
        if self._save_project_as_action is not None:
            self._save_project_as_action.setEnabled(has_data)

    def _update_project_title(self) -> None:
        title = self._base_title
        if isinstance(self._project_path, Path):
            title = f"{self._base_title} - {self._project_path.name}"
        self.setWindowTitle(title)

    def _setup_settings_menu(self, menu_bar: QtWidgets.QMenuBar) -> None:
        settings_menu = menu_bar.addMenu("Settings")
        auto_open_action = QtGui.QAction("Open last project on startup", self)
        auto_open_action.setCheckable(True)
        auto_open_action.setChecked(self._auto_open_last)
        auto_open_action.toggled.connect(self._toggle_auto_open_last)
        settings_menu.addAction(auto_open_action)
        self._auto_open_last_action = auto_open_action

    def _setup_data_menu(self, menu_bar: QtWidgets.QMenuBar) -> None:
        data_menu = menu_bar.addMenu("Data")
        import_action = QtGui.QAction("Import workbook…", self)
        import_action.triggered.connect(self._handle_import_data)
        data_menu.addAction(import_action)

        show_imported_action = QtGui.QAction("Show imported data", self)
        show_imported_action.setCheckable(True)
        show_imported_action.setChecked(True)
        show_imported_action.toggled.connect(self._toggle_show_imported)
        data_menu.addAction(show_imported_action)

        separate_action = QtGui.QAction("Separate imported data", self)
        separate_action.setCheckable(True)
        initial_separate = bool(
            self.settings.value(self._project_settings_key("separate_imported"), False)
        )
        separate_action.setChecked(initial_separate)
        separate_action.toggled.connect(self._toggle_separate_imported)
        data_menu.addAction(separate_action)

        remove_action = QtGui.QAction("Remove imported data", self)
        remove_action.triggered.connect(self._remove_imported_data)
        data_menu.addAction(remove_action)

        self._data_menu = data_menu
        self._show_imported_action = show_imported_action
        self._separate_imported_action = separate_action
        self._remove_imported_action = remove_action

    def _handle_import_data(self) -> None:
        assembly = getattr(self, "assembly_section", None)
        if isinstance(assembly, AssemblySection):
            assembly.open_import_dialog()
            self._update_imported_data_item()
            self._update_project_actions()

    def _toggle_show_imported(self, enabled: bool) -> None:
        assembly = getattr(self, "assembly_section", None)
        if isinstance(assembly, AssemblySection):
            assembly.set_show_imported(bool(enabled))
            self._update_project_actions()

    def _toggle_separate_imported(self, enabled: bool) -> None:
        try:
            self.settings.setValue(
                self._project_settings_key("separate_imported"), bool(enabled)
            )
        except Exception:
            pass
        fabrication = getattr(self, "fabrication_section", None)
        if isinstance(fabrication, FabricationSection):
            fabrication.set_import_separation(bool(enabled))

    def _remove_imported_data(self) -> None:
        assembly = getattr(self, "assembly_section", None)
        if isinstance(assembly, AssemblySection):
            assembly.clear_imported_data()
            self._update_imported_data_item()
            self._update_project_actions()

    def _update_imported_data_item(self) -> None:
        item = self._imported_item
        if not isinstance(item, QtWidgets.QTreeWidgetItem):
            return
        item.takeChildren()
        assembly = getattr(self, "assembly_section", None)
        sources: List[str] = []
        if isinstance(assembly, AssemblySection):
            try:
                sources = assembly.imported_sources()
            except Exception:
                sources = []
        item.setText(1, f"{len(sources)} file(s)" if sources else "")
        for source in sources:
            child = QtWidgets.QTreeWidgetItem(["", source])
            child.setToolTip(1, source)
            item.addChild(child)
        if sources:
            item.setExpanded(False)

    def _mark_dirty(self) -> None:
        if self._suppress_dirty:
            return
        self._dirty = True

    def _has_project_data_to_save(self) -> bool:
        for section in self.sections.values():
            has_data = getattr(section, "has_project_data", None)
            if callable(has_data) and has_data():
                return True
        assembly = getattr(self, "assembly_section", None)
        if assembly is not None:
            has_data = getattr(assembly, "has_project_data", None)
            if callable(has_data) and has_data():
                return True
        return False

    def _project_dialog_start_directory(self) -> Path:
        stored = _sanitise_existing_directory(
            self.settings.value(self._project_settings_key("last_dir"), "")
        )
        if stored:
            return Path(stored)
        if isinstance(self._project_path, Path):
            return self._project_path.parent
        try:
            return Path(self._last_output_dir)
        except Exception:
            return _dialog_start_directory()

    def _default_project_filename(self) -> str:
        return f"microwire_project{self.PROJECT_EXTENSION}"

    def _build_project_payload(self) -> Dict[str, Any]:
        sections_payload: Dict[str, Any] = {}
        for key, section in self.sections.items():
            exporter = getattr(section, "export_project_payload", None)
            if not callable(exporter):
                continue
            try:
                sections_payload[key] = exporter()
            except Exception as exc:
                self.logger.error("Failed to export section %s: %s", key, exc)
        assembly = getattr(self, "assembly_section", None)
        if assembly is not None:
            exporter = getattr(assembly, "export_project_payload", None)
            if callable(exporter):
                try:
                    sections_payload["assemble"] = exporter()
                except Exception as exc:
                    self.logger.error("Failed to export section assemble: %s", exc)
        return {
            "version": self.PROJECT_VERSION,
            "kind": self.PROJECT_KIND,
            "saved_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
            "sections": sections_payload,
        }

    def _save_project(self) -> None:
        if not self._has_project_data_to_save():
            QtWidgets.QMessageBox.information(
                self,
                "Save Project",
                "There is no processed data to save yet.",
            )
            return
        if self._project_path is None:
            self._save_project_as()
            return
        self._write_project_file(self._project_path)

    def _save_project_as(self) -> None:
        if not self._has_project_data_to_save():
            QtWidgets.QMessageBox.information(
                self,
                "Save Project As",
                "Process or import data before saving a project.",
            )
            return
        start_dir = self._project_dialog_start_directory()
        suggested = start_dir / self._default_project_filename()
        filters = f"Microwire Project (*{self.PROJECT_EXTENSION});;All files (*)"
        path_str, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Project As",
            str(suggested),
            filters,
        )
        if not path_str:
            return
        target = Path(path_str)
        if target.suffix.lower() != self.PROJECT_EXTENSION:
            target = target.with_suffix(self.PROJECT_EXTENSION)
        self._write_project_file(target)

    def _write_project_file(self, target: Path) -> None:
        payload = self._build_project_payload()
        sections = payload.get("sections", {})
        if not sections:
            QtWidgets.QMessageBox.information(
                self,
                "Save Project",
                "No processed sections are available to save.",
            )
            return
        try:
            target.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        except Exception as exc:
            QtWidgets.QMessageBox.critical(
                self,
                "Save Project",
                f"Failed to write project file:\\n{exc}",
            )
            return
        self._project_path = target
        self._remember_project_directory(target.parent)
        self._remember_recent_project(target)
        try:
            self.settings.setValue(self._project_settings_key("last_path"), str(target))
        except Exception:
            pass
        self._update_project_title()
        self._update_project_actions()
        self._dirty = False
        self.logger.info("Project saved to %s", target)
        QtWidgets.QMessageBox.information(
            self,
            "Save Project",
            f"Project saved to {target}",
        )

    def _load_recent_projects_setting(self) -> None:
        raw = self.settings.value(self._project_settings_key("recent"), "[]")
        entries: List[str]
        if isinstance(raw, str):
            try:
                decoded = json.loads(raw)
                entries = [str(item) for item in decoded if isinstance(item, str)]
            except json.JSONDecodeError:
                entries = []
        elif isinstance(raw, (list, tuple)):
            entries = [str(item) for item in raw if isinstance(item, str)]
        else:
            entries = []
        original_entries = list(entries)
        seen: set[str] = set()
        ordered: List[str] = []
        for entry in entries:
            candidate = str(entry).strip()
            if (
                candidate
                and not _looks_like_test_path(candidate)
                and candidate not in seen
            ):
                seen.add(candidate)
                ordered.append(candidate)
        self._recent_projects = ordered[:8]
        if self._recent_projects != original_entries[:8]:
            self._save_recent_projects_setting()

    def _save_recent_projects_setting(self) -> None:
        try:
            payload = json.dumps(self._recent_projects[:8], ensure_ascii=False)
        except (TypeError, ValueError):
            payload = "[]"
        self.settings.setValue(self._project_settings_key("recent"), payload)

    def _clear_recent_projects(self) -> None:
        self._recent_projects = []
        self._save_recent_projects_setting()
        self._update_recent_projects_menu()

    def _toggle_auto_open_last(self, enabled: bool) -> None:
        self._auto_open_last = bool(enabled)
        try:
            self.settings.setValue(self._project_settings_key("auto_open_last"), int(bool(enabled)))
        except Exception:
            pass

    def _maybe_auto_open_last_project(self) -> None:
        if not self._auto_open_last:
            return
        last_path = _sanitise_existing_file(
            self.settings.value(self._project_settings_key("last_path"), "")
        )
        candidate: Optional[Path] = None
        if last_path:
            path_obj = Path(last_path)
            if path_obj.exists():
                candidate = path_obj
        else:
            try:
                self.settings.remove(self._project_settings_key("last_path"))
            except Exception:
                pass
        if candidate is None and self._recent_projects:
            fallback = Path(self._recent_projects[0])
            if fallback.exists():
                candidate = fallback
        if candidate is None:
            return
        try:
            self._load_project_from_path(candidate)
        except Exception:
            self.logger.exception("Failed to auto-open last project %s", candidate)

    def _new_project(self) -> None:
        if getattr(self, "_dirty", False):
            box = QtWidgets.QMessageBox(self)
            box.setWindowTitle("Unsaved project")
            box.setText("Save changes to this Microwire Data Builder project before starting a new one?")
            save_btn = box.addButton(QtWidgets.QMessageBox.StandardButton.Save)
            discard_btn = box.addButton("Discard", QtWidgets.QMessageBox.ButtonRole.DestructiveRole)
            cancel_btn = box.addButton(QtWidgets.QMessageBox.StandardButton.Cancel)
            box.setDefaultButton(save_btn)
            box.exec()
            clicked = box.clickedButton()
            if clicked is cancel_btn:
                return
            if clicked is save_btn:
                self._save_project()
                if getattr(self, "_dirty", False):
                    return
        self._suppress_dirty = True
        for section in self.sections.values():
            if isinstance(section, MiniDatabaseSection):
                section.reset_to_blank()
        self._project_path = None
        self._dirty = False
        self._suppress_dirty = False
        self._update_project_title()
        self._update_project_actions()
        self._refresh_sections_after_project_load()

    # Compatibility alias used by menu wiring in some launch contexts
    def new_project(self) -> None:
        self._new_project()

    def _remember_recent_project(self, path: Path) -> None:
        try:
            resolved = str(path.resolve())
        except Exception:
            resolved = str(path)
        self._recent_projects = [entry for entry in self._recent_projects if entry != resolved]
        self._recent_projects.insert(0, resolved)
        self._recent_projects = self._recent_projects[:8]
        self._save_recent_projects_setting()
        self._update_recent_projects_menu()

    def _update_recent_projects_menu(self) -> None:
        menu = self._recent_projects_menu
        if not isinstance(menu, QtWidgets.QMenu):
            return
        menu.clear()
        if not self._recent_projects:
            placeholder = menu.addAction("No recent projects")
            placeholder.setEnabled(False)
            return
        for entry in self._recent_projects:
            display = Path(entry).name or entry
            action = menu.addAction(display)
            action.setToolTip(entry)
            action.triggered.connect(partial(self._open_recent_project, entry))
        menu.addSeparator()
        clear_action = menu.addAction("Clear list")
        clear_action.triggered.connect(self._clear_recent_projects)

    def _open_recent_project(self, entry: str) -> None:
        candidate = Path(entry)
        if not candidate.exists():
            QtWidgets.QMessageBox.warning(
                self,
                "Open Project",
                f"The project file {entry} could not be found. It has been removed from the recent list.",
            )
            self._recent_projects = [item for item in self._recent_projects if item != entry]
            self._save_recent_projects_setting()
            self._update_recent_projects_menu()
            return
        self._load_project_from_path(candidate)

    def _open_project(self) -> None:
        start_dir = self._project_dialog_start_directory()
        filters = f"Microwire Project (*{self.PROJECT_EXTENSION});;All files (*)"
        path_str, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Open Project",
            str(start_dir),
            filters,
        )
        if not path_str:
            return
        target = Path(path_str)
        self._load_project_from_path(target)

    def _load_project_from_path(self, target: Path) -> None:
        progress_dialog: Optional[QtWidgets.QProgressDialog] = None
        total_steps = max(len(self.sections) + 1, 1)
        last_pump = 0.0

        def _pump_events(step: int | None = None, label: str | None = None) -> None:
            """Keep the UI responsive while loading a project."""

            nonlocal last_pump
            if progress_dialog is not None:
                try:
                    if label:
                        progress_dialog.setLabelText(label)
                    if step is not None:
                        progress_dialog.setValue(step)
                except Exception:
                    pass
            now = time.monotonic()
            if last_pump == 0.0 or now - last_pump >= 0.05:
                try:
                    QtWidgets.QApplication.processEvents(
                        QtCore.QEventLoop.ProcessEventsFlag.AllEvents
                    )
                except Exception:
                    pass
                last_pump = now

        try:
            progress_dialog = QtWidgets.QProgressDialog(
                "Loading project…", "", 0, total_steps, self
            )
            progress_dialog.setWindowModality(QtCore.Qt.WindowModality.ApplicationModal)
            progress_dialog.setCancelButton(None)
            progress_dialog.setMinimumDuration(150)
            progress_dialog.setAutoClose(False)
            progress_dialog.setAutoReset(False)
            progress_dialog.show()
            _pump_events(0, "Loading project…")
        except Exception:
            progress_dialog = None

        self._suppress_dirty = True
        try:
            payload = json.loads(target.read_text(encoding="utf-8"))
            _pump_events(0)

            if payload.get("kind") != self.PROJECT_KIND:
                QtWidgets.QMessageBox.critical(
                    self,
                    "Open Project",
                    "The selected file is not a Microwire Data Builder project.",
                )
                return

            sections_payload = payload.get("sections", {})
            if not isinstance(sections_payload, Mapping):
                sections_payload = {}

            for index, (key, section) in enumerate(self.sections.items(), start=1):
                label = getattr(section, "section_title", key)
                _pump_events(index - 1, f"Loading {label}…")
                importer = getattr(section, "import_project_payload", None)
                if callable(importer):
                    if isinstance(section, MiniDatabaseSection):
                        section.reset_to_blank()
                    section_payload = sections_payload.get(key)
                    try:
                        importer(section_payload or {})
                    except Exception as exc:
                        self.logger.error("Failed to load section %s from project: %s", key, exc)
                _pump_events(index)

            assembly_payload = sections_payload.get("assemble")
            assembly = getattr(self, "assembly_section", None)
            if assembly is not None:
                importer = getattr(assembly, "import_project_payload", None)
                if callable(importer):
                    try:
                        importer(assembly_payload or {})
                    except Exception as exc:
                        self.logger.error("Failed to load section assemble: %s", exc)
            self._update_imported_data_item()
            if isinstance(assembly, AssemblySection):
                show_imported = getattr(assembly, "_show_imported", True)
                if self._show_imported_action is not None:
                    self._show_imported_action.setChecked(bool(show_imported))
            if self._separate_imported_action is not None:
                separate = bool(
                    self.settings.value(self._project_settings_key("separate_imported"), False)
                )
                self._separate_imported_action.setChecked(separate)
                fabrication = getattr(self, "fabrication_section", None)
                if isinstance(fabrication, FabricationSection):
                    fabrication.set_import_separation(separate)

            self._project_path = target
            self._remember_project_directory(target.parent)
            self._remember_recent_project(target)
            try:
                self.settings.setValue(self._project_settings_key("last_path"), str(target))
            except Exception:
                pass
            self._update_project_title()
            self._refresh_sections_after_project_load()
            self._update_project_actions()
            self._dirty = False
            self.logger.info("Project loaded from %s", target)
            _pump_events(total_steps, "Finishing…")
            QtWidgets.QMessageBox.information(
                self,
                "Open Project",
                f"Loaded project from {target}",
            )
        except Exception as exc:
            self.logger.exception("Failed to load project %s", target, exc_info=exc)
            QtWidgets.QMessageBox.critical(
                self,
                "Open Project",
                f"Failed to load project file:\n{exc}",
            )
        finally:
            self._suppress_dirty = False
            if progress_dialog is not None:
                try:
                    progress_dialog.close()
                except Exception:
                    try:
                        progress_dialog.cancel()
                    except Exception:
                        pass

    def _refresh_sections_after_project_load(self) -> None:
        for key, section in self.sections.items():
            status_text = ""
            status_label = getattr(section, "status_label", None)
            if isinstance(status_label, QtWidgets.QLabel):
                status_text = status_label.text()
            self._handle_section_status_changed(key, status_text)
            sources: Iterable[str] = []
            if isinstance(section, MiniDatabaseSection):
                sources = section.data.sources
            self._handle_section_sources_changed(key, sources)
        self._handle_fabrication_sources_changed(self.fabrication_section.data.sources)

    def _remember_project_directory(self, directory: Path) -> None:
        try:
            resolved = directory.resolve()
        except Exception:
            resolved = directory
        self.settings.setValue(self._project_settings_key("last_dir"), str(resolved))

def run_app() -> None:
    main()


def main() -> QtWidgets.QWidget | None:
    app = QtWidgets.QApplication.instance()
    owns_app = False
    if app is None:
        app = QtWidgets.QApplication(sys.argv)
        ensure_app_theme(app)
        owns_app = True
    placeholder = QtWidgets.QMainWindow()
    placeholder.setWindowTitle("Microwire Data Builder")
    placeholder.resize(420, 260)
    loading_label = QtWidgets.QLabel("Loading Microwire Data Builder...", placeholder)
    loading_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
    loading_label.setStyleSheet("font-size: 16px; font-weight: 600;")
    placeholder.setCentralWidget(loading_label)
    placeholder.show()
    try:
        app.processEvents()
    except Exception:
        pass

    window_holder: dict[str, QtWidgets.QWidget] = {}

    def _launch() -> None:
        window = BuilderWindow()
        window_holder["window"] = window
        try:
            window.show()
        except Exception:
            window.show()
        placeholder.close()

    if owns_app:
        QtCore.QTimer.singleShot(0, _launch)
        app.exec()
        return window_holder.get("window")

    _launch()
    return window_holder.get("window")


__all__ = ["BuilderWindow", "main", "run_app"]
