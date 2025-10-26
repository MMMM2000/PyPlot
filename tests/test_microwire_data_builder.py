"""Tests for the microwire data builder core logic."""

from __future__ import annotations

from pathlib import Path

import importlib.util
import logging
import sys

import pandas as pd
import pytest

CORE_PATH = Path(__file__).resolve().parent.parent / "microwire_data_builder" / "core.py"

spec = importlib.util.spec_from_file_location("microwire_data_builder_core", CORE_PATH)
assert spec and spec.loader
core = importlib.util.module_from_spec(spec)
sys.modules["microwire_data_builder_core"] = core
spec.loader.exec_module(core)

BuilderConfig = core.BuilderConfig
build_database = core.build_database
_canonical_dimension_field = core._canonical_dimension_field
_header_key = core._header_key
_load_annealing = core._load_annealing
_metadata_from_path = core._metadata_from_path
_resistance_sanity_check = core._resistance_sanity_check
_safe_plot_stem = core._safe_plot_stem
OriginArtifact = core.OriginArtifact
FabricationIndex = core.FabricationIndex
_merged_header_row = core._merged_header_row
_parse_piece_rows = core._parse_piece_rows


def test_filename_parser_extracts_metadata(tmp_path: Path) -> None:
    path = tmp_path / "Ni50Fe27Ga23 6_4a s2 30mA.txt"
    path.write_text("0.1 0.2 2.0\n")
    metadata = _metadata_from_path(path)
    assert metadata.composition_token == "Ni50Fe27Ga23"
    assert metadata.draw_x == 6
    assert metadata.piece_y == 4
    assert metadata.alt_variant is True
    assert metadata.setpoint_mA == 30
    assert metadata.file_name == path.name
    assert metadata.measurement_id


def test_annealing_loader_and_sanity_check(tmp_path: Path) -> None:
    content = "0.1 0.2 2.0\n0.2 0.4 2.0\n0.3 0.6 2.0\n"
    path = tmp_path / "anneal.txt"
    path.write_text(content)
    df = _load_annealing(path)
    assert list(df.columns) == ["I_A", "V_V", "R_ohm", "I_mA"]
    expected_A = [0.1, 0.2, 0.3]
    assert df["I_A"].tolist() == pytest.approx(expected_A)
    assert df["I_mA"].tolist() == pytest.approx([value * 1_000.0 for value in expected_A])
    ok, error = _resistance_sanity_check(df)
    assert ok is True
    assert error is not None
    assert error < 1e-6


def test_annealing_loader_trims_burnthrough_spike(tmp_path: Path) -> None:
    path = tmp_path / "burn.txt"
    path.write_text("0.05 0.10 2.0\n0.06 0.12 2.0\n0.03 0.50 20.0\n")
    df = _load_annealing(path)
    assert len(df) == 2
    assert df["I_A"].tolist() == pytest.approx([0.05, 0.06])
    assert df["I_mA"].tolist() == pytest.approx([50.0, 60.0])
    assert df["R_ohm"].tolist() == pytest.approx([2.0, 2.0])


def test_header_normaliser_variants() -> None:
    assert _header_key("hmotnosť") == "mass_g"
    assert _header_key("P.Č") == "piece_y"
    assert _header_key("d (µm)") == "d_um"
    assert _header_key("D (µm)") == "D_um"
    assert _header_key("d/D") == "d_over_D"
    assert _header_key("Poznámka") == "notes"


def test_piece_header_backfill_extracts_diameters(tmp_path: Path) -> None:
    df = pd.DataFrame(
        [
            [
                "1.Ni46Fe23Ga23Co8 13.03.2025 09:15",
                None,
                None,
                "odpor",
                "d",
                "D",
                "d/D",
            ],
            ["P.Č", "Dátum", "Dĺžka (m)", None, None, None, None],
            ["1.", "45729", "6.4056", "2.15", "7", "25", "0.28"],
        ],
        dtype=object,
    )

    header_idx = 1
    header_values = _merged_header_row(df, header_idx)
    headers = [_header_key(value) for value in header_values]
    index = FabricationIndex()
    _parse_piece_rows(
        df.iloc[header_idx + 1 :],
        headers,
        "Ni46Fe23Ga23Co8",
        1,
        index,
        logging.getLogger("test"),
        tmp_path / "piece.xlsx",
    )
    record = index.get_piece("Ni46Fe23Ga23Co8", 1, 1)
    assert record["d_um"] == pytest.approx(7.0)
    assert record["D_um"] == pytest.approx(25.0)
    assert record["d_over_D"] == pytest.approx(0.28)
    assert record["fabrication_resistance_ohm"] == pytest.approx(2.15)
    display = record.get("d_um__display")
    assert isinstance(display, list) and "7" in display[0]


def test_canonical_dimension_field_filters_non_diameter_columns() -> None:
    assert _canonical_dimension_field("glass_feed_mm_per_min") is None
    assert _canonical_dimension_field("core_diameter_um") == "d_um"
    assert _canonical_dimension_field("glass_diameter_um_raw") == "D_um"
    assert _canonical_dimension_field("ratio_d_core_to_D_glass") == "d_over_D"


def test_safe_plot_stem_removes_path_separators() -> None:
    stem = _safe_plot_stem("Ni55Fe18Ga27 4/1 s1 1000mA")
    assert "/" not in stem
    assert stem.endswith("1000mA")


def test_build_database_integration(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    base = Path("sample_data/database_builder")
    anneal_files = [
        base / "current annealing data" / "Ni55Fe18Ga27 4_1 s1 1000mA.txt",
        base / "current annealing data" / "Ni55Fe18Ga27 4_1 s2 100mA.txt",
    ]
    composition_file = base / "microwire data" / "Ni55Fe18Ga27" / "Ni55Fe18Ga27.xlsx"
    piece_dir = base / "microwire data" / "Ni55Fe18Ga27" / "4.Ni55Fe18Ga27 26112024 0850"
    piece_file = sorted(piece_dir.glob("*.xlsx"))[0]

    config = BuilderConfig(
        fabrication_files=[composition_file, piece_file],
        annealing_files=anneal_files,
        output_dir=tmp_path / "out",
    )

    result = build_database(config)
    df = result.dataframe
    assert len(df) == 1
    row = df.iloc[0]
    expected_columns = list(core.OUTPUT_COLUMNS)
    expected_columns.insert(expected_columns.index("d (µm)") + 1, "d (µm) image")
    expected_columns.insert(expected_columns.index("D (µm)") + 1, "D (µm) image")
    assert list(df.columns) == expected_columns
    assert row["Composition"] == "Ni55Fe18Ga27"
    assert row["Microwire"] == "4/1"
    assert row["File 1000 mA"] == anneal_files[0].name
    assert row["File low mA"] == anneal_files[1].name
    assert pd.isna(row[core.STRAIN_COLUMN])
    assert row["Low mA value (mA)"] == 100
    assert pd.isna(row["d (µm)"])
    assert pd.isna(row["D (µm)"])
    assert row["Production datetime"] == "2024-11-26 08:50:00"
    assert "csv" in result.exports
    assert Path(result.exports["csv"]).exists()


def test_build_database_populates_plot_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    high = tmp_path / "Ni55Fe18Ga27 1_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 1_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")

    produced: dict[str, Path] = {}

    def fake_plot(df, source: Path, plot_dir: Path, figsize: tuple[float, float]) -> Path:
        plot_dir.mkdir(parents=True, exist_ok=True)
        out_path = plot_dir / f"{source.stem}.png"
        out_path.write_text("stub")
        produced[source.name] = out_path
        return out_path

    monkeypatch.setattr(core, "_plot_measurement_matplotlib", fake_plot)

    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        make_plots=True,
    )

    result = build_database(config)
    assert result.plot_paths
    assert not result.origin_artifacts
    row = result.dataframe.iloc[0]
    assert row["Figure — 1000 mA"] == produced[high.name].name
    assert row["Figure — low mA"] == produced[low.name].name
    assert pd.isna(row["Figure — 1000 mA (Origin)"])
    assert pd.isna(row["Figure — low mA (Origin)"])
    assert set(result.plot_paths) == {produced[high.name].name, produced[low.name].name}
    assert row["Low mA value (mA)"] == 120
    assert pd.isna(row[core.STRAIN_COLUMN])


def test_build_database_origin_backend(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    high = tmp_path / "Ni55Fe18Ga27 1_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 1_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")

    origin_records: dict[str, OriginArtifact] = {}

    def fake_origin(df, source: Path, origin_dir: Path, log: logging.Logger | None) -> OriginArtifact:
        origin_dir.mkdir(parents=True, exist_ok=True)
        descriptor = f"{source.stem}.oggu"
        artifact_path = origin_dir / descriptor
        artifact = OriginArtifact(descriptor=descriptor, object_path=artifact_path)
        origin_records[source.name] = artifact
        return artifact

    monkeypatch.setattr(core, "_plot_measurement_origin", fake_origin)

    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        make_plots=True,
        plot_backends=("origin",),
    )

    result = build_database(config)
    assert not result.plot_paths
    assert set(result.origin_artifacts.keys()) == {artifact.descriptor for artifact in origin_records.values()}
    row = result.dataframe.iloc[0]
    assert row["Figure — 1000 mA (Origin)"] == origin_records[high.name].descriptor
    assert row["Figure — low mA (Origin)"] == origin_records[low.name].descriptor
    assert pd.isna(row["Figure — 1000 mA"])
    assert pd.isna(row["Figure — low mA"])
    assert pd.isna(row[core.STRAIN_COLUMN])


def test_excel_export_embeds_plot_images(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    pytest.importorskip("openpyxl")
    from PIL import Image as PILImage

    high = tmp_path / "Ni55Fe18Ga27 1_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 1_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")

    def fake_plot(df, source: Path, plot_dir: Path, figsize: tuple[float, float]) -> Path:
        plot_dir.mkdir(parents=True, exist_ok=True)
        out_path = plot_dir / f"{source.stem}.png"
        PILImage.new("RGB", (320, 200), color=(255, 0, 0)).save(out_path)
        return out_path

    monkeypatch.setattr(core, "_plot_measurement_matplotlib", fake_plot)

    custom_figsize = (5.5, 3.5)
    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        make_plots=True,
        export_formats=("excel",),
        matplotlib_figsize=custom_figsize,
    )

    result = build_database(config)
    excel_path = result.exports["excel"]
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    images = getattr(worksheet, "_images", [])
    assert images, "Expected embedded plot images in the Excel export"

    figure_col_idx = result.dataframe.columns.get_loc("Figure — 1000 mA")
    col_letter = get_column_letter(figure_col_idx + 1)
    assert worksheet[f"{col_letter}2"].value is None

    expected_row_height = core._excel_row_height(custom_figsize[1])
    expected_col_width = core._excel_column_width(custom_figsize[0])
    assert worksheet.row_dimensions[2].height == pytest.approx(expected_row_height, rel=0.01)
    assert worksheet.column_dimensions[col_letter].width == pytest.approx(expected_col_width, rel=0.05)

    anchor = images[0].anchor
    if hasattr(anchor, "_from"):
        assert anchor._from.col == figure_col_idx
        assert anchor._from.row == 1

    workbook.close()


def test_excel_export_respects_high_dpi_images(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    pytest.importorskip("openpyxl")
    from PIL import Image as PILImage
    from zipfile import ZipFile
    from xml.etree import ElementTree as ET

    high = tmp_path / "Ni55Fe18Ga27 1_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 1_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")

    def fake_plot(df, source: Path, plot_dir: Path, figsize: tuple[float, float]) -> Path:
        plot_dir.mkdir(parents=True, exist_ok=True)
        out_path = plot_dir / f"{source.stem}.png"
        PILImage.new("RGB", (1650, 1050), color=(0, 128, 0)).save(out_path, dpi=(300, 300))
        return out_path

    monkeypatch.setattr(core, "_plot_measurement_matplotlib", fake_plot)

    custom_figsize = (5.5, 3.5)
    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        make_plots=True,
        export_formats=("excel",),
        matplotlib_figsize=custom_figsize,
    )

    result = build_database(config)
    excel_path = result.exports["excel"]

    with ZipFile(excel_path, "r") as archive:
        drawing_xml = archive.read("xl/drawings/drawing1.xml")

    tree = ET.fromstring(drawing_xml)
    ns_main = "http://schemas.openxmlformats.org/drawingml/2006/main"
    ns_sheet = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    ext = tree.find(f".//{{{ns_main}}}ext")
    if ext is None:
        ext = tree.find(f".//{{{ns_sheet}}}ext")
    assert ext is not None, "Expected drawing metadata for embedded figure"
    width_emu = int(ext.get("cx"))
    height_emu = int(ext.get("cy"))
    emu_per_inch = 914400
    width_in = width_emu / emu_per_inch
    height_in = height_emu / emu_per_inch
    assert width_in == pytest.approx(custom_figsize[0], rel=0.01)
    assert height_in == pytest.approx(custom_figsize[1], rel=0.01)


def test_microscope_images_populate_diameters(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    high = tmp_path / "Ni55Fe18Ga27 1_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 1_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")
    core_img = tmp_path / "Ni55Fe18Ga27 1_1 core.jpg"
    glass_img = tmp_path / "Ni55Fe18Ga27 1_1 glass.jpg"
    core_img.write_bytes(b"core")
    glass_img.write_bytes(b"glass")

    def fake_extract(path: Path, logger: logging.Logger) -> core.MicroscopeOCRResult:
        result = core.MicroscopeOCRResult()
        name = path.name.lower()
        if 'core' in name:
            result.append_value(16.7)
            result.detections.append(core.MicroscopeDetection(value=16.7, image_path=core_img))
        elif 'glass' in name:
            result.append_value(134.4)
            result.append_value(212.4)
            result.detections.append(core.MicroscopeDetection(value=134.4, image_path=glass_img))
            result.detections.append(core.MicroscopeDetection(value=212.4, image_path=glass_img))
        return result

    monkeypatch.setattr(core, "_extract_microscope_diameters", fake_extract)

    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        microscope_files=[core_img, glass_img],
    )

    result = build_database(config)
    row = result.dataframe.iloc[0]
    d_col = core.OUTPUT_COLUMNS[2]
    D_col = core.OUTPUT_COLUMNS[3]
    ratio_col = core.OUTPUT_COLUMNS[4]
    assert float(row[d_col]) == pytest.approx(16.7)
    assert float(row[D_col]) == pytest.approx(212.4)
    expected_ratio = round(16.7 / 212.4, 3)
    assert float(row[ratio_col]) == pytest.approx(expected_ratio)


def test_microscope_ocr_extracts_bracketed_values(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from PIL import Image as PILImage

    image_path = tmp_path / "Ni55Fe18Ga27 4_1 core.jpg"
    PILImage.new("RGB", (320, 180), color="white").save(image_path)

    class FakeOCR:
        def ocr(self, image, cls: bool = True):  # pragma: no cover - simple stub
            return [
                [
                    (
                        [[0, 0], [160, 0], [160, 40], [0, 40]],
                        ("[1]6.7um", 0.95),
                    ),
                    (
                        [[0, 60], [220, 60], [220, 110], [0, 110]],
                        ("[2]134.5um", 0.94),
                    ),
                ]
            ]

    monkeypatch.setattr(core, "get_paddle_ocr", lambda logger=None: FakeOCR())

    result = core._extract_microscope_diameters(image_path, logging.getLogger("test"))
    assert any(abs(value - 6.7) < 1e-3 for value in result.values)
    assert any(abs(value - 134.5) < 1e-3 for value in result.values)

    grouped = core._group_microscope_measurements([image_path], logging.getLogger("test"))
    key = core._microscope_key(image_path)
    assert key in grouped
    measurements = grouped[key]
    assert measurements.best_core() == pytest.approx(6.7, rel=1e-3)
    assert measurements.best_glass() == pytest.approx(134.5, rel=1e-3)


def test_parse_microscope_candidates_prefers_primary_marker() -> None:
    values = core._parse_microscope_candidates([
        "2025/09/25 [116.7um extra [2] 20.0um",
    ])
    assert values == [pytest.approx(6.7)]


def test_parse_microscope_candidates_ignores_secondary() -> None:
    values = core._parse_microscope_candidates([
        "[2] 44.1um 18.5um",
    ])
    assert values == [pytest.approx(18.5)]


def test_parse_microscope_candidates_filters_outliers() -> None:
    sample_text = """5001000 . 7235.0um\n11]65.1um .\n25.0um"""
    values = core._parse_microscope_candidates([sample_text])
    assert values == [pytest.approx(65.1)]


def test_microscope_key_handles_additional_delimiters() -> None:
    dashed = Path("Ni50Fe27Ga23 5-4 core.jpg")
    spaced = Path("Ni50Fe27Ga23 5 4 glass.png")
    assert core._microscope_key(dashed) == ("Ni50Fe27Ga23", 5, 4)
    assert core._microscope_key(spaced) == ("Ni50Fe27Ga23", 5, 4)


def test_video_metrics_populate_draw_fields(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    high = tmp_path / "Ni55Fe18Ga27 4_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 4_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")
    video_dir = tmp_path / "Ni55Fe18Ga27" / "4.Ni55Fe18Ga27 01012024 0800"
    video_dir.mkdir(parents=True, exist_ok=True)
    video_path = video_dir / "2025-07-02 11-44-34.mkv"
    video_path.write_bytes(b"video")

    class FakeVideoResult:
        def median_temperature(self) -> float | None:
            return 382.5

        def median_underpressure(self) -> float | None:
            return -0.85

        def median_winding_speed(self) -> float | None:
            return 12.5

        def median_glass_feed(self) -> float | None:
            return 37.2

    monkeypatch.setattr(core, "extract_video_metrics", lambda *args, **kwargs: FakeVideoResult())

    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        video_files=[video_path],
        highlight_ocr_values=True,
    )

    result = build_database(config)
    row = result.dataframe.iloc[0]
    temperature_column = "Temperature (°C)"
    underpressure_column = "Underpressure"
    assert float(row[temperature_column]) == pytest.approx(382.5)
    assert float(row[underpressure_column]) == pytest.approx(-0.85)
    assert float(row["Winding speed (m/min)"]) == pytest.approx(12.5)
    assert float(row["Glass feeding (mm/min)"]) == pytest.approx(37.2)
    highlights = result.ocr_highlights
    for column in (
        "Temperature (°C)",
        "Underpressure",
        "Winding speed (m/min)",
        "Glass feeding (mm/min)",
    ):
        assert column in highlights
        assert 0 in highlights[column]


def test_highlight_and_crop_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    high = tmp_path / "Ni55Fe18Ga27 4_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 4_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.1\n0.1 0.2 2.1\n")
    core_img = tmp_path / "Ni55Fe18Ga27 4_1 core.png"
    glass_img = tmp_path / "Ni55Fe18Ga27 4_1 glass.png"

    from PIL import Image

    Image.new("RGB", (40, 40), color="white").save(core_img)
    Image.new("RGB", (40, 40), color="white").save(glass_img)

    def fake_extract(path: Path, logger: logging.Logger) -> core.MicroscopeOCRResult:
        result = core.MicroscopeOCRResult()
        name = path.name.lower()
        if "core" in name:
            detection = core.MicroscopeDetection(
                value=10.0,
                image_path=core_img,
                bbox=(5, 5, 25, 25),
            )
            result.append_value(10.0)
            result.detections.append(detection)
        elif "glass" in name:
            detection = core.MicroscopeDetection(
                value=50.0,
                image_path=glass_img,
                bbox=(4, 4, 30, 30),
            )
            result.append_value(50.0)
            result.detections.append(detection)
        return result

    monkeypatch.setattr(core, "_extract_microscope_diameters", fake_extract)

    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        microscope_files=[core_img, glass_img],
        include_microscope_crops=True,
        highlight_ocr_values=True,
    )

    result = build_database(config)
    row = result.dataframe.iloc[0]
    assert "d (µm) image" in result.dataframe.columns
    assert "D (µm) image" in result.dataframe.columns
    crop_key = row["d (µm) image"]
    assert isinstance(crop_key, str) and crop_key in result.microscope_crops
    assert "d (µm)" in result.ocr_highlights
    assert 0 in result.ocr_highlights["d (µm)"]
    assert "D (µm)" in result.ocr_highlights
    assert 0 in result.ocr_highlights["D (µm)"]

def test_build_database_uses_strain_records(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    high = tmp_path / "Ni55Fe18Ga27 4_1 1000mA.txt"
    low = tmp_path / "Ni55Fe18Ga27 4_1 120mA.txt"
    high.write_text("0.1 0.2 2.0\n0.2 0.4 2.0\n")
    low.write_text("0.05 0.1 2.0\n0.1 0.2 2.0\n")
    strain_path = tmp_path / "strain.xlsx"
    pd.DataFrame(
        {
            "Composition": ["Ni55Fe18Ga27"],
            "Microwire": ["4/1"],
            "M length": [32],
            "A length": [30],
            "Strain %": [6.25],
        }
    ).to_excel(strain_path, index=False)
    config = BuilderConfig(
        fabrication_files=[],
        annealing_files=[high, low],
        output_dir=tmp_path / "out",
        strain_files=[strain_path],
    )
    result = build_database(config)
    row = result.dataframe.iloc[0]
    assert row[core.STRAIN_COLUMN] == "6.250%"
    columns = result.dataframe.columns.tolist()
    figure_idx = columns.index("Figure — 1000 mA")
    assert columns[figure_idx + 1] == "Figure — low mA"
    assert columns[figure_idx + 2] == core.STRAIN_COLUMN


def test_update_existing_exports_with_strain(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    strain_path = tmp_path / "strain.xlsx"
    pd.DataFrame(
        {
            "Composition": ["Ni55Fe18Ga27"],
            "Microwire": ["4/1"],
            "M length": [32],
            "A length": [30],
            "Strain %": [6.25],
        }
    ).to_excel(strain_path, index=False)
    strain_records = core._load_strain_records([strain_path], logging.getLogger("test"))

    legacy_columns = [
        "Composition",
        "Microwire",
        "d (µm)",
        "D (µm)",
        "d/D",
        "Length (m)",
        "Figure — 1000 mA",
        "Figure — low mA",
    ]
    legacy_row = {
        "Composition": "Ni55Fe18Ga27",
        "Microwire": "4/1",
        "d (µm)": 8.0,
        "D (µm)": 40.0,
        "d/D": 0.2,
        "Length (m)": 5.0,
        "Figure — 1000 mA": "high.png",
        "Figure — low mA": "low.png",
    }

    csv_path = tmp_path / "legacy.csv"
    pd.DataFrame([legacy_row], columns=legacy_columns).to_csv(csv_path, index=False)
    core._update_existing_csv_with_strain(csv_path, strain_records, core.OUTPUT_COLUMNS, logging.getLogger("test"))
    updated_csv = pd.read_csv(csv_path)
    assert updated_csv[core.STRAIN_COLUMN].iloc[0] == "6.250%"

    excel_path = tmp_path / "legacy.xlsx"
    pd.DataFrame([legacy_row], columns=legacy_columns).to_excel(excel_path, index=False)
    core._update_existing_excel_with_strain(excel_path, strain_records, logging.getLogger("test"))
    updated_excel = pd.read_excel(excel_path)
    columns = updated_excel.columns.tolist()
    figure_idx = columns.index("Figure — 1000 mA")
    assert columns[figure_idx - 1] == "d/D"
    assert columns[figure_idx + 1] == "Figure — low mA"
    assert columns[figure_idx + 2] == core.STRAIN_COLUMN
    assert updated_excel[core.STRAIN_COLUMN].iloc[0] == "6.250%"
