"""Core data processing for the microwire database builder."""

from __future__ import annotations

import csv
import hashlib
import importlib.util
import sys
import logging
import math
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import numpy as np
import pandas as pd

try:
    from .video import extract_video_metrics
except ImportError:
    module_name = "experiments.microwire_data_builder.video"
    module_path = Path(__file__).with_name("video.py")
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec and spec.loader:
        module = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = module
        spec.loader.exec_module(module)
        extract_video_metrics = module.extract_video_metrics
    else:
        raise

try:
    from .ocr import ensure_tesseract_available
except ImportError:
    module_name = "experiments.microwire_data_builder.ocr"
    module_path = Path(__file__).with_name("ocr.py")
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec and spec.loader:
        module = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = module
        spec.loader.exec_module(module)
        ensure_tesseract_available = module.ensure_tesseract_available
    else:
        raise

sys.modules.setdefault("experiments.microwire_data_builder.core", sys.modules.get(__name__))

os.environ.setdefault("MPLBACKEND", "Agg")

LOGGER_NAME = "microwire_data_builder"
R_CHECK_THRESHOLD = 0.05
DEFAULT_OUTPUT_NAME = "microwire_database"
PLOT_DIR_NAME = "plots"
ORIGIN_DIR_NAME = "origin_objects"

MICRO_SIGN = "µ"

OUTPUT_COLUMNS = [
    "Composition",
    "Microwire",
    "d (µm)",
    "D (µm)",
    "d/D",
    "Length (m)",
    "Production datetime",
    "Mass (g)",
    "Resistance (Ω)",
    "Temperature (°C)",
    "Winding speed (m/min)",
    "Glass feeding (mm/min)",
    "Underpressure",
    "Notes",
    "Figure — 1000 mA",
    "Figure — low mA",
    "Figure — 1000 mA (Origin)",
    "Figure — low mA (Origin)",
    "Low mA value (mA)",
    "File 1000 mA",
    "File low mA",
]

FIGURE_COLUMNS = (
    "Figure — 1000 mA",
    "Figure — low mA",
)

ORIGIN_FIGURE_COLUMNS = tuple(
    column
    for column in OUTPUT_COLUMNS
    if column.startswith("Figure") and "(Origin)" in column
)

_INVALID_FILENAME_CHARS = set('<>:"/\\|?*')

DRAW_NUMERIC_FIELDS = {
    "mass_g",
    "fabrication_resistance_ohm",
    "winding_speed_m_per_min",
    "glass_feed_mm_per_min",
    "underpressure",
}
PIECE_NUMERIC_FIELDS = {
    "piece_turns",
    "length_m",
    "d_um",
    "D_um",
    "d_over_D",
}
DATETIME_FIELDS = {"production_datetime"}
DATE_FIELDS = {"piece_date"}
RAW_VALUE_FIELDS = (
    DRAW_NUMERIC_FIELDS
    | PIECE_NUMERIC_FIELDS
    | DATETIME_FIELDS
    | DATE_FIELDS
    | {"fabrication_temperature_c"}
)

HEADER_HINTS: Dict[str, str] = {
    "zloženie": "composition_label",
    "zlozenie": "composition_label",
    "composition": "composition_label",
    "dátum a čas výroby": "production_datetime",
    "datum a cas vyroby": "production_datetime",
    "datumacasvyroby": "production_datetime",
    "hmotnosť": "mass_g",
    "hmotnost": "mass_g",
    "mass": "mass_g",
    "odpor": "fabrication_resistance_ohm",
    "resistance": "fabrication_resistance_ohm",
    "teplota": "fabrication_temperature_c",
    "temperature": "fabrication_temperature_c",
    "winding speed (m/min)": "winding_speed_m_per_min",
    "glass feeding (mm/min)": "glass_feed_mm_per_min",
    "underpressure": "underpressure",
    "bistabilny/nebistabilny": "bistable_status",
    "poznámka": "notes",
    "Poznámka": "notes",
    "poznamka": "notes",
    "Poznamka": "notes",
    "pozn.": "notes",
    "pozn": "notes",
    "poznámky": "notes",
    "p.č": "piece_y",
    "p.c": "piece_y",
    "p.č.": "piece_y",
    "počet otáčok": "piece_turns",
    "pocet otacok": "piece_turns",
    "dĺžka (m)": "length_m",
    "dlzka (m)": "length_m",
    "d (µm)": "d_um",
    "d (um)": "d_um",
    "d (μm)": "d_um",
    "d(µm)": "d_um",
    "d(um)": "d_um",
    "d(μm)": "d_um",
    "D (µm)": "D_um",
    "D (um)": "D_um",
    "D (μm)": "D_um",
    "D(µm)": "D_um",
    "D(um)": "D_um",
    "D(μm)": "D_um",
    "d/D": "d_over_D",
    "d/d": "d_over_D",
    "Datum": "piece_date",
    "Dátum": "piece_date",
    "datum": "piece_date",
    "dátum": "piece_date",
}

ANNEALING_COLUMNS = ["I_A", "V_V", "R_ohm"]

DRAW_PATTERN = re.compile(r"^(?P<draw>\d+)")
PIECE_PATTERN = re.compile(r"^(?P<piece>\d+)")
XY_PATTERN = re.compile(r"(\d+)_+(\d+)")
SETPOINT_PATTERN = re.compile(r"(\d{1,4})mA", re.IGNORECASE)
ALT_VARIANT_PATTERN = re.compile(r"(?:s\d+|\d+_\d+)a(?!\w)", re.IGNORECASE)
DOT_DATE_PATTERN = re.compile(r"\d{1,2}\.\d{1,2}\.\d{2,4}")
SLASH_DATE_PATTERN = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")

MICROSCOPE_VALUE_PATTERN = re.compile(
    rf"(?P<value>\d+(?:[.,]\d+)?)\s*(?:u?m|{MICRO_SIGN}m)",
    re.IGNORECASE,
)

MICROSCOPE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp")


def _normalise_figsize(value: Sequence[float] | Tuple[float, float] | None) -> Tuple[float, float]:
    default = (4.0, 2.25)
    if not value:
        return default
    try:
        width, height = value  # type: ignore[misc]
    except Exception:
        return default
    try:
        width_f = float(width)
        height_f = float(height)
    except Exception:
        return default
    if not math.isfinite(width_f) or width_f <= 0:
        width_f = default[0]
    if not math.isfinite(height_f) or height_f <= 0:
        height_f = default[1]
    return (max(0.5, width_f), max(0.5, height_f))

EXCEL_EMBED_DPI = 150.0


def _excel_pixel_limits(figure_size: Tuple[float, float]) -> Tuple[int, int]:
    width_in, height_in = figure_size
    width_px = max(int(round(width_in * EXCEL_EMBED_DPI)), 1)
    height_px = max(int(round(height_in * EXCEL_EMBED_DPI)), 1)
    return width_px, height_px


def _excel_row_height(height_in: float) -> float:
    return max(height_in * 72.0, 18.0)


def _excel_column_width(width_in: float) -> float:
    return max(width_in * 7.0, 12.0)


@dataclass
class BuilderConfig:
    """Configuration for the database builder."""

    fabrication_files: List[Path]
    annealing_files: List[Path]
    output_dir: Path
    microscope_files: List[Path] = field(default_factory=list)
    video_files: List[Path] = field(default_factory=list)
    make_plots: bool = False
    export_formats: Tuple[str, ...] = ("csv",)
    plot_dir_name: str = PLOT_DIR_NAME
    origin_dir_name: str = ORIGIN_DIR_NAME
    output_name: str = DEFAULT_OUTPUT_NAME
    plot_backends: Tuple[str, ...] = ()
    export_behaviour: Optional[Dict[str, str]] = None
    matplotlib_figsize: Tuple[float, float] = (4.0, 2.25)


@dataclass
class BuildStats:
    """Accumulates processing statistics."""

    parsed: int = 0
    skipped: int = 0
    missing_draw: int = 0
    missing_piece: int = 0
    resistance_checks_failed: int = 0
    rows_built: int = 0
    missing_high_measurement: int = 0
    missing_low_measurement: int = 0


@dataclass
class OriginArtifact:
    """Metadata describing an Origin object exported for Excel embedding."""

    descriptor: str
    object_path: Optional[Path]
    graph_name: Optional[str] = None
    workbook_name: Optional[str] = None
    worksheet_name: Optional[str] = None
    display_text: Optional[str] = None


@dataclass
class BuildResult:
    """Return value from :func:`build_database`."""

    dataframe: pd.DataFrame
    exports: Dict[str, Path]
    plot_paths: List[Path]
    origin_artifacts: Dict[str, OriginArtifact]
    stats: BuildStats


class FabricationIndex:
    """Lookup tables populated from fabrication spreadsheets."""

    def __init__(self) -> None:
        self.draw_level: Dict[Tuple[str, int], Dict[str, object]] = {}
        self.piece_level: Dict[Tuple[str, int, int], Dict[str, object]] = {}

    def set_draw(self, composition: str, draw_x: int, data: Dict[str, object]) -> None:
        key = (composition, draw_x)
        existing = self.draw_level.get(key, {})
        existing.update(data)
        self.draw_level[key] = existing

    def set_piece(self, composition: str, draw_x: int, piece_y: int, data: Dict[str, object]) -> None:
        key = (composition, draw_x, piece_y)
        existing = self.piece_level.get(key, {})
        existing.update(data)
        self.piece_level[key] = existing

    def get_draw(self, composition: str, draw_x: Optional[int]) -> Dict[str, object]:
        if draw_x is None:
            return {}
        return self.draw_level.get((composition, draw_x), {})

    def get_piece(self, composition: str, draw_x: Optional[int], piece_y: Optional[int]) -> Dict[str, object]:
        if draw_x is None or piece_y is None:
            return {}
        return self.piece_level.get((composition, draw_x, piece_y), {})


@dataclass
class MicroscopeMeasurements:
    """Diameter samples gathered from microscope images."""

    core: List[float] = field(default_factory=list)
    glass: List[float] = field(default_factory=list)
    other: List[float] = field(default_factory=list)

    def extend(self, category: str, values: Iterable[float]) -> None:
        target: List[float]
        if category == "core":
            target = self.core
        elif category == "glass":
            target = self.glass
        else:
            target = self.other
        for value in values:
            if isinstance(value, (int, float)) and math.isfinite(value) and value > 0:
                target.append(float(value))

    def best_core(self) -> Optional[float]:
        for pool in (self.core, self.other, self.glass):
            candidate = _select_microscope_value(pool, prefer="min")
            if candidate is not None:
                return candidate
        return None

    def best_glass(self) -> Optional[float]:
        for pool in (self.glass, self.other, self.core):
            candidate = _select_microscope_value(pool, prefer="max")
            if candidate is not None:
                return candidate
        return None


@dataclass
class VideoMetricsSummary:
    """Aggregated metrics extracted from fabrication videos."""

    temperatures: List[float] = field(default_factory=list)
    underpressures: List[float] = field(default_factory=list)

    def record(self, result) -> None:
        temp = getattr(result, "median_temperature", None)
        if callable(temp):
            value = temp()
        else:
            value = None
        if value is not None and isinstance(value, (int, float)) and math.isfinite(float(value)):
            self.temperatures.append(float(value))
        under_fn = getattr(result, "median_underpressure", None)
        if callable(under_fn):
            under_value = under_fn()
        else:
            under_value = None
        if under_value is not None and isinstance(under_value, (int, float)) and math.isfinite(float(under_value)):
            self.underpressures.append(float(under_value))

    def temperature(self) -> Optional[float]:
        return _select_microscope_value(self.temperatures, "median", allow_negative=True)

    def underpressure(self) -> Optional[float]:
        return _select_microscope_value(self.underpressures, "median", allow_negative=True)


def _logger(logger: Optional[logging.Logger]) -> logging.Logger:
    if logger is not None:
        return logger
    return logging.getLogger(LOGGER_NAME)


def _normalise_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    return text


def _header_key(value: object) -> Optional[str]:
    text = _normalise_text(value)
    if not text:
        return None
    hint = HEADER_HINTS.get(text)
    if hint:
        return hint
    ascii_text = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(ch for ch in ascii_text if not unicodedata.combining(ch))
    lowered = ascii_text.lower()
    simple = lowered.replace("\u00a0", " ")
    simple = re.sub(r"\s+", " ", simple)
    simple_compact = re.sub(r"[^a-z0-9]+", "", lowered)
    if "zlozen" in lowered:
        return "composition_label"
    if "datum" in lowered and "cas" in lowered:
        return "production_datetime"
    if "hmotn" in lowered or "mass" in lowered:
        return "mass_g"
    if "odpor" in lowered or "resistance" in lowered:
        return "fabrication_resistance_ohm"
    if "teplot" in lowered or "temp" in lowered:
        return "fabrication_temperature_c"
    if "winding" in lowered:
        return "winding_speed_m_per_min"
    if "glass" in lowered and ("feed" in lowered or "feeding" in lowered):
        return "glass_feed_mm_per_min"
    if "underpressure" in lowered:
        return "underpressure"
    if "bistabil" in lowered:
        return "bistable_status"
    if "p.c" in lowered or "p.č" in lowered or simple_compact == "pc":
        return "piece_y"
    if "pocet" in lowered and "otac" in lowered:
        return "piece_turns"
    if "dlzk" in lowered or "dlžk" in lowered:
        return "length_m"
    if lowered.strip().startswith("d") and "µm" in lowered:
        first = ascii_text.strip()[:1]
        if first == "D":
            return "D_um"
        return "d_um"
    if lowered.strip().startswith("d") and "um" in lowered and "µ" not in lowered:
        first = ascii_text.strip()[:1]
        if first == "D":
            return "D_um"
        return "d_um"
    if "d/d" in lowered or simple_compact == "dd":
        return "d_over_D"
    if ("datum" in lowered or "dátum" in lowered) and "cas" not in lowered:
        return "piece_date"
    return None


def _is_nan(value: object) -> bool:
    return isinstance(value, float) and math.isnan(value)


def _raw_string(value: object) -> Optional[str]:
    if value is None or _is_nan(value):
        return None
    text = _normalise_text(value)
    return text if text else None


def _parse_numeric(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not _is_nan(value):
        return float(value)
    text = _normalise_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    number = match.group(0).replace(",", ".")
    try:
        return float(number)
    except ValueError:
        return None


def _parse_datetime(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    text = _normalise_text(value)
    if not text:
        return None
    try:
        dayfirst: Optional[bool]
        if DOT_DATE_PATTERN.search(text):
            dayfirst = True
        elif SLASH_DATE_PATTERN.search(text):
            dayfirst = False
        else:
            dayfirst = None
        if dayfirst is None:
            dt = pd.to_datetime(text, dayfirst=False, errors="coerce")
            if pd.isna(dt):
                dt = pd.to_datetime(text, dayfirst=True, errors="coerce")
        else:
            dt = pd.to_datetime(text, dayfirst=dayfirst, errors="coerce")
    except (TypeError, ValueError):
        return None
    if pd.isna(dt):
        return None
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    if isinstance(dt, datetime):
        return dt.isoformat(sep=" ")
    return None


def _parse_date(value: object) -> Optional[str]:
    parsed = _parse_datetime(value)
    if not parsed:
        return None
    try:
        dt = datetime.fromisoformat(parsed)
    except ValueError:
        return parsed
    return dt.date().isoformat()


def _select_microscope_value(
    values: Sequence[float], prefer: str, *, allow_negative: bool = False
) -> Optional[float]:
    cleaned: List[float] = []
    for v in values:
        if not isinstance(v, (int, float)):
            continue
        value = float(v)
        if not math.isfinite(value):
            continue
        if not allow_negative and value <= 0:
            continue
        cleaned.append(value)
    if not cleaned:
        return None
    if prefer == "min":
        return float(min(cleaned))
    if prefer == "max":
        return float(max(cleaned))
    cleaned.sort()
    mid = len(cleaned) // 2
    if len(cleaned) % 2:
        return float(cleaned[mid])
    return float((cleaned[mid - 1] + cleaned[mid]) / 2.0)


def _extract_field_value(field: str, value: object) -> Tuple[Optional[object], Optional[str]]:
    raw = _raw_string(value)
    if field in DATETIME_FIELDS:
        return _parse_datetime(value), raw
    if field in DATE_FIELDS:
        return _parse_date(value), raw
    if field in DRAW_NUMERIC_FIELDS or field in PIECE_NUMERIC_FIELDS:
        return _parse_numeric(value), raw
    if field == "fabrication_temperature_c":
        numeric = _parse_numeric(value)
        return numeric if numeric is not None else raw, raw
    if value is None or _is_nan(value):
        return None, raw
    return _normalise_text(value), raw


def _composition_from_path(path: Path) -> str:
    stem = path.stem
    return stem.split()[0]


def _extract_composition_token(text: str) -> Optional[str]:
    if not text:
        return None
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9]+", text)
    if not tokens:
        return None
    return tokens[0]


def _microscope_key(path: Path) -> Optional[Tuple[str, int, int]]:
    def _match(text: str) -> Optional[Tuple[str, int, int]]:
        if not text:
            return None
        match = XY_PATTERN.search(text)
        if not match:
            return None
        composition = _extract_composition_token(text)
        if not composition:
            return None
        try:
            draw_x = int(match.group(1))
            piece_y = int(match.group(2))
        except (TypeError, ValueError):
            return None
        return composition, draw_x, piece_y

    for candidate in (path.stem, path.name):
        result = _match(candidate)
        if result is not None:
            return result
    for parent in path.parents:
        result = _match(parent.name)
        if result is not None:
            return result
    return None


def _microscope_category(path: Path) -> str:
    stem = path.stem.lower()
    if "core" in stem:
        return "core"
    if "glass" in stem:
        return "glass"
    return "other"


def _iter_fragment_files(
    root: Path,
    fragment: str,
    extensions: Sequence[str],
    *,
    max_depth: int = 3,
    limit: Optional[int] = None,
) -> List[Path]:
    if root is None:
        return []
    try:
        if not root.exists():
            return []
    except OSError:
        return []

    fragment_lower = fragment.lower()
    matches: List[Path] = []
    stack: List[Tuple[Path, int]] = [(root, 0)]
    visited: Set[Path] = set()
    while stack:
        current, depth = stack.pop()
        try:
            entries = list(current.iterdir())
        except OSError:
            continue
        for entry in entries:
            try:
                if entry.is_dir():
                    if depth >= max_depth:
                        continue
                    try:
                        key = entry.resolve()
                    except OSError:
                        key = entry
                    if key in visited:
                        continue
                    visited.add(key)
                    name_lower = entry.name.lower()
                    if depth == 0 or fragment_lower in name_lower:
                        stack.append((entry, depth + 1))
                elif entry.is_file():
                    if entry.suffix.lower() in extensions and fragment_lower in entry.name.lower():
                        matches.append(entry)
                        if limit is not None and len(matches) >= limit:
                            return matches
            except OSError:
                continue
    return matches


def _auto_discover_microscope_paths(
    annealing_files: Sequence[Path], logger: logging.Logger
) -> List[Path]:
    discovered: List[Path] = []
    seen: Set[Path] = set()
    for raw_path in annealing_files:
        path = Path(raw_path)
        key = _microscope_key(path)
        if key is None:
            continue
        composition, draw_x, piece_y = key
        fragment = f"{draw_x}_{piece_y}"
        try:
            resolved = path.resolve()
            parents = list(resolved.parents)
        except OSError:
            parents = list(path.parents)
        candidate_roots: List[Path] = []
        for parent in parents:
            for name in ("microscope", "Microscope"):
                candidate = parent / name
                try:
                    if candidate.is_dir():
                        candidate_roots.append(candidate)
                except OSError:
                    continue
            if candidate_roots:
                break
        if not candidate_roots:
            continue
        lowered = composition.lower()
        search_dirs: List[Path] = []
        for root_dir in candidate_roots:
            search_dirs.append(root_dir)
            try:
                direct = root_dir / composition
                if direct.is_dir():
                    search_dirs.append(direct)
                for child in root_dir.iterdir():
                    if child.is_dir() and child.name.lower().startswith(lowered):
                        search_dirs.append(child)
            except OSError:
                continue
        for search_dir in dict.fromkeys(search_dirs):
            matches = _iter_fragment_files(search_dir, fragment, MICROSCOPE_EXTENSIONS, limit=20)
            for match in matches:
                try:
                    resolved_match = match.resolve()
                except OSError:
                    resolved_match = match
                if resolved_match in seen:
                    continue
                seen.add(resolved_match)
                discovered.append(match)
    if discovered:
        logger.debug("Auto-discovered %s microscope image(s)", len(discovered))
    return discovered


def _extract_microscope_diameters(path: Path, logger: logging.Logger) -> List[float]:
    """Attempt to OCR diameter annotations from a microscope capture."""

    log = logger or logging.getLogger(LOGGER_NAME)
    try:
        import pytesseract  # type: ignore[import-not-found]
    except ImportError:
        log.debug("pytesseract is not installed; skipping microscope OCR for %s", path)
        return []

    if not ensure_tesseract_available(pytesseract, log):
        log.debug("Tesseract executable unavailable; skipping microscope OCR for %s", path)
        return []

    try:
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps  # type: ignore[import-not-found]
    except ImportError:
        log.debug("Pillow is not installed; skipping microscope OCR for %s", path)
        return []

    try:
        with Image.open(path) as img:
            base = img.convert("RGB")
    except Exception:
        log.warning("Failed to open microscope image %s", path, exc_info=True)
        return []

    def _resample(image):
        width, height = image.size
        target = 1600
        longest = max(width, height)
        if longest >= target:
            return image
        scale = target / float(longest)
        new_size = (int(width * scale), int(height * scale))
        resample_attr = getattr(Image, "Resampling", None)
        if resample_attr is not None:
            resample_filter = getattr(resample_attr, "LANCZOS", Image.BICUBIC)
        else:
            resample_filter = getattr(Image, "LANCZOS", Image.BICUBIC)
        return image.resize(new_size, resample_filter)

    candidates: List[str] = []
    grayscale = ImageOps.grayscale(base)
    resized = _resample(grayscale)
    enhanced = ImageEnhance.Contrast(resized).enhance(2.0)
    sharpened = enhanced.filter(ImageFilter.UnsharpMask(radius=2, percent=175))
    variants = [resized, enhanced, sharpened, ImageOps.autocontrast(resized)]

    tesseract_not_found = getattr(pytesseract, "TesseractNotFoundError", RuntimeError)
    for variant in variants:
        try:
            text = pytesseract.image_to_string(variant, config="--psm 6")
        except tesseract_not_found:
            log.warning(
                "Tesseract executable vanished while processing %s; aborting microscope OCR",
                path,
            )
            return []
        except Exception:
            continue
        if text:
            candidates.append(text)

    unique: Dict[float, None] = {}
    for text in candidates:
        for match in MICROSCOPE_VALUE_PATTERN.finditer(text):
            raw = match.group("value").replace(",", ".")
            try:
                value = float(raw)
            except ValueError:
                continue
            if not math.isfinite(value) or value <= 0:
                continue
            key = round(value, 2)
            unique.setdefault(key, value)

    return [float(v) for v in unique.values()]


def _group_microscope_measurements(
    microscope_files: Sequence[Path],
    annealing_files: Sequence[Path],
    logger: Optional[logging.Logger],
) -> Dict[Tuple[str, int, int], MicroscopeMeasurements]:
    log = _logger(logger)
    auto_paths = _auto_discover_microscope_paths(annealing_files, log)
    combined = list(dict.fromkeys([Path(p) for p in microscope_files] + auto_paths))
    grouped: Dict[Tuple[str, int, int], MicroscopeMeasurements] = {}
    for raw_path in combined:
        path = raw_path.expanduser()
        try:
            if not path.exists():
                log.debug("Microscope image %s does not exist; skipping", path)
                continue
        except OSError:
            continue
        key = _microscope_key(path)
        if key is None:
            log.debug("Unable to derive microwire key from microscope image %s", path)
            continue
        category = _microscope_category(path)
        try:
            values = _extract_microscope_diameters(path, log)
        except Exception:
            log.exception("Microscope OCR failed for %s", path)
            continue
        if not values:
            continue
        record = grouped.setdefault(key, MicroscopeMeasurements())
        record.extend(category, values)
    return grouped


def _draw_key(path: Path) -> Optional[Tuple[str, int]]:
    candidates: List[str] = [parent.name for parent in path.parents]
    candidates.extend([path.stem, path.name])
    for candidate in candidates:
        if not candidate:
            continue
        composition = _extract_composition_token(candidate)
        if not composition:
            continue
        draw_match = re.search(r"(\d+)", candidate)
        if not draw_match:
            continue
        try:
            draw_x = int(draw_match.group(1))
        except (TypeError, ValueError):
            continue
        return composition, draw_x
    return None


def _collect_video_metrics(
    video_files: Sequence[Path], logger: Optional[logging.Logger]
) -> Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary]:
    log = _logger(logger)
    aggregated: Dict[Tuple[str, int, Optional[int]], VideoMetricsSummary] = {}
    for raw_path in dict.fromkeys(Path(p) for p in video_files):
        path = raw_path.expanduser()
        try:
            if not path.exists():
                log.debug("Video file %s does not exist; skipping", path)
                continue
        except OSError:
            continue
        key = _microscope_key(path)
        if key is not None:
            composition, draw_x, piece_y = key
        else:
            draw_key = _draw_key(path)
            if draw_key is None:
                log.debug("Unable to derive microwire key from video %s", path)
                continue
            composition, draw_x = draw_key
            piece_y = None
        try:
            result = extract_video_metrics(path, logger=log)
        except Exception:
            log.exception("Failed to analyse video %s", path)
            continue
        summary = aggregated.setdefault((composition, draw_x, piece_y), VideoMetricsSummary())
        summary.record(result)
    return aggregated


def _parse_draw_rows(
    df: pd.DataFrame,
    headers: List[Optional[str]],
    composition: str,
    index: FabricationIndex,
    logger: logging.Logger,
) -> None:
    seen_draws: set[int] = set()
    for _, row in df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if first_cell is None or _is_nan(first_cell):
            continue
        text = _normalise_text(first_cell)
        if not text:
            continue
        draw_x: Optional[int] = None
        m = DRAW_PATTERN.match(text)
        if m:
            draw_x = int(m.group("draw"))
        elif text.lower().startswith(composition.lower()):
            draw_x = 1
        if draw_x is None:
            continue
        if draw_x in seen_draws:
            logger.debug("Duplicate draw %s in %s", draw_x, composition)
        seen_draws.add(draw_x)
        record: Dict[str, object] = {}
        for col_idx, field in enumerate(headers):
            if col_idx == 0 or not field:
                continue
            value = row.iloc[col_idx] if col_idx < len(row) else None
            parsed, raw = _extract_field_value(field, value)
            record[field] = parsed
            if field in RAW_VALUE_FIELDS:
                record[f"{field}_raw"] = raw
        index.set_draw(composition, draw_x, record)


def _parse_piece_rows(
    df: pd.DataFrame,
    headers: List[Optional[str]],
    composition: str,
    draw_x: Optional[int],
    index: FabricationIndex,
    logger: logging.Logger,
) -> None:
    if draw_x is None:
        logger.warning("Could not determine draw number for piece workbook %s", composition)
        return
    for _, row in df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if first_cell is None or _is_nan(first_cell):
            continue
        text = _normalise_text(first_cell)
        if not text:
            continue
        m = PIECE_PATTERN.match(text)
        if not m:
            continue
        piece_y = int(m.group("piece"))
        record: Dict[str, object] = {}
        for col_idx, field in enumerate(headers):
            if col_idx == 0 or not field:
                continue
            value = row.iloc[col_idx] if col_idx < len(row) else None
            parsed, raw = _extract_field_value(field, value)
            record[field] = parsed
            if field in RAW_VALUE_FIELDS:
                record[f"{field}_raw"] = raw
        index.set_piece(composition, draw_x, piece_y, record)


def _read_excel(path: Path) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, header=None, dtype=object)
    except ImportError as exc:  # pragma: no cover - pandas provides helpful message
        raise
    except ValueError as exc:
        raise
    return df


def _parse_composition_workbook(path: Path, index: FabricationIndex, logger: logging.Logger) -> None:
    df = _read_excel(path)
    if df.empty:
        logger.warning("%s is empty", path)
        return
    header_row = df.iloc[0]
    headers = [_header_key(value) for value in header_row]
    data = df.iloc[1:].reset_index(drop=True)
    composition = _composition_from_path(path)
    _parse_draw_rows(data, headers, composition, index, logger)


def _parse_piece_workbook(path: Path, index: FabricationIndex, logger: logging.Logger) -> None:
    df = _read_excel(path)
    if df.empty:
        logger.warning("%s is empty", path)
        return
    header_idx = None
    for idx, row in df.iterrows():
        values = [_normalise_text(cell).lower() for cell in row.tolist()]
        if any("p.c" in v or "p.č" in v or v == "pc" for v in values if v):
            header_idx = idx
            break
    if header_idx is None:
        logger.warning("%s: unable to locate header row", path)
        return
    headers = [_header_key(value) for value in df.iloc[header_idx]]
    data = df.iloc[header_idx + 1 :].reset_index(drop=True)
    stem = path.stem
    match = re.search(r"(?P<draw>\d+)[._](?P<comp>[A-Za-z0-9]+)", stem)
    draw_x: Optional[int] = None
    composition = _composition_from_path(path)
    if match:
        draw_x = int(match.group("draw"))
        composition = match.group("comp")
    _parse_piece_rows(data, headers, composition, draw_x, index, logger)


def build_fabrication_index(
    fabrication_files: Sequence[Path],
    logger: Optional[logging.Logger] = None,
) -> FabricationIndex:
    log = _logger(logger)
    index = FabricationIndex()
    for path in fabrication_files:
        if path.suffix.lower() != ".xlsx":
            log.debug("Skipping non-Excel file %s", path)
            continue
        stem = path.stem
        parent_stem = path.parent.name
        if re.match(r"^\d+", stem) or re.match(r"^\d+", parent_stem):
            _parse_piece_workbook(path, index, log)
        else:
            _parse_composition_workbook(path, index, log)
    return index


@dataclass
class MeasurementMetadata:
    composition_token: str
    draw_x: Optional[int]
    piece_y: Optional[int]
    setpoint_mA: Optional[int]
    alt_variant: bool
    measurement_id: str
    file_name: str
    relpath: str
    timestamp_mtime_utc: str


@dataclass
class MeasurementRecord:
    path: Path
    metadata: MeasurementMetadata
    dataframe: pd.DataFrame
    sanity_ok: bool
    sanity_error: Optional[float]


def _hash_file(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _metadata_from_path(path: Path, root: Optional[Path] = None) -> MeasurementMetadata:
    stat = path.stat()
    timestamp = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
    base = path.stem
    parts = base.split()
    composition = parts[0] if parts else base
    xy_match = XY_PATTERN.search(base)
    draw_x: Optional[int] = int(xy_match.group(1)) if xy_match else None
    piece_y: Optional[int] = int(xy_match.group(2)) if xy_match else None
    setpoint_match = SETPOINT_PATTERN.search(base)
    setpoint = int(setpoint_match.group(1)) if setpoint_match else None
    alt_variant = bool(ALT_VARIANT_PATTERN.search(base))
    relpath = os.fspath(path.relative_to(root)) if root and path.is_relative_to(root) else path.as_posix()
    return MeasurementMetadata(
        composition_token=composition,
        draw_x=draw_x,
        piece_y=piece_y,
        setpoint_mA=setpoint,
        alt_variant=alt_variant,
        measurement_id=_hash_file(path),
        file_name=path.name,
        relpath=relpath,
        timestamp_mtime_utc=timestamp,
    )


def _load_annealing(path: Path) -> pd.DataFrame:
    try:
        df = pd.read_csv(path, sep=None, engine="python", names=ANNEALING_COLUMNS, header=None)
    except (csv.Error, pd.errors.ParserError):
        df = pd.read_csv(
            path,
            sep=r"\s+",
            engine="python",
            names=ANNEALING_COLUMNS,
            header=None,
        )
    df = df.iloc[:, :3].copy()
    for column in ANNEALING_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce")
    df = df.dropna(subset=["I_A", "R_ohm"]).reset_index(drop=True)

    try:
        from plotting.current_annealing.burnthrough import trim_burnthrough_glitch
    except ImportError:
        return df

    currents_mA = df["I_A"].to_numpy(dtype=float) * 1e3
    resistances = df["R_ohm"].to_numpy(dtype=float)
    trimmed_currents, trimmed_resistances = trim_burnthrough_glitch(currents_mA, resistances)
    trimmed_count = int(trimmed_currents.shape[0])
    if trimmed_count < currents_mA.shape[0]:
        df = df.iloc[:trimmed_count].copy()
        df.loc[:, "I_A"] = trimmed_currents / 1e3
        df.loc[:, "R_ohm"] = trimmed_resistances
        df = df.reset_index(drop=True)
    return df


def _resistance_sanity_check(df: pd.DataFrame) -> Tuple[bool, Optional[float]]:
    currents = df["I_A"].to_numpy(dtype=float)
    voltages = df["V_V"].to_numpy(dtype=float)
    resistances = df["R_ohm"].to_numpy(dtype=float)
    with np.errstate(divide="ignore", invalid="ignore"):
        expected = np.divide(voltages, currents, out=np.full_like(resistances, np.nan), where=currents != 0)
    denom = np.maximum(np.abs(resistances), 1e-9)
    rel_error = np.abs(expected - resistances) / denom
    finite = rel_error[np.isfinite(rel_error)]
    if finite.size == 0:
        return False, None
    mean_error = float(np.mean(finite))
    return mean_error < R_CHECK_THRESHOLD, mean_error


def _value_for_output(record: Dict[str, object], field: str) -> Optional[object]:
    if not record:
        return None
    value = record.get(field)
    if value is None or _is_nan(value):
        raw = record.get(f"{field}_raw")
        if raw is None:
            return None
        text = str(raw).strip()
        return text or None
    return value


def _compose_notes(*records: Dict[str, object]) -> Optional[str]:
    notes: List[str] = []
    seen: set[str] = set()
    for record in records:
        if not record:
            continue
        for key in ("bistable_status", "notes"):
            candidate = _value_for_output(record, key)
            if candidate is None:
                continue
            text = str(candidate).strip()
            if not text:
                continue
            lowered = text.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            notes.append(text)
    return "; ".join(notes) if notes else None


def _microwire_label(draw_x: Optional[int], piece_y: Optional[int]) -> str:
    if draw_x is None or piece_y is None:
        return ""
    return f"{draw_x}/{piece_y}"


def _select_high_measurement(records: List[MeasurementRecord]) -> Optional[MeasurementRecord]:
    if not records:
        return None

    def key(record: MeasurementRecord) -> Tuple[int, int, int, str]:
        setpoint = record.metadata.setpoint_mA
        priority_exact = 1 if setpoint == 1000 else 0
        magnitude = setpoint if setpoint is not None else -1
        variant_score = 1 if not record.metadata.alt_variant else 0
        return (priority_exact, magnitude, variant_score, record.metadata.file_name.lower())

    return max(records, key=key)


def _select_low_measurement(records: List[MeasurementRecord]) -> Optional[MeasurementRecord]:
    if not records:
        return None
    candidates = [r for r in records if r.metadata.setpoint_mA is not None]
    if not candidates:
        return None

    def key(record: MeasurementRecord) -> Tuple[int, int, str]:
        setpoint = record.metadata.setpoint_mA or 0
        variant_penalty = 0 if not record.metadata.alt_variant else 1
        return (setpoint, variant_penalty, record.metadata.file_name.lower())

    return min(candidates, key=key)


def _plot_measurement_matplotlib(
    df: pd.DataFrame,
    source: Path,
    plot_dir: Path,
    figsize: Tuple[float, float],
) -> Path:
    import matplotlib

    try:
        matplotlib.use("Agg", force=True)
    except Exception:
        pass
    import matplotlib.pyplot as plt

    matplotlib.rcParams["figure.max_open_warning"] = 0

    from plotting.current_annealing.core import plot_one
    from plotting.utils import format_annealing_title

    plot_dir.mkdir(parents=True, exist_ok=True)
    title = format_annealing_title(source.stem)
    plot_df = pd.DataFrame({"I_mA": df["I_A"] * 1e3, "R_Ohm": df["R_ohm"]})
    fig, fname = plot_one(plot_df, title, figsize=figsize)
    safe_stem = _safe_plot_stem(fname)
    plot_path = plot_dir / f"{safe_stem}.png"
    plot_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(plot_path, dpi=300)
    plt.close(fig)
    return plot_path


def _plot_measurement_origin(
    df: pd.DataFrame,
    source: Path,
    origin_dir: Path,
    log: Optional[logging.Logger] = None,
) -> Optional[OriginArtifact]:
    try:
        from plotting.current_annealing.core import plot_one_origin
        from plotting.utils import format_annealing_title, schedule_origin_release
    except ImportError as exc:  # pragma: no cover - depends on optional module
        raise RuntimeError("originpro is not available") from exc

    plot_df = pd.DataFrame({"I_mA": df["I_A"] * 1e3, "R_Ohm": df["R_ohm"]})
    title = format_annealing_title(source.stem)
    handles = plot_one_origin(
        plot_df,
        title,
        source.name,
        return_handles=True,
    )
    try:
        schedule_origin_release()
    except Exception:
        pass

    if not isinstance(handles, dict):
        return None

    description, graph_name, workbook_name, worksheet_name = _describe_origin_handles(handles)
    safe_stem = _safe_plot_stem(source.stem)
    origin_dir.mkdir(parents=True, exist_ok=True)
    artifact_name = f"{safe_stem}.oggu"
    target_path = origin_dir / artifact_name
    exported_path = _export_origin_object(handles, target_path, log)

    return OriginArtifact(
        descriptor=artifact_name,
        object_path=exported_path,
        graph_name=graph_name,
        workbook_name=workbook_name,
        worksheet_name=worksheet_name,
        display_text=description,
    )


def _describe_origin_handles(
    handles: Dict[str, object],
) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    graph = handles.get("graph")
    workbook = handles.get("workbook")
    worksheet = handles.get("worksheet")
    legend_label = handles.get("legend_label")

    parts: list[str] = []
    if legend_label:
        parts.append(str(legend_label))
    graph_name = _origin_object_name(graph)
    if graph_name:
        parts.append(f"Graph: {graph_name}")
    workbook_name = _origin_object_name(workbook)
    if workbook_name:
        parts.append(f"Book: {workbook_name}")
    worksheet_name = _origin_object_name(worksheet)
    if worksheet_name:
        parts.append(f"Sheet: {worksheet_name}")
    description = " | ".join(parts) if parts else None
    return description, graph_name, workbook_name, worksheet_name


def _export_origin_object(
    handles: Dict[str, object], target_path: Path, log: Optional[logging.Logger]
) -> Optional[Path]:
    origin_any = handles.get("origin")
    graph = handles.get("graph")
    if origin_any is None or graph is None:
        return None

    target_path.parent.mkdir(parents=True, exist_ok=True)
    attempted: list[str] = []

    def _path_created() -> Optional[Path]:
        candidates = [
            target_path,
            target_path.with_suffix(".oggu"),
            target_path.with_suffix(".opju"),
        ]
        for candidate in candidates:
            try:
                if candidate.exists():
                    return candidate
            except OSError:
                continue
        return None

    for attr in ("save_copy", "save_as", "save"):
        method = getattr(graph, attr, None)
        if callable(method):
            try:
                method(str(target_path))
                created = _path_created()
                if created is not None:
                    return created
            except Exception:
                attempted.append(attr)

    for attr in ("save_page", "save_window"):
        method = getattr(origin_any, attr, None)
        if callable(method):
            try:
                method(str(target_path))
                created = _path_created()
                if created is not None:
                    return created
            except Exception:
                attempted.append(attr)

    lt_exec = getattr(origin_any, "lt_exec", None)
    if callable(lt_exec):
        graph_name = _origin_object_name(graph) or ""
        base = str(target_path)
        commands = []
        if graph_name:
            commands.append(f"page -s \"{graph_name}\"; save -oggu \"{base}\";")
            commands.append(f"page -s \"{graph_name}\"; save -opju \"{base}\";")
        commands.append(f"save -oggu \"{base}\";")
        commands.append(f"save -opju \"{base}\";")
        for command in commands:
            try:
                lt_exec(command)
                created = _path_created()
                if created is not None:
                    return created
            except Exception:
                attempted.append(command)

    if log is not None:
        log.warning(
            "Origin graph export failed for %s; attempted %s",
            target_path,
            ", ".join(attempted) if attempted else "no-export",
        )
    return None


def _embed_plots_in_excel_openpyxl(
    excel_path: Path,
    dataframe: pd.DataFrame,
    plot_name_to_path: Dict[str, Path],
    plot_dir: Path,
    log: logging.Logger,
    figure_size: Tuple[float, float],
) -> None:
    """Insert Matplotlib plot images directly into the Excel export."""

    if not excel_path.exists():
        return
    if dataframe.empty:
        return
    figure_columns = [column for column in FIGURE_COLUMNS if column in dataframe.columns]
    if not figure_columns:
        return

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover - optional dependency guard
        log.warning("openpyxl is required to embed plots into Excel workbooks")
        return

    available: Dict[str, Path] = {}
    for name, path in plot_name_to_path.items():
        if path.exists():
            available[name] = path
    if plot_dir and plot_dir.exists():
        for candidate in plot_dir.glob("*.png"):
            available.setdefault(candidate.name, candidate)
    if not available:
        return

    workbook = load_workbook(excel_path)
    try:
        worksheet = workbook.active

        inserted = False
        reset_df = dataframe.reset_index(drop=True)
        for row_idx, row in reset_df.iterrows():
            for column in figure_columns:
                value = row.get(column)
                if not isinstance(value, str):
                    continue
                name = value.strip()
                if not name:
                    continue
                image_path = available.get(name)
                if image_path is None or not image_path.exists():
                    candidate = plot_dir / name if plot_dir else None
                    if candidate is None or not candidate.exists():
                        continue
                    available[name] = candidate
                    image_path = candidate
                try:
                    image = XLImage(str(image_path))
                except Exception:
                    log.exception("Failed to load plot image %s for Excel export", image_path)
                    continue

                # Downscale very large images so they fit within the worksheet cells.
                max_width = max(int(round(figure_size[0] * 96)), 1)
                max_height = max(int(round(figure_size[1] * 96)), 1)
                if image.width and image.height:
                    width_scale = max_width / image.width if image.width > max_width else 1.0
                    height_scale = max_height / image.height if image.height > max_height else 1.0
                    scale = min(width_scale, height_scale)
                    if scale < 1.0:
                        image.width = int(image.width * scale)
                        image.height = int(image.height * scale)

                column_index = dataframe.columns.get_loc(column) + 1
                row_number = row_idx + 2  # account for the header row
                column_letter = get_column_letter(column_index)
                cell_reference = f"{column_letter}{row_number}"

                # Clear the textual filename and embed the image anchored at the cell.
                worksheet[cell_reference].value = None
                worksheet.add_image(image, cell_reference)

                # Adjust the row height and column width to accommodate the scaled figure.
                if image.height:
                    target_height = image.height * 0.75  # approximate px→pt conversion
                    current_height = worksheet.row_dimensions[row_number].height or 0
                    if target_height > current_height:
                        worksheet.row_dimensions[row_number].height = target_height
                if image.width:
                    approx_width = image.width / 7.0
                    current_width = worksheet.column_dimensions[column_letter].width or 0
                    if approx_width > current_width:
                        worksheet.column_dimensions[column_letter].width = approx_width

                inserted = True

        if inserted:
            workbook.save(excel_path)
    finally:
        workbook.close()


def _embed_assets_with_xlsxwriter(
    writer: "pd.ExcelWriter",
    dataframe: pd.DataFrame,
    plot_name_to_path: Dict[str, Path],
    plot_dir: Path,
    origin_artifacts: Dict[str, OriginArtifact],
    figure_size: Tuple[float, float],
    log: logging.Logger,
) -> None:
    try:
        workbook = writer.book
        worksheets = list(writer.sheets.values())
    except Exception:
        return
    if not worksheets:
        return
    worksheet = worksheets[0]

    figure_columns = [column for column in FIGURE_COLUMNS if column in dataframe.columns]
    origin_columns = [
        column
        for column in ("Figure — 1000 mA (Origin)", "Figure — low mA (Origin)")
        if column in dataframe.columns
    ]
    if not figure_columns and not origin_columns:
        return

    reset_df = dataframe.reset_index(drop=True)
    available: Dict[str, Path] = {}
    for name, path in plot_name_to_path.items():
        if path.exists():
            available[name] = path
    if plot_dir and plot_dir.exists():
        for candidate in plot_dir.glob("*.png"):
            available.setdefault(candidate.name, candidate)

    from PIL import Image as PILImage  # local import to avoid hard dependency elsewhere

    max_width = max(int(round(figure_size[0] * 96)), 1)
    max_height = max(int(round(figure_size[1] * 96)), 1)
    row_heights: Dict[int, float] = {}
    column_widths: Dict[int, float] = {}

    for row_idx, row in reset_df.iterrows():
        row_number = row_idx + 1
        for column in figure_columns:
            value = row.get(column)
            if not isinstance(value, str):
                continue
            name = value.strip()
            if not name:
                continue
            image_path = available.get(name)
            if image_path is None and plot_dir:
                candidate = plot_dir / name
                if candidate.exists():
                    available[name] = candidate
                    image_path = candidate
            if image_path is None or not image_path.exists():
                continue
            try:
                with PILImage.open(image_path) as pil_image:
                    width_px, height_px = pil_image.size
            except Exception:
                width_px = height_px = 0
            x_scale = min(1.0, max_width / width_px) if width_px else 1.0
            y_scale = min(1.0, max_height / height_px) if height_px else 1.0
            options: Dict[str, float] = {}
            if x_scale < 1.0:
                options["x_scale"] = x_scale
            if y_scale < 1.0:
                options["y_scale"] = y_scale
            column_index = dataframe.columns.get_loc(column)
            worksheet.write_blank(row_number, column_index, None)
            try:
                worksheet.insert_image(row_number, column_index, str(image_path), options)
            except Exception:
                log.exception("Failed to insert plot image %s", image_path)
                continue
            target_height = (height_px * y_scale * 0.75) if height_px else max_height * 0.75
            if target_height > row_heights.get(row_number, 0.0):
                worksheet.set_row(row_number, target_height)
                row_heights[row_number] = target_height
            approx_width = (width_px * x_scale / 7.0) if width_px else (max_width / 7.0)
            if approx_width > column_widths.get(column_index, 0.0):
                worksheet.set_column(column_index, column_index, approx_width)
                column_widths[column_index] = approx_width

        for column in origin_columns:
            value = row.get(column)
            if not isinstance(value, str):
                continue
            descriptor = value.strip()
            if not descriptor:
                continue
            artifact = origin_artifacts.get(descriptor)
            if artifact is None or artifact.object_path is None:
                continue
            try:
                if not artifact.object_path.exists():
                    continue
            except OSError:
                continue
            column_index = dataframe.columns.get_loc(column)
            worksheet.write_blank(row_number, column_index, None)
            options: Dict[str, object] = {"object_position": 1}
            try:
                worksheet.insert_object(row_number, column_index, str(artifact.object_path), options)
            except Exception:
                log.exception("Failed to insert Origin object %s", artifact.object_path)
                continue
            target_height = max(row_heights.get(row_number, 0.0), max_height * 0.75)
            worksheet.set_row(row_number, target_height)
            row_heights[row_number] = target_height
            approx_width = max(column_widths.get(column_index, 0.0), max_width / 7.0)
            worksheet.set_column(column_index, column_index, approx_width)
            column_widths[column_index] = approx_width

def _origin_object_name(obj: object) -> Optional[str]:
    if obj is None:
        return None
    for attr in ("lt_name", "name", "lname", "long_name"):
        value = getattr(obj, attr, None)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _normalise_output_name(name: str) -> str:
    cleaned = "".join("_" if ch in _INVALID_FILENAME_CHARS else ch for ch in name.strip())
    cleaned = cleaned.strip(".")
    if not cleaned:
        return DEFAULT_OUTPUT_NAME
    return cleaned


def _safe_plot_stem(stem: str) -> str:
    normalised = unicodedata.normalize("NFKC", stem)
    cleaned_chars: list[str] = []
    invalid = _INVALID_FILENAME_CHARS | {os.sep}
    if os.altsep:
        invalid.add(os.altsep)
    for ch in normalised:
        if ch in invalid or ord(ch) < 32:
            cleaned_chars.append("_")
        else:
            cleaned_chars.append(ch)
    cleaned = "".join(cleaned_chars).strip("._ ")
    if not cleaned:
        return "measurement"
    return cleaned


def build_database(
    config: BuilderConfig,
    logger: Optional[logging.Logger] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    root_for_relpaths: Optional[Path] = None,
) -> BuildResult:
    log = _logger(logger)
    output_dir = config.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = _normalise_output_name(getattr(config, "output_name", DEFAULT_OUTPUT_NAME))

    raw_backends = tuple(config.plot_backends) if config.plot_backends else ()
    cleaned_backends = []
    for backend in raw_backends:
        if isinstance(backend, str):
            cleaned_backends.append(backend.lower())
    if config.make_plots and not cleaned_backends:
        cleaned_backends.append("matplotlib")
    deduped_backends = tuple(dict.fromkeys(cleaned_backends))
    recognised_backends = {"matplotlib", "origin"}
    for backend in deduped_backends:
        if backend not in recognised_backends:
            log.warning("Unsupported plot backend '%s'; skipping", backend)
    plot_backends = tuple(b for b in deduped_backends if b in recognised_backends)

    wants_matplotlib = "matplotlib" in plot_backends
    wants_origin_requested = "origin" in plot_backends

    fabrication_index = build_fabrication_index(config.fabrication_files, log)
    stats = BuildStats()
    grouped: Dict[Tuple[str, int, int], List[MeasurementRecord]] = {}
    plot_paths: List[Path] = []
    plot_cache: Dict[str, Path] = {}
    plot_name_to_path: Dict[str, Path] = {}
    origin_artifacts: Dict[str, OriginArtifact] = {}
    origin_cache: Dict[str, OriginArtifact] = {}
    figure_size = _normalise_figsize(getattr(config, "matplotlib_figsize", (4.0, 2.25)))
    plot_dir = output_dir / config.plot_dir_name
    origin_dir = output_dir / config.origin_dir_name
    origin_enabled = wants_origin_requested
    origin_disabled_reason: Optional[str] = None
    microscope_index = _group_microscope_measurements(config.microscope_files, config.annealing_files, log)
    video_index = _collect_video_metrics(config.video_files, log)
    total = len(config.annealing_files)
    for idx, path in enumerate(config.annealing_files, start=1):
        try:
            df = _load_annealing(path)
        except Exception:
            log.exception("Failed to parse %s", path)
            stats.skipped += 1
            if progress_callback:
                progress_callback(idx, total)
            continue
        metadata = _metadata_from_path(path, root_for_relpaths)
        ok, mean_error = _resistance_sanity_check(df)
        if not ok:
            stats.resistance_checks_failed += 1
            if mean_error is None:
                log.warning("R≈V/I sanity check failed for %s", path)
            else:
                log.warning(
                    "R≈V/I sanity check failed for %s (mean error %.2f%%)",
                    path,
                    mean_error * 100,
                )
        if metadata.draw_x is None or metadata.piece_y is None:
            log.warning(
                "Skipping %s because the microwire draw/piece identifiers could not be parsed",
                path,
            )
            stats.skipped += 1
            if progress_callback:
                progress_callback(idx, total)
            continue
        key = (metadata.composition_token, metadata.draw_x, metadata.piece_y)
        record = MeasurementRecord(
            path=path,
            metadata=metadata,
            dataframe=df,
            sanity_ok=ok,
            sanity_error=mean_error,
        )
        grouped.setdefault(key, []).append(record)
        stats.parsed += 1
        if progress_callback:
            progress_callback(idx, total)
    rows: List[Dict[str, object]] = []
    for (composition, draw_x, piece_y), records in sorted(grouped.items()):
        draw_info = fabrication_index.get_draw(composition, draw_x)
        piece_info = fabrication_index.get_piece(composition, draw_x, piece_y)
        row: Dict[str, object] = {column: None for column in OUTPUT_COLUMNS}
        row["Composition"] = composition
        row["Microwire"] = _microwire_label(draw_x, piece_y)
        row["d (µm)"] = _value_for_output(piece_info, "d_um")
        row["D (µm)"] = _value_for_output(piece_info, "D_um")
        row["d/D"] = _value_for_output(piece_info, "d_over_D")
        row["Length (m)"] = _value_for_output(piece_info, "length_m")
        row["Production datetime"] = _value_for_output(draw_info, "production_datetime")
        row["Mass (g)"] = _value_for_output(draw_info, "mass_g")
        row["Resistance (Ω)"] = _value_for_output(draw_info, "fabrication_resistance_ohm")
        row["Temperature (°C)"] = _value_for_output(draw_info, "fabrication_temperature_c")
        row["Winding speed (m/min)"] = _value_for_output(draw_info, "winding_speed_m_per_min")
        row["Glass feeding (mm/min)"] = _value_for_output(draw_info, "glass_feed_mm_per_min")
        row["Underpressure"] = _value_for_output(draw_info, "underpressure")
        row["Notes"] = _compose_notes(draw_info, piece_info)
        microscope_data = microscope_index.get((composition, draw_x, piece_y))
        if microscope_data:
            d_numeric = _parse_numeric(row["d (µm)"])
            if d_numeric is None:
                d_from_micro = microscope_data.best_core()
                if d_from_micro is not None:
                    row["d (µm)"] = d_from_micro
                    d_numeric = d_from_micro
            D_numeric = _parse_numeric(row["D (µm)"])
            if D_numeric is None:
                D_from_micro = microscope_data.best_glass()
                if D_from_micro is not None:
                    row["D (µm)"] = D_from_micro
                    D_numeric = D_from_micro
            ratio_numeric = _parse_numeric(row["d/D"])
            if ratio_numeric is None and d_numeric is not None and D_numeric not in (None, 0):
                try:
                    ratio = d_numeric / D_numeric
                except ZeroDivisionError:
                    ratio = None
                if ratio is not None and math.isfinite(ratio):
                    row["d/D"] = ratio
        video_data = video_index.get((composition, draw_x, piece_y))
        if video_data is None:
            video_data = video_index.get((composition, draw_x, None))
        if video_data:
            temp_numeric = _parse_numeric(row["Temperature (°C)"])
            if temp_numeric is None:
                temp = video_data.temperature()
                if temp is not None:
                    row["Temperature (°C)"] = temp
            under_numeric = _parse_numeric(row["Underpressure"])
            if under_numeric is None:
                under_value = video_data.underpressure()
                if under_value is not None:
                    row["Underpressure"] = under_value
        if not draw_info:
            stats.missing_draw += 1
        if not piece_info:
            stats.missing_piece += 1
        high_record = _select_high_measurement(records)
        low_record = _select_low_measurement(records)
        if low_record and high_record:
            high_sp = high_record.metadata.setpoint_mA
            low_sp = low_record.metadata.setpoint_mA
            setpoints = {r.metadata.setpoint_mA for r in records if r.metadata.setpoint_mA is not None}
            if high_sp is not None and low_sp == high_sp and len(setpoints) <= 1:
                low_record = None
        if high_record:
            row["File 1000 mA"] = high_record.metadata.file_name
        else:
            stats.missing_high_measurement += 1
            log.warning("No 1000 mA measurement found for %s %s", composition, row["Microwire"] or "(unknown)")
        if low_record:
            row["File low mA"] = low_record.metadata.file_name
            if low_record.metadata.setpoint_mA is not None:
                row["Low mA value (mA)"] = low_record.metadata.setpoint_mA
        else:
            stats.missing_low_measurement += 1
            log.warning("No low-current measurement found for %s %s", composition, row["Microwire"] or "(unknown)")
        if wants_matplotlib:
            if high_record:
                cached = plot_cache.get(high_record.metadata.measurement_id)
                if cached is None:
                    try:
                        cached = _plot_measurement_matplotlib(
                            high_record.dataframe,
                            high_record.path,
                            plot_dir,
                            figure_size,
                        )
                        plot_cache[high_record.metadata.measurement_id] = cached
                        plot_paths.append(cached)
                    except Exception:
                        log.exception("Failed to generate plot for %s", high_record.path)
                        cached = None
                if cached is not None:
                    figure_name = Path(cached).name
                    row["Figure — 1000 mA"] = figure_name
                    plot_name_to_path.setdefault(figure_name, cached)
            if low_record:
                cached = plot_cache.get(low_record.metadata.measurement_id)
                if cached is None:
                    try:
                        cached = _plot_measurement_matplotlib(
                            low_record.dataframe,
                            low_record.path,
                            plot_dir,
                            figure_size,
                        )
                        plot_cache[low_record.metadata.measurement_id] = cached
                        if cached not in plot_paths:
                            plot_paths.append(cached)
                    except Exception:
                        log.exception("Failed to generate plot for %s", low_record.path)
                        cached = None
                if cached is not None:
                    figure_name = Path(cached).name
                    row["Figure — low mA"] = figure_name
                    plot_name_to_path.setdefault(figure_name, cached)
        if origin_enabled:
            if high_record:
                cached_origin = origin_cache.get(high_record.metadata.measurement_id)
                if cached_origin is None:
                    try:
                        cached_origin = _plot_measurement_origin(
                            high_record.dataframe,
                            high_record.path,
                            origin_dir,
                            log,
                        )
                    except RuntimeError as exc:
                        if origin_disabled_reason is None:
                            origin_disabled_reason = str(exc) or exc.__class__.__name__
                            log.warning("Origin plotting disabled: %s", origin_disabled_reason)
                        origin_enabled = False
                        cached_origin = None
                    except Exception:
                        log.exception("Failed to generate Origin plot for %s", high_record.path)
                        cached_origin = None
                    else:
                        if cached_origin is not None:
                            origin_cache[high_record.metadata.measurement_id] = cached_origin
                            origin_artifacts.setdefault(cached_origin.descriptor, cached_origin)
                if cached_origin is not None:
                    row["Figure — 1000 mA (Origin)"] = cached_origin.descriptor
            if origin_enabled and low_record:
                cached_origin = origin_cache.get(low_record.metadata.measurement_id)
                if cached_origin is None:
                    try:
                        cached_origin = _plot_measurement_origin(
                            low_record.dataframe,
                            low_record.path,
                            origin_dir,
                            log,
                        )
                    except RuntimeError as exc:
                        if origin_disabled_reason is None:
                            origin_disabled_reason = str(exc) or exc.__class__.__name__
                            log.warning("Origin plotting disabled: %s", origin_disabled_reason)
                        origin_enabled = False
                        cached_origin = None
                    except Exception:
                        log.exception("Failed to generate Origin plot for %s", low_record.path)
                        cached_origin = None
                    else:
                        if cached_origin is not None:
                            origin_cache[low_record.metadata.measurement_id] = cached_origin
                            origin_artifacts.setdefault(cached_origin.descriptor, cached_origin)
                if cached_origin is not None:
                    row["Figure — low mA (Origin)"] = cached_origin.descriptor
        rows.append(row)
        stats.rows_built += 1
    if rows:
        df_out = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    else:
        df_out = pd.DataFrame(columns=OUTPUT_COLUMNS)
    exports: Dict[str, Path] = {}
    requested_formats = tuple(dict.fromkeys(config.export_formats)) if config.export_formats else ("csv",)
    behaviours = {
        (key.lower() if isinstance(key, str) else ""): str(value).lower()
        for key, value in (config.export_behaviour or {}).items()
    }
    for fmt in requested_formats:
        fmt_lower = fmt.lower()
        if fmt_lower == "csv":
            csv_path = output_dir / f"{output_name}.csv"
            behaviour = behaviours.get("csv", "replace")
            to_write = df_out
            if behaviour == "append" and csv_path.exists():
                try:
                    existing = pd.read_csv(csv_path)
                except Exception:
                    log.exception("Failed to read existing CSV at %s; overwriting", csv_path)
                else:
                    to_write = pd.concat([existing, df_out], ignore_index=True)
                    to_write = to_write.drop_duplicates()
            to_write.to_csv(csv_path, index=False)
            exports["csv"] = csv_path
        elif fmt_lower == "excel":
            excel_path = output_dir / f"{output_name}.xlsx"
            behaviour = behaviours.get("excel", "replace")
            to_write = df_out
            if behaviour == "append" and excel_path.exists():
                try:
                    existing = pd.read_excel(excel_path)
                except Exception:
                    log.exception("Failed to read existing Excel at %s; overwriting", excel_path)
                else:
                    to_write = pd.concat([existing, df_out], ignore_index=True)
                    to_write = to_write.drop_duplicates()
            excel_frame = to_write.copy()
            columns_to_blank = [
                column for column in (FIGURE_COLUMNS + ORIGIN_FIGURE_COLUMNS) if column in excel_frame.columns
            ]
            for column in columns_to_blank:
                excel_frame[column] = None
            try:
                import xlsxwriter  # noqa: F401

                with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
                    excel_frame.to_excel(writer, index=False, sheet_name="Sheet1")
                    try:
                        _embed_assets_with_xlsxwriter(
                            writer,
                            to_write,
                            plot_name_to_path,
                            plot_dir,
                            origin_artifacts,
                            figure_size,
                            log,
                        )
                    except Exception:
                        log.exception("Failed to embed figures into %s", excel_path)
            except ImportError:
                excel_frame.to_excel(excel_path, index=False)
                try:
                    _embed_plots_in_excel_openpyxl(
                        excel_path,
                        to_write,
                        plot_name_to_path,
                        plot_dir,
                        log,
                        figure_size,
                    )
                except Exception:
                    log.exception("Failed to embed figure images into %s", excel_path)
            exports["excel"] = excel_path
        else:
            log.warning("Unsupported export format '%s'; skipping", fmt)
    log.info(
        "Measurements parsed: %s | Skipped: %s | Rows built: %s | Missing draw info: %s | Missing piece info: %s | Missing 1000 mA: %s | Missing low mA: %s | R≈V/I failures: %s",
        stats.parsed,
        stats.skipped,
        stats.rows_built,
        stats.missing_draw,
        stats.missing_piece,
        stats.missing_high_measurement,
        stats.missing_low_measurement,
        stats.resistance_checks_failed,
    )
    return BuildResult(
        dataframe=df_out,
        exports=exports,
        plot_paths=plot_paths,
        origin_artifacts=origin_artifacts,
        stats=stats,
    )


__all__ = [
    "BuilderConfig",
    "BuildResult",
    "BuildStats",
    "OriginArtifact",
    "FabricationIndex",
    "build_database",
    "build_fabrication_index",
    "LOGGER_NAME",
    "DEFAULT_OUTPUT_NAME",
]
